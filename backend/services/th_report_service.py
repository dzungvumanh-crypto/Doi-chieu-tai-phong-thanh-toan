"""Service xử lý báo cáo dữ liệu thanh toán SWIFT — Phòng Tổng hợp.

Luồng: parse IN/OUT files → fill template D00054 → trả về bytes.
"""
import io
import logging
import math
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# ── Đường dẫn template ───────────────────────────────────────────────────────
# Dùng rglob thay vì hardcode path vì tên folder tiếng Việt trên Windows
# lưu dạng NFD nhưng Python string literal dùng NFC — pathlib.exists() trả False.
def _find_template() -> Path:
    templates_dir = Path(__file__).resolve().parent.parent.parent / "templates"
    for path in templates_dir.rglob("D00054*.xlsx"):
        if path.is_file() and not path.name.startswith("~"):
            return path
    raise FileNotFoundError(f"Template D00054 không tìm thấy trong {templates_dir}")

_TEMPLATE_PATH = _find_template()

# ── Hàng và cột trong template ────────────────────────────────────────────────
_PERIOD_ROW, _PERIOD_COL = 5, 3          # Row 5, col C = "202605"
_DATA_ROW_START, _DATA_ROW_END = 20, 214  # Dữ liệu quốc gia
_COUNTRY_COL = 3                          # col C = tên quốc gia
_TOTAL_ROW = 217                          # Tổng cộng

# Mapping cột dữ liệu (1-based): E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12
_COL_IN_COUNT   = 5   # E — Số lượng GD đến
_COL_IN_AMT     = 6   # F — Giá trị GD đến
_COL_CN_COUNT   = 7   # G — GD đi Cá nhân: số lượng
_COL_CN_AMT     = 8   # H — GD đi Cá nhân: giá trị
_COL_DN_COUNT   = 9   # I — GD đi Doanh nghiệp: số lượng
_COL_DN_AMT     = 10  # J — GD đi Doanh nghiệp: giá trị
_COL_FI_COUNT   = 11  # K — GD đi TCTD: số lượng
_COL_FI_AMT     = 12  # L — GD đi TCTD: giá trị

_DATA_COLS = [
    _COL_IN_COUNT, _COL_IN_AMT,
    _COL_CN_COUNT, _COL_CN_AMT,
    _COL_DN_COUNT, _COL_DN_AMT,
    _COL_FI_COUNT, _COL_FI_AMT,
]

# ── Mapping tên quốc gia: IN/OUT (UPPER) → Template (UPPER) ──────────────────
# Một số tên trong file SWIFT khác với tên trong template của NHNN.
# Quốc gia luôn để 0, không lấy từ file IN/OUT
_SKIP_COUNTRIES: set[str] = {"VIET NAM"}

_COUNTRY_NAME_MAP: dict[str, str] = {
    "BOLIVIA, PLURINATIONAL STATE OF":           "BOLIVIA",
    "CABO VERDE":                                "CAPE VERDE",
    "CHINA":                                     "CHINA MAINLAND",
    "CZECHIA":                                   "CZECH REPUBLIC",
    "KOREA (DEMOCRATIC PEOPLE'S REPUBLIC OF)":   "DEMOCRATIC PEOPLE'S REPUBLIC OF KOREA",
    "CONGO, THE DEMOCRATIC REPUBLIC OF THE":     "DEMOCRATIC REPUBLIC OF THE CONGO",
    "ESWATINI":                                  "SWAZILAND",
    "LIBYA":                                     "LIBYAN ARAB JAMAHIRIYA",
    "MICRONESIA (FEDERATED STATES OF)":          "MICRONESIA, FEDERATED STATES OF",
    "MOLDOVA, REPUBLIC OF":                      "REPUBLIC OF MOLDOVA",
    "KOREA, REPUBLIC OF":                        "REPUBLIC OF KOREA",
    "NORTH MACEDONIA":                           "THE FORMER YUGOSLAV REPUBLIC OF MACEDONIA",
    "TANZANIA, UNITED REPUBLIC OF":              "UNITED REPUBLIC OF TANZANIA",
    "TURKIYE":                                   "TURKEY",
    "UNITED STATES OF AMERICA":                  "UNITED STATES",
    "VENEZUELA (BOLIVARIAN REPUBLIC OF)":        "VENEZUELA",
}


def _normalize_country(name: str) -> str:
    """Chuẩn hóa tên quốc gia để tra cứu trong template."""
    upper = name.strip().upper()
    return _COUNTRY_NAME_MAP.get(upper, upper)


# ── Đọc sheet dữ liệu ──────────────────────────────────────────────────────────
# Tên sheet dữ liệu khác nhau tùy công cụ export ("Result" hoặc "Export
# Worksheet"); sheet "SQL" (nếu có) chỉ chứa câu query, không phải dữ liệu.
_KNOWN_DATA_SHEET_NAMES = ("Result", "Export Worksheet")


def _read_data_sheet(file_bytes: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    try:
        sheet_name = next((n for n in _KNOWN_DATA_SHEET_NAMES if n in wb.sheetnames), None)
        if sheet_name is None:
            sheet_name = next((n for n in wb.sheetnames if n.upper() != "SQL"), wb.sheetnames[0])
    finally:
        wb.close()
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, dtype=str)


# ── Parse ─────────────────────────────────────────────────────────────────────

def parse_incoming(file_bytes: bytes) -> dict[str, dict]:
    """Parse file Lệnh đến (IN).

    Trả về {tên đã chuẩn hóa: {"count": int, "amount": float}}.
    amount = tổng STTLM_AMT / 1000, làm tròn 2 chữ số thập phân.
    """
    df = _read_data_sheet(file_bytes)
    result: dict[str, dict] = {}

    for _, row in df.iterrows():
        cthed_raw = str(row.get("CTHED", "") or "").strip()
        if not cthed_raw:
            continue
        cthed = _normalize_country(cthed_raw)
        if cthed in _SKIP_COUNTRIES:
            continue

        try:
            count_raw = float(str(row["TOTAL"]))
            count = 0 if math.isnan(count_raw) else int(count_raw)
        except (ValueError, TypeError, KeyError):
            count = 0
        try:
            amt_raw = float(str(row["STTLM_AMT"]).replace(",", "."))
            amount = 0.0 if math.isnan(amt_raw) else round(amt_raw / 1000, 2)
        except (ValueError, TypeError, KeyError):
            amount = 0.0

        # Bỏ qua dòng không có dữ liệu thực
        if count == 0 and amount == 0.0:
            continue

        if cthed not in result:
            result[cthed] = {"count": 0, "amount": 0.0}
        result[cthed]["count"] += count
        result[cthed]["amount"] = round(result[cthed]["amount"] + amount, 2)

    return result


def parse_outgoing(file_bytes: bytes) -> dict[str, dict]:
    """Parse file Lệnh đi (OUT).

    Trả về {tên đã chuẩn hóa: {
        "cn_count", "cn_amt", "dn_count", "dn_amt", "fi_count", "fi_amt"
    }}.
    CUST_TYPE: CN → cá nhân, DN → doanh nghiệp, TCTD/TCTDO → TCTD.
    amount = TOTAL_AMT / 1000, làm tròn 2 chữ số thập phân.
    """
    df = _read_data_sheet(file_bytes)
    result: dict[str, dict] = {}

    _empty: dict = {
        "cn_count": 0, "cn_amt": 0.0,
        "dn_count": 0, "dn_amt": 0.0,
        "fi_count": 0, "fi_amt": 0.0,
    }

    for _, row in df.iterrows():
        cthed_raw = str(row.get("CTHED", "") or "").strip()
        if not cthed_raw:
            continue
        cthed = _normalize_country(cthed_raw)
        if cthed in _SKIP_COUNTRIES:
            continue

        cust_type = str(row.get("CUST_TYPE", "") or "").strip().upper()
        # Bỏ qua dòng không có loại khách hàng (NaN / rỗng)
        if not cust_type or cust_type == "NAN":
            continue

        try:
            count_raw = float(str(row["TOTAL"]))
            count = 0 if math.isnan(count_raw) else int(count_raw)
        except (ValueError, TypeError, KeyError):
            count = 0
        try:
            amt_raw = float(str(row["TOTAL_AMT"]).replace(",", "."))
            amount = 0.0 if math.isnan(amt_raw) else round(amt_raw / 1000, 2)
        except (ValueError, TypeError, KeyError):
            amount = 0.0

        if cthed not in result:
            result[cthed] = dict(_empty)

        if cust_type == "CN":
            result[cthed]["cn_count"] += count
            result[cthed]["cn_amt"] = round(result[cthed]["cn_amt"] + amount, 2)
        elif cust_type == "DN":
            result[cthed]["dn_count"] += count
            result[cthed]["dn_amt"] = round(result[cthed]["dn_amt"] + amount, 2)
        elif cust_type in ("TCTD", "TCTDO"):
            result[cthed]["fi_count"] += count
            result[cthed]["fi_amt"] = round(result[cthed]["fi_amt"] + amount, 2)
        else:
            logger.warning("CUST_TYPE không nhận dạng được: %s (quốc gia %s)", cust_type, cthed)

    return result


# ── Fill template ─────────────────────────────────────────────────────────────

def fill_template(
    incoming: dict[str, dict],
    outgoing: dict[str, dict],
    period_yyyymm: str,
) -> bytes:
    """Điền dữ liệu vào template D00054, trả về bytes của file Excel đã điền.

    Chỉ ghi giá trị vào cell, không thay đổi format.
    Ô không có dữ liệu được điền 0 (không để trống).
    """
    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    ws = wb.active

    # ── Cập nhật kỳ báo cáo ──────────────────────────────────────────────────
    ws.cell(_PERIOD_ROW, _PERIOD_COL).value = period_yyyymm

    # ── Xây mapping: UPPER(tên quốc gia) → số hàng ───────────────────────────
    country_row: dict[str, int] = {}
    for r in range(_DATA_ROW_START, _DATA_ROW_END + 1):
        raw = ws.cell(r, _COUNTRY_COL).value
        if raw:
            country_row[str(raw).strip().upper()] = r

    # ── Pre-fill tất cả ô dữ liệu bằng 0 ────────────────────────────────────
    for r in range(_DATA_ROW_START, _DATA_ROW_END + 1):
        for col in _DATA_COLS:
            ws.cell(r, col).value = 0

    unmatched: set[str] = set()

    # ── Điền Lệnh đến ────────────────────────────────────────────────────────
    for cthed, data in incoming.items():
        row = country_row.get(cthed)
        if row is None:
            unmatched.add(cthed)
            continue
        ws.cell(row, _COL_IN_COUNT).value = data["count"]
        ws.cell(row, _COL_IN_AMT).value   = data["amount"]

    # ── Điền Lệnh đi ─────────────────────────────────────────────────────────
    for cthed, data in outgoing.items():
        row = country_row.get(cthed)
        if row is None:
            unmatched.add(cthed)
            continue
        ws.cell(row, _COL_CN_COUNT).value = data["cn_count"]
        ws.cell(row, _COL_CN_AMT).value   = data["cn_amt"]
        ws.cell(row, _COL_DN_COUNT).value = data["dn_count"]
        ws.cell(row, _COL_DN_AMT).value   = data["dn_amt"]
        ws.cell(row, _COL_FI_COUNT).value = data["fi_count"]
        ws.cell(row, _COL_FI_AMT).value   = data["fi_amt"]

    if unmatched:
        logger.warning("Không tìm thấy quốc gia trong template: %s", sorted(unmatched))

    # ── Tính tổng hàng 217 ───────────────────────────────────────────────────
    for col in _DATA_COLS:
        total = sum(
            (ws.cell(r, col).value or 0)
            for r in range(_DATA_ROW_START, _DATA_ROW_END + 1)
        )
        ws.cell(_TOTAL_ROW, col).value = round(total, 2)

    # ── Lưu ra bytes ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
