# -*- coding: utf-8 -*-
"""
exporters.py
------------
Xuất kết quả đối chiếu ra file Excel (.xlsx) bằng openpyxl, có format cơ bản
(tô màu header, tự co giãn cột) cho dễ nhìn.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DIFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _write_df(ws: Worksheet, df: pd.DataFrame, highlight_diff_col: str | None = None):
    if df is None or df.empty:
        ws.append(["(Không có dữ liệu)"])
        return
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df.iterrows():
        values = []
        for v in row:
            if pd.isna(v):
                values.append("")
            else:
                values.append(v)
        ws.append(values)

    if highlight_diff_col and highlight_diff_col in df.columns:
        col_idx = list(df.columns).index(highlight_diff_col) + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            try:
                if cell.value not in (None, "", 0) and float(cell.value) != 0:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = DIFF_FILL
            except (ValueError, TypeError):
                pass

    # Auto width cơ bản
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"


def export_summary_excel(
    out_path: str,
    summary_den: pd.DataFrame | None,
    summary_di: pd.DataFrame | None,
):
    """File excel tổng hợp: 1 sheet cho điện đến, 1 sheet cho điện đi
    (chỉ tạo sheet nào có dữ liệu được đối chiếu)."""
    wb = Workbook()
    first = True

    if summary_den is not None:
        ws = wb.active if first else wb.create_sheet()
        ws.title = "Tong_hop_Dien_Den"
        first = False
        _write_df(ws, summary_den, highlight_diff_col="Chênh lệch")

    if summary_di is not None:
        ws = wb.active if first else wb.create_sheet()
        ws.title = "Tong_hop_Dien_Di"
        first = False
        _write_df(ws, summary_di, highlight_diff_col="Chênh lệch")

    if first:
        # không có dữ liệu nào
        wb.active.title = "Tong_hop"
        wb.active.append(["Chưa có dữ liệu đối chiếu nào được nạp."])

    wb.save(out_path)


def export_diff_excel(
    out_path: str,
    saa_den_only: pd.DataFrame | None,
    ql_den_only: pd.DataFrame | None,
    ql_di_only: pd.DataFrame | None,
    saa_di_only: pd.DataFrame | None,
    di_matched_not_ack: pd.DataFrame | None = None,
):
    """File excel chi tiết lệch với tối đa 5 sheet:
       SAADENONLY, QLDENONLY, QLDIONLY, SAADIONLY, DI_KHOPKHOA_CHUA_ACK
    (chỉ tạo sheet nào có dữ liệu nguồn tương ứng đã được nạp).

    di_matched_not_ack: các bản ghi điện đi ĐÃ KHỚP KHOÁ ở cả 2 hệ thống
    nhưng không bên nào báo trạng thái Ack (ACK/NAK ở Quản lý điện đi /
    Netw. Status ở SAA điện đi) - cần giao dịch viên kiểm tra thủ công."""
    wb = Workbook()
    first = True
    sheets = [
        ("SAADENONLY", saa_den_only),
        ("QLDENONLY", ql_den_only),
        ("QLDIONLY", ql_di_only),
        ("SAADIONLY", saa_di_only),
        ("DI_KHOPKHOA_CHUA_ACK", di_matched_not_ack),
    ]
    for name, df in sheets:
        if df is None:
            continue
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        first = False
        _write_df(ws, df)

    if first:
        wb.active.title = "ChiTietLech"
        wb.active.append(["Chưa có dữ liệu đối chiếu nào được nạp."])

    wb.save(out_path)
