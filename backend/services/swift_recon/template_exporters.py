# -*- coding: utf-8 -*-
"""
template_exporters.py
----------------------
Xuất "Excel Tổng hợp theo biểu mẫu" / "Excel Chi tiết lệch theo biểu mẫu" —
dùng ĐÚNG file biểu mẫu (Mẫu 04 / Mẫu 05) người dùng cung cấp làm khung nền
(giữ nguyên quốc hiệu, tiêu đề, ký tên...), chỉ ghi đè phần dữ liệu.

File này KHÔNG đụng đến parsers.py / reconcile.py / exporters.py gốc — chỉ
ĐỌC (import reconcile.match_by_key) để lấy đúng danh sách bản ghi lệch, hoàn
toàn không sửa logic đối chiếu.

┌──────────────────────────────────────────────────────────────────────────┐
│ NGUỒN PHÂN LOẠI HỆ THỐNG (SWIFT/IPCAS/P-HUB) — đã kiểm chứng bằng dữ     │
│ liệu thật cho CẢ điện đến lẫn điện đi (đợt 4)                            │
├──────────────────────────────────────────────────────────────────────────┤
│ Đã xác nhận bằng file mẫu thật cả QL_DEN (MSG_IN_..._QL_DEN.xls) VÀ      │
│ QL_DI (MSG_OUT_..._QL_DI.xls): cả 2 loại file đều có sẵn cột             │
│ "Channel Process" chứa thẳng giá trị "IPCAS"/"PMHUB" — classify_system() │
│ đọc thẳng cột này, không cần đoán qua hoa văn SaSeq/Msg Key nữa (cách cũ │
│ chỉ còn là DỰ PHÒNG cho trường hợp 1 file nào đó về sau không có cột     │
│ này).                                                                     │
│                                                                            │
│ Tên cột thật đã xác nhận và cập nhật vào FIELD_CANDIDATES:               │
│   - QL_DEN : Số tham chiếu="RefNo", Số tiền="Amount", Loại tiền="Curent",│
│              Ngân hàng gửi="Send Bic"                                    │
│   - QL_DI  : Số tham chiếu="Refno" (chú ý viết hoa/thường khác QL_DEN — │
│              đã xử lý so khớp không phân biệt hoa/thường), Số tiền=      │
│              "Amount", Loại tiền="Ccy", Ngân hàng gửi="Send Bic"         │
│                                                                            │
│ Dòng "Tổng chênh lệch số lượng điện theo loại điện" tôi tính bằng tổng   │
│ trị tuyệt đối của (SWIFT − (IPCAS+P-HUB)) theo từng loại điện — quy ước  │
│ audit phổ biến nhất. Nếu công thức đúng của bạn khác, nói rõ để tôi sửa  │
│ lại hàm build_system_summary() bên dưới.                                 │
└──────────────────────────────────────────────────────────────────────────┘
"""
from __future__ import annotations

import os
import re
from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from backend.services.swift_recon import reconcile

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "templates")

TEMPLATE_FILES = {
    ("den", "summary"): "tonghop_den.xlsx",
    ("di", "summary"): "tonghop_di.xlsx",
    ("den", "diff"): "chitiet_den.xlsx",
    ("di", "diff"): "chitiet_di.xlsx",
}


# ──────────────────────────────────────────────────────────────────────────
# Phân loại hệ thống nguồn (SWIFT / IPCAS / PMHUB / ARS) — xem cảnh báo ở
# đầu file. seq_raw PHẢI là giá trị THÔ (SaSeq đầy đủ / Msg Key đầy đủ),
# KHÔNG phải _key đã bị cắt bớt ký tự.
# ──────────────────────────────────────────────────────────────────────────
_SEQ_PATTERN = re.compile(r"^(\d{4})([A-Za-z])(.+)$")

# Tên cột "hệ thống nguồn" khả dĩ trong file Quản lý điện — "Channel Process"
# đã XÁC NHẬN đúng với file QL_DEN thật (chứa thẳng "IPCAS"/"PMHUB"). Thử lần
# lượt các tên trong danh sách, dùng cột đầu tiên tồn tại VÀ có giá trị.
CHANNEL_COL_CANDIDATES = ["Channel Process", "Channel", "System", "Source System"]

_KNOWN_SYSTEM_VALUES = {"SWIFT", "IPCAS", "PMHUB", "P-HUB", "ARS"}


def classify_system(source: str, seq_raw, channel_raw=None) -> str:
    """Phân loại 1 bản ghi về hệ thống nguồn.

    Ưu tiên 1: nếu có giá trị channel_raw (đọc từ cột "Channel Process" hay
    tương đương) và nó là 1 trong các giá trị hệ thống đã biết -> dùng THẲNG
    giá trị đó (đáng tin cậy nhất, không cần đoán).

    Dự phòng (chỉ dùng khi KHÔNG có cột channel, ví dụ file QL_DI chưa xác
    nhận có cột tương tự hay không): suy đoán qua hoa văn chuỗi SaSeq/Msg Key
    thô — xem cảnh báo ở đầu file, phần này CHƯA được kiểm chứng bằng dữ
    liệu QL_DI thật.
    """
    if source.startswith("SAA"):
        return "SWIFT"

    ch = ("" if channel_raw is None else str(channel_raw)).strip().upper()
    if ch in _KNOWN_SYSTEM_VALUES:
        return "PMHUB" if ch in ("PMHUB", "P-HUB") else ch

    # Dự phòng: đoán qua hoa văn <4 số chi nhánh><chữ cái><...> của SaSeq/Msg Key
    s = ("" if seq_raw is None else str(seq_raw)).strip()
    m = _SEQ_PATTERN.match(s)
    if not m:
        return "IPCAS"
    brcd, letter = m.group(1), m.group(2).upper()
    if letter == "O":
        return "IPCAS"
    if letter == "S":
        return "PMHUB"
    if letter == "R":
        return "ARS" if brcd == "0000" else "PMHUB"
    return "PMHUB"


_SEQ_COL_BY_SOURCE = {
    "SAA_DEN": "Reception Info",
    "SAA_DI": "Reception Info",
    "QL_DEN": "SaSeq",
    "QL_DI": "Msg Key",
}

FIELD_CANDIDATES = {
    "ref": {
        "SAA_DEN": ["Reference"], "SAA_DI": ["Reference"],
        "QL_DEN": ["RefNo", "Msg Key"],
        "QL_DI": ["Refno", "RefNo", "Reference", "Trans Ref", "Msg Ref", "Ref"],
    },
    "amount": {
        "QL_DEN": ["Amount", "Số tiền", "Value", "Amt"],
        "QL_DI": ["Amount", "Số tiền", "Value", "Amt"],
    },
    "currency": {
        "QL_DEN": ["Curent", "Currency", "Ccy", "Loại tiền"],
        "QL_DI": ["Ccy", "Curent", "Currency", "Loại tiền"],
    },
    "bank": {
        "SAA_DEN": ["Correspondent"], "SAA_DI": ["Correspondent"],
        "QL_DEN": ["Send Bic", "Correspondent", "Bank", "Sender", "BIC", "Sender BIC"],
        "QL_DI": ["Send Bic", "Correspondent", "Bank", "Sender", "BIC", "Sender BIC"],
    },
}


def _first_present(row: pd.Series, candidates: list) -> str:
    """Thử lần lượt các tên cột khả dĩ — khớp CHÍNH XÁC trước, nếu không có
    thì thử khớp KHÔNG PHÂN BIỆT HOA/THƯỜNG (phòng trường hợp file thật đặt
    tên cột khác cách viết hoa so với candidate, ví dụ "Refno" so với
    "RefNo")."""
    for name in candidates:
        if name in row.index:
            v = row[name]
            if v not in (None, "", "nan"):
                return v
    lower_map = {str(c).strip().lower(): c for c in row.index}
    for name in candidates:
        real_col = lower_map.get(name.strip().lower())
        if real_col is not None:
            v = row[real_col]
            if v not in (None, "", "nan"):
                return v
    return ""


def _find_channel_col(df: pd.DataFrame) -> str | None:
    for name in CHANNEL_COL_CANDIDATES:
        if name in df.columns:
            return name
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    for name in CHANNEL_COL_CANDIDATES:
        real_col = lower_map.get(name.strip().lower())
        if real_col is not None:
            return real_col
    return None


def _system_series(df: pd.DataFrame, source: str) -> pd.Series:
    if source.startswith("SAA"):
        return pd.Series(["SWIFT"] * len(df), index=df.index)
    seq_col = _SEQ_COL_BY_SOURCE[source]
    seq_series = df[seq_col] if seq_col in df.columns else pd.Series([""] * len(df), index=df.index)
    channel_col = _find_channel_col(df)
    if channel_col is not None:
        return df.apply(
            lambda r: classify_system(source, r.get(seq_col, ""), r.get(channel_col)), axis=1
        )
    if seq_col not in df.columns:
        return pd.Series(["PMHUB"] * len(df), index=df.index)
    return seq_series.apply(lambda v: classify_system(source, v))


# ──────────────────────────────────────────────────────────────────────────
# Tổng hợp theo hệ thống (dùng cho "Xuất Excel Tổng hợp theo biểu mẫu")
# ──────────────────────────────────────────────────────────────────────────
def build_system_summary(df_a: pd.DataFrame, source_a: str, df_b: pd.DataFrame, source_b: str) -> pd.DataFrame:
    """Trả về DataFrame: Loại điện | SWIFT | IPCAS | P-HUB | Chênh lệch

    Cột SWIFT/IPCAS/P-HUB đếm TOÀN BỘ bản ghi mỗi bên đã import (không chỉ
    phần lệch) — giống hệt ý nghĩa các cột đó trong biểu mẫu Tổng hợp.

    Cột "Chênh lệch" tính bằng SỐ BẢN GHI THỰC SỰ KHÔNG KHỚP KHOÁ (ONLY_A +
    ONLY_B theo _msg_type, lấy từ reconcile.match_by_key() — ĐÚNG cơ chế
    khoá dùng ở tab "Kết quả đối chiếu" và file "Chi tiết lệch"), KHÔNG PHẢI
    hiệu số lượng (SWIFT − (IPCAS+P-HUB)) như trước — hiệu số lượng có thể
    che giấu sai lệch thật khi 2 bên tình cờ có cùng số lượng nhưng không
    cùng giao dịch.
    """
    parts = []
    for df, source in ((df_a, source_a), (df_b, source_b)):
        if df is None or len(df) == 0:
            continue
        parts.append(pd.DataFrame({
            "_msg_type": df["_msg_type"],
            "_system": _system_series(df, source),
        }))
    if not parts:
        return pd.DataFrame(columns=["Loại điện", "SWIFT", "IPCAS", "P-HUB", "Chênh lệch"])

    combo = pd.concat(parts, ignore_index=True)
    combo["_system"] = combo["_system"].replace({"ARS": "PMHUB"})  # biểu mẫu Tổng hợp chỉ có 3 cột
    pivot = combo.groupby(["_msg_type", "_system"]).size().unstack(fill_value=0)
    for col in ("SWIFT", "IPCAS", "PMHUB"):
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot.rename(columns={"PMHUB": "P-HUB"})[["SWIFT", "IPCAS", "P-HUB"]]
    pivot = pivot.reset_index().rename(columns={"_msg_type": "Loại điện"})

    mr = reconcile.match_by_key(df_a, df_b)
    only_a_types = df_a.loc[mr.only_a, "_msg_type"] if len(mr.only_a) else pd.Series(dtype=object)
    only_b_types = df_b.loc[mr.only_b, "_msg_type"] if len(mr.only_b) else pd.Series(dtype=object)
    diff_counts = pd.concat([only_a_types, only_b_types]).value_counts()
    pivot["Chênh lệch"] = pivot["Loại điện"].map(diff_counts).fillna(0).astype(int)

    pivot = pivot.sort_values("Loại điện").reset_index(drop=True)
    return pivot


# ──────────────────────────────────────────────────────────────────────────
# Chi tiết lệch theo hệ thống (dùng cho "Xuất Excel Chi tiết lệch theo biểu mẫu")
# Dùng lại reconcile.match_by_key() — KHÔNG tự viết lại logic ghép cặp.
# ──────────────────────────────────────────────────────────────────────────
def _row_to_template_dict(row: pd.Series, source: str) -> dict:
    seq_col = _SEQ_COL_BY_SOURCE[source]
    seq_raw = row.get(seq_col, "")
    channel_col = None
    for name in CHANNEL_COL_CANDIDATES:
        if name in row.index:
            channel_col = name
            break
    if channel_col is None:
        lower_map = {str(c).strip().lower(): c for c in row.index}
        for name in CHANNEL_COL_CANDIDATES:
            channel_col = lower_map.get(name.strip().lower())
            if channel_col is not None:
                break
    channel_raw = row.get(channel_col) if channel_col else None

    if source.startswith("SAA"):
        ref = row.get("Reference", "")
        cur_amt = str(row.get("Cur/Amt", "") or "").strip()
        parts = cur_amt.split(" ", 1)
        currency, amount = (parts[0], parts[1]) if len(parts) == 2 else ("", cur_amt)
        bank = row.get("Correspondent", "")
    else:
        ref = _first_present(row, FIELD_CANDIDATES["ref"][source])
        amount = _first_present(row, FIELD_CANDIDATES["amount"][source])
        currency = _first_present(row, FIELD_CANDIDATES["currency"][source])
        bank = _first_present(row, FIELD_CANDIDATES["bank"][source])

    return {
        "Loại điện": row.get("_msg_type", ""),
        "Số tham chiếu": ref,
        "Số SWIFT Seq": seq_raw,
        "Số tiền": amount,
        "Loại tiền": currency,
        "Ngân hàng gửi": bank,
        "Hệ thống": classify_system(source, seq_raw, channel_raw),
        "Ghi chú": "",
    }


def build_system_detail_rows(df_a: pd.DataFrame, source_a: str, df_b: pd.DataFrame, source_b: str) -> list:
    """Danh sách dict — mỗi dòng 1 bản ghi LỆCH (chỉ có ở 1 bên), y hệt tập
    hợp only_a + only_b mà reconcile.match_by_key() trả về (dùng đúng khoá
    đối chiếu gốc, không tính lại)."""
    mr = reconcile.match_by_key(df_a, df_b)
    rows = []
    for idx in mr.only_a:
        rows.append(_row_to_template_dict(df_a.loc[idx], source_a))
    for idx in mr.only_b:
        rows.append(_row_to_template_dict(df_b.loc[idx], source_b))
    return rows


# ──────────────────────────────────────────────────────────────────────────
# Helpers thao tác biểu mẫu Excel (openpyxl): tìm dòng mốc, thêm/bớt dòng dữ
# liệu đúng bằng số bản ghi cần điền, giữ nguyên style + merge của dòng mẫu.
# ──────────────────────────────────────────────────────────────────────────
def _find_row_by_col_value(ws: Worksheet, col: int, text: str, start_row: int = 1) -> int | None:
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip() == text:
            return r
    return None


def _row_merge_spans(ws: Worksheet, row: int) -> list:
    return [
        (m.min_col, m.max_col) for m in ws.merged_cells.ranges
        if m.min_row == row and m.max_row == row and m.min_col != m.max_col
    ]


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int):
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        if isinstance(d, MergedCell):
            continue
        d.font = copy(s.font)
        d.border = copy(s.border)
        d.fill = copy(s.fill)
        d.number_format = s.number_format
        d.alignment = copy(s.alignment)
    src_h = ws.row_dimensions[src_row].height
    if src_h:
        ws.row_dimensions[dst_row].height = src_h


def _ensure_data_rows(ws: Worksheet, data_start: int, footer_col: int, footer_text: str,
                       max_col: int, n_needed: int) -> int:
    """Đảm bảo có đúng n_needed dòng dữ liệu ngay dưới data_start (dùng dòng
    data_start làm khuôn style/merge), thêm/bớt dòng cho khớp, trả về vị trí
    MỚI của dòng mốc footer_text sau khi đã chỉnh."""
    footer_row = _find_row_by_col_value(ws, footer_col, footer_text, start_row=data_start)
    if footer_row is None:
        raise ValueError(f"Không tìm thấy dòng '{footer_text}' trong biểu mẫu")

    n_current = footer_row - data_start
    n_needed = max(n_needed, 0)
    spans = _row_merge_spans(ws, data_start)

    if n_needed > n_current:
        add_n = n_needed - n_current
        ws.insert_rows(footer_row, amount=add_n)
        for i in range(add_n):
            r = data_start + n_current + i
            _copy_row_style(ws, data_start, r, max_col)
            for c1, c2 in spans:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        footer_row += add_n
    elif n_needed < n_current:
        remove_n = n_current - n_needed
        ws.delete_rows(footer_row - remove_n, amount=remove_n)
        footer_row -= remove_n

    return footer_row


_DATE_RE = re.compile(r"(ngày)\s*\d{1,2}(\s*tháng\s*)\d{1,2}(\s*năm\s*)\d{4}", re.IGNORECASE)


def _update_date_cells(ws: Worksheet, dt: datetime | None = None):
    dt = dt or datetime.now()

    def _sub(m: re.Match) -> str:
        return f"{m.group(1)} {dt.day:02d}{m.group(2)}{dt.month:02d}{m.group(3)}{dt.year}"

    for row in ws.iter_rows():
        for c in row:
            if isinstance(c, MergedCell):
                continue
            if isinstance(c.value, str) and _DATE_RE.search(c.value):
                c.value = _DATE_RE.sub(_sub, c.value)


# ──────────────────────────────────────────────────────────────────────────
# API công khai
# ──────────────────────────────────────────────────────────────────────────
def export_summary_template(
    out_path: str, direction: str,
    df_a: pd.DataFrame, source_a: str, df_b: pd.DataFrame, source_b: str,
):
    """direction: 'den' hoặc 'di'."""
    fname = TEMPLATE_FILES[(direction, "summary")]
    wb = load_workbook(os.path.join(TEMPLATE_DIR, fname))
    ws = wb.active

    summary = build_system_summary(df_a, source_a, df_b, source_b)

    header_row = _find_row_by_col_value(ws, 2, "Loại điện")
    if header_row is None:
        raise ValueError("Không tìm thấy dòng tiêu đề 'Loại điện' trong biểu mẫu Tổng hợp")
    data_start = header_row + 1
    max_col = ws.max_column

    total_row = _ensure_data_rows(ws, data_start, 2, "Tổng số lượng điện", max_col, len(summary))

    for i, rec in enumerate(summary.to_dict("records")):
        r = data_start + i
        ws.cell(row=r, column=2, value=rec["Loại điện"])          # B: Loại điện
        ws.cell(row=r, column=4, value=int(rec["SWIFT"]))          # D: SWIFT
        ws.cell(row=r, column=5, value=int(rec["IPCAS"]))          # E: IPCAS
        ws.cell(row=r, column=9, value=int(rec["P-HUB"]))          # I: P-HUB
        ws.cell(row=r, column=12, value=int(rec["Chênh lệch"]))    # L: Chênh lệch

    ws.cell(row=total_row, column=4, value=int(summary["SWIFT"].sum()) if len(summary) else 0)
    ws.cell(row=total_row, column=5, value=int(summary["IPCAS"].sum()) if len(summary) else 0)
    ws.cell(row=total_row, column=9, value=int(summary["P-HUB"].sum()) if len(summary) else 0)
    tong_chenh_row = total_row + 1
    ws.cell(row=tong_chenh_row, column=12,
            value=int(summary["Chênh lệch"].abs().sum()) if len(summary) else 0)

    _update_date_cells(ws)
    wb.save(out_path)


def export_diff_template(
    out_path: str, direction: str,
    df_a: pd.DataFrame, source_a: str, df_b: pd.DataFrame, source_b: str,
):
    """direction: 'den' hoặc 'di'."""
    fname = TEMPLATE_FILES[(direction, "diff")]
    wb = load_workbook(os.path.join(TEMPLATE_DIR, fname))
    ws = wb.active

    rows_data = build_system_detail_rows(df_a, source_a, df_b, source_b)

    header_row = _find_row_by_col_value(ws, 2, "STT")
    if header_row is None:
        raise ValueError("Không tìm thấy dòng tiêu đề 'STT' trong biểu mẫu Chi tiết lệch")
    data_start = header_row + 1
    max_col = ws.max_column

    _ensure_data_rows(ws, data_start, 5, "Lập bảng", max_col, len(rows_data))

    for i, rec in enumerate(rows_data):
        r = data_start + i
        ws.cell(row=r, column=2, value=i + 1)                      # B: STT
        ws.cell(row=r, column=3, value=rec["Loại điện"])           # C: Loại điện
        ws.cell(row=r, column=5, value=rec["Số tham chiếu"])       # E: Số tham chiếu
        ws.cell(row=r, column=6, value=rec["Số SWIFT Seq"])        # F: Số SWIFT Seq
        ws.cell(row=r, column=9, value=rec["Số tiền"])              # I: Số tiền
        ws.cell(row=r, column=10, value=rec["Loại tiền"])          # J: Loại tiền
        ws.cell(row=r, column=12, value=rec["Ngân hàng gửi"])      # L: Ngân hàng gửi
        ws.cell(row=r, column=13, value=rec["Hệ thống"])           # M: Hệ thống
        ws.cell(row=r, column=14, value=rec["Ghi chú"])            # N: Ghi chú

    _update_date_cells(ws)
    wb.save(out_path)
