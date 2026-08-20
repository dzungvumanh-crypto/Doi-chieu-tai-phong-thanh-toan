# -*- coding: utf-8 -*-
"""
exporters.py
------------
Port NGUYÊN xuất Excel từ `citad-fixed/DoiSoatCITAD.py` (`export_doiSoat`,
`_add_filter_sheet`, `_style`, `_apply`, `HEADERS`, `CLR`) — giữ đúng màu,
label, độ rộng cột, freeze pane như bản gốc.

`STATUS_LBL` / `CLR_FG` được gộp thành hằng số dùng chung ở đây (bản gốc
định nghĩa lặp lại y hệt ở 3 nơi: export_doiSoat, _add_filter_sheet, và UI
tkinter) để cả export lẫn trang frontend `doi_soat_citad.py` dùng chung 1
nguồn — không đổi giá trị nào so với bản gốc.
"""
from __future__ import annotations

import datetime
from functools import lru_cache

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ['STT', 'Kết quả', 'Loại GD', 'Chiều', 'Số GD (CITAD)', 'Số GD (Agribank)',
           'Dịch vụ', 'Số tiền', 'Loại tiền', 'Ngày GD', 'Ngân hàng', 'Trạng thái',
           'Ghi chú']
# Cột 13 "Ghi chú" — thêm sau (không có trong bản gốc): số cổng CITAD của
# dòng chỉ có ở CITAD ('only_citad'/'lech_trang_thai', xem parsers.py::cong
# và reconcile.py — chỉ 2 status này giữ nguyên dict gốc từ citad_rows nên
# mới có field 'cong'). Rỗng ở dòng 'only_ipcas'/'only_hub' — hợp lý vì
# không phải lệnh từ CITAD, không có cổng nào để ghi.

CLR = {
    'title_bg': 'EFF6FF', 'title_fg': '1E3A5F',
    'hdr_dark': '1E3A5F', 'hdr_med': '2E5F9E',
    'red1': 'FFCCCC', 'red2': 'FFE5E5',
    'org1': 'FFF3CD', 'org2': 'FFF8E6',
    'white': 'FFFFFF', 'gray': 'F5F7FA',
    'red_txt': 'A32D2D', 'blue_txt': '185FA5',
    'green_txt': '3B6D11', 'org_txt': 'A05A00',
    # Riêng cho export_doiSoat_full() — dòng chênh lệch đẩy lên đầu bảng
    # "Tất cả lệnh" bôi VÀNG (khác đỏ/cam của bảng chi tiết chênh lệch
    # riêng) theo đúng yêu cầu, để nổi bật giữa phần lớn dòng đã khớp.
    'yellow1': 'FFF59D', 'yellow2': 'FFF9C4', 'yellow_txt': '7A5C00',
}

STATUS_LBL = {
    'only_citad': 'Chỉ CITAD', 'only_ipcas': 'Chỉ IPCAS',
    'only_hub': 'Chỉ Hub', 'lech_trang_thai': 'Lệch TT', 'both': 'Khớp',
    'dup_citad': 'Trùng CITAD',
}
CLR_FG = {
    'only_citad': CLR['red_txt'], 'only_ipcas': CLR['blue_txt'],
    'only_hub': CLR['blue_txt'], 'lech_trang_thai': CLR['org_txt'], 'both': CLR['green_txt'],
    'dup_citad': CLR['org_txt'],
}

_COL_WIDTHS = [5, 14, 8, 8, 18, 18, 28, 15, 10, 12, 28, 12, 12]
# Riêng cho export_doiSoat_full() — cột "Số tiền" ở export_doiSoat() hiếm khi
# quá dài (chỉ liệt kê phần lệch, số lượng nhỏ), nhưng "Tất cả lệnh" có cả
# ~38.000 dòng khớp, gặp số tiền cỡ vài trăm tỷ VNĐ (vd "307,195,253,272")
# — độ rộng 15 của _COL_WIDTHS bị hiện "####". Dàn rộng thêm các cột hay bị
# cắt (Số tiền, Dịch vụ, Ghi chú) để đọc đủ không cần tự kéo tay.
_COL_WIDTHS_FULL = [6, 14, 9, 9, 17, 17, 32, 20, 10, 12, 22, 12, 34]


def _ghi_chu(r):
    """Cột 13 — ưu tiên thông báo tường minh (vd. phát hiện dup) nếu có,
    không thì mới rơi về "Cổng X" mặc định (xem parsers.py::cong)."""
    if r.get('ghi_chu'):
        return r['ghi_chu']
    if r.get('cong'):
        return f"Cổng {r['cong']}"
    return ''


def _style(bg='FFFFFF', fg='000000', bold=False, sz=10, h='left'):
    thin = Side(style='thin', color='CCCCCC')
    return {
        'font': Font(name='Arial', size=sz, bold=bold, color=fg),
        'fill': PatternFill('solid', fgColor=bg),
        'alignment': Alignment(horizontal=h, vertical='center'),
        'border': Border(top=thin, bottom=thin, left=thin, right=thin),
    }


@lru_cache(maxsize=None)
def _style_cached(bg='FFFFFF', fg='000000', bold=False, sz=10, h='left'):
    """Như _style() nhưng nhớ lại object đã tạo cho cùng bộ tham số, dùng
    CHUNG 1 instance Font/Fill/... cho nhiều ô thay vì dựng mới từng ô một.
    Chỉ dùng ở export_doiSoat_full() — sheet ở đó có thể tới hàng chục nghìn
    dòng (mọi lệnh, không chỉ lệch), đo thực tế: dựng style mới cho từng ô
    kiểu _style() mất ~85 giây cho ~38.000 dòng x 13 cột — chiếm 1 trong 4
    slot run_heavy() dùng chung cả backend suốt hơn 1 phút mỗi lượt xuất,
    ảnh hưởng người khác đang thao tác việc nặng khác cùng lúc. Cache theo
    tham số làm giảm hẳn số lần openpyxl phải so khớp/thêm style mới vào
    bảng style dùng chung của workbook — export_doiSoat() (báo cáo CHỈ lệch,
    quy mô nhỏ vài trăm dòng) không cần đổi, giữ nguyên _style() như cũ."""
    return _style(bg=bg, fg=fg, bold=bold, sz=sz, h=h)


@lru_cache(maxsize=None)
def _align_cached(h='left'):
    """Chỉ set alignment (không font/fill/border) — dùng riêng cho dòng
    KHỚP ở export_doiSoat_full() để vẫn căn đúng theo cột mà không quay lại
    chi phí full style đã đo ở _style_cached(). Đo thực tế: 400.000 ô chỉ
    set alignment mất ~6 giây (so với ~18-30 giây nếu set đủ 4 thuộc tính,
    hay ~2 giây nếu không set gì) — đủ rẻ để áp cho cả phần dòng khớp mà
    tổng thời gian xuất vẫn ở mức vài giây, không quay lại ~1 phút."""
    return Alignment(horizontal=h, vertical='center')


def _apply(cell, **kw):
    for k, v in kw.items():
        setattr(cell, k, v)


def export_doiSoat(lech_rows, n_khop, ngay_cham, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Đối soát chi tiết'

    n_lech = len(lech_rows)
    n_total = n_khop + n_lech

    # ── Row 1: tiêu đề ────────────────────────────────────────────
    ws.merge_cells('A1:M1')
    c = ws['A1']
    c.value = f'KẾT QUẢ ĐỐI SOÁT LỆNH — CITAD (NHNN) vs AGRIBANK (IPCAS) — Ngày {ngay_cham}'
    _apply(c, **_style(bg=CLR['title_bg'], fg=CLR['title_fg'], bold=True, sz=13, h='center'))

    # ── Row 2: tổng kết ───────────────────────────────────────────
    ws.merge_cells('A2:M2')
    c = ws['A2']
    c.value = (f'Tổng: {n_total:,} lệnh   |   ✓ Khớp: {n_khop:,}   |   '
               f'✗ Lệch: {n_lech:,}   |   Xuất lúc: {datetime.datetime.now().strftime("%H:%M %d/%m/%Y")}')
    _apply(c, **_style(bg=CLR['title_bg'], fg='374151', sz=9, h='center'))

    # ── Row 3: trống ──────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: group headers ──────────────────────────────────────
    ws.merge_cells('A4:D4'); ws['A4'].value = 'THÔNG TIN CHUNG'
    ws.merge_cells('E4:H4'); ws['E4'].value = 'CITAD (NHNN)'
    ws.merge_cells('I4:L4'); ws['I4'].value = 'AGRIBANK (IPCAS)'
    ws['M4'].value = 'GHI CHÚ'  # cột mới, không thuộc nhóm nào — chỉ 1 cột, không cần merge
    for col, bg in [('A', CLR['hdr_dark']), ('E', CLR['hdr_dark']), ('I', CLR['hdr_med']), ('M', CLR['hdr_dark'])]:
        cell = ws[col + '4']
        _apply(cell, **_style(bg=bg, fg='FFFFFF', bold=True, sz=10, h='center'))
    # fill merged cells
    for rng, bg in [('B4:D4', CLR['hdr_dark']), ('F4:H4', CLR['hdr_dark']), ('J4:L4', CLR['hdr_med'])]:
        for cell in ws[rng][0]:
            _apply(cell, **_style(bg=bg, fg='FFFFFF', bold=True, sz=10, h='center'))

    # ── Row 5: column headers ─────────────────────────────────────
    for ci, h in enumerate(HEADERS, 1):
        bg = CLR['hdr_dark'] if ci <= 8 else CLR['hdr_med']
        cell = ws.cell(5, ci, h)
        _apply(cell, **_style(bg=bg, fg='FFFFFF', bold=True, sz=9, h='center'))

    # ── Data rows ─────────────────────────────────────────────────
    for ri, r in enumerate(lech_rows):
        row_num = ri + 6
        st = r.get('status', '')
        even = ri % 2 == 0
        if st in ('lech_trang_thai', 'dup_citad'):
            bg = CLR['org1'] if even else CLR['org2']
        elif st != 'both':
            bg = CLR['red1'] if even else CLR['red2']
        else:
            bg = CLR['white'] if even else CLR['gray']
        st_lbl = STATUS_LBL.get(st, st)
        st_fg = (CLR['red_txt'] if st == 'only_citad' else
                 CLR['blue_txt'] if st in ('only_ipcas', 'only_hub') else
                 CLR['org_txt'] if st in ('lech_trang_thai', 'dup_citad') else
                 CLR['green_txt'])

        so_tien = r.get('so_tien') or r.get('so_tien_agri') or ''
        try:
            so_tien = int(so_tien)
        except (ValueError, TypeError):
            so_tien = ''

        vals = [
            ri + 1, st_lbl,
            (r.get('loai') or '').upper(),
            'Đi' if r.get('chieu') == 'di' else 'Đến',
            r.get('so_gd') or '',
            r.get('key_agri') or '',
            r.get('dich_vu') or '',
            so_tien,
            r.get('loai_tien') or 'VNĐ',
            r.get('ngay') or '',
            r.get('nh_nhan') or '',
            r.get('trang_thai') or '',
            _ghi_chu(r),
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row_num, ci, val)
            h = 'right' if ci == 8 else ('center' if ci in (1, 2, 3, 4, 9, 13) else 'left')
            fg = st_fg if ci == 2 else '111827'
            bold = ci == 2
            _apply(cell, **_style(bg=bg, fg=fg, bold=bold, sz=10, h=h))
            if ci == 8 and isinstance(val, int):
                cell.number_format = '#,##0'

    # ── Độ rộng cột ───────────────────────────────────────────────
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze ────────────────────────────────────────────────────
    ws.freeze_panes = 'A6'

    # ── Sheet 2-5: các sheet lọc ─────────────────────────────────
    _add_filter_sheet(wb, lech_rows, n_khop, ngay_cham, 'Chỉ CITAD', ('only_citad', 'dup_citad'))
    _add_filter_sheet(wb, lech_rows, n_khop, ngay_cham, 'Chỉ Agribank', ('only_ipcas', 'only_hub'))
    _add_filter_sheet(wb, lech_rows, n_khop, ngay_cham, 'Lệch trạng thái', 'lech_trang_thai')

    wb.save(filepath)


def _wo_cell(ws, value, **style_kw):
    cell = WriteOnlyCell(ws, value=value)
    _apply(cell, **_style_cached(**style_kw))
    return cell


def export_doiSoat_full(lech_rows, khop_rows, ngay_cham, filepath):
    """Xuất "Tất cả lệnh" — khác export_doiSoat() ở chỗ liệt kê ĐỦ mọi
    lệnh (khớp lẫn lệch), không chỉ phần lệch. Mỗi dòng vẫn map cặp
    CITAD ↔ Agribank cùng cấu trúc cột HEADERS như export_doiSoat() —
    dễ đối chiếu qua lại giữa 2 file nếu cần. Lệch đẩy lên ĐẦU bảng,
    bôi VÀNG để nổi bật giữa phần lớn dòng đã khớp (yêu cầu Phòng Thanh
    toán) — khác màu đỏ/cam ở export_doiSoat(), vốn dùng cho báo cáo CHỈ
    có lệch nên không cần nổi bật giữa gì cả. Không có các sheet lọc
    theo trạng thái (2-5) — dành riêng cho export_doiSoat(), báo cáo này
    đã liệt kê hết trong 1 sheet duy nhất.

    Dùng `Workbook(write_only=True)` — khác export_doiSoat() (Workbook
    thường). Đo thực tế trên dữ liệu 1 ngày chấm thật (~38.000 dòng khớp):
    Workbook thường + tạo mới Font/Fill/... từng ô mất ~85 giây; đổi
    sang cache style (_style_cached) chỉ còn ~58 giây — vẫn chiếm 1 trong
    4 slot run_heavy() dùng chung cả backend gần 1 phút mỗi lượt xuất.
    write_only ghi từng dòng tuần tự (append), không cho quay lại sửa ô
    đã ghi và KHÔNG hỗ trợ merge_cells() (đã xác nhận bằng code thật, ném
    AttributeError) — nên phần tiêu đề (dòng 1-4) không merge ô như
    export_doiSoat() nữa, chỉ đặt nhãn ở ô đầu mỗi vùng, tô cùng màu nền
    cho cả dải để vẫn đọc được thành từng khối, không phải hy sinh gì về
    số liệu hay việc "map cặp" CITAD-Agribank mà chỉ đổi cách trình bày
    tiêu đề."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet('Tất cả lệnh')

    # PHẢI đặt độ rộng cột VÀ freeze_panes TRƯỚC lượt ws.append() đầu tiên —
    # write_only mode ghi các khối này ra trước `<sheetData>` theo luồng,
    # đặt sau khi đã append thì bị bỏ qua ÂM THẦM (không lỗi, chỉ đơn giản
    # không có tác dụng gì) — đã xác nhận thực tế bằng code thật cho cả 2
    # thuộc tính, không phải suy đoán.
    for i, w in enumerate(_COL_WIDTHS_FULL, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A6'

    n_khop = len(khop_rows)
    n_lech = len(lech_rows)
    n_total = n_khop + n_lech
    all_rows = lech_rows + khop_rows  # lệch trước — sắp lên đầu bảng

    n_cols = len(HEADERS)
    last_col = get_column_letter(n_cols)

    # Merge THẬT bằng `ws.merged_cells.add(...)` — khác lần trước tưởng
    # write_only không hỗ trợ (do thiếu tiện ích `ws.merge_cells()`), nhưng
    # thuộc tính `merged_cells` cấp thấp vẫn dùng được và đã xác nhận bằng
    # code thật: lưu/mở lại đúng, Excel hiển thị full nội dung + màu nền
    # trải khắp cả dải merge (chỉ cần style Ô ĐẦU TIÊN, các ô còn lại
    # trong dải Excel tự dùng định dạng ô đầu khi hiển thị, không cần style
    # riêng từng ô). Sửa lại từ cách cũ (chỉ đặt nhãn ở ô đầu, tô nền rời
    # rạc từng ô, KHÔNG merge thật) — cách cũ khiến chữ dài bị "che mất" vì
    # các ô sau tuy trống nhưng ĐÃ ĐƯỢC GHI (WriteOnlyCell) nên Excel không
    # tự tràn chữ sang được, khác ô thật sự chưa từng ghi.
    ws.append([_wo_cell(
        ws, f'TẤT CẢ LỆNH ĐỐI SOÁT — CITAD (NHNN) vs AGRIBANK (IPCAS) — Ngày {ngay_cham}',
        bg=CLR['title_bg'], fg=CLR['title_fg'], bold=True, sz=13, h='center',
    )])
    ws.merged_cells.add(f'A1:{last_col}1')

    ws.append([_wo_cell(
        ws,
        f'Tổng: {n_total:,} lệnh   |   ✓ Khớp: {n_khop:,}   |   '
        f'✗ Lệch (đẩy lên đầu, bôi vàng): {n_lech:,}   |   '
        f'Xuất lúc: {datetime.datetime.now().strftime("%H:%M %d/%m/%Y")}',
        bg=CLR['title_bg'], fg='374151', sz=9, h='center',
    )])
    ws.merged_cells.add(f'A2:{last_col}2')

    ws.append([])  # dòng 3 — trống, cách dòng (không ghi ô nào, khỏi cần style)

    # Dòng 4 — nhãn nhóm cột, merge thật A4:D4/E4:H4/I4:L4 (M4 đứng riêng,
    # đúng 4 nhóm khớp thứ tự cột trong HEADERS).
    ws.append([
        _wo_cell(ws, 'THÔNG TIN CHUNG', bg=CLR['hdr_dark'], fg='FFFFFF', bold=True, sz=10, h='center'),
        None, None, None,
        _wo_cell(ws, 'CITAD (NHNN)', bg=CLR['hdr_dark'], fg='FFFFFF', bold=True, sz=10, h='center'),
        None, None, None,
        _wo_cell(ws, 'AGRIBANK (IPCAS)', bg=CLR['hdr_med'], fg='FFFFFF', bold=True, sz=10, h='center'),
        None, None, None,
        _wo_cell(ws, 'GHI CHÚ', bg=CLR['hdr_dark'], fg='FFFFFF', bold=True, sz=10, h='center'),
    ])
    ws.merged_cells.add('A4:D4')
    ws.merged_cells.add('E4:H4')
    ws.merged_cells.add('I4:L4')

    ws.append([
        _wo_cell(ws, h, bg=(CLR['hdr_dark'] if ci <= 8 else CLR['hdr_med']), fg='FFFFFF', bold=True, sz=9, h='center')
        for ci, h in enumerate(HEADERS, 1)
    ])

    # Dòng KHỚP (đa số, có thể tới hàng chục nghìn) ghi GIÁ TRỊ THÔ, không
    # set font/fill/border/alignment riêng từng ô — đo thực tế (cProfile +
    # benchmark tách riêng): chi phí KHÔNG nằm ở việc tạo style object (đã
    # cache theo tham số) mà ở chính API `cell.font = ...`/`cell.fill = ...`
    # của openpyxl — mỗi lần gọi đều tính `__hash__`/`__eq__` để tra/thêm
    # vào bảng style dùng chung của workbook. 400.000 ô giá trị thuần ghi
    # trong write_only mode chỉ mất ~2 giây; cũng 400.000 ô có set đủ 4
    # style (kể cả TÁI DÙNG object) mất 18-30 giây. Với ~38.000 dòng khớp
    # thật, việc bỏ style riêng ô đưa export_doiSoat_full() từ ~60 giây
    # (chiếm 1 trong 4 slot run_heavy() dùng chung backend) xuống còn vài
    # giây. Dòng LỆCH (đẩy lên đầu, chỉ vài chục dòng) vẫn tô vàng đầy đủ
    # như yêu cầu — số dòng ít nên không đáng kể về tốc độ.
    for ri, r in enumerate(all_rows):
        st = r.get('status', '')
        is_lech = st != 'both'

        so_tien = r.get('so_tien') or r.get('so_tien_agri') or ''
        try:
            so_tien = int(so_tien)
        except (ValueError, TypeError):
            so_tien = ''

        vals = [
            ri + 1, STATUS_LBL.get(st, st),
            (r.get('loai') or '').upper(),
            'Đi' if r.get('chieu') == 'di' else 'Đến',
            r.get('so_gd') or '',
            r.get('key_agri') or '',
            r.get('dich_vu') or '',
            so_tien,
            r.get('loai_tien') or 'VNĐ',
            r.get('ngay') or '',
            r.get('nh_nhan') or '',
            r.get('trang_thai') or '',
            _ghi_chu(r),
        ]

        # Căn cột: chỉ STT (cột 1) căn giữa, MỌI cột còn lại căn trái — kể cả
        # Số tiền (không right-align theo lệ thường của cột số), theo đúng
        # yêu cầu Phòng Thanh toán.
        if is_lech:
            even = ri % 2 == 0
            bg = CLR['yellow1'] if even else CLR['yellow2']
            row_cells = []
            for ci, val in enumerate(vals, 1):
                h = 'center' if ci == 1 else 'left'
                fg = CLR['yellow_txt'] if ci == 2 else '111827'
                cell = _wo_cell(ws, val, bg=bg, fg=fg, bold=(ci == 2), sz=10, h=h)
                if ci == 8 and isinstance(val, int):
                    cell.number_format = '#,##0'
                row_cells.append(cell)
            ws.append(row_cells)
        else:
            # Dòng khớp: KHÔNG set font/fill/border (giữ tốc độ — xem
            # docstring hàm) nhưng vẫn set alignment cho MỌI ô (rẻ, xem
            # _align_cached()) để căn đúng như dòng lệch, đồng bộ cả bảng.
            row_cells = []
            for ci, val in enumerate(vals, 1):
                cell = WriteOnlyCell(ws, value=val)
                cell.alignment = _align_cached('center' if ci == 1 else 'left')
                if ci == 8 and isinstance(val, int):
                    cell.number_format = '#,##0'
                row_cells.append(cell)
            ws.append(row_cells)

    wb.save(filepath)


def _add_filter_sheet(wb, rows, n_khop, ngay_cham, title, status_filter):
    ws = wb.create_sheet(title)
    if isinstance(status_filter, tuple):
        filtered = [r for r in rows if r.get('status') in status_filter]
    else:
        filtered = [r for r in rows if r.get('status') == status_filter]

    ws.merge_cells('A1:M1')
    ws['A1'].value = f'{title} — Ngày {ngay_cham} — {len(filtered):,} lệnh'
    _apply(ws['A1'], **_style(bg=CLR['title_bg'], fg=CLR['title_fg'], bold=True, sz=12, h='center'))

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(2, ci, h)
        _apply(cell, **_style(bg=CLR['hdr_dark'], fg='FFFFFF', bold=True, sz=9, h='center'))

    for ri, r in enumerate(filtered):
        row_num = ri + 3
        st = r.get('status', '')
        bg = (CLR['org1'] if ri % 2 == 0 else CLR['org2']) if st in ('lech_trang_thai', 'dup_citad') else (
            CLR['red1'] if ri % 2 == 0 else CLR['red2']
        )
        so_tien = r.get('so_tien') or ''
        try:
            so_tien = int(so_tien)
        except Exception:
            so_tien = ''

        vals = [ri + 1, STATUS_LBL.get(st, st),
                (r.get('loai') or '').upper(),
                'Đi' if r.get('chieu') == 'di' else 'Đến',
                r.get('so_gd') or '', r.get('key_agri') or '',
                r.get('dich_vu') or '', so_tien,
                r.get('loai_tien') or 'VNĐ', r.get('ngay') or '',
                r.get('nh_nhan') or '', r.get('trang_thai') or '',
                _ghi_chu(r)]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row_num, ci, val)
            h = 'right' if ci == 8 else ('center' if ci in (1, 2, 3, 4, 9, 13) else 'left')
            fg = CLR_FG.get(st, '111827') if ci == 2 else '111827'
            _apply(cell, **_style(bg=bg, fg=fg, bold=(ci == 2), sz=10, h=h))
            if ci == 8 and isinstance(val, int):
                cell.number_format = '#,##0'

    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'
