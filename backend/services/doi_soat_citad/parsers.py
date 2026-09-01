# -*- coding: utf-8 -*-
"""
parsers.py
----------
Port NGUYÊN logic parse từ `citad-fixed/DoiSoatCITAD.py` (tool desktop độc
lập) — không đổi bất kỳ quy tắc nhận diện cột/ngày/trạng thái nào.

Sửa duy nhất khi port (không phải đổi logic nghiệp vụ): bản gốc giải nén
ZIP của CITAD vào `/tmp/_citad_<name>` — đường dẫn Unix không hợp lệ trên
Windows (nơi app thực tế chạy). Thay bằng `tempfile.mkdtemp()`, cùng cách
`backend/services/swift_recon/upload_utils.py` đã làm cho module Swift
Recon — chỉ là fix để chạy được, không ảnh hưởng kết quả đối soát.
"""
from __future__ import annotations

import csv
import datetime
import logging
import os
import re
import shutil
import tempfile
import zipfile

logger = logging.getLogger(__name__)

try:
    import xlrd
    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False


class UnknownFileFormat(Exception):
    pass


class _XlrdWs:
    """Wrapper xlrd sheet"""
    def __init__(self, ws):
        self._ws = ws
        self.nrows = ws.nrows
        self.ncols = ws.ncols

    def cell_value(self, row, col):
        try:
            v = self._ws.cell_value(row, col)
            return v if v is not None else ''
        except Exception as e:
            # Nuốt lỗi CÓ CHỦ Ý (parse không được phép chết vì 1 ô lỗi) —
            # nhưng phải log, không im lặng hoàn toàn: đúng lớp lỗi mà
            # docstring đầu file kể lại đã từng gây "mọi file .xlsx ra 0
            # dòng, không báo lỗi gì" (nguyên nhân cụ thể đã sửa, nhưng cơ
            # chế nuốt lỗi này có thể che giấu nguyên nhân KHÁC y hệt).
            logger.warning("Lỗi đọc ô CITAD (dòng %s, cột %s): %s", row, col, e)
            return ''


class _OpenpyxlWs:
    """Wrapper openpyxl sheet"""
    def __init__(self, ws):
        self._ws = ws
        self.nrows = ws.max_row or 0
        self.ncols = ws.max_column or 0

    def cell_value(self, row, col):
        try:
            v = self._ws.cell(row=row + 1, column=col + 1).value
            return v if v is not None else ''
        except Exception as e:
            # Xem ghi chú ở _XlrdWs.cell_value() — nuốt lỗi có chủ ý nhưng
            # phải log để không tái diễn lớp lỗi "0 dòng không rõ lý do".
            logger.warning("Lỗi đọc ô CITAD (dòng %s, cột %s): %s", row, col, e)
            return ''


# ──────────────────────────────────────────────────────────────
# PARSE CITAD
# ──────────────────────────────────────────────────────────────
def parse_citad_xls(filepath, ngay_cham=None):
    """Parse 1 file XLS/XLSX CITAD, trả về list lệnh.

    `ngay_cham` (dd/mm/yyyy, tuỳ chọn): mỗi file CITAD có đúng 1 dòng header
    "Ngày giao dịch: dd/mm/yyyy" áp dụng cho TOÀN BỘ file (không có cột ngày
    theo từng dòng) — nếu truyền `ngay_cham` và phát hiện được ngày của file
    KHÁC ngày chấm, bỏ qua CẢ FILE (trả rows rỗng + thông báo), để hỗ trợ
    quy trình thật: người dùng tải nhiều file CITAD của nhiều ngày cùng lúc,
    không cần tự lọc trước — xem thêm parse_citad_files()."""
    ext = os.path.splitext(filepath)[1].lower()
    is_openpyxl = ext == '.xlsx' or not _HAS_XLRD
    try:
        if is_openpyxl:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheets = [_OpenpyxlWs(wb[name]) for name in wb.sheetnames]
        else:
            wb = xlrd.open_workbook(filepath)
            sheets = [_XlrdWs(wb.sheet_by_index(i)) for i in range(wb.nsheets)]
    except Exception as e:
        return [], f"Lỗi đọc file: {e}"

    rows_out = []
    chieu_ref = [None]     # truyen chieu tu sheet dau sang cac sheet sau
    cong_ref = [None]      # truyen cong (tach tu header "Ngan hang:") sang cac sheet sau
    ngay_ref = [None]      # truyen ngay giao dich (tach tu header) sang cac sheet sau
    loai_tien_ref = [None]  # truyen loai_tien (VND/USD/EUR) tu sheet dau sang cac sheet sau —
                             # thieu truoc day (khac chieu_ref/cong_ref): sheet 2+ luon bi coi
                             # la VND du sheet dau la ngoai te, sai ca cot doc so tien lan nhan
    try:
        for shi, ws in enumerate(sheets):
            rows = _parse_sheet(ws, filepath, shi == 0, chieu_ref, cong_ref, ngay_ref, loai_tien_ref)
            rows_out += rows
    finally:
        # Đóng SAU khi đọc xong toàn bộ ô — read_only mode đọc trực tiếp
        # từ ZIP archive theo luồng, đóng sớm (trước khi đọc, như code cũ)
        # làm cell_value() ném ValueError, bị _OpenpyxlWs nuốt im lặng và
        # trả '' cho mọi ô — kết quả: mọi file .xlsx parse ra 0 dòng, không
        # báo lỗi. Đã tái hiện thực tế bằng openpyxl thật để xác nhận.
        if is_openpyxl:
            wb.close()

    if ngay_cham and ngay_ref[0] and ngay_ref[0] != ngay_cham:
        return [], f"Bỏ qua — file thuộc ngày {ngay_ref[0]}, khác ngày chấm {ngay_cham}"

    return rows_out, None


def _parse_sheet(ws, filepath, is_first, chieu_ref=None, cong_ref=None, ngay_ref=None, loai_tien_ref=None):
    # Detect chiều + loaiTien từ 12 dòng đầu (chỉ sheet đầu có header)
    chieu = 'di'
    loai_tien = 'VND'
    loai_file = 'il'
    cong = ''

    if is_first:
        for i in range(12):
            txt = ' '.join(str(ws.cell_value(i, j)) for j in range(ws.ncols)).lower()
            if 'chuyển tiền đến' in txt or 'báo cáo chuyển tiền đến' in txt:
                chieu = 'den'
            if 'chuyển tiền đi' in txt or 'báo cáo chuyển tiền đi' in txt:
                chieu = 'di'
            if 'đô la mỹ' in txt or ' usd' in txt:
                loai_tien = 'USD'
            if 'euro' in txt or ' eur' in txt:
                loai_tien = 'EUR'
            if 'giá trị cao' in txt:
                loai_file = 'ih'
            if 'giá trị thấp' in txt:
                loai_file = 'il'
            # Dòng "Ngân hàng: 01204001 - NH No&PTNT..." — 3 số cuối của mã
            # chính là số CỔNG CITAD (khớp đúng quy ước CONG_MAP dùng ở
            # module Đối chiếu: 01204001→cổng 1, 01204009→cổng 9,
            # 92204012→cổng 12, 79204017→cổng 17, 48204018→cổng 18 — đã
            # xác nhận thực tế trên 5 file mẫu, không phải suy đoán).
            m_cong = re.search(r'ngân hàng:\s*(\d{6,})', txt)
            if m_cong:
                try:
                    cong = str(int(m_cong.group(1)[-3:]))
                except ValueError:
                    pass
            # Dòng "Ngày giao dịch: dd/mm/yyyy" — áp dụng cho TOÀN BỘ file,
            # không phải theo từng dòng (CITAD không có cột ngày trên từng
            # dòng dữ liệu) — dùng để lọc cả file khi khác ngày chấm, xem
            # parse_citad_xls().
            m_ngay = re.search(r'ngày giao dịch:\s*(\d{1,2})/(\d{1,2})/(\d{4})', txt)
            if m_ngay and ngay_ref is not None:
                d, m, y = m_ngay.group(1).zfill(2), m_ngay.group(2).zfill(2), m_ngay.group(3)
                ngay_ref[0] = f'{d}/{m}/{y}'
        if chieu_ref is not None:
            chieu_ref[0] = chieu  # luu chieu de sheet sau dung
        if cong_ref is not None:
            cong_ref[0] = cong  # luu cong de sheet sau dung
        if loai_tien_ref is not None:
            loai_tien_ref[0] = loai_tien  # luu loai_tien de sheet sau dung
    else:
        if chieu_ref is not None and chieu_ref[0] is not None:
            chieu = chieu_ref[0]  # sheet sau ke thua chieu tu sheet dau
        if cong_ref is not None and cong_ref[0]:
            cong = cong_ref[0]  # sheet sau ke thua cong tu sheet dau
        if loai_tien_ref is not None and loai_tien_ref[0] is not None:
            loai_tien = loai_tien_ref[0]  # sheet sau ke thua loai_tien tu sheet dau —
                                            # truoc day luon roi ve mac dinh 'VND' o day

    # Tìm header row
    h_row_idx = -1
    i_so_gd = 2
    i_dich_vu = 5
    i_no_den = 16 if loai_tien == 'VND' else 15
    i_co_di = 24 if loai_tien == 'VND' else 23
    i_ngay = 0

    if is_first:
        for i in range(min(20, ws.nrows)):
            row_txt = ' '.join(str(ws.cell_value(i, j)) for j in range(ws.ncols)).lower()
            if 'số gd' in row_txt or 'số giao dịch' in row_txt:
                h_row_idx = i
                for j in range(ws.ncols):
                    c = str(ws.cell_value(i, j)).strip().lower()
                    if c in ('số gd', 'số giao dịch'):
                        i_so_gd = j
                    if c == 'dịch vụ':
                        i_dich_vu = j
            if 'nợ' in row_txt or 'có' in row_txt:
                for j in range(ws.ncols):
                    c = str(ws.cell_value(i, j)).strip().lower()
                    if c == 'nợ':
                        i_no_den = j
                    if c == 'có':
                        i_co_di = j
            if h_row_idx >= 0 and (i_no_den >= 0 or i_co_di >= 0):
                break

    data_start = (h_row_idx + 3) if (is_first and h_row_idx >= 0) else 0

    # Verify cột tiền: scan data để tìm cột đúng (header có thể lệch 1)
    def _verify_col(col_idx, max_scan=30):
        checked = 0
        for vi in range(data_start, ws.nrows):
            sogd = ''.join(
                c for c in str(ws.cell_value(vi, i_so_gd)).replace("'", "").replace(".0", "") if c.isdigit()
            )
            if len(sogd) < 6:
                continue
            checked += 1
            v = str(ws.cell_value(vi, col_idx)).strip().replace("'", "")
            if v and v not in ('0', '0.0', 'nan', ''):
                if re.search(r'[1-9]', v):
                    return True
            if checked >= max_scan:
                break
        return False

    # File Đi: verify iCoDi (tiền luôn ở Có)
    # File Đến: verify iNoDen (Chuyển có chiếm đa số đầu file)
    if chieu == 'di':
        for d in [0, -1, 1, -2, 2]:
            if 0 <= i_co_di + d < ws.ncols and _verify_col(i_co_di + d):
                i_co_di = i_co_di + d
                break
    else:
        for d in [0, -1, 1, -2, 2]:
            if 0 <= i_no_den + d < ws.ncols and _verify_col(i_no_den + d):
                i_no_den = i_no_den + d
                break

    result = []
    for i in range(data_start, ws.nrows):
        so_gd_raw = str(ws.cell_value(i, i_so_gd)).strip().replace("'", "").replace(".0", "")
        so_gd = ''.join(c for c in so_gd_raw if c.isdigit())
        if len(so_gd) < 6:
            continue

        # Đọc dịch vụ từ col i_dich_vu, fallback col-2
        dich_vu = str(ws.cell_value(i, i_dich_vu)).strip().replace("'", "")
        if not dich_vu and i_dich_vu >= 2:
            dich_vu = str(ws.cell_value(i, i_dich_vu - 2)).strip().replace("'", "")

        dv_low = dich_vu.lower()

        # Cột tiền theo chiều file và dịch vụ:
        # File Đi:  tiền luôn ở cột Có (i_co_di)
        # File Đến: "Chuyển có" → cột Nợ, "Chuyển nợ" → cột Có
        if chieu == 'di':
            col_tien = i_co_di
        elif 'chuyển nợ' in dv_low:
            col_tien = i_co_di
        else:
            col_tien = i_no_den

        tien_raw = str(ws.cell_value(i, col_tien)).strip().replace("'", "")
        so_tien = _parse_so_tien(tien_raw)
        if not so_tien:
            continue

        ngay = str(ws.cell_value(i, i_ngay)).strip()
        loai = 'ih' if 'giá trị cao' in dv_low else loai_file
        dich_vu_final = dich_vu or ('Chuyển có giá trị cao' if loai == 'ih' else 'Chuyển có giá trị thấp')

        result.append({
            'so_gd': so_gd,
            'dich_vu': dich_vu_final,
            'loai': loai,
            'chieu': chieu,
            'loai_tien': loai_tien,
            'so_tien': so_tien,
            'ngay': ngay,
            'cong': cong,
        })
    return result


def parse_citad_files(filepaths, ngay_cham=None, progress_cb=None):
    """Parse nhiều file CITAD (XLS hoặc ZIP chứa XLS).

    Khác bản gốc: giải nén ZIP vào thư mục tạm qua `tempfile.mkdtemp()`
    (bản gốc dùng `/tmp/_citad_<name>` — không hợp lệ trên Windows) và tự
    dọn sạch thư mục tạm sau khi parse xong. Không đổi cách nhận diện dòng.

    `ngay_cham` (tuỳ chọn): truyền xuống `parse_citad_xls()` để tự bỏ qua
    file thuộc ngày khác — hỗ trợ quy trình thật: người dùng tải chung file
    CITAD của nhiều ngày trước/sau ngày chấm (lệnh lập ngày này nhưng đi
    kênh ngày khác), công cụ tự lọc đúng ngày ĐI KÊNH thay vì bắt tự lọc
    tay. Xem docstring `parse_citad_xls()`.
    """
    all_rows = []
    errors = []
    for fp in filepaths:
        ext = os.path.splitext(fp)[1].lower()
        if ext == '.zip':
            extract_dir = tempfile.mkdtemp(prefix='citad_soat_')
            try:
                with zipfile.ZipFile(fp) as z:
                    for name in z.namelist():
                        if name.lower().endswith('.xls') or name.lower().endswith('.xlsx'):
                            tmp = os.path.join(extract_dir, os.path.basename(name))
                            with z.open(name) as src, open(tmp, 'wb') as dst:
                                dst.write(src.read())
                            rows, err = parse_citad_xls(tmp, ngay_cham)
                            all_rows += rows
                            if err:
                                errors.append(f"{name}: {err}")
            except Exception as e:
                errors.append(f"{os.path.basename(fp)}: {e}")
            finally:
                shutil.rmtree(extract_dir, ignore_errors=True)
        elif ext in ('.xls', '.xlsx'):
            rows, err = parse_citad_xls(fp, ngay_cham)
            all_rows += rows
            if err:
                errors.append(f"{os.path.basename(fp)}: {err}")
        else:
            errors.append(f"{os.path.basename(fp)}: định dạng không hỗ trợ (cần XLS/XLSX hoặc ZIP)")
    return all_rows, errors


# ──────────────────────────────────────────────────────────────
# PARSE IPCAS
# ──────────────────────────────────────────────────────────────
def _strip_apos(s):
    return s.lstrip("'").strip() if s else ''


def _parse_so_tien(raw) -> int:
    """Đọc số tiền dạng chuỗi/số — chịu được cả định dạng thường
    ("553,722,000,000") lẫn định dạng KHOA HỌC ("5.53722E+11").

    Bug thật (xác nhận 25/08/2026, Phòng Thanh toán tự phát hiện): GDV mở
    file CSV IPCAS trong Excel để xoá thử 1 dòng rồi lưu lại — Excel TỰ ĐỘNG
    đổi mọi số tiền ĐỦ LỚN (nhóm "cao"/IH — hàng trăm tỷ trở lên) sang định
    dạng khoa học khi lưu CSV, số nhỏ (nhóm "thấp"/IL) không đủ lớn nên
    không bị đổi — đúng khớp quan sát thực tế "chỉ cao mới lỗi, thấp thì
    không". Cách đọc cũ (xoá mọi ký tự không phải chữ số) xử lý SAI hoàn
    toàn với dạng khoa học: "5.53722E+11" bị xoá mất dấu chấm/E/dấu cộng,
    ghép chữ số còn lại thành "55372211" — SAI HẲN so với giá trị thật
    553.722.000.000, và cái đuôi "11" chính là số mũ "E+11" dính vào. Không
    phải lỗi do việc xoá dòng — chỉ cần Excel lưu lại CSV có cột số tiền lớn
    là dính, xoá dòng chỉ là thao tác tình cờ kích hoạt Excel lưu lại file.

    CHỈ bắt theo dấu hiệu 'E'/'e' của khoa học — KHÔNG bắt theo dấu chấm
    nói chung. Bug thật khi sửa lần đầu (tự phát hiện ngay khi kiểm lại):
    số tiền CITAD dùng DẤU CHẤM làm dấu phân cách HÀNG NGHÌN kiểu Việt Nam
    (vd "252.121.572" = 252.121.572 đồng, không phải số thập phân
    252,121572) — bắt theo cả dấu chấm sẽ hiểu "790.840" (790.840 đồng)
    thành số thập phân 790,84 rồi làm tròn ra 791, sai gấp cả nghìn lần.
    Dấu chấm nhiều lần ("252.121.572") thì `float()` tự ném lỗi nên vô
    tình không sao — nhưng đúng 1 dấu chấm ("790.840") thì `float()` CHẤP
    NHẬN được, âm thầm ra số sai. Định dạng khoa học không bao giờ có 'E'
    trong số Việt Nam nên chỉ bắt theo 'E'/'e' là an toàn, không đụng gì
    tới cách đọc số CITAD."""
    if raw is None:
        return 0
    s = str(raw).strip()
    if not s:
        return 0
    cleaned = s.replace(',', '').replace("'", '').replace(' ', '')
    if re.search(r'[Ee]', cleaned):
        try:
            return int(round(float(cleaned)))
        except (ValueError, TypeError):
            pass
    digits = re.sub(r'[^0-9]', '', s)
    return int(digits) if digits else 0


# Chuẩn hoá 1 ô "ngày" đọc từ Excel về "dd/mm/yyyy" để so được với
# `ngay_cham` (luôn dd/mm/yyyy, nhập từ form) — ô ngày có thể ở 3 dạng tuỳ
# nguồn file: (1) chuỗi "dd/mm/yyyy" sẵn có (IPCAS CSV), (2) object
# datetime/date thật (openpyxl với ô định dạng Date), (3) số serial Excel
# (xlrd không tự quy đổi ô Date, và cả 2 khi đọc qua text). Thiếu bước này,
# `str(cell_value)[:10]` của ô kiểu Date ra "yyyy-mm-dd" — KHÔNG BAO GIỜ
# khớp "dd/mm/yyyy", lọc rớt âm thầm toàn bộ dòng của file (xem
# _parse_hub_xls) mà không báo lỗi gì.
def _normalize_date_cell(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d/%m/%Y')
    s = str(v).strip() if v is not None else ''
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f'{d}/{mo}/{y}'
    try:
        serial = float(s)
        if 40000 < serial < 60000:  # khoảng năm ~2009-2064, loại số khác bị hiểu nhầm
            base = datetime.date(1899, 12, 30)
            return (base + datetime.timedelta(days=int(serial))).strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        pass
    # Dạng "yyyymmdd" liền không dấu phân cách (vd cột "Ngày GD" của file Hub
    # PaymentHub, khác với "Ngày quyết toán"/"Ngày hạch toán" dd/mm/yyyy cùng
    # file) — xác nhận thực tế trên file Danh_sach_giao_dich_den thật, KHÔNG
    # phải suy đoán: thiếu bước này khiến _parse_hub_xls() lọc rớt ÂM THẦM
    # toàn bộ dòng của cột này khi có ngay_cham (không bao giờ khớp chuỗi
    # "20260819" != "19/08/2026"), y hệt lớp lỗi ô kiểu Date mô tả ở trên.
    m2 = re.match(r'^(\d{4})(\d{2})(\d{2})$', s)
    if m2:
        y, mo, d = m2.group(1), m2.group(2), m2.group(3)
        if 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            return f'{d}/{mo}/{y}'
    return s[:10]  # không nhận diện được — giữ hành vi cũ, không đoán mò thêm


def _parse_ipcas_text(text, filename, ngay_cham):
    """Parse nội dung text CSV của IPCAS"""
    text = text.lstrip('﻿')
    lines = text.splitlines()
    if not lines:
        return []

    # Build header map
    header = lines[0]
    cols = [c.strip().lstrip("'").lstrip('﻿') for c in header.split(',')]
    hmap = {name: i for i, name in enumerate(cols)}
    n_headers = len(cols)
    has_nkt = 'NGAY_KENH_TRA' in hmap

    # Detect chiều từ 50 dòng đầu
    sample = ' '.join(lines[1:51]).upper()
    n_scnl = sample.count('SCNL')
    n_den = sum(sample.count(s) for s in ['PYED', 'PYEK', 'SBFL', 'SBSC'])
    fname = (filename or '').lower()
    if n_scnl > 0 and n_scnl >= n_den:
        chieu = 'di'
    elif n_den > 0 and n_den > n_scnl:
        chieu = 'den'
    elif 'den' in fname or '_den_' in fname:
        chieu = 'den'
    else:
        chieu = 'di'

    i_nh = hmap.get('NH_NHAN', hmap.get('NH_GUI', -1))
    i_nkt = hmap.get('NGAY_KENH_TRA', -1)

    # Trước đây có 1 bước lọc bỏ ÂM THẦM dòng trùng y hệt (cùng ngày/chi
    # nhánh/txid/số tiền/trace/trạng thái) ngay ở đây — xác nhận thật: IPCAS
    # có thể hạch toán CÙNG 1 lệnh nhiều lần (nghi ngờ lỗi hệ thống lõi),
    # trong khi CITAD chỉ nhận đúng 1 lần; bỏ ngay lúc đọc file khiến
    # reconcile.py không bao giờ thấy để phát hiện, đối soát báo "khớp" bình
    # thường như không có gì. Đã BỎ HẲN bước lọc này — giữ nguyên MỌI dòng
    # kể cả trùng y hệt, để reconcile.py::run_doiSoat_ram() tự đếm số lần
    # trùng theo đúng khoá khớp lệnh (msgref/txid+loai+so_tien) và sinh
    # đúng số dòng "Chỉ Agribank" phản ánh số dư thật — xem ghi chú ở đó.
    result = []

    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue

        # Parse CSV có thể có quoted fields
        try:
            vals = list(csv.reader([line]))[0]
        except Exception:
            vals = line.split(',')

        if len(vals) < 8:
            continue

        # Tính shift do NH_NHAN chứa dấu phẩy
        shift = max(0, len(vals) - n_headers)

        def gv(name):
            idx = hmap.get(name, -1)
            if idx < 0:
                return ''
            if i_nh >= 0 and idx > i_nh:
                idx += shift
            if idx >= len(vals):
                return ''
            return _strip_apos(vals[idx])

        # NGAY_KENH_TRA: đọc qua gv() (không dùng thẳng vals[i_nkt] như trước
        # — thiếu cộng `shift` khi NH_NHAN chứa dấu phẩy làm lệch cột, đọc
        # sai ngày cho các dòng đó; gv() đã xử lý đúng shift này cho mọi
        # trường khác, dùng chung cho nkt luôn thay vì lặp lại logic).
        # Không dùng vals[-1] vì NOI_DUNG có dấu ;.
        nkt = ''
        if i_nkt >= 0:
            last = gv('NGAY_KENH_TRA').strip()
            # Dang chuan: 4/6/2026 hoac 04/06/2026
            m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', last)
            if m:
                d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
                nkt = f'{d}/{mo}/{y}'
            else:
                # Excel serial date (vi du: 46011.34583)
                try:
                    serial = float(last)
                    if 40000 < serial < 60000:
                        base = datetime.date(1899, 12, 30)
                        d_obj = base + datetime.timedelta(days=int(serial))
                        nkt = d_obj.strftime('%d/%m/%Y')
                except (ValueError, TypeError):
                    pass

        tt = gv('TRANG_THAI_LENH')
        ngay_gd_raw = gv('NGAY_GIAO_DICH').strip()
        # Chuan hoa ngay_gd: 4/6/2026 -> 04/06/2026
        _m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', ngay_gd_raw)
        ngay_gd = f'{_m.group(1).zfill(2)}/{_m.group(2).zfill(2)}/{_m.group(3)}' if _m else ngay_gd_raw[:10]
        so_tien = _parse_so_tien(gv('SO_TIEN'))
        kenh = gv('KENH_THANH_TOAN').lower()

        # Filter theo chiều
        if chieu == 'di':
            # Giu: SCNL (thanh cong) + WFPG/SBFL/RFED/SDEB/SBSC/RTSC (dang xu
            # ly, co the lech TT) + ERPO/CALD (IPCAS bao that bai/huy - GIU
            # lai de doi chieu voi CITAD: neu CITAD VAN CO lenh nay thi la
            # bat thuong that su (IPCAS sai, lenh da di kenh thanh cong) ->
            # 'lech_trang_thai', con neu CITAD khong co thi la that bai binh
            # thuong (chua tung di kenh) -> reconcile.py tu bo qua o vong lap
            # "IPCAS Di du", khong tinh vao "Chi Agribank".
            # Bo: rong
            KEEP_DI = {'SCNL', 'WFPG', 'SBFL', 'RFED', 'SDEB', 'SBSC', 'RTSC', 'ERPO', 'CALD'}
            if tt not in KEEP_DI:
                continue
            # Yêu cầu Phòng Thanh toán 27/08/2026: SCNL báo lệnh đã sang kênh
            # thành công, nhưng NGAY_KENH_TRA vẫn trống nghĩa là kênh CHƯA
            # THỰC SỰ xác nhận ngày trả — giữ nguyên coi là khớp (VALID_DI ở
            # reconcile.py) sẽ khớp "khống" với CITAD dù chưa có xác nhận
            # thật. Bỏ khỏi kết quả IPCAS (không chỉ khỏi diện SCNL=khớp) để
            # lệnh CITAD tương ứng (nếu có) rơi đúng vào "Chỉ CITAD" — cần
            # người dùng tự xác minh, không tự động khớp.
            if has_nkt and tt == 'SCNL' and not nkt:
                continue
            if ngay_cham:
                if has_nkt:
                    if nkt and nkt != ngay_cham:
                        # nkt co gia tri nhung khac ngay cham -> bo
                        continue
                    elif not nkt:
                        # nkt trong (cot bi lech hoac chua xu ly) -> fallback ngay_gd
                        if ngay_gd and ngay_gd != ngay_cham:
                            continue
                else:
                    if ngay_gd and ngay_gd != ngay_cham:
                        continue
        else:
            # Chiều Đến: KHÔNG lọc bỏ theo trạng thái nào cả — CITAD ghi
            # nhận hết kể cả SBSC/SBFL/RFED, chỉ lọc theo ngày.
            if ngay_cham and ngay_gd and ngay_gd != ngay_cham:
                continue

        txid_raw = _strip_apos(gv('TXID'))
        if chieu == 'den':
            txid = re.sub(r'[^0-9]', '', txid_raw.split('-')[0].split()[0])
        else:
            txid = txid_raw

        msgref = _strip_apos(gv('MSGREF'))
        nh_nhan = gv('NH_NHAN') or gv('NH_GUI')
        loai = 'ih' if 'cao' in kenh else 'il'

        result.append({
            'txid': txid,
            'msgref': msgref,
            'loai': loai,
            'chieu': chieu,
            'so_tien': so_tien,
            'trang_thai': tt,
            'nkt': nkt,
            'kenh': gv('KENH_THANH_TOAN'),
            'nh_nhan': nh_nhan,
            'ngay': ngay_gd,
            # `chi_nhanh` — CHỈ để reconcile.py so khớp "có phải cùng 1 bản
            # ghi IPCAS bị lặp lại y hệt hay không" (khoá mịn), KHÔNG dùng để
            # khớp lệnh với CITAD (khoá khớp lệnh vẫn là txid/msgref+loai+
            # so_tien như cũ, xem run_doiSoat_ram()). Cần field này vì
            # (txid, loai, so_tien) là khoá CỐ Ý làm thô để khớp lệnh linh
            # hoạt — IPCAS xác nhận thật dùng CHUNG 1 txid cho nhiều lệnh
            # KHÁC NHAU (khác nh_nhan), nên không thể dùng đúng khoá đó để
            # kết luận "IPCAS ghi trùng" (regression thật 23/08/2026: 2 lệnh
            # Đến khác ngân hàng nhận, trùng ngẫu nhiên txid+loai+so_tien, bị
            # hiểu nhầm hàng loạt thành hạch toán trùng — xem ghi chú
            # `_ghi_chu_khop_du_nguon` trong reconcile.py).
            #
            # `trace` giữ lại để tham khảo/audit (không tốn gì thêm, đã đọc
            # sẵn cột này) nhưng KHÔNG đưa vào khoá mịn ở reconcile.py —
            # xác nhận thật (23/08/2026, cùng ngày): 3 dòng cùng 1 lệnh
            # trùng lặp có thể mang 2-3 giá trị trace KHÁC NHAU (IPCAS cấp
            # trace mới mỗi lần ghi sổ) dù mọi trường khác giống hệt — đưa
            # trace vào khoá làm dòng trùng thứ 3 rơi ra khoá riêng, biến
            # mất khỏi báo cáo hoàn toàn (không khớp, không "Chỉ IPCAS").
            'chi_nhanh': gv('CHI_NHANH'),
            'trace': gv('TRACE'),
            # `refhub` — mã tham chiếu gốc của điện đến, phải DUY NHẤT cho
            # mỗi bản ghi thật (khác hẳn txid — có thể trùng giữa nhiều lệnh
            # không liên quan, xem ghi chú `chi_nhanh` ở trên). Dùng để xác
            # nhận chắc chắn "cùng 1 lệnh gốc" khi phát hiện hạch toán nhầm
            # rồi huỷ (xem reconcile.py — GDV xác nhận 23/08/2026).
            'refhub': gv('REFHUB'),
        })
    return result


def parse_ipcas_files(filepaths, ngay_cham, progress_cb=None):
    """Parse nhiều file IPCAS (CSV hoặc ZIP chứa CSV)"""
    all_rows = []
    errors = []

    for fp in filepaths:
        ext = os.path.splitext(fp)[1].lower()
        fname = os.path.basename(fp)
        if progress_cb:
            progress_cb(f'Đọc IPCAS: {fname}')

        if ext == '.zip':
            try:
                with zipfile.ZipFile(fp) as z:
                    for name in z.namelist():
                        n_ext = os.path.splitext(name)[1].lower()
                        if n_ext in ('.csv', '.txt'):
                            raw = z.read(name)
                            text = None
                            for enc in ('utf-8-sig', 'utf-8', 'cp1258', 'latin-1'):
                                try:
                                    text = raw.decode(enc)
                                    break
                                except Exception:
                                    continue  # 'latin-1' ở cuối luôn thành công (không raise)
                            rows = _parse_ipcas_text(text, name, ngay_cham)
                            all_rows += rows
            except Exception as e:
                errors.append(f"{fname}: {e}")

        elif ext in ('.csv', '.txt'):
            try:
                raw = open(fp, 'rb').read()
                text = None
                for enc in ('utf-8-sig', 'utf-8', 'cp1258', 'latin-1'):
                    try:
                        text = raw.decode(enc)
                        break
                    except Exception:
                        continue  # 'latin-1' ở cuối luôn thành công (không raise)
                rows = _parse_ipcas_text(text, fname, ngay_cham)
                all_rows += rows
            except Exception as e:
                errors.append(f"{fname}: {e}")

        else:
            errors.append(f"{fname}: định dạng không hỗ trợ (cần CSV hoặc ZIP)")

    return all_rows, errors


# ──────────────────────────────────────────────────────────────
# PARSE HUB NGOẠI TỆ
# ──────────────────────────────────────────────────────────────
def parse_hub_files(filepaths, ngay_cham):
    """Parse file Hub ngoại tệ (XLS/XLSX) -> list of dicts"""
    all_rows = []
    errors = []
    for fp in filepaths:
        ext = os.path.splitext(fp)[1].lower()
        fname = os.path.basename(fp).lower()
        try:
            rows = _parse_hub_xls(fp, ext, fname, ngay_cham)
            all_rows += rows
        except Exception as e:
            errors.append(f"{os.path.basename(fp)}: {e}")
    return all_rows, errors


def _parse_hub_xls(filepath, ext, fname, ngay_cham):
    """Parse 1 file Hub XLS/XLSX"""
    if ext == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)  # khong read_only de doc dung max_row
        sheets = [(name, _OpenpyxlWs(wb[name])) for name in wb.sheetnames]
        wb.close()
    else:
        if not _HAS_XLRD:
            raise RuntimeError(
                f"Không đọc được file .xls '{fname}': thiếu thư viện xlrd "
                "(chỉ hỗ trợ .xlsx nếu không có xlrd)."
            )
        wb = xlrd.open_workbook(filepath)
        sheets = [(wb.sheet_by_index(i).name, _XlrdWs(wb.sheet_by_index(i)))
                  for i in range(wb.nsheets)]

    # Detect chiều từ tên file
    chieu_file = 'den' if ('den' in fname or 'đến' in fname) else 'di'

    result = []
    for sheet_name, ws in sheets:
        chieu = chieu_file
        if 'den' in sheet_name.lower() or 'đến' in sheet_name.lower():
            chieu = 'den'
        elif 'di' in sheet_name.lower():
            chieu = 'di'

        # Tìm header row
        h_row = -1
        i_so_tc = -1   # Số thành công / Số GD
        i_so_tien = -1
        i_loai_tien = -1
        i_ngay = -1
        i_nh = -1
        i_trang_thai = -1

        for i in range(min(10, ws.nrows)):
            row_txt = ' '.join(str(ws.cell_value(i, j)).lower() for j in range(ws.ncols))
            if any(k in row_txt for k in ['số thành công', 'so thanh cong', 'số giao dịch', 'msgid', 'msg_key']):
                # Dòng ứng viên — có thể là dòng tiêu đề THẬT, nhưng cũng có thể
                # là dòng tổng kết đầu file kiểu "Tổng số giao dịch:12" (cũng
                # chứa cụm "số giao dịch" nên khớp nhầm điều kiện trên). Quét
                # cột vào biến CỤC BỘ trước — chỉ nhận dòng này làm h_row nếu
                # thực sự tìm được cột i_so_tc, nếu không thì bỏ qua và quét
                # tiếp dòng sau thay vì `break` sớm rồi bỏ luôn cả sheet (bug
                # thật đã xác nhận: dòng tổng kết khiến sheet bị bỏ 100%, mất
                # trắng dữ liệu Napas/PSS-MDP mà không báo lỗi gì).
                cand_so_tc = -1
                cand_so_tien = -1
                cand_loai_tien = -1
                cand_ngay = -1
                cand_nh = -1
                cand_trang_thai = -1
                for j in range(ws.ncols):
                    hdr = str(ws.cell_value(i, j)).strip().lower()
                    # Uu tien 'so thanh cong' truoc, sau moi den msgid
                    if any(k in hdr for k in ['thành công', 'so tc', 'số tc']):
                        cand_so_tc = j
                    elif any(k in hdr for k in ['msgid', 'msg_key']) and cand_so_tc < 0:
                        cand_so_tc = j
                    if 'số tiền' in hdr or 'so tien' in hdr or 'amount' in hdr:
                        cand_so_tien = j
                    if 'loại tiền' in hdr or 'currency' in hdr or 'loai tien' in hdr:
                        cand_loai_tien = j
                    if 'kênh trả' in hdr or 'kenh tra' in hdr:
                        cand_ngay = j  # uu tien ngay kenh tra
                    elif 'ngày' in hdr and ('nhận' in hdr or 'gd' in hdr or 'giao' in hdr) and cand_ngay < 0:
                        cand_ngay = j  # fallback ngay nhan
                    if 'ngân hàng' in hdr or 'nh ' in hdr:
                        cand_nh = j
                    if 'trạng thái' in hdr or 'trang thai' in hdr:
                        cand_trang_thai = j
                if cand_so_tc < 0:
                    continue  # dòng khớp từ khoá nhưng không có cột thật — thử dòng sau
                h_row, i_so_tc, i_so_tien = i, cand_so_tc, cand_so_tien
                i_loai_tien, i_ngay, i_nh = cand_loai_tien, cand_ngay, cand_nh
                i_trang_thai = cand_trang_thai
                break

        if h_row < 0 or i_so_tc < 0:
            continue

        data_start = h_row + 1
        for i in range(data_start, ws.nrows):
            so_tc = str(ws.cell_value(i, i_so_tc)).strip().replace("'", "")
            # Chỉ giữ số — khớp đúng cách so_gd bên CITAD được làm sạch
            # (parse_citad_xls(), chỉ isdigit()). Trước đây giữ cả chữ
            # (isalpha()) khiến khoá 2 bên KHÔNG BAO GIỜ khớp được nếu MSGID
            # Hub có lẫn chữ cái — báo nhầm "chỉ CITAD"/"chỉ Hub" dù cùng 1
            # lệnh thật.
            so_tc_clean = ''.join(ch for ch in so_tc if ch.isdigit())
            if not so_tc_clean or len(so_tc_clean) < 4:
                continue

            # Số tiền
            tien_raw = str(ws.cell_value(i, i_so_tien)).strip() if i_so_tien >= 0 else ''
            so_tien = _parse_so_tien(tien_raw)
            if not so_tien:
                continue

            # Loại tiền
            loai_tien = 'USD'
            if i_loai_tien >= 0:
                lt = str(ws.cell_value(i, i_loai_tien)).strip().upper()
                if lt in ('USD', 'EUR', 'GBP', 'JPY', 'CNY'):
                    loai_tien = lt

            # Ngày — cell_value() có thể trả object datetime thật (openpyxl,
            # ô định dạng Date) hoặc số serial (xlrd), không phải chuỗi
            # "dd/mm/yyyy" — phải chuẩn hoá mới so được với ngay_cham, xem
            # _normalize_date_cell().
            ngay = ''
            if i_ngay >= 0:
                ngay = _normalize_date_cell(ws.cell_value(i, i_ngay))

            # Lọc ngày chấm
            if ngay_cham and ngay and ngay != ngay_cham:
                continue

            nh = str(ws.cell_value(i, i_nh)).strip() if i_nh >= 0 else ''
            # Trạng thái — CHỈ để reconcile.py chọn đúng "dòng gốc" khi 1 lệnh
            # chuyển chi nhánh sinh nhiều dòng cùng Số thành công (khác chi
            # nhánh) — xác nhận nghiệp vụ 23/08/2026 (Phòng Thanh toán), cùng
            # hiện tượng đã xử lý cho IPCAS (xem PRIORITY_TT/CGBR trong
            # reconcile.py): "Đã trả KH" là dòng gốc, các trạng thái khác là
            # dòng con (chi nhánh trung gian). KHÔNG dùng để lọc/loại bỏ dòng
            # nào ở đây — chỉ mang theo để chọn ưu tiên lúc khớp lệnh.
            trang_thai = str(ws.cell_value(i, i_trang_thai)).strip() if i_trang_thai >= 0 else ''

            result.append({
                'so_gd': so_tc_clean,   # Số thành công = key ghép với CITAD
                'loai': 'ih',            # Hub ngoại tệ luôn là IH
                'chieu': chieu,
                'loai_tien': loai_tien,
                'so_tien': so_tien,
                'nh_nhan': nh,
                'ngay': ngay,
                'trang_thai': trang_thai,
            })
    return result
