"""Business logic Đối chiếu CITAD ↔ PaymentHub.

- Buffer CITAD/PaymentHub: dict in-memory tạm giữ dữ liệu Extension vừa gửi
  lên cho tới khi người dùng bấm "Nạp" — KHÔNG cần bền vững qua restart,
  giống bản gốc. Khác bản gốc ở 1 điểm bắt buộc: bản gốc chạy 1 server cục
  bộ trên máy từng người nên buffer vốn chỉ có 1 chủ; nay dùng chung 1
  backend cho cả Phòng Thanh toán nên buffer phải tách theo `owner`.

  `owner` KHÔNG do client tự khai (đã sửa sau review bảo mật — trước đây
  Extension tự gửi `owner` trong payload, ai có khoá chung cũng ghi được
  buffer dưới bất kỳ tên nào, kể cả chèn số liệu giả để chênh lệch ra đúng
  0). Giờ `owner` = username suy ra từ 1 "mã kết nối" (extension token) cá
  nhân — mỗi người tự tạo trên `/doi_chieu_citad` sau khi đã đăng nhập thật,
  dán vào Extension 1 lần (xem `generate_extension_token`/
  `resolve_extension_token` bên dưới). Token bị lộ chỉ ảnh hưởng đúng 1
  người, thu hồi riêng lẻ được — không cần đổi khoá chung cho cả phòng.
- `_build_xlsx()`: port NGUYÊN 1:1 từ `citad-fixed/server.py::_build_xlsx`
  — đây là mẫu báo cáo "BÁO CÁO ĐỐI CHIẾU GIAO DỊCH HỆ THỐNG THANH TOÁN
  ĐIỆN TỬ LIÊN NGÂN HÀNG" đã duyệt, KHÔNG được đổi bất kỳ dòng
  format/màu/border/công thức nào khi port. NGOẠI LỆ duy nhất (theo yêu
  cầu bổ sung sau khi port): dòng ngày ở tiêu đề A4 đổi từ "(dd/mm/yyyy)"
  sang "(Ngày d tháng m năm yyyy)" — xem `_format_vn_date()`.
- Session lưu theo `ngay` — 1 bản CHUNG cho cả phòng (bản gốc dùng SQLite
  riêng, khoá theo ngay+user_id tự nhập, vốn chỉ 1 người/máy nên không có
  khái niệm "chung"). Nhiều người cùng chấm 1 ngày: ai lưu sau cùng là bản
  hiện hành (ghi đè `doi_chieu_citad_sessions`), nhưng mỗi lần lưu đều ghi
  thêm 1 dòng vào `doi_chieu_citad_history` (ngay, staff_id, created_at) —
  xem `get_reconciliation_history()` — nút "Lịch sử đối chiếu" trên trang
  hiển thị đúng thứ tự ai đã chấm ngày đó lúc nào.
- **"Lưu bản tạm" / "Lưu bản cuối" (`status`, thêm 2026-08-20)** — xem
  `session_save()`. Bản tạm cho phép NGƯỜI KHÁC người lập bảng (`created_by`)
  vào nạp riêng Napas/PSS-MDP qua Extension, cứu tình huống 1 người chấm 5
  Cổng CITAD/PaymentHub nhưng Napas/PSS-MDP phải người khác quét (trang CITAD
  đó chỉ có ở Cổng 1). Bản cuối CHỐT — không ai sửa được nữa kể cả người lập
  bảng, chỉ Admin mở khoá lại qua `session_admin_unlock()`. `created_by` KHÁC
  `updated_by`: created_by cố định (người lập bảng, đầu tiên lưu ngày đó),
  updated_by đổi theo người lưu sau cùng (kể cả người chỉ nạp Napas).
- `build_extension_zip()`: nén thư mục `extension_citad/` (nằm ở gốc repo,
  cạnh `backend/`) thành 1 file .zip TẠI THỜI ĐIỂM TẢI — không lưu sẵn file
  zip nào, luôn khớp đúng code hiện tại của extension, không cần bước build
  riêng. Phục vụ nút "Tải Extension" trên `/doi_chieu_citad` (Chrome không
  cho web tự cài extension — đây chỉ là tải file để người dùng tự Load
  unpacked, xem `extension_citad/README.md`).
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import secrets
import sqlite3
import threading
import zipfile
from datetime import datetime
from decimal import Decimal

from backend.core.config import BASE_DIR
from backend.database import _vn_now
from backend.schemas.doi_chieu_citad import ExportIn

EXTENSION_DIR = BASE_DIR / "extension_citad"


def _format_vn_date(day_str: str) -> str:
    """'dd/mm/yyyy' -> 'Ngày dd tháng m năm yyyy' (khớp mẫu báo cáo NHNN).
    Nếu không parse được (định dạng lạ), trả nguyên chuỗi gốc trong ngoặc."""
    try:
        d, m, y = day_str.strip().split('/')
        return f'Ngày {int(d)} tháng {int(m)} năm {y}'
    except Exception:
        return day_str

# ── CITAD / PaymentHub buffer (in-memory) — tách theo owner (username) ────
# {owner: {key: data}} — mỗi người dùng chỉ thấy/xoá buffer của chính mình.
# Route dùng các hàm này đều là `def` đồng bộ nên FastAPI chạy chúng trong
# threadpool THẬT (nhiều thread đồng thời) — không có _buffer_lock thì 2 lỗi
# thật xảy ra: (1) buffer_get_*() lặp .values() trong khi thread khác đang
# thêm key mới (đổi kích thước dict) → CPython ném "RuntimeError: dictionary
# changed size during iteration"; (2) buffer_save_*() gồm 2 bước KHÔNG
# nguyên tử (setdefault rồi mới gán) — nếu buffer_clear_*() xen giữa 2 bước
# đó, item vừa ghi rơi vào bucket đã bị pop khỏi dict ngoài, biến mất im
# lặng. Cả 2 tình huống đều khả thi thực tế: Extension bắn nhiều request
# liên tiếp trong khi người dùng bấm "Nạp" gần như đồng thời.
_buffer_lock = threading.Lock()
_citad_buffer: dict[str, dict] = {}
_ph_buffer: dict[str, dict] = {}


def buffer_save_citad(owner: str, data: dict) -> None:
    with _buffer_lock:
        _citad_buffer.setdefault(owner, {})[data["key"]] = data


def buffer_get_citad(owner: str) -> list:
    with _buffer_lock:
        return list(_citad_buffer.get(owner, {}).values())


def buffer_clear_citad(owner: str) -> None:
    with _buffer_lock:
        _citad_buffer.pop(owner, None)


def buffer_save_ph(owner: str, items: list) -> None:
    with _buffer_lock:
        bucket = _ph_buffer.setdefault(owner, {})
        for item in items:
            bucket[item["key"]] = item


def buffer_get_ph(owner: str) -> list:
    with _buffer_lock:
        return list(_ph_buffer.get(owner, {}).values())


def buffer_clear_ph(owner: str) -> None:
    with _buffer_lock:
        _ph_buffer.pop(owner, None)


# ── Extension token — mỗi staff 1 token cá nhân (thay khoá tĩnh dùng chung) ──
# Chỉ lưu SHA-256 hash, không lưu plaintext — không thể xem lại token cũ,
# chỉ tạo mã mới (tự thu hồi mã cũ vì PRIMARY KEY staff_id, 1 token/người).
def _hash_token(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()


def generate_extension_token(db: sqlite3.Connection, staff_id: int) -> str:
    """Tạo token mới cho staff_id, GHI ĐÈ token cũ nếu có (thu hồi tự động).
    Trả về token PLAINTEXT — CHỈ lần này, gọi lại sẽ ra token khác."""
    token = secrets.token_urlsafe(32)
    now = _vn_now()
    db.execute(
        """INSERT INTO doi_chieu_citad_extension_tokens (staff_id, token_hash, created_at, last_used_at)
           VALUES (?,?,?,NULL)
           ON CONFLICT(staff_id) DO UPDATE SET token_hash=excluded.token_hash,
                                                created_at=excluded.created_at,
                                                last_used_at=NULL""",
        (staff_id, _hash_token(token), now),
    )
    db.commit()
    return token


# last_used_at chỉ phục vụ hiển thị trạng thái tương đối trên UI ("lần dùng
# gần nhất") — không cần chính xác từng giây. Extension gọi buffer liên tục
# (mỗi lần MutationObserver bắt được kết quả mới), nếu ghi lại last_used_at
# trên MỌI request thì mỗi request tốn thêm 1 giao dịch ghi (khoá cả file
# SQLite, cạnh tranh trực tiếp với audit_logs của AuditMiddleware và mọi
# module khác dùng chung data/ksnb.db) chỉ để đổi 1 con số hiển thị không ai
# cần xem realtime. Gộp lại: chỉ ghi khi đã quá _LAST_USED_THROTTLE_SECONDS
# kể từ lần ghi trước.
_LAST_USED_THROTTLE_SECONDS = 300


def resolve_extension_token(db: sqlite3.Connection, token: str) -> tuple[int, str] | None:
    """Token hợp lệ -> trả về (staff_id, username) chủ token — staff_id để
    caller ghi audit đúng người (xem `_resolve_extension_owner` trong
    `backend/api/doi_chieu_citad.py`), username để dùng làm khoá buffer như
    trước. Chỉ ghi lại last_used_at nếu đã "cũ" hơn
    _LAST_USED_THROTTLE_SECONDS (xem comment trên) — phần lớn request KHÔNG
    còn tốn giao dịch ghi nào ở đây nữa.
    Token sai/rỗng/đã bị thu hồi -> None (caller trả 403, không đoán bừa)."""
    if not token:
        return None
    row = db.execute(
        """SELECT t.staff_id, t.last_used_at, u.username FROM doi_chieu_citad_extension_tokens t
           JOIN user_tttt u ON u.id = t.staff_id AND u.is_active = 1
           WHERE t.token_hash = ?""",
        (_hash_token(token),),
    ).fetchone()
    if not row:
        return None
    now = _vn_now()
    last_used = row["last_used_at"]
    stale = last_used is None or (
        now - datetime.fromisoformat(str(last_used))
    ).total_seconds() >= _LAST_USED_THROTTLE_SECONDS
    if stale:
        db.execute(
            "UPDATE doi_chieu_citad_extension_tokens SET last_used_at=? WHERE staff_id=?",
            (now, row["staff_id"]),
        )
        db.commit()
    return row["staff_id"], row["username"]


def revoke_extension_token(db: sqlite3.Connection, staff_id: int) -> None:
    db.execute("DELETE FROM doi_chieu_citad_extension_tokens WHERE staff_id=?", (staff_id,))
    db.commit()


def get_extension_token_status(db: sqlite3.Connection, staff_id: int) -> dict:
    row = db.execute(
        "SELECT created_at, last_used_at FROM doi_chieu_citad_extension_tokens WHERE staff_id=?",
        (staff_id,),
    ).fetchone()
    if not row:
        return {"connected": False, "created_at": None, "last_used_at": None}
    return {
        "connected": True,
        "created_at": str(row["created_at"]) if row["created_at"] else None,
        "last_used_at": str(row["last_used_at"]) if row["last_used_at"] else None,
    }


class SessionLockedError(Exception):
    """Ngày đã "Lưu bản cuối" (status='final') — không ai sửa được nữa qua
    đường lưu thường, kể cả người lập bảng. Chỉ Admin gỡ được qua
    session_admin_unlock()."""


class SessionForbiddenError(Exception):
    """Người gọi không phải người lập bảng (created_by) nên không được sửa
    trường ngoài Napas/PSS-MDP, và không được "Lưu bản cuối"."""


# Đúng 4 field người KHÔNG PHẢI người lập bảng được phép sửa trên 1 bản tạm
# — khớp SessionIn (napas_m/t, pssmdp_m/t). Mọi field khác (gD, phD, lap_bang,
# kiem_soat, ebank_m/t — ebank giữ nguyên không ai sửa được nữa, xem
# doi_chieu_citad.py) LUÔN giữ nguyên giá trị đã có trong bản tạm, bất kể
# client gửi lên gì — không tin dữ liệu client cho các field ngoài phạm vi.
_NAPAS_ONLY_FIELDS = ("napas_m", "napas_t", "pssmdp_m", "pssmdp_t")


# ── Session theo ngày — 1 bản CHUNG cho cả phòng (không tách theo người) ────
# "Lưu bản tạm"/"Lưu bản cuối" (status) — xem docstring đầu file. Nhiều người
# cùng chấm 1 ngày: ai lưu sau cùng là bản hiện hành (ghi đè
# doi_chieu_citad_sessions), nhưng mỗi lần lưu đều lưu NGUYÊN VẸN số liệu
# phiên chấm đó vào doi_chieu_citad_history — xem get_reconciliation_history()/
# get_history_entry_data() — nên xem/tải lại đúng bản của từng lần lưu, không
# chỉ biết ai đã sửa lúc nào. Các lần lưu tạm LIÊN TIẾP không đẻ thêm dòng
# lịch sử mới — chỉ UPDATE tại chỗ dòng lịch sử tạm gần nhất (tránh phình
# Lịch sử vì mỗi lần ai đó chỉ nạp thêm Napas cũng gọi lưu). "Lưu bản cuối"
# ĐÓNG đúng dòng tạm đang mở đó — cũng UPDATE tại chỗ (chỉ đổi status
# 'draft' -> 'final'), KHÔNG tách thành 1 dòng lịch sử riêng (trước đây tách
# riêng, gây hiểu lầm "nhảy ra thêm 1 dòng bảng tạm" khi người dùng nhìn
# thấy 2 dòng cho cùng 1 lần chấm — theo yêu cầu người dùng, 1 lần chấm chỉ
# nên là 1 dòng, đổi trạng thái tại chỗ). Dòng lịch sử MỚI chỉ sinh ra khi
# KHÔNG có dòng tạm nào đang mở — ngày chưa từng lưu, hoặc dòng gần nhất đã
# là 'final' (vd sau khi Admin mở khoá rồi lưu tiếp — coi là 1 đợt chấm mới,
# giữ nguyên dòng 'final' cũ làm mốc lịch sử của đợt trước).
def session_save(db: sqlite3.Connection, ngay: str, staff_id: int, data: dict, status: str) -> None:
    if status not in ("draft", "final"):
        raise ValueError(f"status không hợp lệ: {status!r}")

    row = db.execute(
        "SELECT data, status, created_by FROM doi_chieu_citad_sessions WHERE ngay=?", (ngay,)
    ).fetchone()

    created_by = staff_id  # ngày chưa từng có ai lưu — người này là người lập bảng
    if row:
        if row["status"] == "final":
            raise SessionLockedError(
                "Ngày này đã được chốt bản cuối — không thể lưu thêm. Liên hệ Admin nếu cần mở khoá."
            )
        created_by = row["created_by"] or staff_id
        if created_by != staff_id:
            # Không phải người lập bảng — chỉ được lưu tạm, chỉ được đổi đúng
            # 4 field Napas/PSS-MDP, giữ nguyên mọi field khác của bản tạm cũ.
            if status == "final":
                raise SessionForbiddenError(
                    "Chỉ người lập bảng mới được \"Lưu bản cuối\"."
                )
            existing = json.loads(row["data"])
            for f in _NAPAS_ONLY_FIELDS:
                existing[f] = data.get(f, existing.get(f, 0))
            data = existing

    now = _vn_now()
    data_json = json.dumps(data)
    db.execute(
        """INSERT INTO doi_chieu_citad_sessions (ngay, data, updated_at, updated_by, status, created_by)
           VALUES (?,?,?,?,?,?)
           ON CONFLICT(ngay) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at,
                                            updated_by=excluded.updated_by, status=excluded.status,
                                            created_by=excluded.created_by""",
        (ngay, data_json, now, staff_id, status, created_by),
    )

    # Gộp lưu tạm liên tiếp vào CÙNG 1 dòng lịch sử — nhưng CHỈ khi cùng 1
    # người lưu liên tiếp (thêm điều kiện staff_id, xác nhận yêu cầu Phòng
    # Thanh toán 25/08/2026). TRƯỚC ĐÂY chỉ xét status=='draft', không xét
    # ai lưu — 2 người khác nhau lưu tạm nối tiếp nhau (vd A lưu tạm, B bổ
    # sung Napas rồi lưu tạm tiếp) sẽ bị gộp chung 1 dòng, đè mất dấu vết
    # dòng riêng của A, chỉ còn thấy B trong Lịch sử dù cả 2 đều đã lưu
    # thật. Khác người thì tách dòng MỚI — mỗi người 1 dòng riêng cho lần
    # họ lưu, đúng ý "mỗi người chấm là 1 dòng".
    last_hist = db.execute(
        "SELECT id, status, staff_id FROM doi_chieu_citad_history WHERE ngay=? ORDER BY id DESC LIMIT 1", (ngay,)
    ).fetchone()
    if last_hist and last_hist["status"] == "draft" and last_hist["staff_id"] == staff_id:
        db.execute(
            "UPDATE doi_chieu_citad_history SET data=?, created_at=?, status=? WHERE id=?",
            (data_json, now, status, last_hist["id"]),
        )
        hist_id = last_hist["id"]
    else:
        cur = db.execute(
            "INSERT INTO doi_chieu_citad_history (ngay, staff_id, data, created_at, status) VALUES (?,?,?,?,?)",
            (ngay, staff_id, data_json, now, status),
        )
        hist_id = cur.lastrowid

    # Ghi riêng vào nhật ký sửa — KHÔNG gộp như dòng lịch sử ở trên, để giữ
    # đủ dấu vết từng lần lưu tạm dù chúng chung 1 history_id (xem docstring
    # get_history_edits()).
    db.execute(
        "INSERT INTO doi_chieu_citad_history_edits (history_id, staff_id, created_at) VALUES (?,?,?)",
        (hist_id, staff_id, now),
    )
    db.commit()


def session_admin_unlock(db: sqlite3.Connection, ngay: str) -> None:
    """Chỉ Admin gọi được (kiểm tra role ở lớp API) — mở khoá 1 ngày đã
    "Lưu bản cuối" về lại 'draft' để sửa tiếp. Không đổi created_by (người
    lập bảng vẫn là người cũ, vẫn là người duy nhất sửa được đủ mọi field
    sau khi mở khoá — chỉ status đổi)."""
    db.execute("UPDATE doi_chieu_citad_sessions SET status='draft' WHERE ngay=?", (ngay,))
    db.commit()


def session_get(db: sqlite3.Connection, ngay: str) -> dict | None:
    row = db.execute(
        """SELECT s.data, s.status, s.created_by, u.username AS created_by_username
           FROM doi_chieu_citad_sessions s
           LEFT JOIN user_tttt u ON u.id = s.created_by
           WHERE s.ngay=?""",
        (ngay,),
    ).fetchone()
    if not row:
        return None
    data = json.loads(row["data"])
    # Field _meta_* — KHÔNG phải số liệu đối chiếu, chỉ để frontend quyết định
    # ai được sửa gì (xem docstring session_save()). Đặt tiền tố "_meta_" để
    # không lẫn với field nghiệp vụ thật nào của SessionIn.
    data["_meta_status"] = row["status"]
    data["_meta_created_by"] = row["created_by"]
    data["_meta_created_by_username"] = row["created_by_username"]
    return data


_STATUS_CONGS = [1, 9, 18, 17, 12]
_STATUS_CURS = ['VNĐ', 'USD', 'EUR']
_STATUS_FK = ['di_ih_m', 'di_ih_t', 'di_il_m', 'di_il_t', 'den_ih_m', 'den_ih_t', 'den_il_m', 'den_il_t']


def _status_dec(v) -> Decimal:
    """Chuyển sang Decimal CHÍNH XÁC TUYỆT ĐỐI — xem `_dec()` trong
    `frontend/pages/doi_chieu_citad.py` (cùng lý do, đi qua `str(v)` để
    tránh mở khai triển nhị phân của float)."""
    try:
        return Decimal(str(v)) if v not in (None, '') else Decimal(0)
    except Exception:
        return Decimal(0)


def is_reconciliation_matched(sess: dict) -> bool:
    """True nếu tổng CITAD (5 cổng + Napas/PSS-MDP IH Đến) == tổng
    PaymentHub cho ĐỦ 8 trường — đúng công thức dòng "CHÊNH LỆCH" hiện trên
    trang Đối chiếu CITAD (`_compute_totals()` ở
    frontend/pages/doi_chieu_citad.py) — giữ đồng bộ công thức ở 2 nơi vì
    frontend không gọi được service backend trực tiếp (khác tiến trình).

    Cộng dồn bằng Decimal (`_status_dec()`), KHÔNG bằng float — cộng nhiều
    dòng (5 Cổng + Napas + PSS-MDP) bằng số thực có thể sinh dư nhị phân dù
    về bản chất đã khớp tuyệt đối (bug thật 25/08/2026: 0,0078125 dù CITAD
    gốc cộng đúng khớp PaymentHub). Decimal cộng đúng tuyệt đối với số liệu
    gốc — không làm tròn, nên không có nguy cơ che mất lệch thật dù nhỏ."""
    gD = sess.get("gD", {}) or {}
    phD = sess.get("phD", {}) or {}
    ci = {f: Decimal(0) for f in _STATUS_FK}
    for c in _STATUS_CONGS:
        for u in _STATUS_CURS:
            src = (gD.get(str(c), {}) or {}).get(u, {}) or {}
            for f in _STATUS_FK:
                ci[f] += _status_dec(src.get(f, 0))
    ci["den_ih_m"] += _status_dec(sess.get("napas_m", 0)) + _status_dec(sess.get("pssmdp_m", 0))
    ci["den_ih_t"] += _status_dec(sess.get("napas_t", 0)) + _status_dec(sess.get("pssmdp_t", 0))
    ph = {f: Decimal(0) for f in _STATUS_FK}
    for u in _STATUS_CURS:
        src = phD.get(u, {}) or {}
        for f in _STATUS_FK:
            ph[f] += _status_dec(src.get(f, 0))
    return all(ci[f] == ph[f] for f in _STATUS_FK)


def get_reconciliation_status(db: sqlite3.Connection, ngay: str) -> dict:
    """Trạng thái đối chiếu của 1 ngày, dùng để cảnh báo ở module Sổ trực
    (xem `so_truc_service.check_citad_status`) — KHÔNG phải endpoint hiển
    thị số liệu, chỉ trả 2 cờ: có bản LƯU BẢNG CUỐI chưa, và bản đó đã khớp
    (hết chênh lệch) chưa.

    Chỉ tính bản đã "Lưu bảng cuối" (status='final') là "đã có đối chiếu" —
    bảng tạm (status='draft') vẫn có thể còn đang chấm dở/chưa đủ người góp
    Napas-PSS-MDP, coi như CHƯA CÓ để Sổ trực vẫn cảnh báo, không để lọt bản
    tạm chưa hoàn chỉnh."""
    sess = session_get(db, ngay)
    is_final = bool(sess) and sess.get("_meta_status") == "final"
    return {
        "exists": is_final,
        "matched": is_final and is_reconciliation_matched(sess),
    }


def session_list(db: sqlite3.Connection) -> list:
    rows = db.execute(
        "SELECT data FROM doi_chieu_citad_sessions ORDER BY ngay DESC",
    ).fetchall()
    return [json.loads(r["data"]) for r in rows]


def _parse_ngay(ngay: str) -> datetime | None:
    try:
        return datetime.strptime(ngay.strip(), "%d/%m/%Y")
    except Exception:
        return None


def get_reconciliation_days(
    db: sqlite3.Connection,
    tu_ngay: str | None = None,
    den_ngay: str | None = None,
    nguoi_cham: str | None = None,
) -> list:
    """1 dòng/ngày đã có ai chấm — ngày, user chấm đối chiếu (người lập bảng,
    CỐ ĐỊNH), số lần lưu, cập nhật lúc — phục vụ tab "Lịch sử" (bảng nhiều
    ngày cùng lúc, lọc theo khoảng ngày + tên người chấm). Lọc/sắp xếp bằng
    Python vì cột `ngay` lưu dạng text dd/mm/yyyy — so sánh chuỗi trực tiếp
    trong SQL sẽ SAI thứ tự thời gian (ví dụ "01/12/2026" < "05/01/2026"
    theo string nhưng đến sau). `nguoi_cham` so khớp KHÔNG phân biệt hoa/
    thường, khớp theo cả tên đầy đủ lẫn username.

    Cột hiển thị lấy `created_by` (người lập bảng), KHÔNG lấy `updated_by`
    (người lưu sau cùng) — trước đây dùng updated_by khiến cột này bị "ghi
    đè" mỗi khi người KHÁC người lập bảng chỉ nạp thêm Napas/PSS-MDP vào
    bảng tạm (xem _NAPAS_ONLY_FIELDS), gây hiểu lầm đổi cả người phụ trách.
    Ai đã từng sửa gì lúc nào xem qua icon "Ai đã sửa bảng tạm này"
    (get_history_edits()), tách hẳn khỏi cột này."""
    rows = db.execute(
        """SELECT s.ngay, s.updated_at, u.username AS created_by_username,
                  u.full_name AS created_by_name,
                  (SELECT COUNT(*) FROM doi_chieu_citad_history h WHERE h.ngay = s.ngay) AS so_lan_luu
           FROM doi_chieu_citad_sessions s
           LEFT JOIN user_tttt u ON u.id = s.created_by"""
    ).fetchall()

    tu_dt = _parse_ngay(tu_ngay) if tu_ngay else None
    den_dt = _parse_ngay(den_ngay) if den_ngay else None
    nguoi_kw = nguoi_cham.strip().lower() if nguoi_cham else None

    parsed = []
    for r in rows:
        d = _parse_ngay(r["ngay"])
        if d is None:  # bỏ qua dòng ngày lỗi định dạng, không để sập cả danh sách
            continue
        if tu_dt and d < tu_dt:
            continue
        if den_dt and d > den_dt:
            continue
        if nguoi_kw:
            hay = f"{r['created_by_username'] or ''} {r['created_by_name'] or ''}".lower()
            if nguoi_kw not in hay:
                continue
        parsed.append((d, {
            "ngay": r["ngay"],
            "created_by_username": r["created_by_username"],
            "created_by_name": r["created_by_name"],
            "updated_at": str(r["updated_at"]) if r["updated_at"] else None,
            "so_lan_luu": r["so_lan_luu"],
        }))
    parsed.sort(key=lambda t: t[0], reverse=True)
    return [item for _, item in parsed]


def get_reconciliation_history(db: sqlite3.Connection, ngay: str) -> list:
    """Lịch sử từng lần lưu đối chiếu CITAD của 1 ngày cụ thể, theo đúng thứ
    tự thời gian đã lưu (cũ -> mới) — KHÔNG trả kèm số liệu (có thể nặng nếu
    nhiều dòng) — xem từng bản cụ thể qua get_history_entry_data(id).

    `username` mỗi dòng lấy đúng `h.staff_id` — người THỰC SỰ bấm Lưu ra
    đúng dòng lịch sử này (xác nhận yêu cầu Phòng Thanh toán 25/08/2026:
    mỗi dòng bung ra phải hiện đúng người đã lưu dòng đó, không gộp về 1
    tên duy nhất của cả ngày). TRƯỚC ĐÂY override bằng `created_by` của
    `doi_chieu_citad_sessions` (người lập bảng gốc, cố định suốt ngày) với
    lý do tránh "nhảy lung tung" khi bản tạm bị gộp nhiều lần lưu vào cùng
    1 dòng — nhưng đó chính xác lại là điều Phòng Thanh toán muốn THẤY:
    dòng lịch sử nào do ai lưu sau cùng thì hiện đúng người đó, không che
    đi. Ai đã sửa TỪNG PHẦN dữ liệu trong 1 dòng (không chỉ ai bấm Lưu sau
    cùng) xem chi tiết hơn qua get_history_edits()."""
    rows = db.execute(
        """SELECT h.id, h.status, h.created_at, h.staff_id, hu.username
           FROM doi_chieu_citad_history h
           JOIN user_tttt hu ON hu.id = h.staff_id
           WHERE h.ngay = ?
           ORDER BY h.created_at ASC, h.id ASC""",
        (ngay,),
    ).fetchall()
    return [
        {
            "id": r["id"], "staff_id": r["staff_id"], "username": r["username"],
            "created_at": str(r["created_at"]), "status": r["status"],
        }
        for r in rows
    ]


def get_history_entry_data(db: sqlite3.Connection, history_id: int) -> dict | None:
    """Số liệu NGUYÊN VẸN của đúng 1 lần lưu trong lịch sử — phục vụ nút
    "Tải" trên từng dòng lịch sử, xem lại/khôi phục đúng bản của lần lưu
    đó (khác nút "Tải" chính, luôn lấy bản HIỆN HÀNH mới nhất).

    Kèm `_meta_*` như session_get() — `created_by`/`created_by_username` lấy
    từ `doi_chieu_citad_sessions` của đúng `ngay` (người lập bảng tính theo
    NGÀY, không phải theo từng dòng lịch sử — 1 dòng lịch sử có thể do người
    KHÁC người lập bảng lưu, vd người chỉ nạp Napas).

    `_meta_status` lấy từ TRẠNG THÁI HIỆN TẠI của ngày đó (`s.status`), KHÔNG
    phải `h.status` đóng băng lúc lưu đúng dòng này — nếu dùng h.status, mở
    lại 1 dòng tạm từ TRƯỚC một lần Admin mở khoá rồi chốt lại sẽ hiện nhầm
    "sửa/lưu tiếp được" dù ngày đó đã khoá thật, bấm Lưu sẽ ăn lỗi 403 (bug
    đã gặp thực tế, xem lịch sử sửa)."""
    row = db.execute(
        """SELECT h.data, s.status, s.created_by, u.username AS created_by_username
           FROM doi_chieu_citad_history h
           LEFT JOIN doi_chieu_citad_sessions s ON s.ngay = h.ngay
           LEFT JOIN user_tttt u ON u.id = s.created_by
           WHERE h.id=?""",
        (history_id,),
    ).fetchone()
    if not row:
        return None
    data = json.loads(row["data"])
    data["_meta_status"] = row["status"]
    data["_meta_created_by"] = row["created_by"]
    data["_meta_created_by_username"] = row["created_by_username"]
    return data


def get_history_edits(db: sqlite3.Connection, history_id: int) -> list:
    """Danh sách MỌI lần lưu đã góp phần tạo nên dòng lịch sử này, kèm thời
    gian — khác get_reconciliation_history() (1 dòng/lần lưu, đã GỘP các lần
    lưu tạm liên tiếp), ở đây liệt kê đầy đủ từng người từng lưu kể cả những
    lần lưu tạm bị gộp không tạo dòng lịch sử riêng. Phục vụ icon "Ai đã sửa
    bảng tạm này" trên tab Lịch sử."""
    rows = db.execute(
        """SELECT e.staff_id, u.username, u.full_name, e.created_at
           FROM doi_chieu_citad_history_edits e
           JOIN user_tttt u ON u.id = e.staff_id
           WHERE e.history_id = ?
           ORDER BY e.created_at ASC, e.id ASC""",
        (history_id,),
    ).fetchall()
    return [
        {
            "staff_id": r["staff_id"], "username": r["username"],
            "full_name": r["full_name"], "created_at": str(r["created_at"]),
        }
        for r in rows
    ]


def session_delete(db: sqlite3.Connection, ngay: str, staff_id: int) -> None:
    """Chỉ người lập bảng mới xoá được, và không xoá được ngày đã "Lưu bản
    cuối" — cùng nguyên tắc với session_save() (xem docstring đó). Không có
    ngoại lệ Admin ở đây (khác session_admin_unlock() — mở khoá để SỬA TIẾP,
    không phải để xoá trắng số liệu đã chốt)."""
    row = db.execute(
        "SELECT status, created_by FROM doi_chieu_citad_sessions WHERE ngay=?", (ngay,)
    ).fetchone()
    if row:
        if row["status"] == "final":
            raise SessionLockedError("Ngày này đã được chốt bản cuối — không thể xoá.")
        if row["created_by"] and row["created_by"] != staff_id:
            raise SessionForbiddenError("Chỉ người lập bảng mới được xoá bản ghi này.")
    db.execute("DELETE FROM doi_chieu_citad_sessions WHERE ngay=?", (ngay,))
    db.commit()


# ── Xuất Excel — port NGUYÊN 1:1 từ citad-fixed/server.py::_build_xlsx ─────
def build_xlsx(data: ExportIn) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    CONGS = [1, 9, 18, 17, 12]
    CURS = ['VNĐ', 'USD', 'EUR']
    FK = ['di_ih_m', 'di_ih_t', 'di_il_m', 'di_il_t', 'den_ih_m', 'den_ih_t', 'den_il_m', 'den_il_t']
    TNR = 'Times New Roman'

    def nv(v):
        try:
            return float(str(v).replace(',', '')) if v else 0
        except Exception:
            return 0

    def _dec(v) -> Decimal:
        """Chuyển sang Decimal CHÍNH XÁC TUYỆT ĐỐI — xem `_dec()` trong
        `frontend/pages/doi_chieu_citad.py` (cùng lý do: `str(v)` trước khi
        vào Decimal để tránh mở khai triển nhị phân của float). Dùng riêng
        cho `ci`/`ph`/`diff` (dòng CITAD/PaymentHub/Chênh lệch) — cộng dồn
        nhiều dòng (5 Cổng × 3 loại tiền + Napas + PSS-MDP) bằng số thực có
        thể sinh dư nhị phân dù về bản chất đã khớp tuyệt đối (bug thật
        25/08/2026: 0,0078125 dù CITAD gốc cộng đúng khớp PaymentHub)."""
        try:
            return Decimal(str(v)) if v else Decimal(0)
        except Exception:
            return Decimal(0)

    def F(bold=False, size=14, color='000000'):
        return Font(name=TNR, bold=bold, size=size, color=color)

    def AL(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def Bdr(w='thin', bw=None):
        s = Side(style=w)
        bs = Side(style=bw or w)
        return Border(top=s, bottom=bs, left=s, right=s)

    def Fill(h):
        return PatternFill('solid', fgColor=h)

    ci = [Decimal(0)] * 8
    for c in CONGS:
        for u in CURS:
            src = (data.gD.get(str(c), {}) or {}).get(u, {}) or {}
            for i, f in enumerate(FK):
                ci[i] += _dec(src.get(f, 0))
    # Chỉ cộng Napas (nm/nt) vào tổng CITAD — KHÔNG cộng Ebanking (em/et).
    # Đã đối chiếu với DoiChieuCITAD.py::_calc() của tool desktop gốc: gốc
    # CŨNG chỉ cộng napas.den_ih_m/t vào ci['den_ih_m'/'t'], không có dòng
    # tương ứng cho ebank — đây là hành vi gốc, KHÔNG phải sai sót khi port.
    # 20/08/2026: dòng "Ebanking" đã bỏ khỏi Excel xuất ra (kênh này không
    # còn dùng, đồng bộ với việc đã bỏ ô nhập Ebanking khỏi màn hình trước
    # đó) — data.em/et không còn được dùng ở đâu trong build_xlsx nữa,
    # vẫn giữ 2 field trong ExportIn để không phá payload cũ.
    ci[4] += _dec(data.nm)
    ci[5] += _dec(data.nt)
    # PSS - MDP: kênh mới thêm sau, theo yêu cầu Phòng Thanh toán — CÙNG
    # nguyên lý với Napas (cộng vào tổng CITAD), khác Ebanking (không cộng).
    ci[4] += _dec(data.sm)
    ci[5] += _dec(data.st)
    ph = [Decimal(0)] * 8
    for u in CURS:
        src = (data.phD.get(u, {}) or {})
        for i, f in enumerate(FK):
            ph[i] += _dec(src.get(f, 0))
    diff = [ci[i] - ph[i] for i in range(8)]
    wb = Workbook()
    ws = wb.active
    # openpyxl raise lỗi (500 không kiểm soát) nếu tên sheet chứa ký tự Excel
    # cấm (: \ / ? * [ ]) — lọc trước khi gán, không đổi nội dung số liệu.
    safe_sheet_name = re.sub(r'[:\\/?*\[\]]', '_', data.sheet_name) or 'Sheet1'
    ws.title = safe_sheet_name[:31]
    NUM = '#,##0'
    for col, wd in zip('ABCDEFGHIJ', [24.43, 7, 11.43, 30.14, 11.43, 28.57, 10.86, 30.14, 12.43, 28.57]):
        ws.column_dimensions[col].width = wd

    def rh(r, h=18.75):
        ws.row_dimensions[r].height = h

    BLU = '4472C4'

    def hcell(r, c, val, merge_to=None):
        cell = ws.cell(r, c)
        cell.value = val
        cell.font = Font(name=TNR, bold=True, size=11, color='FFFFFF')
        cell.alignment = AL()
        cell.fill = Fill(BLU)
        cell.border = Bdr()
        if merge_to:
            ws.merge_cells(f'{chr(64+c)}{r}:{chr(64+merge_to)}{r}')

    rh(1, 53.25)
    ws.merge_cells('B1:E1')
    ws['B1'] = 'NGÂN HÀNG NÔNG NGHIỆP\nVÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM'
    ws['B1'].font = F()
    ws['B1'].alignment = AL(wrap=True)
    rh(2)
    ws.merge_cells('B2:E2')
    ws['B2'] = 'TRUNG TÂM THANH TOÁN'
    ws['B2'].font = F(bold=True)
    ws['B2'].alignment = AL()
    rh(3)
    rh(4, 57)
    ws.merge_cells('A4:J4')
    ws['A4'] = f'BÁO CÁO ĐỐI CHIẾU GIAO DỊCH HỆ THỐNG THANH TOÁN ĐIỆN TỬ LIÊN NGÂN HÀNG\n({_format_vn_date(data.day_str)})'
    ws['A4'].font = F(bold=True)
    ws['A4'].alignment = AL(wrap=True)
    for c in range(1, 11):
        rh(5)
        rh(6)
        rh(7)
        ws.cell(5, c).fill = Fill(BLU)
        ws.cell(5, c).border = Bdr()
        ws.cell(6, c).fill = Fill(BLU)
        ws.cell(6, c).border = Bdr()
        ws.cell(7, c).fill = Fill(BLU)
        ws.cell(7, c).border = Bdr()
    hcell(5, 3, 'LỆNH ĐI', 6)
    hcell(5, 7, 'LỆNH ĐẾN', 10)
    hcell(6, 3, 'IH', 4)
    hcell(6, 5, 'IL', 6)
    hcell(6, 7, 'IH', 8)
    hcell(6, 9, 'IL', 10)
    for c, lbl in zip(range(3, 11), ['SỐ MÓN', 'SỐ TIỀN'] * 4):
        hcell(7, c, lbl)

    def wr(r, lbl, cur, vals, bold=False, fh=None, lc='000000'):
        rh(r)
        fl = Fill(fh) if fh else None
        for ci2, (val, is_lbl, is_cur) in enumerate(
            [(lbl, True, False), (cur, False, True)] + [(v, False, False) for v in vals]
        ):
            c = ws.cell(r, ci2 + 1)
            if is_lbl:
                c.value = val
                c.font = Font(name=TNR, bold=bold, size=14, color=lc)
                c.alignment = AL()
            elif is_cur:
                c.value = val
                c.font = F()
                c.alignment = AL()
            else:
                if val:
                    c.value = float(val)
                    c.number_format = NUM
                c.font = Font(name=TNR, bold=bold, size=14, color='000000')
                c.alignment = AL('right')
            c.border = Bdr()
            if fl:
                c.fill = fl

    for ri, cur in enumerate(['EUR', 'USD', 'VNĐ']):
        v = [nv((data.phD.get(cur, {}) or {}).get(f, 0)) for f in FK]
        wr(8 + ri, f'Payment {cur}', cur, v, bold=True)
    wr(11, 'CITAD', '', ci, bold=True, fh='BDD7EE', lc='1F3864')
    row = 12
    for cong in CONGS:
        for ci2, cur in enumerate(CURS):
            v = [nv((data.gD.get(str(cong), {}).get(cur, {}) or {}).get(f, 0)) for f in FK]
            wr(row, f'Cổng {cong}' if ci2 == 0 else '', cur, v, bold=(ci2 == 0))
            row += 1
    wr(row, 'Napas', '', [0, 0, 0, 0, data.nm, data.nt, 0, 0])
    row += 1
    wr(row, 'PSS - MDP', '', [0, 0, 0, 0, data.sm, data.st, 0, 0])
    row += 1
    rh(row)
    ws.cell(row, 1).value = 'Chênh lệch'
    ws.cell(row, 1).font = Font(name=TNR, bold=True, size=14, color='7F0000')
    ws.cell(row, 1).alignment = AL()
    ws.cell(row, 1).fill = Fill('FFE699')
    ws.cell(row, 1).border = Bdr('thin', 'medium')
    ws.cell(row, 2).value = 0
    ws.cell(row, 2).font = F(bold=True)
    ws.cell(row, 2).alignment = AL()
    ws.cell(row, 2).fill = Fill('FFE699')
    ws.cell(row, 2).border = Bdr('thin', 'medium')
    for i, v in enumerate(diff):
        c = ws.cell(row, 3 + i)
        # ci/ph cộng bằng Decimal (_dec()) nên `v` ở đây CHÍNH XÁC TUYỆT ĐỐI,
        # không có dư nhị phân nào phải làm tròn/che đi — khớp thật mới ra
        # đúng 0, lệch thật dù nhỏ (kể cả dưới 1 đơn vị) vẫn hiện đúng số,
        # không đánh đổi độ chính xác lấy gọn màn hình. Trước đây `v if v
        # else None` để trống ô khi khớp (0 là falsy) thay vì hiện "0" —
        # giờ luôn ghi giá trị thật.
        c.value = float(v)
        c.number_format = NUM
        c.font = Font(name=TNR, bold=True, size=14, color='006100' if v == 0 else 'FF0000')
        c.alignment = AL('right')
        c.fill = Fill('FFE699')
        c.border = Bdr('thin', 'medium')
    row += 2
    rh(row - 1)
    rh(row)
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row, 1).value = '                  LẬP BẢNG'
    ws.cell(row, 1).font = F(bold=True)
    ws.cell(row, 1).alignment = AL()
    ws.merge_cells(f'H{row}:I{row}')
    ws.cell(row, 8).value = 'KIỂM SOÁT'
    ws.cell(row, 8).font = F(bold=True)
    ws.cell(row, 8).alignment = AL()
    row += 1
    for _ in range(4):
        rh(row)
        row += 1
    rh(row)
    ws.cell(row, 2).value = data.lb
    ws.cell(row, 2).font = F(bold=True)
    ws.cell(row, 2).alignment = AL()
    ws.merge_cells(f'H{row}:I{row}')
    ws.cell(row, 8).value = data.ks
    ws.cell(row, 8).font = F(bold=True)
    ws.cell(row, 8).alignment = AL()
    ws.print_area = f'A1:J{row}'

    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.3, bottom=0.3, header=0.1, footer=0.1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tải Extension Chrome dạng .zip (phục vụ nút "Tải Extension") ──────────
def build_extension_zip() -> bytes:
    """Nén `extension_citad/` thành .zip trong bộ nhớ — không lưu file tạm,
    không cần bước build riêng, luôn khớp đúng code hiện tại trên server.

    File nằm NGAY GỐC zip (không bọc thêm 1 lớp thư mục "extension_citad/"
    bên trong) — nếu bọc thêm, công cụ giải nén của Windows ("Extract All")
    sẽ tạo 1 thư mục ngoài cùng trùng tên file zip (cũng "extension_citad"),
    cộng với lớp bên trong zip → lồng 2 lần
    (extension_citad/extension_citad/manifest.json), khiến Chrome báo lỗi
    "Manifest file is missing or unreadable" vì manifest.json không nằm
    ngay trong thư mục người dùng chọn ở "Load unpacked"."""
    if not EXTENSION_DIR.is_dir():
        raise FileNotFoundError(f"Không tìm thấy thư mục {EXTENSION_DIR}")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(EXTENSION_DIR.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=path.relative_to(EXTENSION_DIR))
    return buf.getvalue()


def get_extension_latest_version() -> str:
    """Đọc field "version" từ manifest.json — LUÔN khớp đúng bản .zip
    `build_extension_zip()` đang phát hành (đọc trực tiếp từ file, không
    hardcode số ở chỗ khác). Frontend dùng để so sánh với version Extension
    đang cài trên máy người dùng (hỏi qua chrome.runtime.sendMessage), báo
    popup nhắc cập nhật nếu khác nhau."""
    manifest_path = EXTENSION_DIR / "manifest.json"
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    return data["version"]
