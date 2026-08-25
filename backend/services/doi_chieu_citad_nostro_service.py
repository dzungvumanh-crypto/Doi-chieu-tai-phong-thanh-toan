"""Business logic Đối chiếu CITAD ↔ PaymentHub — Phòng QLTK Nostro, Vostro.

Song song với `doi_chieu_citad_service.py` (Phòng Thanh toán), KHÔNG sửa file
đó. Extension Chrome là gói RIÊNG (`extension_citad_nv/`, không chung với
`extension_citad/` của Phòng Thanh toán — theo đúng yêu cầu "không trùng với
Phòng Thanh toán") nên `build_extension_zip`/`get_extension_latest_version`
bên dưới đóng gói thư mục riêng đó, KHÔNG import từ `doi_chieu_citad_service`.

Chỉ tái dùng CHUNG duy nhất cơ chế "mã kết nối" (`resolve_extension_token`,
`generate_extension_token`, `revoke_extension_token`,
`get_extension_token_status`) — bảng `doi_chieu_citad_extension_tokens`
trung lập theo staff_id, không có cột phân biệt module/gói Extension nào,
nên 1 mã vẫn xác thực được cho cả 2 Extension riêng biệt.

Buffer CITAD/PaymentHub và bảng session/history đều là bản sao RIÊNG (khoá
theo `ky` — kỳ đối chiếu Từ ngày-Đến ngày, không phải `ngay` đơn — vì Phòng
QLTK Nostro, Vostro có thể chấm gộp nhiều ngày liên tiếp cùng lúc).

Công thức đối chiếu (khác hẳn Phòng Thanh toán — không có IH/IL, không có
chiều Đến, không có ngoại tệ, chỉ 1 nguồn HUB nên chỉ 1 cặp Chênh lệch):
  Tổng CITAD(gtt) = Σ cD[cong]["gtt"] qua 5 cổng; tương tự Tổng CITAD(gtc).
  Tổng HUB(gtt) = phD["gtt"]; Tổng HUB(gtc) = phD["gtc_truoc"] + phD["gtc_tu"].
  Chênh lệch(gtt|gtc) = Tổng CITAD − Tổng HUB.
"""
from __future__ import annotations

import io
import json
import threading
import sqlite3
import zipfile
from datetime import datetime, timedelta

from backend.core.config import BASE_DIR
from backend.database import _vn_now
from backend.schemas.doi_chieu_citad_nostro import ExportIn, CONGS, CONG_LABEL, LOAI_CITAD
# Tái dùng nguyên vẹn cơ chế mã kết nối — trung lập, không gắn riêng Phòng
# Thanh toán (xem docstring đầu file).
from backend.services.doi_chieu_citad_service import (  # noqa: F401
    resolve_extension_token,
    generate_extension_token,
    revoke_extension_token,
    get_extension_token_status,
)

EXTENSION_DIR = BASE_DIR / "extension_citad_nv"


def build_extension_zip() -> bytes:
    """Nén `extension_citad_nv/` — gói Extension RIÊNG của Phòng QLTK
    Nostro, Vostro (không chung `extension_citad/` của Phòng Thanh toán).
    Cùng nguyên tắc đóng gói với `doi_chieu_citad_service.build_extension_zip`
    (file nằm NGAY GỐC zip, không bọc thêm 1 lớp thư mục — xem docstring ở
    đó để biết lý do, tránh lỗi "Manifest file is missing" khi Load unpacked)."""
    if not EXTENSION_DIR.is_dir():
        raise FileNotFoundError(f"Không tìm thấy thư mục {EXTENSION_DIR}")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(EXTENSION_DIR.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=path.relative_to(EXTENSION_DIR))
    return buf.getvalue()


def get_extension_latest_version() -> str:
    manifest_path = EXTENSION_DIR / "manifest.json"
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    return data["version"]

# ── CITAD / PaymentHub buffer (in-memory) — tách theo owner (username) ────
# Bản sao RIÊNG của Phòng QLTK Nostro, Vostro — không dùng chung dict với
# `doi_chieu_citad_service._citad_buffer`/`_ph_buffer` (khác cấu trúc dữ
# liệu: không có cur/chieu, có loai="gtc_truoc"/"gtc_tu" bên HUB). Cùng lý do
# cần khoá (_buffer_lock) như bản gốc — nhiều thread FastAPI threadpool có
# thể đọc/ghi đồng thời (xem giải thích chi tiết trong doi_chieu_citad_service.py).
_buffer_lock = threading.Lock()
_citad_buffer: dict[str, dict] = {}
_ph_buffer: dict[str, dict] = {}


def buffer_save_citad(owner: str, data: dict) -> None:
    with _buffer_lock:
        _citad_buffer.setdefault(owner, {})[data["key"]] = data


def buffer_get_citad(owner: str) -> list:
    with _buffer_lock:
        return list(_citad_buffer.get(owner, {}).values())


def buffer_clear_citad(owner: str) -> None:
    with _buffer_lock:
        _citad_buffer.pop(owner, None)


def buffer_save_ph(owner: str, items: list) -> None:
    with _buffer_lock:
        bucket = _ph_buffer.setdefault(owner, {})
        for item in items:
            bucket[item["key"]] = item


def buffer_get_ph(owner: str) -> list:
    with _buffer_lock:
        return list(_ph_buffer.get(owner, {}).values())


def buffer_clear_ph(owner: str) -> None:
    with _buffer_lock:
        _ph_buffer.pop(owner, None)


# ── Session theo kỳ đối chiếu — 1 bản CHUNG cho cả phòng (không tách theo
# người) ────────────────────────────────────────────────────────────────
def session_save(db: sqlite3.Connection, ky: str, staff_id: int, data: dict) -> None:
    """`created_by` chỉ ghi ở lần lưu ĐẦU TIÊN của kỳ — nhánh DO UPDATE cố ý
    KHÔNG đụng tới cột này, nên người lập bảng giữ nguyên dù người khác lưu
    đè sau đó (`updated_by` mới là người lưu sau cùng). COALESCE để vá cả
    bản ghi cũ lỡ tạo trước khi có cột này, tránh dòng mất tên người chấm."""
    now = _vn_now()
    data_json = json.dumps(data)
    db.execute(
        """INSERT INTO doi_chieu_citad_nostro_sessions (ky, data, updated_at, updated_by, created_by)
           VALUES (?,?,?,?,?)
           ON CONFLICT(ky) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at,
                                          updated_by=excluded.updated_by,
                                          created_by=COALESCE(doi_chieu_citad_nostro_sessions.created_by,
                                                              excluded.created_by)""",
        (ky, data_json, now, staff_id, staff_id),
    )
    db.execute(
        "INSERT INTO doi_chieu_citad_nostro_history (ky, staff_id, data, created_at) VALUES (?,?,?,?)",
        (ky, staff_id, data_json, now),
    )
    db.commit()


def session_get(db: sqlite3.Connection, ky: str) -> dict | None:
    row = db.execute(
        "SELECT data FROM doi_chieu_citad_nostro_sessions WHERE ky=?", (ky,),
    ).fetchone()
    return json.loads(row["data"]) if row else None


def session_delete(db: sqlite3.Connection, ky: str) -> None:
    db.execute("DELETE FROM doi_chieu_citad_nostro_sessions WHERE ky=?", (ky,))
    db.commit()


def session_list(db: sqlite3.Connection) -> list:
    """Sắp xếp bằng Python theo NGÀY BẮT ĐẦU của kỳ. `ORDER BY ky DESC` trong
    SQL là so sánh CHUỖI "dd/mm/yyyy-..." nên sai thứ tự thời gian
    ("01/12/2026" < "05/01/2026" theo chuỗi nhưng đến sau) — đúng lý do
    `_parse_ky_start()` tồn tại."""
    rows = db.execute(
        "SELECT ky, data FROM doi_chieu_citad_nostro_sessions",
    ).fetchall()
    parsed = [(_parse_ky_start(r["ky"]) or datetime.min, json.loads(r["data"])) for r in rows]
    parsed.sort(key=lambda t: t[0], reverse=True)
    return [data for _, data in parsed]


def _nv(v) -> float:
    try:
        return float(v) if v not in (None, "") else 0.0
    except Exception:
        return 0.0


def compute_totals(sess: dict) -> tuple[dict, dict]:
    """Tổng CITAD (5 cổng) và Tổng HUB, riêng gtt/gtc — công thức duy nhất,
    dùng chung cho `is_reconciliation_matched()` và `build_xlsx_nostro()` để
    luôn khớp nhau (frontend tính lại y hệt ở `_compute_totals()` của
    `frontend/pages/doi_chieu_citad_nostro.py` — giữ đồng bộ 3 nơi vì
    frontend không gọi được thẳng service backend, khác tiến trình)."""
    cD = sess.get("cD", {}) or {}
    phD = sess.get("phD", {}) or {}
    ci = {loai: {"soMon": 0.0, "soTien": 0.0} for loai in LOAI_CITAD}
    for cong in CONGS:
        cong_data = cD.get(cong, {}) or {}
        for loai in LOAI_CITAD:
            src = cong_data.get(loai, {}) or {}
            ci[loai]["soMon"] += _nv(src.get("soMon", 0))
            ci[loai]["soTien"] += _nv(src.get("soTien", 0))
    gtt_hub = phD.get("gtt", {}) or {}
    gtc_truoc = phD.get("gtc_truoc", {}) or {}
    gtc_tu = phD.get("gtc_tu", {}) or {}
    hub = {
        "gtt": {"soMon": _nv(gtt_hub.get("soMon", 0)), "soTien": _nv(gtt_hub.get("soTien", 0))},
        "gtc": {
            "soMon": _nv(gtc_truoc.get("soMon", 0)) + _nv(gtc_tu.get("soMon", 0)),
            "soTien": _nv(gtc_truoc.get("soTien", 0)) + _nv(gtc_tu.get("soTien", 0)),
        },
    }
    return ci, hub


def is_reconciliation_matched(sess: dict) -> bool:
    ci, hub = compute_totals(sess)
    return all(
        ci[loai]["soMon"] == hub[loai]["soMon"] and ci[loai]["soTien"] == hub[loai]["soTien"]
        for loai in LOAI_CITAD
    )


def get_reconciliation_status(db: sqlite3.Connection, ky: str) -> dict:
    sess = session_get(db, ky)
    return {
        "exists": sess is not None,
        "matched": bool(sess) and is_reconciliation_matched(sess),
    }


def _parse_ky_start(ky: str) -> datetime | None:
    """`ky` = "dd/mm/yyyy-dd/mm/yyyy" — sắp xếp/lọc theo NGÀY BẮT ĐẦU (vế
    trước dấu '-'). Chuỗi text nên không so sánh được trực tiếp theo thứ tự
    thời gian, giống lý do `_parse_ngay()` trong doi_chieu_citad_service.py."""
    try:
        tu = ky.split("-", 1)[0].strip()
        return datetime.strptime(tu, "%d/%m/%Y")
    except Exception:
        return None


def _parse_ddmmyyyy(s: str) -> datetime | None:
    try:
        return datetime.strptime(s.strip(), "%d/%m/%Y")
    except Exception:
        return None


def _parse_ky_range(ky: str) -> tuple[datetime, datetime] | None:
    """"dd/mm/yyyy-dd/mm/yyyy" -> (ngày bắt đầu, ngày kết thúc). Ngày đơn
    (tu=den) hợp lệ như kỳ nhiều ngày bình thường."""
    parts = ky.split("-")
    if len(parts) != 2:
        return None
    start, end = _parse_ddmmyyyy(parts[0]), _parse_ddmmyyyy(parts[1])
    if not start or not end:
        return None
    return (start, end) if start <= end else (end, start)


def normalize_ky(ky: str) -> str:
    """Chuẩn hoá + kiểm tra `ky` trước khi lưu, ném ValueError nếu sai.

    Không kiểm thì ô ngày bị xoá trắng sẽ lưu thành `ky = "-"`: dòng đó
    VÔ HÌNH ở tab Lịch sử (`_parse_ky_start()` trả None nên bị `continue`)
    nhưng vẫn chiếm PRIMARY KEY trong bảng, và UI không còn đường nào xoá.
    Chuẩn hoá luôn thứ tự 2 vế để "05/08/2026-01/08/2026" và
    "01/08/2026-05/08/2026" không thành 2 bản ghi rời của cùng một kỳ."""
    rng = _parse_ky_range(ky or "")
    if not rng:
        raise ValueError(
            "Kỳ đối chiếu không hợp lệ — cần đủ Từ ngày và Đến ngày dạng dd/mm/yyyy."
        )
    return f"{rng[0].strftime('%d/%m/%Y')}-{rng[1].strftime('%d/%m/%Y')}"


def check_period_overlap(db: sqlite3.Connection, tu_ngay: str, den_ngay: str, exclude_ky: str | None = None) -> dict:
    """Kiểm tra kỳ [tu_ngay, den_ngay] SẮP lưu có CHỒNG lên kỳ nào đã lưu
    trước đó không, và có HỞ khoảng trống giữa kỳ liền trước gần nhất với kỳ
    này không — cả 2 đều là lỗi dễ mắc khi chấm gộp nhiều ngày (chồng =
    tính trùng, hở = bỏ sót ngày không ai chấm). Chỉ CẢNH BÁO, không chặn
    lưu — quyết định cuối vẫn ở người dùng (`exclude_ky`: bỏ qua chính kỳ
    đang sửa khi lưu đè lại kỳ cũ, tránh tự báo trùng với chính nó)."""
    new_start, new_end = _parse_ddmmyyyy(tu_ngay), _parse_ddmmyyyy(den_ngay)
    if not new_start or not new_end:
        return {"overlaps": [], "gap_before": None}
    if new_start > new_end:
        new_start, new_end = new_end, new_start

    rows = db.execute("SELECT ky FROM doi_chieu_citad_nostro_sessions").fetchall()
    ranges = []
    for r in rows:
        ky = r["ky"]
        if exclude_ky and ky == exclude_ky:
            continue
        rng = _parse_ky_range(ky)
        if rng:
            ranges.append((ky, rng[0], rng[1]))

    overlaps = [ky for ky, s, e in ranges if s <= new_end and e >= new_start]

    gap_before = None
    prior_ends = [e for _, _, e in ranges if e < new_start]
    if prior_ends:
        latest_prior_end = max(prior_ends)
        gap_days = (new_start - latest_prior_end).days - 1
        if gap_days > 0:
            gap_before = {
                "tu_ngay": (latest_prior_end + timedelta(days=1)).strftime("%d/%m/%Y"),
                "den_ngay": (new_start - timedelta(days=1)).strftime("%d/%m/%Y"),
                "so_ngay": gap_days,
            }
    return {"overlaps": overlaps, "gap_before": gap_before}


def get_reconciliation_days(
    db: sqlite3.Connection,
    tu_ngay: str | None = None,
    den_ngay: str | None = None,
    nguoi_cham: str | None = None,
) -> list:
    """1 dòng/kỳ đã có ai chấm, phục vụ tab "Lịch sử" (lọc theo khoảng ngày +
    tên người chấm). Cột người chấm lấy `created_by` — NGƯỜI LẬP BẢNG, cố
    định suốt vòng đời bản ghi — KHÔNG lấy `updated_by` (người lưu sau cùng):
    bản chung của cả phòng, ai lưu đè sau cũng sẽ chiếm mất tên người lập
    bảng, đúng lỗi đã sửa ở module Phòng Thanh toán (xem `created_by` trong
    _ensure_indexes()).

    Lọc/sắp xếp bằng Python vì `ky` lưu dạng text — so sánh chuỗi trong SQL
    sai thứ tự thời gian. Ngày lọc do người dùng GÕ TAY (ô lọc là input tự
    do) nên phải parse bằng `_parse_ddmmyyyy()` trả None, KHÔNG dùng thẳng
    `strptime` — gõ dở "01/08" mà ném ValueError là sập cả trang lịch sử."""
    rows = db.execute(
        """SELECT s.ky, s.updated_at, u.username AS created_by_username,
                  u.full_name AS created_by_name,
                  (SELECT COUNT(*) FROM doi_chieu_citad_nostro_history h WHERE h.ky = s.ky) AS so_lan_luu
           FROM doi_chieu_citad_nostro_sessions s
           LEFT JOIN user_tttt u ON u.id = s.created_by"""
    ).fetchall()

    tu_dt = _parse_ddmmyyyy(tu_ngay) if tu_ngay else None
    den_dt = _parse_ddmmyyyy(den_ngay) if den_ngay else None
    nguoi_kw = nguoi_cham.strip().lower() if nguoi_cham else None

    parsed = []
    for r in rows:
        d = _parse_ky_start(r["ky"])
        if d is None:
            continue
        if tu_dt and d < tu_dt:
            continue
        if den_dt and d > den_dt:
            continue
        if nguoi_kw:
            hay = f"{r['created_by_username'] or ''} {r['created_by_name'] or ''}".lower()
            if nguoi_kw not in hay:
                continue
        parsed.append((d, {
            "ky": r["ky"],
            "created_by_username": r["created_by_username"],
            "created_by_name": r["created_by_name"],
            "updated_at": str(r["updated_at"]) if r["updated_at"] else None,
            "so_lan_luu": r["so_lan_luu"],
        }))
    parsed.sort(key=lambda t: t[0], reverse=True)
    return [item for _, item in parsed]


def get_reconciliation_history(db: sqlite3.Connection, ky: str) -> list:
    rows = db.execute(
        """SELECT h.id, h.staff_id, u.username, h.created_at
           FROM doi_chieu_citad_nostro_history h
           JOIN user_tttt u ON u.id = h.staff_id
           WHERE h.ky = ?
           ORDER BY h.created_at ASC, h.id ASC""",
        (ky,),
    ).fetchall()
    return [
        {"id": r["id"], "staff_id": r["staff_id"], "username": r["username"], "created_at": str(r["created_at"])}
        for r in rows
    ]


def get_history_entry_data(db: sqlite3.Connection, history_id: int) -> dict | None:
    row = db.execute(
        "SELECT data FROM doi_chieu_citad_nostro_history WHERE id=?", (history_id,)
    ).fetchone()
    return json.loads(row["data"]) if row else None


# ── Xuất Excel ──────────────────────────────────────────────────────────
def build_xlsx_nostro(data: ExportIn) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import re

    TNR = "Times New Roman"

    def F(bold=False, size=11, color="000000"):
        return Font(name=TNR, bold=bold, size=size, color=color)

    def AL(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def Bdr(w="thin"):
        s = Side(style=w)
        return Border(top=s, bottom=s, left=s, right=s)

    def Fill(h):
        return PatternFill("solid", fgColor=h)

    sess = {"cD": data.cD, "phD": data.phD}
    ci, hub = compute_totals(sess)

    wb = Workbook()
    ws = wb.active
    safe_sheet_name = re.sub(r'[:\\/?*\[\]]', '_', data.sheet_name) or 'Sheet1'
    ws.title = safe_sheet_name[:31]
    NUM = '#,##0'
    BLU = '4472C4'
    for col, wd in zip('ABCDEF', [16, 14, 18, 14, 18, 14]):
        ws.column_dimensions[col].width = wd

    def hcell(r, c, val, fill=BLU, color='FFFFFF', bold=True):
        cell = ws.cell(r, c)
        cell.value = val
        cell.font = F(bold=bold, color=color)
        cell.alignment = AL(wrap=True)
        cell.fill = Fill(fill)
        cell.border = Bdr()
        return cell

    ws.merge_cells('A1:F1')
    ws['A1'] = 'BÁO CÁO ĐỐI CHIẾU CITAD - PAYMENTHUB — PHÒNG QLTK NOSTRO, VOSTRO'
    ws['A1'].font = F(bold=True, size=13)
    ws['A1'].alignment = AL()
    ws.row_dimensions[1].height = 22
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Kỳ đối chiếu: {data.tu_ngay} - {data.den_ngay}"
    ws['A2'].font = F(bold=True)
    ws['A2'].alignment = AL()

    row = 4
    hcell(row, 1, 'Cổng CITAD')
    hcell(row, 2, 'GTT - Số món')
    hcell(row, 3, 'GTT - Số tiền')
    hcell(row, 4, 'GTC - Số món')
    hcell(row, 5, 'GTC - Số tiền')
    row += 1
    cD = data.cD or {}
    for cong in CONGS:
        c = cD.get(cong, {}) or {}
        gtt = c.get('gtt', {}) or {}
        gtc = c.get('gtc', {}) or {}
        vals = [CONG_LABEL.get(cong, f'Cổng {cong}'), _nv(gtt.get('soMon', 0)), _nv(gtt.get('soTien', 0)),
                _nv(gtc.get('soMon', 0)), _nv(gtc.get('soTien', 0))]
        for ci2, v in enumerate(vals, start=1):
            cell = ws.cell(row, ci2)
            cell.value = v
            cell.border = Bdr()
            cell.alignment = AL('left' if ci2 == 1 else 'right')
            if ci2 > 1:
                cell.number_format = NUM
        row += 1
    ws.cell(row, 1).value = 'Tổng cộng 5 cổng'
    ws.cell(row, 1).font = F(bold=True)
    ws.cell(row, 1).border = Bdr()
    for ci2, v in [(2, ci['gtt']['soMon']), (3, ci['gtt']['soTien']),
                   (4, ci['gtc']['soMon']), (5, ci['gtc']['soTien'])]:
        cell = ws.cell(row, ci2)
        cell.value = v
        cell.number_format = NUM
        cell.font = F(bold=True)
        cell.alignment = AL('right')
        cell.border = Bdr()
        cell.fill = Fill('DCE6F1')
    row += 2

    hcell(row, 1, 'HUB (PaymentHub)')
    hcell(row, 2, 'GTT - Số món')
    hcell(row, 3, 'GTT - Số tiền')
    hcell(row, 4, 'GTC - Số món')
    hcell(row, 5, 'GTC - Số tiền')
    row += 1
    phD = data.phD or {}
    gtt_h = phD.get('gtt', {}) or {}
    gtc_truoc = phD.get('gtc_truoc', {}) or {}
    gtc_tu = phD.get('gtc_tu', {}) or {}
    ws.cell(row, 1).value = 'GTT'
    ws.cell(row, 1).border = Bdr()
    ws.cell(row, 2).value = _nv(gtt_h.get('soMon', 0))
    ws.cell(row, 3).value = _nv(gtt_h.get('soTien', 0))
    ws.cell(row, 4).value = None
    ws.cell(row, 5).value = None
    for ci2 in (2, 3):
        ws.cell(row, ci2).number_format = NUM
        ws.cell(row, ci2).border = Bdr()
        ws.cell(row, ci2).alignment = AL('right')
    ws.cell(row, 4).border = Bdr()
    ws.cell(row, 5).border = Bdr()
    row += 1
    ws.cell(row, 1).value = 'GTC — Trước 15h30'
    ws.cell(row, 1).border = Bdr()
    ws.cell(row, 4).value = _nv(gtc_truoc.get('soMon', 0))
    ws.cell(row, 5).value = _nv(gtc_truoc.get('soTien', 0))
    for ci2 in (1, 2, 3, 4, 5):
        ws.cell(row, ci2).border = Bdr()
    ws.cell(row, 4).number_format = NUM
    ws.cell(row, 5).number_format = NUM
    ws.cell(row, 4).alignment = AL('right')
    ws.cell(row, 5).alignment = AL('right')
    row += 1
    ws.cell(row, 1).value = 'GTC — Từ 15h30'
    ws.cell(row, 4).value = _nv(gtc_tu.get('soMon', 0))
    ws.cell(row, 5).value = _nv(gtc_tu.get('soTien', 0))
    for ci2 in (1, 2, 3, 4, 5):
        ws.cell(row, ci2).border = Bdr()
    ws.cell(row, 4).number_format = NUM
    ws.cell(row, 5).number_format = NUM
    ws.cell(row, 4).alignment = AL('right')
    ws.cell(row, 5).alignment = AL('right')
    row += 1
    ws.cell(row, 1).value = 'Tổng HUB'
    ws.cell(row, 1).font = F(bold=True)
    for ci2, v in [(2, hub['gtt']['soMon']), (3, hub['gtt']['soTien']),
                   (4, hub['gtc']['soMon']), (5, hub['gtc']['soTien'])]:
        cell = ws.cell(row, ci2)
        cell.value = v
        cell.number_format = NUM
        cell.font = F(bold=True)
        cell.alignment = AL('right')
        cell.fill = Fill('DCE6F1')
        cell.border = Bdr()
    ws.cell(row, 1).border = Bdr()
    ws.cell(row, 1).fill = Fill('DCE6F1')
    row += 2

    diff_gtt_mon = ci['gtt']['soMon'] - hub['gtt']['soMon']
    diff_gtt_tien = ci['gtt']['soTien'] - hub['gtt']['soTien']
    diff_gtc_mon = ci['gtc']['soMon'] - hub['gtc']['soMon']
    diff_gtc_tien = ci['gtc']['soTien'] - hub['gtc']['soTien']
    hcell(row, 1, 'Chênh lệch (CITAD − HUB)', fill='FFE699', color='7F0000')
    hcell(row, 2, 'GTT - Số món', fill='FFE699', color='7F0000')
    hcell(row, 3, 'GTT - Số tiền', fill='FFE699', color='7F0000')
    hcell(row, 4, 'GTC - Số món', fill='FFE699', color='7F0000')
    hcell(row, 5, 'GTC - Số tiền', fill='FFE699', color='7F0000')
    row += 1
    ws.cell(row, 1).value = ''
    ws.cell(row, 1).fill = Fill('FFE699')
    ws.cell(row, 1).border = Bdr()
    for ci2, v in [(2, diff_gtt_mon), (3, diff_gtt_tien), (4, diff_gtc_mon), (5, diff_gtc_tien)]:
        cell = ws.cell(row, ci2)
        cell.value = v if v else 0
        cell.number_format = NUM
        cell.font = F(bold=True, color='006100' if v == 0 else 'FF0000')
        cell.alignment = AL('right')
        cell.fill = Fill('FFE699')
        cell.border = Bdr()
    row += 2

    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row, 1).value = '                  LẬP BẢNG'
    ws.cell(row, 1).font = F(bold=True)
    ws.cell(row, 1).alignment = AL()
    ws.merge_cells(f'D{row}:F{row}')
    ws.cell(row, 4).value = 'KIỂM SOÁT'
    ws.cell(row, 4).font = F(bold=True)
    ws.cell(row, 4).alignment = AL()
    row += 4
    ws.cell(row, 2).value = data.lb
    ws.cell(row, 2).font = F(bold=True)
    ws.cell(row, 2).alignment = AL()
    ws.merge_cells(f'D{row}:F{row}')
    ws.cell(row, 4).value = data.ks
    ws.cell(row, 4).font = F(bold=True)
    ws.cell(row, 4).alignment = AL()

    ws.print_area = f'A1:F{row}'
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3, header=0.1, footer=0.1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
