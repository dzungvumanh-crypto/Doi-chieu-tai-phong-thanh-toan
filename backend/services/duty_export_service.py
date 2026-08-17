"""
Duty Export Service — tạo file Excel lịch trực theo tuần.

Bám theo mẫu thật `Lịch trực PTT.xlsx` do Phòng Thanh toán đang dùng:
A4 ngang, **5 cột A–E**, Times New Roman, **không màu nền** (in ra giấy trắng
đen), tiêu đề cỡ 24, tiêu đề cột 18, dữ liệu 16.

Bố cục:
    A1:E1  tiêu đề tuần            · hàng 2 để trống
    hàng 3 THỨ | NGÀY | NHÂN VIÊN (C:D gộp) | LÃNH ĐẠO
    hàng 4+ dữ liệu, mỗi ngày ĐÚNG một hàng
    ngay dưới bảng: "Ghi chú :" · cách 2 hàng: chức danh · cách 4 hàng: tên người ký
"""
from io import BytesIO
from datetime import date, timedelta
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.page import PageMargins

WEEKDAY_VI = {0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"}

_thin = Side(style="thin")
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_NCOLS = 5          # A–E, đúng bằng mẫu

FONT_NAME = "Times New Roman"
SZ_TITLE, SZ_HEADER, SZ_DATA, SZ_NOTE = 24, 18, 16, 14

# Độ rộng cột theo ĐÚNG tỉ lệ mẫu, giãn đều hệ số 1,174 cho kín bề ngang A4 xoay
# ngang. Bản gốc rộng 22,2cm trong khi vùng in là 27,2cm nên thừa 5cm dồn về một
# bên; giãn lên 26,0cm rồi bật căn giữa thì hai mép cách đều nhau.
_COL_WIDTH = {"A": 18.2, "B": 16.4, "C": 36.4, "D": 32.3, "E": 36.4}

# Chiều cao hàng phải LỚN HƠN cỡ chữ nhân 1,33 (khoảng cách dòng của Excel), nếu
# không chữ bị ép sát và tràn đè lên vạch kẻ — nhìn thành "ô thiếu kẻ phía trên".
# Mẫu để tiêu đề 30pt cho chữ 24 (cần 31,9) và đầu bảng 22,8pt cho chữ 18 (cần
# 23,9) nên cả hai đều hụt; nới ra vừa đủ, tỉ lệ tổng thể gần như không đổi.
H_TITLE, H_BLANK, H_HEADER, H_DATA = 34, 30, 26, 51

# Nhãn phụ sau tên thứ, chỉ cho loại ca KHÔNG đoán được từ cột "Thứ".
# Ca thứ 6 từng có hậu tố " (T6)" → in ra thành "T6 (T6)", vừa lặp vừa không có
# trong mẫu; đã bỏ.
_SHIFT_SUFFIX = {
    "cutoff":          " (C/O)",
    "settlement_main": " (QT)",
}


def _font(bold=False, size=SZ_DATA, italic=False) -> Font:
    return Font(name=FONT_NAME, bold=bold, size=size, italic=italic)


def _align(h="center", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _ten_lanh_dao(shift: dict) -> str:
    """Một ca có thể có nhiều Lãnh đạo — xuống dòng trong cùng một ô."""
    ten = [p.get("full_name", "") for p in (shift.get("leaders") or [])]
    if not ten and shift.get("leader"):
        ten = [shift["leader"].get("full_name", "")]
    return "\n".join(t for t in ten if t)


def _ten_nhan_vien_chinh(shift: dict) -> List[str]:
    """Trực chính, người giữ vai song phương đứng đầu."""
    sp = (shift.get("sp") or {}).get("full_name", "")
    ten = [nv.get("full_name", "") for nv in (shift.get("nvs") or [])]
    return ([sp] + ten) if sp else ten


def _o_quyet_toan(chinh: List[str], phu: List[str]):
    """
    Ô nhân viên của ngày quyết toán: trực chính IN HOA đậm, xuống dòng rồi tới
    trực phụ chữ nghiêng nhỏ hơn — hai nhóm trong CÙNG một hàng.

    Trực phụ về sớm hơn nên phải nhìn ra ngay là hai nhóm khác nhau. Dùng rich
    text vì một ô Excel chỉ có một Font, không thể trộn hai kiểu bằng cách thường.
    """
    dam    = InlineFont(rFont=FONT_NAME, sz=SZ_DATA, b=True)
    nghieng = InlineFont(rFont=FONT_NAME, sz=SZ_DATA - 3, i=True)

    khoi = []
    for i, ten in enumerate(chinh):
        khoi.append(TextBlock(dam, ("\n" if i else "") + ten.upper()))
    for ten in phu:
        khoi.append(TextBlock(nghieng, ("\n" if khoi else "") + ten))
    return CellRichText(*khoi) if khoi else None


def _apply_row(ws, row: int, values: list, bold=False, size=SZ_DATA,
               italic=False, border=True):
    """
    Ghi một hàng đúng 5 ô. Không tô nền — mẫu là trắng đen.

    Gọi hàm này TRƯỚC khi gộp ô. Ô đã gộp thì không gán được giá trị, và nếu vì
    thế mà bỏ qua luôn phần kẻ viền thì nửa phải của vùng gộp không có cạnh nào
    — Excel vẽ ra một ô hở, đúng lỗi đã gặp ở ô tiêu đề "NHÂN VIÊN". Nên ở đây
    ô đã gộp vẫn được kẻ, chỉ không ghi giá trị.
    """
    from openpyxl.cell.cell import MergedCell
    for col_idx, val in enumerate(values[:_NCOLS], start=1):
        cell = ws.cell(row=row, column=col_idx)
        da_gop = isinstance(cell, MergedCell)
        if not da_gop:
            cell.value = val if val not in ("", None) else None
            cell.font = _font(bold=bold, size=size, italic=italic)
            cell.alignment = _align(h="center" if col_idx <= 2 else "left")
        if border:
            cell.border = BORDER_ALL


def build_week_excel(shifts: list, week_start: date, week_end: date,
                     signer_name: str = "Nguyễn Quốc Hùng",
                     holiday_map: dict = None,
                     signer_title: str = "GIÁM ĐỐC") -> bytes:
    """
    Tạo file .xlsx cho 1 tuần.
    shifts: list shift dict từ duty_schedule_service._enrich_shift
    Trả về bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tuan {week_start:%d%m}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3,
    )
    # Cân bảng vào giữa bề ngang tờ giấy thay vì dồn về mép trái
    ws.print_options.horizontalCentered = True
    for col, w in _COL_WIDTH.items():
        ws.column_dimensions[col].width = w

    is_settlement_week = any(s.get("shift_type") == "settlement_main" for s in shifts)

    shift_by_date: dict = {}
    for s in shifts:
        shift_by_date.setdefault(s.get("shift_date", ""), []).append(s)

    # Tiêu đề chạy tới ngày cuối THỰC TẾ có ca, không phải tới hết khoảng quét:
    # caller quét tới chủ nhật để không bỏ sót ngày làm bù, nhưng tuần thường mà
    # đề "đến chủ nhật" thì sai với thứ người đọc nhìn thấy trong bảng.
    ngay_cuoi = week_start + timedelta(days=4)          # mặc định là thứ 6
    for ds in shift_by_date:
        try:
            d = date.fromisoformat(ds)
        except ValueError:
            continue
        if week_start <= d <= week_end and d > ngay_cuoi:
            ngay_cuoi = d
    week_end = min(week_end, max(ngay_cuoi, week_start))

    # ── Tiêu đề ────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"].value = (f"LỊCH TRỰC TỪ NGÀY {week_start:%d/%m/%Y} "
                      f"ĐẾN NGÀY {week_end:%d/%m/%Y}")
    ws["A1"].font = _font(bold=True, size=SZ_TITLE)
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = H_TITLE
    ws.row_dimensions[2].height = H_BLANK

    # ── Tiêu đề cột ───────────────────────────────────────────
    header_row = 3
    # Kẻ viền TRƯỚC rồi mới gộp — gộp trước thì nửa phải của vùng gộp không được
    # kẻ, ô "NHÂN VIÊN" hở mất cạnh trên
    _apply_row(ws, header_row, ["THỨ", "NGÀY", "NHÂN VIÊN", "", "LÃNH ĐẠO"],
               bold=True, size=SZ_HEADER)
    ws.merge_cells(f"C{header_row}:D{header_row}")
    for col in range(1, _NCOLS + 1):
        ws.cell(row=header_row, column=col).alignment = _align()
    ws.row_dimensions[header_row].height = H_HEADER

    # ── Dữ liệu: mỗi ngày ĐÚNG một hàng ───────────────────────
    current_row = header_row + 1
    current = week_start
    while current <= week_end:
        wd = current.weekday()
        date_str = current.strftime("%Y-%m-%d")
        day_shifts = shift_by_date.get(date_str, [])

        # Thứ 7 / chủ nhật chỉ lên file khi hôm đó thật sự có ca (ngày làm bù),
        # để tuần bình thường không bị thêm hai hàng trống
        if wd >= 5 and not day_shifts:
            current += timedelta(days=1)
            continue

        shift = day_shifts[0] if day_shifts else None
        thu_label = (WEEKDAY_VI.get(wd, "")
                     + _SHIFT_SUFFIX.get((shift or {}).get("shift_type", ""), ""))
        date_label = current.strftime("%d/%m/%Y")

        if shift is None:
            # CHỈ ngày đã khai nghỉ lễ mới ghi nhãn. Ngày không có ca vì lý do
            # khác (chưa xếp, thiếu người) để trống — không được ghi "Nghỉ lễ"
            # cho nó, vì như thế là nói sai: hôm đó không phải ngày được nghỉ.
            # Hỏi "hôm đó CÓ PHẢI ngày lễ không" bằng chính khoá, đừng suy từ giá
            # trị nhãn: ngày lễ khai mà bỏ trống ghi chú có nhãn rỗng/None, suy
            # theo giá trị là ô ra trắng, mất luôn chữ "Nghỉ lễ".
            text = ""
            if date_str in (holiday_map or {}):
                nhan = (holiday_map[date_str] or "").strip()
                text = f"(Nghỉ lễ: {nhan})" if nhan else "(Nghỉ lễ)"
            _apply_row(ws, current_row, [thu_label, date_label, text, "", ""],
                       italic=bool(text))
            if text:
                ws.merge_cells(f"C{current_row}:E{current_row}")
            ws.row_dimensions[current_row].height = H_DATA
            current_row += 1
            current += timedelta(days=1)
            continue

        nv_chinh = _ten_nhan_vien_chinh(shift)
        nv_phu   = [nv.get("full_name", "") for nv in (shift.get("nv_phu") or [])]
        leader   = _ten_lanh_dao(shift)

        if nv_phu:
            # Ngày quyết toán: gộp cả hai nhóm vào một ô, trực chính IN HOA đậm
            # rồi tới trực phụ nghiêng nhỏ. Giữ đúng một hàng như mẫu.
            _apply_row(ws, current_row, [thu_label, date_label, "", "", leader])
            # Ghi vào ô góc trên-trái TRƯỚC khi gộp, không thì nội dung mất
            ws.cell(row=current_row, column=3).value = _o_quyet_toan(nv_chinh, nv_phu)
            ws.merge_cells(f"C{current_row}:D{current_row}")
            so_dong = len(nv_chinh) + len(nv_phu)
        else:
            # Ngày thường: chia đôi danh sách sang hai ô nhân viên
            mid = (len(nv_chinh) + 1) // 2
            _apply_row(ws, current_row,
                       [thu_label, date_label,
                        "\n".join(nv_chinh[:mid]), "\n".join(nv_chinh[mid:]), leader])
            so_dong = max(mid, len(nv_chinh) - mid)

        ws.row_dimensions[current_row].height = max(H_DATA, 22 * max(1, so_dong))
        current_row += 1
        current += timedelta(days=1)

    # ── Ghi chú & chữ ký ──────────────────────────────────────
    # Mẫu để "Ghi chú :" ngay dưới bảng, không chen hàng trống
    note_row = current_row
    ws.cell(row=note_row, column=1).value = "Ghi chú :"
    ws.cell(row=note_row, column=1).font = _font(bold=True, size=SZ_HEADER)
    ws.cell(row=note_row, column=1).alignment = _align(h="left")

    if is_settlement_week:
        for dong in ("- Cán bộ không trực chính làm việc theo giờ của hệ thống là 19h",
                     "- Cán bộ không có tên trong lịch trực làm việc theo giờ làm việc "
                     "của Agribank"):
            ws.cell(row=note_row, column=2).value = dong
            ws.cell(row=note_row, column=2).font = _font(size=SZ_NOTE, italic=True)
            ws.cell(row=note_row, column=2).alignment = _align(h="left")
            ws.merge_cells(f"B{note_row}:E{note_row}")
            note_row += 1
        note_row -= 1

    # Chức danh cách ghi chú 2 hàng, tên người ký cách chức danh 4 hàng — theo mẫu
    row_chuc_danh = note_row + 2
    ws.cell(row=row_chuc_danh, column=5).value = signer_title or "GIÁM ĐỐC"
    ws.cell(row=row_chuc_danh, column=5).font = _font(bold=True, size=SZ_DATA)
    ws.cell(row=row_chuc_danh, column=5).alignment = _align()

    row_ten = row_chuc_danh + 4
    ws.cell(row=row_ten, column=5).value = signer_name
    ws.cell(row=row_ten, column=5).font = _font(size=SZ_HEADER)
    ws.cell(row=row_ten, column=5).alignment = _align()

    ws.print_area = f"A1:E{row_ten}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
