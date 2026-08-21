import io
import os
import shutil
import subprocess
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor
from typing import List
from datetime import datetime, timedelta

import pyzipper
import pandas as pd

from .config import zip_password
from .zip_utils import (
    find_zip_tool as _find_zip_tool,
    build_extract_cmd as _build_extract_cmd,
    detect_encoding_path as _detect_encoding_path,
    detect_encoding_from_bytes as _detect_encoding_from_bytes,
    NULL_SESSION as _NULL_SESSION,
    bao_dung_cong_cu as _bao_dung_cong_cu,
    bao_giai_nen_xong as _bao_giai_nen_xong,
    bao_lui_ve_pyzipper as _bao_lui_ve_pyzipper,
)

_TRANG_THAI_LOAI_TRU = {'CALD', 'ERPO', 'TPER'}

_COLS = [
    'NGAY_GIAO_DICH', 'CHI_NHANH', 'REFHUB', 'MSGREF', 'MSGSEQ', 'TXID',
    'KENH_THANH_TOAN', 'TRANG_THAI_LENH', 'SO_TIEN', 'TRACE',
    'SE_TRACE', 'SESSION', 'LOAI_LENH_OSB', 'NH_NHAN',
    'MA_GIAO_DICH', 'NOI_DUNG', 'NGAY_KENH_TRA',
]


def _doc_zip(zip_path: str, session_filter: str = None, log=None) -> pd.DataFrame:
    result = _find_zip_tool()
    if result:
        tool_path, tool_type = result
        _bao_dung_cong_cu('B4', zip_path, tool_type, tool_path, log)
        try:
            return _doc_zip_tool(zip_path, session_filter, tool_path, tool_type, log)
        except Exception as e:
            _bao_lui_ve_pyzipper('B4', zip_path, str(e), log)
    else:
        _bao_lui_ve_pyzipper('B4', zip_path, '', log)
    return _doc_zip_pyzipper(zip_path, session_filter, log)


def _doc_zip_tool(zip_path: str, session_filter, tool_path: str, tool_type: str,
                  log=None) -> pd.DataFrame:
    tmp_dir = tempfile.mkdtemp(prefix='ach_b4_')
    try:
        _t  = time.perf_counter()
        cmd = _build_extract_cmd(tool_path, tool_type, zip_path, tmp_dir, zip_password().decode())
        r   = subprocess.run(cmd, capture_output=True, timeout=300)
        _bao_giai_nen_xong('B4', zip_path, time.perf_counter() - _t, r.returncode, log)
        if r.returncode != 0:
            raise RuntimeError(r.stderr.decode(errors='replace'))
        frames = []
        for name in sorted(os.listdir(tmp_dir)):
            if not name.lower().endswith('.csv'):
                continue
            path = os.path.join(tmp_dir, name)
            enc  = _detect_encoding_path(path)
            if log:
                log(f'[B4] Đang đọc {name} ({os.path.getsize(path) / 1048576:.0f} MB)...')
            if session_filter:
                sid  = str(session_filter)
                keep = frozenset({sid} | _NULL_SESSION)
                chunk_list = []
                for chunk in pd.read_csv(
                    path, dtype=str, encoding=enc,
                    usecols=lambda c: c in _COLS,
                    chunksize=200_000, low_memory=False,
                ):
                    if 'SESSION' in chunk.columns:
                        sess = chunk['SESSION'].fillna('').astype(str).str.strip().str.lstrip("'")
                        chunk_list.append(chunk[sess.isin(keep)])
                if chunk_list:
                    frames.append(pd.concat(chunk_list, ignore_index=True))
            else:
                frames.append(pd.read_csv(
                    path, dtype=str, encoding=enc,
                    usecols=lambda c: c in _COLS, low_memory=False,
                ))
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=_COLS)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _doc_zip_pyzipper(zip_path: str, session_filter: str = None, log=None) -> pd.DataFrame:
    frames = []
    with pyzipper.AESZipFile(zip_path, 'r') as z:
        z.setpassword(zip_password())
        for name in z.namelist():
            if not name.lower().endswith('.csv'):
                continue
            if log:
                log(f'[B4] Đang nạp {name} vào bộ nhớ (cách dự phòng)...')
            data = z.read(name)
            enc  = _detect_encoding_from_bytes(data[:512])
            df   = pd.read_csv(
                io.BytesIO(data), dtype=str, encoding=enc,
                usecols=lambda c: c in _COLS, low_memory=False,
                encoding_errors='replace',
            )
            if session_filter:
                sid  = str(session_filter)
                keep = frozenset({sid} | _NULL_SESSION)
                if 'SESSION' in df.columns:
                    sess = df['SESSION'].fillna('').astype(str).str.strip().str.lstrip("'")
                    df   = df[sess.isin(keep)]
            if not df.empty:
                frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=_COLS)


def _tao_so_trace(df: pd.DataFrame) -> pd.Series:
    se = df['SE_TRACE'].fillna('').astype(str).str.strip().str.lstrip("'0")
    tr = df['TRACE'].fillna('').astype(str).str.strip().str.lstrip("'0")
    return se.where(se.ne(''), tr)


def _chuan_hoa_co_ban(df: pd.DataFrame) -> pd.DataFrame:
    """Chuẩn hoá SO_TIEN/SO_TRACE/NGAY_KENH_TRA trên MIS_đi thô — dùng chung cho
    `_process_mis_di()`, `tim_giao_dich_bi_loai_session_null()`, và dòng bổ sung
    thủ công ở `ap_dung_confirm_mis_di()` (Bước 2 Checkpoint MIS_đi)."""
    df = df.copy()
    df['SO_TIEN']  = pd.to_numeric(df['SO_TIEN'], errors='coerce').fillna(0).astype('int64')
    df['SO_TRACE'] = _tao_so_trace(df)
    df['NGAY_KENH_TRA'] = pd.to_datetime(
        df['NGAY_KENH_TRA'].str.strip(), format='%d/%m/%Y %H:%M:%S', errors='coerce'
    )
    return df


def _doc_mis_di_raw(zip_paths: List[str], session_id: str, log_callback=None) -> pd.DataFrame:
    """Đọc 2 ZIP MIS_DI song song, trả về DataFrame thô. Dùng cho parallel I/O."""
    _log = log_callback or print
    _log('[B4] Đọc MIS_DI từ 2 ZIP...')
    sid = str(session_id)
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = [ex.submit(_doc_zip, p, sid, _log) for p in zip_paths]
        frames  = [f.result() for f in futures]
    df = pd.concat(frames, ignore_index=True)
    _log(f'[B4] Đọc xong MIS_DI: {len(df):,} dòng khớp session {sid}.')
    return df


def doc_mis_di_khong_loc_session(zip_paths: List[str], log_callback=None) -> pd.DataFrame:
    """Đọc nhiều ZIP MIS_DI KHÔNG lọc theo session (khác `_doc_mis_di_raw()`) —
    dùng riêng cho tra cứu REFHUB bổ sung từ NGÀY KHÁC ở Checkpoint Bước 2
    (2026-08-04, xem `ap_dung_confirm_mis_di()`): người chấm chủ động xác nhận 1
    giao dịch cụ thể theo đúng REFHUB, không cần khớp session của ngày đang chạy."""
    _log = log_callback or print
    _log(f'[B4] Đọc {len(zip_paths)} file MIS_DI (không lọc session, tra REFHUB ngày khác)...')
    if not zip_paths:
        return pd.DataFrame(columns=_COLS)
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = [ex.submit(_doc_zip, p, None, _log) for p in zip_paths]
        frames  = [f.result() for f in futures]
    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=_COLS)
    _log(f'[B4] Đọc xong (không lọc session): {len(df):,} dòng.')
    return df


def _them_cot_khoa(df: pd.DataFrame) -> pd.DataFrame:
    """Thêm cột KEY_HUB (đối chiếu NPO) và 'CN tiền Hub' (đối chiếu GW) — dùng chung
    cho cả MIS_đi lẫn nhóm timeout theo trạng thái."""
    cn_clean       = df['CHI_NHANH'].astype(str).str.strip()
    df['KEY_HUB']  = cn_clean + df['SO_TRACE'] + df['SO_TIEN'].astype(str)
    cn_tien        = cn_clean + df['SO_TIEN'].astype(str)
    loc            = df.columns.get_loc('CHI_NHANH') + 1
    df.insert(loc, 'CN tiền Hub', cn_tien)
    return df


def _xay_dung_tra_cuu_gw_goc(df_gw_goc: pd.DataFrame) -> dict:
    """BR mới (2026-07-23, xử lý SESSION=NULL) — dict MSGREF (đã chuẩn hóa) ->
    SessionId (đã chuẩn hóa) tra trên GW gốc. 1 MSGREF chỉ có 1 giá trị đại diện
    (đã chốt với Business Owner); nếu GW gốc thực tế có ≥2 SessionId khác nhau cho
    cùng MSGREF (dữ liệu chưa dedup theo nhiều sheet) → marker '__NHIEU__', KHÔNG
    tự chọn 1 giá trị theo vị trí (đúng nguyên tắc đã áp dụng cho C.1/C.2)."""
    if df_gw_goc is None or len(df_gw_goc) == 0:
        return {}
    msgref = _chuan_hoa_msgref(df_gw_goc['MSGREF'])
    sess   = df_gw_goc['SessionId'].fillna('').astype(str).str.strip()
    sess   = sess.mask(sess.str.lower().isin(['nan', 'none']), '')
    grouped = pd.DataFrame({'MSGREF': msgref, 'SessionId': sess}).groupby('MSGREF')['SessionId'].unique()
    return {m: (v[0] if len(v) == 1 else '__NHIEU__') for m, v in grouped.items()}


def _loc_session_null_theo_gw_goc(df_null: pd.DataFrame, session_id: str,
                                  ngay_dt: datetime, df_gw_goc: pd.DataFrame):
    """BR mới (2026-07-23) — thay thế hoàn toàn logic khung giờ cũ cho giao dịch
    MIS_đi có SESSION=NULL. Trả về (mask_giu, ly_do), cùng index với `df_null`.

    Bước 1: tra SessionId thật của MSGREF trên GW gốc (chưa lọc session/PrcFlg).
    Bước 2: so NGAY_GIAO_DICH (ngày giá trị) với ngày đối chiếu T:
    - NGAY_GIAO_DICH = T: giữ nếu SessionId tìm được thuộc {rỗng, '0000', session
      đối chiếu}, hoặc MSGREF không có trên GW gốc (đánh dấu riêng để chấm thủ công
      phân biệt với "GW có nhưng SessionId rỗng").
    - NGAY_GIAO_DICH = T-1: CHỈ giữ nếu SessionId = session đối chiếu.
    - Khác cả T và T-1: giữ tạm — Business Owner xác nhận (2026-07-23) chưa có quy
      tắc chính thức cho trường hợp này, đánh dấu riêng để xem xét sau.
    GW gốc có ≥2 SessionId khác nhau cho cùng MSGREF → không xác định, giữ + đánh
    dấu riêng (không tự chọn 1 giá trị).

    Cập nhật 2026-08-04 (audit toàn diện — không thông báo giao dịch bị loại ngầm):
    trước đây các trường hợp bị LOẠI (mask_giu=False) tại T-1 (không tìm thấy/session
    rỗng/'0000'), và trường hợp tìm thấy trên GW gốc nhưng là 1 SessionId thật KHÁC
    (không rỗng/'0000'/đúng phiên) tại CẢ T lẫn T-1, đều rơi qua hết 6 nhánh trên mà
    KHÔNG được gán `ly_do` (giữ chuỗi rỗng) — không có cách nào biết SAU NÀY vì sao
    1 giao dịch cụ thể bị loại. Hành vi GIỮ/LOẠI (`mask_giu`) hoàn toàn KHÔNG đổi
    (case "khác session" đã được Business Owner xác nhận là loại có chủ đích,
    25/07/2026) — chỉ bổ sung nhãn để không còn "vô hình" trong báo cáo/chấm tay.
    """
    tra_cuu = _xay_dung_tra_cuu_gw_goc(df_gw_goc)
    sid = str(session_id).strip()

    msgref_norm = _chuan_hoa_msgref(df_null['MSGREF'])
    session_gw  = msgref_norm.map(tra_cuu)

    ngay_gd = pd.to_datetime(df_null['NGAY_GIAO_DICH'].str.strip(), format='%d/%m/%Y', errors='coerce')
    ngay_T  = ngay_dt.date()
    ngay_T1 = (ngay_dt - timedelta(days=1)).date()
    la_ngay_T  = ngay_gd.dt.date == ngay_T
    la_ngay_T1 = ngay_gd.dt.date == ngay_T1
    la_khac    = ~la_ngay_T & ~la_ngay_T1

    khong_tim_thay = session_gw.isna()
    nhieu_gia_tri  = session_gw == '__NHIEU__'
    la_rong        = session_gw == ''
    la_0000        = session_gw == '0000'
    la_dung        = session_gw == sid

    mask_giu = pd.Series(False, index=df_null.index)
    ly_do    = pd.Series('', index=df_null.index)

    m = la_khac
    mask_giu |= m; ly_do = ly_do.mask(m, 'NGAY_GIA_TRI_KHAC_T_VA_T-1')

    m = ~la_khac & nhieu_gia_tri
    mask_giu |= m; ly_do = ly_do.mask(m, 'GW_GOC_NHIEU_SESSIONID_KHAC_NHAU')

    m = la_ngay_T & ~nhieu_gia_tri & khong_tim_thay
    mask_giu |= m; ly_do = ly_do.mask(m, 'KHONG_TIM_THAY_TREN_GW')

    m = la_ngay_T & ~nhieu_gia_tri & ~khong_tim_thay & la_rong
    mask_giu |= m; ly_do = ly_do.mask(m, 'GW_SESSION_NULL')

    m = la_ngay_T & ~nhieu_gia_tri & ~khong_tim_thay & la_0000
    mask_giu |= m; ly_do = ly_do.mask(m, 'GW_SESSION_0000')

    m = (la_ngay_T | la_ngay_T1) & ~nhieu_gia_tri & ~khong_tim_thay & la_dung
    mask_giu |= m; ly_do = ly_do.mask(m, 'GW_SESSION_DOI_CHIEU')

    # Mọi trường hợp còn lại (T hoặc T-1) không khớp bất kỳ nhánh GIỮ nào ở trên đều
    # đã bị loại đúng theo BR (T-1 chỉ giữ khi đúng session; "khác session thật" tại
    # T/T-1 là loại có chủ đích) — gán nhãn CHUNG để không còn rỗng, tách biệt case cụ
    # thể "tìm thấy nhưng khác session" cho dễ chấm tay hơn case "không tìm thấy".
    m = ~la_khac & ~mask_giu & ~khong_tim_thay
    ly_do = ly_do.mask(m, 'GW_SESSION_KHAC')

    m = ~la_khac & ~mask_giu & khong_tim_thay
    ly_do = ly_do.mask(m, 'KHONG_TIM_THAY_TREN_GW_TAI_T-1')

    return mask_giu, ly_do


def _process_mis_di(df: pd.DataFrame, session_id: str, ngay_dt: datetime,
                    df_gw_goc: pd.DataFrame, log_callback=None):
    """Mục 2 tài liệu đối chiếu — xây "MIS_đi sạch" từ MIS_Hub thô.

    Bước 1 (TRẠNG THÁI): bỏ toàn bộ CALD/ERPO/TPER khỏi MIS_đi — loại hẳn, không
    dùng lại các trạng thái này cho bất kỳ mục đích nào khác (BR chính thức, xem
    project_ach_timeout_rule.md — thay thế cơ chế "Nguồn 1" cũ đã gỡ).

    Bước 2 (SESSION, áp dụng cho phần còn lại sau bước 1):
    - SESSION khác NULL: giữ nguyên cách xử lý hiện tại — chỉ lấy đúng session đối
      chiếu.
    - SESSION = NULL: BR mới (2026-07-23) — KHÔNG còn lọc theo khung giờ nữa, xem
      `_loc_session_null_theo_gw_goc()`.

    Bước 3 (trong `_them_cot_khoa()`): tạo khóa CN TIỀN = CHI_NHANH + SO_TIEN.

    Trả về df_mis_di — có thêm cột `LY_DO_GIU_SESSION_NULL` (rỗng với dòng SESSION
    khác NULL, dùng để chấm thủ công phân biệt lý do giữ giao dịch SESSION=NULL).
    """
    _log = log_callback or print
    sid  = str(session_id)

    df = _chuan_hoa_co_ban(df)
    df['SESSION']      = df['SESSION'].fillna('').astype(str).str.strip().str.lstrip("'")
    df['SESSION_NULL'] = df['SESSION'].isin(['', 'nan', 'None', 'NaN'])

    # Bước 1 — TRẠNG THÁI: bỏ CALD/ERPO/TPER khỏi MIS_đi.
    mask_trang_thai_loai_tru = df['TRANG_THAI_LENH'].isin(_TRANG_THAI_LOAI_TRU)

    # Bước 2 — SESSION.
    mask_khac_null = ~df['SESSION_NULL'] & (df['SESSION'] == sid)

    df['LY_DO_GIU_SESSION_NULL'] = ''
    mask_session_null = pd.Series(False, index=df.index)
    df_null = df[df['SESSION_NULL']]
    if len(df_null) > 0:
        mask_null_giu, ly_do_null = _loc_session_null_theo_gw_goc(
            df_null, session_id, ngay_dt, df_gw_goc,
        )
        df.loc[df_null.index, 'LY_DO_GIU_SESSION_NULL'] = ly_do_null
        mask_session_null.loc[df_null.index] = mask_null_giu

    mask_session = mask_khac_null | mask_session_null

    df_mis_di = _them_cot_khoa(df[~mask_trang_thai_loai_tru & mask_session].copy())

    _log(f'[B4] MIS_đi (bước 1+2): {len(df_mis_di):,} dòng '
         f'(SESSION=NULL giữ lại: {int((mask_session_null & ~mask_trang_thai_loai_tru).sum()):,})')
    return df_mis_di.reset_index(drop=True)


def _chuan_hoa_msgref(s: pd.Series) -> pd.Series:
    return s.fillna('').astype(str).str.strip().str.lstrip("'")


_COLS_CAN_KIEM_TRA_THU_CONG = [
    'NGAY_GIAO_DICH', 'CHI_NHANH', 'CN tiền Hub', 'REFHUB', 'MSGREF', 'TRANG_THAI_LENH',
    'SO_TIEN', 'SESSION', 'NGAY_KENH_TRA', 'LY_DO_CAN_KIEM_TRA',
]


def tim_giao_dich_bi_loai_session_null(df: pd.DataFrame, session_id: str, ngay_dt: datetime,
                                       df_gw_goc: pd.DataFrame, log_callback=None) -> pd.DataFrame:
    """Điểm 1 (2026-07-31, xem project_ach_4diem_pr_plan) — thay cho
    `tim_can_kiem_tra_thu_cong()` cũ (Milestone F Option C, 2 nhánh). Nhánh
    "NGAY_GIA_TRI_KHAC_T_VA_T-1" đã BỎ khỏi hàm này vì giờ tự nhiên hiển thị trong
    file confirm MIS_đi (vẫn nằm trong MIS_đi, người chấm thấy trực tiếp — xem
    `ap_dung_confirm_mis_di()`), không cần hiển thị lại riêng nữa.

    Chỉ còn đúng 1 nhánh: MSGREF rỗng hoặc không tra được SessionId trên GW gốc tại
    NGAY_GIAO_DICH=T-1 — nhánh này ĐANG bị `_process_mis_di()` loại thẳng khỏi
    MIS_đi (không đổi hành vi loại đó). Hàm này tạo BẢN SAO độc lập để hiển thị CHỈ
    ĐỂ XEM (sheet phụ trong file confirm MIS_đi, xem
    `pipeline.py::xuat_excel_confirm_mis_di()`) — không có cột chấm (vì dòng này
    không thuộc MIS_đi nên không có gì để "loại bỏ"), không ảnh hưởng gì tới luồng
    chính.

    Trả về DataFrame (có thể rỗng) — KHÔNG rỗng nghĩa là còn giao dịch bị loại cần
    người chấm biết, không phải dấu hiệu lỗi.
    """
    _log = log_callback or print

    df = _chuan_hoa_co_ban(df)
    df['SESSION']      = df['SESSION'].fillna('').astype(str).str.strip().str.lstrip("'")
    df['SESSION_NULL'] = df['SESSION'].isin(['', 'nan', 'None', 'NaN'])

    # CALD/ERPO/TPER không bao giờ vào diện hiển thị (BR đã chốt, không đổi).
    mask_trang_thai_loai_tru = df['TRANG_THAI_LENH'].isin(_TRANG_THAI_LOAI_TRU)
    df_null = df[df['SESSION_NULL'] & ~mask_trang_thai_loai_tru]

    if len(df_null) == 0:
        _log('[Điểm 1] Giao dịch bị loại (SESSION=NULL, không tra được GW tại T-1): 0')
        return _them_cot_khoa(df_null.copy())[_COLS_CAN_KIEM_TRA_THU_CONG[:-1]].assign(LY_DO_CAN_KIEM_TRA='')

    mask_giu, _ = _loc_session_null_theo_gw_goc(df_null, session_id, ngay_dt, df_gw_goc)

    ngay_gd     = pd.to_datetime(df_null['NGAY_GIAO_DICH'].str.strip(), format='%d/%m/%Y', errors='coerce')
    ngay_T1     = (ngay_dt - timedelta(days=1)).date()
    la_ngay_T1  = ngay_gd.dt.date == ngay_T1
    msgref_norm = _chuan_hoa_msgref(df_null['MSGREF'])
    tra_cuu     = _xay_dung_tra_cuu_gw_goc(df_gw_goc)
    session_gw  = msgref_norm.map(tra_cuu)
    khong_tim_thay = session_gw.isna()

    mask_bi_loai = la_ngay_T1 & khong_tim_thay & ~mask_giu
    la_rong      = msgref_norm == ''
    ly_do = pd.Series('', index=df_null.index)
    ly_do = ly_do.mask(mask_bi_loai & la_rong, 'MSGREF_RONG_TAI_T-1')
    ly_do = ly_do.mask(mask_bi_loai & ~la_rong, 'MSGREF_KHONG_TIM_THAY_TREN_GW_TAI_T-1')

    df_ket_qua = _them_cot_khoa(df_null[mask_bi_loai].copy())
    df_ket_qua['LY_DO_CAN_KIEM_TRA'] = ly_do[mask_bi_loai].values

    _log(f'[Điểm 1] Giao dịch bị loại (SESSION=NULL, không tra được GW tại T-1): {len(df_ket_qua):,}')
    return df_ket_qua[_COLS_CAN_KIEM_TRA_THU_CONG].reset_index(drop=True)


def tim_toan_bo_giao_dich_bi_loai_session_null(df: pd.DataFrame, session_id: str, ngay_dt: datetime,
                                               df_gw_goc: pd.DataFrame, log_callback=None) -> pd.DataFrame:
    """Audit 2026-08-04 — bản MỞ RỘNG của `tim_giao_dich_bi_loai_session_null()`:
    hàm cũ chỉ bắt lại đúng 1 nhánh loại (T-1, không tìm thấy trên GW gốc) để hiển
    thị TẠM trong file confirm MIS_đi lúc Checkpoint (`dung_sau_mis_di=True`) — các
    nhánh loại khác (T-1 rỗng/'0000', "khác session thật" tại cả T/T-1) hoàn toàn
    KHÔNG xuất hiện ở BẤT KỲ báo cáo nào, kể cả báo cáo CUỐI `doi_chieu_<ngày>.xlsx`,
    vì `tim_giao_dich_bi_loai_session_null()` chỉ được gọi có điều kiện.

    Hàm này trả về TOÀN BỘ giao dịch SESSION=NULL bị `_process_mis_di()` loại
    (mọi lý do, dùng đúng `ly_do` đã hoàn thiện ở `_loc_session_null_theo_gw_goc()`
    — xem cập nhật 2026-08-04 ở đó), để gọi UNCONDITIONALLY trong `main_from_dir()`
    và luôn xuất hiện trong báo cáo cuối — không còn giao dịch nào biến mất hoàn
    toàn khỏi mọi báo cáo. KHÔNG thay thế `tim_giao_dich_bi_loai_session_null()`
    (vẫn giữ nguyên cho file confirm Checkpoint, có nhãn chi tiết hơn cho riêng case
    T-1/MSGREF rỗng).

    Trả về DataFrame (có thể rỗng).
    """
    _log = log_callback or print

    df = _chuan_hoa_co_ban(df)
    df['SESSION']      = df['SESSION'].fillna('').astype(str).str.strip().str.lstrip("'")
    df['SESSION_NULL'] = df['SESSION'].isin(['', 'nan', 'None', 'NaN'])

    mask_trang_thai_loai_tru = df['TRANG_THAI_LENH'].isin(_TRANG_THAI_LOAI_TRU)
    df_null = df[df['SESSION_NULL'] & ~mask_trang_thai_loai_tru]

    if len(df_null) == 0:
        _log('[Audit] Giao dịch SESSION=NULL bị loại khỏi MIS_đi (mọi lý do): 0')
        return _them_cot_khoa(df_null.copy())[_COLS_CAN_KIEM_TRA_THU_CONG[:-1]].assign(LY_DO_CAN_KIEM_TRA='')

    mask_giu, ly_do = _loc_session_null_theo_gw_goc(df_null, session_id, ngay_dt, df_gw_goc)
    mask_bi_loai = ~mask_giu

    df_ket_qua = _them_cot_khoa(df_null[mask_bi_loai].copy())
    df_ket_qua['LY_DO_CAN_KIEM_TRA'] = ly_do[mask_bi_loai].values

    _log(f'[Audit] Giao dịch SESSION=NULL bị loại khỏi MIS_đi (mọi lý do): {len(df_ket_qua):,}')
    return df_ket_qua[_COLS_CAN_KIEM_TRA_THU_CONG].reset_index(drop=True)


def khop_voi_gw(df_mis_di: pd.DataFrame, dict_gw_count: dict, df_gw: pd.DataFrame,
               log_callback=None):
    """Mục 3 tài liệu đối chiếu — so khớp CN TIỀN (CHI_NHANH+SO_TIEN) giữa MIS_đi và GW.

    BR-ACH-001 (chốt 2026-07-21, qua nhiều vòng xác nhận với Business Owner) — thứ tự
    xử lý BẮT BUỘC, không được đảo:

    Requirement C.1 — xác định NHÓM CN_TIỀN chênh lệch (mức NHÓM, không chọn dòng):
    so tổng COUNT_MIS (mọi trạng thái) với COUNT_GW của cả nhóm. KHÔNG dùng cumcount
    để chọn ra dòng cụ thể nào trong nhóm — đó là suy luận không có căn cứ khi 1 nhóm
    có nhiều dòng trùng CN_TIỀN (y hệt vấn đề đã tránh ở Trường hợp 2/GW-thừa, xem
    `tim_nhom_gw_thua()`).

    Requirement C.2 — CHỈ trong các nhóm đã xác định thừa ở C.1, tra MSGREF (đã chuẩn
    hóa — bỏ dấu nháy đơn đầu) cho TỪNG dòng để phân loại (không so MSGREF trên toàn
    bộ dữ liệu trước khi có nhóm chênh lệch):
    - MSGREF KHÔNG có trên GW sạch → nhánh 1A, "Timeout không đi kênh" → df_timeout.
    - MSGREF CÓ trên GW sạch → nhánh 1B, "Timeout thật" (đã được kênh xác nhận, KHÔNG
      phải thừa) → df_khop_dung, đánh dấu cột `MATCH_TYPE = 'TIMEOUT'`.
    Tuyệt đối KHÔNG dùng `PrcFlg` của GW để quyết định — chỉ dựa vào MSGREF có tồn
    tại hay không, bất kể GW đang ở trạng thái gì.

    Phạm vi trạng thái xét ở C.2 (chốt 2026-08-03, thay quy tắc TPAY-only cũ): TOÀN
    BỘ trạng thái còn lại trong nhóm thừa đều được xét như TPAY — vì CALD/ERPO/TPER
    (biến `_TRANG_THAI_LOAI_TRU`) đã bị loại hẳn ở Bước 1 (`_process_mis_di()`), không
    cần lọc lại theo trạng thái ở đây nữa. Business Owner xác nhận sẽ có nhiều trạng
    thái khác ngoài TPAY phát sinh theo thời gian (đã thấy TXPR/TXCA thực tế) và muốn
    xử lý đồng nhất, không liệt kê danh sách trắng cố định.

    Trả về (df_khop_dung, df_timeout).
    """
    _log   = log_callback or print
    cn_col = 'CN tiền Hub'

    # C.1 — nhóm nào thừa (mức nhóm, KHÔNG chọn dòng).
    cnt_mis_nhom   = df_mis_di.groupby(cn_col, sort=False)[cn_col].transform('size')
    cnt_gw_nhom    = df_mis_di[cn_col].map(dict_gw_count).fillna(0).astype(int)
    mask_nhom_thua = cnt_mis_nhom > cnt_gw_nhom

    # C.2 — trong nhóm thừa, tra MSGREF cho MỌI dòng (CALD/ERPO/TPER đã loại ở Bước 1).
    msgref_mis    = _chuan_hoa_msgref(df_mis_di['MSGREF'])
    msgref_gw_set = frozenset(_chuan_hoa_msgref(df_gw['MSGREF']))

    mask_xet_trong_nhom_thua = mask_nhom_thua
    mask_co_tren_gw          = msgref_mis.isin(msgref_gw_set)

    mask_timeout      = mask_xet_trong_nhom_thua & ~mask_co_tren_gw   # nhánh 1A
    mask_timeout_that = mask_xet_trong_nhom_thua & mask_co_tren_gw    # nhánh 1B

    df = df_mis_di.copy()
    df['MATCH_TYPE'] = ''
    df.loc[mask_timeout_that, 'MATCH_TYPE'] = 'TIMEOUT'

    df_timeout   = df[mask_timeout].drop(columns=['MATCH_TYPE']).copy()
    df_khop_dung = df[~mask_timeout].copy()

    _log(
        f'[B4][Mục 3] Nhóm CN_TIỀN thừa: {int(mask_nhom_thua.sum()):,} dòng | '
        f'Timeout không đi kênh (nhánh 1A): {len(df_timeout):,} | '
        f'Timeout thật đã khớp GW (nhánh 1B): {int(mask_timeout_that.sum()):,} | '
        f'Khớp đúng (chuyển sang đối chiếu NPO): {len(df_khop_dung):,}'
    )
    return df_khop_dung.reset_index(drop=True), df_timeout.reset_index(drop=True)


def tim_nhom_gw_thua(df_mis_di: pd.DataFrame, df_gw: pd.DataFrame, log_callback=None):
    """Requirement C.1a (Change Plan, Trường hợp 2 của BR-ACH-001) — xác định các
    NHÓM CN_TIỀN có GW nhiều hơn MIS_đi (COUNT_GW > COUNT_MIS), trả về DỮ LIỆU GỐC
    (đầy đủ cột) để người dùng chấm thủ công thẳng, không phải tự map ngược lại.

    Chỉ dựa vào việc bên MIS_đi có bản ghi nào không cho mỗi nhóm CN_TIỀN (KHÔNG
    dùng `PrcFlg`/`SessionId`, KHÔNG ghép cặp MSGREF hay bất kỳ kỹ thuật suy luận
    nào khác — đúng nguyên tắc BR-ACH-001 "không được tự suy luận"):
    - COUNT_MIS == 0 (nhóm chỉ tồn tại ở GW, không trùng khóa với bên nào) → xác
      định CHẮC CHẮN toàn bộ dòng GW của nhóm đó là chênh lệch (Trường hợp 1 — đã
      làm rõ 2026-07-21).
    - COUNT_MIS >= 1 và COUNT_GW > COUNT_MIS (cả 2 bên đều có, lệch số lượng) →
      KHÔNG xác định được bản ghi cụ thể nào là thừa (Trường hợp 2) → xuất TOÀN BỘ
      dòng gốc của CẢ 2 nguồn thuộc nhóm CN_TIỀN đó.

    Trả về (df_xac_dinh, df_can_doi_chieu) — cả 2 đều là dòng dữ liệu GỐC (nguyên
    cột từ df_gw/df_mis_di), kèm thêm cột `SOURCE` ('GW'/'MIS') và `NHOM_CN_TIEN`
    (khóa CN_TIỀN, để gom/lọc theo nhóm khi đối chiếu thủ công).
    """
    _log    = log_callback or print
    mis_col = 'CN tiền Hub'
    gw_col  = 'KEY_GW'

    cnt_mis  = df_mis_di.groupby(mis_col, sort=False).size()
    cnt_gw   = df_gw.groupby(gw_col, sort=False).size()
    all_keys = set(cnt_mis.index) | set(cnt_gw.index)
    cnt_mis  = cnt_mis.reindex(all_keys, fill_value=0)
    cnt_gw   = cnt_gw.reindex(all_keys, fill_value=0)

    gw_thua_keys       = [k for k in all_keys if cnt_gw[k] > cnt_mis[k]]
    keys_xac_dinh      = [k for k in gw_thua_keys if cnt_mis[k] == 0]
    keys_can_doi_chieu = [k for k in gw_thua_keys if cnt_mis[k] > 0]

    def _gan_nguon(df, key_col, keys, source_label):
        sub = df[df[key_col].isin(keys)].copy()
        sub['SOURCE']       = source_label
        sub['NHOM_CN_TIEN'] = sub[key_col]
        return sub

    df_xac_dinh = _gan_nguon(df_gw, gw_col, keys_xac_dinh, 'GW')

    df_can_doi_chieu = pd.concat([
        _gan_nguon(df_gw,     gw_col,  keys_can_doi_chieu, 'GW'),
        _gan_nguon(df_mis_di, mis_col, keys_can_doi_chieu, 'MIS'),
    ], ignore_index=True)

    _log(
        f'[B4][Mục 3 - C.1a] GW-thừa đã xác định: {len(keys_xac_dinh):,} nhóm '
        f'({len(df_xac_dinh):,} dòng) | GW-thừa cần đối chiếu thủ công: '
        f'{len(keys_can_doi_chieu):,} nhóm ({len(df_can_doi_chieu):,} dòng)'
    )
    return df_xac_dinh.reset_index(drop=True), df_can_doi_chieu.reset_index(drop=True)


# ─── Bước 2 — Checkpoint xác nhận thủ công tại MIS_đi: đọc file confirm (Bước 1) ──

_GIA_TRI_LOAI_BO_HOP_LE = {'', 'loại bỏ'}


def _doc_sheet_confirm_mis_di(xac_nhan_path: str):
    """Đọc sheet MIS_DI_CONFIRM của file confirm MIS_đi do
    `pipeline.py::xuat_excel_confirm_mis_di()` (Bước 1) sinh ra. Trả về
    (df_confirm, refhub_bo_sung) — df có cột LOAI_BO (vùng dữ liệu chính, phía trên
    dòng ghi chú "BỔ SUNG..."); refhub_bo_sung là list các REFHUB paste vào vùng bổ
    sung bên dưới.

    Dùng openpyxl trực tiếp (không phải pd.read_excel) — pandas yêu cầu
    openpyxl>=3.1.5 để đọc qua engine='openpyxl', nhưng dự án pin 3.1.2."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(xac_nhan_path, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Không đọc được file xác nhận: {xac_nhan_path} ({e})") from e

    if 'MIS_DI_CONFIRM' not in wb.sheetnames:
        raise ValueError(f"File xác nhận thiếu sheet 'MIS_DI_CONFIRM': {xac_nhan_path}")

    rows = [list(r) for r in wb['MIS_DI_CONFIRM'].iter_rows(values_only=True)]
    if not rows:
        raise ValueError(f'Sheet MIS_DI_CONFIRM rỗng: {xac_nhan_path}')

    o_dau = rows[0][0] if rows[0] else None
    chi_1_cot = all(len(r) <= 1 or all(v is None for v in r[1:]) for r in rows)
    if chi_1_cot and isinstance(o_dau, str) and 'mis_đi rỗng' in o_dau.lower():
        return pd.DataFrame(columns=['REFHUB', 'LOAI_BO']), []

    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    if 'LOAI_BO' not in header:
        raise ValueError(
            f"Sheet MIS_DI_CONFIRM thiếu cột LOAI_BO — file có thể bị sửa cấu trúc: {xac_nhan_path}"
        )
    if 'REFHUB' not in header:
        raise ValueError(
            f"Sheet MIS_DI_CONFIRM thiếu cột REFHUB — file có thể bị sửa cấu trúc: {xac_nhan_path}"
        )

    note_idx = None
    for i in range(1, len(rows)):
        cell0 = rows[i][0] if rows[i] else None
        if isinstance(cell0, str) and cell0.strip().upper().startswith('BỔ SUNG'):
            note_idx = i
            break
    if note_idx is None:
        raise ValueError(
            "Không tìm thấy vùng 'BỔ SUNG GIAO DỊCH BỊ BỎ SÓT' trong sheet MIS_DI_CONFIRM — "
            f"file có thể bị sửa cấu trúc: {xac_nhan_path}"
        )

    n_col = len(header)
    data_rows = [
        (r + [None] * (n_col - len(r)))[:n_col]
        for r in rows[1:note_idx]
        if any(v is not None for v in r)
    ]
    df = pd.DataFrame(data_rows, columns=header)

    refhub_bo_sung = []
    for r in rows[note_idx + 2:]:
        v = r[0] if r else None
        if v is not None and str(v).strip() != '':
            refhub_bo_sung.append(str(v).strip())

    return df, refhub_bo_sung


def ap_dung_confirm_mis_di(xac_nhan_path: str, df_mis_di: pd.DataFrame,
                          df_mis_di_raw: pd.DataFrame, log_callback=None,
                          doc_them_ngay_khac=None) -> pd.DataFrame:
    """Bước 2 — Checkpoint xác nhận thủ công tại MIS_đi (Điểm 1, 2026-07-31 — thay
    hẳn cơ chế Timeout-confirm cũ, xem project_ach_4diem_pr_plan). File confirm
    (Bước 1, `xuat_excel_confirm_mis_di()`) cho người chấm 2 việc trên MIS_đi (đầu
    ra bước 5, TRƯỚC khi so khớp GW):
    - Cột LOAI_BO trên từng dòng: '' (mặc định, giữ) hoặc 'loại bỏ' (bỏ dòng khỏi
      MIS_đi).
    - Khu vực bổ sung: dán REFHUB của giao dịch bị lọc oan (có trên MIS_hub thô,
      chưa có trong MIS_đi) — tra trên `df_mis_di_raw` để kéo về. KHÔNG áp ràng
      buộc 3 trạng thái loại cố định (CALD/ERPO/TPER) cho phần bổ sung (khác cơ chế
      Timeout-confirm cũ) — trạng thái tại thời điểm xuất file chưa phải trạng thái
      cuối cùng, lệnh người chấm thêm vào luôn được tôn trọng tuyệt đối.

    Công thức: MIS_đi chuẩn = (MIS_đi ban đầu − dòng LOAI_BO='loại bỏ') + (REFHUB
    hợp lệ ở khu bổ sung). REFHUB trùng ≥2 dòng trên MIS_hub thô → báo lỗi, từ chối
    chạy tiếp (không tự chọn 1 dòng theo vị trí).

    doc_them_ngay_khac — callback KHÔNG tham số (2026-08-04, mở rộng theo yêu cầu
    Business Owner: cần thêm giao dịch timeout của vài ngày trước vào báo cáo ngày
    hiện tại), trả về DataFrame MIS_đi thô bổ sung từ NGÀY KHÁC (không lọc
    session). CHỈ được gọi (tối đa 1 lần, lazy) khi có REFHUB bổ sung KHÔNG tìm
    thấy trong `df_mis_di_raw` của chính ngày đang chạy. Mặc định `None` — hành vi
    y hệt trước đây, không đổi gì nếu không dùng.

    Trả về df_mis_di_chuan — dùng thay `df_mis_di` ở phần còn lại của pipeline
    (trước khi gọi `khop_voi_gw()`).
    """
    _log = log_callback or print

    df_confirm, refhub_bo_sung = _doc_sheet_confirm_mis_di(xac_nhan_path)

    if len(df_confirm) == 0:
        gia_tri = pd.Series(dtype=str)
    else:
        gia_tri = df_confirm['LOAI_BO'].apply(lambda v: str(v).strip() if pd.notna(v) else '')
        sai = df_confirm[~gia_tri.isin(_GIA_TRI_LOAI_BO_HOP_LE)]
        if len(sai) > 0:
            chi_tiet = list(zip(sai['REFHUB'].astype(str), gia_tri[sai.index]))
            raise ValueError(f'Giá trị LOAI_BO không hợp lệ: {chi_tiet}')

    refhub_loai_bo = set(
        df_confirm.loc[gia_tri == 'loại bỏ', 'REFHUB'].astype(str).str.strip().str.lstrip("'")
    ) if len(df_confirm) else set()

    refhub_mis_di_chuan = df_mis_di['REFHUB'].astype(str).str.strip().str.lstrip("'")
    mask_loai_bo = refhub_mis_di_chuan.isin(refhub_loai_bo)
    df_giu_lai   = df_mis_di[~mask_loai_bo].copy()
    _log(f'[Bước 2] MIS_đi: loại bỏ {int(mask_loai_bo.sum()):,} dòng theo LOAI_BO, '
         f'còn {len(df_giu_lai):,} dòng')

    df_bo_sung = df_mis_di.iloc[0:0].copy()
    if refhub_bo_sung:
        da_paste = pd.Series(refhub_bo_sung)
        trung_trong_paste = da_paste[da_paste.duplicated()].tolist()
        if trung_trong_paste:
            raise ValueError(f'REFHUB bị paste trùng lặp trong vùng bổ sung: {trung_trong_paste}')

        refhub_da_co_chuan = set(refhub_mis_di_chuan) - refhub_loai_bo
        refhub_raw_chuan   = df_mis_di_raw['REFHUB'].astype(str).str.strip().str.lstrip("'")

        df_raw_ngay_khac = None  # lazy cache — chỉ đọc 1 lần nếu thật sự cần

        rows, khong_tim_thay, trung_voi_da_co, trung_lap_tren_raw = [], [], [], []
        for r in refhub_bo_sung:
            r_norm = str(r).strip().lstrip("'")
            if r_norm in refhub_da_co_chuan:
                trung_voi_da_co.append(r)
                continue
            match = df_mis_di_raw[refhub_raw_chuan == r_norm]
            nguon = 'dữ liệu ngày đang chạy'
            if len(match) == 0 and doc_them_ngay_khac is not None:
                if df_raw_ngay_khac is None:
                    df_raw_ngay_khac = doc_them_ngay_khac()
                if len(df_raw_ngay_khac) > 0:
                    refhub_khac_chuan = (
                        df_raw_ngay_khac['REFHUB'].astype(str).str.strip().str.lstrip("'")
                    )
                    match = df_raw_ngay_khac[refhub_khac_chuan == r_norm]
                    nguon = 'dữ liệu ngày khác'
            if len(match) == 0:
                khong_tim_thay.append(r)
                continue
            if len(match) > 1:
                trung_lap_tren_raw.append(r)
                continue
            _log(f'[Bước 2] REFHUB bổ sung {r_norm}: tìm thấy từ {nguon}.')
            rows.append(match)

        if khong_tim_thay:
            raise ValueError(f'Không tìm thấy REFHUB trên MIS_đi RAW: {khong_tim_thay}')
        if trung_lap_tren_raw:
            raise ValueError(
                f'REFHUB bổ sung khớp ≥2 dòng trên MIS_hub thô — không tự chọn 1 dòng theo vị trí: '
                f'{trung_lap_tren_raw}'
            )
        if trung_voi_da_co:
            raise ValueError(f'REFHUB bổ sung đã có sẵn trong MIS_đi: {trung_voi_da_co}')

        if rows:
            df_bo_sung = pd.concat(rows, ignore_index=True)
            df_bo_sung = _them_cot_khoa(_chuan_hoa_co_ban(df_bo_sung))
            df_bo_sung['LY_DO_GIU_SESSION_NULL'] = ''

    _log(f'[Bước 2] Bổ sung từ REFHUB paste thêm: {len(df_bo_sung)} giao dịch')

    df_mis_di_chuan = pd.concat([df_giu_lai, df_bo_sung], ignore_index=True, sort=False)
    _log(f'[Bước 2] MIS_đi chuẩn (final): {len(df_mis_di_chuan):,} dòng')

    return df_mis_di_chuan
