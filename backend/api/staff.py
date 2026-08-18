"""KSNB Staff management endpoints"""
import io
import os
import re
import sqlite3
import tempfile
import unicodedata
from datetime import date as _date_cls, datetime as _datetime_cls, timedelta as _timedelta
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from backend.core.concurrency import run_heavy
from backend.database import DB_PATH, get_db, write_audit, compute_annual_leave, _vn_now

# Mốc hiệu lực gốc cho dòng lịch sử phòng đầu tiên của mỗi cán bộ
_DEPT_HISTORY_EPOCH = "2000-01-01"
from backend.schemas.staff import StaffCreate, StaffUpdate, StaffOut
from backend.core.security import get_password_hash
from backend.core.uploads import read_limited, read_limited_sync
from backend.core.deps import get_current_staff, require_feature
from backend.core.enums import ROLE_RANK, VALID_ROLES
from backend.core.net import client_ip as _client_ip

router = APIRouter(prefix="/api/staff", tags=["Staff"])

# Quản trị viên (cấp 1 + cấp 2) — không thuộc phòng nào
_ADMIN_ROLES = frozenset(("admin", "admin_l2"))
# Role được xem toàn bộ nhân viên không bị giới hạn phòng
_BROAD_VIEW_ROLES = frozenset(("admin", "admin_l2", "hau_kiem_vien", "giam_doc", "pho_giam_doc"))
# Cột do tính năng Nghỉ phép làm chủ — /import-db không đè lên tài khoản đã có
_COT_NGHI_PHEP = frozenset(("used_leave_days", "annual_leave_days", "carryover_notice_year"))


def _chan_leo_thang_quyen(current: dict, role_moi, row_cu=None) -> None:
    """Chặn người dùng tự nâng mình — hoặc nâng người khác — lên quyền cao hơn.

    Trước đây chỗ này chỉ có đúng một luật: "QTV cấp 2 không được đụng QTV cấp
    1". Nghĩa là rào chắn nằm ở CHỖ AI ĐANG ĐƯỢC GÁN feature `staff.edit`, chứ
    không nằm trong mã. Ai được cấp feature đó — kể cả một chuyên viên — đều gọi
    được `PUT /api/staff/{id_của_chính_mình}` với `role: "admin"` và trở thành
    quản trị viên toàn quyền. Hiện tại feature mới chỉ nằm ở nhóm quản trị nên
    chưa ai khai thác được, nhưng đó là may chứ không phải thiết kế.

    Giữ NGUYÊN cách hành xử của admin và admin_l2 (đang chạy tốt, đổi là gãy
    việc quản trị hằng ngày); chỉ siết những vai trò còn lại theo bậc quyền
    trong docs/DESIGN.md: chỉ được thao tác với vai trò THẤP HƠN mình.
    Bậc bằng nhau cũng bị chặn — đó chính là trường hợp tự sửa vai trò mình.
    """
    # QTV cấp 1: toàn quyền, như cũ.
    if current["role"] == "admin":
        return

    # QTV cấp 2: giữ nguyên hai luật cũ.
    if current["role"] == "admin_l2":
        if row_cu is not None and row_cu["role"] == "admin":
            raise HTTPException(403, "Quản trị viên cấp 2 không được sửa tài khoản Quản trị viên cấp 1")
        if role_moi == "admin":
            raise HTTPException(
                403, "Quản trị viên cấp 2 không được nâng lên Quản trị viên cấp 1")
        return

    # Mọi vai trò khác có feature staff.create / staff.edit.
    bac_minh = ROLE_RANK.get(current["role"], 0)
    if row_cu is not None:
        bac_cu = ROLE_RANK.get(row_cu["role"], 0)
        if bac_cu >= bac_minh:
            raise HTTPException(
                403,
                "Không được sửa tài khoản có vai trò ngang hoặc cao hơn mình"
                + (" (kể cả tài khoản của chính mình)" if row_cu["id"] == current["id"] else ""),
            )
    if role_moi is not None and ROLE_RANK.get(role_moi, 0) >= bac_minh:
        raise HTTPException(
            403, "Không được gán vai trò ngang hoặc cao hơn vai trò của mình")

_ROLE_ORDER_SQL = """
    CASE role
        WHEN 'giam_doc'      THEN 0
        WHEN 'pho_giam_doc'  THEN 1
        WHEN 'admin'         THEN 2
        WHEN 'admin_l2'      THEN 3
        WHEN 'truong_phong'  THEN 4
        WHEN 'pho_phong'     THEN 5
        WHEN 'hau_kiem_vien' THEN 6
        WHEN 'chuyen_vien'   THEN 7
        ELSE 9
    END
"""


def _validate_dept(db: sqlite3.Connection, role: str, department_id):
    # Quản trị viên (cấp 1 + cấp 2) không thuộc phòng nào — bỏ qua yêu cầu chọn phòng
    if role in _ADMIN_ROLES:
        return
    if not department_id:
        raise HTTPException(400, "Phải chọn phòng ban")
    dept = db.execute(
        "SELECT * FROM departments WHERE id = ? AND is_active = 1", (department_id,)
    ).fetchone()
    if not dept:
        raise HTTPException(400, "Phòng ban không tồn tại hoặc không còn hoạt động")
    if role in ("giam_doc", "pho_giam_doc") and dept["is_source"]:
        raise HTTPException(400, "Giám đốc / Phó Giám đốc phải thuộc Ban Giám đốc")


@router.get("/", response_model=List[StaffOut])
def list_staff(
    active_only: bool = True,
    department_id: Optional[int] = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    clauses = ["(is_deleted = 0 OR is_deleted IS NULL)"]
    params = []
    if active_only:
        clauses.append("is_active = 1")

    # ── Scope theo role ──
    is_broad = current["role"] in _BROAD_VIEW_ROLES
    if not is_broad:
        # Kiểm tra có phải nhân viên phòng Tổng hợp không (cần xem GĐ/PGĐ để chọn approver)
        dept_row = db.execute(
            "SELECT code FROM departments WHERE id = ?", (current.get("department_id"),)
        ).fetchone() if current.get("department_id") else None
        is_broad = bool(dept_row and dept_row["code"].upper() in ("TH", "TONGHOP", "TONG_HOP"))

    if not is_broad:
        # truong_phong / pho_phong / chuyen_vien: chỉ xem phòng mình
        own_dept = current.get("department_id")
        if own_dept:
            clauses.append("department_id = ?")
            params.append(own_dept)
    elif department_id:
        # Broad role: filter theo phòng nếu client yêu cầu
        clauses.append("department_id = ?")
        params.append(department_id)

    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    rows = db.execute(
        f"SELECT * FROM user_tttt {where} ORDER BY {_ROLE_ORDER_SQL}, full_name", params
    ).fetchall()
    return [_enrich(dict(r)) for r in rows]


# ─── Export Excel / Import DB ────────────────────────────────────────────────
# Đặt trước /{staff_id} để tránh FastAPI match "export" như int

_ROLE_VN = {
    "chuyen_vien":   "Chuyên viên",
    "pho_phong":     "Phó phòng",
    "truong_phong":  "Trưởng phòng",
    "hau_kiem_vien": "Hậu kiểm viên",
    "giam_doc":      "Giám đốc",
    "pho_giam_doc":  "Phó Giám đốc",
    "admin":         "Quản trị viên cấp 1",
    "admin_l2":      "Quản trị viên cấp 2",
}


@router.get("/export")
def export_staff_excel(
    request: Request,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.export")),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from datetime import date, datetime as _dt

    rows = db.execute("""
        SELECT u.full_name, u.employee_code, u.role, d.name AS dept_name,
               u.username, u.ipcas_code, u.payment_username, u.phone,
               u.is_active, u.created_at
        FROM user_tttt u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.is_deleted = 0 OR u.is_deleted IS NULL
        ORDER BY d.name, u.full_name
    """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách cán bộ"

    hdr_fill = PatternFill("solid", fgColor="C62828")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["STT", "Họ và tên", "Mã cán bộ", "Quyền", "Phòng", "Username",
               "User IPCAS", "User Payment", "Điện thoại", "Trạng thái", "Ngày tạo"]
    widths   = [6, 28, 14, 18, 24, 18, 14, 20, 14, 12, 14]
    ws.append(headers)
    for cell, w in zip(ws[1], widths):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for idx, r in enumerate(rows, 1):
        created = ""
        try:
            created = _dt.fromisoformat(str(r["created_at"])).strftime("%d/%m/%Y")
        except Exception:
            pass
        ws.append([
            idx,
            r["full_name"] or "",
            r["employee_code"] or "",
            _ROLE_VN.get(r["role"], r["role"]),
            r["dept_name"] or "",
            r["username"] or "",
            r["ipcas_code"] or "",
            r["payment_username"] or "",
            r["phone"] or "",
            "Hoạt động" if r["is_active"] else "Tạm khóa",
            created,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Mang cả bảng nhân sự ra ngoài — phải để lại vết. AuditMiddleware chỉ ghi
    # POST/PUT/PATCH/DELETE và còn bỏ qua hẳn tiền tố /api/staff, nên nếu ở đây
    # không tự ghi thì việc này KHÔNG xuất hiện ở bất kỳ nhật ký nào.
    write_audit(db, current["id"], "staff_export", "staff", None,
                f"Xuất danh sách {len(rows)} cán bộ ra Excel", _client_ip(request))
    db.commit()
    fname = f"danh_sach_can_bo_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@router.get("/export-db")
def export_users_db(
    request: Request,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.export")),
):
    """Xuất bảng user_tttt thành file SQLite để chép sang hệ thống khác.

    File này chứa NGUYÊN cột `pwd_hash` — đó là chủ đích, vì mục đích của nó là
    chuyển người dùng sang hệ thống khác mà không bắt ai đặt lại mật khẩu. Hệ
    quả: ai cầm được file là cầm toàn bộ mã băm mật khẩu để dò ngoại tuyến.

    Vì thế hai lớp chặn, không phải một:
      - Chỉ quản trị viên (cấp 1 hoặc cấp 2) được gọi, dù feature `staff.export`
        có trót được gán cho nhóm nào khác đi nữa.
      - Luôn ghi một dòng nhật ký. Đây là GET nên AuditMiddleware không đụng tới.
    """
    if current["role"] not in _ADMIN_ROLES:
        raise HTTPException(
            403,
            "Chỉ quản trị viên được xuất file DB người dùng — file này chứa mã băm mật khẩu.",
        )
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    try:
        src = sqlite3.connect(DB_PATH)
        exp = sqlite3.connect(tmp.name)
        ddl = src.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='user_tttt'"
        ).fetchone()[0]
        exp.execute(ddl)
        src.row_factory = sqlite3.Row
        rows_db = src.execute(
            "SELECT * FROM user_tttt WHERE is_deleted = 0 OR is_deleted IS NULL"
        ).fetchall()
        if rows_db:
            cols = list(rows_db[0].keys())
            ph = ",".join("?" * len(cols))
            exp.executemany(f"INSERT INTO user_tttt VALUES ({ph})", [tuple(r) for r in rows_db])
        exp.commit()
        src.close()
        exp.close()
        data = open(tmp.name, "rb").read()
    finally:
        os.unlink(tmp.name)

    # _client_ip(): request.client.host luôn là 127.0.0.1 vì frontend gọi
    # backend qua loopback — IP thật của trình duyệt nằm ở X-Client-IP.
    write_audit(db, current["id"], "staff_export_db", "staff", None,
                f"Xuất file DB người dùng ({len(rows_db)} tài khoản, GỒM mã băm mật khẩu)",
                _client_ip(request))
    db.commit()

    from datetime import date as _date
    fname_db = f"users_{_date.today().strftime('%Y%m%d')}.db"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{fname_db}"'},
    )


def _enrich(row: dict) -> dict:
    """Inject annual_leave_days tính từ join_industry_date nếu có."""
    if row.get("join_industry_date"):
        row["annual_leave_days"] = compute_annual_leave(row["join_industry_date"])
    return row


@router.get("/{staff_id}", response_model=StaffOut)
def get_staff(
    staff_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    row = db.execute("SELECT * FROM user_tttt WHERE id = ?", (staff_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy cán bộ")
    return _enrich(dict(row))


@router.post("/", response_model=StaffOut)
def create_staff(
    body: StaffCreate,
    request: Request,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.create")),
):
    if db.execute("SELECT id FROM user_tttt WHERE username = ?", (body.username,)).fetchone():
        raise HTTPException(400, "Username đã tồn tại")
    emp_code = body.employee_code or body.username
    if db.execute("SELECT id FROM user_tttt WHERE employee_code = ?", (emp_code,)).fetchone():
        raise HTTPException(400, "Mã nhân viên đã tồn tại")
    _chan_leo_thang_quyen(current, body.role)
    # Admin (cấp 1 + cấp 2) không thuộc phòng nào → ép department_id về None dù client có gửi
    dept_id = None if body.role in _ADMIN_ROLES else body.department_id
    _validate_dept(db, body.role, dept_id)
    join_date_iso = body.join_industry_date.isoformat() if body.join_industry_date else None
    cur = db.execute(
        """INSERT INTO user_tttt
           (employee_code, full_name, role, department_id, username, pwd_hash,
            phone, email, start_date, join_industry_date, ipcas_code, payment_username, is_active)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,1)""",
        (emp_code, body.full_name, body.role, dept_id, body.username,
         get_password_hash(body.password), body.phone, body.email,
         body.start_date.isoformat() if body.start_date else None,
         join_date_iso, body.ipcas_code, body.payment_username),
    )
    new_id = cur.lastrowid
    # Dòng lịch sử phòng đầu tiên — hiệu lực từ epoch để phủ mọi tháng trước đó
    if dept_id:
        db.execute(
            "INSERT INTO staff_department_history (staff_id, department_id, effective_from, created_at) VALUES (?,?,?,?)",
            (new_id, dept_id, _DEPT_HISTORY_EPOCH, str(_vn_now())),
        )
    client_ip = request.client.host if request.client else "unknown"
    write_audit(db, current["id"], "staff_create", "staff", new_id,
                f"Tạo tài khoản {body.username} ({body.full_name})", client_ip)
    db.commit()
    row = db.execute("SELECT * FROM user_tttt WHERE id = ?", (new_id,)).fetchone()
    return _enrich(dict(row))


@router.put("/{staff_id}", response_model=StaffOut)
def update_staff(
    staff_id: int,
    body: StaffUpdate,
    request: Request,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.edit")),
):
    row = db.execute("SELECT * FROM user_tttt WHERE id = ?", (staff_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy cán bộ")
    _chan_leo_thang_quyen(current, body.role, row)
    update_data = body.model_dump(exclude_none=True)
    # Serialize date object → ISO string cho sqlite3
    if "join_industry_date" in update_data and update_data["join_industry_date"]:
        update_data["join_industry_date"] = update_data["join_industry_date"].isoformat()
    if "employee_code" in update_data:
        dup = db.execute(
            "SELECT id FROM user_tttt WHERE employee_code = ? AND id != ?",
            (update_data["employee_code"], staff_id),
        ).fetchone()
        if dup:
            raise HTTPException(400, "Mã cán bộ đã tồn tại")
    new_role = update_data.get("role", row["role"])
    new_dept = update_data.get("department_id", row["department_id"])
    # Admin (cấp 1 + cấp 2) không thuộc phòng nào → ép department_id về None
    if new_role in _ADMIN_ROLES:
        new_dept = None
        update_data["department_id"] = None
    _validate_dept(db, new_role, new_dept)
    if update_data:
        sets = ", ".join(f"{k} = ?" for k in update_data)
        params = list(update_data.values()) + [staff_id]
        db.execute(f"UPDATE user_tttt SET {sets} WHERE id = ?", params)
        # Đổi phòng → ghi mốc lịch sử, hiệu lực từ ngày quản trị viên thực hiện (hôm nay)
        # Bỏ qua khi chuyển sang admin (new_dept = None) vì history.department_id NOT NULL
        if new_dept != row["department_id"] and new_dept:
            db.execute(
                "INSERT INTO staff_department_history (staff_id, department_id, effective_from, created_at) VALUES (?,?,?,?)",
                (staff_id, new_dept, _vn_now().date().isoformat(), str(_vn_now())),
            )
        client_ip = request.client.host if request.client else "unknown"
        changed = ", ".join(f"{k}={v}" for k, v in update_data.items() if k != "pwd_hash")
        write_audit(db, current["id"], "staff_update", "staff", staff_id,
                    f"Cập nhật {row['username']}: {changed}", client_ip)
        db.commit()
    row = db.execute("SELECT * FROM user_tttt WHERE id = ?", (staff_id,)).fetchone()
    return _enrich(dict(row))


@router.delete("/{staff_id}")
def delete_staff(
    staff_id: int,
    request: Request,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.delete")),
):
    if staff_id == current["id"]:
        raise HTTPException(400, "Không thể xóa tài khoản của chính mình")
    # `id` phải có trong SELECT: _chan_leo_thang_quyen() đọc nó để nói rõ
    # "kể cả tài khoản của chính mình" trong thông báo từ chối.
    row = db.execute("SELECT id, username, full_name, role FROM user_tttt WHERE id = ?", (staff_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy cán bộ")
    _chan_leo_thang_quyen(current, None, row)
    db.execute("UPDATE user_tttt SET is_deleted = 1, is_active = 0 WHERE id = ?", (staff_id,))
    client_ip = request.client.host if request.client else "unknown"
    write_audit(db, current["id"], "staff_delete", "staff", staff_id,
                f"Xóa tài khoản {row['username']} ({row['full_name']})", client_ip)
    db.commit()
    return {"message": "Đã xóa tài khoản"}


@router.post("/import-db")
def import_users_db(
    file: UploadFile = File(...),
    request: Request = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.import_db")),
):
    """Nhập bảng user_tttt từ file SQLite xuất bởi hệ thống cùng schema.

    Cặp đôi với `/export-db`, nên file mang theo NGUYÊN `pwd_hash` và `role`:
    ai gọi được endpoint này là đặt được mật khẩu VÀ vai trò cho bất kỳ ai —
    kể cả tự nâng mình lên `admin`. Vì thế chặn theo VAI TRÒ THẬT chứ không chỉ
    theo feature, y hệt lý do đã ghi ở `/export-db`.

    Ba thứ trong file cố ý KHÔNG được tin:
      - `role` sai chính tả → tài khoản rớt khỏi mọi kiểm tra quyền (xem
        `_kiem_tra_role` trong schemas/staff.py) → bỏ dòng, báo lại.
      - `department_id` trỏ vào phòng không tồn tại → bỏ dòng, báo lại. Nhờ vậy
        không cần tắt kiểm tra khoá ngoại nữa: bản cũ tắt `foreign_keys` rồi ghi
        bừa, để lại đúng loại dữ liệu hỏng mà `StaffOut._null_la_khoa` đang phải
        chữa cháy.
      - Số liệu phép của tài khoản ĐÃ CÓ (xem `_COT_NGHI_PHEP`) — đó là dữ liệu
        của tính năng Nghỉ phép, nơi có batch + rollback riêng. Nhập file cũ đè
        lên là xoá sổ ngày phép đã dùng của cả cơ quan mà không hoàn tác được.
        Tài khoản MỚI thì vẫn lấy — di trú người sang thì số liệu đi theo người.
    """
    if current["role"] not in _ADMIN_ROLES:
        raise HTTPException(
            403,
            "Chỉ quản trị viên được nhập file DB người dùng — file này đặt được "
            "mật khẩu và vai trò cho mọi tài khoản.",
        )

    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    try:
        tmp.write(read_limited_sync(file, ten="File DB người dùng"))
        tmp.close()
        imp = sqlite3.connect(tmp.name)
        imp.row_factory = sqlite3.Row
        rows = imp.execute("SELECT * FROM user_tttt").fetchall()
        imp.close()
    finally:
        os.unlink(tmp.name)

    if not rows:
        raise HTTPException(400, "File không có dữ liệu user_tttt")

    cols = list(rows[0].keys())
    non_id_cols = [c for c in cols if c != "id"]
    dept_ids = {r["id"] for r in db.execute("SELECT id FROM departments").fetchall()}

    inserted = updated = 0
    bo_qua: list[str] = []

    for row in rows:
        ma = row["employee_code"] if "employee_code" in cols else None
        if not ma:
            # Cột này là khoá đối chiếu duy nhất và NOT NULL UNIQUE trong lược đồ.
            # Không chặn ở đây thì INSERT ném IntegrityError → 500 giữa mẻ.
            bo_qua.append("(dòng không có mã cán bộ)")
            continue
        existing = db.execute(
            "SELECT id, role FROM user_tttt WHERE employee_code = ?", (ma,)
        ).fetchone()

        # ── Validate từng dòng — dòng hỏng bị bỏ, không làm hỏng cả mẻ ──
        role_moi = row["role"] if "role" in cols else None
        if role_moi is not None and role_moi not in VALID_ROLES:
            bo_qua.append(f"{ma}: vai trò '{role_moi}' không hợp lệ")
            continue
        if current["role"] == "admin_l2" and (
            role_moi == "admin" or (existing and existing["role"] == "admin")
        ):
            bo_qua.append(f"{ma}: QTV cấp 2 không được đụng tài khoản QTV cấp 1")
            continue
        dept = row["department_id"] if "department_id" in cols else None
        if dept not in (None, ""):
            # int(): SQLite không ép kiểu, file do người khác chỉnh tay có thể
            # để id phòng dưới dạng chuỗi — "1" != 1 sẽ loại nhầm cả mẻ dữ liệu đúng
            try:
                dept = int(dept)
            except (TypeError, ValueError):
                bo_qua.append(f"{ma}: phòng ban '{dept}' không phải số")
                continue
            if dept not in dept_ids:
                bo_qua.append(f"{ma}: phòng ban id={dept} không tồn tại trên hệ thống này")
                continue

        vals = {c: row[c] for c in non_id_cols}
        if "department_id" in vals:
            vals["department_id"] = dept if dept not in (None, "") else None
        # NULL lọt vào is_active là nguồn của validator vá lỗi ở StaffOut — chặn tại đây
        if "is_active" in vals and vals["is_active"] is None:
            vals["is_active"] = 0

        if existing:
            ghi = [c for c in non_id_cols
                   if c != "employee_code" and c not in _COT_NGHI_PHEP]
            if not ghi:
                bo_qua.append(f"{ma}: file không có cột nào được phép cập nhật")
                continue
            sets = ", ".join(f"{c} = ?" for c in ghi)
            db.execute(
                f"UPDATE user_tttt SET {sets} WHERE employee_code = ?",
                [vals[c] for c in ghi] + [ma],
            )
            updated += 1
        else:
            ph = ",".join("?" * len(non_id_cols))
            db.execute(
                f"INSERT INTO user_tttt ({','.join(non_id_cols)}) VALUES ({ph})",
                [vals[c] for c in non_id_cols],
            )
            inserted += 1

    db.commit()
    chi_tiet = f"Import DB: +{inserted} mới, ~{updated} cập nhật"
    if bo_qua:
        chi_tiet += f", bỏ qua {len(bo_qua)} dòng ({'; '.join(bo_qua[:5])}"
        chi_tiet += " …)" if len(bo_qua) > 5 else ")"
    write_audit(db, current["id"], "staff_import_db", "staff", None, chi_tiet, _client_ip(request))
    db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": bo_qua}


# ─── Nhập Ngày vào ngành hàng loạt từ Excel ─────────────────────────────────

def _fold_hdr(s) -> str:
    """Hạ chữ + bỏ dấu + gộp khoảng trắng để dò tên cột bất kể cách gõ."""
    t = unicodedata.normalize("NFD", str(s or "")).replace("đ", "d").replace("Đ", "D")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", t).strip().lower()


def _parse_join_date(v):
    """Excel trả về datetime, date, chuỗi dd/mm/yyyy hoặc số serial — chuẩn hoá về ISO."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, _datetime_cls):
        d = v.date()
    elif isinstance(v, _date_cls):
        d = v
    elif isinstance(v, (int, float)):
        # Serial Excel (epoch 1899-12-30) — chỉ gặp khi ô không được định dạng ngày
        try:
            d = _date_cls(1899, 12, 30) + _timedelta(days=int(v))
        except Exception:
            return None
    else:
        s = str(v).strip().split()[0]
        d = None
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%y"):
            try:
                d = _datetime_cls.strptime(s, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            return None
    # Chặn ngày vô lý: gõ nhầm năm làm số ngày phép sai suốt về sau
    if not (1950 <= d.year <= _date_cls.today().year):
        return None
    return d.isoformat()


def _parse_join_date_workbook(content: bytes):
    """Đọc file Excel → [(ma_can_bo, ho_ten, ngay_iso|None, so_dong)]."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active

    # ── Dò dòng tiêu đề (file mẫu có dòng trống ở trên) ──
    col_code = col_date = col_name = None
    header_row = 0
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        found = {}
        for ci, cell in enumerate(row):
            h = _fold_hdr(cell)
            if not h:
                continue
            if "ma can bo" in h or h in ("ma cb", "ma nhan vien"):
                found["code"] = ci
            elif "vao nganh" in h:
                found["date"] = ci
            elif "ho va ten" in h or h == "ho ten":
                found["name"] = ci
        if "code" in found and "date" in found:
            col_code, col_date = found["code"], found["date"]
            col_name = found.get("name")
            header_row = ri
            break
    if col_code is None:
        wb.close()
        raise HTTPException(400, "Không tìm thấy cột 'Mã cán bộ' và 'Ngày vào ngành' trong file")

    items = []
    for ri, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        code = row[col_code] if col_code < len(row) else None
        if code is None or not str(code).strip():
            continue  # dòng tiêu đề nhóm phòng (chỉ có tên phòng ở cột B)
        # Mã cán bộ đọc từ ô số sẽ ra float ("2.00733664e+08") → ép về int trước
        if isinstance(code, float) and code.is_integer():
            code = int(code)
        raw_date = row[col_date] if col_date < len(row) else None
        name = row[col_name] if (col_name is not None and col_name < len(row)) else ""
        items.append((str(code).strip(), str(name or "").strip(),
                      _parse_join_date(raw_date), ri))
    wb.close()
    return items


@router.post("/import-join-dates")
async def import_join_dates(
    file: UploadFile = File(...),
    overwrite: bool = Query(False, description="Ghi đè cả những người đã có ngày vào ngành"),
    dry_run: bool = Query(False, description="Chỉ xem trước, không ghi DB"),
    request: Request = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("staff.import_join_date")),
):
    """Nhập hàng loạt Ngày vào ngành từ Excel, khớp theo Mã cán bộ.

    Mặc định chỉ điền vào ô đang trống — người đã có ngày (sửa tay trên máy chính)
    không bị file cũ đè lên. Muốn đè thì gọi với `overwrite=true`.
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Chỉ nhận file Excel .xlsx")

    items = await run_heavy(_parse_join_date_workbook,
                            await read_limited(file, ten="File Excel ngày vào ngành"))
    if not items:
        raise HTTPException(400, "Không đọc được dòng dữ liệu nào trong file")

    existing = {
        str(r["employee_code"] or "").strip(): r
        for r in db.execute(
            """SELECT id, employee_code, full_name, join_industry_date FROM user_tttt
               WHERE is_deleted = 0 OR is_deleted IS NULL"""
        )
    }

    updated = unchanged = 0
    not_found, bad_date, kept = [], [], []
    seen = set()
    for code, name, iso, ri in items:
        if code in seen:
            continue  # trùng mã trong file — giữ dòng đầu
        seen.add(code)
        row = existing.get(code)
        if row is None:
            not_found.append(f"Dòng {ri}: {code} — {name}")
            continue
        if iso is None:
            bad_date.append(f"Dòng {ri}: {code} — {name}")
            continue
        cur = (row["join_industry_date"] or "")[:10]
        if cur == iso:
            unchanged += 1
            continue
        if cur and not overwrite:
            kept.append(f"{code} — {row['full_name']}: giữ {cur}, file ghi {iso}")
            continue
        if not dry_run:
            db.execute("UPDATE user_tttt SET join_industry_date = ? WHERE id = ?", (iso, row["id"]))
        updated += 1

    if not dry_run:
        db.commit()
        client_ip = request.client.host if request and request.client else "unknown"
        write_audit(db, current["id"], "staff_import_join_dates", "staff", None,
                    f"{file.filename}: cập nhật {updated}, giữ nguyên {unchanged}, "
                    f"không khớp {len(not_found)}", client_ip)
        db.commit()

    return {
        "dry_run": dry_run,
        "total_rows": len(items),
        "updated": updated,
        "unchanged": unchanged,
        "not_found": not_found,
        "bad_date": bad_date,
        "kept_existing": kept,
    }


# ─── Leave records (DEPRECATED — dùng /api/leaves/ thay thế) ────────────────
@router.get("/{staff_id}/leaves", deprecated=True)
def list_leaves_deprecated(
    staff_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    rows = db.execute(
        "SELECT * FROM leave_records WHERE staff_id = ? ORDER BY start_date DESC", (staff_id,)
    ).fetchall()
    return [dict(r) for r in rows]
