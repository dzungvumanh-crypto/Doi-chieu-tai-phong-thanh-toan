"""Bundle management endpoints"""
import io
import json
import logging
import sqlite3
from collections import defaultdict
from datetime import date as date_type
from typing import List, Optional, Tuple
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl.styles import Alignment, Border, Font, Side

from backend.core.concurrency import run_heavy
from backend.core.deps import get_current_staff, require_feature
from backend.core.enums import StaffRole
from backend.database import get_db, _vn_now
from backend.schemas.bundles import (
    ArchiveCoverParseResponse, ArchiveCoverPrintRequest, ArchiveCoverRow,
    BundleGenerateRequest, BundleUpdateRequest,
    StorageViewResponse, StorageViewRow, StorageViewUpdateRequest,
    StorageSummaryCell, StorageSummaryDept, StorageSummaryResponse, StorageSummaryRow,
)
from backend.schemas.handovers import ArchiveRecord, HandoverArchiveResponse
from backend.services import archive_cover_service
from backend.services.bundle_service import BundleResult, EntryUnit, generate_bundles_for_entries
from backend.services.cover_service import generate_covers_docx

_log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/bundles", tags=["Bundles"])


def _download_headers(filename: str) -> dict:
    fallback = "".join(ch if ord(ch) < 128 and ch not in '\\"' else "_" for ch in filename)
    return {
        "Content-Disposition": (
            f"attachment; filename=\"{fallback}\"; "
            f"filename*=UTF-8''{quote(filename, safe='')}"
        )
    }


# ─── Helpers cho bundle data ──────────────────────────────────────────────────

def _get_dates_for_bundle(bundle: dict) -> frozenset:
    """Lấy tập hợp ngày GD — ưu tiên cover_units JSON, fallback qua items."""
    if bundle.get("cover_units"):
        try:
            data = json.loads(bundle["cover_units"])
            return frozenset(date_type.fromisoformat(u["date"]) for u in data)
        except Exception:
            _log.warning("bundle %s: cover_units JSON lỗi, dùng items fallback", bundle["id"])
    return frozenset(
        date_type.fromisoformat(item["entry"]["transaction_date"])
        for item in bundle.get("items", [])
        if item.get("entry") and item["entry"].get("transaction_date")
    )


def _units_from_bundle(bundle: dict) -> List[EntryUnit]:
    """Build EntryUnit list từ cover_units JSON hoặc fallback qua items."""
    if bundle.get("cover_units"):
        try:
            data = json.loads(bundle["cover_units"])
            return [
                EntryUnit(
                    entry_ids=[], source_user_id=0,
                    user_code=u["user_code"], full_name=u.get("full_name"),
                    transaction_date=date_type.fromisoformat(u["date"]),
                    sheet_count=u["sheet_count"], is_large=u.get("is_large", False),
                )
                for u in data
            ]
        except Exception:
            _log.warning("bundle %s: cover_units JSON lỗi khi build units, dùng fallback", bundle["id"])
    units = []
    for item in bundle.get("items", []):
        e = item.get("entry")
        s = e.get("staff") if e else None
        if e and s:
            units.append(EntryUnit(
                entry_ids=[e["id"]],
                source_user_id=e["staff_id"],
                user_code=s.get("ipcas_code") or "",
                full_name=s.get("full_name"),
                transaction_date=date_type.fromisoformat(e["transaction_date"]),
                sheet_count=e["sheet_count"],
            ))
    return units


def _get_bundle_label(bundle: dict, all_bundles: list) -> Tuple[int, int]:
    bundle_dates = {b["id"]: _get_dates_for_bundle(b) for b in all_bundles}
    single_date_groups: dict = defaultdict(list)
    for b in all_bundles:
        dates = bundle_dates[b["id"]]
        if len(dates) == 1:
            single_date_groups[next(iter(dates))].append(b)
    my_dates = bundle_dates.get(bundle["id"], frozenset())
    if len(my_dates) == 1:
        day = next(iter(my_dates))
        same = sorted(single_date_groups[day], key=lambda b: b["sequence"])
        if len(same) > 1:
            idx = next(i + 1 for i, b in enumerate(same) if b["id"] == bundle["id"])
            return idx, len(same)
    return 1, 1


def _load_bundle_group(db: sqlite3.Connection, group_id: int) -> dict:
    g = db.execute("SELECT * FROM bundle_groups WHERE id = ?", (group_id,)).fetchone()
    if not g:
        raise HTTPException(404, "Không tìm thấy nhóm tập")

    dept = db.execute("SELECT * FROM departments WHERE id = ?", (g["department_id"],)).fetchone()
    creator = db.execute("SELECT * FROM user_tttt WHERE id = ?", (g["created_by_id"],)).fetchone()

    bundle_rows = db.execute(
        """SELECT b.*, ks.full_name AS cust_name
           FROM bundles b
           LEFT JOIN user_tttt ks ON b.custodian_id = ks.id
           WHERE b.group_id = ? ORDER BY b.sequence""",
        (group_id,),
    ).fetchall()

    bundles_out = []
    for b in bundle_rows:
        item_rows = db.execute(
            """SELECT bi.id, bi.entry_id,
                      de.transaction_date, de.sheet_count, de.notes, de.staff_id,
                      ks.employee_code AS s_emp, ks.full_name AS s_name, ks.role AS s_role,
                      ks.department_id AS s_dept, ks.username AS s_user,
                      ks.phone AS s_phone, ks.email AS s_email, ks.start_date AS s_start,
                      ks.ipcas_code, ks.payment_username, ks.is_active AS s_active
               FROM bundle_items bi
               JOIN document_entries de ON bi.entry_id = de.id
               LEFT JOIN user_tttt ks ON de.staff_id = ks.id
               WHERE bi.bundle_id = ?""",
            (b["id"],),
        ).fetchall()

        items_out = []
        for item in item_rows:
            staff = None
            if item["staff_id"]:
                staff = {
                    "id": item["staff_id"],
                    "employee_code": item["s_emp"] or "",
                    "full_name": item["s_name"] or "",
                    "role": item["s_role"] or "",
                    "department_id": item["s_dept"],
                    "username": item["s_user"] or "",
                    "phone": item["s_phone"],
                    "email": item["s_email"],
                    "start_date": item["s_start"],
                    "ipcas_code": item["ipcas_code"],
                    "payment_username": item["payment_username"],
                    "is_active": bool(item["s_active"]),
                }
            items_out.append({
                "id": item["id"],
                "entry_id": item["entry_id"],
                "entry": {
                    "id": item["entry_id"],
                    "staff_id": item["staff_id"],
                    "transaction_date": item["transaction_date"],
                    "sheet_count": item["sheet_count"],
                    "notes": item["notes"],
                    "staff": staff,
                },
            })

        bundles_out.append({
            "id": b["id"],
            "group_id": b["group_id"],
            "sequence": b["sequence"],
            "total_sheets": b["total_sheets"],
            "custodian_id": b["custodian_id"],
            "custodian_name": b["cust_name"],
            "storage_box": b["storage_box"],
            "storage_location": b["storage_location"],
            "cover_printed_at": b["cover_printed_at"],
            "status": b["status"] or "pending",
            "cover_units": b["cover_units"],
            "items": items_out,
        })

    def _staff_dict(r):
        if not r:
            return None
        return {
            "id": r["id"],
            "employee_code": r["employee_code"] or "",
            "full_name": r["full_name"] or "",
            "role": r["role"] or "",
            "department_id": r["department_id"],
            "username": r["username"] or "",
            "phone": r["phone"],
            "email": r["email"],
            "start_date": r["start_date"],
            "ipcas_code": r["ipcas_code"],
            "payment_username": r["payment_username"],
            "is_active": bool(r["is_active"]),
        }

    return {
        "id": g["id"],
        "department_id": g["department_id"],
        "total_bundles": g["total_bundles"],
        "created_at": g["created_at"],
        "notes": g["notes"],
        "department": dict(dept) if dept else None,
        "created_by_staff": _staff_dict(creator),
        "bundles": bundles_out,
    }


def _delete_group_cascade(db: sqlite3.Connection, group_id: int):
    bundle_ids = [r["id"] for r in db.execute("SELECT id FROM bundles WHERE group_id = ?", (group_id,)).fetchall()]
    if bundle_ids:
        ph = ",".join("?" * len(bundle_ids))
        db.execute(f"DELETE FROM bundle_items WHERE bundle_id IN ({ph})", bundle_ids)
    db.execute("DELETE FROM bundles WHERE group_id = ?", (group_id,))
    db.execute("DELETE FROM bundle_groups WHERE id = ?", (group_id,))


# ─── Storage / archive helpers ────────────────────────────────────────────────

def _decompose_bundles_to_rows(bundles_data: list) -> list:
    """Shared logic: bundles_data là list dict có id, sequence, total_sheets, cover_units, items."""
    bundle_dates = {b["id"]: sorted(_get_dates_for_bundle(b)) for b in bundles_data}
    single_day: dict = defaultdict(list)
    multi_day: list = []
    for b in bundles_data:
        dates = bundle_dates[b["id"]]
        if len(dates) == 1:
            single_day[dates[0]].append(b)
        elif len(dates) > 1:
            multi_day.append((dates, b))

    rows = []
    for dates, b in multi_day:
        rows.append(StorageViewRow(
            days=sorted(d.day for d in dates),
            bundle_ids=[b["id"]],
            bundle_sheets=[b["total_sheets"]],
            n_bundles=1,
        ))
    for d, date_bundles in single_day.items():
        date_bundles = sorted(date_bundles, key=lambda b: b["sequence"])
        rows.append(StorageViewRow(
            days=[d.day],
            bundle_ids=[b["id"] for b in date_bundles],
            bundle_sheets=[b["total_sheets"] for b in date_bundles],
            n_bundles=len(date_bundles),
        ))
    rows.sort(key=lambda r: min(r.days) if r.days else 0)
    return rows


def _load_bundles_for_storage(db: sqlite3.Connection, group_id: int) -> list:
    """Tải bundles + dates cho storage/archive view (minimal — chỉ cần dates)."""
    bundles = db.execute(
        "SELECT id, sequence, total_sheets, cover_units FROM bundles WHERE group_id = ? ORDER BY sequence",
        (group_id,),
    ).fetchall()
    result = []
    for b in bundles:
        item_dates = db.execute(
            "SELECT de.transaction_date FROM bundle_items bi JOIN document_entries de ON bi.entry_id = de.id WHERE bi.bundle_id = ?",
            (b["id"],),
        ).fetchall()
        result.append({
            "id": b["id"],
            "sequence": b["sequence"],
            "total_sheets": b["total_sheets"],
            "cover_units": b["cover_units"],
            "items": [{"entry": {"transaction_date": r["transaction_date"]}} for r in item_dates],
        })
    return result


def _get_storage_rows_for_month(
    db: sqlite3.Connection, department_id: int, year: int, month: int
) -> tuple:
    notes_key = f"Tháng {month:02d}/{year}"
    groups = db.execute(
        "SELECT id FROM bundle_groups WHERE department_id = ? AND notes = ? ORDER BY created_at",
        (department_id, notes_key),
    ).fetchall()

    dept = db.execute("SELECT name FROM departments WHERE id = ?", (department_id,)).fetchone()
    dept_name = dept["name"] if dept else ""

    all_rows: list = []
    for g in groups:
        bundles_data = _load_bundles_for_storage(db, g["id"])
        all_rows.extend(_decompose_bundles_to_rows(bundles_data))
    return dept_name, all_rows


def _generate_archive_records(
    db: sqlite3.Connection, department_id: int, year: int, tieu_de_dau: str, tu_tap: str
) -> list:
    dept = db.execute("SELECT name FROM departments WHERE id = ?", (department_id,)).fetchone()
    dept_name = dept["name"] if dept else str(department_id)

    groups = db.execute(
        "SELECT id, notes FROM bundle_groups WHERE department_id = ? AND notes LIKE ? ORDER BY created_at",
        (department_id, f"Tháng %/{year}"),
    ).fetchall()

    by_month: dict = defaultdict(list)
    for g in groups:
        try:
            month = int((g["notes"] or "").split("/")[0].split(" ")[1])
        except (IndexError, ValueError):
            continue
        by_month[month].append(g["id"])

    records = []
    for month in range(1, 13):
        group_ids = by_month.get(month, [])
        if not group_ids:
            continue

        all_rows: list = []
        for gid in group_ids:
            bundles_data = _load_bundles_for_storage(db, gid)
            all_rows.extend(_decompose_bundles_to_rows(bundles_data))

        tieu_de_cuoi = f"{dept_name} tháng {month:02d}/{year}"
        for r in all_rows:
            days = r.days
            ds_ngay_full = ", ".join(f"{d:02d}/{month:02d}/{year}" for d in days)
            ngay_mo = f"{min(days):02d}/{month:02d}/{year}"
            ngay_kt = f"{max(days):02d}/{month:02d}/{year}"
            for stt in range(1, r.n_bundles + 1):
                if r.n_bundles == 1:
                    tieu_de = f"{tieu_de_dau} {ds_ngay_full} {tieu_de_cuoi}"
                else:
                    tieu_de = f"{tieu_de_dau} {ds_ngay_full} {tieu_de_cuoi} {tu_tap} {stt}/{r.n_bundles}"
                records.append(ArchiveRecord(ngay_mo=ngay_mo, ngay_kt=ngay_kt, tieu_de=tieu_de))
    return records


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/groups")
def list_groups(
    department_id: Optional[int] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("menu.bundles")),
):
    clauses = []
    params: list = []
    if department_id:
        clauses.append("bg.department_id = ?")
        params.append(department_id)
    if month:
        clauses.append("bg.notes LIKE ?")
        params.append(f"Tháng {month:02d}/%")
    if year:
        clauses.append("bg.notes LIKE ?")
        params.append(f"%/{year}")

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = db.execute(
        f"""SELECT bg.*, d.code AS dept_code, d.name AS dept_name,
                   d.is_source AS d_is_source, d.is_active AS d_is_active
            FROM bundle_groups bg LEFT JOIN departments d ON bg.department_id = d.id
            {where} ORDER BY bg.created_at DESC""",
        params,
    ).fetchall()

    result = []
    for g in rows:
        creator = db.execute("SELECT * FROM user_tttt WHERE id = ?", (g["created_by_id"],)).fetchone()
        bundle_rows = db.execute(
            """SELECT b.id, b.sequence, b.total_sheets, b.custodian_id, b.storage_box,
                      b.storage_location, b.cover_printed_at, b.status, b.cover_units,
                      b.group_id
               FROM bundles b WHERE b.group_id = ? ORDER BY b.sequence""",
            (g["id"],),
        ).fetchall()

        bundles_out = []
        for b in bundle_rows:
            item_rows = db.execute(
                "SELECT id, entry_id FROM bundle_items WHERE bundle_id = ?", (b["id"],)
            ).fetchall()
            bundles_out.append({
                "id": b["id"],
                "group_id": b["group_id"],
                "sequence": b["sequence"],
                "total_sheets": b["total_sheets"],
                "custodian_id": b["custodian_id"],
                "storage_box": b["storage_box"],
                "storage_location": b["storage_location"],
                "cover_printed_at": b["cover_printed_at"],
                "status": b["status"] or "pending",
                "cover_units": b["cover_units"],
                "items": [{"id": r["id"], "entry_id": r["entry_id"], "entry": None} for r in item_rows],
            })

        dept_dict = None
        if g["dept_name"]:
            dept_dict = {
                "id": g["department_id"],
                "code": g["dept_code"] or "",
                "name": g["dept_name"] or "",
                "is_source": bool(g["d_is_source"]),
                "is_active": bool(g["d_is_active"]),
            }

        creator_dict = None
        if creator:
            creator_dict = {
                "id": creator["id"],
                "employee_code": creator["employee_code"] or "",
                "full_name": creator["full_name"] or "",
                "role": creator["role"] or "",
                "department_id": creator["department_id"],
                "username": creator["username"] or "",
                "phone": creator["phone"],
                "email": creator["email"],
                "start_date": creator["start_date"],
                "ipcas_code": creator["ipcas_code"],
                "payment_username": creator["payment_username"],
                "is_active": bool(creator["is_active"]),
            }

        result.append({
            "id": g["id"],
            "department_id": g["department_id"],
            "total_bundles": g["total_bundles"],
            "created_at": g["created_at"],
            "notes": g["notes"],
            "department": dept_dict,
            "created_by_staff": creator_dict,
            "bundles": bundles_out,
        })

    return result


@router.get("/groups/{group_id}")
def get_group(
    group_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("menu.bundles")),
):
    return _load_bundle_group(db, group_id)


@router.post("/generate")
def generate_bundles(
    req: BundleGenerateRequest,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("bundles.generate")),
):
    """Tự động gom tập từ danh sách entry IDs"""
    dept = db.execute("SELECT id FROM departments WHERE id = ?", (req.department_id,)).fetchone()
    if not dept:
        raise HTTPException(404, "Không tìm thấy phòng")

    entry_ids = list(dict.fromkeys(req.entry_ids))
    if not entry_ids:
        raise HTTPException(400, "Không có chứng từ để gom")

    ph = ",".join("?" * len(entry_ids))
    entries = db.execute(
        f"""SELECT de.id, de.handover_id, de.staff_id, de.transaction_date, de.sheet_count,
                   de.entry_status, h.department_id AS h_dept_id, ks.role AS staff_role,
                   ks.ipcas_code, ks.full_name, ks.payment_username
            FROM document_entries de
            JOIN handovers h ON de.handover_id = h.id
            LEFT JOIN user_tttt ks ON de.staff_id = ks.id
            WHERE de.id IN ({ph})""",
        entry_ids,
    ).fetchall()

    if not entries:
        raise HTTPException(400, "Không có chứng từ để gom")

    # Chỉ gom các entry đã xác nhận
    entries = [e for e in entries if (e["entry_status"] or "confirmed") == "confirmed"]
    if not entries:
        raise HTTPException(400, "Không có chứng từ đã xác nhận để gom tập")

    # Chỉ gom chứng từ của giao dịch viên (chuyen_vien) — loại entry rác của trưởng/phó phòng
    entries = [e for e in entries if e["staff_role"] == StaffRole.CHUYEN_VIEN.value]
    if not entries:
        raise HTTPException(400, "Không có chứng từ của giao dịch viên để gom tập")

    # Xóa bundle group cũ cùng phòng+tháng để regenerate
    if req.notes:
        old_groups = db.execute(
            "SELECT id FROM bundle_groups WHERE department_id = ? AND notes = ?",
            (req.department_id, req.notes),
        ).fetchall()
        for og in old_groups:
            _delete_group_cascade(db, og["id"])

    # Kiểm tra tất cả entry thuộc đúng phòng
    invalid = [e["id"] for e in entries if e["h_dept_id"] != req.department_id]
    if invalid:
        raise HTTPException(400, "Có chứng từ không thuộc phòng đã chọn")

    # Kiểm tra entry đã thuộc tập đang hiệu lực (tránh gom trùng)
    entry_ids = [e["id"] for e in entries]
    placeholders = ",".join("?" * len(entry_ids))
    already_bundled = db.execute(
        f"""SELECT bi.entry_id, bg.notes
            FROM bundle_items bi
            JOIN bundles b ON bi.bundle_id = b.id
            JOIN bundle_groups bg ON b.group_id = bg.id
            WHERE bi.entry_id IN ({placeholders})""",
        entry_ids,
    ).fetchall()
    if already_bundled:
        conflicts = ", ".join(str(r["entry_id"]) for r in already_bundled)
        raise HTTPException(
            409, f"Chứng từ #{conflicts} đã thuộc tập khác. Hủy tập cũ trước khi gom lại."
        )

    entries_data = [
        {
            "id": e["id"],
            "source_user_id": e["staff_id"],
            "user_code": e["ipcas_code"] or str(e["staff_id"]),
            "full_name": e["full_name"],
            "transaction_date": date_type.fromisoformat(e["transaction_date"]),
            "sheet_count": e["sheet_count"],
        }
        for e in entries
    ]

    bundle_results = generate_bundles_for_entries(entries_data)
    if not bundle_results:
        raise HTTPException(400, "Không thể gom tập")

    # Lưu vào DB
    cur = db.execute(
        "INSERT INTO bundle_groups (department_id, total_bundles, created_by_id, notes, created_at) VALUES (?,?,?,?,?)",
        (req.department_id, len(bundle_results), current["id"], req.notes, str(_vn_now())),
    )
    group_id = cur.lastrowid

    for br in bundle_results:
        cover_units_json = json.dumps(
            [
                {
                    "user_code": unit.user_code,
                    "full_name": unit.full_name,
                    "date": unit.transaction_date.isoformat(),
                    "sheet_count": unit.sheet_count,
                    "is_large": unit.is_large,
                }
                for unit in br.units
            ],
            ensure_ascii=False,
        )
        cur2 = db.execute(
            "INSERT INTO bundles (group_id, sequence, total_sheets, custodian_id, status, cover_units) VALUES (?,?,?,?,?,?)",
            (group_id, br.sequence, br.total_sheets, req.custodian_id, "pending", cover_units_json),
        )
        bundle_id = cur2.lastrowid
        for unit in br.units:
            for eid in unit.entry_ids:
                db.execute("INSERT INTO bundle_items (bundle_id, entry_id) VALUES (?,?)", (bundle_id, eid))

    db.commit()
    return _load_bundle_group(db, group_id)


@router.put("/{bundle_id}")
def update_bundle(
    bundle_id: int,
    req: BundleUpdateRequest,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("bundles.generate")),
):
    b = db.execute("SELECT * FROM bundles WHERE id = ?", (bundle_id,)).fetchone()
    if not b:
        raise HTTPException(404, "Không tìm thấy tập")
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if update_data:
        sets = ", ".join(f"{k}=?" for k in update_data)
        db.execute(f"UPDATE bundles SET {sets} WHERE id=?", list(update_data.values()) + [bundle_id])
        db.commit()
    return _load_bundle_group(db, b["group_id"])


@router.get("/{bundle_id}/cover")
def download_cover(
    bundle_id: int,
    custodian_id: Optional[int] = None,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    b_row = db.execute("SELECT * FROM bundles WHERE id = ?", (bundle_id,)).fetchone()
    if not b_row:
        raise HTTPException(404, "Không tìm thấy tập")

    group = _load_bundle_group(db, b_row["group_id"])
    bundle = next((b for b in group["bundles"] if b["id"] == bundle_id), None)
    if not bundle:
        raise HTTPException(404, "Không tìm thấy tập")

    label_seq, label_total = _get_bundle_label(bundle, group["bundles"])

    cust_name = "..."
    cid = custodian_id or b_row["custodian_id"]
    if cid:
        cust = db.execute("SELECT full_name FROM user_tttt WHERE id = ?", (cid,)).fetchone()
        if cust:
            cust_name = cust["full_name"]

    units = _units_from_bundle(bundle)
    br = BundleResult(
        sequence=bundle["sequence"],
        total_bundles_in_group=group["total_bundles"],
        total_sheets=bundle["total_sheets"],
        units=units,
        label_seq=label_seq,
        label_total=label_total,
        custodian_name=cust_name,
    )
    dept_name = group["department"]["name"] if group.get("department") else "Phòng"
    docx_bytes = generate_covers_docx(dept_name, [br])
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers=_download_headers(f"bia_tap_{bundle['sequence']}.docx"),
    )


@router.get("/groups/{group_id}/cover-all")
async def download_all_covers(
    group_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("bundles.download_cover")),
):
    # ~100ms mỗi tập (một lần render docxtpl) — nhóm 31 tập mất ~3,1s
    def _work() -> bytes:
        group = _load_bundle_group(db, group_id)
        bundle_results = []
        for bundle in sorted(group["bundles"], key=lambda b: b["sequence"]):
            cust_name = bundle.get("custodian_name") or "..."
            label_seq, label_total = _get_bundle_label(bundle, group["bundles"])
            bundle_results.append(BundleResult(
                sequence=bundle["sequence"],
                total_bundles_in_group=group["total_bundles"],
                total_sheets=bundle["total_sheets"],
                units=_units_from_bundle(bundle),
                label_seq=label_seq,
                label_total=label_total,
                custodian_name=cust_name,
            ))
        dept_name = group["department"]["name"] if group.get("department") else "Phòng"
        return generate_covers_docx(dept_name, bundle_results)

    docx_bytes = await run_heavy(_work)
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers=_download_headers(f"bia_tat_ca_tap_{group_id}.docx"),
    )


@router.post("/{bundle_id}/mark-printed")
def mark_bundle_printed(
    bundle_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("bundles.mark_printed")),
):
    b = db.execute("SELECT group_id FROM bundles WHERE id = ?", (bundle_id,)).fetchone()
    if not b:
        raise HTTPException(404, "Không tìm thấy tập")
    db.execute(
        "UPDATE bundles SET cover_printed_at=?, status='printed' WHERE id=?",
        (str(_vn_now()), bundle_id),
    )
    db.commit()
    return _load_bundle_group(db, b["group_id"])


@router.post("/groups/{group_id}/mark-printed")
def mark_group_printed(
    group_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("bundles.mark_printed")),
):
    g = db.execute("SELECT id FROM bundle_groups WHERE id = ?", (group_id,)).fetchone()
    if not g:
        raise HTTPException(404, "Không tìm thấy nhóm tập")
    db.execute(
        "UPDATE bundles SET cover_printed_at=?, status='printed' WHERE group_id=?",
        (str(_vn_now()), group_id),
    )
    db.commit()
    return _load_bundle_group(db, group_id)


@router.get("/cover-bulk")
async def download_bulk_covers(
    department_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("bundles.download_cover")),
):
    # Cả phòng — nặng hơn cover-all vì gộp mọi nhóm tập
    def _work() -> bytes:
        dept = db.execute("SELECT * FROM departments WHERE id = ?", (department_id,)).fetchone()
        if not dept:
            raise HTTPException(404, "Không tìm thấy phòng")

        group_rows = db.execute(
            "SELECT id FROM bundle_groups WHERE department_id = ? ORDER BY created_at ASC",
            (department_id,),
        ).fetchall()
        if not group_rows:
            raise HTTPException(404, "Không có nhóm tập nào cho phòng này")

        all_bundle_results = []
        for gr in group_rows:
            group = _load_bundle_group(db, gr["id"])
            for bundle in sorted(group["bundles"], key=lambda b: b["sequence"]):
                cust_name = bundle.get("custodian_name") or "..."
                label_seq, label_total = _get_bundle_label(bundle, group["bundles"])
                all_bundle_results.append(BundleResult(
                    sequence=bundle["sequence"],
                    total_bundles_in_group=group["total_bundles"],
                    total_sheets=bundle["total_sheets"],
                    units=_units_from_bundle(bundle),
                    label_seq=label_seq,
                    label_total=label_total,
                    custodian_name=cust_name,
                ))

        if not all_bundle_results:
            raise HTTPException(404, "Không có tập nào để tải bìa")

        return generate_covers_docx(dept["name"], all_bundle_results)

    docx_bytes = await run_heavy(_work)
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers=_download_headers(f"bia_phong_{department_id}.docx"),
    )


@router.get("/storage-view", response_model=StorageViewResponse)
def storage_view(
    department_id: int,
    year: int,
    month: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("menu.storage")),
):
    dept_name, all_rows = _get_storage_rows_for_month(db, department_id, year, month)
    notes_key = f"Tháng {month:02d}/{year}"
    return StorageViewResponse(
        department_name=dept_name,
        period=notes_key,
        rows=all_rows,
        total_sheets=sum(sum(r.bundle_sheets) for r in all_rows),
        total_bundles=sum(r.n_bundles for r in all_rows),
    )


@router.get("/storage-summary", response_model=StorageSummaryResponse)
def storage_summary(
    year: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("menu.storage")),
):
    depts = db.execute(
        "SELECT id, name FROM departments WHERE is_source = 1 ORDER BY id"
    ).fetchall()

    # Dùng lại đúng hàm dựng bảng chi tiết để tổng hợp luôn khớp với /storage-view.
    # Không gộp thành 1 câu SQL SUM/COUNT: "tập có ngày" phụ thuộc cover_units JSON
    # (xem _get_dates_for_bundle), không diễn đạt được bằng SQL.
    rows = []
    for month in range(1, 13):
        cells = []
        for d in depts:
            _, det_rows = _get_storage_rows_for_month(db, d["id"], year, month)
            cells.append(StorageSummaryCell(
                department_id=d["id"],
                total_sheets=sum(sum(r.bundle_sheets) for r in det_rows),
                total_bundles=sum(r.n_bundles for r in det_rows),
            ))
        rows.append(StorageSummaryRow(
            month=month,
            cells=cells,
            total_sheets=sum(c.total_sheets for c in cells),
            total_bundles=sum(c.total_bundles for c in cells),
        ))

    return StorageSummaryResponse(
        year=year,
        departments=[StorageSummaryDept(id=d["id"], name=d["name"]) for d in depts],
        rows=rows,
    )


def _bundle_dates(db: sqlite3.Connection, bundle_id: int) -> list:
    """Ngày của một bundle — dùng cover_units, fallback qua items (giống _get_dates_for_bundle)."""
    b = db.execute("SELECT id, cover_units FROM bundles WHERE id = ?", (bundle_id,)).fetchone()
    if not b:
        return []
    item_dates = db.execute(
        "SELECT de.transaction_date FROM bundle_items bi "
        "JOIN document_entries de ON bi.entry_id = de.id WHERE bi.bundle_id = ?",
        (bundle_id,),
    ).fetchall()
    return sorted(_get_dates_for_bundle({
        "id": b["id"],
        "cover_units": b["cover_units"],
        "items": [{"entry": {"transaction_date": r["transaction_date"]}} for r in item_dates],
    }))


def _delete_bundle(db: sqlite3.Connection, bundle_id: int):
    db.execute("DELETE FROM bundle_items WHERE bundle_id = ?", (bundle_id,))
    db.execute("DELETE FROM bundles WHERE id = ?", (bundle_id,))


def _group_year_month(db: sqlite3.Connection, group_id: int) -> Tuple[int, int]:
    """Tháng/năm của nhóm tập, đọc từ notes 'Tháng MM/YYYY' — không tin tham số client."""
    g = db.execute("SELECT notes FROM bundle_groups WHERE id = ?", (group_id,)).fetchone()
    try:
        mm, yyyy = (g["notes"] or "").split(" ")[1].split("/")
        return int(yyyy), int(mm)
    except (TypeError, IndexError, ValueError):
        raise HTTPException(400, "Không xác định được tháng của nhóm tập, không thể sửa ngày")


def _row_dates(days: list, year: int, month: int) -> list:
    """Ngày người dùng nhập (1–31) → list date đã sắp xếp, bỏ trùng."""
    out = set()
    for d in days:
        try:
            out.add(date_type(year, month, int(d)))
        except (TypeError, ValueError):
            raise HTTPException(400, f"Ngày {d} không hợp lệ trong tháng {month:02d}/{year}")
    if not out:
        raise HTTPException(400, "Mỗi dòng phải có ít nhất một ngày")
    return sorted(out)


def _bundle_units_raw(db: sqlite3.Connection, bundle_id: int, cover_units) -> list:
    """Unit của tập dạng dict (date là object) — ưu tiên cover_units, fallback qua items."""
    if cover_units:
        try:
            return [
                {"user_code": u.get("user_code") or "", "full_name": u.get("full_name") or "",
                 "date": date_type.fromisoformat(u["date"]),
                 "sheet_count": u.get("sheet_count") or 0, "is_large": u.get("is_large", False)}
                for u in json.loads(cover_units)
            ]
        except Exception:
            _log.warning("bundle %s: cover_units JSON lỗi, dựng lại unit từ items", bundle_id)
    rows = db.execute(
        "SELECT de.transaction_date, de.sheet_count, ks.ipcas_code, ks.full_name "
        "FROM bundle_items bi JOIN document_entries de ON bi.entry_id = de.id "
        "LEFT JOIN user_tttt ks ON de.staff_id = ks.id WHERE bi.bundle_id = ?",
        (bundle_id,),
    ).fetchall()
    return [
        {"user_code": r["ipcas_code"] or "", "full_name": r["full_name"] or "",
         "date": date_type.fromisoformat(r["transaction_date"]),
         "sheet_count": r["sheet_count"] or 0, "is_large": False}
        for r in rows
    ]


def _dump_units(units: list) -> str:
    return json.dumps(
        [{"user_code": u["user_code"], "full_name": u["full_name"], "date": u["date"].isoformat(),
          "sheet_count": u["sheet_count"], "is_large": u["is_large"]} for u in units],
        ensure_ascii=False,
    )


def _rewrite_bundle_dates(db: sqlite3.Connection, bundle_id: int, new_dates: list):
    """Đổi ngày của tập bằng cách ghi lại cover_units — KHÔNG sửa document_entries."""
    b = db.execute("SELECT id, total_sheets, cover_units FROM bundles WHERE id = ?", (bundle_id,)).fetchone()
    if not b:
        return
    units = _bundle_units_raw(db, bundle_id, b["cover_units"])
    old_dates = sorted({u["date"] for u in units})
    if old_dates == new_dates:
        return

    if units:
        # Ánh xạ theo vị trí: ngày cũ thứ i → ngày mới thứ i; ngày cũ dư dồn vào ngày mới cuối
        remap = {d: new_dates[min(i, len(new_dates) - 1)] for i, d in enumerate(old_dates)}
        units = [{**u, "date": remap[u["date"]]} for u in units]
    else:
        units = [{"user_code": "", "full_name": "", "date": new_dates[0],
                  "sheet_count": b["total_sheets"], "is_large": False}]

    # Ngày mới chưa có unit nào → thêm unit rỗng, nếu không ngày đó biến mất khỏi bảng
    covered = {u["date"] for u in units}
    units += [{"user_code": "", "full_name": "", "date": d, "sheet_count": 0, "is_large": False}
              for d in new_dates if d not in covered]
    units.sort(key=lambda u: u["date"])

    db.execute("UPDATE bundles SET cover_units=? WHERE id=?", (_dump_units(units), bundle_id))


@router.patch("/storage-view")
def update_storage_view(
    req: StorageViewUpdateRequest,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("menu.storage")),
):
    touched_groups: set = set()

    for row in req.rows:
        anchor_id = row.bundle_ids[0] if row.bundle_ids else None
        anchor = (
            db.execute("SELECT group_id FROM bundles WHERE id = ?", (anchor_id,)).fetchone()
            if anchor_id else None
        )
        group_id = anchor["group_id"] if anchor else None
        if group_id is not None:
            touched_groups.add(group_id)

        # ── Sửa ngày: ghi trước khi thêm tập mới để tập mới nhận ngày mới ──
        new_dates = None
        if row.days is not None and group_id is not None:
            year, month = _group_year_month(db, group_id)
            new_dates = _row_dates(row.days, year, month)
            for i, bundle_id in enumerate(row.bundle_ids):
                # Tập sắp bị xoá (số chứng từ = 0) thì không cần đổi ngày
                if i < len(row.bundle_sheets) and row.bundle_sheets[i] <= 0:
                    continue
                _rewrite_bundle_dates(db, bundle_id, new_dates)

        # ── Thêm tập mới cho các ô trống được nhập (dùng ngày của dòng) ──
        adds = [s for s in row.new_sheets if s and s > 0]
        if adds and anchor_id and group_id is not None:
            dates = new_dates or _bundle_dates(db, anchor_id)
            seq = (db.execute(
                "SELECT COALESCE(MAX(sequence), 0) FROM bundles WHERE group_id = ?", (group_id,)
            ).fetchone()[0])
            for sheets in adds:
                seq += 1
                cover_units_json = json.dumps(
                    [{"user_code": "", "full_name": "", "date": d.isoformat(),
                      "sheet_count": sheets if i == 0 else 0, "is_large": False}
                     for i, d in enumerate(dates)],
                    ensure_ascii=False,
                ) if dates else None
                db.execute(
                    "INSERT INTO bundles (group_id, sequence, total_sheets, status, cover_units) "
                    "VALUES (?,?,?,?,?)",
                    (group_id, seq, sheets, "pending", cover_units_json),
                )

        # ── Cập nhật / xoá tập hiện có (0 = xoá) ──
        for i, bundle_id in enumerate(row.bundle_ids):
            if i >= len(row.bundle_sheets):
                continue
            val = row.bundle_sheets[i]
            if val <= 0:
                _delete_bundle(db, bundle_id)
            else:
                db.execute("UPDATE bundles SET total_sheets=? WHERE id=?", (val, bundle_id))

    # ── Tính lại tổng số tập của các group bị ảnh hưởng ──
    for gid in touched_groups:
        cnt = db.execute("SELECT COUNT(*) FROM bundles WHERE group_id = ?", (gid,)).fetchone()[0]
        db.execute("UPDATE bundle_groups SET total_bundles=? WHERE id=?", (cnt, gid))

    db.commit()
    return {"ok": True}


@router.get("/handover-archive", response_model=HandoverArchiveResponse)
def handover_archive_preview(
    department_id: int,
    year: int,
    tieu_de_dau: str = "Hồ sơ ngày",
    tu_tap: str = "tập",
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    records = _generate_archive_records(db, department_id, year, tieu_de_dau, tu_tap)
    return HandoverArchiveResponse(records=records, total=len(records))


@router.get("/handover-archive-excel")
async def handover_archive_excel(
    department_id: int,
    year: int,
    tieu_de_dau: str = "Hồ sơ ngày",
    tu_tap: str = "tập",
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    def _work() -> tuple[bytes, str]:
        records = _generate_archive_records(db, department_id, year, tieu_de_dau, tu_tap)
        dept = db.execute("SELECT name FROM departments WHERE id = ?", (department_id,)).fetchone()
        dept_name = dept["name"] if dept else str(department_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bàn giao lưu trữ"

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_h = Font(name="Times New Roman", size=11, bold=True)
        font_d = Font(name="Times New Roman", size=11)
        align_c = Alignment(horizontal="center", vertical="center")
        align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col, hdr in enumerate(["NGAY_MO_HS", "NGAY_KT_HS", "TIEUDE_HS"], 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.font = font_h
            cell.border = border
            cell.alignment = align_c

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 75

        for row_i, rec in enumerate(records, 2):
            for col, (val, aln) in enumerate(
                [(rec.ngay_mo, align_c), (rec.ngay_kt, align_c), (rec.tieu_de, align_l)], 1
            ):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.font = font_d
                cell.border = border
                cell.alignment = aln

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read(), dept_name

    content, dept_name = await run_heavy(_work)
    safe_name = dept_name.replace("/", "-").replace("\\", "-")
    filename = f"ban_giao_luu_tru_{safe_name}_{year}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_download_headers(filename),
    )


# ─── In bìa hồ sơ lưu trữ (mẫu M01/LHS) ──────────────────────────────────────

_ARCHIVE_COVER_MAX_UPLOAD = 20 * 1024 * 1024
_ARCHIVE_COVER_MAX_ROWS = 2000
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


@router.post("/archive-cover-parse", response_model=ArchiveCoverParseResponse)
async def archive_cover_parse(
    file: UploadFile = File(...),
    _: dict = Depends(require_feature("menu.storage")),
):
    """Đọc file Excel tra cứu hồ sơ → danh sách dữ liệu để in bìa."""
    name = file.filename or ""
    if not name.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(400, "Chỉ nhận file Excel (.xls hoặc .xlsx)")

    content = await file.read()
    if not content:
        raise HTTPException(400, "File rỗng")
    if len(content) > _ARCHIVE_COVER_MAX_UPLOAD:
        raise HTTPException(400, "File quá lớn (tối đa 20MB)")

    def _work():
        return archive_cover_service.parse_lookup_excel(content, name)

    try:
        records = await run_heavy(_work)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        _log.exception("archive-cover-parse: đọc file %s thất bại", name)
        raise HTTPException(400, f"Không đọc được file Excel: {e}")

    if not records:
        raise HTTPException(400, "File không có dòng dữ liệu nào trong sheet 'Data'")

    rows, warnings = [], []
    for i, r in enumerate(records, 1):
        rows.append(ArchiveCoverRow(
            stt=i, ma_vach=r.ma_vach, ngay_mo=r.ngay_mo,
            tieu_de=r.tieu_de, ngay_cvkt=r.ngay_cvkt, so_to=r.so_to,
        ))
    no_date = [r.stt for r in rows if not r.ngay_mo]
    if no_date:
        warnings.append(
            f"{len(no_date)} dòng không tìm thấy ngày trong tên hồ sơ "
            f"(STT {', '.join(map(str, no_date[:10]))}"
            f"{'…' if len(no_date) > 10 else ''}) — ô Ngày mở sẽ để trống"
        )
    no_code = [r.stt for r in rows if not r.ma_vach]
    if no_code:
        warnings.append(
            f"{len(no_code)} dòng thiếu mã vạch "
            f"(STT {', '.join(map(str, no_code[:10]))}{'…' if len(no_code) > 10 else ''})"
        )

    return ArchiveCoverParseResponse(rows=rows, total=len(rows), warnings=warnings)


@router.post("/archive-cover-print")
async def archive_cover_print(
    payload: ArchiveCoverPrintRequest,
    _: dict = Depends(require_feature("menu.storage")),
):
    """Sinh bìa hồ sơ: 1 file Word nhiều trang, hoặc ZIP mỗi hồ sơ một file."""
    if not payload.rows:
        raise HTTPException(400, "Chưa chọn hồ sơ nào để in bìa")
    if len(payload.rows) > _ARCHIVE_COVER_MAX_ROWS:
        raise HTTPException(400, f"Tối đa {_ARCHIVE_COVER_MAX_ROWS} hồ sơ một lần in")

    records = [
        archive_cover_service.ArchiveCoverRecord(
            ma_vach=r.ma_vach, ngay_mo=r.ngay_mo, tieu_de=r.tieu_de,
            ngay_cvkt=r.ngay_cvkt, so_to=r.so_to or "1",
        )
        for r in payload.rows
    ]
    as_zip = payload.as_zip

    def _work() -> bytes:
        if as_zip:
            return archive_cover_service.generate_covers_zip(records)
        return archive_cover_service.generate_covers(records)

    try:
        content = await run_heavy(_work)
    except (FileNotFoundError, archive_cover_service.TemplateMismatchError) as e:
        _log.error("archive-cover-print: %s", e)
        raise HTTPException(500, str(e))

    if as_zip:
        return Response(content=content, media_type="application/zip",
                        headers=_download_headers("bia_ho_so.zip"))
    return Response(content=content, media_type=_DOCX_MIME,
                    headers=_download_headers("bia_ho_so.docx"))


@router.delete("/groups/{group_id}")
def delete_group(
    group_id: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(require_feature("bundles.delete")),
):
    g = db.execute("SELECT id FROM bundle_groups WHERE id = ?", (group_id,)).fetchone()
    if not g:
        raise HTTPException(404, "Không tìm thấy nhóm tập")
    _delete_group_cascade(db, group_id)
    db.commit()
    return {"message": "Đã xóa nhóm tập"}
