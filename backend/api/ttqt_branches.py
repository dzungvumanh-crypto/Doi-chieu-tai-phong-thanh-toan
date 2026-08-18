"""Danh sách CN thực hiện TTQT — tra cứu, CRUD, import/export Excel.

Nguồn gốc dữ liệu là file Excel "Danh sách CN thực hiện TTQT" do Phòng KSNB
phát hành. File có 1 dòng phân cách ghi "Đóng BICCODE": mọi dòng phía dưới là
chi nhánh đã đóng BIC. Import đọc đúng quy ước đó, export ghi lại y hệt để file
xuất ra nhập lại được (round-trip).
"""
import io
import logging
import sqlite3
import unicodedata
from datetime import date, datetime
from typing import Optional
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.core.concurrency import run_heavy
from backend.core.uploads import read_limited
from backend.core.deps import require_feature
from backend.database import get_db, write_audit, _vn_now
from backend.schemas.ttqt_branches import BranchCreate, BranchOut, BranchUpdate, ImportResult

_log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/ttqt-branches", tags=["TTQT Branches"])

# Thứ tự cột dùng chung cho cả import lẫn export
_COLS = [
    ("ma_cn",      "MÃ CN",              12),
    ("ten_cn",     "TÊN CN",             28),
    ("swift_bic",  "SWIFT BIC",          16),
    ("loai_cn",    "LOẠI CN",            10),
    ("duoc_phep",  "ĐƯỢC PHÉP",          14),
    ("cn_quan_ly", "CN LOẠI I QUẢN LÝ",  20),
    ("ghi_chu",    "GHI CHÚ",            42),
    ("sdt",        "SĐT",                20),
    ("dia_chi",    "ĐỊA CHỈ",            50),
    ("dia_chi_en", "ĐỊA CHỈ TIẾNG ANH",  50),
]
_FIELDS = [f for f, _, _ in _COLS]
_CLOSED_MARKER = "ĐÓNG BICCODE"
_MAX_BLANK_RUN = 50      # số dòng trống liên tiếp thì coi như hết dữ liệu


def _download_headers(filename: str) -> dict:
    fallback = "".join(ch if ord(ch) < 128 and ch not in '\\"' else "_" for ch in filename)
    return {
        "Content-Disposition": (
            f'attachment; filename="{fallback}"; '
            f"filename*=UTF-8''{quote(filename, safe='')}"
        )
    }


def _fold(s) -> str:
    """Hạ chữ theo Unicode để so khớp.

    KHÔNG dùng `LOWER()` của SQLite: hàm đó chỉ hạ được A-Z, để nguyên 'Đ'
    (U+0110) và mọi chữ hoa có dấu. Hậu quả đã xảy ra thật: `LOWER(ten_cn)` ra
    "Điện biên" trong khi từ khoá gửi lên là "điện biên" → 43/218 chi nhánh tìm
    không ra, danh sách rỗng, không một dòng cảnh báo nào.
    """
    return unicodedata.normalize("NFC", str(s or "")).casefold()


def _no_tone(s: str) -> str:
    """Bỏ dấu để gõ không dấu vẫn tìm ra: 'điện biên' → 'dien bien'.
    'đ'/'Đ' không có phân giải NFD nên phải thay tay."""
    s = "".join(ch for ch in unicodedata.normalize("NFD", s)
                if not unicodedata.combining(ch))
    return s.replace("đ", "d").replace("Đ", "D")


def _row_out(r: sqlite3.Row) -> dict:
    d = dict(r)
    d["is_closed"] = bool(d.get("is_closed"))
    d["updated_at"] = str(d["updated_at"]) if d.get("updated_at") else None
    return d


# ─── Tra cứu ─────────────────────────────────────────────────────────────────
@router.get("/", response_model=list[BranchOut])
def list_branches(
    q: Optional[str] = Query(None, description="Tìm theo mã CN / tên CN / SWIFT BIC"),
    loai_cn: Optional[int] = Query(None, description="1 hoặc 2; bỏ trống = tất cả"),
    status: str = Query("active", pattern="^(active|closed|all)$"),
    _: dict = Depends(require_feature("menu.ttqt_branches")),
    db: sqlite3.Connection = Depends(get_db),
):
    where, params = [], []
    if status == "active":
        where.append("is_closed = 0")
    elif status == "closed":
        where.append("is_closed = 1")
    if loai_cn in (1, 2):
        where.append("loai_cn = ?")
        params.append(loai_cn)
    sql = "SELECT * FROM ttqt_branches"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY is_closed, IFNULL(sort_order, 999999), ma_cn"
    rows = [_row_out(r) for r in db.execute(sql, params).fetchall()]

    # Lọc từ khoá bằng Python, không bằng SQL LIKE — xem _fold(). Bảng chỉ vài
    # trăm dòng nên đọc hết rồi lọc rẻ hơn nhiều so với việc phải cài hàm
    # so-sánh Unicode riêng cho SQLite trên mọi kết nối.
    if q and q.strip():
        kw = _fold(q.strip())
        kw_plain = _no_tone(kw)

        def _hit(b: dict) -> bool:
            hay = _fold(f"{b['ma_cn']} {b['ten_cn']} {b.get('swift_bic') or ''}")
            # Khớp có dấu HOẶC khớp sau khi bỏ dấu — gõ "dien bien" cũng ra
            return kw in hay or kw_plain in _no_tone(hay)

        rows = [b for b in rows if _hit(b)]
    return rows


# ─── Thêm / sửa / xoá ────────────────────────────────────────────────────────
@router.post("/", response_model=BranchOut)
def create_branch(
    body: BranchCreate,
    current: dict = Depends(require_feature("ttqt_branches.create")),
    db: sqlite3.Connection = Depends(get_db),
):
    if db.execute("SELECT 1 FROM ttqt_branches WHERE ma_cn = ?", (body.ma_cn,)).fetchone():
        raise HTTPException(400, f"Mã CN {body.ma_cn} đã tồn tại")
    # sort_order đẩy xuống cuối nhóm (đang hoạt động / đã đóng) tương ứng
    max_order = db.execute(
        "SELECT IFNULL(MAX(sort_order), 0) FROM ttqt_branches WHERE is_closed = ?",
        (int(body.is_closed),),
    ).fetchone()[0]
    vals = [getattr(body, f) for f in _FIELDS]
    cur = db.execute(
        f"INSERT INTO ttqt_branches ({','.join(_FIELDS)}, is_closed, sort_order, updated_at)"
        f" VALUES ({','.join('?' * len(_FIELDS))}, ?, ?, ?)",
        (*vals, int(body.is_closed), max_order + 1, _vn_now()),
    )
    write_audit(db, current["id"], "ttqt_branch.create", "ttqt_branch", cur.lastrowid,
                f"{body.ma_cn} — {body.ten_cn}")
    db.commit()
    row = db.execute("SELECT * FROM ttqt_branches WHERE id = ?", (cur.lastrowid,)).fetchone()
    return _row_out(row)


@router.patch("/{branch_id}", response_model=BranchOut)
def update_branch(
    branch_id: int,
    body: BranchUpdate,
    current: dict = Depends(require_feature("ttqt_branches.edit")),
    db: sqlite3.Connection = Depends(get_db),
):
    old = db.execute("SELECT * FROM ttqt_branches WHERE id = ?", (branch_id,)).fetchone()
    if not old:
        raise HTTPException(404, "Không tìm thấy chi nhánh")
    dup = db.execute(
        "SELECT 1 FROM ttqt_branches WHERE ma_cn = ? AND id <> ?", (body.ma_cn, branch_id)
    ).fetchone()
    if dup:
        raise HTTPException(400, f"Mã CN {body.ma_cn} đã thuộc chi nhánh khác")
    db.execute(
        f"UPDATE ttqt_branches SET {','.join(f + '=?' for f in _FIELDS)},"
        f" is_closed=?, updated_at=? WHERE id = ?",
        (*[getattr(body, f) for f in _FIELDS], int(body.is_closed), _vn_now(), branch_id),
    )
    changed = [f for f in _FIELDS if (old[f] or None) != getattr(body, f)]
    if bool(old["is_closed"]) != body.is_closed:
        changed.append("is_closed")
    write_audit(db, current["id"], "ttqt_branch.update", "ttqt_branch", branch_id,
                f"{body.ma_cn} — sửa: {', '.join(changed) or 'không đổi'}")
    db.commit()
    row = db.execute("SELECT * FROM ttqt_branches WHERE id = ?", (branch_id,)).fetchone()
    return _row_out(row)


@router.delete("/{branch_id}")
def delete_branch(
    branch_id: int,
    current: dict = Depends(require_feature("ttqt_branches.delete")),
    db: sqlite3.Connection = Depends(get_db),
):
    row = db.execute("SELECT ma_cn, ten_cn FROM ttqt_branches WHERE id = ?", (branch_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy chi nhánh")
    db.execute("DELETE FROM ttqt_branches WHERE id = ?", (branch_id,))
    write_audit(db, current["id"], "ttqt_branch.delete", "ttqt_branch", branch_id,
                f"{row['ma_cn']} — {row['ten_cn']}")
    db.commit()
    return {"ok": True}


# ─── Import Excel ────────────────────────────────────────────────────────────
def _norm_header(v) -> str:
    return unicodedata.normalize("NFC", " ".join(str(v).split()).upper()) if v is not None else ""


def _cell_str(v) -> Optional[str]:
    """Ô Excel → chuỗi. Ngày (một số ô GHI CHÚ là datetime) → dd/mm/yyyy;
    số nguyên dạng float (1000.0) → '1000' để mã CN không dính đuôi '.0'."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    # Chuẩn hoá NFC: file gốc ghi tiếng Việt dạng tổ hợp ("ế" = "ê" + dấu sắc
    # rời) ở 11/218 dòng. Bộ gõ tiếng Việt thông thường sinh NFC, nên để nguyên
    # thì tìm theo tên KHÔNG ra kết quả — sai im lặng, không báo lỗi.
    s = unicodedata.normalize("NFC", str(v)).strip()
    return s or None


def _parse_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    """Đọc file Excel → danh sách dict theo _FIELDS + is_closed. Trả kèm cảnh báo."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Không đọc được file Excel: {e}")
    ws = wb.worksheets[0]

    # ── Tìm dòng tiêu đề cột (không cố định — file có dòng tên bảng ở trên) ──
    header_idx, col_map = None, {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        labels = {_norm_header(v): j for j, v in enumerate(row) if v is not None}
        if "MÃ CN" in labels and "TÊN CN" in labels:
            header_idx = i
            col_map = {f: labels[lbl] for f, lbl, _ in _COLS if lbl in labels}
            break
    if header_idx is None:
        raise HTTPException(400, "Không tìm thấy dòng tiêu đề có cột 'MÃ CN' và 'TÊN CN'")
    missing = [lbl for f, lbl, _ in _COLS if f not in col_map]
    warnings = [f"Thiếu cột trong file, để trống: {', '.join(missing)}"] if missing else []

    # ── Đọc dữ liệu; dòng "Đóng BICCODE" bật cờ cho toàn bộ phần còn lại ──
    # Dừng sau _MAX_BLANK_RUN dòng trống liên tiếp. File thật khai vùng dữ liệu
    # tới 1.048.573 dòng (cả cột được định dạng sẵn) trong khi chỉ có 218 dòng
    # thật — duyệt hết mất 13 giây chỉ để đọc dòng rỗng. Ngưỡng đủ rộng để vượt
    # qua vài dòng cách giữa hai khối dữ liệu, hẹp hơn nhiều so với vùng rỗng đuôi.
    items, is_closed, ma_col = [], False, col_map["ma_cn"]
    blank_run = 0
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            blank_run += 1
            if blank_run >= _MAX_BLANK_RUN:
                break
            continue
        blank_run = 0
        first = _cell_str(row[0])
        if first and _CLOSED_MARKER in first.upper():
            is_closed = True
            continue
        ma_cn = _cell_str(row[ma_col]) if ma_col < len(row) else None
        if not ma_cn:
            continue
        rec = {f: (_cell_str(row[j]) if j < len(row) else None) for f, j in col_map.items()}
        if not rec.get("ten_cn"):
            warnings.append(f"Bỏ qua mã CN {ma_cn}: thiếu TÊN CN")
            continue
        # LOẠI CN là số 1/2; giá trị lạ giữ None thay vì làm hỏng cả lần nhập
        loai = rec.get("loai_cn")
        try:
            rec["loai_cn"] = int(float(loai)) if loai else None
        except ValueError:
            warnings.append(f"Mã CN {ma_cn}: LOẠI CN '{loai}' không phải số, để trống")
            rec["loai_cn"] = None
        if rec.get("swift_bic"):
            rec["swift_bic"] = rec["swift_bic"].upper()
        rec["is_closed"] = is_closed
        items.append(rec)
    wb.close()
    return items, warnings


@router.post("/import", response_model=ImportResult)
async def import_branches(
    file: UploadFile = File(...),
    delete_missing: bool = Query(
        False, description="Xoá các CN có trong DB nhưng không có trong file"
    ),
    current: dict = Depends(require_feature("ttqt_branches.import")),
    db: sqlite3.Connection = Depends(get_db),
):
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Chỉ nhận file Excel .xlsx")
    # openpyxl là Python thuần, giữ GIL suốt — gọi thẳng ở đây sẽ chặn event
    # loop và treo mọi request khác trong lúc đọc file.
    items, warnings = await run_heavy(_parse_workbook,
                                      await read_limited(file, ten="File Excel chi nhánh"))
    if not items:
        raise HTTPException(400, "Không đọc được dòng dữ liệu nào trong file")

    # Trùng mã CN ngay trong file: giữ dòng đầu, báo dòng sau — nếu để cả hai
    # cùng UPSERT thì kết quả phụ thuộc thứ tự dòng, rất khó truy vết về sau.
    seen, uniq, skipped = set(), [], 0
    for rec in items:
        if rec["ma_cn"] in seen:
            warnings.append(f"Trùng mã CN {rec['ma_cn']} trong file — bỏ qua dòng sau")
            skipped += 1
            continue
        seen.add(rec["ma_cn"])
        uniq.append(rec)

    existing = {r["ma_cn"]: r["id"] for r in db.execute("SELECT id, ma_cn FROM ttqt_branches")}
    inserted = updated = 0
    for order, rec in enumerate(uniq, start=1):
        vals = [rec.get(f) for f in _FIELDS]
        if rec["ma_cn"] in existing:
            db.execute(
                f"UPDATE ttqt_branches SET {','.join(f + '=?' for f in _FIELDS)},"
                f" is_closed=?, sort_order=?, updated_at=? WHERE id = ?",
                (*vals, int(rec["is_closed"]), order, _vn_now(), existing[rec["ma_cn"]]),
            )
            updated += 1
        else:
            db.execute(
                f"INSERT INTO ttqt_branches ({','.join(_FIELDS)}, is_closed, sort_order, updated_at)"
                f" VALUES ({','.join('?' * len(_FIELDS))}, ?, ?, ?)",
                (*vals, int(rec["is_closed"]), order, _vn_now()),
            )
            inserted += 1

    deleted = 0
    if delete_missing:
        stale = [mid for ma, mid in existing.items() if ma not in seen]
        for mid in stale:
            db.execute("DELETE FROM ttqt_branches WHERE id = ?", (mid,))
        deleted = len(stale)

    write_audit(
        db, current["id"], "ttqt_branch.import", "ttqt_branch", None,
        f"{file.filename}: +{inserted} mới, ~{updated} cập nhật, -{deleted} xoá",
    )
    db.commit()
    return ImportResult(inserted=inserted, updated=updated, deleted=deleted,
                        skipped=skipped, errors=warnings)


# ─── Export Excel ────────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _build_workbook(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date.today().strftime("%d.%m.%y")
    n_col = len(_COLS) + 1  # +1 cột STT

    ws.cell(row=1, column=1, value="DANH SÁCH CHI NHÁNH THỰC HIỆN TTQT TRỰC TIẾP")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    hdr_fill = PatternFill("solid", fgColor="DBEAFE")
    for j, label in enumerate(["STT"] + [lbl for _, lbl, _ in _COLS], start=1):
        c = ws.cell(row=3, column=j, value=label)
        c.font, c.fill, c.border = Font(bold=True), hdr_fill, _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 6
    for j, (_, _, w) in enumerate(_COLS, start=2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    def _write(r_idx: int, stt: int, rec: dict):
        ws.cell(row=r_idx, column=1, value=stt).border = _BORDER
        for j, f in enumerate(_FIELDS, start=2):
            c = ws.cell(row=r_idx, column=j, value=rec.get(f))
            c.border = _BORDER
            c.alignment = Alignment(vertical="top", wrap_text=f in ("ghi_chu", "dia_chi", "dia_chi_en"))

    r = 4
    for stt, rec in enumerate([x for x in rows if not x["is_closed"]], start=1):
        _write(r, stt, rec)
        r += 1
    closed = [x for x in rows if x["is_closed"]]
    if closed:
        c = ws.cell(row=r, column=1, value="Đóng BICCODE")
        c.font = Font(bold=True, color="B91C1C")
        r += 1
        for stt, rec in enumerate(closed, start=1):
            _write(r, stt, rec)
            r += 1

    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export")
async def export_branches(
    q: Optional[str] = Query(None),
    loai_cn: Optional[int] = Query(None),
    status: str = Query("active", pattern="^(active|closed|all)$"),
    current: dict = Depends(require_feature("ttqt_branches.export")),
    db: sqlite3.Connection = Depends(get_db),
):
    rows = list_branches(q=q, loai_cn=loai_cn, status=status, _=current, db=db)
    content = await run_heavy(_build_workbook, rows)
    write_audit(db, current["id"], "ttqt_branch.export", "ttqt_branch", None,
                f"{len(rows)} chi nhánh (status={status})")
    db.commit()
    fname = f"danh_sach_cn_ttqt_{date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_download_headers(fname),
    )
