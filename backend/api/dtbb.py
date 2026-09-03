"""Đối chiếu số liệu DTBB — Phòng Kế toán. Raw SQL thuần, không ORM."""
import asyncio
import io
import sqlite3
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from backend.core.deps import require_feature
from backend.database import get_db, write_audit, _vn_now
from backend.schemas.dtbb import (
    DtbbCalculateResult, DtbbCurrencyOut, DtbbHistoryItem,
    DtbbReportDetailOut, DtbbSaveRequest, DtbbSaveResponse,
)
from backend.services.dtbb.calculator import DtbbResult, calculate_from_uploads
from backend.services.dtbb.reader import DtbbFileError

router = APIRouter()

_ALLOWED_EXT = ".xls"

# Quyền vào màn hình + mọi thao tác tính/lưu/xoá/xuất: "menu.dtbb". Riêng bước XÁC NHẬN
# (vàng→xanh) và bỏ xác nhận tách "dtbb.confirm" — chốt số liệu là việc của người kiểm
# soát, không phải việc thường ngày. Cả hai gán qua màn Phân quyền theo nhóm; không gate
# theo phòng/chức danh nữa (xem mục "Phân quyền" trong docs/DESIGN.md).
_QUYEN_DUNG    = require_feature("menu.dtbb")
_QUYEN_XAC_NHAN = require_feature("dtbb.confirm")


def _result_to_schema(result: DtbbResult) -> DtbbCalculateResult:
    return DtbbCalculateResult(
        report_date=result.report_date,
        branch_code=result.branch_code,
        file_count=result.file_count,
        vnd_duoi12=result.vnd_duoi12,
        vnd_tu12=result.vnd_tu12,
        usd_duoi12=result.usd_duoi12,
        usd_tu12=result.usd_tu12,
        tk413_usd=result.tk413_usd,
        rate_usd_to_vnd=result.rate_usd_to_vnd,
        all_ccy_codes=result.all_ccy_codes,
        currencies_used=result.currencies_used,
        unconverted_ccy=result.unconverted_ccy,
        netted_9300_ccy=result.netted_9300_ccy,
        details=[
            DtbbCurrencyOut(ccy=d.ccy, rate_to_vnd=d.rate_to_vnd, group1_native=d.group1_native,
                             group2_native=d.group2_native, tk413_native=d.tk413_native)
            for d in result.details
        ],
    )


# ─── Tính toán (chỉ xem trước, KHÔNG ghi DB) ──────────────────────────────────
@router.post("/calculate", response_model=DtbbCalculateResult)
async def calculate_dtbb(
    files: List[UploadFile] = File(...),
    current: dict = Depends(_QUYEN_DUNG),
    db: sqlite3.Connection = Depends(get_db),
):
    if not files:
        raise HTTPException(400, "Chưa chọn file nào")

    contents = []
    for f in files:
        name = f.filename or ""
        if name.startswith("~$"):
            continue  # file khoá tạm của Office, bỏ qua
        if not name.lower().endswith(_ALLOWED_EXT):
            raise HTTPException(400, f"File '{name}' không phải định dạng .XLS")
        data = await f.read()
        contents.append((name, data))

    try:
        # Parse xlrd là CPU-bound đồng bộ — bọc asyncio.to_thread để không chặn event
        # loop backend khi nhiều file/nhiều người dùng tính toán cùng lúc (đúng loại
        # lỗi "quên bọc I/O sync trong async" đã từng làm treo cả server ở PR22).
        result = await asyncio.to_thread(calculate_from_uploads, contents)
    except DtbbFileError as e:
        # detail dạng dict (không chỉ chuỗi) — kèm filenames để FE tô đỏ đúng file lỗi
        # trong danh sách đã chọn, không phải regex lại chuỗi thông báo.
        raise HTTPException(400, {"message": str(e), "filenames": e.filenames})
    return _result_to_schema(result)


# ─── Lưu vào lịch sử ───────────────────────────────────────────────────────────
@router.post("/save", response_model=DtbbSaveResponse)
def save_dtbb(
    body: DtbbSaveRequest,
    current: dict = Depends(_QUYEN_DUNG),
    db: sqlite3.Connection = Depends(get_db),
):
    report_date_str = body.report_date.isoformat()
    existing = db.execute(
        """SELECT id, status, created_by, created_at, updated_by, updated_at
           FROM dtbb_reports WHERE report_date=? AND branch_code=?""",
        (report_date_str, body.branch_code),
    ).fetchone()

    if existing and existing["status"] == "confirmed":
        # Kỳ đã xác nhận (xanh) — chặn hẳn, không đi qua flow needs_confirmation cũ.
        # Phải KSV bỏ xác nhận (POST .../unconfirm) trước mới ghi đè lại được.
        raise HTTPException(
            400,
            f"Kỳ {report_date_str} (chi nhánh {body.branch_code}) đã được xác nhận — "
            "cần người có quyền xác nhận bỏ xác nhận trước khi ghi đè.",
        )

    if existing and not body.confirm_overwrite:
        touched_by_id = existing["updated_by"] or existing["created_by"]
        touched_at = existing["updated_at"] or existing["created_at"]
        toucher = db.execute("SELECT full_name FROM user_tttt WHERE id=?", (touched_by_id,)).fetchone()
        return DtbbSaveResponse(
            report_id=existing["id"], report_date=body.report_date, branch_code=body.branch_code,
            needs_confirmation=True,
            existing_touched_by_name=toucher["full_name"] if toucher else None,
            existing_touched_at=touched_at,
        )

    now = str(_vn_now())
    if existing:
        db.execute(
            """UPDATE dtbb_reports SET vnd_duoi12=?, vnd_tu12=?, usd_duoi12=?, usd_tu12=?,
                   tk413_usd=?, rate_usd_to_vnd=?, file_count=?, status='pending', confirmed_by=NULL,
                   confirmed_at=NULL, updated_by=?, updated_at=? WHERE id=?""",
            (body.vnd_duoi12, body.vnd_tu12, body.usd_duoi12, body.usd_tu12,
             body.tk413_usd, body.rate_usd_to_vnd, body.file_count, current["id"], now, existing["id"]),
        )
        report_id = existing["id"]
        db.execute("DELETE FROM dtbb_report_details WHERE report_id=?", (report_id,))
        action = "dtbb_save_overwrite"
    else:
        try:
            cur = db.execute(
                """INSERT INTO dtbb_reports
                       (report_date, branch_code, vnd_duoi12, vnd_tu12, usd_duoi12, usd_tu12, tk413_usd,
                        rate_usd_to_vnd, file_count, status, created_by, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,'pending',?,?)""",
                (report_date_str, body.branch_code, body.vnd_duoi12, body.vnd_tu12, body.usd_duoi12, body.usd_tu12,
                 body.tk413_usd, body.rate_usd_to_vnd, body.file_count, current["id"], now),
            )
        except sqlite3.IntegrityError:
            # 2 request lưu kỳ mới gần như đồng thời (bấm nhanh 2 lần) — cả 2 đều
            # thấy existing=None ở bước SELECT phía trên, request thứ 2 vi phạm
            # UNIQUE(report_date, branch_code). Dữ liệu không bị nhân đôi (UNIQUE
            # đã chặn), chỉ cần trả thông báo nghiệp vụ thay vì để lộ 500 thô.
            raise HTTPException(
                400,
                f"Kỳ {report_date_str} (chi nhánh {body.branch_code}) vừa được lưu bởi "
                "thao tác khác — tải lại trang và thử lưu lại.",
            )
        report_id = cur.lastrowid
        action = "dtbb_save_new"

    for d in body.details:
        db.execute(
            """INSERT INTO dtbb_report_details
                   (report_id, ccy, rate_to_vnd, group1_native, group2_native, tk413_native)
               VALUES (?,?,?,?,?,?)""",
            (report_id, d.ccy, d.rate_to_vnd, d.group1_native, d.group2_native, d.tk413_native),
        )

    write_audit(
        db, current["id"], action, "dtbb_report", report_id,
        detail=(f"Kỳ {report_date_str} chi nhánh {body.branch_code}: VND<12={body.vnd_duoi12:,.0f} "
                f"(đã gộp TK413-VND), VND≥12={body.vnd_tu12:,.0f}, USD<12={body.usd_duoi12:,.2f}, "
                f"USD≥12={body.usd_tu12:,.2f}, TK413-USD={body.tk413_usd:,.2f}"),
    )
    db.commit()
    return DtbbSaveResponse(report_id=report_id, report_date=body.report_date,
                             branch_code=body.branch_code, overwritten=bool(existing))


# ─── Xác nhận / bỏ xác nhận (vàng ↔ xanh) ──────────────────────────────────────
def _load_report_or_404(report_id: int, db: sqlite3.Connection) -> dict:
    row = db.execute("SELECT * FROM dtbb_reports WHERE id=?", (report_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy kỳ DTBB")
    return dict(row)


@router.post("/{report_id}/confirm", response_model=DtbbSaveResponse)
def confirm_dtbb(
    report_id: int,
    current: dict = Depends(_QUYEN_XAC_NHAN),
    db: sqlite3.Connection = Depends(get_db),
):
    report = _load_report_or_404(report_id, db)
    if report["status"] == "confirmed":
        raise HTTPException(400, "Kỳ này đã được xác nhận rồi")
    toucher_id = report["updated_by"] or report["created_by"]
    if current["id"] == toucher_id:
        raise HTTPException(403, "Không được tự xác nhận kỳ do chính mình tạo/sửa gần nhất")

    now = str(_vn_now())
    db.execute(
        "UPDATE dtbb_reports SET status='confirmed', confirmed_by=?, confirmed_at=? WHERE id=?",
        (current["id"], now, report_id),
    )
    write_audit(
        db, current["id"], "dtbb_confirm", "dtbb_report", report_id,
        detail=f"Xác nhận kỳ {report['report_date']} chi nhánh {report['branch_code']}",
    )
    db.commit()
    return DtbbSaveResponse(report_id=report_id, report_date=report["report_date"],
                             branch_code=report["branch_code"])


@router.post("/{report_id}/unconfirm", response_model=DtbbSaveResponse)
def unconfirm_dtbb(
    report_id: int,
    current: dict = Depends(_QUYEN_XAC_NHAN),
    db: sqlite3.Connection = Depends(get_db),
):
    report = _load_report_or_404(report_id, db)
    if report["status"] != "confirmed":
        raise HTTPException(400, "Kỳ này chưa được xác nhận")

    db.execute(
        "UPDATE dtbb_reports SET status='pending', confirmed_by=NULL, confirmed_at=NULL WHERE id=?",
        (report_id,),
    )
    write_audit(
        db, current["id"], "dtbb_unconfirm", "dtbb_report", report_id,
        detail=f"Bỏ xác nhận kỳ {report['report_date']} chi nhánh {report['branch_code']}",
    )
    db.commit()
    return DtbbSaveResponse(report_id=report_id, report_date=report["report_date"],
                             branch_code=report["branch_code"])


@router.delete("/{report_id}")
def delete_dtbb(
    report_id: int,
    current: dict = Depends(_QUYEN_DUNG),
    db: sqlite3.Connection = Depends(get_db),
):
    """Xoá kỳ DTBB đã lưu. Đi cùng quyền tính/lưu ("menu.dtbb"), KHÔNG đòi thêm
    "dtbb.confirm" — chỉ chặn khi kỳ đang ở trạng thái xanh (đã xác nhận), lúc đó phải
    bỏ xác nhận trước (đưa về vàng) rồi mới xoá được."""
    report = _load_report_or_404(report_id, db)
    if report["status"] == "confirmed":
        raise HTTPException(
            400, "Kỳ đã xác nhận (xanh) — cần người có quyền xác nhận bỏ xác nhận trước khi xoá"
        )
    db.execute("DELETE FROM dtbb_reports WHERE id=?", (report_id,))
    write_audit(
        db, current["id"], "dtbb_delete", "dtbb_report", report_id,
        detail=f"Xoá kỳ {report['report_date']} chi nhánh {report['branch_code']}",
    )
    db.commit()
    return {"ok": True}


# ─── Lịch sử ───────────────────────────────────────────────────────────────────
@router.get("/history", response_model=List[DtbbHistoryItem])
def list_history(
    current: dict = Depends(_QUYEN_DUNG),
    db: sqlite3.Connection = Depends(get_db),
):
    rows = db.execute(
        """SELECT r.*, cu.full_name AS created_by_name, uu.full_name AS updated_by_name,
                  cf.full_name AS confirmed_by_name
           FROM dtbb_reports r
           JOIN user_tttt cu ON cu.id = r.created_by
           LEFT JOIN user_tttt uu ON uu.id = r.updated_by
           LEFT JOIN user_tttt cf ON cf.id = r.confirmed_by
           ORDER BY r.report_date DESC, r.branch_code"""
    ).fetchall()
    return [dict(r) for r in rows]


def _load_report_detail(report_date: str, branch_code: str, db: sqlite3.Connection) -> dict:
    row = db.execute(
        """SELECT r.*, cu.full_name AS created_by_name, uu.full_name AS updated_by_name,
                  cf.full_name AS confirmed_by_name
           FROM dtbb_reports r
           JOIN user_tttt cu ON cu.id = r.created_by
           LEFT JOIN user_tttt uu ON uu.id = r.updated_by
           LEFT JOIN user_tttt cf ON cf.id = r.confirmed_by
           WHERE r.report_date = ? AND r.branch_code = ?""",
        (report_date, branch_code),
    ).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy kỳ DTBB đã lưu")
    details = db.execute(
        """SELECT ccy, rate_to_vnd, group1_native, group2_native, tk413_native
           FROM dtbb_report_details WHERE report_id=? ORDER BY ccy""",
        (row["id"],),
    ).fetchall()
    d = dict(row)
    d["details"] = [dict(x) for x in details]
    return d


@router.get("/history/{report_date}/{branch_code}", response_model=DtbbReportDetailOut)
def get_history_detail(
    report_date: str,
    branch_code: str,
    current: dict = Depends(_QUYEN_DUNG),
    db: sqlite3.Connection = Depends(get_db),
):
    return _load_report_detail(report_date, branch_code, db)


# ─── Xuất Excel ────────────────────────────────────────────────────────────────
@router.get("/export/{report_date}/{branch_code}")
def export_dtbb(
    report_date: str,
    branch_code: str,
    current: dict = Depends(_QUYEN_DUNG),
    db: sqlite3.Connection = Depends(get_db),
):
    data = _load_report_detail(report_date, branch_code, db)

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tong hop DTBB"
    FONT_NAME = "Times New Roman"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="C62828")
    hdr_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFFFF")

    y, m, d = data["report_date"].split("-")
    report_date_vn = f"{d}/{m}/{y}"
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Cơ sở tính DTBB ngày {report_date_vn} — chi nhánh {data['branch_code']} (từ {data['file_count']} file)"
    c.font = Font(name=FONT_NAME, size=13, bold=True)
    c.alignment = Alignment(horizontal="center")

    # DB lưu số đầy đủ (VND/USD thật) — chỉ quy đổi Triệu VND / Ngàn USD lúc xuất
    # Excel để dễ đọc, không đổi giá trị lưu trữ. TK413-VND đã gộp vào cột VND
    # dưới 12 tháng (không tách riêng); TK413-USD vẫn giữ cột riêng.
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "Đơn vị: cột VND — Triệu VND; cột USD — Ngàn USD"
    c.font = Font(name=FONT_NAME, size=10, italic=True)
    c.alignment = Alignment(horizontal="center")

    headers = ["VND dưới 12 tháng (triệu)", "VND từ 12 tháng (triệu)", "USD dưới 12 tháng (ngàn)",
               "USD từ 12 tháng (ngàn)", "TK413-USD (ngàn)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = 20

    divisors = [1_000_000, 1_000_000, 1_000, 1_000, 1_000]
    values = [data["vnd_duoi12"], data["vnd_tu12"], data["usd_duoi12"],
              data["usd_tu12"], data["tk413_usd"]]
    for i, (v, div) in enumerate(zip(values, divisors), start=1):
        cell = ws.cell(row=4, column=i, value=round(v / div, 2))
        cell.font = Font(name=FONT_NAME, size=11)
        cell.border = border
        cell.number_format = "#,##0.00"

    ws2 = wb.create_sheet("Chi tiet theo loai tien")
    d_headers = ["Mã tiền", "Tỷ giá → VND", "Dưới 12 tháng (nguyên tệ)",
                 "Từ 12 tháng (nguyên tệ)", "TK413 (nguyên tệ)"]
    for i, h in enumerate(d_headers, start=1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
        ws2.column_dimensions[get_column_letter(i)].width = 22
    for r, d in enumerate(data["details"], start=2):
        ws2.cell(row=r, column=1, value=d["ccy"]).border = border
        ws2.cell(row=r, column=2, value=d["rate_to_vnd"]).border = border
        ws2.cell(row=r, column=3, value=round(d["group1_native"], 2)).border = border
        ws2.cell(row=r, column=4, value=round(d["group2_native"], 2)).border = border
        ws2.cell(row=r, column=5, value=round(d["tk413_native"], 2)).border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''dtbb_{report_date}_{branch_code}.xlsx"},
    )
