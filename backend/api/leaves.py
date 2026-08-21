"""Quản lý nghỉ phép — đăng ký, phê duyệt, tải phiếu"""
import base64
import io
import json
import os
import re
import sqlite3
import threading
import unicodedata
from datetime import date, timedelta
from typing import FrozenSet, List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from backend.core.concurrency import run_heavy
from backend.core.uploads import read_limited_sync
from backend.core.deps import TONG_HOP_CODES, get_current_staff, require_feature
from backend.core.enums import LeaveStatus
from backend.core.paths import template_path
from backend.database import get_db, write_audit, _vn_now, compute_annual_leave, compute_carry_over
from backend.schemas.leaves import (
    LeaveCreate, LeaveReview, TongHopReview,
    DirectLeaveCreate, RecallCreate, LeaveQuotaUpsert,
)
from backend.services import leave_pdf

router = APIRouter()

LEAVE_TYPE_LABELS = {
    "bat_buoc":  "Nghỉ phép bắt buộc",
    "annual":    "Nghỉ phép năm",
    "thai_san":  "Nghỉ thai sản",
    "bao_hiem":  "Nghỉ bảo hiểm",
    "other":     "Khác",
    # Legacy types (giữ để tương thích với data cũ)
    "sick":      "Nghỉ ốm",
    "personal":  "Nghỉ việc riêng",
    "dot_xuat":  "Nghỉ đột xuất",
}

_VALID_LEAVE_TYPES = frozenset(LEAVE_TYPE_LABELS.keys())
# Các loại nghỉ không tính vào/trừ hạn mức phép năm
_NO_QUOTA_TYPES = frozenset({"thai_san", "bao_hiem"})

ACTION_LABELS = {
    "create":         ("Nộp đơn",            "blue"),
    "ksv_approve":    ("KSV phê duyệt",      "green"),
    "ksv_reject":     ("KSV từ chối",        "red"),
    "th_forward":     ("TH chuyển GĐ",       "blue"),
    "th_reject":      ("TH từ chối",         "red"),
    "gd_approve":     ("GĐ phê duyệt",       "green"),
    "gd_reject":      ("GĐ từ chối",         "red"),
    "resubmit":       ("Nộp lại",            "orange"),
    "cancel":         ("Hủy đơn",            "grey"),
    "direct_create":  ("Khai báo hộ",        "purple"),
    "recall_request": ("Yêu cầu rút đơn",    "orange"),
    "recall_approve": ("Xác nhận rút đơn",   "grey"),
    "th_ack_gd":      ("TH xác nhận đã biết", "green"),
}

_LEAVE_STATUS_VN = {
    LeaveStatus.PENDING_KSV:      "Chờ KSV duyệt",
    LeaveStatus.PENDING_TONG_HOP: "Chờ Tổng hợp",
    LeaveStatus.PENDING_GD:       "Chờ Ban lãnh đạo duyệt",
    LeaveStatus.APPROVED:         "Hoàn thành",
    LeaveStatus.REJECTED:         "Bị từ chối",
    LeaveStatus.CANCELLED:        "Đã hủy",
}

# Các role cấp cao — bỏ qua bước KSV, vào thẳng pending_tong_hop
_HIGH_ROLES = frozenset(("giam_doc", "pho_giam_doc", "admin", "truong_phong"))


# ─── Helpers ────────────────────────────────────────────────────────────────

def _calc_used_days(staff_id: int, year: int, db: sqlite3.Connection,
                    exclude_id: int | None = None,
                    include_pending: bool = False) -> float:
    """Số ngày đã dùng trong năm — nguồn sự thật duy nhất.

    Đếm tất cả loại TRỪ thai_san/bao_hiem.
    include_pending=True: cộng thêm đơn đang chờ duyệt (dùng khi kiểm tra quota lúc nộp lại).
    exclude_id: bỏ qua đơn đang xem (dùng khi in phiếu).
    """
    if include_pending:
        statuses = ("'approved','pending_ksv','pending_tong_hop','pending_gd'")
    else:
        statuses = "'approved'"
    excl = "AND id != ?" if exclude_id is not None else ""
    params: list = [staff_id]
    if exclude_id is not None:
        params.append(exclude_id)
    # Lấy theo overlap với năm (không chỉ start_date cùng năm) — đơn vắt qua
    # ranh giới năm (vd 29/12 → 02/01) vẫn phải được xét để đếm đúng phần ngày
    # rơi vào năm đang tính, xem thêm vòng lặp clip theo d.year bên dưới.
    params += [f"{year}-12-31", f"{year}-01-01"]
    rows = db.execute(
        f"""SELECT spread_dates, start_date, end_date FROM leave_records
            WHERE staff_id=? {excl} AND status IN ({statuses})
              AND leave_type NOT IN ('thai_san','bao_hiem')
              AND start_date <= ? AND end_date >= ?""",
        params,
    ).fetchall()
    total = 0.0
    _holidays: frozenset | None = None  # lazy load khi cần
    for row in rows:
        if row["spread_dates"]:
            total += len([d for d in json.loads(row["spread_dates"]) if d.startswith(str(year))])
        else:
            if _holidays is None:
                hrows = db.execute(
                    "SELECT date FROM public_holidays WHERE date >= ? AND date <= ?",
                    (f"{year}-01-01", f"{year}-12-31"),
                ).fetchall()
                _holidays = frozenset(date.fromisoformat(r["date"]) for r in hrows)
            d = date.fromisoformat(row["start_date"])
            e = date.fromisoformat(row["end_date"])
            while d <= e:
                if d.year == year and d.weekday() < 5 and d not in _holidays:
                    total += 1
                d += timedelta(days=1)
    return total


def _calc_used_days_bulk(staff_ids: list, year: int, db: sqlite3.Connection,
                         include_pending: bool = False) -> dict:
    """Bản gộp truy vấn của _calc_used_days cho NHIỀU nhân viên cùng lúc (1 query
    thay vì N) — dùng cho các màn hình liệt kê toàn bộ nhân viên (get_quotas,
    export_quotas, stats_annual) để tránh N+1. Phải giữ đúng logic tính giống hệt
    _calc_used_days (đơn lẻ)."""
    if not staff_ids:
        return {}
    statuses = ("'approved','pending_ksv','pending_tong_hop','pending_gd'"
                if include_pending else "'approved'")
    placeholders = ",".join("?" * len(staff_ids))
    rows = db.execute(
        f"""SELECT staff_id, spread_dates, start_date, end_date FROM leave_records
            WHERE staff_id IN ({placeholders}) AND status IN ({statuses})
              AND leave_type NOT IN ('thai_san','bao_hiem')
              AND start_date <= ? AND end_date >= ?""",
        list(staff_ids) + [f"{year}-12-31", f"{year}-01-01"],
    ).fetchall()
    result = {sid: 0.0 for sid in staff_ids}
    _holidays: frozenset | None = None
    for row in rows:
        if row["spread_dates"]:
            result[row["staff_id"]] += len([d for d in json.loads(row["spread_dates"]) if d.startswith(str(year))])
        else:
            if _holidays is None:
                hrows = db.execute(
                    "SELECT date FROM public_holidays WHERE date >= ? AND date <= ?",
                    (f"{year}-01-01", f"{year}-12-31"),
                ).fetchall()
                _holidays = frozenset(date.fromisoformat(r["date"]) for r in hrows)
            d = date.fromisoformat(row["start_date"])
            e = date.fromisoformat(row["end_date"])
            while d <= e:
                if d.year == year and d.weekday() < 5 and d not in _holidays:
                    result[row["staff_id"]] += 1
                d += timedelta(days=1)
    return result


def _carry_over_bulk(staff_ids: list, year: int, db: sqlite3.Connection,
                     effective: bool = True, ref_date=None) -> dict:
    """Bản gộp truy vấn của compute_carry_over (backend/database.py) cho NHIỀU
    nhân viên cùng lúc — tránh N+1. Phải giữ đúng logic tính giống hệt bản đơn lẻ."""
    if not staff_ids:
        return {}
    if effective:
        check_date = ref_date if ref_date else date.today()
        if check_date > date(year, 3, 31):
            return {sid: 0.0 for sid in staff_ids}
    prev_year = year - 1
    placeholders = ",".join("?" * len(staff_ids))
    quota_by_staff = {
        r["staff_id"]: float(r["quota_days"])
        for r in db.execute(
            f"SELECT staff_id, quota_days FROM leave_quotas WHERE year=? AND staff_id IN ({placeholders})",
            [prev_year] + list(staff_ids),
        ).fetchall()
    }
    join_by_staff = {
        r["id"]: r["join_industry_date"]
        for r in db.execute(
            f"SELECT id, join_industry_date FROM user_tttt WHERE id IN ({placeholders})",
            list(staff_ids),
        ).fetchall()
    }
    used_by_staff: dict = {}
    _holidays = None
    for r in db.execute(
        f"""SELECT staff_id, start_date, end_date, spread_dates FROM leave_records
           WHERE staff_id IN ({placeholders}) AND status='approved'
             AND leave_type NOT IN ('thai_san','bao_hiem')
             AND start_date <= ? AND end_date >= ?""",
        list(staff_ids) + [f"{prev_year}-12-31", f"{prev_year}-01-01"],
    ).fetchall():
        sid = r["staff_id"]
        if r["spread_dates"]:
            used_by_staff[sid] = used_by_staff.get(sid, 0.0) + len(
                [d for d in json.loads(r["spread_dates"]) if d.startswith(str(prev_year))]
            )
        else:
            if _holidays is None:
                hrows = db.execute(
                    "SELECT date FROM public_holidays WHERE date >= ? AND date <= ?",
                    (f"{prev_year}-01-01", f"{prev_year}-12-31"),
                ).fetchall()
                _holidays = frozenset(date.fromisoformat(hr["date"]) for hr in hrows)
            d = date.fromisoformat(r["start_date"])
            e = date.fromisoformat(r["end_date"])
            while d <= e:
                if d.year == prev_year and d.weekday() < 5 and d not in _holidays:
                    used_by_staff[sid] = used_by_staff.get(sid, 0.0) + 1
                d += timedelta(days=1)
    result = {}
    for sid in staff_ids:
        prev_quota = quota_by_staff.get(sid)
        if prev_quota is None:
            prev_quota = float(compute_annual_leave(join_by_staff.get(sid), prev_year))
        result[sid] = max(0.0, prev_quota - used_by_staff.get(sid, 0.0))
    return result


def _load_holidays(db: sqlite3.Connection, start: date, end: date) -> FrozenSet[date]:
    rows = db.execute(
        "SELECT date FROM public_holidays WHERE date >= ? AND date <= ?",
        (start.isoformat(), end.isoformat()),
    ).fetchall()
    return frozenset(date.fromisoformat(r["date"]) for r in rows)


def calculate_leave_days(
    start: date, end: date,
    holiday_dates: FrozenSet[date] = frozenset(),
) -> int:
    """Đếm ngày làm việc thực (trừ T7, CN và ngày lễ). Tối thiểu 1 ngày."""
    count = 0
    d = start
    while d <= end:
        if d.weekday() < 5 and d not in holiday_dates:
            count += 1
        d += timedelta(days=1)
    return max(count, 1)


def _period_days(
    start: date, end: date,
    holiday_dates: FrozenSet[date] = frozenset(),
    leave_type: Optional[str] = None,
) -> int:
    """Số ngày của khoảng nghỉ liên tục (khi không dùng spread_dates).

    thai_san/bao_hiem: tính theo ngày lịch liên tục (kể cả T7, CN, lễ).
    Các loại khác: chỉ tính ngày làm việc (calculate_leave_days).
    """
    if leave_type in _NO_QUOTA_TYPES:
        return (end - start).days + 1
    return calculate_leave_days(start, end, holiday_dates)


def _norm_vn(s) -> str:
    """Chuẩn hoá text tiếng Việt để so khớp: bỏ dấu, chữ thường, gọn khoảng trắng.

    Đ/đ không tự tách dấu qua NFD (không giống ă/â/ê...) nên phải thay tay.
    """
    if not s:
        return ""
    s = str(s).replace("Đ", "D").replace("đ", "d")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def _is_tong_hop_staff(staff: dict, db: sqlite3.Connection) -> bool:
    dept_id = staff.get("department_id")
    if not dept_id:
        return False
    r = db.execute("SELECT code FROM departments WHERE id = ?", (dept_id,)).fetchone()
    return bool(r and r["code"].upper() in TONG_HOP_CODES)


def _can_gd_review(current: dict, db: sqlite3.Connection) -> bool:
    if current["role"] == "giam_doc":
        return True
    if current["role"] == "pho_giam_doc":
        today = _vn_now().date().isoformat()
        return db.execute(
            "SELECT id FROM delegation_records WHERE pho_giam_doc_id = ? AND is_active = 1 AND start_date <= ? AND end_date >= ?",
            (current["id"], today, today),
        ).fetchone() is not None
    return False


def _apply_status_transition(
    leave_id: int, old_status: str, new_status: str,
    start: date, end: date, staff_id: int,
    holiday_dates: FrozenSet[date],
    db: sqlite3.Connection,
    days: Optional[int] = None,
    is_new: bool = False,
):
    """Cập nhật status và điều chỉnh used_leave_days (idempotent).

    is_new=True: dùng khi tạo đơn đã approved ngay (GĐ tự duyệt) — old_status
    truyền vào phải là status THẬT sự đang có trong DB (để khoá lạc quan khớp),
    còn is_new mới là cờ báo "coi như mới approved" để cộng used_leave_days dù
    old_status == new_status.
    """
    if days is None:
        rec = db.execute("SELECT spread_dates, leave_type FROM leave_records WHERE id=?", (leave_id,)).fetchone()
        if rec and rec["spread_dates"]:
            days = len(json.loads(rec["spread_dates"]))
        else:
            days = calculate_leave_days(start, end, holiday_dates)
        leave_type = rec["leave_type"] if rec else None
    else:
        rec2 = db.execute("SELECT leave_type FROM leave_records WHERE id=?", (leave_id,)).fetchone()
        leave_type = rec2["leave_type"] if rec2 else None

    # Khoá lạc quan: chỉ chuyển trạng thái nếu status hiện tại trong DB đúng
    # bằng old_status — chặn 2 request duyệt trùng cùng lúc (double-click, 2
    # tab) cộng/trừ used_leave_days 2 lần cho cùng 1 đơn.
    cur = db.execute(
        "UPDATE leave_records SET status = ?, updated_at = ? WHERE id = ? AND status = ?",
        (new_status, str(_vn_now()), leave_id, old_status),
    )
    if cur.rowcount == 0:
        raise HTTPException(409, "Đơn đã được xử lý bởi một yêu cầu khác, vui lòng tải lại trang")

    # Chỉ điều chỉnh used_leave_days khi leave_type xác định và không miễn quota
    if leave_type is not None and leave_type not in _NO_QUOTA_TYPES:
        if (is_new or old_status != LeaveStatus.APPROVED) and new_status == LeaveStatus.APPROVED:
            db.execute(
                "UPDATE user_tttt SET used_leave_days = COALESCE(used_leave_days, 0) + ? WHERE id = ?",
                (days, staff_id),
            )
        elif old_status == LeaveStatus.APPROVED and new_status in (LeaveStatus.CANCELLED, LeaveStatus.REJECTED):
            db.execute(
                "UPDATE user_tttt SET used_leave_days = MAX(0, COALESCE(used_leave_days, 0) - ?) WHERE id = ?",
                (days, staff_id),
            )


def _log_action(
    db: sqlite3.Connection, leave_id: int, actor_id: int, action: str,
    comment: Optional[str], from_status: str, to_status: str,
):
    db.execute(
        "INSERT INTO leave_action_logs (leave_id, actor_id, action, comment, from_status, to_status, created_at) VALUES (?,?,?,?,?,?,?)",
        (leave_id, actor_id, action, comment, from_status, to_status, str(_vn_now())),
    )


def _validate_ksv(ksv_id: Optional[int], current: dict, db: sqlite3.Connection) -> dict:
    if not ksv_id:
        raise HTTPException(400, "Vui lòng chọn người phê duyệt bước KSV")
    ksv = db.execute("SELECT * FROM user_tttt WHERE id = ? AND is_active = 1", (ksv_id,)).fetchone()
    if not ksv:
        raise HTTPException(400, "Người phê duyệt không tồn tại hoặc đã bị vô hiệu")
    if ksv["role"] not in ("truong_phong", "pho_phong", "admin"):
        raise HTTPException(400, "Người phê duyệt phải là Trưởng phòng hoặc Phó phòng")
    if ksv_id == current["id"]:
        raise HTTPException(400, "Không thể tự phê duyệt")
    # KSV phải cùng phòng — khớp đúng get_approvers() (đã sửa: Hậu kiểm viên
    # cũng lọc theo phòng, không còn cross-department nữa).
    if ksv["role"] != "admin":
        staff_dept = current.get("department_id")
        ksv_dept   = ksv["department_id"]
        if staff_dept and ksv_dept and staff_dept != ksv_dept:
            raise HTTPException(400, "Người phê duyệt phải thuộc cùng phòng ban với người nộp đơn")
    return dict(ksv)


def _leave_to_out(leave_id: int, db: sqlite3.Connection) -> dict:
    today = _vn_now().date().isoformat()
    r = db.execute(
        """SELECT lr.*,
                  s.full_name AS staff_name, s.department_id AS s_dept_id, s.role AS staff_role,
                  kv.full_name AS ksv_name,
                  th.full_name AS th_name,
                  gd.full_name AS gd_approver_name, gd.role AS gd_role,
                  d.name AS dept_name,
                  db_user.full_name AS declarer_name,
                  -- PGĐ chỉ duyệt được khi còn ủy quyền hiệu lực HÔM NAY. Tính sẵn ở
                  -- đây (không phải chỉ trong gd_review) để màn hình nói được lý do
                  -- vì sao người duyệt không có nút, thay vì im lặng.
                  CASE WHEN gd.role = 'pho_giam_doc' THEN
                       (SELECT COUNT(*) FROM delegation_records dr
                         WHERE dr.pho_giam_doc_id = gd.id AND dr.is_active = 1
                           AND dr.start_date <= ? AND dr.end_date >= ?)
                       ELSE 1 END AS gd_can_review
           FROM leave_records lr
           LEFT JOIN user_tttt s       ON lr.staff_id             = s.id
           LEFT JOIN user_tttt kv      ON lr.ksv_approver_id      = kv.id
           LEFT JOIN user_tttt th      ON lr.tong_hop_approver_id = th.id
           LEFT JOIN user_tttt gd      ON lr.gd_approver_id       = gd.id
           LEFT JOIN departments d     ON s.department_id          = d.id
           LEFT JOIN user_tttt db_user ON lr.direct_by             = db_user.id
           WHERE lr.id = ?""",
        (today, today, leave_id),
    ).fetchone()
    if not r:
        return {}

    start = date.fromisoformat(r["start_date"])
    end   = date.fromisoformat(r["end_date"])
    _h    = _load_holidays(db, start, end)
    _days = len(json.loads(r["spread_dates"])) if r["spread_dates"] else _period_days(start, end, _h, r["leave_type"])

    return {
        "id":                     r["id"],
        "staff_id":               r["staff_id"],
        "staff_name":             r["staff_name"] or "",
        "staff_role":             r["staff_role"],
        "department_name":        r["dept_name"],
        "start_date":             r["start_date"],
        "end_date":               r["end_date"],
        "leave_days":             _days,
        "leave_type":             r["leave_type"],
        "reason":                 r["reason"],
        "status":                 r["status"],
        "ksv_approver_id":        r["ksv_approver_id"],
        "ksv_approver_name":      r["ksv_name"],
        "ksv_approved_at":        r["ksv_approved_at"],
        "ksv_comment":            r["ksv_comment"],
        "tong_hop_approver_id":   r["tong_hop_approver_id"],
        "tong_hop_approver_name": r["th_name"],
        "tong_hop_approved_at":   r["tong_hop_approved_at"],
        "tong_hop_comment":       r["tong_hop_comment"],
        "gd_approver_id":         r["gd_approver_id"],
        "gd_approver_name":       r["gd_approver_name"],
        "gd_is_pgd":              (r["gd_role"] == "pho_giam_doc") if r["gd_role"] else False,
        "gd_can_review":          bool(r["gd_can_review"]),
        "gd_approved_at":         r["gd_approved_at"],
        "gd_comment":             r["gd_comment"],
        "is_direct":              bool(r["is_direct"]) if r["is_direct"] is not None else False,
        "declarer_name":          r["declarer_name"] or "",
        "spread_dates":           json.loads(r["spread_dates"]) if r["spread_dates"] else None,
        "recall_reason":          r["recall_reason"],
        "created_at":             r["created_at"],
        "rejected_step":          (
            "GĐ"  if r["status"] == "rejected" and r["gd_approved_at"]
            else "TH"  if r["status"] == "rejected" and r["tong_hop_approved_at"]
            else "KSV" if r["status"] == "rejected"
            else None
        ),
        "is_resubmitted": bool(db.execute(
            "SELECT 1 FROM leave_action_logs WHERE leave_id=? AND action='resubmit' LIMIT 1",
            (leave_id,)
        ).fetchone()),
    }


# ─── Endpoints ──────────────────────────────────────────────────────────────

@router.get("/approvers")
def get_approvers(
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    _ROLE_LABEL = {
        "truong_phong":  "Trưởng phòng",
        "pho_phong":     "Phó phòng",
        "admin":         "Quản trị viên cấp 1",
    }
    dept_id = current.get("department_id")
    if dept_id:
        # Chỉ lấy KSV cùng phòng (Trưởng/Phó phòng) với người tạo đơn. Hậu kiểm viên
        # KHÔNG duyệt nghỉ phép — ngang chuyên viên ở quy trình này.
        rows = db.execute(
            """SELECT id, full_name, role FROM user_tttt
               WHERE is_active = 1 AND id != ?
                 AND role IN ('truong_phong','pho_phong')
                 AND department_id = ?
               ORDER BY full_name""",
            (current["id"], dept_id),
        ).fetchall()
    else:
        rows = db.execute(
            """SELECT id, full_name, role FROM user_tttt
               WHERE is_active = 1 AND role IN ('truong_phong','pho_phong','admin')
                 AND id != ? ORDER BY full_name""",
            (current["id"],),
        ).fetchall()
    return [{"id": r["id"], "full_name": r["full_name"], "role_label": _ROLE_LABEL.get(r["role"], r["role"])} for r in rows]


@router.get("/gd-list")
def get_gd_list(
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    rows = db.execute(
        "SELECT id, full_name, role FROM user_tttt WHERE is_active = 1 AND role IN ('giam_doc','pho_giam_doc')"
    ).fetchall()
    _ROLE_LABEL = {"giam_doc": "Giám đốc", "pho_giam_doc": "Phó Giám đốc"}
    return [{"id": r["id"], "full_name": r["full_name"], "role_label": _ROLE_LABEL.get(r["role"], r["role"])} for r in rows]


@router.post("/")
def create_leave(
    body: LeaveCreate,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.create")),
):
    if current.get("role") == "admin":
        raise HTTPException(403, "Admin không tham gia quy trình nghỉ phép")
    if body.leave_type not in _VALID_LEAVE_TYPES:
        raise HTTPException(400, f"Loại nghỉ phép không hợp lệ: {body.leave_type}")

    # ── Xử lý spread_dates (nghỉ ngày lẻ không liên tục) ──
    if body.spread_dates:
        spread = sorted(set(body.spread_dates))
        if len(spread) < 1:
            raise HTTPException(400, "spread_dates phải có ít nhất 1 ngày")
        try:
            eff_start = date.fromisoformat(spread[0])
            eff_end   = date.fromisoformat(spread[-1])
        except ValueError:
            raise HTTPException(400, "Định dạng ngày không hợp lệ (yêu cầu YYYY-MM-DD)")
        leave_days = len(spread)
        spread_json = json.dumps(spread)
        _h = _load_holidays(db, eff_start, eff_end)
    else:
        if body.end_date < body.start_date:
            raise HTTPException(400, "Ngày kết thúc phải sau ngày bắt đầu")
        eff_start   = body.start_date
        eff_end     = body.end_date
        _h          = _load_holidays(db, eff_start, eff_end)
        leave_days  = _period_days(eff_start, eff_end, _h, body.leave_type)
        spread_json = None

    if body.leave_type == "annual":
        if eff_start < _vn_now().date():
            raise HTTPException(400, "Nghỉ phép năm phải từ hôm nay trở đi")

    # Kiểm tra hạn mức (không áp dụng cho bat_buoc, thai_san, bao_hiem)
    if body.leave_type not in _NO_QUOTA_TYPES and body.leave_type != "bat_buoc":
        ref_year  = eff_start.year
        carry_eff = compute_carry_over(current["id"], ref_year, db,
                                       effective=True, ref_date=eff_start)
        # Ưu tiên hạn mức nhập tay (leave_quotas), khớp đúng get_quotas/download_leave_form
        # — trước đây dùng thẳng current["annual_leave_days"] (công thức thô, sai năm),
        # bỏ qua override của admin.
        _q_row = db.execute(
            "SELECT quota_days FROM leave_quotas WHERE staff_id=? AND year=?",
            (current["id"], ref_year),
        ).fetchone()
        quota = (
            float(_q_row["quota_days"]) if _q_row
            else float(compute_annual_leave(current.get("join_industry_date"), ref_year))
        )
        used_total = _calc_used_days(current["id"], ref_year, db, include_pending=True)
        remaining  = quota + carry_eff - used_total
        if leave_days > remaining:
            raise HTTPException(400, f"Vượt quá số ngày phép còn lại ({remaining:.0f} ngày)")

    if body.leave_type == "bat_buoc" and leave_days < 5:
        raise HTTPException(400, "Nghỉ phép bắt buộc phải từ 5 ngày làm việc trở lên")

    # Kiểm tra trùng ngày theo spread_dates thực tế (không dùng envelope khi có spread)
    if body.spread_dates:
        _existing = db.execute(
            """SELECT start_date, end_date, spread_dates FROM leave_records
               WHERE staff_id=? AND status NOT IN ('rejected','cancelled')
                 AND NOT (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')""",
            (current["id"],)
        ).fetchall()
        _new_days = set(spread)
        for _el in _existing:
            if _el["spread_dates"]:
                if _new_days & set(json.loads(_el["spread_dates"])):
                    raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")
            elif any(_el["start_date"] <= d <= _el["end_date"] for d in spread):
                raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")
    else:
        # Đơn mới là khoảng liên tục — với đơn hiện có CÓ spread_dates, phải so
        # với các ngày thực tế (không phải envelope start/end) để tránh báo
        # trùng oan khi khoảng mới chỉ chồng lên envelope chứ không chạm ngày
        # thực nào.
        _existing2 = db.execute(
            """SELECT start_date, end_date, spread_dates FROM leave_records
               WHERE staff_id=? AND status NOT IN ('rejected','cancelled')
                 AND NOT (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')
                 AND start_date<=? AND end_date>=?""",
            (current["id"], eff_end.isoformat(), eff_start.isoformat())
        ).fetchall()
        if _existing2:
            _new_range_days = None
            for _el in _existing2:
                if _el["spread_dates"]:
                    if _new_range_days is None:
                        _new_range_days, _d = set(), eff_start
                        while _d <= eff_end:
                            _new_range_days.add(_d.isoformat())
                            _d += timedelta(days=1)
                    if set(json.loads(_el["spread_dates"])) & _new_range_days:
                        raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")
                else:
                    raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")

    if current["role"] == "giam_doc":
        # GĐ là cấp cao nhất — tự duyệt ngay, không qua quy trình
        initial_status  = LeaveStatus.APPROVED
        ksv_approver_id = None
    elif current["role"] in _HIGH_ROLES:
        initial_status  = LeaveStatus.PENDING_TONG_HOP
        ksv_approver_id = None
    else:
        ksv = _validate_ksv(body.ksv_approver_id, current, db)
        initial_status  = LeaveStatus.PENDING_KSV
        ksv_approver_id = ksv["id"]

    # Nếu user chọn trước Ban lãnh đạo, validate
    gd_approver_id = None
    gd_approved_at = None
    if current["role"] == "giam_doc":
        # GĐ tự duyệt đơn của chính mình — ghi nhận GĐ là người duyệt
        gd_approver_id = current["id"]
        gd_approved_at = str(_vn_now())
    elif body.gd_approver_id:
        gd_staff = db.execute(
            "SELECT id, role FROM user_tttt WHERE id=? AND is_active=1", (body.gd_approver_id,)
        ).fetchone()
        if gd_staff and gd_staff["role"] in ("giam_doc", "pho_giam_doc"):
            gd_approver_id = gd_staff["id"]

    cur = db.execute(
        """INSERT INTO leave_records
               (staff_id, start_date, end_date, leave_type, reason, status,
                ksv_approver_id, gd_approver_id, gd_approved_at, spread_dates, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (current["id"], eff_start.isoformat(), eff_end.isoformat(),
         body.leave_type, body.reason, initial_status, ksv_approver_id,
         gd_approver_id, gd_approved_at, spread_json, str(_vn_now()), str(_vn_now())),
    )
    leave_id = cur.lastrowid
    if body.signature:
        _save_signature(db, leave_id, "nguoi_de_nghi", current["id"], body.signature)
    _log_action(db, leave_id, current["id"], "create", None, "", initial_status)
    if initial_status == LeaveStatus.APPROVED:
        _apply_status_transition(leave_id, LeaveStatus.APPROVED, LeaveStatus.APPROVED,
                                 eff_start, eff_end, current["id"], _h, db, is_new=True)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.get("/")
def list_leaves(
    scope: str = "mine",
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    role = current["role"]
    # Bản ghi tổng hợp giả (nhập Excel / sửa tay hạn mức, xem update_used_days và
    # import_quota_apply) chỉ phục vụ tính _calc_used_days — không phải đơn nghỉ
    # phép thật, không có người nộp/KSV/TH/GĐ nào cả. Ẩn khỏi mọi danh sách đơn để
    # khỏi hiện như 1 đơn thật trong "Đơn của tôi"/"Chờ duyệt"/lịch/toàn trung tâm.
    clauses: list = ["NOT (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')"]
    params: list  = []

    if scope == "mine":
        clauses.append("staff_id = ?")
        params.append(current["id"])

    elif scope == "pending":
        # Đơn của GĐ đã tự động approved nhưng Tổng hợp chưa "biết" — chỉ mang tính
        # thông báo (xem tong-hop-ack), không phải điều kiện duyệt.
        _gd_unack = (
            "(status = 'approved' AND tong_hop_approver_id IS NULL "
            "AND staff_id IN (SELECT id FROM user_tttt WHERE role = 'giam_doc'))"
        )
        if role == "admin":
            clauses.append(f"(status IN ('pending_ksv','pending_tong_hop','pending_gd') OR {_gd_unack})")
        elif role in ("giam_doc", "pho_giam_doc"):
            if not _can_gd_review(current, db):
                return []
            clauses.append("gd_approver_id = ? AND status = 'pending_gd'")
            params.append(current["id"])
        elif _is_tong_hop_staff(current, db):
            if role in ("truong_phong", "pho_phong"):
                # PP/TP Tổng hợp: duyệt bước TH cho toàn trung tâm
                # VÀ duyệt bước KSV cho nhân viên phòng mình
                clauses.append(
                    f"(status = 'pending_tong_hop' OR "
                    f"(ksv_approver_id = ? AND status = 'pending_ksv') OR {_gd_unack})"
                )
                params.append(current["id"])
            else:
                clauses.append(f"(status = 'pending_tong_hop' OR {_gd_unack})")
        elif role in ("truong_phong", "pho_phong"):
            clauses.append("ksv_approver_id = ? AND status = 'pending_ksv'")
            params.append(current["id"])
        else:
            return []

    elif scope == "declared":
        # Người khai báo hộ xem các đơn mình đã khai báo
        clauses.append("direct_by = ?")
        params.append(current["id"])

    elif scope == "dept":
        # Phó phòng / Trưởng phòng xem tất cả đơn của nhân viên trong phòng mình
        if role not in ("truong_phong", "pho_phong", "admin"):
            raise HTTPException(403, "Không có quyền xem đơn phòng")
        dept_id = current.get("department_id")
        if not dept_id:
            return []
        clauses.append("staff_id IN (SELECT id FROM user_tttt WHERE department_id = ? AND is_active = 1)")
        params.append(dept_id)

    elif scope == "all":
        if role not in ("admin", "giam_doc", "pho_giam_doc"):
            if not _is_tong_hop_staff(current, db):
                raise HTTPException(403, "Không có quyền xem tất cả đơn")

    else:
        raise HTTPException(400, "scope phải là mine | pending | declared | dept | all")

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = db.execute(
        f"SELECT id FROM leave_records {where} ORDER BY created_at DESC", params
    ).fetchall()
    return [_leave_to_out(r["id"], db) for r in rows]


@router.get("/today")
def leaves_today(
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    """Thống kê nghỉ phép hôm nay: tổng + phân theo phòng."""
    # Chỉ đếm đơn ĐÃ DUYỆT — cố ý khác /calendar (lịch tháng hiện cả đơn chờ,
    # nhưng có nhãn trạng thái đi kèm). Endpoint này chỉ phục vụ card "Nghỉ phép
    # hôm nay" ở Trang chủ — con số trần, không nhãn — nên gộp đơn chưa duyệt vào
    # sẽ báo người vẫn đang đi làm là đã nghỉ. Đừng "đồng bộ" hai chỗ này.
    today = _vn_now().date().isoformat()
    rows = db.execute(
        """SELECT lr.id, lr.spread_dates, lr.start_date, lr.end_date,
                  u.full_name, d.name AS dept_name
           FROM leave_records lr
           JOIN user_tttt u ON lr.staff_id = u.id
           LEFT JOIN departments d ON u.department_id = d.id
           WHERE lr.status = 'approved'
             AND NOT (lr.reason LIKE '[Import]%' OR lr.reason LIKE '[Điều chỉnh]%')
             AND lr.start_date <= ? AND lr.end_date >= ?""",
        (today, today),
    ).fetchall()

    result = []
    for r in rows:
        if r["spread_dates"]:
            if today not in json.loads(r["spread_dates"]):
                continue
        result.append({
            "staff_name": r["full_name"] or "",
            "dept_name":  r["dept_name"] or "—",
        })

    by_dept: dict = {}
    for item in result:
        d = item["dept_name"]
        by_dept[d] = by_dept.get(d, 0) + 1

    return {
        "total":   len(result),
        "by_dept": [{"dept_name": d, "count": c}
                    for d, c in sorted(by_dept.items(), key=lambda x: -x[1])],
    }


@router.get("/calendar")
def leave_calendar(
    year: int,
    month: int,
    db: sqlite3.Connection = Depends(get_db),
    _: dict = Depends(get_current_staff),
):
    import calendar as _cal
    if not (1 <= month <= 12):
        raise HTTPException(400, "month phải từ 1 đến 12")
    # date(year, ...) bên dưới ném ValueError nếu năm ngoài 1..9999 → HTTP 500
    if not (2000 <= year <= 2100):
        raise HTTPException(400, "year phải từ 2000 đến 2100")
    last_day = _cal.monthrange(year, month)[1]
    start = date(year, month, 1)
    end   = date(year, month, last_day)

    leaves = db.execute(
        """SELECT lr.id, lr.start_date, lr.end_date, lr.leave_type, lr.status,
                  ks.full_name
           FROM leave_records lr
           LEFT JOIN user_tttt ks ON lr.staff_id = ks.id
           WHERE lr.status NOT IN ('rejected','cancelled')
             AND NOT (lr.reason LIKE '[Import]%' OR lr.reason LIKE '[Điều chỉnh]%')
             AND lr.start_date <= ? AND lr.end_date >= ?""",
        (end.isoformat(), start.isoformat()),
    ).fetchall()

    day_map: dict = {}
    cur = start
    while cur <= end:
        day_map[cur.isoformat()] = []
        cur += timedelta(days=1)

    for lv in leaves:
        cur = max(date.fromisoformat(lv["start_date"]), start)
        lv_end = min(date.fromisoformat(lv["end_date"]), end)
        while cur <= lv_end:
            day_map[cur.isoformat()].append({
                "staff_name": lv["full_name"] or "",
                "leave_type": lv["leave_type"],
                "status":     lv["status"],
                "leave_id":   lv["id"],
            })
            cur += timedelta(days=1)

    return {"year": year, "month": month, "days": day_map}


@router.get("/export")
def export_leaves(
    scope: str = "all",
    ids: str = "",        # danh sách id cách nhau bởi dấu phẩy, ưu tiên hơn scope
    date_from: str = "",  # ISO yyyy-mm-dd, lọc thêm theo khoảng ngày nghỉ (tuỳ chọn)
    date_to: str = "",
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    role = current["role"]
    # Bản ghi tổng hợp giả (nhập Excel / sửa tay hạn mức) không phải đơn nghỉ phép
    # thật — ẩn khỏi file xuất, xem chú thích tương tự ở list_leaves().
    clauses: list = ["NOT (lr.reason LIKE '[Import]%' OR lr.reason LIKE '[Điều chỉnh]%')"]
    params: list  = []

    # Nếu có danh sách ID cụ thể → xuất đúng những dòng đó
    if ids:
        id_list = [int(i.strip()) for i in ids.split(",") if i.strip().isdigit()]
        if not id_list:
            raise HTTPException(400, "ids không hợp lệ")
        placeholders = ",".join("?" * len(id_list))
        clauses.append(f"lr.id IN ({placeholders})")
        params.extend(id_list)
    elif scope == "mine":
        clauses.append("lr.staff_id = ?")
        params.append(current["id"])
    elif scope == "all":
        if role not in ("admin", "giam_doc", "pho_giam_doc"):
            if not _is_tong_hop_staff(current, db):
                raise HTTPException(403, "Không có quyền xuất tất cả đơn")
    elif scope == "pending":
        if _is_tong_hop_staff(current, db) and role in ("truong_phong", "pho_phong"):
            clauses.append(
                "(lr.status = 'pending_tong_hop' OR "
                "(lr.ksv_approver_id = ? AND lr.status = 'pending_ksv'))"
            )
            params.append(current["id"])
        elif role in ("truong_phong", "pho_phong"):
            clauses.append("lr.ksv_approver_id = ? AND lr.status = 'pending_ksv'")
            params.append(current["id"])
        elif _is_tong_hop_staff(current, db):
            clauses.append("lr.status = 'pending_tong_hop'")
        elif role in ("giam_doc", "pho_giam_doc", "admin"):
            clauses.append("lr.status = 'pending_gd'")
        else:
            clauses.append("1=0")
    elif scope in ("dept", "declared"):
        clauses.append("1=1")  # allow all, scope already handled by client
    else:
        raise HTTPException(400, "scope phải là mine | pending | all | dept | declared")

    # Lọc thêm theo khoảng ngày nghỉ nếu có — trước đây frontend tính fv/tv chỉ để
    # đặt TÊN FILE (vd "..._01012026-31012026...") nhưng không gửi lên đây, khiến nội
    # dung file luôn là toàn bộ scope bất kể tên file ghi gì — dễ hiểu lầm báo cáo sai kỳ.
    if date_from:
        clauses.append("lr.end_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("lr.start_date <= ?")
        params.append(date_to)

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    leaves = db.execute(
        f"""SELECT lr.*, s.full_name AS staff_name, s.department_id AS s_dept_id,
                   kv.full_name AS ksv_name, th.full_name AS th_name,
                   gd.full_name AS gd_name, gd.role AS gd_role,
                   d.name AS dept_name
            FROM leave_records lr
            LEFT JOIN user_tttt s  ON lr.staff_id             = s.id
            LEFT JOIN user_tttt kv ON lr.ksv_approver_id      = kv.id
            LEFT JOIN user_tttt th ON lr.tong_hop_approver_id = th.id
            LEFT JOIN user_tttt gd ON lr.gd_approver_id       = gd.id
            LEFT JOIN departments d ON s.department_id          = d.id
            {where}
            ORDER BY lr.created_at DESC""",
        params,
    ).fetchall()

    # Tải tất cả ngày lễ 1 lần cho toàn bộ khoảng
    all_holidays: FrozenSet[date] = frozenset()
    if leaves:
        min_d = min(date.fromisoformat(lv["start_date"]) for lv in leaves)
        max_d = max(date.fromisoformat(lv["end_date"])   for lv in leaves)
        all_holidays = _load_holidays(db, min_d, max_d)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nghỉ phép"

    hdr_fill = PatternFill("solid", fgColor="C62828")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers  = ["STT", "Họ và tên", "Phòng", "Loại nghỉ",
                "Ngày nghỉ", "Số ngày", "Trạng thái",
                "KSV duyệt", "Phòng Tổng hợp", "GĐ/PGĐ", "Ngày tạo"]
    widths   = [6, 25, 20, 18, 40, 9, 18, 22, 22, 22, 14]
    ws.append(headers)
    for cell, w in zip(ws[1], widths):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for idx, lv in enumerate(leaves, 1):
        s_date = date.fromisoformat(lv["start_date"])
        e_date = date.fromisoformat(lv["end_date"])
        days   = _period_days(s_date, e_date, all_holidays, lv["leave_type"])
        gd_name = ""
        if lv["gd_name"]:
            gd_name = lv["gd_name"]
            if lv["gd_role"] == "pho_giam_doc":
                gd_name += " (TUQ)"
        created = lv["created_at"] or ""
        try:
            from datetime import datetime as _dt
            created = _dt.fromisoformat(str(created)).strftime("%d/%m/%Y")
        except Exception:
            pass
        # Ngày nghỉ: liệt kê từng ngày nếu không liên nhau
        import json as _j
        if lv["spread_dates"]:
            try:
                spread = _j.loads(lv["spread_dates"])
                dates_str = ", ".join(date.fromisoformat(d).strftime("%d/%m/%Y") for d in sorted(spread))
            except Exception:
                dates_str = s_date.strftime("%d/%m/%Y")
        else:
            dates_str = f"{s_date.strftime('%d/%m/%Y')} → {e_date.strftime('%d/%m/%Y')}"

        row_data = [
            idx,
            lv["staff_name"] or "",
            lv["dept_name"] or "",
            LEAVE_TYPE_LABELS.get(lv["leave_type"] or "", ""),
            dates_str,
            days,
            _LEAVE_STATUS_VN.get(lv["status"] or "", lv["status"] or ""),
            lv["ksv_name"] or "",
            lv["th_name"]  or "",
            gd_name,
            created,
        ]
        ws.append(row_data)
        # Wrap text cho cột Ngày nghỉ (cột E = index 5)
        ws.cell(ws.max_row, 5).alignment = Alignment(wrap_text=True, horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''danh_sach_nghi_phep.xlsx"},
    )


# ─── Thông báo carry-over hết hiệu lực sau Q1 ──────────────────────────────────
# Phải khai báo TRƯỚC "/{leave_id}" bên dưới — nếu không, FastAPI/Starlette sẽ
# khớp "/carryover-notice" vào route "/{leave_id}" trước (cùng 1 segment, cùng
# method GET) rồi báo lỗi 422 khi ép kiểu int, route đúng bên dưới không bao
# giờ được gọi tới.

_CARRYOVER_NOTICE_CUTOFF = (3, 31)  # hết Q1 (31/3)


@router.get("/carryover-notice")
def get_carryover_notice(
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    """Có cần hiện popup thông báo carry-over hết hiệu lực không — 1 lần/năm/user,
    hiện cho mọi user (không phân biệt có carry-over hay không) kể từ sau 31/3."""
    today = _vn_now().date()
    year  = today.year
    cutoff = date(year, *_CARRYOVER_NOTICE_CUTOFF)
    if today <= cutoff:
        return {"show": False}
    row = db.execute(
        "SELECT carryover_notice_year FROM user_tttt WHERE id=?", (current["id"],)
    ).fetchone()
    already_seen = bool(row and row["carryover_notice_year"] == year)
    return {"show": not already_seen, "year": year}


@router.post("/carryover-notice/ack")
def ack_carryover_notice(
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    """Đánh dấu user đã xem thông báo carry-over hết hiệu lực trong năm nay."""
    year = _vn_now().date().year
    db.execute(
        "UPDATE user_tttt SET carryover_notice_year=? WHERE id=?", (year, current["id"]),
    )
    db.commit()
    return {"ok": True}


@router.get("/my-balance")
def get_my_balance(
    year: int = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    """Hạn mức phép của CHÍNH người đang đăng nhập — dùng cho banner "Phép còn lại"
    ở đầu trang. Khác với get_quotas (yêu cầu quyền leaves.quota_admin, xem toàn bộ
    nhân viên), endpoint này ai đăng nhập cũng gọi được nhưng chỉ trả về đúng 1
    người (current["id"]) — tái dùng cùng công thức (override trong leave_quotas
    + carry-over + used) để không lệch với tab Hạn mức phép.

    Đặt TRƯỚC /{leave_id} — nếu không, FastAPI khớp "/my-balance" vào path param
    leave_id (giống lý do staff.py đặt /export trước /{staff_id})."""
    from datetime import date as _today_d
    yr = year or _today_d.today().year
    staff_id = current["id"]

    q = db.execute(
        "SELECT quota_days FROM leave_quotas WHERE year=? AND staff_id=?", (yr, staff_id)
    ).fetchone()
    quota = float(q["quota_days"]) if q else float(compute_annual_leave(current.get("join_industry_date"), yr))
    carry = _carry_over_bulk([staff_id], yr, db, effective=True).get(staff_id, 0.0)
    used  = _calc_used_days_bulk([staff_id], yr, db, include_pending=True).get(staff_id, 0.0)

    return {
        "year":        yr,
        "quota_days":  quota,
        "carry_over":  carry,
        "used_days":   used,
        "remaining":   max(0.0, quota + carry - used),
    }


@router.get("/{leave_id}")
def get_leave(
    leave_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    r = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not r:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    is_th = _is_tong_hop_staff(current, db)
    if (r["staff_id"] != current["id"]
            and r["ksv_approver_id"] != current["id"]
            and r["tong_hop_approver_id"] != current["id"]
            and r["gd_approver_id"] != current["id"]
            and r["direct_by"] != current["id"]
            and not is_th
            and current["role"] not in ("admin", "giam_doc", "pho_giam_doc")):
        raise HTTPException(403, "Không có quyền xem đơn này")
    return _leave_to_out(leave_id, db)


@router.delete("/{leave_id}")
def delete_leave(
    leave_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    """Xóa đơn khai báo hộ — chỉ người khai báo hoặc admin mới được xóa."""
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn")
    if not leave["is_direct"]:
        raise HTTPException(403, "Chỉ có thể xóa đơn khai báo hộ")
    if leave["direct_by"] != current["id"] and current["role"] != "admin":
        raise HTTPException(403, "Không có quyền xóa đơn này")
    # Hoàn trả used_leave_days nếu đơn đã approved và không phải loại miễn quota
    if leave["status"] == "approved" and leave["leave_type"] not in _NO_QUOTA_TYPES:
        if leave["spread_dates"]:
            days = len(json.loads(leave["spread_dates"]))
        else:
            from datetime import date as _d
            s = _d.fromisoformat(leave["start_date"])
            e = _d.fromisoformat(leave["end_date"])
            _h = _load_holidays(db, s, e)
            days = calculate_leave_days(s, e, _h)
        db.execute(
            "UPDATE user_tttt SET used_leave_days = MAX(0, COALESCE(used_leave_days,0) - ?) WHERE id = ?",
            (days, leave["staff_id"]),
        )
    db.execute("DELETE FROM leave_action_logs WHERE leave_id = ?", (leave_id,))
    db.execute("DELETE FROM leave_records WHERE id = ?", (leave_id,))
    db.commit()
    return {"message": "Đã xóa đơn khai báo hộ"}


@router.put("/{leave_id}/ksv-review")
def ksv_review(
    leave_id: int,
    body: LeaveReview,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.approve_ksv")),
):
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    if leave["status"] != LeaveStatus.PENDING_KSV:
        raise HTTPException(400, f"Đơn đang ở trạng thái '{leave['status']}'")
    if leave["ksv_approver_id"] != current["id"] and current["role"] != "admin":
        raise HTTPException(403, "Bạn không phải người được chỉ định duyệt bước này")
    if body.action == "reject" and not body.comment:
        raise HTTPException(400, "Vui lòng nhập lý do từ chối")

    old = leave["status"]
    new_status = LeaveStatus.PENDING_TONG_HOP if body.action == "approve" else LeaveStatus.REJECTED
    # Nếu admin thực hiện: ghi lại chính admin là người duyệt bước này
    if current["role"] == "admin":
        db.execute("UPDATE leave_records SET ksv_approver_id=? WHERE id=?", (current["id"], leave_id))
    db.execute(
        "UPDATE leave_records SET ksv_approved_at=?, ksv_comment=? WHERE id=?",
        (str(_vn_now()), body.comment, leave_id),
    )
    if body.action == "approve" and body.signature:
        _save_signature(db, leave_id, "ksv", current["id"], body.signature)
    start = date.fromisoformat(leave["start_date"])
    end   = date.fromisoformat(leave["end_date"])
    _h    = _load_holidays(db, start, end)
    _apply_status_transition(leave_id, old, new_status, start, end, leave["staff_id"], _h, db)
    _log_action(db, leave_id, current["id"],
                "ksv_approve" if body.action == "approve" else "ksv_reject",
                body.comment, old, new_status)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.post("/{leave_id}/tong-hop-review")
def tong_hop_review(
    leave_id: int,
    body: TongHopReview,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    if not _is_tong_hop_staff(current, db) and current["role"] != "admin":
        raise HTTPException(403, "Chỉ nhân viên Phòng Tổng hợp hoặc Admin mới thực hiện được")
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    # [TẠM TẮT theo yêu cầu 2026-07-18 — bật lại khi được báo] Chặn tự xử lý đơn
    # của chính mình ở bước Tổng hợp — hiện cho phép Phòng Tổng hợp duyệt tất cả
    # đơn kể cả đơn của họ để test luồng duyệt trung tâm.
    # if leave["staff_id"] == current["id"] and current["role"] != "admin":
    #     raise HTTPException(403, "Không thể tự xử lý đơn nghỉ phép của chính mình")
    if leave["status"] != LeaveStatus.PENDING_TONG_HOP:
        raise HTTPException(400, f"Đơn đang ở trạng thái '{leave['status']}'")
    if leave["recall_reason"]:
        raise HTTPException(400, "Đây là yêu cầu rút đơn — dùng /recall-approve để xử lý")
    if body.action == "reject" and not body.comment:
        raise HTTPException(400, "Vui lòng nhập lý do từ chối")

    old = leave["status"]
    db.execute(
        "UPDATE leave_records SET tong_hop_approver_id=?, tong_hop_approved_at=?, tong_hop_comment=? WHERE id=?",
        (current["id"], str(_vn_now()), body.comment, leave_id),
    )

    if body.action == "forward":
        if not body.gd_approver_id:
            raise HTTPException(400, "Vui lòng chọn GĐ/PGĐ phê duyệt")
        gd = db.execute("SELECT * FROM user_tttt WHERE id = ? AND is_active = 1", (body.gd_approver_id,)).fetchone()
        if not gd or gd["role"] not in ("giam_doc", "pho_giam_doc"):
            raise HTTPException(400, "Người được chọn không phải GĐ hoặc PGĐ")
        db.execute("UPDATE leave_records SET gd_approver_id=? WHERE id=?", (body.gd_approver_id, leave_id))
        new_status = LeaveStatus.PENDING_GD
        action_key = "th_forward"
    else:
        new_status = LeaveStatus.REJECTED
        action_key = "th_reject"

    start = date.fromisoformat(leave["start_date"])
    end   = date.fromisoformat(leave["end_date"])
    _h    = _load_holidays(db, start, end)
    _apply_status_transition(leave_id, old, new_status, start, end, leave["staff_id"], _h, db)
    _log_action(db, leave_id, current["id"], action_key, body.comment, old, new_status)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.put("/{leave_id}/tong-hop-ack")
def tong_hop_ack_gd_leave(
    leave_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    """Tổng hợp xác nhận ĐÃ BIẾT đơn của Giám đốc — chỉ mang tính thông báo/ghi nhận,
    không phải điều kiện duyệt (đơn GĐ đã tự động approved ngay khi tạo)."""
    if not _is_tong_hop_staff(current, db) and current["role"] != "admin":
        raise HTTPException(403, "Chỉ Phòng Tổng hợp hoặc Admin mới xác nhận được")
    leave = db.execute("SELECT * FROM leave_records WHERE id=?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn")
    staff = db.execute("SELECT role FROM user_tttt WHERE id=?", (leave["staff_id"],)).fetchone()
    if not staff or staff["role"] != "giam_doc":
        raise HTTPException(400, "Chỉ áp dụng cho đơn nghỉ phép của Giám đốc")
    if leave["tong_hop_approver_id"]:
        raise HTTPException(400, "Đơn đã được Tổng hợp xác nhận trước đó")

    db.execute(
        "UPDATE leave_records SET tong_hop_approver_id=?, tong_hop_approved_at=? WHERE id=?",
        (current["id"], str(_vn_now()), leave_id),
    )
    _log_action(db, leave_id, current["id"], "th_ack_gd", None, leave["status"], leave["status"])
    db.commit()
    return _leave_to_out(leave_id, db)


@router.put("/{leave_id}/gd-review")
def gd_review(
    leave_id: int,
    body: LeaveReview,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.approve_gd")),
):
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    if leave["status"] != LeaveStatus.PENDING_GD:
        raise HTTPException(400, f"Đơn đang ở trạng thái '{leave['status']}'")
    if leave["staff_id"] == current["id"] and current["role"] != "admin":
        raise HTTPException(403, "Không thể tự duyệt đơn nghỉ phép của chính mình")
    if current["role"] != "admin":
        if not _can_gd_review(current, db):
            raise HTTPException(403, "Phó Giám đốc chưa được ủy quyền phê duyệt")
        if leave["gd_approver_id"] != current["id"]:
            raise HTTPException(403, "Bạn không phải người được chỉ định duyệt bước này")
    if body.action == "reject" and not body.comment:
        raise HTTPException(400, "Vui lòng nhập lý do từ chối")

    old = leave["status"]
    new_status = LeaveStatus.APPROVED if body.action == "approve" else LeaveStatus.REJECTED
    db.execute(
        "UPDATE leave_records SET gd_approved_at=?, gd_comment=?, gd_approver_id=? WHERE id=?",
        (str(_vn_now()), body.comment, current["id"], leave_id),
    )
    if body.action == "approve" and body.signature:
        _save_signature(db, leave_id, "gd", current["id"], body.signature)
    start = date.fromisoformat(leave["start_date"])
    end   = date.fromisoformat(leave["end_date"])
    _h    = _load_holidays(db, start, end)
    _apply_status_transition(leave_id, old, new_status, start, end, leave["staff_id"], _h, db)
    _log_action(db, leave_id, current["id"],
                "gd_approve" if body.action == "approve" else "gd_reject",
                body.comment, old, new_status)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.put("/{leave_id}/resubmit")
def resubmit_leave(
    leave_id: int,
    body: LeaveCreate,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.resubmit")),
):
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    if leave["staff_id"] != current["id"]:
        raise HTTPException(403, "Chỉ chủ nhân đơn mới được nộp lại")
    if leave["status"] != LeaveStatus.REJECTED:
        raise HTTPException(400, "Chỉ có thể nộp lại đơn đã bị từ chối")

    if body.spread_dates:
        spread = sorted(set(body.spread_dates))
        if len(spread) < 1:
            raise HTTPException(400, "spread_dates phải có ít nhất 1 ngày")
        try:
            eff_start  = date.fromisoformat(spread[0])
            eff_end    = date.fromisoformat(spread[-1])
        except ValueError:
            raise HTTPException(400, "Định dạng ngày không hợp lệ (yêu cầu YYYY-MM-DD)")
        leave_days = len(spread)
        spread_json = json.dumps(spread)
        _h = _load_holidays(db, eff_start, eff_end)
    else:
        if body.end_date < body.start_date:
            raise HTTPException(400, "Ngày kết thúc phải sau ngày bắt đầu")
        eff_start   = body.start_date
        eff_end     = body.end_date
        _h          = _load_holidays(db, eff_start, eff_end)
        leave_days  = _period_days(eff_start, eff_end, _h, body.leave_type)
        spread_json = None

    if body.leave_type not in _NO_QUOTA_TYPES and body.leave_type != "bat_buoc":
        ref_year  = eff_start.year
        carry_eff = compute_carry_over(current["id"], ref_year, db,
                                       effective=True, ref_date=eff_start)
        _q_row = db.execute(
            "SELECT quota_days FROM leave_quotas WHERE staff_id=? AND year=?",
            (current["id"], ref_year),
        ).fetchone()
        quota = (
            float(_q_row["quota_days"]) if _q_row
            else float(compute_annual_leave(current.get("join_industry_date"), ref_year))
        )
        used_cur  = _calc_used_days(current["id"], ref_year, db, include_pending=True)
        remaining = quota + carry_eff - used_cur
        if leave_days > remaining:
            raise HTTPException(400, f"Vượt quá số ngày phép còn lại ({remaining:.0f} ngày)")

    # Kiểm tra trùng ngày theo spread_dates thực tế
    if body.spread_dates:
        _existing = db.execute(
            """SELECT start_date, end_date, spread_dates FROM leave_records
               WHERE staff_id=? AND id!=? AND status NOT IN ('rejected','cancelled')
                 AND NOT (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')""",
            (current["id"], leave_id)
        ).fetchall()
        _new_days = set(spread)
        for _el in _existing:
            if _el["spread_dates"]:
                if _new_days & set(json.loads(_el["spread_dates"])):
                    raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")
            elif any(_el["start_date"] <= d <= _el["end_date"] for d in spread):
                raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")
    else:
        _existing2 = db.execute(
            """SELECT start_date, end_date, spread_dates FROM leave_records
               WHERE staff_id=? AND id!=? AND status NOT IN ('rejected','cancelled')
                 AND NOT (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')
                 AND start_date<=? AND end_date>=?""",
            (current["id"], leave_id, eff_end.isoformat(), eff_start.isoformat())
        ).fetchall()
        if _existing2:
            _new_range_days = None
            for _el in _existing2:
                if _el["spread_dates"]:
                    if _new_range_days is None:
                        _new_range_days, _d = set(), eff_start
                        while _d <= eff_end:
                            _new_range_days.add(_d.isoformat())
                            _d += timedelta(days=1)
                    if set(json.loads(_el["spread_dates"])) & _new_range_days:
                        raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")
                else:
                    raise HTTPException(409, "Khoảng ngày nghỉ bị trùng với đơn hiện có")

    if current["role"] == "giam_doc":
        # GĐ là cấp cao nhất — tự duyệt ngay, không qua quy trình (nhất quán với create_leave)
        new_status      = LeaveStatus.APPROVED
        ksv_approver_id = None
    elif current["role"] in _HIGH_ROLES:
        new_status      = LeaveStatus.PENDING_TONG_HOP
        ksv_approver_id = None
    else:
        ksv = _validate_ksv(body.ksv_approver_id, current, db)
        new_status      = LeaveStatus.PENDING_KSV
        ksv_approver_id = ksv["id"]

    # Validate và lưu GĐ/PGĐ approver nếu người dùng chọn
    new_gd_approver_id = None
    new_gd_approved_at = None
    if current["role"] == "giam_doc":
        new_gd_approver_id = current["id"]
        new_gd_approved_at = str(_vn_now())
    elif body.gd_approver_id:
        gd_row = db.execute(
            "SELECT id FROM user_tttt WHERE id=? AND role IN ('giam_doc','pho_giam_doc') AND is_active=1",
            (body.gd_approver_id,)
        ).fetchone()
        if gd_row:
            new_gd_approver_id = gd_row["id"]

    old = leave["status"]
    _log_action(db, leave_id, current["id"], "resubmit", None, old, new_status)

    db.execute(
        """UPDATE leave_records SET
               ksv_approver_id=?, ksv_approved_at=NULL, ksv_comment=NULL,
               tong_hop_approver_id=NULL, tong_hop_approved_at=NULL, tong_hop_comment=NULL,
               gd_approver_id=?, gd_approved_at=?, gd_comment=NULL,
               leave_type=?, start_date=?, end_date=?, reason=?, spread_dates=?
           WHERE id=?""",
        (ksv_approver_id, new_gd_approver_id, new_gd_approved_at, body.leave_type, eff_start.isoformat(),
         eff_end.isoformat(), body.reason, spread_json, leave_id),
    )
    # Đơn quay lại từ đầu → chữ ký của người duyệt cũ không còn giá trị. Ngày tháng
    # và số ngày phép trên phiếu đã đổi, giữ lại là để chữ ký thật nằm trên tờ đơn khác.
    db.execute("DELETE FROM leave_signatures WHERE leave_id=? AND slot IN ('ksv','gd')", (leave_id,))
    if body.signature:
        _save_signature(db, leave_id, "nguoi_de_nghi", current["id"], body.signature)
    _apply_status_transition(leave_id, old, new_status, eff_start, eff_end, leave["staff_id"], _h, db)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.patch("/{leave_id}/cancel")
def cancel_leave(
    leave_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    if leave["staff_id"] != current["id"] and current["role"] != "admin":
        raise HTTPException(403, "Chỉ chủ nhân đơn hoặc Admin mới được hủy")
    if leave["status"] == LeaveStatus.CANCELLED:
        raise HTTPException(400, "Đơn đã được hủy trước đó")
    if leave["recall_reason"] and current["role"] != "admin":
        raise HTTPException(
            400,
            "Đơn đang chờ Phòng Tổng hợp xác nhận rút — dùng chức năng xác nhận rút đơn để xử lý",
        )
    # Hủy thẳng đơn đã approved (bỏ qua luồng Rút đơn/Recall) chỉ dành cho:
    # GĐ tự hủy đơn của chính mình (toàn quyền), hoặc người có quyền
    # "leaves.cancel" được cấp riêng — khớp đúng điều kiện hiện nút ở frontend
    # (_can_cancel_now). Người khác phải dùng /recall cho đơn đã duyệt.
    if leave["status"] == LeaveStatus.APPROVED and current["role"] != "admin":
        _is_gd_self = current["role"] == "giam_doc" and leave["staff_id"] == current["id"]
        _has_cancel_feature = db.execute(
            """SELECT 1 FROM group_features gf
               JOIN group_members gm ON gm.group_id = gf.group_id
               JOIN user_groups g ON g.id = gm.group_id AND g.is_active = 1
               WHERE gm.staff_id = ? AND gf.feature_code = 'leaves.cancel'
               LIMIT 1""",
            (current["id"],),
        ).fetchone() is not None
        if not (_is_gd_self or _has_cancel_feature):
            raise HTTPException(
                403,
                "Đơn đã duyệt — vui lòng dùng chức năng Rút đơn (cần Phòng Tổng hợp xác nhận)",
            )

    old   = leave["status"]
    start = date.fromisoformat(leave["start_date"])
    end   = date.fromisoformat(leave["end_date"])
    _h    = _load_holidays(db, start, end)
    _apply_status_transition(leave_id, old, LeaveStatus.CANCELLED, start, end, leave["staff_id"], _h, db)
    _log_action(db, leave_id, current["id"], "cancel", None, old, LeaveStatus.CANCELLED)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.get("/{leave_id}/history")
def get_leave_history(
    leave_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    leave = db.execute("SELECT * FROM leave_records WHERE id = ?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    is_th = _is_tong_hop_staff(current, db)
    if (leave["staff_id"] != current["id"]
            and leave["ksv_approver_id"] != current["id"]
            and leave["gd_approver_id"] != current["id"]
            and leave["direct_by"] != current["id"]
            and not is_th
            and current["role"] not in ("admin", "giam_doc", "pho_giam_doc")):
        raise HTTPException(403)

    logs = db.execute(
        """SELECT lal.*, ks.full_name AS actor_name
           FROM leave_action_logs lal
           LEFT JOIN user_tttt ks ON lal.actor_id = ks.id
           WHERE lal.leave_id = ? ORDER BY lal.created_at""",
        (leave_id,),
    ).fetchall()

    return [
        {
            "id":           log["id"],
            "actor_name":   log["actor_name"] or "",
            "action":       log["action"],
            "action_label": ACTION_LABELS.get(log["action"], (log["action"], "grey"))[0],
            "action_color": ACTION_LABELS.get(log["action"], (log["action"], "grey"))[1],
            "comment":      log["comment"],
            "from_status":  log["from_status"],
            "to_status":    log["to_status"],
            "created_at":   log["created_at"],
        }
        for log in logs
    ]


# ─── Word template download ──────────────────────────────────────────────────

_ROLE_VN = {
    "chuyen_vien":   "Chuyên viên",
    "pho_phong":     "Phó phòng",
    "truong_phong":  "Trưởng phòng",
    "hau_kiem_vien": "Hậu kiểm viên",
    "giam_doc":      "Giám đốc",
    "pho_giam_doc":  "Phó Giám đốc",
    "admin":         "Quản trị viên cấp 1",
    "admin_l2":      "Quản trị viên cấp 2",
}

# template_path() chứ không phải os.path.join(): tên thư mục có dấu trên đĩa đang
# ở dạng NFD, ghép chuỗi NFC từ mã nguồn sẽ không khớp — xem backend/core/paths.py
_TPL_DIR = template_path("Phòng Tổng hợp", "Nghỉ phép")
_TPL_PATH = template_path("don_xin_nghi_phep_tpl.docx")

def _pick_template(staff_role: str) -> str:
    """Chọn file docx theo role. Fallback về template gốc nếu chưa có."""
    _map = {
        "chuyen_vien":   "don_xin_nghi_phep_nv.docx",
        "pho_phong":     "don_xin_nghi_phep_tp.docx",
        "truong_phong":  "don_xin_nghi_phep_tp.docx",
        "giam_doc":      "don_xin_nghi_phep_gd.docx",
        "pho_giam_doc":  "don_xin_nghi_phep_pgd.docx",
        "hau_kiem_vien": "don_xin_nghi_phep_nv.docx",
        "admin":         "don_xin_nghi_phep_tp.docx",
    }
    fname = _map.get(staff_role or "", "don_xin_nghi_phep_tp.docx")
    path  = os.path.join(_TPL_DIR, fname)
    return path if os.path.exists(path) else _TPL_PATH


def _is_business_contiguous(sorted_dates: list) -> bool:
    """True nếu các ngày liền nhau, hoặc khoảng cách giữa 2 ngày kế tiếp chỉ
    toàn thứ 7/CN (nghỉ hết tuần rồi nghỉ tiếp tuần sau) — coi như 1 khoảng
    liên tục để ghi gọn "Từ ngày...đến hết ngày..." thay vì liệt kê từng ngày."""
    for prev, cur in zip(sorted_dates, sorted_dates[1:]):
        gap = (cur - prev).days
        if gap == 1:
            continue
        if gap > 1 and all(
            (prev + timedelta(days=i)).weekday() >= 5
            for i in range(1, gap)
        ):
            continue
        return False
    return True


def _fmt_leave_period(start: date, end: date, days: int, spread_dates: list = None) -> str:
    if spread_dates and len(spread_dates) > 1:
        parsed = sorted(date.fromisoformat(d) for d in spread_dates)
        if _is_business_contiguous(parsed):
            # Liền nhau (hoặc chỉ cách bởi T7/CN) — ghi gọn thành 1 khoảng.
            return (
                f"{days:02d} ngày "
                f"(Từ ngày {parsed[0].strftime('%d/%m/%Y')} đến hết ngày {parsed[-1].strftime('%d/%m/%Y')})"
            )
        # Ngày lẻ không liên tục — liệt kê đủ từng ngày, không ghi "Từ...đến..."
        # (dễ hiểu lầm là nghỉ liên tục cả khoảng min→max).
        dates_str = ", ".join(d.strftime("%d/%m/%Y") for d in parsed)
        return f"{days:02d} ngày (các ngày: {dates_str})"
    if days == 1:
        return f"{days:02d} ngày (ngày {start.day:02d} tháng {start.month:02d} năm {start.year})"
    return (
        f"{days:02d} ngày "
        f"(Từ ngày {start.strftime('%d/%m/%Y')} đến ngày {end.strftime('%d/%m/%Y')})"
    )


_FORM_ROW_SQL = """SELECT lr.*,
          s.full_name AS staff_name, s.role AS staff_role,
          s.employee_code, s.annual_leave_days, s.used_leave_days,
          s.join_industry_date,
          d.name AS dept_name,
          kv.full_name AS ksv_name,
          gd.full_name AS gd_approver_name, gd.role AS gd_role
   FROM leave_records lr
   LEFT JOIN user_tttt s  ON lr.staff_id             = s.id
   LEFT JOIN departments d ON s.department_id          = d.id
   LEFT JOIN user_tttt kv ON lr.ksv_approver_id      = kv.id
   LEFT JOIN user_tttt gd ON lr.gd_approver_id       = gd.id
   WHERE lr.id = ?"""


def _load_form_row(leave_id: int, db: sqlite3.Connection):
    r = db.execute(_FORM_ROW_SQL, (leave_id,)).fetchone()
    if not r:
        raise HTTPException(404, "Không tìm thấy đơn nghỉ phép")
    return r


def _can_view_form(r, current: dict, db: sqlite3.Connection) -> bool:
    return (r["staff_id"] == current["id"]
            or r["ksv_approver_id"] == current["id"]
            or r["gd_approver_id"] == current["id"]
            or r["direct_by"] == current["id"]
            or _is_tong_hop_staff(current, db)
            or current["role"] in ("admin", "giam_doc", "pho_giam_doc"))


def _draft_form_row(body: LeaveCreate, current: dict, db: sqlite3.Connection) -> dict:
    """Dòng dữ liệu giả lập cho đơn CHƯA tạo — để xem trước trước khi gửi.

    Giữ đúng bộ khoá mà _build_form_ctx đọc; đơn thật thì các khoá này do câu
    SELECT ở _FORM_ROW_SQL cấp.
    """
    s = db.execute(
        """SELECT s.*, d.name AS dept_name
           FROM user_tttt s LEFT JOIN departments d ON s.department_id = d.id
           WHERE s.id = ?""",
        (current["id"],),
    ).fetchone()
    if not s:
        raise HTTPException(404, "Không tìm thấy hồ sơ nhân sự của bạn")

    def _name_role(sid):
        if not sid:
            return None, None
        row = db.execute("SELECT full_name, role FROM user_tttt WHERE id=?", (sid,)).fetchone()
        return (row["full_name"], row["role"]) if row else (None, None)

    ksv_name, _ = _name_role(body.ksv_approver_id)
    gd_name, gd_role = _name_role(body.gd_approver_id)
    if body.spread_dates:
        spread = sorted(set(body.spread_dates))
        eff_start, eff_end = spread[0], spread[-1]
        spread_json = json.dumps(spread)
    else:
        eff_start, eff_end = body.start_date.isoformat(), body.end_date.isoformat()
        spread_json = None
    return {
        "staff_id": current["id"], "staff_name": s["full_name"], "staff_role": s["role"],
        "employee_code": s["employee_code"], "annual_leave_days": s["annual_leave_days"],
        "used_leave_days": s["used_leave_days"], "join_industry_date": s["join_industry_date"],
        "dept_name": s["dept_name"], "direct_by": None,
        "start_date": eff_start, "end_date": eff_end, "spread_dates": spread_json,
        "leave_type": body.leave_type, "reason": body.reason,
        "ksv_approver_id": body.ksv_approver_id, "ksv_name": ksv_name,
        "gd_approver_id": body.gd_approver_id, "gd_approver_name": gd_name, "gd_role": gd_role,
    }


def _build_form_ctx(r, leave_id: Optional[int], db: sqlite3.Connection) -> tuple:
    """(ctx cho docxtpl, đường dẫn template). `r` là dòng thật hoặc dòng giả lập."""
    tpl_path = _pick_template(r["staff_role"])
    if not os.path.exists(tpl_path):
        raise HTTPException(500, "Chưa có template đơn nghỉ phép")

    start = date.fromisoformat(r["start_date"])
    end   = date.fromisoformat(r["end_date"])
    now   = _vn_now()
    _h    = _load_holidays(db, start, end)

    leave_days = len(json.loads(r["spread_dates"])) if r["spread_dates"] else _period_days(start, end, _h, r["leave_type"])
    # Ưu tiên hạn mức nhập tay (leave_quotas) — khớp đúng quy tắc get_quotas/export_quotas/
    # stats_annual; chỉ tính theo ngày vào ngành khi năm đó chưa có ai nhập tay.
    _q_row = db.execute(
        "SELECT quota_days FROM leave_quotas WHERE staff_id=? AND year=?",
        (r["staff_id"], start.year),
    ).fetchone()
    tong_phep = (
        float(_q_row["quota_days"]) if _q_row
        else (compute_annual_leave(r["join_industry_date"], start.year) if r["join_industry_date"] else (r["annual_leave_days"] or 12))
    )
    carry_original = compute_carry_over(r["staff_id"], start.year, db, effective=False)
    # Carryover hiệu lực theo ngày bắt đầu của đơn (Q1 → có carryover, sau Q1 → 0)
    carry_eff_doc  = compute_carry_over(r["staff_id"], start.year, db, effective=True, ref_date=start)
    # Số ngày đã nghỉ TRONG CÙNG NĂM (trừ đơn hiện tại)
    da_nghi = _calc_used_days(r["staff_id"], start.year, db, exclude_id=leave_id)
    con_lai = max(0.0, tong_phep + carry_eff_doc - da_nghi - leave_days)

    # ── Biến 2-năm cho trường hợp có ngày dư ──
    has_carryover   = carry_eff_doc > 0
    carryover_used  = min(float(carry_eff_doc), float(leave_days))
    new_year_days   = leave_days - carryover_used
    if has_carryover:
        _prev_year = start.year - 1
        _q_prev = db.execute(
            "SELECT quota_days FROM leave_quotas WHERE staff_id=? AND year=?",
            (r["staff_id"], _prev_year),
        ).fetchone()
        tong_so_phep_prev = float(_q_prev["quota_days"]) if _q_prev else float(
            compute_annual_leave(r["join_industry_date"], _prev_year) if r["join_industry_date"] else 12
        )
        da_nghi_prev = _calc_used_days(r["staff_id"], _prev_year, db)
        con_lai_prev = max(0.0, tong_so_phep_prev - da_nghi_prev - carryover_used)
        con_lai_cur  = max(0.0, tong_phep - da_nghi - new_year_days)
    else:
        _prev_year = start.year - 1
        tong_so_phep_prev = 0.0
        da_nghi_prev = 0.0
        con_lai_prev = 0.0
        con_lai_cur  = con_lai

    # Tên GĐ/PGĐ
    gd_name = r["gd_approver_name"] or ""

    # Tên phòng ngắn (bỏ tiền tố "Phòng ")
    dept_raw  = r["dept_name"] or ""
    dept_short = dept_raw.upper()
    if dept_short.startswith("PHÒNG "):
        dept_short = dept_short[6:]

    # Nhãn KSV — luôn hiện với role thường, kể cả khai báo hộ (ksv_approver_id=NULL)
    ksv_sign = r["staff_role"] not in ("truong_phong", "giam_doc", "pho_giam_doc")
    ksv_role = ""
    if r["ksv_approver_id"]:
        ksv_r = db.execute("SELECT role FROM user_tttt WHERE id=?", (r["ksv_approver_id"],)).fetchone()
        if ksv_r:
            ksv_role = ksv_r["role"]

    if ksv_role == "pho_phong":
        ksv_dept_label      = f"TUQ. TRƯỞNG PHÒNG {dept_short}"
        ksv_dept_label_line2 = "PHÓ TRƯỞNG PHÒNG"
    else:  # truong_phong hoặc không xác định (khai báo hộ)
        ksv_dept_label      = f"TRƯỞNG PHÒNG {dept_short}"
        ksv_dept_label_line2 = ""

    # Nhãn ký tên cho Ban lãnh đạo
    gd_role = r["gd_role"] or ""
    if gd_role == "pho_giam_doc":
        gd_title_line1 = "TUQ. GIÁM ĐỐC TTTT"
        gd_title_line2 = "PHÓ GIÁM ĐỐC"
    elif gd_role == "giam_doc":
        gd_title_line1 = "GIÁM ĐỐC TTTT"
        gd_title_line2 = ""
    else:
        gd_title_line1 = "GIÁM ĐỐC TTTT"
        gd_title_line2 = ""

    # annual/bat_buoc → ngày làm đơn; dot_xuat và các loại khác → ngày bắt đầu nghỉ
    _doc_date = now.date() if r["leave_type"] in ("annual", "bat_buoc") else start
    ctx = {
        "ngay_thang_nam":   f"{_doc_date.day:02d} tháng {_doc_date.month:02d} năm {_doc_date.year}",
        "ho_va_ten":        r["staff_name"] or "",
        "chuc_vu":          ("Giám đốc - Trung tâm Thanh toán Agribank"
                            if r["staff_role"] == "giam_doc"
                            else "Phó Giám đốc - Trung tâm Thanh toán Agribank"
                            if r["staff_role"] == "pho_giam_doc"
                            else _ROLE_VN.get(r["staff_role"] or "", r["staff_role"] or "")),
        "don_vi":           r["dept_name"] or "",
        "nam_phep":         str(start.year),
        "tong_so_phep":      str(tong_phep),
        "so_ngay_da_nghi":   f"{da_nghi:g}",
        "so_ngay_xin_nghi":  _fmt_leave_period(
            start, end, leave_days,
            json.loads(r["spread_dates"]) if r["spread_dates"] else None,
        ),
        "so_ngay_con_lai":   f"{con_lai_cur:g}",
        # 2-năm (carryover)
        "has_carryover":     has_carryover,
        "prev_year":         str(_prev_year),
        "tong_so_phep_prev": f"{int(tong_so_phep_prev)}",
        "da_nghi_prev":      f"{da_nghi_prev:g}",
        "da_nghi_cur":       f"{da_nghi:g}",
        "con_lai_prev":      f"{con_lai_prev:g}",
        "con_lai_cur":       f"{con_lai_cur:g}",
        "ly_do":            r["reason"] or "",
        "ksv_name":         (r["ksv_name"] or "") if ksv_sign else "",
        "gd_name":          gd_name,
        "gd_title_line1":        gd_title_line1,
        "gd_title_line2":        gd_title_line2,
    }

    ctx["ksv_dept_label"] = ksv_dept_label if ksv_sign else ""
    ctx["ksv_dept_label_line2"] = ksv_dept_label_line2 if ksv_sign else ""

    # Override chuc_vu để thêm phòng + tên ngân hàng cho các role cấp trung
    if r["staff_role"] in ("truong_phong", "pho_phong"):
        role_prefix = _ROLE_VN.get(r["staff_role"], "")
        dept_suffix = (r["dept_name"] or "").replace("Phòng ", "")
        if dept_suffix:
            ctx["chuc_vu"] = f"{role_prefix} {dept_suffix} - Trung tâm Thanh toán Agribank"
        else:
            ctx["chuc_vu"] = f"{role_prefix} - Trung tâm Thanh toán Agribank"

    _LEAVE_TYPE_VN = {
        "thai_san": "thai sản theo chế độ", "bao_hiem": "bảo hiểm theo chế độ",
        "annual": "phép năm", "bat_buoc": "phép bắt buộc",
        "other": "phép khác",
    }
    is_no_quota = r["leave_type"] in _NO_QUOTA_TYPES
    ctx["is_no_quota"]    = is_no_quota
    ctx["leave_type_vn"]  = _LEAVE_TYPE_VN.get(r["leave_type"], r["leave_type"] or "")
    # thai_san / bao_hiem không tính hạn mức — để trống các ô quota trên phiếu
    if is_no_quota:
        ctx.update({
            "tong_so_phep":      "",
            "so_ngay_da_nghi":   "",
            "so_ngay_con_lai":   "",
            "has_carryover":     False,
            "tong_so_phep_prev": "",
            "da_nghi_prev":      "",
            "da_nghi_cur":       "",
            "con_lai_prev":      "",
            "con_lai_cur":       "",
        })
    return ctx, tpl_path


def _render_form_docx(ctx: dict, tpl_path: str) -> bytes:
    from docxtpl import DocxTemplate
    tpl = DocxTemplate(tpl_path)
    tpl.render(ctx)
    buf = io.BytesIO()
    tpl.save(buf)
    return buf.getvalue()


# ─── Bản PDF + chữ ký ────────────────────────────────────────────────────────
# Nhãn để dò vị trí gợi ý cho từng ô ký. Dò PHÂN BIỆT HOA-THƯỜNG: chữ thường
# "Trưởng phòng" còn nằm ở dòng "Chức vụ:" phía trên, khớp trúng dòng đó thì chữ
# ký nhảy lên giữa trang.
_SIG_SLOTS = {
    "nguoi_de_nghi": "NGƯỜI ĐỀ NGHỊ",
    "ksv":           "TRƯỞNG PHÒNG",
    "gd":            "GIÁM ĐỐC TTTT",
}
_SIG_SLOT_VN = {"nguoi_de_nghi": "Người đề nghị", "ksv": "Trưởng phòng", "gd": "Ban lãnh đạo"}
_SIG_W_MM = 38.0          # bề ngang mặc định của khung chữ ký
_SIG_RATIO = 0.38         # cao/rộng khi không đọc được kích thước ảnh


def _form_pdf(r, leave_id: Optional[int], db: sqlite3.Connection) -> bytes:
    """PDF gốc (chưa có chữ ký). Word chỉ chạy khi cache trượt — khoá cache là
    nội dung đơn nên sửa người duyệt / số ngày là tự dựng lại."""
    ctx, tpl_path = _build_form_ctx(r, leave_id, db)
    try:
        mtime = os.path.getmtime(tpl_path)
    except OSError:
        mtime = 0
    key = leave_pdf.cache_key(tpl_path, mtime, json.dumps(ctx, sort_keys=True, default=str))
    return leave_pdf.base_pdf(key, lambda: _render_form_docx(ctx, tpl_path))


def _sig_image(staff_id: Optional[int], db: sqlite3.Connection) -> Optional[bytes]:
    if not staff_id:
        return None
    row = db.execute("SELECT image FROM user_signatures WHERE staff_id=?", (staff_id,)).fetchone()
    return row["image"] if row else None


def _suggest_sig_box(pdf: bytes, slot: str, image: Optional[bytes]) -> dict:
    """Khung gợi ý: ngay dưới nhãn ô ký, giữa theo chiều ngang của nhãn."""
    ratio = _SIG_RATIO
    if image:
        w_px, h_px = leave_pdf.png_size(image)
        if w_px and h_px:
            ratio = h_px / w_px
    w = _SIG_W_MM
    h = max(6.0, min(45.0, w * ratio))
    anchor = None
    needle = _SIG_SLOTS.get(slot)
    if needle:
        anchor = leave_pdf.find_text_box(pdf, needle, match_case=True)
    if anchor:
        x = anchor["x_mm"] + anchor["w_mm"] / 2 - w / 2
        y = anchor["y_mm"] + anchor["h_mm"] + 1.5
        page = anchor["page"]
    else:
        # Không dò được nhãn (mẫu đơn đổi chữ) — thả xuống giữa nửa dưới trang,
        # người ký tự kéo. Thà lệch còn hơn không hiện chữ ký nào.
        x, y, page = 85.0, 200.0, 0
    return {"page": page, "x_mm": round(x, 2), "y_mm": round(y, 2),
            "w_mm": round(w, 2), "h_mm": round(h, 2)}


def _placed_signatures(leave_id: Optional[int], db: sqlite3.Connection) -> list:
    if not leave_id:
        return []
    rows = db.execute(
        """SELECT slot, page, x_mm, y_mm, w_mm, h_mm, image, signed_at
           FROM leave_signatures WHERE leave_id=? ORDER BY id""",
        (leave_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def _validate_placement(pl) -> dict:
    """Chặn toạ độ vô lý trước khi ghi DB (khung ngoài trang, bé/to bất thường)."""
    if pl is None:
        return None
    box = {"page": int(pl.page or 0), "x_mm": float(pl.x_mm), "y_mm": float(pl.y_mm),
           "w_mm": float(pl.w_mm), "h_mm": float(pl.h_mm)}
    if box["page"] < 0 or box["page"] > 20:
        raise HTTPException(400, "Số trang đặt chữ ký không hợp lệ")
    if not (5.0 <= box["w_mm"] <= 150.0 and 3.0 <= box["h_mm"] <= 150.0):
        raise HTTPException(400, "Kích thước chữ ký không hợp lệ (rộng 5–150mm, cao 3–150mm)")
    if not (-5.0 <= box["x_mm"] <= 420.0 and -5.0 <= box["y_mm"] <= 594.0):
        raise HTTPException(400, "Vị trí chữ ký nằm ngoài trang")
    return box


def _save_signature(db: sqlite3.Connection, leave_id: int, slot: str,
                    staff_id: int, pl) -> None:
    """Ghi chữ ký vào đơn. Ảnh được SAO LẠI vào leave_signatures ngay lúc ký —
    sau này người ký đổi hoặc xoá ảnh cá nhân thì đơn cũ không đổi theo."""
    box = _validate_placement(pl)
    if box is None:
        return
    image = _sig_image(staff_id, db)
    if not image:
        raise HTTPException(400, "Chưa có ảnh chữ ký — vào Quản lý người dùng để tải lên")
    db.execute("DELETE FROM leave_signatures WHERE leave_id=? AND slot=?", (leave_id, slot))
    db.execute(
        """INSERT INTO leave_signatures
               (leave_id, slot, staff_id, page, x_mm, y_mm, w_mm, h_mm, image, signed_at)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (leave_id, slot, staff_id, box["page"], box["x_mm"], box["y_mm"],
         box["w_mm"], box["h_mm"], image, str(_vn_now())),
    )


def _data_url(image: bytes) -> str:
    return "data:image/png;base64," + base64.b64encode(image).decode()


def _preview_payload(r, leave_id: Optional[int], slot: str,
                     signer_id: Optional[int], db: sqlite3.Connection) -> dict:
    """Ảnh trang đơn + chữ ký đã có + khung gợi ý cho ô ký sắp tới."""
    try:
        pdf = _form_pdf(r, leave_id, db)
    except leave_pdf.PdfConvertError as e:
        raise HTTPException(503, f"Không dựng được bản xem trước: {e}")

    all_sigs = _placed_signatures(leave_id, db)
    placed = [p for p in all_sigs if p["slot"] != slot]
    my_image = _sig_image(signer_id, db) if slot else None
    # Ô này đã ký rồi (duyệt lại, xem lại) → mở đúng chỗ cũ, đừng kéo về chỗ gợi ý
    earlier = next((p for p in all_sigs if p["slot"] == slot), None)
    if not slot:
        suggest = None
    elif earlier:
        suggest = {k: earlier[k] for k in ("page", "x_mm", "y_mm", "w_mm", "h_mm")}
    else:
        suggest = _suggest_sig_box(pdf, slot, my_image)
    page_no = suggest["page"] if suggest else (placed[0]["page"] if placed else 0)
    png, w_mm, h_mm, n_pages = leave_pdf.page_png(pdf, page_no)
    return {
        "page": page_no,
        "pages": n_pages,
        "page_png": _data_url(png),
        "page_w_mm": round(w_mm, 2),
        "page_h_mm": round(h_mm, 2),
        "slot": slot or None,
        "slot_label": _SIG_SLOT_VN.get(slot, ""),
        "signature": ({"data_url": _data_url(my_image)} if my_image else None),
        "suggest": suggest,
        "placed": [
            {"slot": p["slot"], "label": _SIG_SLOT_VN.get(p["slot"], p["slot"]),
             "x_mm": p["x_mm"], "y_mm": p["y_mm"], "w_mm": p["w_mm"], "h_mm": p["h_mm"],
             "data_url": _data_url(p["image"])}
            for p in placed if p["page"] == page_no
        ],
    }


def _sig_slot_for(r, current: dict, db: sqlite3.Connection) -> str:
    """Ô ký mà người đang đăng nhập được quyền ký trên đơn này."""
    if r["staff_id"] == current["id"]:
        return "nguoi_de_nghi"
    if r["status"] == LeaveStatus.PENDING_KSV and (
            r["ksv_approver_id"] == current["id"] or current["role"] == "admin"):
        return "ksv"
    if r["status"] == LeaveStatus.PENDING_GD and (
            r["gd_approver_id"] == current["id"] or current["role"] == "admin"):
        return "gd"
    return ""


# ── Ba endpoint dưới đây gọi Word qua PowerShell (5–7 giây khi cache lạnh, tối
# đa 150 giây nếu Word treo) và tuần tự hoá trên `leave_pdf._word_lock`.
#
# Chúng PHẢI là `async def` + `await run_heavy(...)`. Để `def` thì FastAPI đẩy
# vào threadpool CHUNG 40 token của anyio: mấy người cùng bấm "Xem trước" là
# xếp hàng ở `_word_lock` mà vẫn mỗi người giữ một token, Word treo một lần là
# bể cạn — và lúc đó MỌI endpoint khác của hệ thống (chấm công, bàn giao, sổ
# trực... đều là `def`) cùng đứng chờ theo. Đo được trên hệ thống thật: 40 việc
# nặng đồng thời làm `/api/auth/me` kẹt 38 giây.
#
# `run_heavy()` giới hạn ở MAX_HEAVY=4, phần token còn lại luôn dành cho request
# nhẹ. Xem backend/core/concurrency.py.
#
# Kết nối SQLite dùng lại được trong luồng phụ vì `get_db()` mở với
# `check_same_thread=False`, và ở đây chỉ có một luồng đụng vào tại một thời điểm.


@router.post("/preview/warmup")
def warm_up_preview(current: dict = Depends(get_current_staff)):
    """Bật sẵn Word ở nền — gọi ngay khi mở màn nghỉ phép.

    Mở Word tốn ~1,5 giây, chuyển một file chỉ tốn ~0,35 giây. Bật trước lúc người
    dùng còn đang điền đơn thì đến khi bấm "Xem trước" gần như không phải chờ.

    Trả lời ngay, không đợi Word lên: đây là việc dọn đường, hỏng cũng không sao
    (đường xem trước thật vẫn tự bật Word khi cần).
    """
    threading.Thread(target=leave_pdf.warm_up, name="word-warmup", daemon=True).start()
    return {"ok": True}


@router.post("/preview")
async def preview_draft_form(
    body: LeaveCreate,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.create")),
):
    """Xem trước đơn CHƯA gửi — để người làm đơn đặt chữ ký rồi mới bấm gửi."""
    if body.leave_type not in _VALID_LEAVE_TYPES:
        raise HTTPException(400, f"Loại nghỉ phép không hợp lệ: {body.leave_type}")

    def _work():
        r = _draft_form_row(body, current, db)
        return _preview_payload(r, None, "nguoi_de_nghi", current["id"], db)

    return await run_heavy(_work)


@router.get("/{leave_id}/preview")
async def preview_leave_form(
    leave_id: int,
    slot: str = "",
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    """Xem trước đơn đã có. `slot` để trống → tự chọn ô ký hợp lệ của người gọi."""
    if slot and slot not in _SIG_SLOTS:
        raise HTTPException(400, "Ô ký không hợp lệ")

    def _work():
        r = _load_form_row(leave_id, db)
        if not _can_view_form(r, current, db):
            raise HTTPException(403, "Không có quyền xem đơn này")
        o = slot or _sig_slot_for(r, current, db)
        return _preview_payload(r, leave_id, o, current["id"] if o else None, db)

    return await run_heavy(_work)


@router.get("/{leave_id}/download")
async def download_leave_form(
    leave_id: int,
    fmt: str = "pdf",
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(get_current_staff),
):
    def _work():
        r = _load_form_row(leave_id, db)
        if not _can_view_form(r, current, db):
            raise HTTPException(403, "Không có quyền tải đơn này")

        ec = r["employee_code"] or "staff"
        if fmt == "docx":
            # Đường lui khi máy chủ không chuyển được PDF (chưa cài Word / Word treo).
            ctx, tpl_path = _build_form_ctx(r, leave_id, db)
            return (_render_form_docx(ctx, tpl_path),
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    f"don_nghi_phep_{ec}_{r['start_date']}.docx")

        try:
            pdf = _form_pdf(r, leave_id, db)
            pdf = leave_pdf.stamp(pdf, _placed_signatures(leave_id, db))
        except leave_pdf.PdfConvertError as e:
            raise HTTPException(503, f"Không tạo được PDF: {e}")
        return pdf, "application/pdf", f"don_nghi_phep_{ec}_{r['start_date']}.pdf"

    noi_dung, kieu, ten_file = await run_heavy(_work)
    return StreamingResponse(
        io.BytesIO(noi_dung),
        media_type=kieu,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{ten_file}"},
    )


# ─── Hạn mức phép (quota) ──────────────────────────────────────────────────────

@router.get("/quotas/{year}")
def get_quotas(
    year: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Trả danh sách hạn mức phép của tất cả nhân viên trong năm."""
    staffs = db.execute(
        """SELECT u.id, u.full_name, u.employee_code, u.join_industry_date, d.name AS dept_name
           FROM user_tttt u
           LEFT JOIN departments d ON u.department_id = d.id
           WHERE u.is_active=1 AND (u.is_deleted=0 OR u.is_deleted IS NULL)
           ORDER BY d.name, u.full_name"""
    ).fetchall()
    staff_ids = [s["id"] for s in staffs]
    from datetime import date as _today_d
    is_current_year = (year == _today_d.today().year)
    # Gộp truy vấn hạn mức/carry-over/đã dùng cho TOÀN BỘ nhân viên thành vài query
    # thay vì mỗi nhân viên ~9-10 query (N+1) — quan trọng vì endpoint này load mỗi
    # lần mở trang Hạn mức phép.
    quota_by_staff = {
        r["staff_id"]: float(r["quota_days"])
        for r in db.execute(
            f"SELECT staff_id, quota_days FROM leave_quotas WHERE year=? AND staff_id IN "
            f"({','.join('?' * len(staff_ids))})", [year] + staff_ids
        ).fetchall()
    } if staff_ids else {}
    carry_by_staff      = _carry_over_bulk(staff_ids, year, db, effective=True)
    carry_orig_by_staff = _carry_over_bulk(staff_ids, year, db, effective=False) if is_current_year else {}
    # include_pending=True: khớp đúng enforcement thật lúc tạo đơn (create_leave
    # dùng include_pending=True), tránh hiện "còn lại" cao hơn thực tế cho phép.
    used_by_staff        = _calc_used_days_bulk(staff_ids, year, db, include_pending=True)
    result = []
    for s in staffs:
        quota = quota_by_staff.get(s["id"], float(compute_annual_leave(s["join_industry_date"], year)))
        carry          = carry_by_staff.get(s["id"], 0.0)    # hiển thị Chuyển kỳ
        # Chỉ dùng carry_original để bù khi đang xem năm hiện tại (sau Q1).
        # Năm cũ → Q1 đã qua lâu rồi, carry-over không còn liên quan nữa.
        carry_original = carry_orig_by_staff.get(s["id"], 0.0) if is_current_year else 0.0
        used = used_by_staff.get(s["id"], 0.0)
        result.append({
            "staff_id":         s["id"],
            "staff_name":       s["full_name"],
            "employee_code":    s["employee_code"] or "",
            "dept_name":        s["dept_name"] or "",
            "join_industry_date": str(s["join_industry_date"])[:10] if s["join_industry_date"] else "",
            "year":             year,
            "quota_days":       quota,
            "carry_over":       carry,
            "carry_original":   carry_original,
            "used_days":        used,
            # "remaining" phải khớp số ngày enforcement thực sự cho phép đặt tiếp
            # (create_leave dùng carry hiệu lực theo Q1, không phải carry_original thô).
            "remaining":        max(0.0, quota + carry - used),
        })
    return result


@router.post("/quotas")
def upsert_quota(
    body: LeaveQuotaUpsert,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Ghi đè hạn mức phép cho nhân viên trong năm cụ thể."""
    if body.quota_days < 0:
        raise HTTPException(400, "quota_days không được âm")
    staff = db.execute("SELECT id FROM user_tttt WHERE id=? AND is_active=1", (body.staff_id,)).fetchone()
    if not staff:
        raise HTTPException(404, "Nhân viên không tồn tại")
    db.execute(
        "INSERT INTO leave_quotas (staff_id, year, quota_days) VALUES (?,?,?) ON CONFLICT(staff_id, year) DO UPDATE SET quota_days=excluded.quota_days",
        (body.staff_id, body.year, body.quota_days),
    )
    db.commit()
    return {"ok": True, "staff_id": body.staff_id, "year": body.year, "quota_days": body.quota_days}


@router.patch("/quotas/staff/{staff_id}/join-date")
def update_join_date(
    staff_id: int,
    body: dict,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Cập nhật ngày vào ngành — chỉ dành cho quota admin.

    Đây là đường ghi thứ ba vào `user_tttt.join_industry_date` (hai đường kia ở
    `backend/api/staff.py`: sửa cán bộ và nhập Excel hàng loạt). Cột này là hồ
    sơ nhân sự chứ không phải số liệu phép, nên phải giữ hai thứ ngang bằng hai
    đường kia:

      - **Validate ISO.** Trước đây nhận thẳng chuỗi client gửi và ghi nguyên
        vào cột DATE. Gõ "01/07/2020" hay "hôm qua" đều vào được, rồi
        `compute_annual_leave()` và mọi chỗ `date.fromisoformat()` đọc cột này
        sẽ vỡ — ở nơi khác, muộn hơn, không ai lần ra nguyên nhân.
      - **Ghi nhật ký kèm giá trị cũ.** AuditMiddleware có ghi, nhưng chỉ ghi
        được `PATCH <đường dẫn>` — không biết ai đổi từ ngày nào sang ngày nào.
        Số ngày phép năm tính từ cột này, nên đổi nó là đổi hạn mức phép.
    """
    join_date = (body.get("join_industry_date") or "").strip()
    if not join_date:
        raise HTTPException(400, "join_industry_date không được để trống")
    try:
        join_date = date.fromisoformat(join_date).isoformat()
    except ValueError:
        raise HTTPException(400, "Ngày vào ngành phải theo định dạng YYYY-MM-DD")
    if date.fromisoformat(join_date) > _vn_now().date():
        raise HTTPException(400, "Ngày vào ngành không được ở tương lai")

    staff = db.execute(
        "SELECT id, full_name, join_industry_date FROM user_tttt WHERE id=? AND is_active=1",
        (staff_id,),
    ).fetchone()
    if not staff:
        raise HTTPException(404, "Không tìm thấy nhân viên")

    db.execute("UPDATE user_tttt SET join_industry_date=? WHERE id=?", (join_date, staff_id))
    write_audit(
        db, current["id"], "staff_join_date_update", "staff", staff_id,
        f"{staff['full_name']}: ngày vào ngành {staff['join_industry_date'] or '(trống)'} → {join_date}",
    )
    db.commit()
    return {"ok": True, "staff_id": staff_id, "join_industry_date": join_date}


@router.patch("/quotas/staff/{staff_id}/used-days")
def update_used_days(
    staff_id: int,
    body: dict,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Sửa tay số ngày đã dùng cho 1 nhân viên/năm.

    _calc_used_days (nguồn sự thật duy nhất) không đọc user_tttt.used_leave_days —
    nó đếm từ leave_records — nên sửa tay cũng phải ghi qua 1 bản ghi nghỉ bat_buoc
    approved tổng hợp, giống hệt cơ chế "Nhập file hạn mức" (import_quota_apply).
    Xoá bản ghi tổng hợp cũ (dù tạo bởi import hay sửa tay trước đó) trước khi tạo
    mới — 2 cách nhập luôn thay thế lẫn nhau, không cộng dồn.
    """
    try:
        year = int(body.get("year"))
        used_days = float(body.get("used_days"))
    except (TypeError, ValueError):
        raise HTTPException(400, "year/used_days không hợp lệ")
    if used_days < 0:
        raise HTTPException(400, "Số ngày đã dùng không được âm")
    staff = db.execute("SELECT id FROM user_tttt WHERE id=? AND is_active=1", (staff_id,)).fetchone()
    if not staff:
        raise HTTPException(404, "Không tìm thấy nhân viên")

    db.execute(
        "DELETE FROM leave_records WHERE staff_id=? AND leave_type='bat_buoc' "
        "AND strftime('%Y', start_date)=? AND (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')",
        (staff_id, str(year)),
    )
    # used_days là TỔNG mong muốn (khớp giá trị đang hiển thị/tiền điền trên dialog —
    # vốn đã gồm cả đơn nghỉ THẬT), không phải số ngày cộng thêm. Sau khi xoá bản ghi
    # tổng hợp cũ ở trên, phần còn lại trong leave_records là đơn THẬT — phải trừ đi
    # phần này rồi mới chèn đúng phần chênh lệch. Thiếu bước trừ này thì mỗi lần lưu
    # (kể cả không đổi gì, vì dialog tự điền sẵn tổng hiện tại) sẽ cộng dồn thêm đúng
    # bằng tổng cũ — tăng vô hạn qua từng lần bấm Lưu.
    real_used = _calc_used_days(staff_id, year, db, include_pending=True)
    delta = used_days - real_used
    n_days = int(delta + 0.5) if delta > 0 else 0  # round-half-up, không âm
    if n_days >= 1:
        sd = _import_spread_dates(n_days, year)
        if sd:
            now = _vn_now()
            db.execute(
                """INSERT INTO leave_records
                       (staff_id, leave_type, start_date, end_date, spread_dates,
                        status, reason, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (staff_id, "bat_buoc", sd[0], sd[-1], json.dumps(sd), "approved",
                 f"[Điều chỉnh] Đặt số ngày đã dùng = {n_days} (năm {year})",
                 now, now),
            )
    final_used = real_used + n_days
    db.execute("UPDATE user_tttt SET used_leave_days=? WHERE id=?", (final_used, staff_id))
    db.commit()
    return {
        "ok": True, "staff_id": staff_id, "year": year,
        "used_days": final_used,
        # requested < real_used (vd đơn thật đã nhiều hơn số muốn đặt) → không thể trừ
        # bớt đơn thật bằng bản ghi tổng hợp, kết quả bị chặn ở real_used — báo cho
        # frontend biết để thông báo thay vì im lặng lệch số.
        "capped": used_days < real_used,
    }


# ─── Nhập file hạn mức (Excel) ─────────────────────────────────────────────────
# File có thể khác cấu trúc/thứ tự cột giữa các lần — dò cột theo TIÊU ĐỀ (dòng
# header) thay vì cố định vị trí, chỉ cần đủ các trường: STT, Họ và tên,
# Mã cán bộ, Hạn mức, Đã nghỉ. Khớp nhân viên theo Mã cán bộ trước (duy nhất
# theo từng người) — nếu không khớp thì thử theo tên (chuẩn hoá bỏ dấu, chỉ
# nhận khi tên đó chỉ khớp đúng 1 nhân viên). Chỉ nhập "Hạn mức" + "Đã nghỉ"
# (ghi đè used_leave_days) — không nhập "Chuyển năm" vì hệ thống tính động
# từ hạn mức + số ngày đã dùng của năm trước.

def _qi_detect_columns(ws) -> Optional[dict]:
    """Dò dòng tiêu đề trong 6 dòng đầu, trả về map field -> chỉ số cột."""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        col_map: dict = {}
        for idx, cell in enumerate(row or ()):
            n = _norm_vn(cell)
            if not n:
                continue
            if "stt" not in col_map and n == "stt":
                col_map["stt"] = idx
            elif "ma_can_bo" not in col_map and ("ma can bo" in n or "ma cb" in n or "ma nv" in n or "ma nhan vien" in n):
                col_map["ma_can_bo"] = idx
            elif "ho_ten" not in col_map and ("ho ten" in n or "ho va ten" in n or n == "ten"):
                col_map["ho_ten"] = idx
            elif "phong" not in col_map and "phong" in n:
                col_map["phong"] = idx
            elif "han_muc" not in col_map and "han muc" in n:
                col_map["han_muc"] = idx
            elif "da_nghi" not in col_map and "da nghi" in n:
                col_map["da_nghi"] = idx
        if {"stt", "ho_ten", "han_muc"} <= col_map.keys():
            return col_map
    return None


@router.post("/quotas/{year}/import/preview")
def import_quota_preview(
    year: int,
    file: UploadFile = File(...),
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Đọc file Excel hạn mức, khớp nhân viên theo Mã cán bộ / tên — KHÔNG ghi DB."""
    import openpyxl
    content = read_limited_sync(file, ten="File Excel hạn mức")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "File không hợp lệ — vui lòng chọn đúng file Excel (.xlsx)")
    ws = wb.active

    try:
        col_map = _qi_detect_columns(ws)
        if not col_map:
            raise HTTPException(
                400,
                "Không tìm thấy dòng tiêu đề hợp lệ trong file — cần có tối thiểu các cột "
                "STT, Họ và tên, Hạn mức (còn Mã cán bộ, Phòng, Đã nghỉ nếu có).",
            )

        staffs_by_code: dict = {}
        staffs_by_name: dict = {}
        staff_names: dict = {}
        for r in db.execute("SELECT id, employee_code, full_name FROM user_tttt WHERE is_active=1").fetchall():
            code = (r["employee_code"] or "").strip()
            if code:
                staffs_by_code[code] = r["id"]
            staff_names[r["id"]] = r["full_name"] or ""
            nm = _norm_vn(r["full_name"])
            if nm:
                staffs_by_name.setdefault(nm, []).append(r["id"])

        # Gộp truy vấn hạn mức/đã dùng hiện tại thành 1 lần thay vì mỗi dòng 1 query.
        old_quota_by_staff = {
            r["staff_id"]: float(r["quota_days"])
            for r in db.execute("SELECT staff_id, quota_days FROM leave_quotas WHERE year=?", (year,)).fetchall()
        }
        old_used_by_staff = {
            r["id"]: float(r["used_leave_days"]) if r["used_leave_days"] is not None else 0.0
            for r in db.execute("SELECT id, used_leave_days FROM user_tttt WHERE is_active=1").fetchall()
        }

        def _cell(row, key):
            idx = col_map.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        def _to_stt(v):
            """Chấp nhận STT dạng int, float nguyên (1.0), hoặc text số ("1")."""
            if isinstance(v, bool):
                return None
            if isinstance(v, int):
                return v
            if isinstance(v, float) and v.is_integer():
                return int(v)
            if isinstance(v, str) and v.strip().isdigit():
                return int(v.strip())
            return None

        rows_out = []
        for row in ws.iter_rows(values_only=True):
            stt = _to_stt(_cell(row, "stt"))
            if stt is None:
                continue
            ho_ten  = _cell(row, "ho_ten")
            ma_cb   = _cell(row, "ma_can_bo")
            phong   = _cell(row, "phong")
            han_muc = _cell(row, "han_muc")
            da_nghi = _cell(row, "da_nghi")
            if not (ho_ten and str(ho_ten).strip()) or han_muc is None:
                continue

            try:
                new_quota = float(str(han_muc).strip().replace(",", "."))
                new_used  = float(str(da_nghi).strip().replace(",", ".")) if da_nghi not in (None, "") else 0.0
            except (ValueError, TypeError):
                # Ô số liệu không hợp lệ — vẫn hiện dòng này để người dùng biết, nhưng
                # đánh dấu lỗi và không cho tick áp dụng (matched=False).
                rows_out.append({
                    "stt": stt, "ho_ten": str(ho_ten).strip(),
                    "ma_can_bo": str(ma_cb).strip() if ma_cb else "",
                    "phong": str(phong).strip() if phong else "",
                    "matched": False, "match_method": None, "staff_id": None,
                    "new_quota_days": 0, "new_used_leave_days": 0,
                    "old_quota_days": None, "old_used_leave_days": None,
                    "row_error": "Hạn mức / Đã nghỉ không phải số hợp lệ",
                    "rounded_warning": False,
                })
                continue

            ma_cb_s = str(ma_cb).strip() if ma_cb else ""
            staff_id = None
            match_method = None
            if ma_cb_s and ma_cb_s in staffs_by_code:
                staff_id = staffs_by_code[ma_cb_s]
                match_method = "ma_can_bo"
            else:
                cands = staffs_by_name.get(_norm_vn(ho_ten)) or []
                if len(cands) == 1:
                    staff_id = cands[0]
                    match_method = "ten"

            item = {
                "stt":                 stt,
                "ho_ten":              str(ho_ten).strip(),
                "ma_can_bo":           ma_cb_s,
                "phong":               str(phong).strip() if phong else "",
                "matched":             staff_id is not None,
                "match_method":        match_method,
                "matched_name":        staff_names.get(staff_id) if staff_id else None,
                "staff_id":            staff_id,
                "new_quota_days":      new_quota,
                "new_used_leave_days": new_used,
                "old_quota_days":      old_quota_by_staff.get(staff_id) if staff_id else None,
                "old_used_leave_days": old_used_by_staff.get(staff_id) if staff_id else None,
                # Hệ thống không có khái niệm "nửa ngày phép" — khi áp dụng, "Đã nghỉ"
                # có phần thập phân sẽ bị làm tròn lên nguyên ngày (xem import_quota_apply).
                # Cảnh báo trước để người nhập biết, tránh lệch 0.5 ngày âm thầm.
                "rounded_warning":     new_used != int(new_used),
            }
            rows_out.append(item)
    finally:
        wb.close()

    matched_count = sum(1 for r in rows_out if r["matched"])
    return {"filename": file.filename, "rows": rows_out, "total": len(rows_out), "matched": matched_count}


def _import_spread_dates(n_days: int, year: int) -> list:
    """Sinh n_days ngày làm việc (T2–T6) từ 02/01/year — dùng cho bản ghi nghỉ tổng hợp khi import."""
    out = []
    d = date(year, 1, 2)
    while len(out) < n_days and d.year == year:
        if d.weekday() < 5:
            out.append(d.isoformat())
        d += timedelta(days=1)
    return out


@router.post("/quotas/{year}/import/apply")
def import_quota_apply(
    year: int,
    body: dict,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Áp dụng dữ liệu hạn mức đã xem trước.

    - quota_days → ghi vào leave_quotas.
    - "Đã nghỉ" → tạo bản ghi nghỉ bat_buoc approved tổng hợp (nguồn sự thật cho
      _calc_used_days), thay vì chỉ ghi used_leave_days (trường này không được đọc
      khi tính hạn mức). Import lại sẽ thay bản ghi tổng hợp cũ (idempotent).
    - Giá trị cũ + id bản ghi tạo ra lưu vào quota_import_items để hoàn tác."""
    rows = body.get("rows") or []
    filename = body.get("filename") or ""
    rows = [r for r in rows if r.get("staff_id")]
    if not rows:
        raise HTTPException(400, "Không có dòng nào khớp nhân viên để áp dụng")

    for r in rows:
        try:
            qd, ud = float(r.get("new_quota_days") or 0), float(r.get("new_used_leave_days") or 0)
        except (ValueError, TypeError):
            raise HTTPException(400, "Dữ liệu hạn mức/đã nghỉ không hợp lệ")
        if qd < 0 or ud < 0:
            raise HTTPException(400, "Hạn mức và số ngày đã nghỉ không được âm")

    staff_ids = [r["staff_id"] for r in rows]
    placeholders = ",".join("?" * len(staff_ids))
    active_staff_ids = {
        r["id"] for r in db.execute(
            f"SELECT id FROM user_tttt WHERE id IN ({placeholders}) AND is_active=1", staff_ids
        ).fetchall()
    }
    old_quota_by_staff = {
        r["staff_id"]: float(r["quota_days"])
        for r in db.execute(
            f"SELECT staff_id, quota_days FROM leave_quotas WHERE year=? AND staff_id IN ({placeholders})",
            [year] + staff_ids,
        ).fetchall()
    }
    old_used_by_staff = {
        r["id"]: float(r["used_leave_days"]) if r["used_leave_days"] is not None else 0.0
        for r in db.execute(
            f"SELECT id, used_leave_days FROM user_tttt WHERE id IN ({placeholders})", staff_ids
        ).fetchall()
    }

    now = _vn_now()
    cur = db.execute(
        "INSERT INTO quota_import_batches (year, filename, imported_by, imported_at, row_count) VALUES (?,?,?,?,?)",
        (year, filename, current["id"], now, len(rows)),
    )
    batch_id = cur.lastrowid

    # Xoá TRƯỚC toàn bộ bản ghi tổng hợp cũ (import HOẶC sửa tay trước đó) của
    # từng nhân viên trong đợt này, rồi mới tính _calc_used_days_bulk — để
    # real_used_by_staff phản ánh đúng phần đơn THẬT còn lại, không lẫn số của
    # lần import/sửa tay trước. Phải tách thành 2 lượt (xoá rồi mới tính hàng
    # loạt) thay vì tính trong cùng vòng lặp per-row như update_used_days, vì ở
    # đây xử lý nhiều nhân viên cùng lúc — tính bulk 1 lần rẻ hơn N lần gọi
    # _calc_used_days đơn lẻ.
    for staff_id in staff_ids:
        if staff_id not in active_staff_ids:
            continue
        db.execute(
            "DELETE FROM leave_records WHERE staff_id=? AND leave_type='bat_buoc' "
            "AND strftime('%Y', start_date)=? AND (reason LIKE '[Import]%' OR reason LIKE '[Điều chỉnh]%')",
            (staff_id, str(year)),
        )
    real_used_by_staff = _calc_used_days_bulk(staff_ids, year, db, include_pending=True)

    applied = 0
    capped_staff: list = []
    for r in rows:
        staff_id = r["staff_id"]
        if staff_id not in active_staff_ids:
            continue
        new_quota = float(r.get("new_quota_days") or 0)
        new_used  = float(r.get("new_used_leave_days") or 0)
        old_quota = old_quota_by_staff.get(staff_id)
        old_used  = old_used_by_staff.get(staff_id, 0.0)

        # ── Hạn mức ──
        db.execute(
            "INSERT INTO leave_quotas (staff_id, year, quota_days) VALUES (?,?,?) "
            "ON CONFLICT(staff_id, year) DO UPDATE SET quota_days=excluded.quota_days",
            (staff_id, year, new_quota),
        )

        # ── "Đã nghỉ" → bản ghi nghỉ tổng hợp (để _calc_used_days đếm được) ──
        # new_used là TỔNG mong muốn (số trong file Excel), không phải số ngày
        # cộng thêm — trừ đi phần đơn THẬT (real_used, đã tính ở trên sau khi
        # xoá bản ghi tổng hợp cũ) rồi mới chèn đúng phần chênh lệch. Thiếu bước
        # trừ này thì nhân viên đã có đơn thật trong năm sẽ bị cộng dồn sai
        # ngay từ lần import đầu tiên (giống lỗi đã sửa ở update_used_days).
        real_used = real_used_by_staff.get(staff_id, 0.0)
        delta = new_used - real_used
        # int(round()) dùng banker's rounding (4.5→4, 5.5→6) — không nhất quán.
        # delta có thể âm nên chặn >=0 trước khi +0.5 làm tròn half-up.
        n_days = int(delta + 0.5) if delta > 0 else 0
        if delta < 0:
            capped_staff.append(r.get("matched_name") or r.get("ho_ten") or f"#{staff_id}")
        created_leave_id = None
        if n_days >= 1:
            _sd = _import_spread_dates(n_days, year)
            if _sd:
                _c = db.execute(
                    """INSERT INTO leave_records
                           (staff_id, leave_type, start_date, end_date, spread_dates,
                            status, reason, created_at, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (staff_id, "bat_buoc", _sd[0], _sd[-1], json.dumps(_sd), "approved",
                     f"[Import] Tổng hợp {n_days} ngày đã nghỉ năm {year} (batch #{batch_id})",
                     now, now),
                )
                created_leave_id = _c.lastrowid

        final_used = real_used + n_days

        # ── Lưu item để hoàn tác (kèm id bản ghi vừa tạo) ──
        db.execute(
            """INSERT INTO quota_import_items
                   (batch_id, staff_id, old_quota_days, old_used_leave_days,
                    new_quota_days, new_used_leave_days, created_leave_id)
               VALUES (?,?,?,?,?,?,?)""",
            (batch_id, staff_id, old_quota, old_used, new_quota, final_used, created_leave_id),
        )
        # Trường cache — hiển thị hạn mức không đọc, giữ đồng bộ cho tương thích.
        db.execute("UPDATE user_tttt SET used_leave_days=? WHERE id=?", (final_used, staff_id))
        applied += 1

    db.execute("UPDATE quota_import_batches SET matched_count=? WHERE id=?", (applied, batch_id))
    db.commit()
    return {
        "batch_id": batch_id, "applied": applied,
        # Nhân viên có "Đã nghỉ" trong file thấp hơn số ngày đơn thật đã ghi
        # nhận — không thể trừ bớt đơn thật bằng bản ghi tổng hợp, kết quả bị
        # giữ ở mức thật, báo cho frontend biết thay vì im lặng lệch số.
        "capped_staff": capped_staff,
    }


@router.get("/quotas/import/history")
def import_quota_history(
    year: Optional[int] = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Lịch sử các lần nhập file hạn mức — dùng để hoàn tác."""
    clause = "WHERE b.year=?" if year else ""
    params = (year,) if year else ()
    rows = db.execute(
        f"""SELECT b.*, u.full_name AS imported_by_name, rb.full_name AS rolled_back_by_name
            FROM quota_import_batches b
            LEFT JOIN user_tttt u  ON b.imported_by     = u.id
            LEFT JOIN user_tttt rb ON b.rolled_back_by  = rb.id
            {clause}
            ORDER BY b.imported_at DESC""",
        params,
    ).fetchall()
    return [dict(r) for r in rows]


@router.post("/quotas/import/{batch_id}/rollback")
def import_quota_rollback(
    batch_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Hoàn tác 1 lần nhập file — khôi phục quota_days, xoá bản ghi nghỉ tổng hợp đã
    tạo, và khôi phục used_leave_days về giá trị trước khi nhập. Nếu có lần nhập sau
    đó cũng đổi cùng nhân viên, giá trị sẽ bị ghi đè theo lần hoàn tác này (không tự
    động dồn nhiều lần hoàn tác)."""
    batch = db.execute("SELECT * FROM quota_import_batches WHERE id=?", (batch_id,)).fetchone()
    if not batch:
        raise HTTPException(404, "Không tìm thấy lần nhập")
    if batch["status"] == "rolled_back":
        raise HTTPException(400, "Lần nhập này đã được hoàn tác trước đó")

    items = db.execute("SELECT * FROM quota_import_items WHERE batch_id=?", (batch_id,)).fetchall()
    skipped_superseded = 0
    for it in items:
        # Nếu nhân viên này đã có lần nhập KHÁC (chưa hoàn tác) mới hơn cùng năm
        # đè lên sau batch này, thì hoàn tác batch cũ sẽ ghi sai giá trị (batch
        # mới mới là đúng) — bỏ qua nhân viên này, chỉ hoàn tác cho ai chưa bị
        # đè bởi lần nhập nào mới hơn.
        _newer = db.execute(
            """SELECT 1 FROM quota_import_items qi
               JOIN quota_import_batches b2 ON b2.id = qi.batch_id
               WHERE qi.staff_id = ? AND b2.year = ? AND b2.id != ?
                 AND b2.status != 'rolled_back'
                 AND (b2.imported_at > ? OR (b2.imported_at = ? AND b2.id > ?))
               LIMIT 1""",
            (it["staff_id"], batch["year"], batch_id,
             batch["imported_at"], batch["imported_at"], batch_id),
        ).fetchone()
        if _newer:
            skipped_superseded += 1
            continue
        if it["old_quota_days"] is not None:
            db.execute(
                "INSERT INTO leave_quotas (staff_id, year, quota_days) VALUES (?,?,?) "
                "ON CONFLICT(staff_id, year) DO UPDATE SET quota_days=excluded.quota_days",
                (it["staff_id"], batch["year"], it["old_quota_days"]),
            )
        else:
            db.execute(
                "DELETE FROM leave_quotas WHERE staff_id=? AND year=?", (it["staff_id"], batch["year"])
            )
        # Xoá bản ghi nghỉ tổng hợp mà lần nhập này đã tạo (nếu có)
        _clid = it["created_leave_id"] if "created_leave_id" in it.keys() else None
        if _clid:
            db.execute("DELETE FROM leave_records WHERE id=?", (_clid,))
        if it["old_used_leave_days"] is not None:
            db.execute(
                "UPDATE user_tttt SET used_leave_days=? WHERE id=?",
                (it["old_used_leave_days"], it["staff_id"]),
            )

    db.execute(
        "UPDATE quota_import_batches SET status='rolled_back', rolled_back_by=?, rolled_back_at=? WHERE id=?",
        (current["id"], str(_vn_now()), batch_id),
    )
    db.commit()
    return {
        "ok": True, "batch_id": batch_id,
        "restored": len(items) - skipped_superseded,
        "skipped_superseded": skipped_superseded,
    }


@router.get("/quotas/{year}/export")
def export_quotas(
    year: int,
    ids: str = "",   # staff_id cách nhau bởi dấu phẩy
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.quota_admin")),
):
    """Xuất bảng hạn mức phép năm ra Excel."""
    import openpyxl, io
    from openpyxl.styles import Alignment, Font, PatternFill
    from fastapi.responses import Response

    id_filter = {int(i.strip()) for i in ids.split(",") if i.strip().isdigit()} if ids else None

    staffs = db.execute(
        """SELECT u.id, u.full_name, u.employee_code, u.join_industry_date, d.name AS dept_name
           FROM user_tttt u
           LEFT JOIN departments d ON u.department_id = d.id
           WHERE u.is_active=1 AND (u.is_deleted=0 OR u.is_deleted IS NULL)
           ORDER BY d.name, u.full_name"""
    ).fetchall()

    staffs = [s for s in staffs if not id_filter or s["id"] in id_filter]
    staff_ids = [s["id"] for s in staffs]
    # Gộp truy vấn thay vì N+1 (mỗi nhân viên trước đây ~4-9 query riêng).
    quota_by_staff = {
        r["staff_id"]: float(r["quota_days"])
        for r in db.execute(
            f"SELECT staff_id, quota_days FROM leave_quotas WHERE year=? AND staff_id IN "
            f"({','.join('?' * len(staff_ids))})", [year] + staff_ids
        ).fetchall()
    } if staff_ids else {}
    carry_by_staff = _carry_over_bulk(staff_ids, year, db, effective=True)
    # include_pending=True: khớp đúng enforcement thật lúc tạo đơn.
    used_by_staff  = _calc_used_days_bulk(staff_ids, year, db, include_pending=True)

    data = []
    for s in staffs:
        quota  = quota_by_staff.get(s["id"], float(compute_annual_leave(s["join_industry_date"], year)))
        # carry-over hết hiệu lực sau Q1 (31/3) — cả cột hiển thị lẫn "remaining" đều dùng
        # cùng một giá trị "còn hiệu lực" để nhất quán với enforcement lúc tạo đơn.
        carry  = carry_by_staff.get(s["id"], 0.0)
        used   = used_by_staff.get(s["id"], 0.0)
        join_date = s["join_industry_date"] or ""
        data.append({"name": s["full_name"], "dept": s["dept_name"] or "",
                     "join_date": join_date,
                     "quota": quota, "carry": carry, "used": used,
                     "remaining": max(0.0, quota + carry - used)})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Han muc phep {year}"

    # Dòng tiêu đề năm
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"BẢNG HẠN MỨC NGHỈ PHÉP NĂM {year}")
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="8B0000")
    hdr_font    = Font(color="FFFFFF", bold=True)
    headers     = ["STT", "Họ và tên", "Phòng ban", "Ngày vào ngành", "Hạn mức", "Chuyển kỳ", "Đã dùng", "Ngày phép của năm"]
    col_widths  = [6, 30, 30, 18, 12, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font, cell.fill = hdr_font, header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w
    for ri, r in enumerate(data, 3):
        ws.cell(ri, 1, ri - 2).alignment = Alignment(horizontal="center")
        ws.cell(ri, 2, r["name"]).alignment = Alignment(horizontal="left")
        ws.cell(ri, 3, r["dept"]).alignment = Alignment(horizontal="left")
        ws.cell(ri, 4, r["join_date"]).alignment = Alignment(horizontal="center")
        for ci, val in enumerate([r["quota"], r["carry"], r["used"]], 5):
            ws.cell(ri, ci, round(val, 1)).alignment = Alignment(horizontal="center")
        cell_rem = ws.cell(ri, 8, round(r["remaining"], 1))
        cell_rem.font = Font(color="157A3A" if r["remaining"] > 0 else "CC0000", bold=True)
        cell_rem.alignment = Alignment(horizontal="center")
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="han_muc_phep_{year}.xlsx"'},
    )


@router.get("/export/annual")
def export_all_leaves_annual(
    year: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.stats_export")),
):
    """Xuất tất cả đơn nghỉ phép trong năm ra Excel."""
    import openpyxl, io
    from openpyxl.styles import Alignment, Font, PatternFill
    from fastapi.responses import Response

    # feature leaves.stats_export chỉ gate theo group_features, không tự ràng buộc
    # phòng ban — nếu admin lỡ gán quyền này cho nhóm không phải Tổng hợp/lãnh đạo,
    # phải tự lọc ở đây để không lộ dữ liệu toàn trung tâm (khớp leader_dashboard).
    _scope_required = (current["role"] not in ("admin", "giam_doc", "pho_giam_doc")
                       and not _is_tong_hop_staff(current, db))
    _dept_sql    = " AND s.department_id = ?" if _scope_required else ""
    _dept_params = [current.get("department_id")] if _scope_required else []

    rows = db.execute(
        f"""SELECT lr.*, s.full_name AS staff_name, d.name AS dept_name,
                  kv.full_name AS ksv_name, th.full_name AS th_name,
                  gd.full_name AS gd_name
           FROM leave_records lr
           LEFT JOIN user_tttt s  ON lr.staff_id = s.id
           LEFT JOIN departments d ON s.department_id = d.id
           LEFT JOIN user_tttt kv ON lr.ksv_approver_id = kv.id
           LEFT JOIN user_tttt th ON lr.tong_hop_approver_id = th.id
           LEFT JOIN user_tttt gd ON lr.gd_approver_id = gd.id
           WHERE strftime('%Y', lr.start_date) = ?{_dept_sql}
           ORDER BY lr.created_at DESC""",
        [str(year)] + _dept_params
    ).fetchall()

    import json as _json

    def _fmt_date(d_str):
        """YYYY-MM-DD → DD/MM/YYYY"""
        try:
            y, mo, dd = d_str[:10].split("-")
            return f"{dd}/{mo}/{y}"
        except Exception:
            return d_str or ""

    _STATUS_VN = {
        "pending_ksv": "Chờ KSV duyệt", "pending_tong_hop": "Chờ Tổng hợp",
        "pending_gd": "Chờ Ban lãnh đạo duyệt", "approved": "Hoàn thành",
        "rejected": "Từ chối", "cancelled": "Đã hủy"
    }
    headers = ["STT", "Ngày tạo", "Họ và tên", "Phòng", "Loại nghỉ", "Ngày nghỉ",
               "Số ngày", "Trạng thái", "KSV duyệt", "Phòng Tổng hợp", "GĐ/PGĐ"]
    widths  = [6, 14, 28, 28, 18, 40, 10, 18, 22, 22, 20]
    hfill = PatternFill("solid", fgColor="8B0000")
    hfont = Font(bold=True, color="FFFFFF")

    # Sắp xếp theo ngày bắt đầu tăng dần
    rows_sorted = sorted(rows, key=lambda r: r["start_date"] or "")

    # "Số ngày" phải khớp cách tính dùng để trừ hạn mức (_period_days: chỉ tính
    # ngày làm việc, trừ T7/CN/lễ — trừ thai_san/bao_hiem tính ngày lịch), không
    # phải đếm ngày lịch thô như trước.
    _holidays_yr = _load_holidays(db, date(year, 1, 1), date(year, 12, 31))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Don nghi phep {year}"

    ws.merge_cells("A1:L1")
    tc = ws.cell(1, 1, f"TẤT CẢ ĐƠN NGHỈ PHÉP NĂM {year}")
    tc.font = Font(bold=True, size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(2, ci, h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for idx, r in enumerate(rows_sorted, 1):
        try:
            nd = len(_json.loads(r["spread_dates"])) if r["spread_dates"] else _period_days(
                date.fromisoformat(r["start_date"]), date.fromisoformat(r["end_date"]),
                _holidays_yr, r["leave_type"],
            )
        except Exception:
            nd = 1
        ri = idx + 2
        ws.cell(ri, 1, idx).alignment = Alignment(horizontal="center")
        # Ngày nghỉ: ghi từng ngày riêng nếu không liên nhau
        if r["spread_dates"]:
            try:
                spread = _json.loads(r["spread_dates"])
                dates_str = ", ".join(_fmt_date(d) for d in sorted(spread))
            except Exception:
                dates_str = _fmt_date(r["start_date"])
        else:
            dates_str = f"{_fmt_date(r['start_date'])} → {_fmt_date(r['end_date'])}"

        ws.cell(ri, 2, _fmt_date((r["created_at"] or "")[:10])).alignment = Alignment(horizontal="center")
        ws.cell(ri, 3, r["staff_name"] or "").alignment = Alignment(horizontal="left")
        ws.cell(ri, 4, r["dept_name"] or "").alignment = Alignment(horizontal="left")
        ws.cell(ri, 5, LEAVE_TYPE_LABELS.get(r["leave_type"] or "", r["leave_type"] or ""))
        ws.cell(ri, 6, dates_str).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(ri, 7, nd).alignment = Alignment(horizontal="center")
        ws.cell(ri, 8, _STATUS_VN.get(r["status"] or "", r["status"] or ""))
        ws.cell(ri, 9, r["ksv_name"] or "")
        ws.cell(ri, 10, r["th_name"] or "")
        ws.cell(ri, 11, r["gd_name"] or "")

    buf = io.BytesIO(); wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="bao_cao_don_nghi_phep_{year}.xlsx"'},
    )


# ─── Báo cáo năm ────────────────────────────────────────────────────────────

@router.get("/stats/annual/{year}")
def stats_annual(
    year: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.stats_export")),
):
    """Tổng hợp phép năm — số ngày quota, carry-over, đã dùng, còn lại theo từng nhân viên."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    # Cùng lý do như export_all_leaves_annual: feature leaves.stats_export không tự
    # ràng buộc phòng ban — lọc thủ công để tránh lộ dữ liệu toàn trung tâm.
    _scope_required = (current["role"] not in ("admin", "giam_doc", "pho_giam_doc")
                       and not _is_tong_hop_staff(current, db))
    _dept_sql    = " AND u.department_id = ?" if _scope_required else ""
    _dept_params = [current.get("department_id")] if _scope_required else []

    staffs = db.execute(
        f"""SELECT u.id, u.full_name, u.employee_code, u.join_industry_date, d.name AS dept_name
           FROM user_tttt u
           LEFT JOIN departments d ON u.department_id = d.id
           WHERE u.is_active=1 AND (u.is_deleted=0 OR u.is_deleted IS NULL){_dept_sql}
           ORDER BY d.name, u.full_name""",
        _dept_params
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tổng hợp phép {year}"

    # Dòng tiêu đề năm
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"BÁO CÁO TỔNG HỢP NGHỈ PHÉP NĂM {year}")
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    hdr_fill = PatternFill("solid", fgColor="8B0000")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["STT", "Họ và tên", "Phòng ban", "Ngày vào ngành", "Hạn mức", "Chuyển kỳ", "Đã dùng", "Ngày phép của năm"]
    widths  = [6, 28, 28, 18, 12, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    staff_ids = [s["id"] for s in staffs]
    # Gộp truy vấn thay vì N+1.
    quota_by_staff = {
        r["staff_id"]: float(r["quota_days"])
        for r in db.execute(
            f"SELECT staff_id, quota_days FROM leave_quotas WHERE year=? AND staff_id IN "
            f"({','.join('?' * len(staff_ids))})", [year] + staff_ids
        ).fetchall()
    } if staff_ids else {}
    carry_by_staff = _carry_over_bulk(staff_ids, year, db, effective=True)
    # include_pending=True: khớp đúng enforcement thật lúc tạo đơn.
    used_by_staff  = _calc_used_days_bulk(staff_ids, year, db, include_pending=True)

    for idx, s in enumerate(staffs, 1):
        quota = quota_by_staff.get(s["id"], float(compute_annual_leave(s["join_industry_date"], year)))
        # carry-over hết hiệu lực sau Q1 (31/3) — cả cột hiển thị lẫn "remaining" đều dùng
        # cùng một giá trị "còn hiệu lực" để nhất quán với enforcement lúc tạo đơn.
        carry     = carry_by_staff.get(s["id"], 0.0)
        used      = used_by_staff.get(s["id"], 0.0)
        remaining = max(0.0, quota + carry - used)
        ri = idx + 2
        ws.cell(ri, 1, idx).alignment = Alignment(horizontal="center")
        ws.cell(ri, 2, s["full_name"]).alignment = Alignment(horizontal="left")
        ws.cell(ri, 3, s["dept_name"] or "").alignment = Alignment(horizontal="left")
        ws.cell(ri, 4, s["join_industry_date"] or "").alignment = Alignment(horizontal="center")
        for ci, val in enumerate([quota, carry, used], 5):
            ws.cell(ri, ci, round(val, 1)).alignment = Alignment(horizontal="center")
        cell_rem = ws.cell(ri, 8, round(remaining, 1))
        cell_rem.font = Font(color="157A3A" if remaining > 0 else "CC0000", bold=True)
        cell_rem.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''bao_cao_phep_{year}.xlsx"},
    )


# ─── Dashboard lãnh đạo ─────────────────────────────────────────────────────

@router.get("/stats/leader-dashboard")
def leader_dashboard(
    year: int = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.dashboard")),
):
    """Thống kê tổng quan: số đơn theo trạng thái, top nghỉ nhiều, chờ duyệt."""
    from datetime import date as _date
    yr = year or _date.today().year

    # Đơn chờ duyệt — filter theo role, giống list_leaves(scope="pending")
    role = current["role"]
    p_where = ""
    p_params: list = []
    # Đơn của GĐ đã tự động approved nhưng Tổng hợp chưa "biết" — giống hệt
    # _gd_unack trong list_leaves(scope="pending"), thiếu nhánh này khiến widget
    # "Chờ KSV/Chờ Tổng hợp" trên dashboard không khớp số với tab Chờ xác nhận TT.
    _gd_unack = (
        "(lr.status = 'approved' AND lr.tong_hop_approver_id IS NULL "
        "AND lr.staff_id IN (SELECT id FROM user_tttt WHERE role = 'giam_doc'))"
    )
    if role == "admin":
        p_where = f"(lr.status IN ('pending_ksv','pending_tong_hop','pending_gd') OR {_gd_unack})"
    elif role in ("giam_doc", "pho_giam_doc"):
        if _can_gd_review(current, db):
            p_where = "lr.gd_approver_id = ? AND lr.status = 'pending_gd'"
            p_params.append(current["id"])
    elif _is_tong_hop_staff(current, db) and role in ("truong_phong", "pho_phong"):
        # PP/TP Tổng hợp: xem cả KSV phòng mình + TH/GĐ toàn trung tâm — khớp đúng
        # phạm vi "toàn trung tâm" đã áp dụng cho approved/top_staff bên dưới
        # (_scope_required). Thiếu pending_gd ở đây khiến ô "Chờ GĐ" luôn hiện 0
        # dù đơn đang thật sự chờ GĐ (nhìn thấy trong bảng danh sách bên dưới).
        p_where = (f"(lr.status IN ('pending_tong_hop', 'pending_gd') "
                   f"OR (lr.ksv_approver_id = ? AND lr.status = 'pending_ksv') OR {_gd_unack})")
        p_params.append(current["id"])
    elif _is_tong_hop_staff(current, db):
        p_where = f"(lr.status IN ('pending_tong_hop', 'pending_gd') OR {_gd_unack})"
    elif role in ("truong_phong", "pho_phong"):
        p_where = "lr.ksv_approver_id = ? AND lr.status = 'pending_ksv'"
        p_params.append(current["id"])

    if p_where:
        pending_rows = db.execute(
            f"""SELECT lr.id, s.full_name, lr.status, lr.start_date, lr.end_date, lr.leave_type
               FROM leave_records lr
               JOIN user_tttt s ON lr.staff_id = s.id
               WHERE {p_where}
               ORDER BY lr.created_at""",
            p_params,
        ).fetchall()
    else:
        pending_rows = []

    # Số liệu tổng quan (đã duyệt, khai báo hộ, top nghỉ nhiều) chỉ tính người
    # cùng phòng, TRỪ admin / Giám đốc / PGĐ / Phòng Tổng hợp — những vai trò
    # này thấy toàn trung tâm. Áp dụng cho mọi role còn lại (kể cả chuyên viên).
    # Dùng cờ riêng để biết CÓ cần scope hay không — không được suy ra từ giá trị
    # department_id, vì nếu user chưa gán phòng (department_id NULL) thì falsy
    # sẽ vô tình tắt luôn bộ lọc và lộ số liệu toàn trung tâm cho họ.
    _scope_required = (role not in ("admin", "giam_doc", "pho_giam_doc")
                       and not _is_tong_hop_staff(current, db))
    if _scope_required:
        _dept_sql    = " AND s.department_id = ?"
        # department_id NULL → "= NULL" không khớp dòng nào (đúng ý: ẩn hết thay
        # vì lộ toàn trung tâm) — an toàn hơn là bỏ lọc.
        _dept_params = [current.get("department_id")]
    else:
        _dept_sql    = ""
        _dept_params = []

    # by_status: đếm từ danh sách pending đã filter + approved trong năm
    by_status: dict = {}
    for r in pending_rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    # Bản ghi tổng hợp giả (nhập Excel/sửa tay hạn mức) không phải đơn nghỉ phép
    # thật — loại khỏi "Hoàn thành"/"Khai báo hộ" để khớp với list_leaves() (đã ẩn
    # tương tự) và không đếm 1 thao tác đối soát hạn mức như 1 đơn đã hoàn thành.
    _not_synthetic = "AND NOT (lr.reason LIKE '[Import]%' OR lr.reason LIKE '[Điều chỉnh]%')"
    approved_cnt = db.execute(
        f"""SELECT COUNT(*) FROM leave_records lr JOIN user_tttt s ON lr.staff_id = s.id
           WHERE lr.status='approved' AND strftime('%Y', lr.start_date)=?{_dept_sql} {_not_synthetic}""",
        [str(yr)] + _dept_params,
    ).fetchone()[0]
    by_status["approved"] = approved_cnt
    by_status["direct"] = db.execute(
        f"""SELECT COUNT(*) FROM leave_records lr JOIN user_tttt s ON lr.staff_id = s.id
           WHERE lr.is_direct=1 AND lr.status='approved' AND strftime('%Y', lr.start_date)=?{_dept_sql} {_not_synthetic}""",
        [str(yr)] + _dept_params,
    ).fetchone()[0]
    pending = [
        {
            "id": r["id"], "staff_name": r["full_name"],
            "status": r["status"], "status_label": _LEAVE_STATUS_VN.get(r["status"], r["status"]),
            "start_date": r["start_date"], "end_date": r["end_date"],
            "leave_type": LEAVE_TYPE_LABELS.get(r["leave_type"], r["leave_type"]),
        }
        for r in pending_rows
    ]

    # Top 10 nhân viên nghỉ nhiều nhất trong năm (approved) — tính "Số ngày" bằng
    # _period_days (ngày làm việc, trừ T7/CN/lễ) giống mọi nơi khác trong hệ
    # thống, không đếm ngày lịch thô (julianday) như trước.
    top_raw = db.execute(
        f"""SELECT lr.staff_id, s.full_name, lr.leave_type, lr.spread_dates,
                  lr.start_date, lr.end_date
           FROM leave_records lr
           JOIN user_tttt s ON lr.staff_id = s.id
           WHERE lr.status='approved' AND strftime('%Y', lr.start_date)=?{_dept_sql} {_not_synthetic}""",
        [str(yr)] + _dept_params,
    ).fetchall()
    _holidays_yr = _load_holidays(db, date(yr, 1, 1), date(yr, 12, 31))
    _top_totals: dict = {}
    _top_names: dict = {}
    for r in top_raw:
        if r["spread_dates"]:
            nd = len(json.loads(r["spread_dates"]))
        else:
            nd = _period_days(
                date.fromisoformat(r["start_date"]), date.fromisoformat(r["end_date"]),
                _holidays_yr, r["leave_type"],
            )
        _top_totals[r["staff_id"]] = _top_totals.get(r["staff_id"], 0) + nd
        _top_names[r["staff_id"]] = r["full_name"]
    top_staff = [
        {"staff_name": _top_names[sid], "total_days": total}
        for sid, total in sorted(_top_totals.items(), key=lambda kv: kv[1], reverse=True)[:10]
    ]

    return {
        "year": yr,
        "by_status": by_status,
        "pending": pending,
        "top_staff": top_staff,
    }


# ─── Khai báo hộ (direct leave) ──────────────────────────────────────────────

@router.post("/direct")
def create_direct_leave(
    body: DirectLeaveCreate,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.declare_direct")),
):
    """Khai báo hộ — tạo đơn approved ngay, cộng used_leave_days thủ công."""
    staff = db.execute(
        "SELECT * FROM user_tttt WHERE id=? AND is_active=1", (body.staff_id,)
    ).fetchone()
    if not staff:
        raise HTTPException(404, "Nhân viên không tồn tại")

    if body.leave_type not in _VALID_LEAVE_TYPES:
        raise HTTPException(400, f"Loại nghỉ phép không hợp lệ: {body.leave_type}")

    if body.spread_dates:
        spread = sorted(set(body.spread_dates))
        if len(spread) < 1:
            raise HTTPException(400, "spread_dates phải có ít nhất 1 ngày")
        try:
            eff_start  = date.fromisoformat(spread[0])
            eff_end    = date.fromisoformat(spread[-1])
        except ValueError:
            raise HTTPException(400, "Định dạng ngày không hợp lệ (yêu cầu YYYY-MM-DD)")
        leave_days = len(spread)
        spread_json = json.dumps(spread)
    else:
        if body.end_date < body.start_date:
            raise HTTPException(400, "Ngày kết thúc phải sau ngày bắt đầu")
        eff_start   = body.start_date
        eff_end     = body.end_date
        _h          = _load_holidays(db, eff_start, eff_end)
        leave_days  = _period_days(eff_start, eff_end, _h, body.leave_type)
        spread_json = None

    # Kiểm tra trùng ngày với đơn đã tồn tại (kể cả khai báo hộ và đơn thường)
    all_dates = json.loads(spread_json) if spread_json else [
        (eff_start + __import__('datetime').timedelta(days=i)).isoformat()
        for i in range((eff_end - eff_start).days + 1)
    ]
    existing = db.execute(
        """SELECT lr.id, lr.start_date, lr.end_date, lr.spread_dates
           FROM leave_records lr
           WHERE lr.staff_id=? AND lr.status NOT IN ('cancelled','rejected')
             AND NOT (lr.reason LIKE '[Import]%' OR lr.reason LIKE '[Điều chỉnh]%')""",
        (body.staff_id,)
    ).fetchall()
    conflict_dates = []
    for ex in existing:
        ex_dates = json.loads(ex["spread_dates"]) if ex["spread_dates"] else [
            (date.fromisoformat(ex["start_date"]) + __import__('datetime').timedelta(days=i)).isoformat()
            for i in range((date.fromisoformat(ex["end_date"]) - date.fromisoformat(ex["start_date"])).days + 1)
        ]
        overlap = set(all_dates) & set(ex_dates)
        if overlap:
            conflict_dates.extend(sorted(overlap))
    if conflict_dates:
        conflict_str = ", ".join(sorted(set(conflict_dates))[:5])
        raise HTTPException(409, f"Nhân viên đã có đơn nghỉ vào ngày: {conflict_str}. Vui lòng kiểm tra lại.")

    # Kiểm tra hạn mức (không áp dụng cho thai_san, bao_hiem)
    if body.leave_type not in _NO_QUOTA_TYPES:
        # sqlite3.Row không có .get() — (row or {}).get(...) crash 500 nếu đã có
        # dòng leave_quotas (kể cả quota_days=0). Ưu tiên leave_quotas (nhập tay/
        # upload file), chỉ fallback công thức ngày vào ngành khi chưa có.
        _q_row = db.execute(
            "SELECT quota_days FROM leave_quotas WHERE staff_id=? AND year=?",
            (body.staff_id, eff_start.year),
        ).fetchone()
        quota = (
            float(_q_row["quota_days"]) if _q_row
            else float(compute_annual_leave(staff["join_industry_date"], eff_start.year))
        )
        carry     = compute_carry_over(body.staff_id, eff_start.year, db, effective=True, ref_date=eff_start)
        used      = _calc_used_days(body.staff_id, eff_start.year, db, include_pending=True)
        remaining = quota + carry - used
        if leave_days > remaining:
            raise HTTPException(400,
                f"Vượt quá hạn mức phép năm {eff_start.year}. "
                f"Còn lại {remaining:.0f} ngày, khai báo {leave_days} ngày.")

    cur = db.execute(
        """INSERT INTO leave_records
               (staff_id, start_date, end_date, leave_type, reason, status,
                is_direct, direct_by, spread_dates, created_at, updated_at)
           VALUES (?,?,?,?,?,'approved',1,?,?,?,?)""",
        (body.staff_id, eff_start.isoformat(), eff_end.isoformat(),
         body.leave_type, body.reason,
         current["id"], spread_json, str(_vn_now()), str(_vn_now())),
    )
    leave_id = cur.lastrowid
    if body.leave_type not in _NO_QUOTA_TYPES:
        db.execute(
            "UPDATE user_tttt SET used_leave_days = COALESCE(used_leave_days,0) + ? WHERE id=?",
            (leave_days, body.staff_id),
        )
    _log_action(db, leave_id, current["id"], "direct_create", None, "", LeaveStatus.APPROVED)
    db.commit()
    return _leave_to_out(leave_id, db)


# ─── Recall (rút đơn đã duyệt) ───────────────────────────────────────────────

@router.post("/{leave_id}/recall")
def request_recall(
    leave_id: int,
    body: RecallCreate,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.recall")),
):
    """Yêu cầu rút đơn đã duyệt — ghi recall_reason, chuyển sang pending_tong_hop để xác nhận."""
    leave = db.execute("SELECT * FROM leave_records WHERE id=?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn")
    if leave["status"] != LeaveStatus.APPROVED:
        raise HTTPException(400, "Chỉ rút được đơn đã duyệt")
    if leave["staff_id"] != current["id"] and current["role"] != "admin":
        raise HTTPException(403, "Chỉ chủ nhân đơn hoặc Admin mới được yêu cầu rút")
    if not body.reason:
        raise HTTPException(400, "Vui lòng nhập lý do rút đơn")

    old = leave["status"]
    db.execute(
        "UPDATE leave_records SET recall_reason=?, status=?, updated_at=? WHERE id=?",
        (body.reason, LeaveStatus.PENDING_TONG_HOP, str(_vn_now()), leave_id),
    )
    _log_action(db, leave_id, current["id"], "recall_request", body.reason, old, LeaveStatus.PENDING_TONG_HOP)
    db.commit()
    return _leave_to_out(leave_id, db)


@router.put("/{leave_id}/recall-approve")
def approve_recall(
    leave_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("leaves.recall")),
):
    """Phòng Tổng hợp xác nhận rút đơn — chuyển sang cancelled."""
    if not _is_tong_hop_staff(current, db) and current["role"] != "admin":
        raise HTTPException(403, "Chỉ Phòng Tổng hợp hoặc Admin mới xác nhận rút đơn")
    leave = db.execute("SELECT * FROM leave_records WHERE id=?", (leave_id,)).fetchone()
    if not leave:
        raise HTTPException(404, "Không tìm thấy đơn")

    # Đơn của Giám đốc: GĐ có toàn quyền với đơn của mình, Tổng hợp có thể xác nhận
    # rút/hủy bất cứ lúc nào sau khi tạo — không cần GĐ gọi request_recall trước
    # và không bắt buộc đơn phải đang ở pending_tong_hop.
    staff_row  = db.execute("SELECT role FROM user_tttt WHERE id=?", (leave["staff_id"],)).fetchone()
    is_gd_leave = bool(staff_row and staff_row["role"] == "giam_doc")

    if not is_gd_leave and leave["staff_id"] == current["id"] and current["role"] != "admin":
        raise HTTPException(403, "Không thể tự xác nhận rút đơn của chính mình")

    if is_gd_leave:
        if leave["status"] in (LeaveStatus.CANCELLED, LeaveStatus.REJECTED):
            raise HTTPException(400, "Đơn đã ở trạng thái kết thúc, không thể xác nhận rút")
    else:
        if leave["status"] != LeaveStatus.PENDING_TONG_HOP or not leave["recall_reason"]:
            raise HTTPException(400, "Đơn này không trong trạng thái chờ xác nhận rút")
    old = leave["status"]
    # Trừ used_leave_days khi xác nhận rút (approved → cancelled) — trừ thai_san/bao_hiem
    # vì loại này chưa từng được cộng vào used_leave_days lúc duyệt.
    if leave["leave_type"] not in _NO_QUOTA_TYPES:
        start = date.fromisoformat(leave["start_date"])
        end   = date.fromisoformat(leave["end_date"])
        _h    = _load_holidays(db, start, end)
        days  = len(json.loads(leave["spread_dates"])) if leave["spread_dates"] else _period_days(start, end, _h, leave["leave_type"])
        db.execute(
            "UPDATE user_tttt SET used_leave_days = MAX(0, COALESCE(used_leave_days,0) - ?) WHERE id=?",
            (days, leave["staff_id"]),
        )
    db.execute(
        "UPDATE leave_records SET status=?, updated_at=? WHERE id=?",
        (LeaveStatus.CANCELLED, str(_vn_now()), leave_id),
    )
    _log_action(db, leave_id, current["id"], "recall_approve", None, old, LeaveStatus.CANCELLED)
    db.commit()
    return _leave_to_out(leave_id, db)
