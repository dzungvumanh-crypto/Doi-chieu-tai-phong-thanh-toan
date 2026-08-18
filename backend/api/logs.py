"""Log viewer API — chỉ dành cho Admin"""
import io
import os
import re
import sqlite3
import tempfile
from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from backend.database import get_db, write_audit, _vn_now
from backend.core.config import settings
from backend.core.deps import require_admin, require_feature
from backend.core.net import client_ip as _client_ip

router = APIRouter()

_LOG_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "logs", "app.log",
)

_LOG_RE = re.compile(
    r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\s+(\w+)\s+(\S+)\s+—\s+(.*)$"
)

PAGE_SIZE = 50
# Thời hạn lưu nhật ký + việc dọn nằm ở backend/services/log_cleanup_service.py
# (chạy nền theo lịch, không dọn trong request xem nhật ký nữa)


def _parse_log_file(level_filter: str = "", page: int = 1):
    if not os.path.exists(_LOG_PATH):
        return [], 0

    with open(_LOG_PATH, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    parsed = []
    current = None
    for line in lines:
        m = _LOG_RE.match(line.rstrip())
        if m:
            if current:
                parsed.append(current)
            ts, level, logger, msg = m.groups()
            current = {"ts": ts, "level": level.upper(), "logger": logger, "msg": msg}
        else:
            stripped = line.strip()
            if current and stripped:
                current["msg"] += "\n" + stripped

    if current:
        parsed.append(current)

    parsed.reverse()

    if level_filter and level_filter.upper() not in ("ALL", ""):
        parsed = [e for e in parsed if e["level"] == level_filter.upper()]

    total = len(parsed)
    offset = (page - 1) * PAGE_SIZE
    return parsed[offset: offset + PAGE_SIZE], total


@router.get("/backup-info")
def get_backup_info(_: dict = Depends(require_feature("menu.logs"))):
    from backend.services.backup_service import last_backup_info
    return last_backup_info()


@router.get("/backup")
def backup_db(
    request: Request,
    current: dict = Depends(require_feature("menu.logs")),
    db: sqlite3.Connection = Depends(get_db),
):
    db_path = settings.DATABASE_URL.replace("sqlite:///", "")
    if not os.path.exists(db_path):
        from fastapi import HTTPException
        raise HTTPException(404, "Không tìm thấy file database")

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".db")
    os.close(tmp_fd)
    try:
        src = sqlite3.connect(db_path)
        dst = sqlite3.connect(tmp_path)
        src.backup(dst)
        dst.close()
        src.close()
        with open(tmp_path, "rb") as f:
            data = f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    # Ghi vào audit_logs, KHÔNG phải login_logs: tải backup không phải sự kiện
    # đăng nhập. Dòng cũ nằm trong login_logs với success=1 làm mọi thống kê
    # "số lượt đăng nhập" đếm dôi, và đẩy nhật ký đăng nhập thật ra khỏi trang
    # đầu. Đây là GET nên AuditMiddleware không đụng tới — phải tự ghi.
    # _client_ip(): frontend gọi backend qua loopback nên request.client.host
    # luôn là 127.0.0.1; IP thật của trình duyệt nằm ở X-Client-IP.
    stamp = _vn_now().strftime("%Y%m%d_%H%M%S")
    write_audit(db, current["id"], "db_backup_download", "system", None,
                f"Tải bản sao cơ sở dữ liệu ({stamp}) — file chứa TOÀN BỘ dữ liệu, gồm cả mã băm mật khẩu",
                _client_ip(request))
    db.commit()

    filename = f"ksnb_backup_{stamp}.db"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/logins/export")
def export_login_logs(
    success: str = Query(""),
    _: dict = Depends(require_feature("menu.logs")),
    db: sqlite3.Connection = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    clauses = []
    params: list = []
    if success == "true":
        clauses.append("ll.success = 1")
    elif success == "false":
        clauses.append("ll.success = 0")
    where = "WHERE " + " AND ".join(clauses) if clauses else ""

    rows = db.execute(
        f"""SELECT ll.*, ks.full_name
            FROM login_logs ll
            LEFT JOIN user_tttt ks ON ll.staff_id = ks.id
            {where}
            ORDER BY ll.created_at DESC""",
        params,
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhật ký đăng nhập"

    hdr_fill = PatternFill("solid", fgColor="37474F")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["STT", "Thời gian", "Kết quả", "Username", "Họ và tên", "IP", "Chi tiết"]
    widths  = [6, 18, 12, 20, 28, 18, 40]
    ws.append(headers)
    for cell, w in zip(ws[1], widths):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for idx, r in enumerate(rows, 1):
        ts = (r["created_at"] or "")[:16].replace("T", " ")
        ws.append([
            idx,
            ts,
            "Thành công" if r["success"] else "Thất bại",
            r["username"] or "",
            r["full_name"] or "",
            r["ip_address"] or "",
            r["detail"] or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date
    fname = f"nhat_ky_dang_nhap_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@router.get("/logins")
def get_login_logs(
    page:    int = Query(1, ge=1),
    success: str = Query(""),
    _: dict = Depends(require_feature("menu.logs")),
    db: sqlite3.Connection = Depends(get_db),
):
    clauses = []
    params: list = []
    if success == "true":
        clauses.append("ll.success = 1")
    elif success == "false":
        clauses.append("ll.success = 0")
    where = "WHERE " + " AND ".join(clauses) if clauses else ""

    total = db.execute(
        f"SELECT COUNT(*) FROM login_logs ll {where}", params
    ).fetchone()[0]

    offset = (page - 1) * PAGE_SIZE
    rows = db.execute(
        f"""SELECT ll.*, ks.full_name
            FROM login_logs ll
            LEFT JOIN user_tttt ks ON ll.staff_id = ks.id
            {where}
            ORDER BY ll.created_at DESC
            LIMIT ? OFFSET ?""",
        params + [PAGE_SIZE, offset],
    ).fetchall()

    return {
        "entries": [
            {
                "id":         r["id"],
                "username":   r["username"],
                "full_name":  r["full_name"],
                "ip_address": r["ip_address"],
                "success":    bool(r["success"]),
                "detail":     r["detail"],
                "created_at": r["created_at"],
            }
            for r in rows
        ],
        "total":     total,
        "page":      page,
        "page_size": PAGE_SIZE,
        "pages":     max(1, -(-total // PAGE_SIZE)),  # ceiling division
    }


def _audit_where(method: str, q: str):
    clauses, params = [], []
    if method and method.upper() in ("POST", "PUT", "PATCH", "DELETE"):
        clauses.append("al.action = ?")
        params.append(method.upper())
    if q.strip():
        like = f"%{q.strip()}%"
        clauses.append("(al.target_type LIKE ? OR al.detail LIKE ? OR ks.full_name LIKE ? OR ks.username LIKE ?)")
        params += [like, like, like, like]
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    return where, params


@router.get("/audit")
def get_audit_logs(
    page:   int = Query(1, ge=1),
    method: str = Query(""),
    q:      str = Query(""),
    _: dict = Depends(require_feature("menu.logs")),
    db: sqlite3.Connection = Depends(get_db),
):
    where, params = _audit_where(method, q)
    base = f"""FROM audit_logs al
               LEFT JOIN user_tttt ks ON al.actor_id = ks.id
               {where}"""

    total = db.execute(f"SELECT COUNT(*) {base}", params).fetchone()[0]
    offset = (page - 1) * PAGE_SIZE
    rows = db.execute(
        f"""SELECT al.*, ks.full_name, ks.username {base}
            ORDER BY al.created_at DESC, al.id DESC
            LIMIT ? OFFSET ?""",
        params + [PAGE_SIZE, offset],
    ).fetchall()

    from backend.services.audit_labels import describe_work, describe_result, result_ok
    return {
        "entries": [
            {
                "id":          r["id"],
                "created_at":  r["created_at"],
                "work":        describe_work(r["action"], r["target_type"]),
                "result":      describe_result(r["detail"], r["action"]),
                "result_ok":   result_ok(r["detail"], r["action"]),
                "target_type": r["target_type"],   # path thô — cho tooltip tra cứu
                "ip_address":  r["ip_address"],
                "username":    r["username"],
                "full_name":   r["full_name"],
            }
            for r in rows
        ],
        "total":     total,
        "page":      page,
        "page_size": PAGE_SIZE,
        "pages":     max(1, -(-total // PAGE_SIZE)),
    }


@router.get("/audit/export")
def export_audit_logs(
    method: str = Query(""),
    q:      str = Query(""),
    _: dict = Depends(require_feature("menu.logs")),
    db: sqlite3.Connection = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from datetime import date
    from backend.services.audit_labels import describe_work, describe_result

    where, params = _audit_where(method, q)
    rows = db.execute(
        f"""SELECT al.*, ks.full_name, ks.username
            FROM audit_logs al
            LEFT JOIN user_tttt ks ON al.actor_id = ks.id
            {where}
            ORDER BY al.created_at DESC, al.id DESC""",
        params,
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhật ký hệ thống"
    hdr_fill = PatternFill("solid", fgColor="37474F")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["STT", "Thời gian", "Người thao tác", "Username", "Công việc", "Kết quả", "IP", "Đường dẫn kỹ thuật"]
    widths  = [6, 18, 26, 18, 40, 20, 18, 40]
    ws.append(headers)
    for cell, w in zip(ws[1], widths):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for idx, r in enumerate(rows, 1):
        ts = (r["created_at"] or "")[:19].replace("T", " ")
        ws.append([
            idx, ts,
            r["full_name"] or "", r["username"] or "",
            describe_work(r["action"], r["target_type"]),
            describe_result(r["detail"], r["action"]),
            r["ip_address"] or "", r["target_type"] or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"nhat_ky_he_thong_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@router.get("/time-sync")
def get_time_sync(_: dict = Depends(require_feature("menu.logs"))):
    """Trạng thái lệch giờ máy chủ so với nguồn NTP + giờ máy hiện tại."""
    from backend.services.time_sync import check_drift
    r = check_drift()
    r["server_time"] = _vn_now().strftime("%Y-%m-%d %H:%M:%S")
    return r


@router.get("/")
def get_logs(
    level: str = Query(""),
    page:  int  = Query(1, ge=1),
    _: dict = Depends(require_feature("menu.logs")),
):
    entries, total = _parse_log_file(level, page)
    return {
        "entries":   entries,
        "total":     total,
        "page":      page,
        "page_size": PAGE_SIZE,
        "pages":     max(1, -(-total // PAGE_SIZE)),
    }
