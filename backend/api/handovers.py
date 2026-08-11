"""Handover (Bàn giao chứng từ) endpoints"""
import calendar
import io
import logging
import sqlite3
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill

from backend.core.concurrency import run_heavy
from backend.core.deps import get_current_staff, require_feature
from backend.core.enums import EntryStatus, StaffRole
from backend.database import get_db, _vn_now
from backend.schemas.handovers import (
    BorrowRequest, EntryHistoryItem, EntryHistoryOut,
    EntryUpsertRequest, GridEntryOut, GridResponse, HandbackRequest,
    RejectRequest, ReturnToStaffRequest,
)
from backend.services.handover_report_service import submitted_at_from_logs

# `log` là tên biến lặp trong get_entry_history → logger phải mang tên khác
log_ = logging.getLogger(__name__)

router = APIRouter(prefix="/api/handovers", tags=["Handovers"])


# ─── Phạm vi phòng được truy cập ─────────────────────────────────────────────
# Hai trục độc lập: XEM phòng nào, và CÓ được ghi hay không.
#   Xem mọi phòng — admin, GĐ, PGĐ (giám sát toàn đơn vị) và người hậu kiểm
#                   (có feature confirm_entry, nghiệp vụ chạy trên mọi phòng nguồn).
#   Còn lại        — kể cả trưởng/phó phòng — chỉ phòng của chính mình.
#   Ghi            — admin/GĐ/PGĐ bị cấm hoàn toàn; ai còn lại thì theo feature nhóm
#                    và chỉ trong phòng mình (trừ hậu kiểm).
_READ_ALL_DEPT_ROLES = (StaffRole.ADMIN.value, StaffRole.GIAM_DOC.value,
                        StaffRole.PHO_GIAM_DOC.value)

# Vai trò CHỈ ĐỌC ở màn bàn giao — chặn cứng ở tầng dependency, không qua feature.
# Admin bypass toàn bộ require_feature() nên nếu không chặn riêng ở đây thì admin
# vẫn ghi được mọi thứ. GĐ/PGĐ chỉ giám sát, không nhập liệu nghiệp vụ.
_NO_WRITE_ROLES = (StaffRole.ADMIN.value, StaffRole.GIAM_DOC.value,
                   StaffRole.PHO_GIAM_DOC.value)


def _has_feature(db: sqlite3.Connection, staff_id: int, feature_code: str) -> bool:
    return bool(db.execute(
        """SELECT 1 FROM group_features gf
           JOIN group_members gm ON gm.group_id = gf.group_id
           JOIN user_groups g ON g.id = gm.group_id AND g.is_active = 1
           WHERE gm.staff_id = ? AND gf.feature_code = ?
           LIMIT 1""",
        (staff_id, feature_code),
    ).fetchone())


def _can_view_all_depts(db: sqlite3.Connection, current: dict) -> bool:
    return (current["role"] in _READ_ALL_DEPT_ROLES
            or _has_feature(db, current["id"], "handovers.confirm_entry"))


def _can_write_all_depts(db: sqlite3.Connection, current: dict) -> bool:
    # Không còn vai trò nào được ghi xuyên phòng theo role — chỉ hậu kiểm (theo feature).
    return _has_feature(db, current["id"], "handovers.confirm_entry")


def require_handover_write(feature_code: str):
    """require_feature + chặn cứng vai trò chỉ đọc. Dùng cho mọi endpoint ghi."""
    def _check(current: dict = Depends(require_feature(feature_code))) -> dict:
        if current["role"] in _NO_WRITE_ROLES:
            raise HTTPException(403, "Vai trò của bạn chỉ được xem, không được thao tác trên chứng từ bàn giao")
        return current
    return _check


def _assert_dept_allowed(db: sqlite3.Connection, current: dict, department_id: Optional[int]) -> None:
    """Chặn user chỉ được XEM phòng mình khi đụng tới dữ liệu phòng khác."""
    if _can_view_all_depts(db, current):
        return
    if current.get("department_id") is None:
        raise HTTPException(403, "Tài khoản chưa được gán phòng — không thể xem chứng từ")
    if department_id != current["department_id"]:
        raise HTTPException(403, "Chỉ được xem chứng từ của phòng mình")


def _assert_dept_write_allowed(db: sqlite3.Connection, current: dict, department_id: Optional[int]) -> None:
    """Chặn thao tác (lưu / xác nhận / mượn / trả / từ chối) lên phòng khác."""
    if _can_write_all_depts(db, current):
        return
    if current.get("department_id") is None:
        raise HTTPException(403, "Tài khoản chưa được gán phòng — không thể thao tác chứng từ")
    if department_id != current["department_id"]:
        raise HTTPException(403, "Chỉ được thao tác trên chứng từ của phòng mình")


def _entry_dept(db: sqlite3.Connection, entry_id: int) -> Optional[int]:
    row = db.execute(
        """SELECT h.department_id FROM document_entries de
           JOIN handovers h ON h.id = de.handover_id WHERE de.id = ?""",
        (entry_id,),
    ).fetchone()
    return row["department_id"] if row else None


# ─── Phòng của cán bộ theo lịch sử đổi phòng ─────────────────────────────────
def _dept_at(db: sqlite3.Connection, staff_id: int, date_iso: str):
    """Phòng cán bộ thuộc về tại một ngày (dòng lịch sử mới nhất còn hiệu lực)."""
    row = db.execute(
        """SELECT department_id FROM staff_department_history
           WHERE staff_id = ? AND effective_from <= ?
           ORDER BY effective_from DESC, id DESC LIMIT 1""",
        (staff_id, date_iso),
    ).fetchone()
    if row:
        return row["department_id"]
    u = db.execute("SELECT department_id FROM user_tttt WHERE id = ?", (staff_id,)).fetchone()
    return u["department_id"] if u else None

# ─── Mapping hiển thị lịch sử ────────────────────────────────────────────────
_ACTION_LABEL = {
    "handover":           ("Bàn giao chứng từ",       "blue"),
    "edited_cv":          ("Sửa số tờ (CV)",           "blue"),
    "confirmed":          ("Xác nhận đã nhận",          "green"),
    "borrow_requested":   ("Yêu cầu mượn lại",          "orange"),
    "borrowed":           ("Xác nhận cho mượn",         "orange"),
    "returned":           ("Bàn giao lại",               "blue"),
    "returned_to_staff":  ("Trả lại chứng từ cho cán bộ", "orange"),
    "edited_hkv":         ("HKV sửa trực tiếp",          "purple"),
    "rejected_borrow":    ("Từ chối yêu cầu mượn",      "red"),
    "rejected_return":    ("Từ chối bàn giao lại",       "red"),
    "rejected_handover":  ("Từ chối chứng từ mới",       "red"),
}

_STATUS_LABEL = {
    EntryStatus.PENDING:   "Chờ xác nhận",
    EntryStatus.CONFIRMED: "Đã xác nhận",
    EntryStatus.BORROWED:  "Đang mượn",
    EntryStatus.REJECTED:  "Bị từ chối",
}

_ROLE_LABEL = {
    "admin":         "Quản trị viên cấp 1",
    "admin_l2":      "Quản trị viên cấp 2",
    "hau_kiem_vien": "Hậu kiểm viên",
    "giam_doc":      "Giám đốc",
    "pho_giam_doc":  "Phó Giám đốc",
    "truong_phong":  "Trưởng phòng",
    "pho_phong":     "Phó phòng",
    "chuyen_vien":   "Chuyên viên",
}


# ─── Grid ─────────────────────────────────────────────────────────────────────
@router.get("/grid", response_model=GridResponse)
def get_handover_grid(
    department_id: int,
    # Chặn ở tầng khai báo: monthrange()/date() bên dưới ném IllegalMonthError
    # và ValueError với giá trị ngoài khoảng → lọt lên thành HTTP 500.
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("menu.handovers")),
):
    _assert_dept_allowed(db, current, department_id)

    days_in_month = calendar.monthrange(year, month)[1]
    period_start = date(year, month, 1).isoformat()
    period_end = (date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)).isoformat()

    # Entries trong tháng — lọc theo phòng ĐÓNG BĂNG trong phiếu bàn giao
    # (handovers.department_id), KHÔNG theo phòng hiện tại của user. Nhờ vậy chứng
    # từ của cán bộ đã chuyển phòng vẫn nằm ở phòng cũ cho các tháng trước khi chuyển.
    entry_rows = db.execute(
        """SELECT de.id, de.handover_id, de.staff_id, de.transaction_date, de.sheet_count,
                  de.entry_status, de.entered_by_id
           FROM document_entries de
           JOIN handovers h ON de.handover_id = h.id
           WHERE h.department_id = ?
             AND de.staff_id IS NOT NULL
             AND de.transaction_date >= ? AND de.transaction_date < ?""",
        (department_id, period_start, period_end),
    ).fetchall()

    staff_ids_with_entries = {e["staff_id"] for e in entry_rows if e["staff_id"]}
    entered_by_ids = {e["entered_by_id"] for e in entry_rows if e["entered_by_id"]}

    # Cán bộ TỪNG thuộc phòng này trong bất kỳ ngày nào của tháng (theo lịch sử đổi
    # phòng) → hiện dòng trống để có thể nhập bù. Khoảng của mỗi mốc lịch sử là
    # [effective_from, mốc kế tiếp); giao với tháng ⇔ eff < period_end và (kế tiếp
    # NULL hoặc kế tiếp > period_start).
    overlap_rows = db.execute(
        """WITH hist AS (
               SELECT staff_id, department_id, effective_from,
                      LEAD(effective_from) OVER (
                          PARTITION BY staff_id ORDER BY effective_from, id
                      ) AS next_from
               FROM staff_department_history
           )
           SELECT DISTINCT staff_id FROM hist
           WHERE department_id = ?
             AND effective_from < ?
             AND (next_from IS NULL OR next_from > ?)""",
        (department_id, period_end, period_start),
    ).fetchall()
    overlap_ids = {r["staff_id"] for r in overlap_rows}

    # Staff rows: (thuộc phòng tháng đó & active) HOẶC có chứng từ (bất kể active).
    # CHỈ giao dịch viên (chuyen_vien) — trưởng/phó phòng, HKV, GĐ, admin không lên lưới.
    _GRID_ROLE = StaffRole.CHUYEN_VIEN.value
    eligible_ids = overlap_ids | staff_ids_with_entries
    if eligible_ids:
        ph = ",".join("?" * len(eligible_ids))
        if staff_ids_with_entries:
            ph2 = ",".join("?" * len(staff_ids_with_entries))
            user_rows = db.execute(
                f"SELECT * FROM user_tttt WHERE role = '{_GRID_ROLE}'"
                f"   AND id IN ({ph})"
                f"   AND (is_active = 1 OR id IN ({ph2}))"
                f" ORDER BY ipcas_code",
                list(eligible_ids) + list(staff_ids_with_entries),
            ).fetchall()
        else:
            user_rows = db.execute(
                f"SELECT * FROM user_tttt WHERE role = '{_GRID_ROLE}'"
                f"   AND id IN ({ph}) AND is_active = 1 ORDER BY ipcas_code",
                list(eligible_ids),
            ).fetchall()
    else:
        user_rows = []

    # Tên người nhập
    entered_by_map: dict = {}
    if entered_by_ids:
        ph2 = ",".join("?" * len(entered_by_ids))
        for r in db.execute(
            f"SELECT id, full_name FROM user_tttt WHERE id IN ({ph2})", list(entered_by_ids)
        ).fetchall():
            entered_by_map[r["id"]] = r["full_name"]

    grid_entries = [
        GridEntryOut(
            staff_id=e["staff_id"],
            day=int(e["transaction_date"].split("-")[2]),
            sheet_count=e["sheet_count"],
            entry_id=e["id"],
            entry_status=e["entry_status"] or EntryStatus.CONFIRMED,
            entered_by_name=entered_by_map.get(e["entered_by_id"]) if e["entered_by_id"] else None,
        )
        for e in entry_rows
    ]
    return GridResponse(users=[dict(u) for u in user_rows], entries=grid_entries, days_in_month=days_in_month)


# ─── Upsert (nhập/sửa ô grid) ────────────────────────────────────────────────
@router.put("/entry-upsert")
def upsert_entry(
    body: EntryUpsertRequest,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.save_entry")),
):
    staff_row = db.execute("SELECT * FROM user_tttt WHERE id = ?", (body.staff_id,)).fetchone()
    if not staff_row:
        raise HTTPException(404, "Không tìm thấy cán bộ")

    # Chỉ giao dịch viên mới có chứng từ bàn giao — chặn tận gốc, không để lọt entry rác
    if staff_row["role"] != StaffRole.CHUYEN_VIEN.value:
        raise HTTPException(400, "Chỉ giao dịch viên mới có chứng từ bàn giao")

    # Người không phải hậu kiểm chỉ nhập cho cán bộ phòng mình. So theo phòng HIỆN TẠI
    # của cán bộ (không phải phòng tại ngày giao dịch) để GDV vẫn nhập bù được cho
    # tháng trước khi chính mình chuyển phòng.
    if body.staff_id != current["id"]:
        _assert_dept_write_allowed(db, current, staff_row["department_id"])

    # User có quyền confirm entry → tạo entry ở trạng thái CONFIRMED ngay
    can_confirm = (current["role"] == StaffRole.ADMIN.value
                   or _has_feature(db, current["id"], "handovers.confirm_entry"))

    # Phòng của chứng từ = phòng cán bộ thuộc về TẠI NGÀY GIAO DỊCH (không phải phòng
    # hiện tại) → nhập bù cho cán bộ đã chuyển phòng vẫn vào đúng phòng cũ.
    dept_id = _dept_at(db, body.staff_id, body.date.isoformat()) or staff_row["department_id"]

    # ── Get/create handover (INSERT OR IGNORE để chống race condition) ──
    db.execute(
        "INSERT OR IGNORE INTO handovers (department_id, handover_date, received_by_id, status, created_at) VALUES (?,?,?,?,?)",
        (dept_id, body.date, None if not can_confirm else current["id"], "draft", str(_vn_now())),
    )
    h_row = db.execute(
        "SELECT * FROM handovers WHERE department_id = ? AND handover_date = ?",
        (dept_id, body.date),
    ).fetchone()
    if not h_row:
        raise HTTPException(500, "Không thể tạo phiếu bàn giao — hãy báo admin")
    handover_id = h_row["id"]

    # ── Entry hiện tại ──
    entry_row = db.execute(
        "SELECT * FROM document_entries WHERE handover_id = ? AND staff_id = ? AND transaction_date = ?",
        (handover_id, body.staff_id, body.date),
    ).fetchone()

    if body.sheet_count > 0:
        if entry_row:
            old_count = entry_row["sheet_count"]
            if not can_confirm:
                if entry_row["entry_status"] in (EntryStatus.CONFIRMED, EntryStatus.BORROWED):
                    raise HTTPException(403, "Chứng từ đã được xác nhận, không thể sửa trực tiếp")
                action = "handover"
                new_status = EntryStatus.PENDING
            else:
                action = "edited_hkv"
                new_status = EntryStatus.CONFIRMED

            set_extra, params_extra = "", []
            if can_confirm and entry_row["confirmed_by_id"] is None:
                set_extra = ", confirmed_by_id=?, confirmed_at=?"
                params_extra = [current["id"], str(_vn_now())]

            db.execute(
                f"UPDATE document_entries SET sheet_count=?, entry_status=?, entered_by_id=?{set_extra} WHERE id=?",
                [body.sheet_count, new_status, current["id"]] + params_extra + [entry_row["id"]],
            )
            db.execute(
                "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, old_sheet_count, new_sheet_count, timestamp) VALUES (?,?,?,?,?,?)",
                (entry_row["id"], action, current["id"], old_count, body.sheet_count, str(_vn_now())),
            )
        else:
            new_status = EntryStatus.PENDING if not can_confirm else EntryStatus.CONFIRMED
            # Dùng INSERT thường (không IGNORE) để lỗi duplicate nổi lên thay vì âm thầm bỏ qua
            db.execute(
                "INSERT INTO document_entries (handover_id, staff_id, transaction_date, sheet_count, entry_status, entered_by_id, confirmed_by_id, confirmed_at) VALUES (?,?,?,?,?,?,?,?)",
                (handover_id, body.staff_id, body.date, body.sheet_count, new_status, current["id"],
                 None if not can_confirm else current["id"], None if not can_confirm else str(_vn_now())),
            )
            entry_row = db.execute(
                "SELECT * FROM document_entries WHERE handover_id = ? AND staff_id = ? AND transaction_date = ?",
                (handover_id, body.staff_id, body.date),
            ).fetchone()
            if entry_row:
                db.execute(
                    "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, old_sheet_count, new_sheet_count, timestamp) VALUES (?,?,?,?,?,?)",
                    (entry_row["id"], "handover" if not can_confirm else "edited_hkv", current["id"], None, body.sheet_count, str(_vn_now())),
                )
            else:
                raise HTTPException(500, "Không thể lưu chứng từ — vui lòng thử lại")
    else:
        if entry_row:
            if not can_confirm and entry_row["entry_status"] not in (EntryStatus.PENDING, EntryStatus.REJECTED):
                raise HTTPException(400, "Không thể xóa chứng từ đã được xác nhận. Vui lòng liên hệ HKV/KSV.")
            db.execute("DELETE FROM entry_change_logs WHERE entry_id = ?", (entry_row["id"],))
            db.execute("DELETE FROM document_entries WHERE id = ?", (entry_row["id"],))

    db.commit()
    return {"ok": True}


# ─── Xác nhận đã nhận / xác nhận cho mượn (HKV/KSV) ────────────────────────
@router.post("/entries/{entry_id}/confirm-received")
def confirm_received(
    entry_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.confirm_entry")),
):
    entry = db.execute("SELECT * FROM document_entries WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_write_allowed(db, current, _entry_dept(db, entry_id))
    if entry["entry_status"] != EntryStatus.PENDING:
        raise HTTPException(409, f"Trạng thái không hợp lệ: '{_STATUS_LABEL.get(entry['entry_status'], entry['entry_status'])}'. Chỉ xác nhận được khi đang chờ xác nhận.")

    if entry["borrow_reason"]:
        db.execute(
            "UPDATE document_entries SET entry_status=?, borrowed_at=?, borrow_reason=NULL WHERE id=?",
            (EntryStatus.BORROWED, str(_vn_now()), entry_id),
        )
        action = "borrowed"
        message = "Đã xác nhận cho mượn chứng từ"
    else:
        db.execute(
            "UPDATE document_entries SET entry_status=?, confirmed_by_id=?, confirmed_at=? WHERE id=?",
            (EntryStatus.CONFIRMED, current["id"], str(_vn_now()), entry_id),
        )
        action = "confirmed"
        message = "Đã xác nhận nhận chứng từ"
        # Cập nhật received_by_id của handover nếu chưa có
        db.execute(
            "UPDATE handovers SET received_by_id=? WHERE id=(SELECT handover_id FROM document_entries WHERE id=?) AND received_by_id IS NULL",
            (current["id"], entry_id),
        )

    db.execute(
        "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, timestamp) VALUES (?,?,?,?,?)",
        (entry_id, action, current["id"], entry["sheet_count"], str(_vn_now())),
    )
    db.commit()
    return {"ok": True, "message": message}


# ─── Gửi yêu cầu mượn lại (Chuyên viên) ─────────────────────────────────────
@router.post("/entries/{entry_id}/borrow")
def borrow_entry(
    entry_id: int,
    body: BorrowRequest,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.borrow")),
):
    entry = db.execute("SELECT * FROM document_entries WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_write_allowed(db, current, _entry_dept(db, entry_id))
    if entry["entry_status"] != EntryStatus.CONFIRMED:
        raise HTTPException(409, f"Trạng thái không hợp lệ: '{_STATUS_LABEL.get(entry['entry_status'], entry['entry_status'])}'. Chỉ mượn được chứng từ đã xác nhận.")

    reason = body.reason.strip()
    if not reason:
        raise HTTPException(400, "Vui lòng nhập lý do mượn")

    db.execute(
        "UPDATE document_entries SET entry_status=?, borrow_reason=? WHERE id=?",
        (EntryStatus.PENDING, reason, entry_id),
    )
    db.execute(
        "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, timestamp) VALUES (?,?,?,?,?)",
        (entry_id, "borrow_requested", current["id"], entry["sheet_count"], str(_vn_now())),
    )
    db.commit()
    return {"ok": True, "message": "Đã gửi yêu cầu mượn chứng từ"}


# ─── Bàn giao lại (Chuyên viên, sau khi mượn) ────────────────────────────────
@router.post("/entries/{entry_id}/handback")
def handback_entry(
    entry_id: int,
    body: HandbackRequest,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.handback")),
):
    entry = db.execute("SELECT * FROM document_entries WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_write_allowed(db, current, _entry_dept(db, entry_id))
    if entry["entry_status"] != EntryStatus.BORROWED:
        raise HTTPException(409, f"Trạng thái không hợp lệ: '{_STATUS_LABEL.get(entry['entry_status'], entry['entry_status'])}'. Chỉ bàn giao lại được khi đang mượn.")

    old_count = entry["sheet_count"]
    db.execute(
        "UPDATE document_entries SET entry_status=?, borrow_reason=NULL, sheet_count=?, entered_by_id=? WHERE id=?",
        (EntryStatus.PENDING, body.sheet_count, current["id"], entry_id),
    )
    db.execute(
        "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, old_sheet_count, new_sheet_count, timestamp) VALUES (?,?,?,?,?,?)",
        (entry_id, "returned", current["id"], old_count, body.sheet_count, str(_vn_now())),
    )
    db.commit()
    return {"ok": True, "message": "Đã bàn giao lại chứng từ"}


# ─── Trả lại chứng từ cho cán bộ (HKV/KSV chủ động) ──────────────────────────
# Đường tắt của luồng mượn: bên đang giữ chứng từ đẩy thẳng đã xác nhận → đang mượn,
# không qua bước GDV gửi yêu cầu + HKV duyệt.
@router.post("/entries/{entry_id}/return-to-staff")
def return_entry_to_staff(
    entry_id: int,
    body: ReturnToStaffRequest,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.return_entry")),
):
    # Chỉ bên đang giữ chứng từ mới trả lại được. GDV muốn lấy chứng từ phải đi
    # đường "Mượn lại" để còn bước xác nhận của hậu kiểm — nếu không, GDV tự cấp
    # cho mình quyền rút chứng từ đã chốt.
    if current["role"] == StaffRole.CHUYEN_VIEN.value:
        raise HTTPException(403, "Giao dịch viên không trả lại được — hãy dùng chức năng Mượn lại")

    entry = db.execute("SELECT * FROM document_entries WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_write_allowed(db, current, _entry_dept(db, entry_id))
    if entry["entry_status"] != EntryStatus.CONFIRMED:
        raise HTTPException(409, f"Trạng thái không hợp lệ: '{_STATUS_LABEL.get(entry['entry_status'], entry['entry_status'])}'. Chỉ trả lại được chứng từ đã xác nhận.")

    reason = body.reason.strip()
    if not reason:
        raise HTTPException(400, "Vui lòng nhập lý do trả lại")

    # Lý do chỉ ghi vào log. KHÔNG ghi vào borrow_reason: cột đó đang là cờ
    # "có yêu cầu mượn chờ duyệt" mà confirm_received() và reject_entry() dựa vào
    # để phân nhánh — set ở đây sẽ làm hai endpoint kia hiểu sai luồng.
    db.execute(
        "UPDATE document_entries SET entry_status=?, borrowed_at=?, borrow_reason=NULL WHERE id=?",
        (EntryStatus.BORROWED, str(_vn_now()), entry_id),
    )
    db.execute(
        "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, notes, timestamp) VALUES (?,?,?,?,?,?)",
        (entry_id, "returned_to_staff", current["id"], entry["sheet_count"], reason, str(_vn_now())),
    )
    db.commit()
    return {"ok": True, "message": "Đã trả lại chứng từ cho cán bộ"}


# ─── Endpoint đã ngừng sử dụng ───────────────────────────────────────────────
@router.post("/entries/{entry_id}/confirm-returned")
def confirm_returned(entry_id: int, _: dict = Depends(get_current_staff)):
    raise HTTPException(410, "Endpoint đã ngừng sử dụng. Dùng /handback + /confirm-received")


# ─── Từ chối (HKV/KSV) ───────────────────────────────────────────────────────
@router.post("/entries/{entry_id}/reject")
def reject_entry(
    entry_id: int,
    body: RejectRequest,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.reject_entry")),
):
    entry = db.execute("SELECT * FROM document_entries WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_write_allowed(db, current, _entry_dept(db, entry_id))
    if entry["entry_status"] != EntryStatus.PENDING:
        raise HTTPException(409, f"Trạng thái không hợp lệ: '{_STATUS_LABEL.get(entry['entry_status'], entry['entry_status'])}'. Chỉ từ chối được khi đang chờ xác nhận.")

    reason = body.reason.strip()
    if not reason:
        raise HTTPException(400, "Vui lòng nhập lý do từ chối")

    if entry["borrow_reason"]:
        # Từ chối yêu cầu mượn → quay về confirmed
        db.execute(
            "UPDATE document_entries SET entry_status=?, borrow_reason=NULL WHERE id=?",
            (EntryStatus.CONFIRMED, entry_id),
        )
        db.execute(
            "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, notes, timestamp) VALUES (?,?,?,?,?,?)",
            (entry_id, "rejected_borrow", current["id"], entry["sheet_count"], reason, str(_vn_now())),
        )
        db.commit()
        return {"ok": True, "message": "Đã từ chối yêu cầu mượn chứng từ"}

    # Xét log cuối để phân loại
    last_log = db.execute(
        "SELECT * FROM entry_change_logs WHERE entry_id = ? ORDER BY timestamp DESC LIMIT 1",
        (entry_id,),
    ).fetchone()
    last_action = last_log["action"] if last_log else None

    if last_action == "returned":
        # Từ chối bàn giao lại → quay về borrowed, khôi phục số tờ cũ
        restore_count = last_log["old_sheet_count"] if last_log["old_sheet_count"] is not None else entry["sheet_count"]
        db.execute(
            "UPDATE document_entries SET entry_status=?, sheet_count=? WHERE id=?",
            (EntryStatus.BORROWED, restore_count, entry_id),
        )
        db.execute(
            "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, notes, timestamp) VALUES (?,?,?,?,?,?)",
            (entry_id, "rejected_return", current["id"], restore_count, reason, str(_vn_now())),
        )
        db.commit()
        return {"ok": True, "message": "Đã từ chối bàn giao lại. Chứng từ vẫn đang mượn."}

    # Chứng từ mới chưa xác nhận → từ chối → giữ entry + log để lưu lịch sử
    db.execute(
        "UPDATE document_entries SET entry_status=? WHERE id=?",
        (EntryStatus.REJECTED, entry_id),
    )
    db.execute(
        "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, notes, timestamp) VALUES (?,?,?,?,?,?)",
        (entry_id, "rejected_handover", current["id"], entry["sheet_count"], reason, str(_vn_now())),
    )
    db.commit()
    return {"ok": True, "message": "Đã từ chối chứng từ. Cán bộ có thể nộp lại."}


# ─── Nộp lại chứng từ bị từ chối (GDV) ───────────────────────────────────────
@router.post("/entries/{entry_id}/resubmit")
def resubmit_entry(
    entry_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_handover_write("handovers.save_entry")),
):
    entry = db.execute("SELECT * FROM document_entries WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_write_allowed(db, current, _entry_dept(db, entry_id))
    if entry["entry_status"] != EntryStatus.REJECTED:
        raise HTTPException(409, f"Trạng thái không hợp lệ: '{_STATUS_LABEL.get(entry['entry_status'], entry['entry_status'])}'. Chỉ nộp lại được chứng từ bị từ chối.")

    db.execute(
        "UPDATE document_entries SET entry_status=?, entered_by_id=? WHERE id=?",
        (EntryStatus.PENDING, current["id"], entry_id),
    )
    db.execute(
        "INSERT INTO entry_change_logs (entry_id, action, performed_by_id, new_sheet_count, timestamp) VALUES (?,?,?,?,?)",
        (entry_id, "handover", current["id"], entry["sheet_count"], str(_vn_now())),
    )
    db.commit()
    return {"ok": True, "message": "Đã nộp lại chứng từ, chờ HKV xác nhận"}


# ─── Lịch sử thay đổi ────────────────────────────────────────────────────────
@router.get("/entries/{entry_id}/history", response_model=EntryHistoryOut)
def get_entry_history(
    entry_id: int,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("menu.handovers")),
):
    entry = db.execute(
        "SELECT de.*, ks.full_name AS s_name, ks.ipcas_code, ks.payment_username, ks.department_id AS s_dept_id "
        "FROM document_entries de LEFT JOIN user_tttt ks ON de.staff_id = ks.id WHERE de.id = ?",
        (entry_id,),
    ).fetchone()
    if not entry:
        raise HTTPException(404, "Không tìm thấy chứng từ")
    _assert_dept_allowed(db, current, _entry_dept(db, entry_id))

    logs = db.execute(
        """SELECT ecl.*, ks.full_name AS p_name, ks.role AS p_role
           FROM entry_change_logs ecl
           LEFT JOIN user_tttt ks ON ecl.performed_by_id = ks.id
           WHERE ecl.entry_id = ? ORDER BY ecl.timestamp DESC""",
        (entry_id,),
    ).fetchall()

    history_items = []
    for log in logs:
        action_key = log["action"] or ""
        label, color = _ACTION_LABEL.get(action_key, (action_key, "blue"))

        if log["old_sheet_count"] is not None and log["new_sheet_count"] is not None and log["old_sheet_count"] != log["new_sheet_count"]:
            label = f"{label}: {log['old_sheet_count']} → {log['new_sheet_count']} tờ"
        elif log["new_sheet_count"] is not None and log["old_sheet_count"] is None:
            label = f"{label} — {log['new_sheet_count']} tờ"

        if log["notes"]:
            label = f"{label} · Lý do: {log['notes']}"

        p_name = log["p_name"] or "?"
        role_label = _ROLE_LABEL.get(log["p_role"] or "", log["p_role"] or "?")
        performer_display = f"{p_name} · {role_label}"

        ts = log["timestamp"] or ""
        if ts:
            from datetime import datetime
            try:
                ts_dt = datetime.fromisoformat(str(ts))
                ts = ts_dt.strftime("%H:%M:%S  %d/%m/%Y")
            except Exception:
                pass

        history_items.append(EntryHistoryItem(
            id=log["id"],
            action=action_key,
            action_label=label,
            action_color=color,
            performed_by=p_name,
            performed_by_role=performer_display,
            timestamp=ts,
            old_sheet_count=log["old_sheet_count"],
            new_sheet_count=log["new_sheet_count"],
        ))

    s_name = entry["s_name"] or entry["payment_username"] or entry["ipcas_code"] or f"ID {entry['staff_id']}"
    current_status = entry["entry_status"] or EntryStatus.CONFIRMED
    tx_date = date.fromisoformat(entry["transaction_date"]).strftime("%d/%m/%Y") if entry["transaction_date"] else ""

    # ── Ngày nộp thật ──
    # Log nộp sớm nhất còn hiệu lực (bỏ các lần đã bị từ chối). Không lấy log mới nhất
    # (thao tác sau như xác nhận / mượn sẽ đẩy ngày trôi đi) và không lấy
    # `handovers.handover_date` (luôn bằng transaction_date khi nhập qua lưới).
    submit_ts = submitted_at_from_logs(logs)
    submit_date = None
    if submit_ts:
        try:
            submit_date = date.fromisoformat(submit_ts[:10]).strftime("%d/%m/%Y")
        except ValueError:
            log_.warning("Timestamp nộp chứng từ không đọc được: entry=%s raw=%r", entry_id, submit_ts)

    return EntryHistoryOut(
        entry_id=entry_id,
        source_user_name=s_name,
        transaction_date=tx_date,
        submit_date=submit_date,
        sheet_count=entry["sheet_count"],
        current_status=current_status,
        current_status_label=_STATUS_LABEL.get(current_status, current_status),
        borrow_reason=entry["borrow_reason"],
        logs=history_items,
    )


@router.get("/export")
async def export_handovers(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    department_id: Optional[int] = None,
    db: sqlite3.Connection = Depends(get_db),
    current: dict = Depends(require_feature("menu.handovers")),
):
    """Xuất danh sách chứng từ bàn giao ra Excel."""
    # Không được xem mọi phòng → ép về phòng mình, bỏ qua department_id client gửi lên.
    # (Không lọc = xuất toàn bộ chứng từ mọi phòng.)
    if not _can_view_all_depts(db, current):
        if current.get("department_id") is None:
            raise HTTPException(403, "Tài khoản chưa được gán phòng — không thể xuất dữ liệu")
        department_id = current["department_id"]

    clauses = []
    params: list = []
    if department_id:
        clauses.append("h.department_id = ?")
        params.append(department_id)
    if from_date:
        clauses.append("h.handover_date >= ?")
        params.append(from_date.isoformat())
    if to_date:
        clauses.append("h.handover_date <= ?")
        params.append(to_date.isoformat())

    where = "WHERE " + " AND ".join(clauses) if clauses else ""

    def _work() -> io.BytesIO:
        entries = db.execute(
            f"""SELECT de.id, de.transaction_date, de.sheet_count, de.entry_status,
                       h.handover_date, h.delivered_by,
                       d.name AS dept_name,
                       ks.ipcas_code, ks.full_name, ks.payment_username
                FROM document_entries de
                JOIN handovers h ON de.handover_id = h.id
                JOIN departments d ON h.department_id = d.id
                LEFT JOIN user_tttt ks ON de.staff_id = ks.id
                {where}
                ORDER BY h.handover_date DESC, d.name""",
            params,
        ).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bàn giao chứng từ"

        hdr_fill = PatternFill("solid", fgColor="1565C0")
        hdr_font = Font(bold=True, color="FFFFFF")
        headers = ["STT", "Phòng", "Ngày bàn giao", "Ngày giao dịch",
                   "User IPCAS", "Họ và tên", "Số tờ", "Trạng thái", "Người nộp"]
        widths  = [6, 20, 14, 14, 16, 25, 9, 16, 22]
        ws.append(headers)
        for cell, w in zip(ws[1], widths):
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w

        for idx, e in enumerate(entries, 1):
            ho_date = date.fromisoformat(e["handover_date"]).strftime("%d/%m/%Y") if e["handover_date"] else ""
            tx_date = date.fromisoformat(e["transaction_date"]).strftime("%d/%m/%Y") if e["transaction_date"] else ""
            display_name = e["full_name"] or e["payment_username"] or ""
            ws.append([
                idx,
                e["dept_name"] or "",
                ho_date,
                tx_date,
                e["ipcas_code"] or "",
                display_name,
                e["sheet_count"],
                _STATUS_LABEL.get(e["entry_status"] or "confirmed", e["entry_status"] or ""),
                e["delivered_by"] or "",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = await run_heavy(_work)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''chung_tu_ban_giao.xlsx"},
    )
