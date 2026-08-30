"""API Quản lý nhân sự — hồ sơ cán bộ, tra cứu, thống kê, nhắc lịch.

Quyền (đặt qua màn Phân quyền theo nhóm, xem backend/core/features.py):

| Việc                              | Hồ sơ của chính mình | Hồ sơ người khác             |
|-----------------------------------|----------------------|------------------------------|
| Xem                               | `menu.hr_profiles`   | + `hr.view_all`              |
| Sửa phần tự khai (cá nhân, bằng   | `menu.hr_profiles`   | + `hr.edit_all`              |
| cấp, đào tạo, công cụ)            |                      |                              |
| Sửa phần công tác, bổ nhiệm,      | `hr.edit_all`        | `hr.edit_all`                |
| quá trình công tác, nghỉ gián đoạn|                      |                              |
| Xem hồ sơ lương                   | `menu.hr_profiles`   | + `hr.salary_view`           |
| Sửa hồ sơ lương                   | `hr.salary_edit`     | `hr.salary_edit`             |

Lý do "sửa lương luôn cần quyền riêng, kể cả hồ sơ mình": bậc lương và hệ số là
số liệu do người làm chế độ nhập theo quyết định, để người ta tự sửa của mình
thì bảng lương mất giá trị đối chiếu.
"""
import io
import logging
import os
import sqlite3
from datetime import date
from typing import Optional
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl.styles import Alignment, Font, PatternFill

from backend.core.concurrency import run_heavy
from backend.core.deps import require_feature
from backend.core.enums import StaffRole
from backend.core.uploads import read_limited, safe_filename
from backend.database import _vn_now, get_db, write_audit
from backend.schemas.hr import DirectoryRow, ReminderOut, StaffBrief, StatsOut
from backend.services import hr_service as hr

_log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/hr", tags=["Quản lý nhân sự"])

# Ảnh thẻ và file quyết định nằm trong SQLite (BLOB), giống ảnh chữ ký đã có.
# Cả hai đều nhỏ và phải đi kèm bản sao lưu DB — để ngoài đĩa là backup DB xong
# vẫn mất file. Trần dưới đây chặn ngay tại cửa, không đợi trần chung 200 MB.
_TRAN_ANH = 5 * 1024 * 1024
_TRAN_FILE = 15 * 1024 * 1024

# Định dạng nhận vào khoá theo PHẦN MỞ RỘNG, và kiểu MIME lưu lại cũng lấy từ
# bảng này chứ KHÔNG lấy `file.content_type`. Kiểu MIME là chuỗi do trình duyệt
# (hay bất cứ ai gọi API) tự khai: tin nó rồi lưu lại thì người tải lên chọn được
# luôn kiểu mà máy chủ sẽ phát ngược ra cho người khác — ảnh thẻ được trả về
# INLINE nên đó là đường đưa nội dung lạ chạy trên chính tên miền của hệ thống.
_ANH_MO_RONG = {".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".png": "image/png", ".webp": "image/webp"}
_FILE_MO_RONG = {**_ANH_MO_RONG, ".pdf": "application/pdf"}


# ── Tiện ích quyền ───────────────────────────────────────────────────────────
def _co_quyen(db: sqlite3.Connection, current: dict, code: str) -> bool:
    """Admin luôn có; còn lại tra quyền được gán qua nhóm (giống require_feature)."""
    if current.get("role") == StaffRole.ADMIN:
        return True
    return bool(db.execute(
        """SELECT 1 FROM group_features gf
           JOIN group_members gm ON gm.group_id = gf.group_id
           JOIN user_groups g ON g.id = gm.group_id AND g.is_active = 1
           WHERE gm.staff_id = ? AND gf.feature_code = ? LIMIT 1""",
        (current["id"], code),
    ).fetchone())


def _quyen_cua_toi(db: sqlite3.Connection, current: dict) -> dict:
    return {
        "view_all":    _co_quyen(db, current, "hr.view_all"),
        "edit_all":    _co_quyen(db, current, "hr.edit_all"),
        "salary_view": _co_quyen(db, current, "hr.salary_view"),
        "salary_edit": _co_quyen(db, current, "hr.salary_edit"),
        "export":      _co_quyen(db, current, "hr.export"),
    }


def _cam(chi_tiet: str):
    return HTTPException(403, chi_tiet)


def _kiem_xem(db, current: dict, staff_id: int, section: str | None = None) -> None:
    la_minh = staff_id == current["id"]
    if not la_minh and not _co_quyen(db, current, "hr.view_all"):
        raise _cam("Không có quyền xem hồ sơ của cán bộ khác")
    if section and hr.SECTIONS.get(section, {}).get("quyen_xem") and not la_minh:
        if not _co_quyen(db, current, hr.SECTIONS[section]["quyen_xem"]):
            raise _cam("Không có quyền xem hồ sơ lương của cán bộ khác")


def _kiem_sua(db, current: dict, staff_id: int, section: str | None = None,
              tu_sua: bool = False) -> None:
    """`tu_sua`: phần này cán bộ có được tự sửa hồ sơ của mình không."""
    spec = hr.SECTIONS.get(section, {}) if section else {}
    if spec.get("quyen_sua"):          # lương: luôn cần quyền riêng
        if not _co_quyen(db, current, spec["quyen_sua"]):
            raise _cam("Không có quyền sửa hồ sơ lương")
        return
    if tu_sua and staff_id == current["id"]:
        return
    if not _co_quyen(db, current, "hr.edit_all"):
        raise _cam("Không có quyền sửa hồ sơ nhân sự")


def _spec(section: str) -> dict:
    spec = hr.SECTIONS.get(section)
    if not spec:
        raise HTTPException(404, f"Không có phân hệ '{section}'")
    return spec


def _staff(db, staff_id: int) -> sqlite3.Row:
    row = db.execute(
        """SELECT u.*, d.name AS department, d.code AS department_code
           FROM user_tttt u LEFT JOIN departments d ON d.id = u.department_id
           WHERE u.id = ?""",
        (staff_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy cán bộ")
    # Chốt duy nhất chặn hồ sơ cho tài khoản quản trị. Đặt ở đây vì MỌI đường ghi
    # (sửa hồ sơ, ảnh thẻ, thêm dòng phân hệ) đều đi qua hàm này — chặn rải rác ở
    # từng endpoint thì thêm endpoint mới là quên.
    if row["role"] in hr.ROLES_KHONG_HO_SO:
        raise HTTPException(
            404, "Tài khoản quản trị viên không có hồ sơ nhân sự — hồ sơ chỉ dành "
                 "cho cán bộ nghiệp vụ.")
    return row


def _kieu_file(filename: str, cho_phep: dict, mo_ta: str) -> tuple[str, str]:
    """Tên file client gửi lên → (tên đã làm sạch, kiểu MIME tự xác định).

    Từ chối ngay nếu phần mở rộng không nằm trong danh sách — người dùng nhìn
    thấy đuôi file, nên báo lỗi theo đuôi mới hiểu được ("chỉ nhận .pdf, .jpg…").
    """
    ten = safe_filename(filename, "tep.dat")
    duoi = os.path.splitext(ten)[1].lower()
    if duoi not in cho_phep:
        raise HTTPException(
            400, f"{mo_ta} chỉ nhận {', '.join(sorted(cho_phep))} — file gửi lên "
                 f"là '{ten}'")
    return ten, cho_phep[duoi]


def _download_headers(filename: str) -> dict:
    fallback = "".join(ch if ord(ch) < 128 and ch not in '\\"' else "_" for ch in filename)
    return {"Content-Disposition": (f'attachment; filename="{fallback}"; '
                                    f"filename*=UTF-8''{quote(filename, safe='')}")}


# ── Đặc tả phân hệ cho frontend dựng form ────────────────────────────────────
@router.get("/meta")
def get_meta(
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Nhãn cột + kiểu dữ liệu của mọi phân hệ. Frontend dựng form từ đây nên
    thêm cột chỉ phải sửa `hr_service.SECTIONS`, không phải sửa cả hai bên."""
    return {
        "sections": {
            ten: {
                "nhan": s["nhan"],
                "tu_sua": s["tu_sua"],
                "co_file": ten in hr.SECTIONS_CO_FILE,
                "fields": s["fields"],
            }
            for ten, s in hr.SECTIONS.items()
        },
        "profile": {
            "tu_khai":   hr.PROFILE_FIELDS_TU_KHAI,
            "cong_tac":  hr.PROFILE_FIELDS_CONG_TAC,
            "tai_khoan": hr.PROFILE_FIELDS_TAI_KHOAN,
            # Nằm ở user_tttt nhưng thuộc khối "công tác" về mặt quyền
            "tai_khoan_hr": hr.PROFILE_FIELDS_TAI_KHOAN_HR,
        },
        "nhom_tra_cuu": hr.NHOM_TRA_CUU,
        # Nhãn chức vụ để danh sách hiển thị đúng chữ — frontend không khai lại
        "chuc_vu": hr.NHAN_CHUC_VU,
        # Danh sách đuôi file + trần dung lượng gửi kèm để ô chọn file ngoài giao
        # diện lọc đúng thứ backend nhận. Trước đây frontend khai `image/*` rộng
        # hơn backend: chọn ảnh GIF thì tải xong mới báo lỗi.
        "tep": {
            "anh_accept":   ",".join(sorted(_ANH_MO_RONG)),
            "file_accept":  ",".join(sorted(_FILE_MO_RONG)),
            "tran_anh_mb":  _TRAN_ANH // (1024 * 1024),
            "tran_file_mb": _TRAN_FILE // (1024 * 1024),
        },
        "quyen": _quyen_cua_toi(db, current),
    }


# ── Danh sách cán bộ ─────────────────────────────────────────────────────────
@router.get("/profiles", response_model=list[StaffBrief])
def list_profiles(
    q: Optional[str] = Query(None, description="Tìm theo mã cán bộ / họ tên"),
    department_id: Optional[int] = Query(None),
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Không có `hr.view_all` thì danh sách chỉ có đúng hồ sơ của mình — không
    trả 403, vì màn hình vẫn dùng được để tự khai hồ sơ."""
    sql = """SELECT u.id AS staff_id, u.employee_code, u.full_name, u.role,
                    u.phone, u.email, d.name AS department,
                    p.position_title, p.gender, p.dob,
                    (p.staff_id IS NOT NULL) AS co_ho_so
             FROM user_tttt u
             LEFT JOIN departments d ON d.id = u.department_id
             LEFT JOIN hr_profiles p ON p.staff_id = u.id
             WHERE u.is_active = 1 AND IFNULL(u.is_deleted, 0) = 0
               AND {chi_can_bo}"""
    # Quản trị viên không nằm trong danh sách hồ sơ — xem hr_service.ROLES_KHONG_HO_SO
    sql = sql.format(chi_can_bo=hr.SQL_CHI_CAN_BO)
    params: list = []
    if not _co_quyen(db, current, "hr.view_all"):
        sql += " AND u.id = ?"
        params.append(current["id"])
    if department_id:
        sql += " AND u.department_id = ?"
        params.append(department_id)
    rows = [dict(r) for r in db.execute(sql, params).fetchall()]
    # Sắp bằng Python chứ không bằng ORDER BY: SQLite so sánh theo mã byte nên
    # tên có dấu bị dồn xuống cuối danh sách (xem hr_service.khoa_ten).
    rows.sort(key=lambda r: hr.khoa_phong_ten(r["department"], r["full_name"], r["role"]))

    if q and q.strip():
        # Lọc bằng Python: LOWER() của SQLite không hạ được chữ có dấu, và bảng
        # nhân sự chỉ vài trăm dòng (xem cùng lý do ở backend/api/ttqt_branches.py).
        kw = hr.bo_dau(q.strip())
        rows = [r for r in rows
                if kw in hr.bo_dau(f"{r['employee_code'] or ''} {r['full_name']}")]
    for r in rows:
        r["co_ho_so"] = bool(r["co_ho_so"])
    return rows


# ── Một hồ sơ đầy đủ ─────────────────────────────────────────────────────────
def _doc_section(db, section: str, staff_id: int) -> list[dict]:
    spec = hr.SECTIONS[section]
    rows = [dict(r) for r in db.execute(
        f"SELECT * FROM {spec['table']} WHERE staff_id = ? ORDER BY {spec['order_by']}",
        (staff_id,),
    ).fetchall()]
    if not rows:
        return rows
    dinh_kem: dict[int, list] = {}
    for a in db.execute(
        f"""SELECT id, item_id, filename, mime, size_bytes, uploaded_at
            FROM hr_attachments WHERE section = ?
              AND item_id IN ({','.join('?' * len(rows))})""",
        (section, *[r["id"] for r in rows]),
    ).fetchall():
        dinh_kem.setdefault(a["item_id"], []).append(dict(a))
    for r in rows:
        r["files"] = dinh_kem.get(r["id"], [])
        for k, v in list(r.items()):
            if isinstance(v, bytes):        # không bao giờ trả BLOB trong JSON
                r.pop(k)
    return rows


@router.get("/profiles/{staff_id}")
def get_profile(
    staff_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    u = _staff(db, staff_id)      # chặn luôn tài khoản quản trị (404 kèm lý do)
    _kiem_xem(db, current, staff_id)
    p = db.execute("SELECT * FROM hr_profiles WHERE staff_id = ?", (staff_id,)).fetchone()
    ho_so = {k: (dict(p).get(k) if p else None) for k in hr.PROFILE_FIELDS}
    quyen = _quyen_cua_toi(db, current)
    la_minh = staff_id == current["id"]

    sections = {}
    for ten, spec in hr.SECTIONS.items():
        if spec.get("quyen_xem") and not la_minh and not quyen["salary_view"]:
            continue        # ẩn hẳn phân hệ lương thay vì trả dòng rỗng gây hiểu nhầm
        sections[ten] = _doc_section(db, ten, staff_id)

    return {
        "staff": {
            "staff_id": staff_id,
            "employee_code": u["employee_code"],
            "full_name": u["full_name"],
            "role": u["role"],
            "department": u["department"],
            "department_id": u["department_id"],
            "phone": u["phone"],
            "email": u["email"],
            "join_industry_date": u["join_industry_date"],
            "is_active": bool(u["is_active"]),
        },
        "profile": ho_so,
        "co_anh": bool(p and p["photo"]),
        "sections": sections,
        "quyen": {
            **quyen,
            "sua_tu_khai": la_minh or quyen["edit_all"],
            "sua_cong_tac": quyen["edit_all"],
            "xem_luong": la_minh or quyen["salary_view"],
        },
    }


@router.put("/profiles/{staff_id}")
def put_profile(
    staff_id: int,
    body: dict = Body(...),
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Cập nhật hồ sơ cá nhân + thông tin công tác.

    Bốn nhóm trường, hai mức quyền, hai bảng đích:

    | Nhóm            | Bảng        | Quyền cần        |
    |-----------------|-------------|------------------|
    | tu_khai         | hr_profiles | tự khai          |
    | tai_khoan       | user_tttt   | tự khai          |
    | cong_tac        | hr_profiles | `hr.edit_all`    |
    | tai_khoan_hr    | user_tttt   | `hr.edit_all`    |

    Kiểm riêng từng nhóm chứ không lọc bỏ im lặng: người chỉ được tự khai mà gửi
    kèm `join_industry_date` sẽ nhận 403, không phải "lưu xong mà không đổi gì".
    """
    _staff(db, staff_id)
    tu_khai = {k: v for k, v in body.items() if k in hr.PROFILE_FIELDS_TU_KHAI}
    cong_tac = {k: v for k, v in body.items() if k in hr.PROFILE_FIELDS_CONG_TAC}
    tai_khoan = {k: v for k, v in body.items() if k in hr.PROFILE_FIELDS_TAI_KHOAN}
    tk_hr = {k: v for k, v in body.items() if k in hr.PROFILE_FIELDS_TAI_KHOAN_HR}
    la = sorted(set(body) - set(hr.PROFILE_FIELDS) - set(hr.PROFILE_FIELDS_TAI_KHOAN)
                - set(hr.PROFILE_FIELDS_TAI_KHOAN_HR))
    if la:
        raise HTTPException(400, f"Trường không thuộc hồ sơ: {', '.join(la)}")

    if tu_khai or tai_khoan:
        _kiem_sua(db, current, staff_id, tu_sua=True)
    if cong_tac or tk_hr:
        _kiem_sua(db, current, staff_id, tu_sua=False)

    try:
        gia_tri = {
            **hr.chuan_hoa(hr.PROFILE_FIELDS_TU_KHAI, tu_khai, mot_phan=True),
            **hr.chuan_hoa(hr.PROFILE_FIELDS_CONG_TAC, cong_tac, mot_phan=True),
        }
        tk = {
            **hr.chuan_hoa(hr.PROFILE_FIELDS_TAI_KHOAN, tai_khoan, mot_phan=True),
            **hr.chuan_hoa(hr.PROFILE_FIELDS_TAI_KHOAN_HR, tk_hr, mot_phan=True),
        }
    except hr.LoiDuLieu as e:
        raise HTTPException(400, str(e))

    db.execute("INSERT OR IGNORE INTO hr_profiles (staff_id) VALUES (?)", (staff_id,))
    if gia_tri:
        db.execute(
            f"UPDATE hr_profiles SET {','.join(k + '=?' for k in gia_tri)},"
            " updated_at=?, updated_by=? WHERE staff_id = ?",
            (*gia_tri.values(), _vn_now(), current["id"], staff_id),
        )
    # Điện thoại / email / ngày vào ngành ghi thẳng vào user_tttt — hồ sơ không
    # giữ bản thứ hai. Đổi `join_industry_date` là đổi luôn SỐ NGÀY PHÉP NĂM của
    # người đó (compute_annual_leave), nên nó nằm trong nhật ký ở dòng dưới.
    if tk:
        db.execute(f"UPDATE user_tttt SET {','.join(k + '=?' for k in tk)} WHERE id = ?",
                   (*tk.values(), staff_id))
    write_audit(db, current["id"], "hr.profile.update", "hr_profile", staff_id,
                f"Cập nhật: {', '.join(sorted({**gia_tri, **tk})) or 'không có gì'}")
    db.commit()
    return get_profile(staff_id, current, db)


# ── Ảnh thẻ ──────────────────────────────────────────────────────────────────
@router.post("/profiles/{staff_id}/photo")
async def upload_photo(
    staff_id: int,
    file: UploadFile = File(...),
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    _staff(db, staff_id)
    _kiem_sua(db, current, staff_id, tu_sua=True)
    ten, mime = _kieu_file(file.filename, _ANH_MO_RONG, "Ảnh cán bộ")
    noi_dung = await read_limited(file, _TRAN_ANH, ten="Ảnh cán bộ")
    db.execute("INSERT OR IGNORE INTO hr_profiles (staff_id) VALUES (?)", (staff_id,))
    db.execute("UPDATE hr_profiles SET photo=?, photo_mime=?, updated_at=?, updated_by=?"
               " WHERE staff_id=?",
               (noi_dung, mime, _vn_now(), current["id"], staff_id))
    write_audit(db, current["id"], "hr.photo.update", "hr_profile", staff_id, ten)
    db.commit()
    return {"ok": True, "size": len(noi_dung)}


@router.get("/profiles/{staff_id}/photo")
def get_photo(
    staff_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    _kiem_xem(db, current, staff_id)
    row = db.execute("SELECT photo, photo_mime FROM hr_profiles WHERE staff_id = ?",
                     (staff_id,)).fetchone()
    if not row or not row["photo"]:
        raise HTTPException(404, "Cán bộ chưa có ảnh")
    return Response(content=row["photo"], media_type=row["photo_mime"] or "image/jpeg",
                    headers={"Cache-Control": "no-store"})


@router.delete("/profiles/{staff_id}/photo")
def delete_photo(
    staff_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    _staff(db, staff_id)
    _kiem_sua(db, current, staff_id, tu_sua=True)
    db.execute("UPDATE hr_profiles SET photo=NULL, photo_mime=NULL, updated_at=?,"
               " updated_by=? WHERE staff_id=?", (_vn_now(), current["id"], staff_id))
    write_audit(db, current["id"], "hr.photo.delete", "hr_profile", staff_id, None)
    db.commit()
    return {"ok": True}


# ── CRUD dùng chung cho 7 phân hệ ────────────────────────────────────────────
@router.get("/sections/{section}/{staff_id}")
def list_items(
    section: str,
    staff_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    _spec(section)
    _kiem_xem(db, current, staff_id, section)
    return _doc_section(db, section, staff_id)


@router.post("/sections/{section}/{staff_id}")
def create_item(
    section: str,
    staff_id: int,
    body: dict = Body(...),
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    spec = _spec(section)
    _staff(db, staff_id)
    _kiem_sua(db, current, staff_id, section, tu_sua=spec["tu_sua"])
    try:
        gia_tri = hr.chuan_hoa(spec["fields"], body)
    except hr.LoiDuLieu as e:
        raise HTTPException(400, str(e))
    cot = list(gia_tri)
    cur = db.execute(
        f"INSERT INTO {spec['table']} (staff_id, {','.join(cot)}, created_at, updated_at, updated_by)"
        f" VALUES (?, {','.join('?' * len(cot))}, ?, ?, ?)",
        (staff_id, *gia_tri.values(), _vn_now(), _vn_now(), current["id"]),
    )
    write_audit(db, current["id"], f"hr.{section}.create", spec["table"], cur.lastrowid,
                f"staff_id={staff_id}")
    db.commit()
    row = db.execute(f"SELECT * FROM {spec['table']} WHERE id = ?", (cur.lastrowid,)).fetchone()
    return {**dict(row), "files": []}


def _lay_dong(db, spec: dict, item_id: int) -> sqlite3.Row:
    row = db.execute(f"SELECT * FROM {spec['table']} WHERE id = ?", (item_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy dòng hồ sơ")
    return row


@router.patch("/items/{section}/{item_id}")
def update_item(
    section: str,
    item_id: int,
    body: dict = Body(...),
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    spec = _spec(section)
    row = _lay_dong(db, spec, item_id)
    _kiem_sua(db, current, row["staff_id"], section, tu_sua=spec["tu_sua"])
    try:
        gia_tri = hr.chuan_hoa(spec["fields"], body, mot_phan=True)
    except hr.LoiDuLieu as e:
        raise HTTPException(400, str(e))
    if not gia_tri:
        raise HTTPException(400, "Không có trường nào để cập nhật")
    # Trường bắt buộc đang có giá trị mà PATCH xoá trắng → chặn, nếu không dòng
    # hồ sơ sẽ mất tên/ngày mà vẫn nằm trong danh sách.
    for ten, v in gia_tri.items():
        if spec["fields"][ten]["bat_buoc"] and v in (None, ""):
            raise HTTPException(400, f"Thiếu {spec['fields'][ten]['nhan']}")
    db.execute(
        f"UPDATE {spec['table']} SET {','.join(k + '=?' for k in gia_tri)},"
        " updated_at=?, updated_by=? WHERE id = ?",
        (*gia_tri.values(), _vn_now(), current["id"], item_id),
    )
    write_audit(db, current["id"], f"hr.{section}.update", spec["table"], item_id,
                f"staff_id={row['staff_id']} — sửa: {', '.join(sorted(gia_tri))}")
    db.commit()
    return dict(_lay_dong(db, spec, item_id))


def _xoa_dinh_kem(db, section: str, item_id: int) -> int:
    """Xoá file của một dòng hồ sơ. `hr_attachments` trỏ tới chủ sở hữu bằng
    (section, item_id) — khoá ngoại đa hình nên SQLite không tự dọn theo được."""
    cur = db.execute("DELETE FROM hr_attachments WHERE section = ? AND item_id = ?",
                     (section, item_id))
    return cur.rowcount


@router.delete("/items/{section}/{item_id}")
def delete_item(
    section: str,
    item_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    spec = _spec(section)
    row = _lay_dong(db, spec, item_id)
    _kiem_sua(db, current, row["staff_id"], section, tu_sua=spec["tu_sua"])
    n_file = _xoa_dinh_kem(db, section, item_id)
    db.execute(f"DELETE FROM {spec['table']} WHERE id = ?", (item_id,))
    write_audit(db, current["id"], f"hr.{section}.delete", spec["table"], item_id,
                f"staff_id={row['staff_id']}, xoá kèm {n_file} file")
    db.commit()
    return {"ok": True}


# ── File đính kèm ────────────────────────────────────────────────────────────
@router.post("/attachments/{section}/{item_id}")
async def upload_attachment(
    section: str,
    item_id: int,
    file: UploadFile = File(...),
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    spec = _spec(section)
    if section not in hr.SECTIONS_CO_FILE:
        raise HTTPException(400, f"Phân hệ '{spec['nhan']}' không đính kèm file")
    row = _lay_dong(db, spec, item_id)
    _kiem_sua(db, current, row["staff_id"], section, tu_sua=spec["tu_sua"])
    ten, mime = _kieu_file(file.filename, _FILE_MO_RONG, "Tệp đính kèm")
    noi_dung = await read_limited(file, _TRAN_FILE, ten="Tệp đính kèm")
    cur = db.execute(
        """INSERT INTO hr_attachments (section, item_id, filename, mime, size_bytes,
                                       content, uploaded_by, uploaded_at)
           VALUES (?,?,?,?,?,?,?,?)""",
        (section, item_id, ten, mime, len(noi_dung), noi_dung,
         current["id"], _vn_now()),
    )
    write_audit(db, current["id"], "hr.attachment.upload", "hr_attachments",
                cur.lastrowid, f"{section}#{item_id}: {ten}")
    db.commit()
    return {"id": cur.lastrowid, "filename": ten, "mime": mime,
            "size_bytes": len(noi_dung)}


def _dinh_kem(db, att_id: int) -> sqlite3.Row:
    row = db.execute("SELECT * FROM hr_attachments WHERE id = ?", (att_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy file")
    return row


def _chu_so_huu(db, row: sqlite3.Row) -> int:
    spec = _spec(row["section"])
    owner = db.execute(f"SELECT staff_id FROM {spec['table']} WHERE id = ?",
                       (row["item_id"],)).fetchone()
    if not owner:
        raise HTTPException(404, "Dòng hồ sơ của file này không còn")
    return owner["staff_id"]


@router.get("/attachments/{att_id}/download")
def download_attachment(
    att_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    row = _dinh_kem(db, att_id)
    _kiem_xem(db, current, _chu_so_huu(db, row), row["section"])
    return Response(content=row["content"], media_type=row["mime"] or "application/pdf",
                    headers=_download_headers(row["filename"]))


@router.delete("/attachments/{att_id}")
def delete_attachment(
    att_id: int,
    current: dict = Depends(require_feature("menu.hr_profiles")),
    db: sqlite3.Connection = Depends(get_db),
):
    row = _dinh_kem(db, att_id)
    spec = _spec(row["section"])
    _kiem_sua(db, current, _chu_so_huu(db, row), row["section"], tu_sua=spec["tu_sua"])
    db.execute("DELETE FROM hr_attachments WHERE id = ?", (att_id,))
    write_audit(db, current["id"], "hr.attachment.delete", "hr_attachments", att_id,
                f"{row['section']}#{row['item_id']}: {row['filename']}")
    db.commit()
    return {"ok": True}


# ── Tra cứu, thống kê, nhắc lịch ─────────────────────────────────────────────
@router.get("/directory", response_model=list[DirectoryRow])
def directory(
    nhom: str = Query("tat_ca"),
    as_of: Optional[str] = Query(None, description="Tra cứu tại thời điểm, YYYY-MM-DD"),
    department_id: Optional[int] = Query(None),
    current: dict = Depends(require_feature("menu.hr_lookup")),
    db: sqlite3.Connection = Depends(get_db),
):
    if nhom not in hr.NHOM_TRA_CUU:
        raise HTTPException(400, f"Nhóm tra cứu không hợp lệ: {nhom}")
    try:
        moc = date.fromisoformat(as_of) if as_of else date.today()
    except ValueError:
        raise HTTPException(400, f"Ngày tra cứu không hợp lệ: {as_of}")
    return hr.tra_cuu_danh_sach(db, nhom, moc, department_id)


@router.get("/stats", response_model=StatsOut)
def stats(
    current: dict = Depends(require_feature("menu.hr_lookup")),
    db: sqlite3.Connection = Depends(get_db),
):
    return hr.tinh_thong_ke(db)


@router.get("/reminders", response_model=list[ReminderOut])
def reminders(
    loai: Optional[str] = Query(None, description="nang_luong | bo_nhiem_lai | cap_moi"),
    current: dict = Depends(require_feature("menu.hr_reminders")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Nhắc lịch nâng lương / bổ nhiệm lại / cấp công cụ mới.

    Chỉ người có `hr.view_all` mới thấy toàn cơ quan; còn lại thấy đúng phần
    của mình — cùng nguyên tắc với danh sách hồ sơ.
    """
    ra = hr.tinh_nhac_lich(db)
    if loai:
        ra = [x for x in ra if x["loai"] == loai]
    if not _co_quyen(db, current, "hr.view_all"):
        ra = [x for x in ra if x["staff_id"] == current["id"]]
    return ra


# ── Xuất Excel danh sách hồ sơ ───────────────────────────────────────────────
_COT_XUAT = [
    ("employee_code", "Mã cán bộ", 14),
    ("full_name",     "Họ và tên", 26),
    ("department",    "Phòng", 26),
    ("chuc_vu",       "Chức vụ", 20),
    ("quy_hoach",     "Quy hoạch", 20),
    ("gender",        "Giới tính", 10),
    ("dob",           "Ngày sinh", 13),
]


def _dung_workbook(rows: list[dict], tieu_de: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách"
    ws.cell(row=1, column=1, value=tieu_de).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COT_XUAT) + 1)
    fill = PatternFill("solid", fgColor="FEE2E2")
    for j, nhan in enumerate(["STT"] + [n for _, n, _ in _COT_XUAT], start=1):
        c = ws.cell(row=3, column=j, value=nhan)
        c.font, c.fill = Font(bold=True), fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 6
    for j, (_, _, w) in enumerate(_COT_XUAT, start=2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    for i, r in enumerate(rows, start=1):
        ws.cell(row=3 + i, column=1, value=i)
        for j, (f, _, _) in enumerate(_COT_XUAT, start=2):
            ws.cell(row=3 + i, column=j, value=r.get(f))
    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export")
async def export_directory(
    nhom: str = Query("tat_ca"),
    as_of: Optional[str] = Query(None),
    department_id: Optional[int] = Query(None),
    current: dict = Depends(require_feature("hr.export")),
    db: sqlite3.Connection = Depends(get_db),
):
    rows = directory(nhom=nhom, as_of=as_of, department_id=department_id,
                     current=current, db=db)
    moc = as_of or date.today().isoformat()
    noi_dung = await run_heavy(
        _dung_workbook, rows, f"DANH SÁCH CÁN BỘ — {hr.NHOM_TRA_CUU[nhom].upper()} "
                              f"(tại ngày {moc})")
    write_audit(db, current["id"], "hr.export", "hr_profile", None,
                f"{len(rows)} cán bộ (nhóm={nhom}, tại ngày {moc})")
    db.commit()
    return Response(
        content=noi_dung,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_download_headers(f"danh_sach_can_bo_{moc.replace('-', '')}.xlsx"),
    )
