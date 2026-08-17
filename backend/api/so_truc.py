"""API Sổ trực cuối ngày — Phòng Thanh toán.

Xem docstring luồng nghiệp vụ đầy đủ ở backend/services/so_truc_service.py.

Phân quyền 2 lớp:
- `require_feature("menu.so_truc")` — lớp NGOÀI, chỉ xác định ai được VÀO
  module này (đọc lịch sử, xem 1 ngày, xuất hiện trong danh sách GDV/KSV để
  CHỌN). Xem lịch sử KHÔNG giới hạn thêm — mọi người có feature này đều xem
  được mọi ngày.
- Ràng buộc nghiệp vụ "đúng người trực mới được sửa/xác nhận" — lớp TRONG,
  nằm trong service (`NotAllowedError`), áp dụng cho THAO TÁC (save-draft,
  forward-ksv, gdv-ack, ksv-confirm, ksv-reject, ksv-cancel, request-edit,
  ksv-finalize-edit): chỉ đúng 2 tài khoản gdv1_id/gdv2_id (một khi đã
  chọn) và đúng tài khoản ksv_id mới làm được, KHÔNG phải bất kỳ ai có
  feature. `NotAllowedError` map sang 403.
"""
from __future__ import annotations

import datetime
import io

from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.api.bundles import _download_headers
from backend.database import get_db
from backend.core.deps import require_feature
from backend.schemas.so_truc import ForwardKsvIn, RejectIn, SaveDraftIn
from backend.services import so_truc_service as svc
from backend.services.so_truc_service import NotAllowedError

router = APIRouter(prefix="/api/so-truc", tags=["so-truc"])


def _iso_to_ddmmyyyy(iso: str) -> str:
    """'2026-08-17' -> '17/08/2026' — dùng cho tiêu đề + cột "Ngày trực"
    trong file Excel xuất ra, khớp định dạng dd/mm/yyyy dùng thống nhất ở
    frontend (trước đây Excel xuất thẳng ISO, lệch với cách hiển thị trên
    web)."""
    y, m, d = iso.split("-")
    return f"{d}/{m}/{y}"


def _iso_dt_to_ddmmyyyy(dt_str) -> str:
    """'2026-08-17 10:35:45.031208' -> '17/08/2026 10:35:45.031208' — chỉ
    đổi phần NGÀY sang dd/mm/yyyy, giữ nguyên giờ:phút:giây(.micro) — dùng
    cho cột "Cập nhật lúc" trong Excel xuất ra (cùng lý do với
    `_iso_to_ddmmyyyy()`, cột này cũng đang lưu/hiện ISO)."""
    if not dt_str:
        return dt_str
    date_part, _, time_part = str(dt_str).partition(" ")
    try:
        y, m, d = date_part.split("-")
    except ValueError:
        return dt_str
    return f"{d}/{m}/{y} {time_part}".rstrip()


def _validate_truc_date(truc_date: str) -> None:
    """Chặn `truc_date` sai định dạng NGAY TẠI API — trước đây không chặn ở
    đâu cả, `POST /so-truc/abc/save-draft` tạo được bản ghi truc_date='abc'
    trong DB, sau đó `_iso_to_ddmmyyyy()` ở frontend (tab Lịch sử) làm
    `y, m, d = iso.split("-")` cho MỌI bản ghi trả về → ValueError, hỏng cả
    tab Lịch sử cho mọi người xem, không chỉ người gọi API sai."""
    try:
        datetime.date.fromisoformat(truc_date)
    except ValueError:
        raise HTTPException(422, f"Ngày trực không hợp lệ: '{truc_date}' (đúng định dạng YYYY-MM-DD)")


@router.get("/history")
def get_history(
    tu_ngay: str | None = None,
    den_ngay: str | None = None,
    db=Depends(get_db),
    current: dict = Depends(require_feature("menu.so_truc")),
):
    """tu_ngay/den_ngay dạng YYYY-MM-DD. Để cả 2 trống → chỉ trả về 1 phiên
    trực gần nhất (xem get_history() trong service); truyền ít nhất 1 mốc
    ngày → trả đầy đủ các dòng khớp bộ lọc."""
    return svc.get_history(db, tu_ngay, den_ngay)


# ── Xuất Excel lịch sử ────────────────────────────────────────────────────────
# Đăng ký TRƯỚC route catch-all "/{truc_date}" bên dưới — Starlette khớp path
# param theo thứ tự đăng ký, để sau "/export" sẽ bị nuốt thành truc_date="export".
_EXPORT_COLS = [
    ("truc_date",  "Ngày trực",  14),
    ("gdv1_name",  "GDV 1",      20),
    ("gdv2_name",  "GDV 2",      20),
    ("truc_phu_names", "Trực phụ", 24),
    ("ksv_name",   "KSV",        20),
    ("status",     "Trạng thái", 18),
    ("ghi_chu",    "Ghi chú",    30),
    ("reject_reason", "Lý do từ chối/huỷ gần nhất", 30),
    ("updated_at", "Cập nhật lúc", 20),
]
_STATUS_VI = {
    "draft": "Đang lập",
    "pending_ksv": "Chờ KSV duyệt", "approved": "Hoàn thành", "cancelled": "Đã huỷ",
}


def _build_history_workbook(rows: list[dict], tu_ngay: str, den_ngay: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lich su so truc"

    ws.cell(
        row=1, column=1,
        value=f"LỊCH SỬ SỔ TRỰC CUỐI NGÀY ({_iso_to_ddmmyyyy(tu_ngay)} → {_iso_to_ddmmyyyy(den_ngay)})",
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_EXPORT_COLS))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    hdr_fill = PatternFill("solid", fgColor="DBEAFE")
    for j, (_, label, width) in enumerate(_EXPORT_COLS, start=1):
        c = ws.cell(row=3, column=j, value=label)
        c.font, c.fill = Font(bold=True), hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = width

    for i, r in enumerate(rows, start=4):
        for j, (field, _, _) in enumerate(_EXPORT_COLS, start=1):
            val = r.get(field)
            if field == "truc_date":
                val = _iso_to_ddmmyyyy(val)
            elif field == "updated_at":
                val = _iso_dt_to_ddmmyyyy(val)
            elif field == "status":
                val = _STATUS_VI.get(val, val)
            elif field == "truc_phu_names":
                val = ", ".join(val) if val else "—"
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=field in ("ghi_chu", "reject_reason", "truc_phu_names"))

    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export")
def export_history(
    tu_ngay: str, den_ngay: str,
    db=Depends(get_db), current: dict = Depends(require_feature("menu.so_truc")),
):
    """Bắt buộc truyền khoảng ngày (khác /history — không có mặc định LIMIT 1,
    xuất Excel mà chỉ ra 1 dòng thì gây hiểu nhầm là lỗi)."""
    if not tu_ngay or not den_ngay:
        raise HTTPException(400, "Phải chọn khoảng ngày (Từ ngày / Đến ngày) để xuất Excel")
    rows = svc.get_history(db, tu_ngay, den_ngay)
    content = _build_history_workbook(rows, tu_ngay, den_ngay)
    fname = f"so_truc_lich_su_{tu_ngay}_{den_ngay}.xlsx".replace("-", "")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_download_headers(fname),
    )


@router.get("/gdv-candidates")
def get_gdv_candidates(
    db=Depends(get_db), current: dict = Depends(require_feature("menu.so_truc"))
):
    return svc.list_gdv_candidates(db)


@router.get("/ksv-candidates")
def get_ksv_candidates(
    db=Depends(get_db), current: dict = Depends(require_feature("menu.so_truc"))
):
    return svc.list_ksv_candidates(db)


@router.get("/{truc_date}/citad-status")
def get_citad_status(
    truc_date: str, db=Depends(get_db), current: dict = Depends(require_feature("menu.so_truc"))
):
    """Đối chiếu CITAD ngày này đã có bản lưu khớp chưa — frontend gọi
    TRƯỚC khi GDV/KSV bấm xác nhận để cảnh báo nếu chưa xong (xem
    so_truc_service.check_citad_status). Đăng ký TRƯỚC route trần
    "/{truc_date}" bên dưới — cùng lý do path converter khớp theo thứ tự
    đăng ký đã ghi trong doi_chieu_citad.py."""
    _validate_truc_date(truc_date)
    return svc.check_citad_status(db, truc_date)


@router.get("/{truc_date}")
def get_so_truc(
    truc_date: str, db=Depends(get_db), current: dict = Depends(require_feature("menu.so_truc"))
):
    _validate_truc_date(truc_date)
    return svc.get_active_by_date(db, truc_date) or {
        "truc_date": truc_date, "status": "draft",
        "gdv1_id": None, "gdv2_id": None, "gdv1_name": "", "gdv2_name": "", "ghi_chu": "",
        "truc_phu_ids": [], "truc_phu_names": [],
    }


@router.post("/{truc_date}/save-draft")
def save_draft(
    truc_date: str, data: SaveDraftIn, db=Depends(get_db),
    current: dict = Depends(require_feature("menu.so_truc")),
):
    _validate_truc_date(truc_date)
    try:
        return svc.save_draft(
            db, truc_date, current["id"], data.gdv1_id, data.gdv2_id, data.ghi_chu, data.truc_phu_ids,
        )
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/draft-cancel")
def draft_cancel(
    truc_date: str, data: RejectIn, db=Depends(get_db),
    current: dict = Depends(require_feature("menu.so_truc")),
):
    """1 trong 2 GDV huỷ hẳn phiên trực đang lập (kể cả sau khi bị GDV kia
    "Từ chối xác nhận") — xem docstring draft_cancel() trong service."""
    _validate_truc_date(truc_date)
    if not data.reason.strip():
        raise HTTPException(400, "Phải nhập lý do huỷ")
    try:
        return svc.draft_cancel(db, truc_date, current["id"], data.reason.strip())
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/forward-ksv")
def forward_ksv(
    truc_date: str, data: ForwardKsvIn, db=Depends(get_db),
    current: dict = Depends(require_feature("menu.so_truc")),
):
    _validate_truc_date(truc_date)
    try:
        return svc.forward_to_ksv(
            db, truc_date, current["id"], data.gdv1_id, data.gdv2_id, data.ghi_chu, data.ksv_id,
            data.truc_phu_ids,
        )
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/gdv-ack")
def gdv_ack(
    truc_date: str, db=Depends(get_db), current: dict = Depends(require_feature("menu.so_truc"))
):
    """GDV còn lại bấm "Xác nhận phiên trực" — không chặn/không đổi trạng
    thái gì, chỉ ghi nhận để hiện dấu tick khi xem lại (xem docstring
    gdv_ack() trong service)."""
    _validate_truc_date(truc_date)
    try:
        return svc.gdv_ack(db, truc_date, current["id"])
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/ksv-confirm")
def ksv_confirm(
    truc_date: str, db=Depends(get_db),
    current: dict = Depends(require_feature("so_truc.ksv_confirm")),
):
    _validate_truc_date(truc_date)
    try:
        return svc.ksv_confirm(db, truc_date, current["id"])
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/ksv-reject")
def ksv_reject(
    truc_date: str, data: RejectIn, db=Depends(get_db),
    current: dict = Depends(require_feature("so_truc.ksv_confirm")),
):
    """"Từ chối để sửa" — quay lại draft, 2 GDV sửa rồi đẩy lại ĐÚNG KSV này."""
    _validate_truc_date(truc_date)
    if not data.reason.strip():
        raise HTTPException(400, "Phải nhập lý do từ chối")
    try:
        return svc.ksv_reject(db, truc_date, current["id"], data.reason.strip())
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/ksv-cancel")
def ksv_cancel(
    truc_date: str, data: RejectIn, db=Depends(get_db),
    current: dict = Depends(require_feature("so_truc.ksv_confirm")),
):
    """"Từ chối để huỷ" — chỉ ĐỀ NGHỊ huỷ, quay về 'draft' (như "để sửa"),
    KHÔNG tự đóng phiên. Phải để 1 trong 2 GDV tự bấm "Huỷ phiên trực"
    (draft-cancel) mới thật sự thành 'cancelled' — xem docstring
    ksv_cancel() trong so_truc_service.py."""
    _validate_truc_date(truc_date)
    if not data.reason.strip():
        raise HTTPException(400, "Phải nhập lý do huỷ")
    try:
        return svc.ksv_cancel(db, truc_date, current["id"], data.reason.strip())
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/request-edit")
def request_edit(
    truc_date: str, data: RejectIn, db=Depends(get_db),
    current: dict = Depends(require_feature("menu.so_truc")),
):
    """1 trong 2 GDV chính hoặc đúng KSV của phiên đã "Hoàn thành" mở lại
    để sửa — chỉ dùng ở tab Lịch sử. Gọi được bởi cả GDV lẫn KSV nên gác
    bằng `menu.so_truc` (rộng), service tự phân biệt đúng người + nhánh xử
    lý (xem docstring request_edit() trong so_truc_service.py)."""
    _validate_truc_date(truc_date)
    if not data.reason.strip():
        raise HTTPException(400, "Phải nhập lý do yêu cầu chỉnh sửa")
    try:
        return svc.request_edit(db, truc_date, current["id"], data.reason.strip())
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/{truc_date}/ksv-finalize-edit")
def ksv_finalize_edit(
    truc_date: str, data: SaveDraftIn, db=Depends(get_db),
    current: dict = Depends(require_feature("so_truc.ksv_confirm")),
):
    """KSV tự sửa xong sau khi tự mở lại (`request_edit`, `ksv_decision=
    'self_edit'`) — lưu thẳng thành 'approved', không qua forward/confirm
    lại vòng nữa."""
    _validate_truc_date(truc_date)
    try:
        return svc.ksv_finalize_edit(
            db, truc_date, current["id"], data.gdv1_id, data.gdv2_id, data.ghi_chu, data.truc_phu_ids,
        )
    except NotAllowedError as e:
        raise HTTPException(403, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))
