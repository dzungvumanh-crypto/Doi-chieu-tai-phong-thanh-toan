"""Ôn tập trắc nghiệm (Quizz) — bộ câu hỏi dùng chung, thi thử, xem lại bài.

Bộ câu hỏi nhập từ Excel **một lần**, lưu thẳng vào DB, mọi người sau đó chỉ
việc chọn bộ có sẵn để ôn — đúng yêu cầu "người sau không phải tải lên lại".

Chấm điểm luôn đọc `correct_no` từ bảng `quiz_questions`, **không** nhận từ
client. Client chỉ gửi lên "tôi chọn ô số mấy".
"""
import hashlib
import io
import json
import logging
import random
import sqlite3
import unicodedata
from typing import Optional
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response

from backend.core.concurrency import run_heavy
from backend.core.deps import require_feature
from backend.core.enums import StaffRole
from backend.core.uploads import read_limited
from backend.database import get_db, write_audit, _vn_now
from backend.schemas.quiz import (
    AttemptCreate, AttemptOut, AttemptResult, AttemptRow, AttemptSettings,
    AttemptSubmit, ImportResult, ProgressIn, QuestionOut, QuizSetOut, QuizSetUpdate,
    ResumeRow, ReviewItem,
)

_log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/quiz", tags=["quiz"])

# Trần một bộ. 550 câu là cỡ file thật đang dùng; 5.000 là chặn file rác /
# file nhầm định dạng chứ không phải giới hạn nghiệp vụ.
_MAX_QUESTIONS = 5000
_MAX_UPLOAD = 20 * 1024 * 1024


def _txt(v) -> str:
    """Ô Excel → chuỗi NFC đã strip. Ô trống / khoảng trắng → ''."""
    if v is None:
        return ""
    return unicodedata.normalize("NFC", str(v)).strip()


def _hash_questions(rows: list[tuple]) -> str:
    """Vân tay của bộ câu hỏi — băm NỘI DUNG đã đọc, không băm byte của file.

    Băm file thất bại ngay ở ca thường gặp nhất: mở file ra xem rồi bấm lưu.
    Excel (và cả openpyxl) nhúng dấu thời gian vào `docProps/core.xml`, nên
    cùng một bộ câu hỏi lưu hai lần cho ra hai chuỗi byte khác nhau → hai mã
    băm khác nhau → chặn trùng lặp coi như không có. Băm nội dung thì đổi tên
    file, lưu lại, xoá bớt cột trống đều không ảnh hưởng.

    Ký tự phân cách dùng US (0x1f) và RS (0x1e) — không xuất hiện trong ô Excel,
    nên không có chuyện hai bộ khác nhau ghép chuỗi lại thành giống nhau.
    """
    h = hashlib.sha256()
    for r in rows:
        h.update("\x1f".join("" if v is None else str(v) for v in r).encode("utf-8"))
        h.update(b"\x1e")
    return h.hexdigest()


def _download_headers(filename: str) -> dict:
    fallback = "".join(ch if ord(ch) < 128 and ch not in '\\"' else "_" for ch in filename)
    return {
        "Content-Disposition": (
            f'attachment; filename="{fallback}"; '
            f"filename*=UTF-8''{quote(filename, safe='')}"
        )
    }


# ─── Đọc file Excel ──────────────────────────────────────────────────────────
def _parse_xlsx(raw: bytes) -> tuple[list[tuple], list[str], int]:
    """bytes → (danh sách câu hợp lệ, danh sách lỗi, số dòng bị bỏ).

    Quy ước cột theo file mẫu: A=Câu hỏi, B..E=Đáp án 1..4, F=Đáp án đúng (1-4).
    Dòng 1 là tiêu đề nếu ô A chứa chữ "câu hỏi" — file do người dùng tự gõ có
    thể không có tiêu đề, nên dò chứ không bỏ cứng dòng đầu.

    Dòng sai bị BỎ QUA kèm thông báo chỉ rõ số dòng, không huỷ cả file: một bộ
    550 câu mà hỏng ở câu 300 thì bắt sửa xong mới được nhập là quá đắt. Nhưng
    tuyệt đối không "đoán" đáp án đúng — sai một câu là người học nhớ sai.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Không đọc được file Excel: {e}")
    ws = wb.worksheets[0]

    rows: list[tuple] = []
    errors: list[str] = []
    skipped = 0
    for idx, r in enumerate(ws.iter_rows(min_row=1, max_col=6, values_only=True), start=1):
        cells = list(r) + [None] * (6 - len(r))
        content = _txt(cells[0])
        if idx == 1 and "câu hỏi" in content.casefold():
            continue                                   # dòng tiêu đề
        opts = [_txt(cells[i]) for i in range(1, 5)]
        correct_raw = _txt(cells[5])
        if not content and not any(opts) and not correct_raw:
            continue                                   # dòng trống — bỏ im lặng

        if not content:
            skipped += 1
            errors.append(f"Dòng {idx}: thiếu nội dung câu hỏi")
            continue
        if sum(1 for o in opts if o) < 2:
            skipped += 1
            errors.append(f"Dòng {idx}: phải có ít nhất 2 đáp án")
            continue
        # "2", "2.0" (ô số của Excel) đều nhận; "B" thì không — file mẫu ghi số.
        try:
            correct = int(float(correct_raw))
        except ValueError:
            skipped += 1
            errors.append(
                f"Dòng {idx}: cột 'Đáp án đúng' phải là số 1-4, đang là '{correct_raw}'")
            continue
        if not 1 <= correct <= 4 or not opts[correct - 1]:
            skipped += 1
            errors.append(f"Dòng {idx}: đáp án đúng số {correct} không tồn tại hoặc để trống")
            continue

        rows.append((content, opts[0] or None, opts[1] or None,
                     opts[2] or None, opts[3] or None, correct))
        if len(rows) > _MAX_QUESTIONS:
            raise HTTPException(400, f"Bộ câu hỏi vượt quá {_MAX_QUESTIONS} câu")
    wb.close()
    return rows, errors, skipped


def _resumable_rows(db, staff_id: int) -> list[sqlite3.Row]:
    """Các bài đang làm dở của một người — mới lưu trước, cũ sau.

    "Dở" = `status = 'in_progress'`, không phân biệt tạm dừng chủ động với rớt
    mạng: với người dùng cả hai đều là "bài còn đó, vào làm tiếp". Phân biệt
    được cũng không tin nổi — máy tắt đột ngột thì không ai kịp ghi cờ nào.
    """
    return db.execute(
        """SELECT a.id, a.set_id, s.name AS set_name, a.mode, a.total_questions,
                  a.current_idx, a.elapsed_ms, a.started_at, a.saved_at,
                  (SELECT COUNT(*) FROM quiz_attempt_items i
                    WHERE i.attempt_id = a.id AND i.chosen_no IS NOT NULL) AS answered
             FROM quiz_attempts a
             JOIN quiz_sets s ON s.id = a.set_id AND s.is_active = 1
            WHERE a.staff_id = ? AND a.status = 'in_progress'
            ORDER BY IFNULL(a.saved_at, a.started_at) DESC""",
        (staff_id,),
    ).fetchall()


# ─── Danh sách bộ câu hỏi ────────────────────────────────────────────────────
@router.get("/sets", response_model=list[QuizSetOut])
def list_sets(
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    rows = db.execute(
        """SELECT s.*, u.full_name AS created_by_name,
                  (SELECT COUNT(*) FROM quiz_attempts a
                    WHERE a.set_id = s.id AND a.staff_id = ? AND a.status = 'finished')
                      AS my_attempts,
                  (SELECT MAX(a.score) FROM quiz_attempts a
                    WHERE a.set_id = s.id AND a.staff_id = ? AND a.status = 'finished')
                      AS my_best_score
             FROM quiz_sets s
             LEFT JOIN user_tttt u ON u.id = s.created_by
            WHERE s.is_active = 1
            ORDER BY s.created_at DESC""",
        (current["id"], current["id"]),
    ).fetchall()
    # Bài đang làm dở — truy vấn riêng, không nhét thêm vào SELECT ở trên:
    # gộp vào sẽ cần ba truy vấn con nữa cho MỖI bộ, trong khi số bài dở của
    # một người luôn đếm trên đầu ngón tay (mỗi bộ nhiều nhất một bài).
    #
    # `setdefault` chứ không phải `{r["set_id"]: r for r in ...}`: dict
    # comprehension giữ dòng CUỐI, mà truy vấn sắp mới-trước-cũ-sau nên dòng
    # cuối là bài cũ nhất — nút "Làm tiếp" sẽ nối vào đúng bài người dùng đã
    # bỏ lâu nhất. Bất biến "mỗi bộ một bài dở" chỉ đúng với bài tạo từ khi có
    # tính năng này; DB đã chạy từ trước có thể còn nhiều bài dở cùng bộ.
    resume: dict[int, sqlite3.Row] = {}
    for r in _resumable_rows(db, current["id"]):
        resume.setdefault(r["set_id"], r)
    out = []
    for r in rows:
        d = dict(r)
        d["created_at"] = str(d["created_at"]) if d.get("created_at") else None
        rs = resume.get(d["id"])
        if rs:
            d["resume_attempt_id"] = rs["id"]
            d["resume_answered"] = rs["answered"]
            d["resume_total"] = rs["total_questions"]
            d["resume_saved_at"] = rs["saved_at"]
        out.append(d)
    return out


@router.post("/sets/upload", response_model=ImportResult)
async def upload_set(
    file: UploadFile = File(...),
    name: Optional[str] = Query(None, description="Tên bộ; bỏ trống thì lấy tên file"),
    description: Optional[str] = Query(None),
    current: dict = Depends(require_feature("quiz.upload")),
    db: sqlite3.Connection = Depends(get_db),
):
    raw = await read_limited(file, _MAX_UPLOAD, "File bộ câu hỏi")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Chỉ nhận file Excel .xlsx")

    set_name = unicodedata.normalize("NFC", (name or "").strip())
    if not set_name:
        set_name = (file.filename or "Bộ câu hỏi").rsplit(".", 1)[0].strip()
    if not set_name:
        raise HTTPException(400, "Không xác định được tên bộ câu hỏi")

    if db.execute("SELECT 1 FROM quiz_sets WHERE name = ?", (set_name,)).fetchone():
        raise HTTPException(409, f"Đã có bộ câu hỏi tên '{set_name}'. Đặt tên khác hoặc xoá bộ cũ.")

    rows, errors, skipped = await run_heavy(_parse_xlsx, raw)
    if not rows:
        raise HTTPException(
            400,
            "Không có câu hỏi hợp lệ nào trong file. "
            + ("; ".join(errors[:5]) if errors else "Kiểm tra lại thứ tự cột: "
               "Câu hỏi | Đáp án 1 | Đáp án 2 | Đáp án 3 | Đáp án 4 | Đáp án đúng"),
        )

    # Cùng nội dung nhưng khác tên vẫn chặn: mục đích của màn này là mỗi bộ chỉ
    # tải lên MỘT lần, hai bản sao chỉ làm bảng xếp hạng tách đôi vô nghĩa.
    content_hash = _hash_questions(rows)
    dup = db.execute(
        "SELECT name FROM quiz_sets WHERE content_hash = ? AND is_active = 1", (content_hash,)
    ).fetchone()
    if dup:
        raise HTTPException(409, f"Bộ câu hỏi này đã có sẵn với tên '{dup['name']}'")

    now = _vn_now()
    cur = db.execute(
        """INSERT INTO quiz_sets (name, description, source_file, content_hash,
                                  question_count, created_by, created_at, is_active)
           VALUES (?,?,?,?,?,?,?,1)""",
        (set_name, (description or "").strip() or None, file.filename, content_hash,
         len(rows), current["id"], now),
    )
    set_id = cur.lastrowid
    db.executemany(
        """INSERT INTO quiz_questions (set_id, order_no, content, opt1, opt2, opt3, opt4, correct_no)
           VALUES (?,?,?,?,?,?,?,?)""",
        [(set_id, i, *r) for i, r in enumerate(rows, start=1)],
    )
    write_audit(db, current["id"], "quiz.upload_set", "quiz_set", set_id,
                f"{set_name} — {len(rows)} câu, bỏ qua {skipped}")
    db.commit()
    # Chỉ trả 20 lỗi đầu: file hỏng cột có thể sinh hàng trăm dòng lỗi giống hệt
    # nhau, đổ hết lên màn hình thì người dùng không đọc nổi dòng nào.
    return {"set_id": set_id, "name": set_name, "imported": len(rows),
            "skipped": skipped, "errors": errors[:20]}


@router.patch("/sets/{set_id}")
def rename_set(
    set_id: int,
    body: QuizSetUpdate,
    current: dict = Depends(require_feature("quiz.upload")),
    db: sqlite3.Connection = Depends(get_db),
):
    row = db.execute("SELECT * FROM quiz_sets WHERE id = ? AND is_active = 1", (set_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy bộ câu hỏi")
    clash = db.execute(
        "SELECT 1 FROM quiz_sets WHERE name = ? AND id <> ?", (body.name, set_id)
    ).fetchone()
    if clash:
        raise HTTPException(409, f"Đã có bộ câu hỏi tên '{body.name}'")
    db.execute("UPDATE quiz_sets SET name = ?, description = ? WHERE id = ?",
               (body.name, body.description, set_id))
    write_audit(db, current["id"], "quiz.rename_set", "quiz_set", set_id,
                f"{row['name']} → {body.name}")
    db.commit()
    return {"ok": True}


@router.delete("/sets/{set_id}")
def delete_set(
    set_id: int,
    current: dict = Depends(require_feature("quiz.delete")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Xoá HẲN — câu hỏi, mọi lượt làm bài và bảng xếp hạng của bộ đi theo (FK CASCADE)."""
    row = db.execute("SELECT * FROM quiz_sets WHERE id = ?", (set_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy bộ câu hỏi")
    n_att = db.execute(
        "SELECT COUNT(*) c FROM quiz_attempts WHERE set_id = ?", (set_id,)
    ).fetchone()["c"]
    db.execute("DELETE FROM quiz_sets WHERE id = ?", (set_id,))
    write_audit(db, current["id"], "quiz.delete_set", "quiz_set", set_id,
                f"{row['name']} — {row['question_count']} câu, {n_att} lượt làm bài")
    db.commit()
    return {"ok": True, "deleted_attempts": n_att}


@router.get("/template")
def download_template(_: dict = Depends(require_feature("quiz.upload"))):
    """File Excel mẫu — đúng thứ tự cột mà _parse_xlsx() chờ đợi."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Câu hỏi", "Đáp án 1", "Đáp án 2", "Đáp án 3", "Đáp án 4", "Đáp án đúng"])
    ws.append(["Trụ sở chính của Agribank đặt tại đâu?", "Hà Nội", "Đà Nẵng",
               "TP. Hồ Chí Minh", "Cần Thơ", 1])
    ws.append(["Bỏ trống Đáp án 4 nếu câu chỉ có 3 lựa chọn", "Lựa chọn A",
               "Lựa chọn B", "Lựa chọn C", None, 2])
    for col, width in zip("ABCDEF", (70, 30, 30, 30, 30, 12)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_download_headers("Mau_bo_cau_hoi.xlsx"),
    )


# ─── Lượt làm bài ────────────────────────────────────────────────────────────
def _options_of(q: sqlite3.Row, order: list[int]) -> list[dict]:
    """[{no, text}] theo đúng thứ tự đã lưu. `no` là số thứ tự GỐC (1-4)."""
    return [{"no": n, "text": q[f"opt{n}"]} for n in order]


def _questions_of_attempt(db, attempt_id: int, reveal: bool) -> list[QuestionOut]:
    """Dựng lại đề của một lượt từ quiz_attempt_items.

    `reveal` mở `correct_no` cho chế độ ôn tập (hiện đúng/sai ngay sau khi chọn).
    Chế độ thi thử để False — đáp án chỉ về client sau khi nộp bài. Frontend là
    NiceGUI chạy phía máy chủ nên giá trị này không rời khỏi server, nhưng vẫn
    giữ đúng ranh giới để API còn dùng được cho client khác.
    """
    rows = db.execute(
        """SELECT i.id AS item_id, i.question_id, i.order_no, i.option_order,
                  i.chosen_no, q.content, q.opt1, q.opt2, q.opt3, q.opt4, q.correct_no
             FROM quiz_attempt_items i
             JOIN quiz_questions q ON q.id = i.question_id
            WHERE i.attempt_id = ?
            ORDER BY i.order_no""",
        (attempt_id,),
    ).fetchall()
    out = []
    for r in rows:
        order = json.loads(r["option_order"])
        out.append(QuestionOut(
            item_id=r["item_id"], question_id=r["question_id"], order_no=r["order_no"],
            content=r["content"], options=_options_of(r, order),
            correct_no=r["correct_no"] if reveal else None,
            chosen_no=r["chosen_no"],
        ))
    return out


@router.post("/attempts", response_model=AttemptOut)
def start_attempt(
    body: AttemptCreate,
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    s = db.execute(
        "SELECT * FROM quiz_sets WHERE id = ? AND is_active = 1", (body.set_id,)
    ).fetchone()
    if not s:
        raise HTTPException(404, "Không tìm thấy bộ câu hỏi")
    qs = db.execute(
        "SELECT * FROM quiz_questions WHERE set_id = ? ORDER BY order_no", (body.set_id,)
    ).fetchall()
    if not qs:
        raise HTTPException(400, "Bộ câu hỏi này chưa có câu nào")

    cfg = body.settings
    picked = list(qs)
    if cfg.shuffle_questions:
        random.shuffle(picked)
    # num_questions=0 hoặc lớn hơn số câu có thật → lấy hết. Không lặp lại câu để
    # cho đủ số: đề 20 câu mà bộ chỉ có 12 thì người học gặp lại câu cũ, vô nghĩa.
    if cfg.num_questions:
        picked = picked[: cfg.num_questions]

    # Bắt đầu bài mới cho cùng một bộ = người dùng đã quyết định bỏ bài dở cũ.
    # Xoá luôn, để mỗi người mỗi bộ nhiều nhất MỘT bài dở: nếu tích lại thì nút
    # "Làm tiếp" không biết nối vào bài nào, và bài dở bỏ quên nằm lại DB mãi.
    db.execute(
        "DELETE FROM quiz_attempts WHERE staff_id = ? AND set_id = ? AND status = 'in_progress'",
        (current["id"], body.set_id),
    )

    now = _vn_now()
    cur = db.execute(
        """INSERT INTO quiz_attempts (set_id, staff_id, mode, settings, total_questions,
                                      status, started_at, saved_at)
           VALUES (?,?,?,?,?, 'in_progress', ?, ?)""",
        (body.set_id, current["id"], cfg.mode, cfg.model_dump_json(), len(picked), now, now),
    )
    attempt_id = cur.lastrowid

    items = []
    for i, q in enumerate(picked, start=1):
        order = [n for n in (1, 2, 3, 4) if q[f"opt{n}"]]
        if cfg.shuffle_options:
            random.shuffle(order)
        items.append((attempt_id, q["id"], i, json.dumps(order)))
    db.executemany(
        """INSERT INTO quiz_attempt_items (attempt_id, question_id, order_no, option_order)
           VALUES (?,?,?,?)""",
        items,
    )
    db.commit()

    return AttemptOut(
        id=attempt_id, set_id=body.set_id, set_name=s["name"], mode=cfg.mode,
        status="in_progress", settings=cfg, total_questions=len(picked),
        started_at=str(now),
        questions=_questions_of_attempt(db, attempt_id, reveal=(cfg.mode == "practice")),
    )


# ─── Tạm dừng & làm tiếp ─────────────────────────────────────────────────────
# `/attempts/resumable` PHẢI đứng trên `/attempts/{attempt_id}`: FastAPI khớp
# route theo thứ tự đăng ký, để sau thì chữ "resumable" rơi vào {attempt_id},
# ép sang int thất bại và trả 422 — nút *Làm tiếp* chết mà không rõ vì sao.
@router.get("/attempts/resumable", response_model=list[ResumeRow])
def list_resumable(
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Bài đang làm dở của tôi — nguồn cho nút *Làm tiếp*."""
    return [
        {**dict(r),
         "started_at": str(r["started_at"]) if r["started_at"] else None,
         "saved_at": str(r["saved_at"]) if r["saved_at"] else None}
        for r in _resumable_rows(db, current["id"])
    ]


@router.get("/attempts/{attempt_id}", response_model=AttemptOut)
def get_attempt(
    attempt_id: int,
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Nạp lại đề đang làm dở — trang làm bài là route riêng, F5 không được mất đề."""
    a = db.execute(
        """SELECT a.*, s.name AS set_name FROM quiz_attempts a
             JOIN quiz_sets s ON s.id = a.set_id WHERE a.id = ?""",
        (attempt_id,),
    ).fetchone()
    if not a:
        raise HTTPException(404, "Không tìm thấy lượt làm bài")
    if a["staff_id"] != current["id"]:
        raise HTTPException(403, "Đây không phải bài làm của bạn")
    cfg = AttemptSettings(**json.loads(a["settings"] or "{}"))
    reveal = a["status"] == "finished" or cfg.mode == "practice"
    return AttemptOut(
        id=a["id"], set_id=a["set_id"], set_name=a["set_name"], mode=a["mode"],
        status=a["status"], settings=cfg, total_questions=a["total_questions"],
        current_idx=a["current_idx"] or 0, elapsed_ms=a["elapsed_ms"] or 0,
        started_at=str(a["started_at"]) if a["started_at"] else None,
        questions=_questions_of_attempt(db, attempt_id, reveal=reveal),
    )


@router.patch("/attempts/{attempt_id}/progress")
def save_progress(
    attempt_id: int,
    body: ProgressIn,
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Ghi lại tiến độ giữa chừng — gọi sau MỖI câu trả lời, không đợi nộp bài.

    KHÔNG chấm điểm ở đây: `is_correct` để nguyên NULL, chỉ ghi `chosen_no`.
    Chấm là việc của lúc nộp bài, và chỉ ở đó mới đọc `correct_no` — giữ đúng
    một chỗ duy nhất quyết định đúng/sai.

    `elapsed_ms` nhận từ client và chỉ ĐƯỢC PHÉP TĂNG. Client tính sai hoặc gửi
    lại gói cũ thì cũng không làm đồng hồ chạy lùi, mà đó mới là hướng nguy
    hiểm: lùi được nghĩa là làm bài có giới hạn giờ vô thời hạn.
    """
    a = db.execute(
        "SELECT staff_id, status, elapsed_ms, total_questions FROM quiz_attempts WHERE id = ?",
        (attempt_id,),
    ).fetchone()
    if not a:
        raise HTTPException(404, "Không tìm thấy lượt làm bài")
    if a["staff_id"] != current["id"]:
        raise HTTPException(403, "Đây không phải bài làm của bạn")
    if a["status"] == "finished":
        raise HTTPException(409, "Bài này đã nộp rồi")

    if body.answers:
        hop_le = {
            r["id"] for r in db.execute(
                "SELECT id FROM quiz_attempt_items WHERE attempt_id = ?", (attempt_id,)
            ).fetchall()
        }
        db.executemany(
            "UPDATE quiz_attempt_items SET chosen_no = ?, time_ms = ? WHERE id = ?",
            [(x.chosen_no, x.time_ms, x.item_id) for x in body.answers if x.item_id in hop_le],
        )
    idx = min(max(body.current_idx, 0), max((a["total_questions"] or 1) - 1, 0))
    db.execute(
        """UPDATE quiz_attempts
              SET current_idx = ?, elapsed_ms = MAX(elapsed_ms, ?), saved_at = ?
            WHERE id = ?""",
        (idx, body.elapsed_ms, _vn_now(), attempt_id),
    )
    db.commit()
    return {"ok": True, "saved": len(body.answers)}


@router.delete("/attempts/{attempt_id}")
def abandon_attempt(
    attempt_id: int,
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Bỏ hẳn một bài đang làm dở. Bài ĐÃ NỘP thì không xoá được — đó là lịch sử."""
    a = db.execute(
        "SELECT staff_id, status FROM quiz_attempts WHERE id = ?", (attempt_id,)
    ).fetchone()
    if not a:
        raise HTTPException(404, "Không tìm thấy lượt làm bài")
    if a["staff_id"] != current["id"]:
        raise HTTPException(403, "Đây không phải bài làm của bạn")
    if a["status"] == "finished":
        raise HTTPException(409, "Bài đã nộp không xoá được")
    db.execute("DELETE FROM quiz_attempts WHERE id = ?", (attempt_id,))
    db.commit()
    return {"ok": True}


@router.post("/attempts/{attempt_id}/submit", response_model=AttemptResult)
def submit_attempt(
    attempt_id: int,
    body: AttemptSubmit,
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    a = db.execute(
        """SELECT a.*, s.name AS set_name FROM quiz_attempts a
             JOIN quiz_sets s ON s.id = a.set_id WHERE a.id = ?""",
        (attempt_id,),
    ).fetchone()
    if not a:
        raise HTTPException(404, "Không tìm thấy lượt làm bài")
    if a["staff_id"] != current["id"]:
        raise HTTPException(403, "Đây không phải bài làm của bạn")
    if a["status"] == "finished":
        raise HTTPException(409, "Bài này đã nộp rồi")

    # Ghi nốt những câu client còn giữ trong tay mà chưa kịp lưu. Ràng buộc
    # `attempt_id = ?` khiến item của lượt khác gửi nhầm vào đây không ăn thua.
    db.executemany(
        "UPDATE quiz_attempt_items SET chosen_no = ?, time_ms = ? WHERE id = ? AND attempt_id = ?",
        [(x.chosen_no, x.time_ms, x.item_id, attempt_id) for x in body.answers],
    )

    # Chấm TOÀN BỘ bài từ DB, không chấm theo những gì client vừa gửi: phần lớn
    # câu trả lời đã được lưu dần trong lúc làm (xem save_progress), nên nếu chỉ
    # xét `body.answers` thì bài làm dở rồi vào lại nộp sẽ mất sạch điểm của các
    # câu trả lời trước lúc tạm dừng.
    db.execute(
        """UPDATE quiz_attempt_items
              SET is_correct = CASE
                    WHEN chosen_no IS NOT NULL AND chosen_no =
                         (SELECT q.correct_no FROM quiz_questions q WHERE q.id = question_id)
                    THEN 1 ELSE 0 END
            WHERE attempt_id = ?""",
        (attempt_id,),
    )
    correct = db.execute(
        "SELECT IFNULL(SUM(is_correct), 0) c FROM quiz_attempt_items WHERE attempt_id = ?",
        (attempt_id,),
    ).fetchone()["c"]

    total = a["total_questions"] or 1
    score = round(correct * 100.0 / total, 1)
    # Đồng hồ chỉ được tăng — cùng lý do với save_progress.
    duration = max(body.duration_ms, a["elapsed_ms"] or 0)
    now = _vn_now()
    db.execute(
        """UPDATE quiz_attempts
              SET correct_count = ?, score = ?, duration_ms = ?, elapsed_ms = ?,
                  status = 'finished', finished_at = ?
            WHERE id = ?""",
        (correct, score, duration, duration, now, attempt_id),
    )
    write_audit(db, current["id"], "quiz.submit", "quiz_attempt", attempt_id,
                f"{a['set_name']} — {correct}/{total} ({score}%)")
    db.commit()
    return _build_result(db, attempt_id)


def _build_result(db, attempt_id: int) -> AttemptResult:
    a = db.execute(
        """SELECT a.*, s.name AS set_name FROM quiz_attempts a
             JOIN quiz_sets s ON s.id = a.set_id WHERE a.id = ?""",
        (attempt_id,),
    ).fetchone()
    rows = db.execute(
        """SELECT i.*, q.content, q.opt1, q.opt2, q.opt3, q.opt4, q.correct_no
             FROM quiz_attempt_items i
             JOIN quiz_questions q ON q.id = i.question_id
            WHERE i.attempt_id = ? ORDER BY i.order_no""",
        (attempt_id,),
    ).fetchall()
    review, skipped = [], 0
    for r in rows:
        if r["chosen_no"] is None:
            skipped += 1
        review.append(ReviewItem(
            order_no=r["order_no"], content=r["content"],
            options=_options_of(r, json.loads(r["option_order"])),
            chosen_no=r["chosen_no"], correct_no=r["correct_no"],
            is_correct=bool(r["is_correct"]), time_ms=r["time_ms"],
        ))
    correct = a["correct_count"] or 0
    total = a["total_questions"] or 0
    return AttemptResult(
        id=a["id"], set_id=a["set_id"], set_name=a["set_name"], mode=a["mode"],
        total_questions=total, correct_count=correct,
        wrong_count=total - correct - skipped, skipped_count=skipped,
        score=a["score"] or 0.0, duration_ms=a["duration_ms"] or 0,
        finished_at=str(a["finished_at"]) if a["finished_at"] else None,
        review=review,
    )


@router.get("/attempts/{attempt_id}/result", response_model=AttemptResult)
def get_result(
    attempt_id: int,
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    a = db.execute("SELECT staff_id, status FROM quiz_attempts WHERE id = ?",
                   (attempt_id,)).fetchone()
    if not a:
        raise HTTPException(404, "Không tìm thấy lượt làm bài")
    # Người khác chỉ xem được điểm trên bảng xếp hạng, không xem được đề + đáp án
    # của bài đã làm — nếu không thì bảng xếp hạng thành chỗ tra đáp án.
    if a["staff_id"] != current["id"] and current["role"] != StaffRole.ADMIN:
        raise HTTPException(403, "Đây không phải bài làm của bạn")
    if a["status"] != "finished":
        raise HTTPException(409, "Bài này chưa nộp")
    return _build_result(db, attempt_id)


# ─── Lịch sử & bảng xếp hạng ─────────────────────────────────────────────────
@router.get("/attempts", response_model=list[AttemptRow])
def my_attempts(
    set_id: Optional[int] = Query(None),
    limit: int = Query(30, ge=1, le=200),
    current: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    sql = """SELECT a.id, a.set_id, s.name AS set_name, a.mode, a.total_questions,
                    a.correct_count, a.score, a.duration_ms, a.finished_at
               FROM quiz_attempts a
               JOIN quiz_sets s ON s.id = a.set_id
              WHERE a.staff_id = ? AND a.status = 'finished'"""
    params: list = [current["id"]]
    if set_id:
        sql += " AND a.set_id = ?"
        params.append(set_id)
    sql += " ORDER BY a.finished_at DESC LIMIT ?"
    params.append(limit)
    return [
        {**dict(r), "finished_at": str(r["finished_at"]) if r["finished_at"] else None}
        for r in db.execute(sql, params).fetchall()
    ]


@router.get("/sets/{set_id}/leaderboard", response_model=list[AttemptRow])
def leaderboard(
    set_id: int,
    limit: int = Query(10, ge=1, le=100),
    _: dict = Depends(require_feature("menu.quiz")),
    db: sqlite3.Connection = Depends(get_db),
):
    """Mỗi người MỘT dòng — lượt tốt nhất. Cùng điểm thì ai nhanh hơn đứng trên.

    Chỉ tính bài thi thử (`mode='exam'`): chế độ ôn tập hiện đáp án ngay sau
    mỗi câu nên điểm 100% là chuyện đương nhiên, trộn chung thì bảng xếp hạng
    mất hết ý nghĩa.
    """
    rows = db.execute(
        """SELECT a.id, a.set_id, s.name AS set_name, u.full_name AS staff_name,
                  a.mode, a.total_questions, a.correct_count, a.score, a.duration_ms,
                  a.finished_at
             FROM quiz_attempts a
             JOIN quiz_sets s ON s.id = a.set_id
             LEFT JOIN user_tttt u ON u.id = a.staff_id
            WHERE a.set_id = ? AND a.status = 'finished' AND a.mode = 'exam'
              AND a.id = (SELECT b.id FROM quiz_attempts b
                           WHERE b.set_id = a.set_id AND b.staff_id = a.staff_id
                             AND b.status = 'finished' AND b.mode = 'exam'
                           ORDER BY b.score DESC, b.duration_ms ASC, b.id ASC LIMIT 1)
            ORDER BY a.score DESC, a.duration_ms ASC
            LIMIT ?""",
        (set_id, limit),
    ).fetchall()
    return [
        {**dict(r), "finished_at": str(r["finished_at"]) if r["finished_at"] else None}
        for r in rows
    ]
