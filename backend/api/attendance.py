"""Chấm công — Phòng Kế toán (ACCT). Raw SQL thuần, không ORM."""
import calendar
import io
import sqlite3
from datetime import date
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from backend.core.deps import get_current_staff, require_admin, require_feature
from backend.services.lich_lam_viec import LichLamViec, la_ngay_lam_viec, tai_lich
from backend.database import get_db, _vn_now
from backend.schemas.attendance import (
    AttendanceSymbolCreate, AttendanceSymbolUpdate, AttendanceSymbolOut,
    AttendanceDayUpdate, AdjustmentCreate, AdjustmentReview,
)

router = APIRouter()

_MANAGER_ROLES = ("truong_phong", "pho_phong")


# Quyền vào màn Chấm công — gán qua màn Phân quyền theo nhóm, không gate theo phòng
# (xem mục "Phân quyền" trong docs/DESIGN.md). Bảng công vẫn CHỈ chứa nhân viên phòng
# ACCT: đó là phạm vi DỮ LIỆU của module, không phải phạm vi quyền.
_QUYEN_CHAM_CONG = require_feature("menu.attendance")


def _is_manager(current: dict) -> bool:
    return current["role"] == "admin" or current["role"] in _MANAGER_ROLES


def _has_group_feature(db: sqlite3.Connection, staff_id: int, code: str) -> bool:
    """Check quyền theo nhóm (group_features) — dùng để CỘNG THÊM cho GDV được
    admin cấp quyền riêng, không thay thế check role trưởng/phó phòng/admin."""
    row = db.execute(
        """SELECT 1 FROM group_features gf
           JOIN group_members gm ON gm.group_id = gf.group_id
           JOIN user_groups g ON g.id = gm.group_id AND g.is_active = 1
           WHERE gm.staff_id = ? AND gf.feature_code = ? LIMIT 1""",
        (staff_id, code),
    ).fetchone()
    return bool(row)


def _can_view_dept(current: dict, db: sqlite3.Connection) -> bool:
    """Trưởng/phó phòng/admin luôn xem được cả phòng; GDV thường chỉ xem được
    nếu admin đã gán quyền attendance.view_dept qua màn Phân quyền theo nhóm."""
    return _is_manager(current) or _has_group_feature(db, current["id"], "attendance.view_dept")


def _can_export(current: dict, db: sqlite3.Connection) -> bool:
    """Tương tự _can_view_dept nhưng cho quyền xuất Excel (attendance.export)."""
    return _is_manager(current) or _has_group_feature(db, current["id"], "attendance.export")


def _load_lich(db: sqlite3.Connection, start: date, end: date) -> LichLamViec:
    """Lịch làm việc trong khoảng — ngày lễ + ngày làm bù.

    Thứ Bảy đi làm bù là ngày làm việc: có mặc định tính công, có xin điều chỉnh
    được, và không tô màu cuối tuần. Xem `backend/services/lich_lam_viec.py`."""
    return tai_lich(db, start, end)


def _default_work_value(db: sqlite3.Connection) -> float:
    """Số công của ký hiệu 'x' (ngày làm việc bình thường, chưa có dữ liệu ghi
    nhận) — tra bảng thay vì hardcode 1.0, vì put_day/create_adjustment đều cho
    admin đổi work_value của bất kỳ ký hiệu nào qua PUT /symbols/x (rà soát tiếp
    theo: 3 chỗ dùng giá trị mặc định trước đây hardcode, không đồng bộ nếu admin
    đổi số công của 'x'). Fallback 1.0 nếu ký hiệu 'x' vì lý do nào đó không còn
    active/tồn tại — không để vỡ trang vì thiếu dữ liệu cấu hình."""
    row = db.execute(
        "SELECT work_value FROM attendance_symbols WHERE symbol='x' AND is_active=1"
    ).fetchone()
    return row["work_value"] if row else 1.0


# Thứ hạng chức vụ để sắp bảng công — khớp đúng _ROLE_ORDER_SQL ở backend/api/staff.py:24-36
# (viết lại dạng dict Python vì cần sort kèm "tên" tách ra, SQL thuần không tách gọn được).
_ROLE_RANK = {
    "giam_doc": 0, "pho_giam_doc": 1, "admin": 2, "admin_l2": 3,
    "truong_phong": 4, "pho_phong": 5, "hau_kiem_vien": 6, "chuyen_vien": 7,
}


def _given_name(full_name: str) -> str:
    """"Tên" (từ cuối cùng) để sắp theo kiểu roster VN — vd "Nguyễn Văn A" sắp theo "A".
    Giới hạn đã biết: không xử lý collation dấu tiếng Việt (vd "Ánh" vs "Bình" có thể lệch
    thứ tự) — cùng hạn chế đã có sẵn ở backend/api/staff.py, không mở rộng sửa ở đây."""
    parts = (full_name or "").split()
    return parts[-1] if parts else (full_name or "")


def _require_staff_in_acct(staff_id: int, db: sqlite3.Connection) -> None:
    """Chặn thao tác lên nhân viên KHÔNG thuộc Phòng Kế toán — vd trưởng phòng KT lỡ chấm
    công/duyệt điều chỉnh cho staff_id phòng khác (phát hiện ở review PR #22)."""
    row = db.execute(
        """SELECT 1 FROM user_tttt u JOIN departments d ON d.id = u.department_id
           WHERE u.id = ? AND d.code = 'ACCT' AND u.is_active = 1""",
        (staff_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "Nhân viên không thuộc Phòng Kế toán")


def _acct_staff_rows(db: sqlite3.Connection):
    rows = db.execute(
        """SELECT u.id, u.full_name, u.employee_code, u.role FROM user_tttt u
           JOIN departments d ON d.id = u.department_id
           WHERE d.code = 'ACCT' AND u.is_active = 1 AND (u.is_deleted = 0 OR u.is_deleted IS NULL)"""
    ).fetchall()
    return sorted(
        rows,
        key=lambda r: (_ROLE_RANK.get(r["role"], 9), _given_name(r["full_name"]), r["employee_code"] or ""),
    )


def _adjustment_to_out(db: sqlite3.Connection, adj_id: int) -> dict:
    row = db.execute(
        """SELECT a.*, u.full_name AS staff_name, r.full_name AS reviewer_name,
                  att.staff_id AS att_staff_id, att.date AS att_date
           FROM attendance_adjustments a
           JOIN attendances att ON att.id = a.attendance_id
           JOIN user_tttt u ON u.id = att.staff_id
           LEFT JOIN user_tttt r ON r.id = a.reviewer_id
           WHERE a.id = ?""",
        (adj_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy yêu cầu điều chỉnh")
    d = dict(row)
    d["staff_id"] = d.pop("att_staff_id")
    d["date"] = d.pop("att_date")
    return d


# ─── Bảng công theo tháng / theo ngày ─────────────────────────────────────────
@router.get("/month")
def get_month(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    scope: str = Query("mine"),
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    # BUG đã sửa (phát hiện khi test thật): trước đây ép scope="mine" cho MỌI
    # chuyên viên vô điều kiện — kể cả GDV đã được admin cấp attendance.view_dept,
    # khiến quyền vừa gán hoàn toàn vô tác dụng ở endpoint này. Giờ chỉ ép về
    # "mine" khi thực sự KHÔNG có quyền xem cả phòng (role manager/admin HOẶC có
    # feature — _can_view_dept đã gộp cả 2 điều kiện).
    if current["role"] == "chuyen_vien" and not _can_view_dept(current, db):
        scope = "mine"
    elif scope not in ("mine", "dept"):
        scope = "mine"
    if scope == "dept" and not _can_view_dept(current, db):
        raise HTTPException(403, "Không có quyền xem bảng công cả phòng")

    days_in_month = calendar.monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, days_in_month)
    lich = _load_lich(db, start, end)

    if scope == "dept":
        staff_rows = _acct_staff_rows(db)
    else:
        staff_rows = db.execute(
            "SELECT id, full_name, employee_code FROM user_tttt WHERE id = ?", (current["id"],)
        ).fetchall()

    staff_ids = [r["id"] for r in staff_rows]
    att_map = {}
    if staff_ids:
        placeholders = ",".join("?" for _ in staff_ids)
        att_rows = db.execute(
            f"""SELECT id, staff_id, date, symbol, work_value, status, note
                FROM attendances WHERE staff_id IN ({placeholders}) AND date >= ? AND date <= ?""",
            (*staff_ids, start.isoformat(), end.isoformat()),
        ).fetchall()
        for r in att_rows:
            att_map[(r["staff_id"], r["date"])] = r

    # Ngày lễ và ngày chưa tới không mặc định tính công (quyết định nghiệp vụ 2026-08-10,
    # review PR #22 — trước đây chỉ check thứ trong tuần, bỏ sót cả 2 trường hợp này).
    today = _vn_now().date()
    default_wv = _default_work_value(db)

    result_staff = []
    for srow in staff_rows:
        sid = srow["id"]
        days = {}
        total = 0.0
        for day_num in range(1, days_in_month + 1):
            d = date(year, month, day_num)
            arow = att_map.get((sid, d.isoformat()))
            if arow:
                days[d.isoformat()] = {
                    "attendance_id": arow["id"], "staff_id": sid, "date": d.isoformat(),
                    "symbol": arow["symbol"], "work_value": arow["work_value"],
                    "status": arow["status"], "note": arow["note"], "is_virtual": False,
                }
                total += arow["work_value"] or 0
            elif la_ngay_lam_viec(d, lich) and d <= today:
                days[d.isoformat()] = {
                    "attendance_id": None, "staff_id": sid, "date": d.isoformat(),
                    "symbol": "x", "work_value": default_wv,
                    "status": None, "note": None, "is_virtual": True,
                }
                total += default_wv
            else:
                days[d.isoformat()] = {
                    "attendance_id": None, "staff_id": sid, "date": d.isoformat(),
                    "symbol": None, "work_value": 0.0,
                    "status": None, "note": None, "is_virtual": True,
                }
        result_staff.append({
            "staff_id": sid, "full_name": srow["full_name"], "employee_code": srow["employee_code"],
            "days": days, "total_work_value": total,
        })

    return {
        "year": year, "month": month, "days_in_month": days_in_month,
        "holidays": [d.isoformat() for d in sorted(lich.ngay_le)],
        # Ngày làm bù phải xuống tới frontend, không thì lưới tô T7 màu cuối tuần
        # trong khi cột Tổng vẫn cộng công của hôm đó — nhìn là thấy vênh.
        "makeup_days": [d.isoformat() for d in sorted(lich.ngay_bu)],
        "staff": result_staff,
    }


@router.get("/day")
def get_day(
    staff_id: int = Query(...),
    date_: str = Query(..., alias="date"),
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    # Đồng nhất với /month: chuyên viên được cấp attendance.view_dept cũng xem
    # được công của người khác qua endpoint tra 1 ngày này, không chỉ của mình.
    # Rà soát tiếp theo: trước đây chỉ chặn đúng role "chuyen_vien" — vai trò khác
    # không phải quản lý (vd hau_kiem_vien) nếu được gán vào ACCT sẽ lọt qua, xem
    # được công bất kỳ ai. Đổi sang _is_manager() để chặn đúng MỌI vai trò không
    # phải quản lý, khớp đúng cách get_month đang làm.
    if not _is_manager(current) and staff_id != current["id"] and not _can_view_dept(current, db):
        raise HTTPException(403, "Chỉ được xem công của chính mình")
    row = db.execute(
        "SELECT * FROM attendances WHERE staff_id=? AND date=?", (staff_id, date_)
    ).fetchone()
    if row:
        return dict(row)
    try:
        d = date.fromisoformat(date_)
    except ValueError:
        raise HTTPException(400, f"Ngày '{date_}' không hợp lệ, định dạng phải là YYYY-MM-DD")
    # Ngày lễ và ngày chưa tới không mặc định tính công — cùng fix với /month
    # (review PR #22, trước đây chỗ này chỉ check thứ trong tuần).
    if la_ngay_lam_viec(d, _load_lich(db, d, d)) and d <= _vn_now().date():
        return {"staff_id": staff_id, "date": date_, "symbol": "x", "work_value": _default_work_value(db),
                "status": None, "note": None, "is_virtual": True}
    return {"staff_id": staff_id, "date": date_, "symbol": None, "work_value": 0.0,
            "status": None, "note": None, "is_virtual": True}


@router.put("/day")
def put_day(
    body: AttendanceDayUpdate,
    staff_id: int = Query(...),
    date_: str = Query(..., alias="date"),
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    if not _is_manager(current):
        raise HTTPException(403, "Chỉ Trưởng/Phó phòng hoặc Admin được sửa công trực tiếp")
    # Đồng bộ với create_adjustment (rà soát vòng 3 PR #22, theo quyết định của
    # người dùng 2026-08-11): không ai — kể cả Trưởng/Phó phòng — được ghi công
    # cho ngày chưa tới.
    try:
        d = date.fromisoformat(date_)
    except ValueError:
        raise HTTPException(400, f"Ngày '{date_}' không hợp lệ, định dạng phải là YYYY-MM-DD")
    if d > _vn_now().date():
        raise HTTPException(400, "Không thể ghi công cho ngày chưa tới")
    sym = db.execute(
        "SELECT work_value FROM attendance_symbols WHERE symbol=? AND is_active=1", (body.symbol,)
    ).fetchone()
    if not sym:
        raise HTTPException(400, f"Ký hiệu '{body.symbol}' không hợp lệ")
    _require_staff_in_acct(staff_id, db)
    now = str(_vn_now())
    # Rà soát review PR #22 (Người 1, 17/08): TP/PP ghi đè công trực tiếp lên một
    # ngày do trigger đồng bộ nghỉ phép tự tạo (status='auto', source_leave_id trỏ
    # về đơn nghỉ) — dòng công không còn "thuộc về" đơn nghỉ gốc nữa sau khi ghi đè
    # tay. Không reset thì huỷ/xoá đơn nghỉ đó sau này sẽ dính FOREIGN KEY constraint
    # (source_leave_id vẫn trỏ về đơn đã xoá).
    db.execute(
        """INSERT INTO attendances
               (staff_id, date, symbol, work_value, status, note, confirmed_by, confirmed_at, created_at, updated_at)
           VALUES (?, ?, ?, ?, 'confirmed', ?, ?, ?, ?, ?)
           ON CONFLICT(staff_id, date) DO UPDATE SET
               symbol=excluded.symbol, work_value=excluded.work_value, status='confirmed',
               note=excluded.note, confirmed_by=excluded.confirmed_by, confirmed_at=excluded.confirmed_at,
               updated_at=excluded.updated_at, source_leave_id=NULL""",
        (staff_id, date_, body.symbol, sym["work_value"], body.note, current["id"], now, now, now),
    )
    db.commit()
    row = db.execute("SELECT * FROM attendances WHERE staff_id=? AND date=?", (staff_id, date_)).fetchone()
    return dict(row)


# ─── Danh sách trưởng/phó phòng ACCT (cho dropdown "Người kiểm soát" khi xuất) ─
@router.get("/managers")
def list_managers(
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    rows = db.execute(
        """SELECT u.id, u.full_name FROM user_tttt u
           JOIN departments d ON d.id = u.department_id
           WHERE d.code = 'ACCT' AND u.role IN ('truong_phong','pho_phong')
             AND u.is_active = 1 AND (u.is_deleted = 0 OR u.is_deleted IS NULL)
           ORDER BY u.full_name"""
    ).fetchall()
    return [dict(r) for r in rows]


# ─── Điều chỉnh công ──────────────────────────────────────────────────────────
@router.post("/adjustments")
def create_adjustment(
    body: AdjustmentCreate,
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    staff_id = body.staff_id or current["id"]
    # Rà soát tiếp theo: chỉ chặn đúng role "chuyen_vien" trước đây bỏ sót vai trò
    # khác không phải quản lý (vd hau_kiem_vien) — họ xin điều chỉnh hộ được tuỳ ý
    # ai trong ACCT. Chỉ "Quản lý" (xem comment dưới) mới được làm hộ người khác.
    if not _is_manager(current) and staff_id != current["id"]:
        raise HTTPException(403, "Chỉ được xin điều chỉnh công của chính mình")
    # Quản lý xin điều chỉnh hộ cho staff_id tuỳ ý — chặn nếu người đó không thuộc ACCT
    # (vd trưởng phòng KT lỡ chọn nhầm nhân viên phòng khác, phát hiện ở review PR #22).
    if staff_id != current["id"]:
        _require_staff_in_acct(staff_id, db)
    # Ngày chưa tới không có công để điều chỉnh — nhất quán với quy tắc "ngày chưa
    # tới không tính công mặc định" ở /month, /day, /export (rà soát vòng 2 PR #22).
    if body.date > _vn_now().date():
        raise HTTPException(400, "Không thể xin điều chỉnh công cho ngày chưa tới")
    # Rà soát vòng 3 PR #22: reason trước đây chỉ được chặn rỗng ở frontend — gọi
    # thẳng API sẽ tạo được yêu cầu không có lý do, ảnh hưởng dấu vết kiểm toán.
    # review_adjustment (nhánh reject) đã có check tương tự cho reject_reason.
    if not body.reason or not body.reason.strip():
        raise HTTPException(400, "Cần nhập lý do xin điều chỉnh")
    sym = db.execute(
        "SELECT work_value FROM attendance_symbols WHERE symbol=? AND is_active=1", (body.new_symbol,)
    ).fetchone()
    if not sym:
        raise HTTPException(400, f"Ký hiệu '{body.new_symbol}' không hợp lệ")

    date_str = body.date.isoformat()
    now = str(_vn_now())
    row = db.execute("SELECT * FROM attendances WHERE staff_id=? AND date=?", (staff_id, date_str)).fetchone()
    if row:
        attendance_id = row["id"]
        old_symbol, old_work_value = row["symbol"], row["work_value"]
    else:
        _lich = _load_lich(db, body.date, body.date)
        if not la_ngay_lam_viec(body.date, _lich):
            _loai = "ngày nghỉ lễ" if body.date in _lich.ngay_le else "ngày nghỉ cuối tuần"
            raise HTTPException(400, f"Không thể xin điều chỉnh cho {_loai} chưa có dữ liệu")
        init_symbol, init_value = "x", 1.0
        cur = db.execute(
            """INSERT INTO attendances (staff_id, date, symbol, work_value, status, created_at, updated_at)
               VALUES (?, ?, ?, ?, 'auto', ?, ?)""",
            (staff_id, date_str, init_symbol, init_value, now, now),
        )
        attendance_id = cur.lastrowid
        old_symbol, old_work_value = init_symbol, init_value

    # Chặn lặp yêu cầu: bấm nhanh 2 lần hoặc client tự retry khi mạng chậm đều có
    # thể tạo 2 dòng attendance_adjustments giống hệt nhau cho cùng 1 ngày (đã gặp
    # thực tế — 2 dòng cùng attendance_id, cách nhau 2 giây). Chặn ở đây thay vì chỉ
    # debounce phía frontend vì debounce không chặn được trường hợp do mạng.
    existing_pending = db.execute(
        "SELECT id FROM attendance_adjustments WHERE attendance_id=? AND status='pending'",
        (attendance_id,),
    ).fetchone()
    if existing_pending:
        raise HTTPException(409, "Đã có yêu cầu điều chỉnh đang chờ duyệt cho ngày này — vui lòng đợi xử lý xong trước khi gửi lại")

    db.execute(
        """INSERT INTO attendance_adjustments
               (attendance_id, requested_by, old_symbol, new_symbol, old_work_value, new_work_value, reason, status, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?)""",
        (attendance_id, current["id"], old_symbol, body.new_symbol, old_work_value, sym["work_value"], body.reason, now),
    )
    db.commit()
    new_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return _adjustment_to_out(db, new_id)


@router.put("/adjustments/{adj_id}/review")
def review_adjustment(
    adj_id: int,
    body: AdjustmentReview,
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    if not _is_manager(current):
        raise HTTPException(403, "Chỉ Trưởng/Phó phòng hoặc Admin được duyệt điều chỉnh công")
    row = db.execute("SELECT * FROM attendance_adjustments WHERE id=?", (adj_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy yêu cầu điều chỉnh")
    if row["status"] != "pending":
        raise HTTPException(400, "Yêu cầu này đã được xử lý")

    now = str(_vn_now())
    if body.action == "approve":
        # Cùng lý do như put_day: duyệt điều chỉnh ghi đè dòng công, dòng không còn
        # là "auto" từ đồng bộ nghỉ phép nữa nên phải bỏ liên kết source_leave_id —
        # tránh FOREIGN KEY constraint khi sau này huỷ/xoá đơn nghỉ gốc.
        db.execute(
            "UPDATE attendances SET symbol=?, work_value=?, status='adjusted', updated_at=?, source_leave_id=NULL WHERE id=?",
            (row["new_symbol"], row["new_work_value"], now, row["attendance_id"]),
        )
        db.execute(
            "UPDATE attendance_adjustments SET status='approved', reviewer_id=?, reviewed_at=? WHERE id=?",
            (current["id"], now, adj_id),
        )
    else:
        if not body.reject_reason or not body.reject_reason.strip():
            raise HTTPException(400, "Cần nhập lý do từ chối")
        db.execute(
            "UPDATE attendance_adjustments SET status='rejected', reviewer_id=?, reviewed_at=?, reject_reason=? WHERE id=?",
            (current["id"], now, body.reject_reason, adj_id),
        )
    db.commit()
    return _adjustment_to_out(db, adj_id)


@router.get("/adjustments")
def list_adjustments(
    scope: str = Query("mine"),
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    clauses, params = [], []
    # Rà soát tiếp theo: chỉ chặn đúng role "chuyen_vien" trước đây bỏ sót vai trò
    # khác không phải quản lý (vd hau_kiem_vien) — họ xem được scope="pending" (toàn
    # bộ yêu cầu đang chờ duyệt của cả phòng) mà không cần quyền gì. review_adjustment
    # (duyệt/từ chối) cũng chỉ dành cho _is_manager() — dùng lại đúng điều kiện đó.
    if not _is_manager(current):
        clauses.append("a.requested_by = ?")
        params.append(current["id"])
    elif scope == "pending":
        clauses.append("a.status = 'pending'")
    else:
        # "mine" hoặc giá trị scope lạ/không nhận diện được — mặc định an toàn về
        # đúng yêu cầu của người gọi, tránh trả toàn bộ hệ thống khi thiếu nhánh này
        # (phát hiện ở review PR #22: scope lạ trước đây khiến clauses rỗng → không
        # có WHERE → trả hết mọi yêu cầu điều chỉnh của cả phòng).
        clauses.append("a.requested_by = ?")
        params.append(current["id"])
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db.execute(
        f"""SELECT a.id FROM attendance_adjustments a
            JOIN attendances att ON att.id = a.attendance_id
            {where}
            ORDER BY a.created_at DESC""",
        params,
    ).fetchall()
    return [_adjustment_to_out(db, r["id"]) for r in rows]


# ─── Xuất Excel bảng công tháng ───────────────────────────────────────────────
@router.get("/export")
def export_month(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    kiem_soat_id: int = Query(..., description="staff_id trưởng/phó phòng ACCT ký tên Kiểm soát"),
    current: dict = Depends(_QUYEN_CHAM_CONG),
    db: sqlite3.Connection = Depends(get_db),
):
    # Cộng thêm quyền theo nhóm (attendance.export) — GDV được admin cấp quyền
    # riêng cũng xuất được, không chỉ trưởng/phó phòng/admin.
    if not _can_export(current, db):
        raise HTTPException(403, "Không có quyền xuất bảng công")

    # "Kiểm soát" bắt buộc là trưởng/phó phòng đang active của đúng phòng ACCT —
    # validate lại ở backend (không tin dữ liệu từ dropdown frontend).
    kiem_soat = db.execute(
        """SELECT full_name FROM user_tttt
           WHERE id = ? AND role IN ('truong_phong','pho_phong') AND is_active = 1
             AND department_id = (SELECT id FROM departments WHERE code='ACCT')""",
        (kiem_soat_id,),
    ).fetchone()
    if not kiem_soat:
        raise HTTPException(400, "Người kiểm soát không hợp lệ (phải là trưởng/phó phòng Kế toán đang hoạt động)")

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    days_in_month = calendar.monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, days_in_month)
    lich = _load_lich(db, start, end)

    staff_rows = _acct_staff_rows(db)
    staff_ids = [r["id"] for r in staff_rows]
    att_map = {}
    if staff_ids:
        placeholders = ",".join("?" for _ in staff_ids)
        att_rows = db.execute(
            f"""SELECT staff_id, date, symbol, work_value FROM attendances
                WHERE staff_id IN ({placeholders}) AND date >= ? AND date <= ?""",
            (*staff_ids, start.isoformat(), end.isoformat()),
        ).fetchall()
        for r in att_rows:
            att_map[(r["staff_id"], r["date"])] = r

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bang cong {month:02d}-{year}"

    # Font/màu khớp đúng mẫu giấy thật đang dùng ở Phòng Kế toán
    # (file 5_BANG_CHAM_CONG_PHONG_KE_TOAN_2026.xlsx) — Times New Roman, header
    # ngày thường nền trắng (không tô đỏ), T7/CN tô vàng trơn (không hoạ tiết).
    FONT_NAME = "Times New Roman"
    hdr_fill = PatternFill("solid", fgColor="FFFFFFFF")
    hdr_font = Font(name=FONT_NAME, size=11, bold=True)
    plain_hdr_font = Font(name=FONT_NAME, size=12, bold=False)
    weekend_fill = PatternFill("solid", fgColor="FFFFFF00")
    holiday_fill = PatternFill("solid", fgColor="C8E6C9")
    thin = Side(style="thin", color="000000")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_cols = 2 + days_in_month + 1  # STT + Họ tên + ngày + Số công
    last_col_letter = get_column_letter(total_cols)
    left_end = total_cols // 2       # chia đôi bề ngang cho khối tiêu đề trái/phải
    right_start = left_end + 1

    # ── Khối tiêu đề/quốc hiệu (để in trực tiếp trên A4) — khớp mẫu ảnh thật của
    # Phòng Kế toán: dòng 1 để trống, dòng 2-3 tên ngân hàng (trái) / quốc hiệu
    # (phải, căn giữa), dòng 5 tên phòng, dòng 7-8 tiêu đề "BẢNG CHẤM CÔNG" +
    # tháng/năm (căn giữa, merge hết bề ngang). Bảng dữ liệu bắt đầu từ dòng 10.
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=left_end)
    ws.cell(row=2, column=1, value="NGÂN HÀNG NÔNG NGHIỆP").font = Font(name=FONT_NAME, size=12, bold=True)
    ws.merge_cells(start_row=2, start_column=right_start, end_row=2, end_column=total_cols)
    c = ws.cell(row=2, column=right_start, value="CỘNG HOÀ XÃ HỘI CHỦ NGHĨA VIỆT NAM")
    c.font = Font(name=FONT_NAME, size=12, bold=True)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=left_end)
    ws.cell(row=3, column=1, value="VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM").font = Font(name=FONT_NAME, size=12, bold=True)
    ws.merge_cells(start_row=3, start_column=right_start, end_row=3, end_column=total_cols)
    c = ws.cell(row=3, column=right_start, value="Độc lập - Tự do - Hạnh phúc")
    c.font = Font(name=FONT_NAME, size=12, bold=True)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=left_end)
    c = ws.cell(row=5, column=1, value="Phòng Kế toán - TTTT")
    c.font = Font(name=FONT_NAME, size=12, bold=True)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=total_cols)
    c = ws.cell(row=7, column=1, value="BẢNG CHẤM CÔNG")
    c.font = Font(name=FONT_NAME, bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=total_cols)
    c = ws.cell(row=8, column=1, value=f"Tháng {month} năm {year}")
    c.font = Font(name=FONT_NAME, size=14)
    c.alignment = Alignment(horizontal="center")

    header_row = 10
    headers = ["STT", "Họ và tên"] + [str(d) for d in range(1, days_in_month + 1)] + ["Số công"]
    for col_idx, val in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=val)
    for col_idx, cell in enumerate(ws[header_row], start=1):
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border
        if 3 <= col_idx <= 2 + days_in_month:
            d = date(year, month, col_idx - 2)
            # Xét lễ TRƯỚC cuối tuần: ngày lễ rơi vào T7/CN phải tô màu lễ.
            # Ngày làm bù thì la_ngay_lam_viec() trả True nên tô như ngày thường.
            if d in lich.ngay_le:
                cell.fill, cell.font = holiday_fill, hdr_font
                continue
            if not la_ngay_lam_viec(d, lich):
                cell.fill, cell.font = weekend_fill, hdr_font
                continue
            cell.fill, cell.font = hdr_fill, hdr_font
            continue
        # STT / Họ và tên: không tô nền, không đậm — khớp mẫu thật
        cell.font = plain_hdr_font
    ws.cell(row=header_row, column=total_cols).fill = hdr_fill
    ws.cell(row=header_row, column=total_cols).font = hdr_font
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    for col_idx in range(3, 3 + days_in_month):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4
    ws.column_dimensions[get_column_letter(3 + days_in_month)].width = 10

    name_font = Font(name=FONT_NAME, size=12)
    symbol_font_normal = Font(name=FONT_NAME, size=11, bold=False)
    symbol_font_special = Font(name=FONT_NAME, size=11, bold=True)  # ký hiệu khác 'x' in đậm — khớp mẫu thật
    total_font = Font(name=FONT_NAME, size=12, bold=True)

    # Ngày lễ và ngày chưa tới không mặc định tính công — cùng fix với /month, /day
    # (review PR #22, trước đây chỗ này chỉ check thứ trong tuần).
    today = _vn_now().date()
    default_wv = _default_work_value(db)

    for idx, srow in enumerate(staff_rows, 1):
        sid = srow["id"]
        row_vals = [idx, srow["full_name"]]
        total = 0.0
        for day_num in range(1, days_in_month + 1):
            d = date(year, month, day_num)
            arow = att_map.get((sid, d.isoformat()))
            if arow:
                symbol, value = arow["symbol"], (arow["work_value"] or 0)
            elif la_ngay_lam_viec(d, lich) and d <= today:
                symbol, value = "x", default_wv
            else:
                symbol, value = "", 0.0
            row_vals.append(symbol)
            total += value
        row_vals.append(total)
        ws.append(row_vals)
        r = ws.max_row
        # Viền kẻ ô cho toàn bộ dòng (STT + Họ tên + từng ngày + Số công) —
        # để giống bảng giấy thật, in ra rõ ràng từng ô thay vì chỉ có màu nền.
        for col_idx in range(1, total_cols + 1):
            ws.cell(row=r, column=col_idx).border = cell_border
        ws.cell(row=r, column=1).font = name_font
        ws.cell(row=r, column=2).font = name_font
        ws.cell(row=r, column=total_cols).font = total_font
        for day_num in range(1, days_in_month + 1):
            d = date(year, month, day_num)
            cell = ws.cell(row=r, column=2 + day_num)
            cell.alignment = Alignment(horizontal="center")
            # Không tô màu riêng cho ô nghỉ phép/nghỉ khác — mẫu thật chỉ in đậm,
            # không có nền màu (chỉ T7/CN và ngày lễ mới có nền).
            if d in lich.ngay_le:
                cell.fill = holiday_fill
            elif not la_ngay_lam_viec(d, lich):
                cell.fill = weekend_fill
            cell.font = symbol_font_normal if cell.value in (None, "", "x") else symbol_font_special

    # Dòng chú thích ký hiệu — liệt kê động từ attendance_symbols đang active
    # (trừ 'x' vì là công bình thường, không cần giải thích), khớp đúng câu chữ
    # mẫu thật: "Ghi chú: P: Nghỉ phép. T: Nghỉ thai sản. ...".
    sym_rows = db.execute(
        "SELECT symbol, description FROM attendance_symbols WHERE is_active=1 AND symbol != 'x' ORDER BY rowid"
    ).fetchall()
    legend = "Ghi chú: " + ". ".join(f"{s['symbol']}: {s['description']}" for s in sym_rows)
    note_row = ws.max_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=total_cols)
    c = ws.cell(row=note_row, column=1, value=legend)
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.alignment = Alignment(horizontal="left")

    # 2 dòng ký tên, cách nhau nhiều dòng trống để ký tay khi in — khớp đúng
    # khoảng cách ở mẫu thật (nhãn dòng note_row+1, tên dòng note_row+8).
    # "Lập bảng" tự động = người đang thực hiện xuất (dù là GDV được cấp quyền
    # hay trưởng/phó phòng), "Kiểm soát" = tên đã chọn ở dropdown frontend, đã
    # validate lại phía trên.
    label_row = note_row + 1
    name_row = label_row + 7
    # "Lập bảng" chiếm cột B-D, "Kiểm soát" chiếm từ cột E tới hết bảng — khớp
    # đúng vùng merge ở mẫu thật (B23:D23 và E23:AH23), không phụ thuộc số ngày
    # trong tháng như công thức cũ (từng đẩy "Kiểm soát" dồn sát mép phải).
    sign_col_2 = 5
    sign_label_font = Font(name=FONT_NAME, size=14, bold=True)
    sign_name_font = Font(name=FONT_NAME, size=14, bold=True)
    ws.merge_cells(start_row=label_row, start_column=2, end_row=label_row, end_column=4)
    ws.cell(row=label_row, column=2, value="Lập bảng").font = sign_label_font
    ws.cell(row=label_row, column=2).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=name_row, start_column=2, end_row=name_row, end_column=4)
    c = ws.cell(row=name_row, column=2, value=current["full_name"])
    c.font = sign_name_font
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=label_row, start_column=sign_col_2, end_row=label_row, end_column=total_cols)
    c = ws.cell(row=label_row, column=sign_col_2, value="Kiểm soát")
    c.font = sign_label_font
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=name_row, start_column=sign_col_2, end_row=name_row, end_column=total_cols)
    c = ws.cell(row=name_row, column=sign_col_2, value=kiem_soat["full_name"])
    c.font = sign_name_font
    c.alignment = Alignment(horizontal="center")

    # ── Cấu hình khổ giấy A4 ngang để in trực tiếp — bảng có tới 34 cột (STT +
    # Họ tên + tối đa 31 ngày + Số công) nên bắt buộc landscape mới vừa. fitToWidth=1
    # ép co vừa 1 trang theo bề ngang khi in; fitToHeight=0 cho phép nhiều trang dọc
    # nếu phòng đông người — mỗi trang tự lặp lại phần tiêu đề + header cột.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws.print_area = f"A1:{last_col_letter}{ws.max_row}"
    ws.print_title_rows = f"1:{header_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''bang_cong_{year}_{month:02d}.xlsx"},
    )


# ─── Cấu hình ký hiệu công (admin) ────────────────────────────────────────────
@router.get("/symbols", response_model=List[AttendanceSymbolOut])
def list_symbols(
    _: dict = Depends(get_current_staff),
    db: sqlite3.Connection = Depends(get_db),
):
    rows = db.execute("SELECT * FROM attendance_symbols ORDER BY symbol").fetchall()
    return [dict(r) for r in rows]


@router.post("/symbols", response_model=AttendanceSymbolOut)
def create_symbol(
    body: AttendanceSymbolCreate,
    _: dict = Depends(require_admin),
    db: sqlite3.Connection = Depends(get_db),
):
    existing = db.execute("SELECT symbol FROM attendance_symbols WHERE symbol=?", (body.symbol,)).fetchone()
    if existing:
        raise HTTPException(400, f"Ký hiệu '{body.symbol}' đã tồn tại")
    db.execute(
        "INSERT INTO attendance_symbols (symbol, description, work_value, color, is_active) VALUES (?,?,?,?,?)",
        (body.symbol, body.description, body.work_value, body.color, body.is_active),
    )
    db.commit()
    row = db.execute("SELECT * FROM attendance_symbols WHERE symbol=?", (body.symbol,)).fetchone()
    return dict(row)


@router.put("/symbols/{symbol}", response_model=AttendanceSymbolOut)
def update_symbol(
    symbol: str,
    body: AttendanceSymbolUpdate,
    _: dict = Depends(require_admin),
    db: sqlite3.Connection = Depends(get_db),
):
    row = db.execute("SELECT * FROM attendance_symbols WHERE symbol=?", (symbol,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy ký hiệu")
    updates = body.model_dump(exclude_unset=True)
    if not updates:
        return dict(row)
    set_clause = ", ".join(f"{k}=?" for k in updates)
    db.execute(f"UPDATE attendance_symbols SET {set_clause} WHERE symbol=?", (*updates.values(), symbol))
    db.commit()
    row = db.execute("SELECT * FROM attendance_symbols WHERE symbol=?", (symbol,)).fetchone()
    return dict(row)
