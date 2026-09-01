# -*- coding: utf-8 -*-
"""
test_doi_soat_citad_exporters.py
---------------------------------
Khoá lại 2 lỗi thật phát hiện 23/08/2026 (Phòng Thanh toán, xem screenshot
báo cáo "Tất cả lệnh" thật) trong `backend/services/doi_soat_citad/exporters.py`:

1. Cột "Ngày GD" luôn TRỐNG cho mọi dòng có gốc CITAD (status 'both' /
   'only_citad' — CITAD không có cột ngày riêng từng dòng, chỉ 1 dòng
   header áp dụng cho cả file, xem parsers.py::parse_citad_xls) — phải rơi
   về đúng `ngay_cham` thay vì để trống.
2. Cột "Số tiền" ở dòng "khớp" trong export_doiSoat_full() căn PHẢI + không
   dấu phẩy (Excel dùng style mặc định vì ô ghi giá trị thô không có `s=`,
   xác nhận bằng mổ XML file xuất thật) — ngược hẳn dòng lệch (căn trái +
   dấu phẩy). Đã sửa: style riêng đúng ô Số tiền cho dòng khớp, các cột
   khác giữ nguyên cách nhanh (không style cả 13 cột — đổi tốc độ ~10 lần
   để đồng bộ toàn bộ cột không đáng, theo Phòng Thanh toán).

Không mock: ghi .xlsx thật rồi đọc lại bằng openpyxl, đúng cách các test
khác trong dự án đã làm (xem test_citad_export_layout.py).
"""
import io

from openpyxl import load_workbook

from backend.services.doi_soat_citad.exporters import export_doiSoat_full


def _khop_row(**kw):
    row = {
        'status': 'both', 'chieu': 'den', 'loai': 'il', 'loai_tien': 'VND',
        'so_gd': '10001411', 'key_agri': '10001411', 'so_tien': 1_000_000_000,
        'dich_vu': 'Chuyển có giá trị cao', 'nh_nhan': '01201001',
        'trang_thai': 'PYED', 'ngay': '',  # CITAD không có ngày riêng từng dòng
        'cong': '1',
    }
    row.update(kw)
    return row


def _lech_row(**kw):
    row = {
        'status': 'only_ipcas', 'chieu': 'den', 'loai': 'ih', 'loai_tien': 'VND',
        'so_gd': '', 'key_agri': '10003008', 'so_tien': 307_195_253_272,
        'dich_vu': 'CITAD CAO', 'nh_nhan': '01401001', 'trang_thai': 'SBSC',
        'ngay': '19/08/2026',
    }
    row.update(kw)
    return row


def _export_and_reload(lech_rows, khop_rows, ngay_cham='19/08/2026'):
    buf = io.BytesIO()
    export_doiSoat_full(lech_rows, khop_rows, ngay_cham, buf)
    buf.seek(0)
    wb = load_workbook(buf)
    return wb['Tất cả lệnh']


def test_ngay_gd_rong_o_citad_roi_ve_ngay_cham():
    ws = _export_and_reload([], [_khop_row(ngay='')])
    row = 6  # 0 dong lech -> dong khop dau tien o hang 6
    assert ws.cell(row, 10).value == '19/08/2026'


def test_ngay_gd_that_cua_ipcas_khong_bi_ghi_de():
    """Dòng only_ipcas/only_hub có ngày THẬT riêng (có thể khác ngay_cham —
    lệnh lập ngày khác ngày đi kênh) — không được ghi đè bằng ngay_cham."""
    ws = _export_and_reload([_lech_row(ngay='18/08/2026')], [], ngay_cham='19/08/2026')
    row = 6
    assert ws.cell(row, 10).value == '18/08/2026'


def test_so_tien_dong_khop_can_trai_va_co_dau_phay_giong_dong_lech():
    ws = _export_and_reload([_lech_row()], [_khop_row()])
    row_lech = 6
    row_khop = 7
    for row in (row_lech, row_khop):
        cell = ws.cell(row, 8)
        assert cell.alignment.horizontal == 'left'
        assert cell.number_format == '#,##0'


def test_stt_dong_khop_can_giua_giong_dong_lech():
    """Cùng lỗi gốc với Số tiền (ô có giá trị không kế thừa style cột) —
    phát hiện tiếp qua câu hỏi trực tiếp của người dùng: STT là số nguyên
    nên Excel General tự căn PHẢI thay vì GIỮA như ý định ban đầu."""
    ws = _export_and_reload([_lech_row()], [_khop_row()])
    for row in (6, 7):
        assert ws.cell(row, 1).alignment.horizontal == 'center'
