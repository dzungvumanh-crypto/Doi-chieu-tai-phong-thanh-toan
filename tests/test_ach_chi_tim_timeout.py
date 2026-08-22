"""Test tính năng "chạy thiếu GL02/MIS_đến, chỉ tìm Timeout không đi kênh"
(chi_tim_timeout — 2026-08-21, xem memory project_ach_gl02_optional_tiered_deps).

Business Owner: "Timeout không đi kênh" (khop_voi_gw) chỉ cần PDF+GW+MIS_đi×2,
không đụng GL02/MIS_đến — nhưng main_from_dir() trước đây chặn cứng cả 2. Đợt
này mở khóa CHỈ Tầng 0 khi người dùng tự xác nhận (chi_tim_timeout=True); Tầng 1
(NPO/MIS thừa đi-đến, huỷ, Điểm 2 OSB, Điểm 4 T-2) tắt hẳn, ghi rõ "CHƯA ĐỐI
CHIẾU ĐƯỢC" thay vì số 0 (tránh hiểu nhầm "đã khớp hết").

3 lớp test:
- TestValidateTang0Ok       — validate_required_files() thuần, không cần file thật.
- TestXuatExcelChuaDoiChieuDuoc — xuat_excel() với DataFrame tổng hợp, đúng pattern
  test_ach_excel_export.py (không mock, ghi file .xlsx thật rồi đọc lại).
- TestMainFromDirChiTimTimeout — end-to-end main_from_dir() với file PDF/GW/MIS_đi
  thật (zip AES thu nhỏ, đúng pattern test_ach_gl02_misden_algorithm.py), KHÔNG
  có GL02/MIS_đến — kiểm tra đúng hành vi raise/không-raise theo chi_tim_timeout.

Chạy: .venv\\Scripts\\python.exe -m pytest tests/test_ach_chi_tim_timeout.py -v
"""

import io

import openpyxl
import pandas as pd
import pyzipper
import pytest

from backend.services.ach import config as _cfg
from backend.services.ach.pipeline import xuat_excel, main_from_dir
from backend.services.ach.validate import validate_required_files

_FULL_SET = [
    'ACH_20260711_VBAAVNVN_NRT_16362_N03_1.pdf',
    'GL02_20260711_1000.zip',
    'di GW 11.07.xlsx',
    'doichieugd_20260710__01_DI_9999_N.zip',
    'doichieugd_20260711__01_DI_9999_N.zip',
    'doichieugd_20260710__01_DEN_9999_N.zip',
    'doichieugd_20260711__01_DEN_9999_N.zip',
]


# ─── validate_required_files() — cờ tang0_ok ─────────────────────────────────

class TestValidateTang0Ok:
    def test_du_het_ca_ok_va_tang0_ok_deu_true(self):
        res = validate_required_files(_FULL_SET)
        assert res['ok'] is True
        assert res['tang0_ok'] is True

    def test_thieu_gl02_ok_false_nhung_tang0_ok_true(self):
        res = validate_required_files([f for f in _FULL_SET if not f.startswith('GL02')])
        assert res['ok'] is False
        assert res['tang0_ok'] is True

    def test_thieu_mis_den_ok_false_nhung_tang0_ok_true(self):
        res = validate_required_files([f for f in _FULL_SET if '_DEN_' not in f])
        assert res['ok'] is False
        assert res['tang0_ok'] is True

    def test_thieu_ca_gl02_lan_mis_den_van_tang0_ok_true(self):
        res = validate_required_files([
            f for f in _FULL_SET if not f.startswith('GL02') and '_DEN_' not in f
        ])
        assert res['ok'] is False
        assert res['tang0_ok'] is True

    def test_thieu_mis_di_thi_tang0_ok_cung_false(self):
        """Thiếu MIS_đi (Tầng 0 thật sự) — KHÔNG được đề nghị chế độ chạy thiếu."""
        res = validate_required_files([f for f in _FULL_SET if '_DI_' not in f])
        assert res['tang0_ok'] is False

    def test_thieu_gw_thi_tang0_ok_cung_false(self):
        res = validate_required_files([f for f in _FULL_SET if 'GW' not in f])
        assert res['tang0_ok'] is False

    def test_thieu_pdf_thi_tang0_ok_cung_false(self):
        res = validate_required_files([f for f in _FULL_SET if not f.endswith('.pdf')])
        assert res['tang0_ok'] is False


# ─── xuat_excel() — nhãn "CHƯA ĐỐI CHIẾU ĐƯỢC" khi ly_do_thieu_tang1 ─────────

def _synthetic_dfs_tang0_only():
    """Chỉ có dữ liệu Tầng 0 (GW + Timeout) — mọi df Tầng 1 = None (mô phỏng
    main_from_dir() khi chi_tim_timeout=True bỏ qua Phase 2 phụ thuộc GL02)."""
    df_timeout = pd.DataFrame({'SO_TIEN': [300_000]})
    df_gw_raw = pd.DataFrame({
        'BRCD': ['0001'], 'STTLMAMT': [1_000_000], 'MSGREF': ['REF1'],
        'SessionId': ['16282'], 'PrcFlg': ['Lệnh Hoàn thành'], 'KEY_GW': ['00011000000'],
    })
    return df_timeout, df_gw_raw


class TestXuatExcelChuaDoiChieuDuoc:
    def test_sheet_tang1_hien_nhan_chua_doi_chieu_duoc(self, tmp_path):
        df_timeout, df_gw_raw = _synthetic_dfs_tang0_only()
        output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')

        xuat_excel(
            output_path, '16282',
            None, None, None,   # df_mis_di_khop, df_npo_di_thua, df_mis_di_thua
            df_timeout,
            None, None, None,   # df_mis_den_khop, df_npo_den_thua, df_mis_den_thua
            df_gw_raw,
            ly_do_thieu_tang1='GL02, MIS_đến',
        )

        wb = openpyxl.load_workbook(output_path)
        assert wb['MIS_DI_KHOP']['A1'].value == 'CHƯA ĐỐI CHIẾU ĐƯỢC — thiếu file GL02, MIS_đến'
        assert wb['NPO_DI_THUA']['A1'].value == 'CHƯA ĐỐI CHIẾU ĐƯỢC — thiếu file GL02, MIS_đến'
        assert wb['MIS_DEN_THUA']['A1'].value == 'CHƯA ĐỐI CHIẾU ĐƯỢC — thiếu file GL02, MIS_đến'

    def test_sheet_tang0_khong_bi_anh_huong(self, tmp_path):
        """TIMEOUT_KHONG_KENH (Tầng 0) vẫn hiện dữ liệu thật, không bị gắn nhãn
        thiếu — đây chính là mục tiêu nghiệp vụ của tính năng."""
        df_timeout, df_gw_raw = _synthetic_dfs_tang0_only()
        output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')

        xuat_excel(
            output_path, '16282',
            None, None, None, df_timeout, None, None, None, df_gw_raw,
            ly_do_thieu_tang1='GL02, MIS_đến',
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb['TIMEOUT_KHONG_KENH']
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert 'SO_TIEN' in header
        assert ws.cell(row=2, column=header.index('SO_TIEN') + 1).value == 300_000

    def test_khong_thieu_gi_thi_khong_hien_nhan_canh_bao(self, tmp_path):
        """Regression — ly_do_thieu_tang1=None (mặc định, hành vi cũ) không được
        đổi thông điệp sheet rỗng thật."""
        output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')
        xuat_excel(
            output_path, '16282',
            None, None, None, None, None, None, None, None,
        )
        wb = openpyxl.load_workbook(output_path)
        assert wb['MIS_DI_KHOP']['A1'].value == '(Không có dữ liệu)'

    def test_tong_ket_hien_banner_canh_bao(self, tmp_path):
        df_timeout, df_gw_raw = _synthetic_dfs_tang0_only()
        output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')
        xuat_excel(
            output_path, '16282',
            None, None, None, df_timeout, None, None, None, df_gw_raw,
            ly_do_thieu_tang1='GL02, MIS_đến',
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb['TONG_KET']
        assert 'THIẾU FILE' in ws['A1'].value
        assert 'GL02' in ws['A1'].value

    def test_tong_ket_khong_hien_banner_khi_du_file(self, tmp_path):
        output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')
        xuat_excel(
            output_path, '16282',
            None, None, None, None, None, None, None, None,
        )
        wb = openpyxl.load_workbook(output_path)
        ws = wb['TONG_KET']
        assert ws['A1'].value == 'Ngày đối chiếu'


# ─── main_from_dir() end-to-end — file PDF/GW/MIS_đi thật, KHÔNG GL02/MIS_đến ─

_SID = '16362'
_MIS_DI_COLS = [
    'NGAY_GIAO_DICH', 'CHI_NHANH', 'REFHUB', 'MSGREF', 'MSGSEQ', 'TXID',
    'KENH_THANH_TOAN', 'TRANG_THAI_LENH', 'SO_TIEN', 'TRACE',
    'SE_TRACE', 'SESSION', 'LOAI_LENH_OSB', 'NH_NHAN',
    'MA_GIAO_DICH', 'NOI_DUNG', 'NGAY_KENH_TRA',
]


def _mis_di_row(refhub='REF1', chi_nhanh='1000', so_tien='500000', trang_thai='SCNL'):
    return {
        'NGAY_GIAO_DICH': '11/07/2026', 'CHI_NHANH': chi_nhanh, 'REFHUB': refhub,
        'MSGREF': 'MSG' + refhub, 'MSGSEQ': '1', 'TXID': 'TX' + refhub,
        'KENH_THANH_TOAN': 'ACH-NAPAS', 'TRANG_THAI_LENH': trang_thai,
        'SO_TIEN': so_tien, 'TRACE': "'0000123456", 'SE_TRACE': '',
        'SESSION': _SID, 'LOAI_LENH_OSB': '', 'NH_NHAN': 'NH TEST',
        'MA_GIAO_DICH': 'MGD' + refhub, 'NOI_DUNG': 'nd', 'NGAY_KENH_TRA': '11/07/2026 10:00:00',
    }


def _make_mis_di_zip(tmp_path, name: str, rows: list[dict]):
    df = pd.DataFrame(rows, columns=_MIS_DI_COLS) if rows else pd.DataFrame(columns=_MIS_DI_COLS)
    csv_bytes = df.to_csv(index=False).encode('utf-8-sig')
    path = tmp_path / name
    with pyzipper.AESZipFile(str(path), 'w', compression=pyzipper.ZIP_DEFLATED,
                              encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(_cfg.ZIP_PASSWORD)
        zf.writestr('data.csv', csv_bytes)
    return str(path)


def _make_gw_xlsx(tmp_path, name: str = 'di GW 11.07.xlsx'):
    """GW khớp đúng 1 dòng với MIS_đi (BRCD+STTLMAMT = CHI_NHANH+SO_TIEN =
    '1000'+'500000') — pipeline chạy hết Tầng 0 mà không phát sinh timeout thật
    ngoài dự kiến (đủ đơn giản để không phụ thuộc chi tiết phân loại timeout)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'đi GW 11.07'
    ws.append(['BRCD', 'STTLMAMT', 'MSGREF', 'SessionId', 'PrcFlg'])
    ws.append(['1000', '500000', 'MSGREF1', _SID, 'Lệnh Hoàn thành'])
    path = tmp_path / name
    wb.save(str(path))
    return str(path)


def _make_pdf(tmp_path, name='ACH_20260711_VBAAVNVN_NRT_16362_N03_1.pdf'):
    path = tmp_path / name
    path.write_bytes(b'%PDF-fake-content-not-parsed')
    return str(path)


class TestMainFromDirChiTimTimeout:
    def test_mac_dinh_thieu_gl02_van_raise(self, tmp_path):
        """Regression — không truyền chi_tim_timeout (mặc định False) phải giữ
        đúng hành vi cũ: raise ngay khi thiếu GL02, không chạy gì thêm."""
        _make_pdf(tmp_path)
        _make_gw_xlsx(tmp_path)
        _make_mis_di_zip(tmp_path, 'doichieugd_20260711__01_DI_9999_N.zip', [_mis_di_row()])
        _make_mis_di_zip(tmp_path, 'doichieugd_20260711__02_DI_9999_N.zip', [])

        with pytest.raises(FileNotFoundError, match='GL02'):
            main_from_dir(str(tmp_path), str(tmp_path / 'out'))

    def test_thieu_mis_di_van_raise_du_chi_tim_timeout_true(self, tmp_path):
        """chi_tim_timeout=True KHÔNG miễn trừ mức tối thiểu Tầng 0 thật sự — chỉ
        1 file MIS_đi (cần 2) vẫn phải raise."""
        _make_pdf(tmp_path)
        _make_gw_xlsx(tmp_path)
        _make_mis_di_zip(tmp_path, 'doichieugd_20260711__01_DI_9999_N.zip', [_mis_di_row()])

        with pytest.raises(FileNotFoundError, match='MIS_DI'):
            main_from_dir(str(tmp_path), str(tmp_path / 'out'), chi_tim_timeout=True)

    def test_thieu_gl02_va_mis_den_chi_tim_timeout_true_chay_thanh_cong(self, tmp_path):
        """Kịch bản chính của tính năng: đủ PDF+GW+MIS_đi×2, KHÔNG GL02/MIS_đến,
        chi_tim_timeout=True — chạy xong không raise, sheet Tầng 0 có dữ liệu
        thật, sheet Tầng 1 ghi rõ CHƯA ĐỐI CHIẾU ĐƯỢC."""
        _make_pdf(tmp_path)
        _make_gw_xlsx(tmp_path)
        _make_mis_di_zip(tmp_path, 'doichieugd_20260711__01_DI_9999_N.zip', [_mis_di_row()])
        _make_mis_di_zip(tmp_path, 'doichieugd_20260711__02_DI_9999_N.zip', [])
        out_dir = tmp_path / 'out'

        output_path = main_from_dir(str(tmp_path), str(out_dir), chi_tim_timeout=True)

        assert output_path is not None
        wb = openpyxl.load_workbook(output_path)
        assert 'TIMEOUT_KHONG_KENH' in wb.sheetnames
        assert wb['MIS_DI_KHOP']['A1'].value.startswith('CHƯA ĐỐI CHIẾU ĐƯỢC')
        assert wb['MIS_DEN_KHOP']['A1'].value.startswith('CHƯA ĐỐI CHIẾU ĐƯỢC')
        assert 'THIẾU FILE' in wb['TONG_KET']['A1'].value
