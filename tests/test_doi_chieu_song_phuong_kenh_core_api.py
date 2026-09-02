"""
Test API-level cho "Đối chiếu đến" — Kênh↔Hub + Hub↔Core chạy TỰ ĐỘNG nối tiếp trong 1 job
(quyết định 2026-08-28, xem docstring `doi_chieu_song_phuong_kenh_core_service.py`).

Thay hẳn `test_doi_chieu_song_phuong_kenh_api.py` + `test_doi_chieu_song_phuong_core_api.py`
(test qua 2 router riêng đã xoá). Không kiểm lại toán khớp (đã có 36+34 test ở
`test_doi_chieu_song_phuong_kenh_algorithm.py`/`_core_algorithm.py`) — chỉ kiểm lớp điều phối:
2 bước chạy nối tiếp, lỗi 1 bước không chặn bước kia, CSV có sẵn thì không giải mã lại GL02.

Chạy: .venv\\Scripts\\python.exe -m pytest tests/test_doi_chieu_song_phuong_kenh_core_api.py -v
"""

import csv
import io
import time
import zipfile

import openpyxl
import pandas as pd
import pytest
import pyzipper

import backend.api.doi_chieu_song_phuong_kenh_core as kenh_core_api
from backend.core.uploads import MAX_REQUEST_BYTES
from backend.services import doi_chieu_song_phuong_kenh_core_service as svc
from backend.services import doi_chieu_song_phuong_service as ipcas_svc
from backend.services.doi_chieu_song_phuong_core import pipeline as core_pipeline_mod
from backend.services.doi_chieu_song_phuong_kenh import pipeline as kenh_pipeline_mod

_MK = "matkhau-test-kenh-core-api"


def test_tran_upload_nho_hon_tran_than_request():
    """Nếu trần upload >= trần thân request (MAX_REQUEST_BYTES, mặc định 600MB) thì
    BodySizeLimitMiddleware chặn TRƯỚC — client chỉ thấy socket đứt kiểu "[WinError 10054]",
    không đọc được thông báo 413 rõ ràng của route (cùng bài học đã dính ở ACH, xem
    tests/test_ach_tran_upload.py)."""
    assert kenh_core_api._MAX_UPLOAD < MAX_REQUEST_BYTES


@pytest.fixture(autouse=True)
def _mat_khau_zip(monkeypatch):
    """Mật khẩu ZIP đọc từ .env (zip_password()) — không còn hằng số ipcas_svc.ZIP_PASSWORD."""
    monkeypatch.setenv("DOI_CHIEU_ZIP_PASSWORD", _MK)


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HUB_COLS = ["NGAY_GIAO_DICH", "CHI_NHANH", "REFHUB", "MSGREF", "MSGSEQ", "TXID",
             "KENH_THANH_TOAN", "TRANG_THAI_LENH", "SO_TIEN", "TRACE", "SESSION",
             "LOAI_LENH_OSB", "NH_GUI", "NOI_DUNG"]
_KENH_COLS = ["STT", "Ngày GD", "Giờ truyền nhận", "MtId/MsgId", "Số tiền"]
_GL02_COLS = ["TRBRCD", "CUSTOMER", "CRAMOUNT", "DRAMOUNT", "REFERENCE", "REMARK"]

# Prefix guard KENH_MTID_PREFIX["202"] — kenh SPRT phải khớp 10 ký tự đầu này.
_MSG_202RT = "0200970488TESTRT202AAAA"


def _hub_row(msgref, txid, ktt="SP REALTIME", trang_thai="PYED", so_tien="100000"):
    return {
        "NGAY_GIAO_DICH": "25/08/2026", "CHI_NHANH": "1000", "REFHUB": "REFHUB001",
        "MSGREF": f"'{msgref}", "MSGSEQ": f"'{msgref}", "TXID": f"'{txid}",
        "KENH_THANH_TOAN": ktt, "TRANG_THAI_LENH": trang_thai, "SO_TIEN": so_tien,
        "TRACE": "000353682", "SESSION": "20260825", "LOAI_LENH_OSB": "0",
        "NH_GUI": "01202001", "NOI_DUNG": "TEST",
    }


def _make_hub_zip(rows: list[dict]) -> bytes:
    df = pd.DataFrame(rows, columns=_HUB_COLS)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("data.csv", df.to_csv(index=False).encode("utf-8-sig"))
    return buf.getvalue()


def _kenh_row(mtid, so_tien="100000"):
    return {"STT": "1", "Ngày GD": "25/08/2026", "Giờ truyền nhận": "25/08/2026 00:00:01",
            "MtId/MsgId": mtid, "Số tiền": so_tien}


def _make_gl02_zip(rows: list[dict]) -> bytes:
    """GL02 thật mã hoá AES (đúng mật khẩu `DOI_CHIEU_ZIP_PASSWORD` — xem fixture `_mat_khau_zip`)."""
    df = pd.DataFrame(rows, columns=_GL02_COLS)
    buf = io.BytesIO()
    with pyzipper.AESZipFile(buf, "w", compression=pyzipper.ZIP_DEFLATED,
                              encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(ipcas_svc.zip_password())
        zf.writestr("gl02.csv", df.to_csv(index=False).encode("utf-8-sig"))
    return buf.getvalue()


def _gl02_row(customer="1000-003046328", dramount="500000"):
    return {"TRBRCD": "1000", "CUSTOMER": customer, "CRAMOUNT": "0", "DRAMOUNT": dramount,
            "REFERENCE": "1000API001002080", "REMARK": "TEST"}


# ─── Helper dựng danh sách file cho /start_upload (2026-09-02, bỏ chế độ thư mục máy chủ) ──────
# Cùng nội dung với _setup_hub_kenh/_setup_gl02/_setup_core_csv ở trên nhưng trả thẳng tuple
# multipart thay vì ghi ra đĩa — /start_upload là đường DUY NHẤT còn lại để chạy job qua API.

def _hub_kenh_upload_files(ngay="20260825") -> list[tuple]:
    hub_bytes = _make_hub_zip([_hub_row(_MSG_202RT, "TXID202RT")])
    kenh_buf = io.BytesIO()
    pd.DataFrame([_kenh_row(_MSG_202RT, "100000")], columns=_KENH_COLS).to_excel(
        kenh_buf, index=False, engine="openpyxl"
    )
    return [
        ("files", (f"doichieugd_{ngay}__05_DEN_9999_N.zip", hub_bytes, "application/zip")),
        ("files", ("kênh đến SPRT 202.xlsx", kenh_buf.getvalue(), _XLSX_MIME)),
    ]


def _gl02_upload_file(ngay="20260825") -> tuple:
    return ("files", (f"GL02_{ngay}_1000.zip", _make_gl02_zip([_gl02_row()]), "application/zip"))


def _core_csv_upload_file() -> tuple:
    df = pd.DataFrame([_gl02_row()], columns=_GL02_COLS)
    return ("files", ("202_DEN.csv", df.to_csv(index=False).encode("utf-8-sig"), "text/csv"))


def _doc_csv(content: bytes) -> list[tuple]:
    """Đọc 1 file CSV kết quả (đổi 2026-08-31, xem export.py) — trả list dòng dạng tuple, cùng
    hình dạng `_doc_sheets()` trả cho 1 sheet, để so sánh dễ với assertion cũ."""
    text = content.decode("utf-8-sig")
    return [tuple(row) for row in csv.reader(text.splitlines())]


def _doc_sheets(content: bytes) -> dict[str, list[tuple]]:
    """Đọc thẳng bằng `openpyxl` (KHÔNG qua `pd.read_excel`) — venv hiện có `openpyxl` 3.1.2,
    thấp hơn mức tối thiểu pandas tự đặt ra để đọc (>=3.1.5), dù bản thân openpyxl vẫn đọc/ghi
    đúng. Không nâng cấp dependency chung chỉ để phục vụ 1 test."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return {name: [tuple(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}


def _wait_done(admin_client, job_id, timeout_s=15):
    deadline = time.time() + timeout_s
    prog = None
    while time.time() < deadline:
        r = admin_client.get(f"/api/doi_chieu_song_phuong_kenh_core/poll/{job_id}")
        assert r.status_code == 200
        prog = r.json()
        if prog["status"] in ("done", "error", "cancelled"):
            return prog
        time.sleep(0.05)
    raise AssertionError(f"Job không hoàn thành sau {timeout_s}s: {prog}")


class TestStartFolderEndpoint:
    """2026-09-02: chế độ "chọn thư mục máy chủ" đã bỏ hẳn (review khanhbq693 PR#70 mục A) —
    các test dưới đây trước dùng `/start_folder` chỉ như đường tắt dựng fixture qua `tmp_path`,
    thật ra đang kiểm THUẬT TOÁN điều phối Kênh↔Hub/Hub↔Core (không phải chính chế độ thư mục),
    nên chuyển sang `/start_upload` — không mất độ phủ test nào."""

    def test_full_flow_ca_2_buoc_ra_ket_qua(self, admin_client, monkeypatch, tmp_path):
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[*_hub_kenh_upload_files(), _gl02_upload_file()],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        assert r.status_code == 200
        job_id = r.json()["job_id"]

        prog = _wait_done(admin_client, job_id)
        assert prog["status"] == "done", prog
        assert prog["ket_qua"]["kenh_hub"] is not None
        assert prog["ket_qua"]["hub_core"] is not None
        assert prog["ket_qua"]["trang_thai"]["kenh_hub"] == {"trang_thai": "da_doi_chieu", "ly_do": None}
        assert prog["ket_qua"]["trang_thai"]["hub_core"] == {"trang_thai": "da_doi_chieu", "ly_do": None}
        assert any(f.startswith("doi_chieu_song_phuong_kenh_tonghop") for f in prog["files"])
        assert any(f == "doi_chieu_song_phuong_kenh_hub_chi_tiet.csv" for f in prog["files"])
        assert any(f == "doi_chieu_song_phuong_kenh_kenh_chi_tiet.csv" for f in prog["files"])
        assert any(f.startswith("hub_core_202_") and f.endswith(".xlsx") for f in prog["files"])
        assert any(f == "hub_core_202_20260825_core_chi_tiet.csv" for f in prog["files"])
        assert any(f == "hub_core_202_20260825_hub_chi_tiet.csv" for f in prog["files"])
        assert any(f.startswith("bao_cao_tong_hop_202_") for f in prog["files"])
        # File gộp phải đứng đầu danh sách (báo cáo chính)
        assert prog["files"][0].startswith("bao_cao_tong_hop_202_")

        # Tải thử 1 file mỗi bước — 2026-08-31: chi tiết đổi sang CSV (đo thật ghi Excel chiếm
        # 60% tổng thời gian job, xem Implementation-notes.html card 98), chỉ bảng tổng hợp còn Excel.
        tonghop = next(f for f in prog["files"] if f.startswith("doi_chieu_song_phuong_kenh_tonghop"))
        r_dl = admin_client.get(f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/{tonghop}")
        assert r_dl.status_code == 200 and r_dl.headers["content-type"] == _XLSX_MIME
        kenh_sheets = _doc_sheets(r_dl.content)
        assert set(kenh_sheets.keys()) == {"Bang1_TongHop"}
        bang1_kenh_rows = kenh_sheets["Bang1_TongHop"]

        r_dl_hub = admin_client.get(
            f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/doi_chieu_song_phuong_kenh_hub_chi_tiet.csv"
        )
        assert r_dl_hub.status_code == 200 and r_dl_hub.headers["content-type"].startswith("text/csv")
        hub_chi_tiet_rows = _doc_csv(r_dl_hub.content)
        r_dl_kenh = admin_client.get(
            f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/doi_chieu_song_phuong_kenh_kenh_chi_tiet.csv"
        )
        assert r_dl_kenh.status_code == 200 and r_dl_kenh.headers["content-type"].startswith("text/csv")
        kenh_chi_tiet_rows = _doc_csv(r_dl_kenh.content)
        # 1 dòng HUB (202-SPRT, xem _setup_hub_kenh) + header — cột "Ngày"/"Ngân hàng"/"Loại" đầu tiên
        assert hub_chi_tiet_rows[0][:3] == ("Ngày", "Ngân hàng", "Loại")
        assert len(hub_chi_tiet_rows) == 2
        assert hub_chi_tiet_rows[1][:3] == ("20260825", "202", "SPRT")
        assert kenh_chi_tiet_rows[0][:3] == ("Ngày", "Ngân hàng", "Loại")
        assert len(kenh_chi_tiet_rows) == 2
        assert kenh_chi_tiet_rows[1][:3] == ("20260825", "202", "SPRT")

        hub_core_file = next(f for f in prog["files"] if f.startswith("hub_core_202_") and f.endswith(".xlsx"))
        r_dl2 = admin_client.get(f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/{hub_core_file}")
        assert r_dl2.status_code == 200 and r_dl2.headers["content-type"] == _XLSX_MIME
        hub_core_sheets = _doc_sheets(r_dl2.content)
        assert set(hub_core_sheets.keys()) == {"TongHop"}
        tonghop_core_rows = hub_core_sheets["TongHop"]

        r_dl_core_ct = admin_client.get(
            f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/hub_core_202_20260825_core_chi_tiet.csv"
        )
        assert r_dl_core_ct.status_code == 200 and r_dl_core_ct.headers["content-type"].startswith("text/csv")
        r_dl_hub_ct = admin_client.get(
            f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/hub_core_202_20260825_hub_chi_tiet.csv"
        )
        assert r_dl_hub_ct.status_code == 200 and r_dl_hub_ct.headers["content-type"].startswith("text/csv")

        # File gộp: 2 sheet, số liệu phải khớp tuyệt đối với 2 file riêng ở trên
        bao_cao = next(f for f in prog["files"] if f.startswith("bao_cao_tong_hop_202_"))
        r_dl3 = admin_client.get(f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/{bao_cao}")
        assert r_dl3.status_code == 200 and r_dl3.headers["content-type"] == _XLSX_MIME
        sheets = _doc_sheets(r_dl3.content)
        assert set(sheets.keys()) == {"TrangThai", "Bang1_KenhHub", "TongHop_HubCore"}
        assert sheets["Bang1_KenhHub"] == bang1_kenh_rows
        assert sheets["TongHop_HubCore"] == tonghop_core_rows
        assert sheets["TrangThai"] == [
            ("Bước", "Trạng thái", "Lý do"),
            ("Kênh↔Hub", "Đã đối chiếu", None),
            ("Hub↔Core", "Đã đối chiếu", None),
        ]

    def test_hub_khong_bi_doc_lai_giua_kenh_hub_va_hub_core(self, admin_client, monkeypatch, tmp_path):
        """2026-08-31, tối ưu hiệu năng — HUB T (202) trước đây bị Bước 2/2 Hub↔Core đọc+giải nén
        lại từ đầu dù Bước 1/2 Kênh↔Hub vừa đọc xong cùng file trong cùng job. Khoá hành vi: đúng
        1 lần gọi `load_hub_zip()` cho cả job (không phải 2)."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")

        calls = []
        real_load_hub_zip = kenh_pipeline_mod.load_hub_zip

        def dem_va_goi(*args, **kwargs):
            calls.append(1)
            return real_load_hub_zip(*args, **kwargs)

        monkeypatch.setattr(kenh_pipeline_mod, "load_hub_zip", dem_va_goi)
        monkeypatch.setattr(core_pipeline_mod, "load_hub_zip", dem_va_goi)

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[*_hub_kenh_upload_files(), _gl02_upload_file()],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        job_id = r.json()["job_id"]
        prog = _wait_done(admin_client, job_id)
        assert prog["status"] == "done", prog
        assert prog["ket_qua"]["kenh_hub"] is not None
        assert prog["ket_qua"]["hub_core"] is not None
        assert len(calls) == 1, f"HUB bị đọc {len(calls)} lần trong job, kỳ vọng đúng 1 lần"

    def test_thieu_gl02_hub_core_loi_khong_chan_kenh_hub(self, admin_client, monkeypatch, tmp_path):
        """Không có GL02/CSV -> Hub↔Core lỗi, nhưng Kênh↔Hub (đủ file) vẫn phải ra kết quả —
        đúng triết lý 'lỗi 1 bước không chặn bước kia'."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")
        # KHÔNG gửi file GL02/CSV — thiếu hẳn dữ liệu CORE

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=_hub_kenh_upload_files(),
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        job_id = r.json()["job_id"]
        prog = _wait_done(admin_client, job_id)
        assert prog["status"] == "done", prog
        assert prog["ket_qua"]["kenh_hub"] is not None
        assert prog["ket_qua"]["hub_core"] is None
        assert prog["ket_qua"]["trang_thai"]["kenh_hub"] == {"trang_thai": "da_doi_chieu", "ly_do": None}
        assert prog["ket_qua"]["trang_thai"]["hub_core"]["trang_thai"] == "chua_doi_chieu"
        assert prog["ket_qua"]["trang_thai"]["hub_core"]["ly_do"]
        assert any(f.startswith("doi_chieu_song_phuong_kenh_tonghop") for f in prog["files"])
        assert not any(f.startswith("hub_core_") for f in prog["files"])

        bao_cao = next(f for f in prog["files"] if f.startswith("bao_cao_tong_hop_202_"))
        r_dl = admin_client.get(f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/{bao_cao}")
        sheets = _doc_sheets(r_dl.content)
        assert set(sheets.keys()) == {"TrangThai", "Bang1_KenhHub"}
        trang_thai_rows = {row[0]: row[1] for row in sheets["TrangThai"][1:]}
        assert trang_thai_rows["Kênh↔Hub"] == "Đã đối chiếu"
        assert trang_thai_rows["Hub↔Core"] == "CHƯA ĐỐI CHIẾU"

    def test_2_file_hub_cung_khop_khong_tu_chon(self, admin_client, monkeypatch, tmp_path):
        """Quyết định 2026-08-30: nhiều người dùng có thể trỏ chung 1 thư mục server (mode 2)
        cùng lúc — 2 file HUB cùng khớp glob (VD 1 file người khác vừa thả vào) KHÔNG được tự
        đoán "mới nhất" như trước, phải báo rõ "chưa đối chiếu" thay vì âm thầm chọn nhầm file."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[
                *_hub_kenh_upload_files(),
                ("files", ("doichieugd_20260825__05_DEN_9999_N_v2.zip",
                           _make_hub_zip([_hub_row(_MSG_202RT, "TXID202RT")]), "application/zip")),
                _gl02_upload_file(),
            ],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        job_id = r.json()["job_id"]
        prog = _wait_done(admin_client, job_id)
        assert prog["ket_qua"]["kenh_hub"] is None
        assert prog["ket_qua"]["hub_core"] is None
        assert prog["ket_qua"]["trang_thai"]["kenh_hub"]["trang_thai"] == "chua_doi_chieu"
        assert "nhiều file hub" in prog["ket_qua"]["trang_thai"]["kenh_hub"]["ly_do"].lower()
        assert any("[LỖI]" in msg and "khớp cùng lúc" in msg for msg in prog["logs"])

    def test_ca_2_buoc_deu_thieu_du_lieu_bao_job_loi(self, admin_client, monkeypatch, tmp_path):
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            # File không khớp bất kỳ pattern HUB/kênh/CORE nào — mô phỏng "thiếu hẳn dữ liệu"
            files=[("files", ("khong_lien_quan.txt", b"khong phai du lieu", "text/plain"))],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        job_id = r.json()["job_id"]
        prog = _wait_done(admin_client, job_id)
        assert prog["status"] == "error"

    def test_co_san_csv_thi_khong_giai_ma_lai_gl02(self, admin_client, monkeypatch, tmp_path):
        """Quyết định 2026-08-28: có `{ma_nh}_DEN.csv` sẵn thì đọc thẳng, KHÔNG được gọi
        `process_zip()` (giải mã AES) dù file GL02 zip cũng tồn tại cùng thư mục."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")

        def _raise_neu_goi(*a, **kw):
            raise AssertionError("process_zip() KHÔNG được gọi khi đã có CSV phân loại sẵn")
        monkeypatch.setattr(ipcas_svc, "process_zip", _raise_neu_goi)

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            # GL02 + CSV cùng gửi lên — CSV phải được ưu tiên, process_zip() không được gọi
            files=[*_hub_kenh_upload_files(), _gl02_upload_file(), _core_csv_upload_file()],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        job_id = r.json()["job_id"]
        prog = _wait_done(admin_client, job_id)
        assert prog["status"] == "done", prog
        assert prog["ket_qua"]["hub_core"] is not None


class TestStartUploadEndpoint:
    """Chế độ tải file qua trình duyệt (quyết định 2026-08-28 đợt 3) — thay cho chỉ đường dẫn
    thư mục server "rất khó khăn". Lưu ra đĩa tại `{output_dir}/_upload/` rồi chạy y hệt chế độ
    thư mục — dùng chung toàn bộ pipeline, không nhân đôi logic."""

    def test_full_flow_qua_upload(self, admin_client, monkeypatch, tmp_path):
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        monkeypatch.setattr(ipcas_svc, "TEMP_DIR", tmp_path / "_out_ipcas")

        hub_bytes = _make_hub_zip([_hub_row(_MSG_202RT, "TXID202RT")])
        kenh_buf = io.BytesIO()
        pd.DataFrame([_kenh_row(_MSG_202RT, "100000")], columns=_KENH_COLS).to_excel(
            kenh_buf, index=False, engine="openpyxl"
        )
        gl02_bytes = _make_gl02_zip([_gl02_row()])

        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[
                ("files", ("doichieugd_20260825__05_DEN_9999_N.zip", hub_bytes, "application/zip")),
                ("files", ("kênh đến SPRT 202.xlsx", kenh_buf.getvalue(), _XLSX_MIME)),
                ("files", ("GL02_20260825_1000.zip", gl02_bytes, "application/zip")),
            ],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        assert r.status_code == 200, r.text
        job_id = r.json()["job_id"]

        prog = _wait_done(admin_client, job_id)
        assert prog["status"] == "done", prog
        assert prog["ket_qua"]["kenh_hub"] is not None
        assert prog["ket_qua"]["hub_core"] is not None

    def test_khong_chon_file_tra_422(self, admin_client):
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[], data={"ngay": "20260825", "ma_nh": "202"},
        )
        assert r.status_code == 422

    def test_ma_nh_khong_hop_le_tra_400(self, admin_client, monkeypatch, tmp_path):
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[("files", ("a.zip", b"x", "application/zip"))],
            data={"ngay": "20260825", "ma_nh": "999"},
        )
        assert r.status_code == 400

    def test_ngay_sai_dinh_dang_tra_400(self, admin_client, monkeypatch, tmp_path):
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[("files", ("a.zip", b"x", "application/zip"))],
            data={"ngay": "25-08-2026", "ma_nh": "202"},
        )
        assert r.status_code == 400

    def test_hai_file_cung_ten_bi_tu_choi_400(self, admin_client, monkeypatch, tmp_path):
        """2 file cùng tên trong 1 lượt upload phải bị chặn ngay — file sau sẽ ghi đè file
        trước trong input_dir mà không ai hay biết nếu không chặn."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[
                ("files", ("a.zip", b"noi dung 1", "application/zip")),
                ("files", ("a.zip", b"noi dung 2", "application/zip")),
            ],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        assert r.status_code == 400
        assert "cùng tên" in r.json()["detail"]
        # bo_job() phải dọn sạch — không để lại thư mục job "pending" mồ côi.
        out_dir = tmp_path / "_out"
        assert not out_dir.exists() or not list(out_dir.iterdir())

    def test_ten_file_co_duong_dan_bi_cat_ve_ten_thuan(self, admin_client, monkeypatch, tmp_path):
        """Tên file client gửi lên chứa `/`/`..` không được dùng thẳng để ghép đường dẫn ghi
        đĩa — chỉ giữ phần tên file (đúng lớp bảo vệ đã áp dụng ở `get_output_file`)."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/start_upload",
            files=[("files", ("../../evil.zip", b"khong-phai-zip-that", "application/zip"))],
            data={"ngay": "20260825", "ma_nh": "202"},
        )
        assert r.status_code == 200
        job_id = r.json()["job_id"]
        prog = _wait_done(admin_client, job_id)
        # Không tìm thấy file HUB/CORE hợp lệ nào (file giả) -> cả 2 bước lỗi -> job lỗi,
        # nhưng quan trọng là KHÔNG được ghi ra ngoài thư mục job (không lỗi 500/ghi lung tung).
        assert prog["status"] in ("error", "done")
        upload_dir = tmp_path / "_out" / job_id / "_upload"
        assert upload_dir.exists()
        assert (upload_dir / "evil.zip").exists()
        assert not (tmp_path / "_out" / "evil.zip").exists()


class TestCheckReadinessEndpoint:
    """Phần 2 (2026-08-30): banner cảnh báo TRƯỚC khi bấm "Chạy" — dò TÊN file (chỉ còn nhận
    `file_names`, không cần upload nội dung thật), không chặn nút Chạy.

    2026-09-02: bỏ hẳn `folder_path` (review khanhbq693 PR#70 mục A) — 2 test trước đây dựng
    fixture qua `tmp_path` rồi gọi `folder_path` đã chuyển sang liệt kê thẳng tên file."""

    def test_du_ca_hai(self, admin_client):
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/check_readiness",
            json={
                "file_names": [
                    "doichieugd_20260825__05_DEN_9999_N.zip",
                    "kênh đến SPRT 202.xlsx",
                    "GL02_20260825_1000.zip",
                ],
                "ngay": "20260825", "ma_nh": "202",
            },
        )
        assert r.status_code == 200, r.text
        assert r.json() == {"kenh_hub": "du", "hub_core": "du"}

    def test_file_names_mode_khong_can_upload_that(self, admin_client):
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/check_readiness",
            json={
                "file_names": [
                    "doichieugd_20260825__05_DEN_9999_N.zip",
                    "kênh đến SPRT 202.xlsx",
                ],
                "ngay": "20260825", "ma_nh": "202",
            },
        )
        assert r.status_code == 200
        body = r.json()
        assert body["kenh_hub"] == "du"
        assert body["hub_core"].startswith("thieu:")

    def test_thieu_file_names_tra_400(self, admin_client):
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/check_readiness",
            json={"file_names": [], "ngay": "20260825", "ma_nh": "202"},
        )
        assert r.status_code == 400

    def test_ngay_sai_dinh_dang_tra_400(self, admin_client):
        r = admin_client.post(
            "/api/doi_chieu_song_phuong_kenh_core/check_readiness",
            json={"file_names": [], "ngay": "25-08-2026", "ma_nh": "202"},
        )
        assert r.status_code == 400


class TestPollEndpoint:
    def test_unknown_job_returns_404(self, admin_client):
        r = admin_client.get("/api/doi_chieu_song_phuong_kenh_core/poll/khong-ton-tai")
        assert r.status_code == 404


class TestCancelEndpoint:
    def test_unknown_job_returns_404(self, admin_client):
        r = admin_client.post("/api/doi_chieu_song_phuong_kenh_core/cancel/khong-ton-tai")
        assert r.status_code == 404


class TestDownloadEndpoint:
    def test_unknown_job_returns_404(self, admin_client, monkeypatch, tmp_path):
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path)
        r = admin_client.get("/api/doi_chieu_song_phuong_kenh_core/download/khong-ton-tai/x.xlsx")
        assert r.status_code == 404

    def test_ten_trung_thu_muc_upload_tra_404_khong_phai_500(self, admin_client, monkeypatch, tmp_path):
        """`input_dir` của tao_job() nằm NGAY TRONG output_dir (`{output_dir}/_upload/`) — GET
        .../download/{job_id}/_upload từng lọt qua exists() rồi vỡ ở read_bytes()
        (IsADirectoryError) thay vì 404 rõ ràng. get_output_file() phải dùng is_file()."""
        monkeypatch.setattr(svc, "TEMP_DIR", tmp_path / "_out")
        job_id, input_dir = svc.tao_job("20260825", "202")
        assert input_dir.is_dir()

        r = admin_client.get(f"/api/doi_chieu_song_phuong_kenh_core/download/{job_id}/_upload")
        assert r.status_code == 404
