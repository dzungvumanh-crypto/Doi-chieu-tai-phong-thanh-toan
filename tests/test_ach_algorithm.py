"""
Synthetic tests cho pipeline Chấm ACH — b4_xu_ly_mis_di.py.

Trạng thái hiện tại (2026-07-23, theo quyết định Business Owner):
- Mục 2 (`_process_mis_di()`) — xây "MIS_đi sạch": bỏ CALD/ERPO/TPER (bước 1, vô
  điều kiện, không dùng lại cho mục đích nào khác) rồi lọc SESSION (bước 2). Trả về
  1 DataFrame duy nhất (KHÔNG phải tuple).
- Mục 2 bước 2, nhánh SESSION=NULL — Requirement mới (2026-07-23, KHÔNG phải bug):
  bỏ hẳn điều kiện lọc theo khung giờ `[23h T-1, 23h T)` cũ (làm bỏ sót giao dịch
  cuối phiên bị xử lý trễ sang session sau). Thay bằng tra SessionId thật của MSGREF
  trên "GW gốc" (dữ liệu GW hoàn toàn chưa lọc session/PrcFlg/làm sạch — xem
  `b3_xu_ly_gw.py::_gop_gw_goc()`), rồi so với ngày đối chiếu T/T-1 — xem
  `TestMisDiBuoc2SessionNullBrMoi`. `_process_mis_di()` đổi signature: bỏ
  `tpay_tu/tpay_den`, thêm `ngay_dt/df_gw_goc`.
- Cơ chế "Nguồn 1" (từng tách CALD/ERPO/TPER + session rỗng + khung giờ + MSGREF
  không có trong GW thành 1 luồng timeout riêng) đã bị GỠ HOÀN TOÀN — Business Owner
  xác nhận BR-ACH-001 là Business Rule chính thức, không giữ lại logic cũ (xem
  memory `project_ach_timeout_rule.md`, `TestNguon1DaGoBo` bên dưới).
- Mục 3 (`khop_voi_gw()`) — ĐÃ triển khai đầy đủ BR-ACH-001 (chốt 2026-07-21 qua
  nhiều vòng xác nhận với Business Owner), thứ tự bắt buộc:
  - **C.1** (mức NHÓM CN_TIỀN, KHÔNG chọn dòng): so tổng `COUNT_MIS` (mọi trạng
    thái) với `COUNT_GW` của cả nhóm — dùng `groupby(...).transform('size')`, KHÔNG
    dùng `cumcount()` (cumcount từng chọn dòng theo vị trí file — bị Business Owner
    bác bỏ vì đó là suy luận không có căn cứ khi 1 nhóm có nhiều dòng trùng khóa).
  - **C.2** (chỉ trong nhóm đã xác định thừa ở C.1): tra MSGREF (chuẩn hóa —
    `.strip().lstrip("'")`) cho TỪNG dòng TPAY, không phụ thuộc `PrcFlg`, không phụ
    thuộc thứ tự xuất hiện trong file. MSGREF không có trên GW sạch → nhánh 1A
    (timeout). MSGREF có trên GW sạch → nhánh 1B ("Timeout thật", vào khớp đúng với
    cột `MATCH_TYPE='TIMEOUT'`).
  - Trạng thái không phải TPAY trong nhóm thừa (SCNL/TXRT) → vẫn khớp đúng, không
    `MATCH_TYPE` (TPAY là điều kiện lọc duy nhất, không đổi).
- `tim_nhom_gw_thua()` (Requirement C.1a, phía GW-thừa) — xác định nhóm CN_TIỀN có
  COUNT_GW > COUNT_MIS, trả về DỮ LIỆU GỐC (không phải bảng tổng hợp COUNT) để chấm
  thủ công thẳng — làm rõ 2026-07-21: (1) COUNT_MIS=0 → "đã xác định chênh lệch"
  (toàn bộ dòng GW gốc), (2) cả 2 bên có dữ liệu nhưng lệch số → "cần đối chiếu thủ
  công" (toàn bộ dòng gốc CẢ 2 nguồn, không tự ghép cặp). Đã wiring vào
  `pipeline.py`/Excel (2 sheet GW_THUA_XAC_DINH/GW_CAN_DOI_CHIEU, 2026-07-23).

Chạy: python -m pytest tests/test_ach_algorithm.py -v
"""

from datetime import datetime

import pandas as pd
import pytest

from backend.services.ach.b4_xu_ly_mis_di import (
    _process_mis_di, khop_voi_gw, tim_nhom_gw_thua, ap_dung_confirm_mis_di,
    tim_giao_dich_bi_loai_session_null,
)
from backend.services.ach.pipeline import xuat_excel_confirm_mis_di


# ── Helper ────────────────────────────────────────────────────────────────────

_NGAY_DT = datetime(2026, 7, 11)  # ngày đối chiếu T mặc định cho các test không quan tâm BR SESSION=NULL
_DF_GW_GOC_RONG = pd.DataFrame(columns=['MSGREF', 'SessionId'])

SID = '16362'


def _row(refhub, chi_nhanh, trang_thai, trace, so_tien='500000',
         session='', ngay_kenh_tra='11/07/2026 10:00:00', msgref=None,
         ngay_giao_dich='11/07/2026'):
    return {
        'NGAY_GIAO_DICH':   ngay_giao_dich,
        'CHI_NHANH':        chi_nhanh,
        'REFHUB':           refhub,
        'MSGREF':           ('MSG' + refhub) if msgref is None else msgref,
        'MSGSEQ':           '1',
        'TXID':             'TX' + refhub,
        'KENH_THANH_TOAN':  'ACH-NAPAS',
        'TRANG_THAI_LENH':  trang_thai,
        'SO_TIEN':          so_tien,
        'TRACE':            trace,
        'SE_TRACE':         '',
        'SESSION':          session,
        'LOAI_LENH_OSB':    '',
        'NH_NHAN':          'NH TEST',
        'MA_GIAO_DICH':     'MGD' + refhub,
        'NOI_DUNG':         'noi dung test',
        'NGAY_KENH_TRA':    ngay_kenh_tra,
    }


def _mis_di(rows, ngay_dt=_NGAY_DT, df_gw_goc=_DF_GW_GOC_RONG):
    """Chạy _process_mis_di() và trả về DataFrame MIS_đi (giá trị đơn, không tuple)."""
    df = pd.DataFrame(rows)
    return _process_mis_di(df, SID, ngay_dt, df_gw_goc)


# ── Mục 2 bước 1 — TRẠNG THÁI: bỏ CALD/ERPO/TPER khỏi MIS_đi ───────────────────

class TestMisDiBuoc1TrangThai:
    """CALD/ERPO/TPER bị loại khỏi MIS_đi ngay bước 1 — không tham gia đối chiếu
    khớp/thừa với NPO, và (từ 2026-07-21) không còn được tách riêng cho bất kỳ mục
    đích nào khác nữa (xem TestNguon1DaGoBo)."""

    @pytest.mark.parametrize('trang_thai', ['CALD', 'ERPO', 'TPER'])
    def test_bi_loai_khoi_mis_di_du_session_dung(self, trang_thai):
        df_mis_di = _mis_di([_row('R1', '1000', trang_thai, '111', session=SID)])
        assert len(df_mis_di) == 0

    @pytest.mark.parametrize('trang_thai', ['CALD', 'ERPO', 'TPER'])
    def test_bi_loai_khoi_mis_di_du_session_rong_trong_khung(self, trang_thai):
        df_mis_di = _mis_di([_row('R1', '1000', trang_thai, '111', session='')])
        assert len(df_mis_di) == 0

    def test_scnl_va_txrt_khong_bi_loai(self):
        df_mis_di = _mis_di([
            _row('R1', '1000', 'SCNL', '111', session=SID),
            _row('R2', '1000', 'TXRT', '222', session=SID),
        ])
        assert len(df_mis_di) == 2

    def test_tpay_khong_bi_loai_boi_buoc_1(self):
        """TPAY không thuộc {CALD, ERPO, TPER} nên sống sót qua bước 1 — mục 3 mới
        là nơi TPAY được xét (thừa so GW → timeout)."""
        df_mis_di = _mis_di([_row('R1', '1000', 'TPAY', '111', session=SID)])
        assert len(df_mis_di) == 1


# ── Mục 2 bước 2 — SESSION khác NULL: không đổi ─────────────────────────────────

class TestMisDiBuoc2SessionKhacNull:
    def test_session_dung_vao_mis_di(self):
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session=SID)])
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['KEY_HUB'] == '1000111500000'

    def test_session_khac_bi_loai(self):
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session='99999')])
        assert len(df_mis_di) == 0


# ── Mục 2 bước 2 — SESSION=NULL: BR mới 2026-07-23 (tra GW gốc, bỏ khung giờ) ───

class TestMisDiBuoc2SessionNullBrMoi:
    """Business Rule mới (2026-07-23, thay thế hoàn toàn logic khung giờ cũ):
    SESSION=NULL được giữ/loại dựa trên SessionId thật tra trên GW gốc (chưa lọc
    session/PrcFlg/làm sạch) + so NGAY_GIAO_DICH với ngày đối chiếu.
    `_NGAY_DT` = 11/07/2026 (ngày T), T-1 = 10/07/2026. Lý do vì sao đã lọc bỏ do
    khung giờ: có giao dịch phát sinh cuối phiên nhưng do độ trễ hệ thống được xử lý
    ở session của phiên SAU — bộ lọc giờ cũ làm bỏ sót các giao dịch này."""

    def test_ngay_T_gw_session_rong_duoc_giu(self):
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': ['']})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'GW_SESSION_NULL'

    def test_ngay_T_gw_session_0000_duoc_giu(self):
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': ['0000']})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'GW_SESSION_0000'

    def test_ngay_T_gw_session_dung_doi_chieu_duoc_giu(self):
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': [SID]})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'GW_SESSION_DOI_CHIEU'

    def test_ngay_T_gw_session_khac_bi_loai(self):
        """SessionId=44 (khác session đối chiếu 16362) — GW đã nhận, xử lý ở session
        khác thật, không phải 'chưa đi kênh' — loại (đúng ví dụ trong yêu cầu gốc)."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': ['44']})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 0

    def test_ngay_T_khong_tim_thay_tren_gw_van_duoc_giu(self):
        """MSGREF không có dòng nào trên GW gốc — coi như chưa đi kênh, giữ nhưng
        đánh dấu riêng để phân biệt với 'GW có SessionId rỗng'."""
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSG_KHONG_CO',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=_DF_GW_GOC_RONG,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'KHONG_TIM_THAY_TREN_GW'

    def test_ngay_T1_chi_giu_khi_dung_session_doi_chieu(self):
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': [SID]})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='10/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'GW_SESSION_DOI_CHIEU'

    def test_ngay_T1_gw_session_rong_bi_loai(self):
        """Ở ngày T-1, KHÔNG áp dụng ngoại lệ rỗng/0000 như ngày T — chỉ chấp nhận
        đúng session đối chiếu (đúng ví dụ trong yêu cầu gốc: T-1 chỉ giữ session=42)."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': ['']})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='10/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 0

    def test_ngay_T1_khong_tim_thay_tren_gw_bi_loai(self):
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSG_KHONG_CO',
                  ngay_giao_dich='10/07/2026')],
            df_gw_goc=_DF_GW_GOC_RONG,
        )
        assert len(df_mis_di) == 0

    def test_gw_goc_nhieu_sessionid_duoc_giu_va_danh_dau(self):
        """MSGREF xuất hiện ≥2 lần trên GW gốc với SessionId khác nhau (dữ liệu chưa
        dedup theo nhiều sheet) — không xác định được 1 giá trị, giữ để chấm thủ
        công, KHÔNG tự chọn 1 giá trị theo vị trí."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1', 'MSGR1'], 'SessionId': [SID, '44']})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'GW_GOC_NHIEU_SESSIONID_KHAC_NHAU'

    def test_ngay_giao_dich_khac_t_va_t1_duoc_giu_tam(self):
        """NGAY_GIAO_DICH không phải T (11/07) cũng không phải T-1 (10/07) — Business
        Owner xác nhận (2026-07-23) chưa có quy tắc chính thức cho trường hợp này,
        giữ tạm + đánh dấu riêng chờ quy tắc."""
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSG_BAT_KY',
                  ngay_giao_dich='05/07/2026')],
            df_gw_goc=_DF_GW_GOC_RONG,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'NGAY_GIA_TRI_KHAC_T_VA_T-1'

    def test_chuan_hoa_msgref_dau_nhay_don_khi_tra_gw_goc(self):
        """MSGREF có dấu nháy đơn đầu ở MIS_đi vẫn phải tra đúng trên GW gốc sau khi
        chuẩn hóa (tái sử dụng `_chuan_hoa_msgref`)."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['0200970405071'], 'SessionId': [SID]})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session='', msgref="'0200970405071",
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'GW_SESSION_DOI_CHIEU'

    def test_session_khac_null_khong_bi_anh_huong_boi_gw_goc(self):
        """SESSION khác NULL — GW gốc dù có dữ liệu 'trông giống' cũng không ảnh
        hưởng, giữ nguyên logic hiện tại (chỉ so SESSION == session đối chiếu)."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSG_KHAC'], 'SessionId': ['99999']})
        df_mis_di = _mis_di(
            [_row('R1', '1000', 'SCNL', '111', session=SID, msgref='MSG_KHAC',
                  ngay_giao_dich='11/07/2026')],
            df_gw_goc=df_gw_goc,
        )
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == ''


# ── Điểm 1 (2026-07-31) — giao dịch SESSION=NULL bị loại, hiển thị CHỈ ĐỂ XEM ──

class TestGiaoDichBiLoaiSessionNull:
    """`tim_giao_dich_bi_loai_session_null()` — thay `tim_can_kiem_tra_thu_cong()`
    cũ (Milestone F Option C, 2 nhánh). Nhánh 'NGAY_GIA_TRI_KHAC_T_VA_T-1' đã BỎ
    khỏi hàm này (giờ tự nhiên hiển thị trong file confirm MIS_đi — xem
    TestApDungConfirmMisDi), chỉ còn nhánh MSGREF rỗng/không tìm thấy tại T-1.
    KHÔNG đổi Rule tự động của `_process_mis_di()` (kiểm chứng bằng cách so
    `len(df_mis_di)` không đổi ở các test dưới)."""

    def _bi_loai(self, rows, ngay_dt=_NGAY_DT, df_gw_goc=_DF_GW_GOC_RONG):
        df = pd.DataFrame(rows)
        return tim_giao_dich_bi_loai_session_null(df, SID, ngay_dt, df_gw_goc)

    def test_msgref_rong_tai_t1_xuat_hien(self):
        rows = [_row('R1', '1000', 'SCNL', '111', session='', msgref='',
                     ngay_giao_dich='10/07/2026')]
        df_bi_loai = self._bi_loai(rows)
        assert len(df_bi_loai) == 1
        assert df_bi_loai.iloc[0]['LY_DO_CAN_KIEM_TRA'] == 'MSGREF_RONG_TAI_T-1'
        # Không đổi Rule tự động — dòng này vẫn bị _process_mis_di() loại như cũ.
        df_mis_di = _mis_di(rows)
        assert len(df_mis_di) == 0

    def test_msgref_khong_tim_thay_tai_t1_xuat_hien(self):
        rows = [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSG_KHONG_CO',
                     ngay_giao_dich='10/07/2026')]
        df_bi_loai = self._bi_loai(rows)
        assert len(df_bi_loai) == 1
        assert df_bi_loai.iloc[0]['LY_DO_CAN_KIEM_TRA'] == 'MSGREF_KHONG_TIM_THAY_TREN_GW_TAI_T-1'
        df_mis_di = _mis_di(rows)
        assert len(df_mis_di) == 0

    def test_ngay_khac_t_va_t1_khong_xuat_hien_o_day_nua(self):
        """Nhánh 'NGAY_GIA_TRI_KHAC_T_VA_T-1' KHÔNG còn hiển thị ở hàm này (Điểm 1)
        — vẫn nằm trong MIS_đi, người chấm thấy trực tiếp qua file confirm."""
        rows = [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSG_BAT_KY',
                     ngay_giao_dich='05/07/2026')]
        df_bi_loai = self._bi_loai(rows)
        assert len(df_bi_loai) == 0
        # Vẫn được _process_mis_di() giữ tạm trong MIS_đi như cũ (Rule tự động không đổi).
        df_mis_di = _mis_di(rows)
        assert len(df_mis_di) == 1
        assert df_mis_di.iloc[0]['LY_DO_GIU_SESSION_NULL'] == 'NGAY_GIA_TRI_KHAC_T_VA_T-1'

    def test_t1_dung_session_doi_chieu_khong_xuat_hien(self):
        """Case đã CHỐT ('khác session → loại', xác nhận 2026-07-25) và case khớp
        đúng session đối chiếu — không thuộc nhánh bị loại còn lại."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': [SID]})
        rows = [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                     ngay_giao_dich='10/07/2026')]
        df_bi_loai = self._bi_loai(rows, df_gw_goc=df_gw_goc)
        assert len(df_bi_loai) == 0

    def test_t1_session_rong_tren_gw_khong_xuat_hien(self):
        """Case 'T-1, GW có nhưng SessionId rỗng' — đã loại theo Rule đã chốt (chỉ
        nhận đúng session đối chiếu ở T-1), không thuộc nhánh bị loại còn lại."""
        df_gw_goc = pd.DataFrame({'MSGREF': ['MSGR1'], 'SessionId': ['']})
        rows = [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSGR1',
                     ngay_giao_dich='10/07/2026')]
        df_bi_loai = self._bi_loai(rows, df_gw_goc=df_gw_goc)
        assert len(df_bi_loai) == 0

    def test_ngay_t_khong_xuat_hien(self):
        """Mọi nhánh ở ngày T đều đã được _process_mis_di() GIỮ với lý do riêng —
        chỉ nhánh T-1 mới có thể bị loại."""
        rows = [_row('R1', '1000', 'SCNL', '111', session='', msgref='MSG_KHONG_CO',
                     ngay_giao_dich='11/07/2026')]
        df_bi_loai = self._bi_loai(rows)
        assert len(df_bi_loai) == 0

    def test_cald_khong_xuat_hien_du_msgref_rong_tai_t1(self):
        """CALD/ERPO/TPER không bao giờ vào diện hiển thị — BR đã chốt, không đổi."""
        rows = [_row('R1', '1000', 'CALD', '111', session='', msgref='',
                     ngay_giao_dich='10/07/2026')]
        df_bi_loai = self._bi_loai(rows)
        assert len(df_bi_loai) == 0

    def test_khong_co_du_lieu_tra_ve_dataframe_rong_dung_cot(self):
        rows = [_row('R1', '1000', 'SCNL', '111', session=SID)]
        df_bi_loai = self._bi_loai(rows)
        assert len(df_bi_loai) == 0
        assert 'LY_DO_CAN_KIEM_TRA' in df_bi_loai.columns
        assert 'REFHUB' in df_bi_loai.columns


# ── Regression: cơ chế "Nguồn 1" đã bị gỡ hoàn toàn (2026-07-21) ────────────────

class TestNguon1DaGoBo:
    """Business Owner xác nhận BR-ACH-001 là Business Rule chính thức của phiên bản
    hiện tại — không giữ lại logic V1 ("Nguồn 1": CALD/ERPO/TPER + session rỗng +
    khung giờ + MSGREF không có trong GW ⇒ timeout riêng). `_process_mis_di()` giờ
    nhận `(df, session_id, ngay_dt, df_gw_goc, log_callback)` (2026-07-23 — thay
    `tpay_tu/tpay_den` bằng `ngay_dt/df_gw_goc` theo BR mới SESSION=NULL) — không còn
    tham số `gw_msgref_set` — và trả về 1 DataFrame duy nhất."""

    def test_process_mis_di_tra_ve_1_dataframe_khong_phai_tuple(self):
        df = pd.DataFrame([_row('R1', '1000', 'SCNL', '111', session=SID)])
        result = _process_mis_di(df, SID, _NGAY_DT, _DF_GW_GOC_RONG)
        assert isinstance(result, pd.DataFrame)

    def test_cald_thoa_dieu_kien_nguon_1_cu_van_khong_sinh_timeout(self):
        """Trước đây (Nguồn 1): CALD + session rỗng + đúng khung giờ + MSGREF thật +
        không có trong GW ⇒ được coi là "timeout không đi kênh" riêng. Giờ: CALD bị
        loại khỏi MIS_đi ngay bước 1, vô điều kiện — không còn đường nào để dòng này
        xuất hiện ở bất kỳ sheet timeout nào của pipeline hiện tại."""
        df_mis_di = _mis_di([_row('R1', '1000', 'CALD', '111', session='',
                                  msgref='MSG_KHONG_CO_O_GW')])
        assert len(df_mis_di) == 0


# ── Mục 3 — `khop_voi_gw` — BR-ACH-001 C.1 (nhóm) + C.2 (MSGREF) ───────────────

class TestKhopVoiGw:
    """C.1: nhóm CN_TIỀN thừa xác định ở MỨC NHÓM (so tổng COUNT_MIS/COUNT_GW),
    KHÔNG dùng cumcount để chọn dòng. C.2: trong nhóm thừa, tra MSGREF (đã chuẩn hóa)
    cho TỪNG dòng TPAY — không phụ thuộc `PrcFlg`, không phụ thuộc thứ tự xuất hiện
    trong file."""

    def test_nhom_khop_du_khong_co_timeout(self):
        """CN TIỀN khớp đủ số lượng với GW (nhóm không thừa) → toàn bộ vào 'khớp
        đúng', không có MATCH_TYPE."""
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session=SID, so_tien='500000')])
        dict_gw = {'1000500000': 1}
        df_gw = pd.DataFrame([_gw_row('1000500000', 'GWMSG1')])
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(khop) == 1
        assert len(timeout) == 0
        assert khop.iloc[0]['MATCH_TYPE'] == ''

    def test_nhom_thua_khong_phai_tpay_van_khop_dung(self):
        """Nhóm thừa nhưng KHÔNG phải TPAY (SCNL) → vẫn 'khớp đúng', không tra
        MSGREF, không có MATCH_TYPE. TPAY là điều kiện lọc duy nhất."""
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session=SID, so_tien='500000')])
        dict_gw = {}  # nhóm thừa: COUNT_MIS=1 > COUNT_GW=0
        df_gw = pd.DataFrame([_gw_row('9999999999', 'KHAC_NHOM')])
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(khop) == 1
        assert len(timeout) == 0
        assert khop.iloc[0]['MATCH_TYPE'] == ''

    def test_tpay_khong_co_msgref_tren_gw_la_timeout(self):
        """Nhóm thừa, TPAY, MSGREF KHÔNG có trên GW sạch → nhánh 1A, timeout."""
        df_mis_di = _mis_di([_row('R1', '1000', 'TPAY', '111', session=SID,
                                  so_tien='500000', msgref='MSGR1')])
        dict_gw = {}  # nhóm thừa: COUNT_MIS=1 > COUNT_GW=0
        df_gw = pd.DataFrame([_gw_row('9999999999', 'KHAC_MSGREF')])
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(khop) == 0
        assert len(timeout) == 1
        assert timeout.iloc[0]['REFHUB'] == 'R1'

    def test_tpay_co_msgref_tren_gw_la_timeout_that(self):
        """Nhóm thừa, TPAY, MSGREF CÓ trên GW sạch → nhánh 1B — KHÔNG phải thừa,
        vào 'khớp đúng' với MATCH_TYPE='TIMEOUT' (không quan tâm GW đang PrcFlg gì)."""
        df_mis_di = _mis_di([_row('R1', '1000', 'TPAY', '111', session=SID,
                                  so_tien='500000', msgref='MSGR1')])
        dict_gw = {}  # vẫn là nhóm "thừa" theo COUNT (GW không có KEY_GW này)
        df_gw = pd.DataFrame([_gw_row('1000999999', 'MSGR1')])  # MSGREF khớp, KEY_GW khác
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(timeout) == 0
        assert len(khop) == 1
        assert khop.iloc[0]['MATCH_TYPE'] == 'TIMEOUT'

    def test_tpay_du_so_gw_khong_can_tra_msgref(self):
        """Nhóm KHÔNG thừa (COUNT_MIS <= COUNT_GW) → TPAY khớp đúng ngay ở bước C.1,
        không cần tra MSGREF (dù MSGREF không hề có trên GW)."""
        df_mis_di = _mis_di([_row('R1', '1000', 'TPAY', '111', session=SID,
                                  so_tien='500000', msgref='MSG_KHONG_CO_O_GW')])
        dict_gw = {'1000500000': 1}
        df_gw = pd.DataFrame([_gw_row('1000500000', 'MSGREF_KHAC_HOAN_TOAN')])
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(khop) == 1
        assert len(timeout) == 0
        assert khop.iloc[0]['MATCH_TYPE'] == ''

    def test_3_tpay_2_gw_phan_loai_theo_msgref_khong_theo_vi_tri(self):
        """Regression quan trọng nhất: nhóm CN_TIỀN có 3 dòng TPAY, GW xác nhận 2.
        Trước đây (cumcount) sẽ chọn 2 dòng CUỐI theo vị trí file làm 'khớp', bất kể
        MSGREF. Giờ: tra MSGREF cho CẢ 3 dòng — dòng nào (bất kể vị trí) có MSGREF
        khớp GW thì là nhánh 1B (khớp đúng), dòng không khớp là nhánh 1A (timeout)."""
        key = '1000500000'
        df_mis_di = _mis_di([
            # Cố ý đặt dòng "không khớp" (R2) ở giữa, không phải cuối — để chắc chắn
            # kết quả không phụ thuộc thứ tự/cumcount.
            _row('R1', '1000', 'TPAY', '111', session=SID, so_tien='500000', msgref='MSGR1'),
            _row('R2', '1000', 'TPAY', '222', session=SID, so_tien='500000', msgref='MSGR2_KHONG_KHOP'),
            _row('R3', '1000', 'TPAY', '333', session=SID, so_tien='500000', msgref='MSGR3'),
        ])
        dict_gw = {key: 2}
        df_gw = pd.DataFrame([_gw_row(key, 'MSGR1'), _gw_row(key, 'MSGR3')])

        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(timeout) == 1
        assert timeout.iloc[0]['REFHUB'] == 'R2'
        assert len(khop) == 2
        assert set(khop['REFHUB']) == {'R1', 'R3'}
        assert (khop['MATCH_TYPE'] == 'TIMEOUT').all()

    def test_chuan_hoa_msgref_dau_nhay_don(self):
        """MSGREF có dấu nháy đơn đầu ở 1 hoặc cả 2 bên vẫn phải so khớp đúng sau
        khi chuẩn hóa (.strip().lstrip("'"))."""
        df_mis_di = _mis_di([_row('R1', '1000', 'TPAY', '111', session=SID,
                                  so_tien='500000', msgref="'0200970405071")])
        dict_gw = {}  # nhóm thừa theo COUNT
        df_gw = pd.DataFrame([_gw_row('9999999999', '0200970405071')])  # không có nháy
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, df_gw)
        assert len(timeout) == 0
        assert len(khop) == 1
        assert khop.iloc[0]['MATCH_TYPE'] == 'TIMEOUT'


# ── Requirement C.1a — GW-thừa, dữ liệu gốc (`tim_nhom_gw_thua`) ────────────────

def _gw_row(key_gw, msgref):
    """Dòng GW tối giản — chỉ cần đủ để test tính nhóm + bảo toàn dữ liệu gốc."""
    return {'KEY_GW': key_gw, 'MSGREF': msgref, 'BRCD': key_gw[:4], 'STTLMAMT': key_gw[4:]}


class TestTimNhomGwThua:
    """Chỉ dựa vào việc bên MIS_đi có bản ghi nào không cho mỗi nhóm CN_TIỀN — KHÔNG
    dùng PrcFlg/SessionId, KHÔNG ghép cặp MSGREF hay suy luận nào khác (đúng làm rõ
    2026-07-21 của Business Owner cho Requirement C.1a — xuất dữ liệu GỐC, không phải
    bảng tổng hợp COUNT, để chấm thủ công thẳng không phải map ngược lại)."""

    def test_gw_thua_xac_dinh_khi_mis_khong_co_ban_ghi(self):
        """COUNT_MIS=0 cho 1 khóa CN_TIỀN → toàn bộ dòng GW gốc của khóa đó là
        'đã xác định chênh lệch' (Trường hợp 1 — không có gì để nhầm lẫn)."""
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session=SID, so_tien='500000')])
        df_gw = pd.DataFrame([
            _gw_row('9999999999', 'GWMSG1'),
            _gw_row('9999999999', 'GWMSG2'),
        ])
        df_xac_dinh, df_can_doi_chieu = tim_nhom_gw_thua(df_mis_di, df_gw)
        assert len(df_xac_dinh) == 2
        assert set(df_xac_dinh['MSGREF']) == {'GWMSG1', 'GWMSG2'}  # dữ liệu gốc còn nguyên
        assert set(df_xac_dinh['SOURCE']) == {'GW'}
        assert set(df_xac_dinh['NHOM_CN_TIEN']) == {'9999999999'}
        assert len(df_can_doi_chieu) == 0

    def test_gw_thua_can_doi_chieu_khi_ca_2_ben_co_du_lieu(self):
        """GW=2, MIS=1 cùng khóa → không xác định được bản ghi cụ thể nào thừa
        (Trường hợp 2) → xuất TOÀN BỘ dòng gốc của cả 2 nguồn (2 GW + 1 MIS = 3)."""
        key = '1000500000'
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session=SID, so_tien='500000')])
        df_gw = pd.DataFrame([_gw_row(key, 'GWMSG1'), _gw_row(key, 'GWMSG2')])

        df_xac_dinh, df_can_doi_chieu = tim_nhom_gw_thua(df_mis_di, df_gw)
        assert len(df_xac_dinh) == 0
        assert len(df_can_doi_chieu) == 3
        assert (df_can_doi_chieu['NHOM_CN_TIEN'] == key).all()
        assert (df_can_doi_chieu['SOURCE'] == 'GW').sum() == 2
        assert (df_can_doi_chieu['SOURCE'] == 'MIS').sum() == 1


# ── Bước 2 — Checkpoint xác nhận thủ công tại MIS_đi (Điểm 1, 2026-07-31) ───────

def _ghi_confirm(path, loai_bo_refhubs=None, refhub_bo_sung=None):
    """Helper test: mở file confirm MIS_đi (Bước 1) bằng openpyxl, điền cột
    LOAI_BO='loại bỏ' cho các REFHUB trong `loai_bo_refhubs`, và paste REFHUB vào
    vùng bổ sung — mô phỏng đúng thao tác người đối chiếu làm bằng tay."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb['MIS_DI_CONFIRM']
    header = [c.value for c in ws[1]]
    col_refhub  = header.index('REFHUB') + 1
    col_loai_bo = header.index('LOAI_BO') + 1

    note_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().upper().startswith('BỔ SUNG'):
            note_row = r
            break

    loai_bo_refhubs = loai_bo_refhubs or []
    data_rows = range(2, note_row) if note_row else range(2, ws.max_row + 1)
    for r in data_rows:
        refhub = ws.cell(row=r, column=col_refhub).value
        if refhub in loai_bo_refhubs:
            ws.cell(row=r, column=col_loai_bo, value='loại bỏ')

    if refhub_bo_sung:
        assert note_row is not None, 'File không có vùng bổ sung REFHUB'
        for i, r in enumerate(refhub_bo_sung):
            ws.cell(row=note_row + 2 + i, column=1, value=r)

    wb.save(path)


def _xuat_confirm(tmp_path, df_mis_di, df_bi_loai=None):
    if df_bi_loai is None:
        df_bi_loai = pd.DataFrame()
    return xuat_excel_confirm_mis_di(str(tmp_path), SID, _NGAY_DT, df_mis_di, df_bi_loai)


class TestApDungConfirmMisDi:
    """Bước 2 — engine đọc file confirm MIS_đi do xuat_excel_confirm_mis_di() (Bước
    1) sinh ra (Điểm 1, 2026-07-31 — thay hẳn cơ chế Timeout-confirm cũ). Công
    thức: MIS_đi chuẩn = (MIS_đi ban đầu − dòng LOAI_BO='loại bỏ') + (REFHUB hợp lệ
    ở khu bổ sung, tra trên MIS_hub RAW). Test round-trip qua file Excel thật
    (không mock I/O) — đúng nguyên tắc doi-chieu skill."""

    def _mis_di_3_dong(self):
        rows = [
            _row('RA', '5000', 'TPAY', '111', session=SID, so_tien='500000', msgref='MSGA'),
            _row('RB', '6000', 'TPAY', '222', session=SID, so_tien='600000', msgref='MSGB'),
            _row('RC', '7000', 'TPAY', '333', session=SID, so_tien='700000', msgref='MSGC'),
        ]
        return _mis_di(rows), rows

    def test_loai_bo_mot_dong(self, tmp_path):
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, loai_bo_refhubs=['RB'])

        df_mis_di_raw = pd.DataFrame(rows)
        df_chuan = ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

        assert set(df_chuan['REFHUB']) == {'RA', 'RC'}
        assert len(df_chuan) == 2

    def test_khong_chon_gi_giu_nguyen_toan_bo(self, tmp_path):
        """Cột LOAI_BO để trống toàn bộ (mặc định) → không dòng nào bị loại."""
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)

        df_mis_di_raw = pd.DataFrame(rows)
        df_chuan = ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

        assert set(df_chuan['REFHUB']) == {'RA', 'RB', 'RC'}
        assert len(df_chuan) == 3

    def test_bo_sung_refhub_bi_loc_oan(self, tmp_path):
        """REFHUB bị lọc oan (có trên MIS_hub thô, chưa có trong MIS_đi) — tra trên
        RAW để kéo về, đủ cột tính toán (CN tiền Hub/KEY_HUB) để dùng tiếp được ở
        khop_voi_gw()/doi_chieu_di()."""
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, refhub_bo_sung=['RD'])

        rows_raw = rows + [_row('RD', '8000', 'TPAY', '444', session='99999',
                                so_tien='800000', msgref='MSGD')]
        df_mis_di_raw = pd.DataFrame(rows_raw)
        df_chuan = ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

        assert set(df_chuan['REFHUB']) == {'RA', 'RB', 'RC', 'RD'}
        bosung = df_chuan[df_chuan['REFHUB'] == 'RD'].iloc[0]
        assert str(bosung['SO_TIEN']) == '800000'
        assert bosung['CN tiền Hub'] == '8000800000'
        assert 'KEY_HUB' in df_chuan.columns and bosung['KEY_HUB'] != ''

    def test_bo_sung_khong_ap_rang_buoc_trang_thai_loai_tru(self, tmp_path):
        """KHÁC cơ chế Timeout-confirm cũ: REFHUB bổ sung có trạng thái
        CALD/ERPO/TPER vẫn được chấp nhận (lệnh người chấm luôn được tôn trọng —
        trạng thái tại thời điểm xuất file chưa phải trạng thái cuối cùng)."""
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, refhub_bo_sung=['RD'])

        rows_raw = rows + [_row('RD', '8000', 'CALD', '444', so_tien='800000', msgref='MSGD')]
        df_mis_di_raw = pd.DataFrame(rows_raw)
        df_chuan = ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

        assert 'RD' in set(df_chuan['REFHUB'])

    def test_loi_refhub_trung_lap_tren_raw(self, tmp_path):
        """REFHUB bổ sung khớp ≥2 dòng trên MIS_hub thô — báo lỗi, không tự chọn 1
        dòng theo vị trí."""
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, refhub_bo_sung=['RD'])

        rows_raw = rows + [
            _row('RD', '8000', 'TPAY', '444', so_tien='800000', msgref='MSGD1'),
            _row('RD', '8000', 'TPAY', '555', so_tien='900000', msgref='MSGD2'),
        ]
        df_mis_di_raw = pd.DataFrame(rows_raw)
        with pytest.raises(ValueError, match='khớp ≥2 dòng'):
            ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

    def test_loi_refhub_bo_sung_khong_tim_thay(self, tmp_path):
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, refhub_bo_sung=['R_KHONG_TON_TAI'])

        df_mis_di_raw = pd.DataFrame(rows)
        with pytest.raises(ValueError, match='Không tìm thấy REFHUB'):
            ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

    def test_loi_refhub_bo_sung_da_co_san(self, tmp_path):
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, refhub_bo_sung=['RA'])  # RA đã có sẵn trong MIS_đi

        df_mis_di_raw = pd.DataFrame(rows)
        with pytest.raises(ValueError, match='đã có sẵn'):
            ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

    def test_loi_gia_tri_loai_bo_khong_hop_le(self, tmp_path):
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb['MIS_DI_CONFIRM']
        header = [c.value for c in ws[1]]
        col_loai_bo = header.index('LOAI_BO') + 1
        ws.cell(row=2, column=col_loai_bo, value='không rõ')
        wb.save(path)

        df_mis_di_raw = pd.DataFrame(rows)
        with pytest.raises(ValueError, match='không hợp lệ'):
            ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

    def test_loi_trung_lap_refhub_bo_sung_trong_paste(self, tmp_path):
        df_mis_di, rows = self._mis_di_3_dong()
        path = _xuat_confirm(tmp_path, df_mis_di)
        _ghi_confirm(path, refhub_bo_sung=['RD', 'RD'])

        rows_raw = rows + [_row('RD', '8000', 'TPAY', '444', so_tien='800000', msgref='MSGD')]
        df_mis_di_raw = pd.DataFrame(rows_raw)
        with pytest.raises(ValueError, match='trùng lặp'):
            ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)

    def test_sheet_khong_co_giao_dich(self, tmp_path):
        """MIS_đi rỗng (0 dòng qua bước 1+2) → sheet MIS_DI_CONFIRM chỉ có thông báo
        rỗng, không có cột LOAI_BO — không lỗi, trả về MIS_đi chuẩn rỗng."""
        rows = [_row('R1', '1000', 'SCNL', '111', session='99999')]  # session khác → loại hết
        df_mis_di = _mis_di(rows)
        assert len(df_mis_di) == 0

        path = _xuat_confirm(tmp_path, df_mis_di)
        df_mis_di_raw = pd.DataFrame(rows)
        df_chuan = ap_dung_confirm_mis_di(path, df_mis_di, df_mis_di_raw)
        assert len(df_chuan) == 0

    def test_khop_du_khong_xuat_hien_o_sheet_nao(self):
        """COUNT_GW == COUNT_MIS (khớp đủ) → không phải chênh lệch, không xuất hiện
        ở df_xac_dinh lẫn df_can_doi_chieu."""
        key = '1000500000'
        df_mis_di = _mis_di([_row('R1', '1000', 'SCNL', '111', session=SID, so_tien='500000')])
        df_gw = pd.DataFrame([_gw_row(key, 'GWMSG1')])

        df_xac_dinh, df_can_doi_chieu = tim_nhom_gw_thua(df_mis_di, df_gw)
        assert len(df_xac_dinh) == 0
        assert len(df_can_doi_chieu) == 0

    def test_mis_thua_khong_duoc_tinh_la_gw_thua(self):
        """MIS thừa (COUNT_MIS > COUNT_GW) là chiều ngược lại (C.1b, khop_voi_gw) —
        không được lẫn vào kết quả của tim_nhom_gw_thua()."""
        key = '1000500000'
        df_mis_di = _mis_di([
            _row('R1', '1000', 'SCNL', '111', session=SID, so_tien='500000'),
            _row('R2', '1000', 'SCNL', '222', session=SID, so_tien='500000'),
        ])
        df_gw = pd.DataFrame([_gw_row(key, 'GWMSG1')])

        df_xac_dinh, df_can_doi_chieu = tim_nhom_gw_thua(df_mis_di, df_gw)
        assert len(df_xac_dinh) == 0
        assert len(df_can_doi_chieu) == 0


# ── Bất biến số học ──────────────────────────────────────────────────────────

class TestBatBienSoHoc:
    _DF_GW_KHONG_LIEN_QUAN = pd.DataFrame([_gw_row('0000000000', 'KHONG_LIEN_QUAN')])

    def test_tong_dong_khop_va_timeout_bang_dau_vao(self):
        rows = [
            _row('R1', '1000', 'SCNL', '111', session=SID, so_tien='100000'),
            _row('R2', '1000', 'TPAY', '222', session=SID, so_tien='100000'),
            _row('R3', '2000', 'TPAY', '333', session=SID, so_tien='200000'),
            _row('R4', '2000', 'TXRT', '444', session=SID, so_tien='200000'),
        ]
        df_mis_di = _mis_di(rows)
        dict_gw = {'1000100000': 1}  # chỉ CN_TIEN của R1/R2 có 1 xác nhận trên GW
        khop, timeout = khop_voi_gw(df_mis_di, dict_gw, self._DF_GW_KHONG_LIEN_QUAN)
        assert len(khop) + len(timeout) == len(df_mis_di)

    def test_tong_tien_bao_toan(self):
        rows = [
            _row('R1', '1000', 'TPAY', '111', session=SID, so_tien='700000'),
            _row('R2', '1000', 'TPAY', '222', session=SID, so_tien='300000'),
        ]
        df_mis_di = _mis_di(rows)
        khop, timeout = khop_voi_gw(df_mis_di, {}, self._DF_GW_KHONG_LIEN_QUAN)
        tong = khop['SO_TIEN'].astype('int64').sum() + timeout['SO_TIEN'].astype('int64').sum()
        assert tong == 700000 + 300000
        assert set(khop['REFHUB']) | set(timeout['REFHUB']) == {'R1', 'R2'}
