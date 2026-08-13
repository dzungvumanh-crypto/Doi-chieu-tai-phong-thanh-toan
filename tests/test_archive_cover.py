"""
Test "In bìa hồ sơ" — đọc Excel tra cứu lưu trữ và điền vào mẫu bìa M01/LHS.

Cột lấy dữ liệu (sheet "Data", dòng tiêu đề ở dòng 2):
  C "Tên hồ sơ/ĐVBQ"  → dòng tiêu đề + Ngày mở (ngày ĐẦU TIÊN trong chuỗi)
  F "Ngày CVKT"       → Ngày công việc kết thúc
  G "Số tờ"           → "Gồm: … tờ"
  I "Mã vạch"         → Ký hiệu thông tin + chuỗi mã vạch *…*

Bìa nằm trong VML textbox nên mọi phép điền đều dựa vào nhãn làm mốc; test này
bắt được ngay khi template đổi cấu trúc (TemplateMismatchError) thay vì lặng lẽ
sinh ra bìa trống.
"""
import io
import re
import sqlite3
import zipfile

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.core.deps import get_current_staff
from backend.database import get_db
from backend.main import app
from backend.services import archive_cover_service as S

_RX_T = re.compile(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", re.S)

_SCHEMA_NHOM = """
CREATE TABLE user_groups   (id INTEGER PRIMARY KEY, name TEXT, is_active INTEGER DEFAULT 1);
CREATE TABLE group_members (group_id INTEGER, staff_id INTEGER);
CREATE TABLE group_features(group_id INTEGER, feature_code TEXT);
"""

_HANG = [
    # (tên hồ sơ, ngày CVKT, số tờ, mã vạch)
    ("Nhật ký chứng từ ngày 01/01/2025 của Phòng Thanh toán", "01/01/2025", "1",
     "1000.P026.177935.1"),
    ("Nhật ký chứng từ ngày 17/12/2025, 18/12/2025, 19/12/2025 của Phòng Thanh toán",
     "19/12/2025", "2", "1000.P026.178068.1"),
    ("Hồ sơ không ghi ngày nào cả", "31/12/2025", "1", "1000.P026.178099.1"),
]


def _text(docx_bytes: bytes) -> str:
    xml = zipfile.ZipFile(io.BytesIO(docx_bytes)).read("word/document.xml").decode("utf-8")
    return "".join(_RX_T.findall(xml))


def _excel_gia_lap() -> bytes:
    """Dựng lại đúng bố cục file LT_HS_TRACUU_*.xls: dòng 1 tiêu đề chung, dòng 2 header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1, column=5, value="Tra cứu hồ sơ tài liệu")
    for col, hdr in enumerate(
        ["STT", "Số hồ sơ/ĐVBQ", "Tên hồ sơ/ĐVBQ", "Thời gian BĐ", "Thời gian KT",
         "Ngày CVKT", "Số tờ", "Trạng thái", "Mã vạch", "Người cập nhật",
         "Ngày cập nhật", "ID"], 1):
        ws.cell(row=2, column=col, value=hdr)
    for i, (ten, cvkt, so_to, ma) in enumerate(_HANG, 3):
        ws.cell(row=i, column=1, value=str(i - 2))
        ws.cell(row=i, column=3, value=ten)
        ws.cell(row=i, column=6, value=cvkt)
        ws.cell(row=i, column=7, value=so_to)
        ws.cell(row=i, column=9, value=ma)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Đọc Excel ────────────────────────────────────────────────────────────────

def test_ngay_mo_lay_ngay_dau_tien_trong_ten_ho_so():
    recs = S.parse_lookup_excel(_excel_gia_lap(), "t.xlsx")
    assert len(recs) == 3
    assert recs[0].ngay_mo == "01/01/2025"
    # 3 ngày trong tiêu đề → lấy ngày đầu, KHÔNG phải ngày CVKT
    assert recs[1].ngay_mo == "17/12/2025"
    assert recs[1].ngay_cvkt == "19/12/2025"


def test_ten_ho_so_khong_co_ngay_thi_ngay_mo_rong():
    recs = S.parse_lookup_excel(_excel_gia_lap(), "t.xlsx")
    assert recs[2].ngay_mo == ""
    assert recs[2].tieu_de == "Hồ sơ không ghi ngày nào cả"


def test_do_dong_tieu_de_theo_noi_dung_khong_theo_so_dong():
    """Chương trình lưu trữ thêm/bớt dòng đầu file thì vẫn phải đọc đủ, không lệch."""
    wb = openpyxl.load_workbook(io.BytesIO(_excel_gia_lap()))
    ws = wb.active
    ws.insert_rows(1, 2)          # đẩy tiêu đề từ dòng 2 xuống dòng 4
    buf = io.BytesIO()
    wb.save(buf)

    recs = S.parse_lookup_excel(buf.getvalue(), "t.xlsx")
    assert len(recs) == 3
    assert recs[0].ngay_mo == "01/01/2025"


def test_thieu_cot_bat_buoc_thi_bao_loi():
    wb = openpyxl.load_workbook(io.BytesIO(_excel_gia_lap()))
    ws = wb.active
    ws.cell(row=2, column=9, value="Cột nào đó")   # mất cột "Mã vạch"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Mã vạch"):
        S.parse_lookup_excel(buf.getvalue(), "t.xlsx")


def test_sai_sheet_thi_bao_loi_ro_rang():
    wb = openpyxl.Workbook()
    wb.active.title = "KhongPhaiData"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Data"):
        S.parse_lookup_excel(buf.getvalue(), "t.xlsx")


# ─── Điền vào template ────────────────────────────────────────────────────────

def test_bia_don_chua_du_bon_thong_tin():
    rec = S.parse_lookup_excel(_excel_gia_lap(), "t.xlsx")[1]
    txt = _text(S.generate_cover_single(rec))

    assert f"Ký hiệu thông tin: {rec.ma_vach}" in txt
    assert f"*{rec.ma_vach}*" in txt          # chuỗi mã vạch font 3 of 9
    assert f"Ngày mở: {rec.ngay_mo}" in txt
    assert rec.tieu_de in txt
    assert f"Gồm: {rec.so_to} tờ" in txt
    assert txt.count(rec.ngay_cvkt) >= 1      # ô dưới nhãn "Ngày công việc kết thúc"
    # Giá trị mẫu trong template phải bị thay hết, không sót
    assert "178074" not in txt


def test_bia_gop_moi_ho_so_mot_trang():
    recs = S.parse_lookup_excel(_excel_gia_lap(), "t.xlsx")
    out = S.generate_covers(recs)
    xml = zipfile.ZipFile(io.BytesIO(out)).read("word/document.xml").decode("utf-8")

    # n hồ sơ → n-1 ngắt trang (trang đầu không cần)
    assert xml.count("<w:pageBreakBefore/>") == len(recs) - 1
    txt = _text(out)
    for rec in recs:
        assert rec.ma_vach in txt
        assert rec.tieu_de in txt


def test_zip_moi_ho_so_mot_file():
    recs = S.parse_lookup_excel(_excel_gia_lap(), "t.xlsx")
    zf = zipfile.ZipFile(io.BytesIO(S.generate_covers_zip(recs)))
    assert sorted(zf.namelist()) == sorted(f"{r.ma_vach}.docx" for r in recs)


def test_khong_co_ho_so_thi_bao_loi():
    with pytest.raises(ValueError):
        S.generate_covers([])


def test_template_doi_cau_truc_thi_bao_loi_thay_vi_sinh_bia_trong(monkeypatch):
    """Mất nhãn mốc → phải nổ TemplateMismatchError, không được im lặng."""
    from docx.oxml.ns import qn

    goc = S._load_template

    def _hong():
        doc = goc()
        for t in doc.element.body.iter(qn("w:t")):
            if t.text and S._LBL_NGAY_BD in t.text:
                t.text = "Đã đổi nhãn"
        return doc

    monkeypatch.setattr(S, "_load_template", _hong)
    rec = S.parse_lookup_excel(_excel_gia_lap(), "t.xlsx")[0]
    with pytest.raises(S.TemplateMismatchError, match="tieu_de"):
        S.generate_cover_single(rec)


# ─── Endpoint ─────────────────────────────────────────────────────────────────

@pytest.fixture
def client_va_db():
    db = sqlite3.connect(":memory:", check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.executescript(_SCHEMA_NHOM)
    db.execute("INSERT INTO user_groups (id, name) VALUES (1, 'Lưu trữ')")
    db.execute("INSERT INTO group_members (group_id, staff_id) VALUES (1, 7)")
    db.commit()

    app.dependency_overrides[get_db] = lambda: db
    yield TestClient(app), db
    app.dependency_overrides.clear()
    db.close()


def _dang_nhap(role="chuyen_vien", staff_id=7):
    app.dependency_overrides[get_current_staff] = lambda: {
        "id": staff_id, "role": role, "username": "u", "full_name": "Người dùng"}


def _cho_quyen(db):
    db.execute("INSERT INTO group_features (group_id, feature_code) VALUES (1, 'menu.storage')")
    db.commit()


def test_khong_co_quyen_menu_storage_thi_403(client_va_db):
    client, _ = client_va_db
    _dang_nhap()
    r = client.post("/api/bundles/archive-cover-parse",
                    files={"file": ("t.xlsx", _excel_gia_lap(), "application/octet-stream")})
    assert r.status_code == 403


def test_parse_roi_in_bia_qua_endpoint(client_va_db):
    client, db = client_va_db
    _cho_quyen(db)
    _dang_nhap()

    r = client.post("/api/bundles/archive-cover-parse",
                    files={"file": ("t.xlsx", _excel_gia_lap(), "application/octet-stream")})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["total"] == 3
    assert data["rows"][1]["ngay_mo"] == "17/12/2025"
    assert any("không tìm thấy ngày" in w for w in data["warnings"])

    r2 = client.post("/api/bundles/archive-cover-print",
                     json={"rows": data["rows"], "as_zip": False})
    assert r2.status_code == 200, r2.text
    assert r2.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml")
    txt = _text(r2.content)
    assert "1000.P026.178068.1" in txt

    r3 = client.post("/api/bundles/archive-cover-print",
                     json={"rows": data["rows"], "as_zip": True})
    assert r3.status_code == 200
    assert r3.headers["content-type"] == "application/zip"
    assert len(zipfile.ZipFile(io.BytesIO(r3.content)).namelist()) == 3


def test_file_sai_dinh_dang_bi_tu_choi(client_va_db):
    client, db = client_va_db
    _cho_quyen(db)
    _dang_nhap()
    r = client.post("/api/bundles/archive-cover-parse",
                    files={"file": ("t.txt", b"khong phai excel", "text/plain")})
    assert r.status_code == 400
    assert "Excel" in r.json()["detail"]


def test_in_bia_khong_chon_dong_nao_thi_400(client_va_db):
    client, db = client_va_db
    _cho_quyen(db)
    _dang_nhap()
    r = client.post("/api/bundles/archive-cover-print", json={"rows": [], "as_zip": False})
    assert r.status_code == 400
