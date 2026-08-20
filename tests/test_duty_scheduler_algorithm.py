"""
Test thuật toán phân lịch trực — ca ngày thường / thứ 6.

Quy tắc nghiệp vụ được kiểm tra:
  1. Mỗi ca bắt buộc 1 Lãnh đạo + 2 người (tổng 3).
  2. Trong ca chỉ ĐÚNG 1 người xử lý song phương (Lãnh đạo hoặc nhân viên).
  3. Ngày thường bốc ngẫu nhiên trong nhóm ít ca nhất; thứ 6 luân phiên tất định.
  4. Người đi dự án / vắng mặt không được phân.
"""
import random
import sqlite3

import pytest

from backend.services import duty_scheduler_engine as eng
from backend.services.duty_scheduler_engine import (
    _generate_ca, _generate_ngay_dac_biet, _save_shift, generate_schedule_for_week,
)

# ── Schema tối thiểu ──────────────────────────────────────────────────────────

_SCHEMA = """
CREATE TABLE departments (
    id INTEGER PRIMARY KEY, code TEXT, name TEXT
);
CREATE TABLE user_tttt (
    id INTEGER PRIMARY KEY, full_name TEXT, role TEXT,
    department_id INTEGER, is_active INTEGER DEFAULT 1, is_deleted INTEGER DEFAULT 0
);
CREATE TABLE duty_staff_meta (
    id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER UNIQUE,
    can_do_sp INTEGER DEFAULT 0, is_sp_backup INTEGER DEFAULT 0,
    is_on_project INTEGER DEFAULT 0, display_order INTEGER DEFAULT 999,
    created_at DATETIME
);
CREATE TABLE duty_absences (
    id INTEGER PRIMARY KEY AUTOINCREMENT, staff_id INTEGER,
    absence_date DATE, created_at DATETIME, UNIQUE(staff_id, absence_date)
);
CREATE TABLE duty_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT, staff_id INTEGER, request_type TEXT,
    specific_date DATE, day_of_week INTEGER, year INTEGER,
    is_active INTEGER DEFAULT 1, created_at DATETIME
);
CREATE TABLE duty_special_days (
    id INTEGER PRIMARY KEY AUTOINCREMENT, date DATE UNIQUE, day_type TEXT,
    label TEXT, is_confirmed INTEGER DEFAULT 0, created_at DATETIME
);
CREATE TABLE duty_rotation_state (
    id INTEGER PRIMARY KEY AUTOINCREMENT, year INTEGER, role TEXT, staff_id INTEGER,
    shift_count INTEGER DEFAULT 0, last_used DATE, position INTEGER DEFAULT 0,
    UNIQUE(year, role, staff_id)
);
CREATE TABLE duty_shifts (
    id INTEGER PRIMARY KEY AUTOINCREMENT, shift_date DATE, shift_type TEXT,
    leader_id INTEGER, leader_ids TEXT DEFAULT '[]',
    sp_id INTEGER, sp_warning TEXT, nv_ids TEXT DEFAULT '[]',
    nv_count INTEGER DEFAULT 0, nv_phu_ids TEXT DEFAULT '[]', nv_phu_count INTEGER DEFAULT 0,
    is_auto INTEGER DEFAULT 1,
    status TEXT DEFAULT 'draft', created_at DATETIME, UNIQUE(shift_date, shift_type)
);
CREATE TABLE duty_shift_config (
    id INTEGER PRIMARY KEY AUTOINCREMENT, year INTEGER UNIQUE,
    ld_count INTEGER DEFAULT 1, nv_count INTEGER DEFAULT 2,
    qt_ld_count INTEGER DEFAULT 1, qt_nv_chinh_count INTEGER DEFAULT 3,
    qt_nv_phu_count INTEGER DEFAULT 2, signer_name TEXT, signer_title TEXT
);
-- Hai bảng dưới KHÔNG thuộc Sổ trực, nhưng lịch trực đọc chúng: ngày lễ chung
-- (nhập ở màn hình Nghỉ phép) và đơn nghỉ phép đã duyệt. Thiếu là mọi test xếp
-- lịch đổ vì "no such table" — xem get_holiday_dates() và get_absences().
CREATE TABLE public_holidays (
    id INTEGER PRIMARY KEY AUTOINCREMENT, date DATE UNIQUE, name TEXT
);
CREATE TABLE leave_records (
    id INTEGER PRIMARY KEY AUTOINCREMENT, staff_id INTEGER,
    start_date DATE, end_date DATE, status TEXT DEFAULT 'pending_ksv'
);
"""

# Ngày mẫu: 2026-08-10 là thứ 2, 2026-08-14 là thứ 6
MONDAY = "2026-08-10"
FRIDAY = "2026-08-14"
YEAR = 2026


def _make_db(staff: list[tuple], ld_count: int = 1, nv_count: int = 2,
             qt_ld: int = 1, qt_chinh: int = 3, qt_phu: int = 2) -> sqlite3.Connection:
    """staff: (id, full_name, role, can_do_sp, is_on_project, display_order)
    Số người mỗi ca lấy từ cấu hình — mặc định 1 Lãnh đạo + 2 nhân viên."""
    db = sqlite3.connect(":memory:")
    db.row_factory = sqlite3.Row
    db.executescript(_SCHEMA)
    db.execute("INSERT INTO departments (id, code, name) VALUES (1,'PAYMENT','Phòng Thanh toán')")
    for sid, name, role, can_sp, on_proj, order in staff:
        db.execute(
            "INSERT INTO user_tttt (id, full_name, role, department_id) VALUES (?,?,?,1)",
            (sid, name, role),
        )
        db.execute(
            "INSERT INTO duty_staff_meta (user_id, can_do_sp, is_on_project, display_order) VALUES (?,?,?,?)",
            (sid, can_sp, on_proj, order),
        )
    db.execute(
        "INSERT INTO duty_shift_config (year, ld_count, nv_count, qt_ld_count, "
        "qt_nv_chinh_count, qt_nv_phu_count) VALUES (?,?,?,?,?,?)",
        (YEAR, ld_count, nv_count, qt_ld, qt_chinh, qt_phu))
    db.commit()
    return db


def _standard_staff() -> list[tuple]:
    """2 LD (1 làm SP) + 6 NV (2 làm SP)."""
    return [
        (1, "LD Một", "truong_phong", 1, 0, 1),   # LD làm được SP
        (2, "LD Hai", "pho_phong",    0, 0, 2),
        (3, "NV Ba",   "chuyen_vien", 1, 0, 3),   # NV làm được SP
        (4, "NV Bốn",  "chuyen_vien", 1, 0, 4),   # NV làm được SP
        (5, "NV Năm",  "chuyen_vien", 0, 0, 5),
        (6, "NV Sáu",  "chuyen_vien", 0, 0, 6),
        (7, "NV Bảy",  "chuyen_vien", 0, 0, 7),
        (8, "NV Tám",  "chuyen_vien", 0, 0, 8),
    ]


def _members(shift: dict) -> list[int]:
    """Toàn bộ id người trong ca: Lãnh đạo + song phương + trực chính + trực phụ."""
    import json
    ids = list(json.loads(shift["leader_ids"] or "[]"))
    if shift["sp_id"]:
        ids.append(shift["sp_id"])
    ids.extend(json.loads(shift["nv_ids"] or "[]"))
    ids.extend(json.loads(shift.get("nv_phu_ids") or "[]"))
    return ids


def _leaders(shift: dict) -> list[int]:
    import json
    return list(json.loads(shift["leader_ids"] or "[]"))


def _sp_capable_ids(db) -> set:
    rows = db.execute("SELECT user_id FROM duty_staff_meta WHERE can_do_sp=1").fetchall()
    return {r["user_id"] for r in rows}


def _gen_one(db, date_str: str, seed: int, shift_type: str = "normal") -> dict:
    """shift=None nghĩa là không lập được ca (vi phạm luật cứng 1 LD + 2 NV)."""
    ld_role, nv_role = ("LD", "NV") if shift_type == "normal" else ("LD_friday", "NV_friday")
    shifts, warns = _generate_ca(
        db, date_str, YEAR, ld_role, nv_role, shift_type, random.Random(seed)
    )
    return {"shift": shifts[0] if shifts else None, "warnings": warns}


# ══════════════════════════════════════════════════════════════
# 1. Cấu trúc ca: đủ 3 người
# ══════════════════════════════════════════════════════════════

def test_ca_ngay_thuong_luon_du_ba_nguoi():
    for seed in range(20):
        db = _make_db(_standard_staff())
        r = _gen_one(db, MONDAY, seed)
        assert len(_members(r["shift"])) == 3, f"seed={seed} không đủ 3 người"
        assert _leaders(r["shift"]), "ca phải có Lãnh đạo"
        assert r["warnings"] == []


def test_ca_thu_sau_luon_du_ba_nguoi():
    db = _make_db(_standard_staff())
    r = _gen_one(db, FRIDAY, 0, shift_type="friday")
    assert len(_members(r["shift"])) == 3
    assert r["shift"]["shift_type"] == "friday"


# ══════════════════════════════════════════════════════════════
# 1b. Số người đi theo khai báo ở tab Cài đặt
# ══════════════════════════════════════════════════════════════

def test_khai_hai_lanh_dao_ba_nhan_vien_thi_sinh_dung_the():
    for seed in range(10):
        db = _make_db(_standard_staff(), ld_count=2, nv_count=3)
        r = _gen_one(db, MONDAY, seed)
        assert len(_leaders(r["shift"])) == 2, f"seed={seed} không đủ 2 Lãnh đạo"
        assert len(_members(r["shift"])) == 5, f"seed={seed} không đủ 5 người"


def test_can_bang_theo_tong_so_ca_khong_theo_tung_loai_ca():
    """Mỗi loại ca đếm một sổ riêng thì người đứng cuối cả ba sổ vẫn có thể có
    tổng số ca cao nhất. Cân bằng phải xét tổng."""
    db = _make_db(_standard_staff())
    # NV Năm đã trực 3 ca thứ 6 và 2 ca cut-off, sổ ngày thường vẫn trắng
    for role, n in (("NV_friday", 3), ("NV_cutoff", 2)):
        db.execute("INSERT INTO duty_rotation_state (year, role, staff_id, shift_count, position) "
                   "VALUES (?,?,5,?,0)", (YEAR, role, n))
    db.commit()

    # Xếp ca ngày thường 20 lần: NV Năm không được ưu tiên dù sổ 'NV' đang là 0
    dem = 0
    for seed in range(20):
        d = _make_db(_standard_staff())
        for role, n in (("NV_friday", 3), ("NV_cutoff", 2)):
            d.execute("INSERT INTO duty_rotation_state (year, role, staff_id, shift_count, position) "
                      "VALUES (?,?,5,?,0)", (YEAR, role, n))
        d.commit()
        if 5 in _members(_gen_one(d, MONDAY, seed)["shift"]):
            dem += 1
    assert dem == 0, f"NV Năm đã trực 5 ca ở loại khác nhưng vẫn được chọn {dem}/20 lần"


def test_tranh_du_song_phuong_khong_duoc_de_len_can_bang():
    """Lãnh đạo đã giữ vai song phương thì thích nhân viên không biết song phương
    hơn — nhưng chỉ khi ngang số ca, không đẩy người ít ca xuống cuối."""
    staff = [
        (1, "LD Một", "truong_phong", 1, 0, 1),   # lãnh đạo biết SP
        (3, "NV Ba",  "chuyen_vien",  1, 0, 3),   # biết SP nhưng chưa trực ca nào
        (5, "NV Năm", "chuyen_vien",  0, 0, 5),
        (6, "NV Sáu", "chuyen_vien",  0, 0, 6),
    ]
    db = _make_db(staff)
    # Hai nhân viên không biết SP đã trực nhiều; NV Ba biết SP nhưng đang 0 ca
    for sid in (5, 6):
        db.execute("INSERT INTO duty_rotation_state (year, role, staff_id, shift_count, position) "
                   "VALUES (?, 'NV', ?, 5, 0)", (YEAR, sid))
    db.commit()
    r = _gen_one(db, MONDAY, 0)
    assert 3 in _members(r["shift"]), (
        "NV Ba đang ít ca nhất mà bị bỏ qua chỉ vì biết song phương")


def test_khong_vo_hai_lanh_dao_cung_biet_song_phuong():
    """Có sẵn Lãnh đạo không biết song phương thì đừng xếp 2 người biết cùng ca."""
    staff = [
        (1, "LD Một", "truong_phong", 1, 0, 1),   # biết SP
        (2, "LD Hai", "pho_phong",    1, 0, 2),   # biết SP
        (9, "LD Ba",  "pho_phong",    0, 0, 9),   # không biết SP
        (5, "NV Năm", "chuyen_vien",  0, 0, 5),
        (6, "NV Sáu", "chuyen_vien",  0, 0, 6),
        (7, "NV Bảy", "chuyen_vien",  0, 0, 7),
    ]
    for seed in range(15):
        db = _make_db(staff, ld_count=2, nv_count=2)
        r = _gen_one(db, MONDAY, seed)
        ld_biet_sp = [i for i in _leaders(r["shift"]) if i in (1, 2)]
        assert len(ld_biet_sp) <= 1, f"seed={seed} vơ 2 lãnh đạo cùng biết song phương"


def test_doi_cau_hinh_thi_ca_sinh_ra_doi_theo():
    db2 = _make_db(_standard_staff(), ld_count=1, nv_count=2)
    db5 = _make_db(_standard_staff(), ld_count=2, nv_count=3)
    assert len(_members(_gen_one(db2, MONDAY, 0)["shift"])) == 3
    assert len(_members(_gen_one(db5, MONDAY, 0)["shift"])) == 5


def test_khai_nhieu_hon_pool_thi_khong_lap_duoc_ca():
    """2 Lãnh đạo trong pool mà khai 3 → luật cứng chặn, không có ca."""
    db = _make_db(_standard_staff(), ld_count=3, nv_count=2)
    r = _gen_one(db, MONDAY, 0)
    assert r["shift"] is None
    assert any(w["type"] == "khong_du_nguoi" for w in r["warnings"])
    assert "3 Lãnh đạo" in r["warnings"][0]["msg"]


def test_khai_bay_nhan_vien_qua_pool_thi_khong_lap_duoc_ca():
    db = _make_db(_standard_staff(), ld_count=1, nv_count=7)   # pool chỉ 6 NV
    r = _gen_one(db, MONDAY, 0)
    assert r["shift"] is None
    assert any(w["type"] == "khong_du_nguoi" for w in r["warnings"])


# ══════════════════════════════════════════════════════════════
# 1c. Ca quyết toán — MỘT ca, nhân viên chia trực chính / trực phụ
# ══════════════════════════════════════════════════════════════

def _gen_quyet_toan(db, date_str=MONDAY, seed=0):
    db.execute("INSERT INTO duty_special_days (date, day_type, is_confirmed) "
               "VALUES (?, 'settlement', 1)", (date_str,))
    db.commit()
    shifts, warns = _generate_ngay_dac_biet(
        db, date_str, YEAR, "settlement_main", "LD", "NV", random.Random(seed))
    return {"shift": shifts[0] if shifts else None, "warnings": warns, "shifts": shifts}


def test_ca_quyet_toan_chi_sinh_mot_ban_ghi():
    """Trước đây quyết toán là 2 dòng (main + sub); nay gộp thành 1 ca."""
    db = _make_db(_standard_staff(), qt_ld=1, qt_chinh=3, qt_phu=2)
    r = _gen_quyet_toan(db)
    assert len(r["shifts"]) == 1, "ca quyết toán phải là MỘT bản ghi"
    assert r["shift"]["shift_type"] == "settlement_main"


def test_ca_quyet_toan_dung_so_truc_chinh_va_truc_phu():
    import json
    db = _make_db(_standard_staff(), qt_ld=1, qt_chinh=3, qt_phu=2)
    s = _gen_quyet_toan(db)["shift"]
    so_chinh = len(json.loads(s["nv_ids"])) + (1 if s["sp_id"] else 0)
    assert len(_leaders(s)) == 1
    assert so_chinh == 3, f"trực chính phải 3 người, đang {so_chinh}"
    assert len(json.loads(s["nv_phu_ids"])) == 2
    assert len(_members(s)) == 6, "tổng 1 lãnh đạo + 3 chính + 2 phụ"


def test_ca_quyet_toan_lanh_dao_dung_chung_khong_phan_chinh_phu():
    db = _make_db(_standard_staff(), qt_ld=2, qt_chinh=2, qt_phu=2)
    s = _gen_quyet_toan(db)["shift"]
    assert len(_leaders(s)) == 2, "lãnh đạo dùng chung cả ca, khai 2 thì phải đủ 2"


def test_ca_quyet_toan_nguoi_song_phuong_nam_o_nhom_truc_chinh():
    import json
    db = _make_db(_standard_staff(), qt_ld=1, qt_chinh=3, qt_phu=2)
    sp_ids = _sp_capable_ids(db)
    s = _gen_quyet_toan(db)["shift"]
    chinh = set(json.loads(s["nv_ids"])) | ({s["sp_id"]} - {None}) | set(_leaders(s))
    assert chinh & sp_ids, "phải có người song phương trong lãnh đạo hoặc trực chính"


def test_ngay_quyet_toan_chua_xac_nhan_thi_khong_sinh():
    db = _make_db(_standard_staff())
    shifts, warns = _generate_ngay_dac_biet(
        db, MONDAY, YEAR, "settlement_main", "LD", "NV", random.Random(0))
    assert shifts == []
    assert any("chưa được xác nhận" in w["msg"] for w in warns)


def test_ca_quyet_toan_thieu_nguoi_thi_khong_lap_duoc():
    staff = _standard_staff()[:4]     # 2 LD + 2 NV
    db = _make_db(staff, qt_ld=1, qt_chinh=3, qt_phu=2)
    r = _gen_quyet_toan(db)
    assert r["shift"] is None
    assert any(w["type"] == "khong_du_nguoi" for w in r["warnings"])


# ══════════════════════════════════════════════════════════════
# 2. Đúng 1 người xử lý song phương
# ══════════════════════════════════════════════════════════════

def test_moi_ca_chi_co_dung_mot_nguoi_song_phuong():
    for seed in range(30):
        db = _make_db(_standard_staff())
        sp_ids = _sp_capable_ids(db)
        r = _gen_one(db, MONDAY, seed)
        in_shift = [i for i in _members(r["shift"]) if i in sp_ids]
        assert len(in_shift) == 1, f"seed={seed} có {len(in_shift)} người SP, cần đúng 1"


def test_thu_sau_cung_giu_dung_mot_nguoi_song_phuong():
    db = _make_db(_standard_staff())
    sp_ids = _sp_capable_ids(db)
    r = _gen_one(db, FRIDAY, 0, shift_type="friday")
    assert len([i for i in _members(r["shift"]) if i in sp_ids]) == 1


def test_lanh_dao_lam_sp_thi_hai_nguoi_con_lai_khong_lam_sp():
    """Chỉ có 1 LD và LD đó làm được SP → LD kiêm SP, 2 NV còn lại phải non-SP."""
    staff = [
        (1, "LD Một", "truong_phong", 1, 0, 1),
        (3, "NV Ba",  "chuyen_vien",  1, 0, 3),
        (5, "NV Năm", "chuyen_vien",  0, 0, 5),
        (6, "NV Sáu", "chuyen_vien",  0, 0, 6),
    ]
    for seed in range(10):
        db = _make_db(staff)
        r = _gen_one(db, MONDAY, seed)
        s = r["shift"]
        assert _leaders(s) == [1]
        assert s["sp_id"] is None, "LD kiêm SP thì không gán SP riêng"
        assert s["sp_warning"] == "leader_sp"
        assert sorted(_members(s)) == [1, 5, 6], "hai người còn lại phải là NV không làm SP"


def test_nhan_vien_lam_sp_khi_lanh_dao_khong_lam_duoc():
    staff = [
        (2, "LD Hai", "pho_phong",   0, 0, 2),
        (3, "NV Ba",  "chuyen_vien", 1, 0, 3),
        (5, "NV Năm", "chuyen_vien", 0, 0, 5),
        (6, "NV Sáu", "chuyen_vien", 0, 0, 6),
    ]
    db = _make_db(staff)
    r = _gen_one(db, MONDAY, 0)
    s = r["shift"]
    assert _leaders(s) == [2]
    assert s["sp_id"] == 3
    assert s["sp_warning"] is None
    assert len(_members(s)) == 3


# ══════════════════════════════════════════════════════════════
# 3. Ngẫu nhiên (ngày thường) vs luân phiên tất định (thứ 6)
# ══════════════════════════════════════════════════════════════

def test_ngay_thuong_bo_c_ngau_nhien_khong_lap_lai_mot_ket_qua():
    ket_qua = set()
    for seed in range(30):
        db = _make_db(_standard_staff())
        ket_qua.add(tuple(sorted(_members(_gen_one(db, MONDAY, seed)["shift"]))))
    assert len(ket_qua) > 1, "ngày thường phải ngẫu nhiên, không ra cùng một tổ hợp"


def test_thu_sau_tat_dinh_khong_phu_thuoc_seed():
    ket_qua = set()
    for seed in range(30):
        db = _make_db(_standard_staff())
        r = _gen_one(db, FRIDAY, seed, shift_type="friday")
        ket_qua.add(tuple(sorted(_members(r["shift"]))))
    assert len(ket_qua) == 1, "thứ 6 luân phiên tất định — seed không được ảnh hưởng"


def test_ngau_nhien_van_can_bang_so_ca():
    """Bốc trong nhóm ít ca nhất → chênh lệch số ca giữa các NV phải nhỏ."""
    db = _make_db(_standard_staff())
    # 4 tuần liên tiếp của tháng 8/2026
    for ws in ("2026-08-03", "2026-08-10", "2026-08-17", "2026-08-24"):
        generate_schedule_for_week(db, ws, seed=42)

    rows = db.execute(
        "SELECT staff_id, shift_count FROM duty_rotation_state WHERE year=? AND role='NV'",
        (YEAR,),
    ).fetchall()
    counts = [r["shift_count"] for r in rows]
    assert counts, "phải có dữ liệu vòng xoay"
    # Ngưỡng 3 cho pool 6 nhân viên: ngoài cân bằng số ca còn hai ràng buộc nữa
    # (đúng 1 người song phương mỗi ca, tránh ê-kíp cố định) nên pool nhỏ không
    # thể khít hơn. Trên bộ nhân sự thật 21 người, lệch nhóm lãnh đạo là 1.
    assert max(counts) - min(counts) <= 3, f"lệch ca quá lớn: {sorted(counts)}"


def test_khong_tao_ra_e_kip_truc_co_dinh():
    """Cân bằng số ca quá chặt sẽ ghép mãi một cặp lãnh đạo–nhân viên với nhau."""
    db = _make_db(_standard_staff())
    for ws in ("2026-08-03", "2026-08-10", "2026-08-17", "2026-08-24"):
        generate_schedule_for_week(db, ws, seed=7)

    import json as _json
    cap: dict = {}
    for r in db.execute("SELECT * FROM duty_shifts"):
        s = dict(r)
        nvs = _json.loads(s["nv_ids"] or "[]") + _json.loads(s["nv_phu_ids"] or "[]")
        if s["sp_id"]:
            nvs.append(s["sp_id"])
        for ld in _json.loads(s["leader_ids"] or "[]"):
            for nv in nvs:
                cap[(ld, nv)] = cap.get((ld, nv), 0) + 1

    so_ca = db.execute("SELECT COUNT(*) FROM duty_shifts").fetchone()[0]
    nhieu_nhat = max(cap.values())
    assert nhieu_nhat < so_ca * 0.75, (
        f"một cặp đi cùng nhau {nhieu_nhat}/{so_ca} ca — thành ê-kíp cố định")


# ══════════════════════════════════════════════════════════════
# 4. Loại người không khả dụng
# ══════════════════════════════════════════════════════════════

def test_nguoi_di_du_an_khong_bao_gio_duoc_phan():
    staff = _standard_staff()
    staff[4] = (5, "NV Năm", "chuyen_vien", 0, 1, 5)   # đi dự án
    for seed in range(20):
        db = _make_db(staff)
        assert 5 not in _members(_gen_one(db, MONDAY, seed)["shift"])


def test_nguoi_vang_mat_khong_duoc_phan_ngay_do():
    for seed in range(20):
        db = _make_db(_standard_staff())
        db.execute(
            "INSERT INTO duty_absences (staff_id, absence_date) VALUES (6, ?)", (MONDAY,)
        )
        db.commit()
        assert 6 not in _members(_gen_one(db, MONDAY, seed)["shift"])


# ══════════════════════════════════════════════════════════════
# 5. Khi không đủ người — ưu tiên đủ 3 người + cảnh báo
# ══════════════════════════════════════════════════════════════

def test_thieu_nguoi_non_sp_van_du_ba_nguoi_va_canh_bao_multi_sp():
    """Pool toàn người làm được SP → không tách được, vẫn phải đủ 3 người."""
    staff = [
        (2, "LD Hai", "pho_phong",   0, 0, 2),
        (3, "NV Ba",  "chuyen_vien", 1, 0, 3),
        (4, "NV Bốn", "chuyen_vien", 1, 0, 4),
    ]
    db = _make_db(staff)
    r = _gen_one(db, MONDAY, 0)
    assert len(_members(r["shift"])) == 3
    assert r["shift"]["sp_warning"] == "multi_sp"
    assert any(w["type"] == "multi_sp" for w in r["warnings"])


def test_khong_ai_lam_sp_thi_canh_bao_no_sp():
    staff = [
        (2, "LD Hai", "pho_phong",   0, 0, 2),
        (5, "NV Năm", "chuyen_vien", 0, 0, 5),
        (6, "NV Sáu", "chuyen_vien", 0, 0, 6),
    ]
    db = _make_db(staff)
    r = _gen_one(db, MONDAY, 0)
    assert len(_members(r["shift"])) == 3
    assert r["shift"]["sp_warning"] == "no_sp_chinh"
    assert any(w["type"] == "no_sp_chinh" for w in r["warnings"])


def test_thieu_nhan_vien_thi_khong_hinh_thanh_ca():
    """Luật cứng: thà không có ca còn hơn có ca 2 người."""
    staff = [
        (2, "LD Hai", "pho_phong",   0, 0, 2),
        (3, "NV Ba",  "chuyen_vien", 1, 0, 3),
    ]
    db = _make_db(staff)
    r = _gen_one(db, MONDAY, 0)
    assert r["shift"] is None, "không đủ 2 nhân viên thì không được lập ca"
    assert any(w["type"] == "khong_du_nguoi" for w in r["warnings"])


def test_khong_co_lanh_dao_thi_khong_hinh_thanh_ca():
    staff = [
        (3, "NV Ba",  "chuyen_vien", 1, 0, 3),
        (5, "NV Năm", "chuyen_vien", 0, 0, 5),
        (6, "NV Sáu", "chuyen_vien", 0, 0, 6),
    ]
    db = _make_db(staff)
    r = _gen_one(db, MONDAY, 0)
    assert r["shift"] is None, "không có Lãnh đạo thì không được lập ca"
    assert any(w["type"] == "khong_du_nguoi" for w in r["warnings"])


def test_thieu_nguoi_thi_khong_ghi_vong_xoay():
    """Ca không hình thành thì không ai được tính số ca."""
    staff = [
        (2, "LD Hai", "pho_phong",   0, 0, 2),
        (3, "NV Ba",  "chuyen_vien", 1, 0, 3),
    ]
    db = _make_db(staff)
    _gen_one(db, MONDAY, 0)
    tong = db.execute(
        "SELECT COALESCE(SUM(shift_count),0) AS t FROM duty_rotation_state WHERE year=?", (YEAR,)
    ).fetchone()["t"]
    assert tong == 0, f"ca không lập được nhưng vẫn ghi {tong} lượt trực"


# ══════════════════════════════════════════════════════════════
# 6. Đăng ký xin trực & không lặp người trong tuần
# ══════════════════════════════════════════════════════════════

def test_dang_ky_dich_danh_ngay_duoc_uu_tien_tuyet_doi():
    """Đăng ký 'once' ép cứng — phải có mặt bất kể vòng xoay/ngẫu nhiên."""
    for seed in range(20):
        db = _make_db(_standard_staff())
        db.execute(
            "INSERT INTO duty_requests (staff_id, request_type, specific_date, year, is_active) "
            "VALUES (8, 'once', ?, ?, 1)", (MONDAY, YEAR)
        )
        db.commit()
        assert 8 in _members(_gen_one(db, MONDAY, seed)["shift"])


def test_ca_trong_tuan_luon_dung_mot_lanh_dao_va_hai_nhan_vien():
    """Cấu trúc cứng: không ca nào được có 2 Lãnh đạo hay 3 nhân viên."""
    for seed in (7, 11, 23, 99):
        db = _make_db(_standard_staff())
        generate_schedule_for_week(db, MONDAY, seed=seed)

        rows = db.execute("SELECT * FROM duty_shifts ORDER BY shift_date").fetchall()
        assert len(rows) == 5, "tuần làm việc phải có 5 ca"

        vai = dict(db.execute(
            "SELECT id, CASE WHEN role IN ('truong_phong','pho_phong') THEN 'LD' ELSE 'NV' END "
            "FROM user_tttt"
        ).fetchall())
        for r in rows:
            ids = _members(dict(r))
            so_ld = sum(1 for i in ids if vai[i] == "LD")
            so_nv = sum(1 for i in ids if vai[i] == "NV")
            assert (so_ld, so_nv) == (1, 2), (
                f"seed={seed} ngày {r['shift_date']}: {so_ld} LD + {so_nv} NV, cần 1 LD + 2 NV"
            )


def test_uu_tien_nguoi_chua_truc_trong_tuan():
    """Pool 8 người / 5 ca — ai cũng phải được dùng ít nhất 1 lần."""
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=7)

    rows = db.execute("SELECT * FROM duty_shifts").fetchall()
    da_truc: set = set()
    for r in rows:
        da_truc.update(_members(dict(r)))
    chua_truc = {s[0] for s in _standard_staff()} - da_truc
    assert not chua_truc, f"còn người chưa được phân ca nào trong tuần: {chua_truc}"


def test_ghi_vong_xoay_dung_so_nguoi_duoc_chon():
    """Ứng viên bị loại trong lúc thử tổ hợp không được cộng số ca."""
    db = _make_db(_standard_staff())
    _gen_one(db, MONDAY, 0)
    tong = db.execute(
        "SELECT COALESCE(SUM(shift_count),0) AS t FROM duty_rotation_state WHERE year=?", (YEAR,)
    ).fetchone()["t"]
    assert tong == 3, f"1 ca = 3 lượt trực, nhưng ghi nhận {tong}"


# ══════════════════════════════════════════════════════════════
# 5. Vòng xoay phải hoàn lại khi ca bị xoá
#
# Sinh ca thì cộng số ca; xoá ca mà không trừ lại là hệ thống vẫn nhớ người đó
# đã trực. Vì cân bằng số ca là tiêu chí CHÍNH, người bị cộng oan sẽ bị đẩy
# xuống cuối hàng ở mọi tuần sau — lịch vẫn "trông hợp lệ" nên không ai thấy.
# ══════════════════════════════════════════════════════════════

def _tong_vong_xoay(db) -> int:
    return db.execute(
        "SELECT COALESCE(SUM(shift_count),0) AS t FROM duty_rotation_state WHERE year=?",
        (YEAR,)
    ).fetchone()["t"]


def _tong_slot_that(db) -> int:
    """Số lượt trực thật đang nằm trong bảng ca — mốc để đối chiếu vòng xoay."""
    import json
    n = 0
    for r in db.execute("SELECT leader_ids, sp_id, nv_ids, nv_phu_ids FROM duty_shifts"):
        n += len(json.loads(r["leader_ids"] or "[]"))
        n += len(json.loads(r["nv_ids"] or "[]"))
        n += len(json.loads(r["nv_phu_ids"] or "[]"))
        if r["sp_id"]:
            n += 1
    return n


def test_sinh_lai_cung_tuan_khong_lam_phinh_so_ca():
    """Bất biến số học: vòng xoay luôn khớp số lượt trực thật, sinh lại bao nhiêu lần cũng vậy."""
    db = _make_db(_standard_staff())
    for lan, seed in enumerate([1, 2, 3], start=1):
        generate_schedule_for_week(db, MONDAY, overwrite_draft=(lan > 1), seed=seed)
        assert _tong_vong_xoay(db) == _tong_slot_that(db), (
            f"sau lần sinh thứ {lan}: vòng xoay {_tong_vong_xoay(db)} "
            f"nhưng chỉ có {_tong_slot_that(db)} lượt trực thật"
        )


def test_xoa_tuan_thi_tra_lai_so_ca_da_cong():
    from backend.services.duty_schedule_service import delete_shifts_for_week
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)
    assert _tong_vong_xoay(db) > 0

    delete_shifts_for_week(db, MONDAY)
    assert db.execute("SELECT COUNT(*) c FROM duty_shifts").fetchone()["c"] == 0
    assert _tong_vong_xoay(db) == 0, "xoá sạch lịch mà hệ thống vẫn nhớ người đã trực"


def test_xoa_mot_ca_chi_tru_dung_nguoi_trong_ca_do():
    from backend.services.duty_schedule_service import delete_shift
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)
    truoc = _tong_vong_xoay(db)

    ca = db.execute("SELECT * FROM duty_shifts ORDER BY shift_date LIMIT 1").fetchone()
    so_nguoi_trong_ca = len(_members(dict(ca)))
    delete_shift(db, ca["id"])

    assert _tong_vong_xoay(db) == truoc - so_nguoi_trong_ca
    assert _tong_vong_xoay(db) == _tong_slot_that(db)


# ══════════════════════════════════════════════════════════════
# 6. Thứ 7 / chủ nhật đi làm — khai bằng "Ngày bù"
# ══════════════════════════════════════════════════════════════

SATURDAY = "2026-08-15"   # thứ 7 của tuần chứa MONDAY
SUNDAY   = "2026-08-16"


def _khai_ngay_bu(db, date_str: str, da_xac_nhan: bool = True) -> None:
    db.execute("INSERT INTO duty_special_days (date, day_type, is_confirmed) "
               "VALUES (?, 'makeup', ?)", (date_str, 1 if da_xac_nhan else 0))
    db.commit()


def test_thu_bay_khai_ngay_bu_thi_sinh_duoc_ca():
    db = _make_db(_standard_staff())
    _khai_ngay_bu(db, SATURDAY)
    generate_schedule_for_week(db, MONDAY, seed=1)

    ngay_co_ca = [r["shift_date"] for r in
                  db.execute("SELECT shift_date FROM duty_shifts ORDER BY shift_date")]
    assert SATURDAY in ngay_co_ca, "khai ngày bù thứ 7 mà không sinh ca"
    assert len(ngay_co_ca) == 6, f"tuần phải có 6 ca, đang có {len(ngay_co_ca)}"


def test_thu_bay_chua_xac_nhan_thi_khong_sinh_ca():
    """Giống cut-off và quyết toán: khai xong còn phải bấm xác nhận."""
    db = _make_db(_standard_staff())
    _khai_ngay_bu(db, SATURDAY, da_xac_nhan=False)
    generate_schedule_for_week(db, MONDAY, seed=1)

    ngay_co_ca = [r["shift_date"] for r in db.execute("SELECT shift_date FROM duty_shifts")]
    assert SATURDAY not in ngay_co_ca
    assert len(ngay_co_ca) == 5


def test_ngay_bu_dung_dung_cau_hinh_ca_thuong():
    """BO chốt: T7/CN không khai số người riêng, dùng chung cấu hình ca thường."""
    import json
    db = _make_db(_standard_staff(), ld_count=2, nv_count=3)
    _khai_ngay_bu(db, SATURDAY)
    generate_schedule_for_week(db, MONDAY, seed=1)

    ca = db.execute("SELECT * FROM duty_shifts WHERE shift_date=?", (SATURDAY,)).fetchone()
    assert ca["shift_type"] == "normal", "ngày bù phải là ca thường, không phải loại ca mới"
    so_ld = len(json.loads(ca["leader_ids"] or "[]"))
    so_nv = len(json.loads(ca["nv_ids"] or "[]")) + (1 if ca["sp_id"] else 0)
    assert (so_ld, so_nv) == (2, 3)


def test_ca_thu_bay_xem_xac_nhan_xoa_duoc_theo_tuan():
    """Chống ca mồ côi: ca T7 phải nằm trong cả 3 thao tác theo tuần."""
    from backend.services.duty_schedule_service import (
        get_shifts_for_week, confirm_shifts_for_week, delete_shifts_for_week,
    )
    db = _make_db(_standard_staff())
    _khai_ngay_bu(db, SATURDAY)
    generate_schedule_for_week(db, MONDAY, seed=1)

    assert SATURDAY in [s["shift_date"] for s in get_shifts_for_week(db, MONDAY)]

    confirm_shifts_for_week(db, MONDAY)
    ca_t7 = db.execute("SELECT status FROM duty_shifts WHERE shift_date=?", (SATURDAY,)).fetchone()
    assert ca_t7["status"] == "confirmed", "xác nhận cả tuần mà bỏ sót ca thứ 7"

    delete_shifts_for_week(db, MONDAY)
    assert db.execute("SELECT COUNT(*) c FROM duty_shifts").fetchone()["c"] == 0


def test_khong_ai_truc_2_lan_trong_tuan_6_ngay():
    """Luật tuần giữ nguyên khi tuần dài ra — BO đã chốt không nới."""
    db = _make_db(_standard_staff())
    _khai_ngay_bu(db, SATURDAY)
    generate_schedule_for_week(db, MONDAY, seed=3)

    dem: dict = {}
    for r in db.execute("SELECT * FROM duty_shifts"):
        for sid in _members(dict(r)):
            dem[sid] = dem.get(sid, 0) + 1
    # 8 người / 6 ca × 3 chỗ = 18 lượt → không tránh được lặp, nhưng phải chia đều
    assert max(dem.values()) - min(dem.values()) <= 1, f"chia ca lệch quá 1: {dem}"


def test_chu_nhat_cung_khai_ngay_bu_duoc():
    db = _make_db(_standard_staff())
    _khai_ngay_bu(db, SATURDAY)
    _khai_ngay_bu(db, SUNDAY)
    generate_schedule_for_week(db, MONDAY, seed=1)

    ngay_co_ca = [r["shift_date"] for r in
                  db.execute("SELECT shift_date FROM duty_shifts ORDER BY shift_date")]
    assert ngay_co_ca[-2:] == [SATURDAY, SUNDAY]
    assert len(ngay_co_ca) == 7, "tuần khai cả T7 lẫn CN phải có 7 ca"


def test_file_excel_chi_them_dong_t7_khi_hom_do_co_ca():
    """Tuần thường phải giữ nguyên hình dạng cũ: 5 dòng, đề đến thứ 6."""
    import io
    from datetime import date as _d
    from openpyxl import load_workbook
    from backend.services.duty_schedule_service import get_shifts_for_week
    from backend.services.duty_export_service import build_week_excel
    from backend.services.duty_calendar_utils import week_span

    def _dung_file(db):
        t2, cn = week_span(MONDAY)
        data = build_week_excel(get_shifts_for_week(db, MONDAY),
                                _d.fromisoformat(t2), _d.fromisoformat(cn))
        ws = load_workbook(io.BytesIO(data)).active
        cot_thu = [ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)]
        return ws["A1"].value, [t for t in cot_thu if t and t != "Ghi chú :"]

    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)
    tieu_de, cot = _dung_file(db)
    assert "ĐẾN NGÀY 14/08/2026" in tieu_de, "tuần thường không được đề tới cuối tuần"
    assert not any(t.startswith("T7") or t.startswith("CN") for t in cot)

    db2 = _make_db(_standard_staff())
    _khai_ngay_bu(db2, SATURDAY)
    generate_schedule_for_week(db2, MONDAY, seed=1)
    tieu_de2, cot2 = _dung_file(db2)
    assert "ĐẾN NGÀY 15/08/2026" in tieu_de2, "tuần có ngày bù phải đề tới thứ 7"
    assert any(t.startswith("T7") for t in cot2), "file thiếu dòng thứ 7"


def test_cutoff_tinh_ca_ngay_lam_bu_cuoi_thang():
    """T7 làm bù cũng là ngày làm việc — bỏ qua thì cut-off bị đẩy lùi lên nhầm ngày."""
    from backend.services.duty_calendar_utils import compute_cutoff_dates

    # 31/10/2026 là thứ 7, 30/10 là thứ 6, 29/10 thứ 5
    khong_bu = compute_cutoff_dates(10, 2026, set())
    assert khong_bu == ["2026-10-29", "2026-10-30"]

    co_bu = compute_cutoff_dates(10, 2026, set(), {"2026-10-31"})
    assert co_bu == ["2026-10-30", "2026-10-31"], "cut-off phải tính cả ngày làm bù"


# ══════════════════════════════════════════════════════════════
# 7. File Excel bám mẫu thật "Lịch trực PTT.xlsx"
#
# Mẫu là 5 cột A–E, trắng đen, và LUÔN có chức danh người ký — kể cả tuần
# thường. Bốn test dưới canh đúng những chỗ dễ trôi lại về bản cũ.
# ══════════════════════════════════════════════════════════════

def _mo_file(db, signer_title=None, holiday_map=None):
    """Dựng file cho tuần chứa MONDAY rồi mở lại bằng openpyxl."""
    import io
    from datetime import date as _d
    from openpyxl import load_workbook
    from backend.services.duty_schedule_service import get_shifts_for_week
    from backend.services.duty_export_service import build_week_excel
    from backend.services.duty_calendar_utils import week_span

    t2, cn = week_span(MONDAY)
    kw = {}
    if signer_title is not None:
        kw["signer_title"] = signer_title
    data = build_week_excel(get_shifts_for_week(db, MONDAY),
                            _d.fromisoformat(t2), _d.fromisoformat(cn),
                            holiday_map=holiday_map, **kw)
    return load_workbook(io.BytesIO(data)).active


def _chu_trong_o(cell) -> str:
    """Ô có thể là chuỗi thường hoặc rich text (ngày quyết toán)."""
    return "" if cell.value is None else str(cell.value)


def test_file_excel_dung_5_cot_va_khong_mau():
    """
    Mẫu chỉ có A–E và không tô nền ô nào.

    Đổi _NCOLS thôi KHÔNG đủ: _apply_row ghi đúng len(values), nên list 8 phần tử
    sót lại vẫn tạo ô F/G/H có viền → max_column = 8. Test này canh đúng chỗ đó.
    """
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)
    ws = _mo_file(db)

    assert ws.max_column == 5, f"mẫu có 5 cột A–E, file đang có {ws.max_column}"

    co_mau = [
        f"{cell.coordinate}"
        for row in ws.iter_rows()
        for cell in row
        if cell.fill is not None and cell.fill.patternType
    ]
    assert not co_mau, f"mẫu là trắng đen, các ô sau bị tô nền: {co_mau[:8]}"


def test_file_excel_luon_co_chuc_danh_nguoi_ky():
    """Trước đây "GIÁM ĐỐC" chỉ hiện ở tuần quyết toán — mẫu thì tuần nào cũng có."""
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)          # tuần THƯỜNG
    ws = _mo_file(db)

    chu = [_chu_trong_o(ws.cell(r, 5)) for r in range(1, ws.max_row + 1)]
    assert "GIÁM ĐỐC" in chu, "tuần thường cũng phải có chức danh người ký"

    # Chức danh phải nằm TRÊN tên người ký, không phải ngược lại
    assert chu.index("GIÁM ĐỐC") < chu.index("Nguyễn Quốc Hùng")


def test_chuc_danh_nguoi_ky_khai_duoc():
    """Đổi chức danh ở cấu hình thì file phải đổi theo, không còn literal trong code."""
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)

    ws = _mo_file(db, signer_title="PHÓ GIÁM ĐỐC")
    chu = [_chu_trong_o(ws.cell(r, 5)) for r in range(1, ws.max_row + 1)]
    assert "PHÓ GIÁM ĐỐC" in chu
    assert "GIÁM ĐỐC" not in chu, "chức danh cũ không được sót lại"


def test_ngay_quyet_toan_gop_mot_hang():
    """
    BO chốt: ngày quyết toán chỉ chiếm MỘT hàng, ô nhân viên chứa cả trực chính
    lẫn trực phụ (chính IN HOA, phụ nghiêng nhỏ). Bản cũ dựng 2 hàng + merge dọc.
    """
    db = _make_db(_standard_staff(), qt_ld=1, qt_chinh=2, qt_phu=2)
    db.execute("INSERT INTO duty_special_days (date, day_type, is_confirmed) "
               "VALUES (?, 'settlement', 1)", (MONDAY,))
    db.commit()
    generate_schedule_for_week(db, MONDAY, seed=1)

    ca = db.execute("SELECT * FROM duty_shifts WHERE shift_date=?", (MONDAY,)).fetchone()
    assert ca["shift_type"] == "settlement_main"
    assert ca["nv_phu_count"] == 2, "cần có nhóm trực phụ thì test mới có nghĩa"

    ws = _mo_file(db)

    # Đúng một hàng mang nhãn ngày quyết toán
    hang_qt = [r for r in range(1, ws.max_row + 1)
               if (ws.cell(r, 1).value or "") and "(QT)" in str(ws.cell(r, 1).value)]
    assert len(hang_qt) == 1, f"ngày quyết toán phải gọn 1 hàng, đang có {len(hang_qt)}"

    # Không còn merge dọc cột A/B (dấu vết của bố cục 2 hàng)
    merge_doc = [str(m) for m in ws.merged_cells.ranges
                 if m.min_col == m.max_col and m.min_row != m.max_row]
    assert not merge_doc, f"không được merge dọc nữa: {merge_doc}"

    # Ô nhân viên chứa CẢ trực chính lẫn trực phụ
    r = hang_qt[0]
    o_nv = _chu_trong_o(ws.cell(r, 3))
    import json as _json
    ten = {p["id"]: p["full_name"] for p in
           [dict(x) for x in db.execute("SELECT id, full_name FROM user_tttt")]}
    chinh = ([ten[ca["sp_id"]]] if ca["sp_id"] else []) + \
            [ten[i] for i in _json.loads(ca["nv_ids"] or "[]")]
    phu = [ten[i] for i in _json.loads(ca["nv_phu_ids"] or "[]")]

    for t in chinh:
        assert t.upper() in o_nv, f"thiếu trực chính {t} (phải IN HOA)"
    for t in phu:
        assert t in o_nv, f"thiếu trực phụ {t}"


def test_moi_o_trong_bang_deu_du_bon_canh_vien():
    """
    Ô nằm trong vùng gộp cũng phải được kẻ đủ 4 cạnh.

    Lỗi đã gặp: code gộp ô TRƯỚC rồi mới kẻ, mà ô đã gộp thì bị bỏ qua — nửa phải
    của ô tiêu đề "NHÂN VIÊN" (C3:D3) không có cạnh nào nên Excel vẽ ra ô hở.
    Test quét cả vùng bảng để lỗi không quay lại khi ai đó thêm vùng gộp mới.
    """
    db = _make_db(_standard_staff(), qt_ld=1, qt_chinh=2, qt_phu=2)
    # Có đủ 3 kiểu hàng: thường, nghỉ lễ (gộp C:E) và quyết toán (gộp C:D)
    db.execute("INSERT INTO duty_special_days (date, day_type, is_confirmed) "
               "VALUES ('2026-08-11', 'settlement', 1)")
    db.commit()
    generate_schedule_for_week(db, MONDAY, seed=1)
    ws = _mo_file(db, holiday_map={"2026-08-12": "Lễ thử"})

    # Vùng bảng: từ hàng tiêu đề cột tới hàng ngay trước "Ghi chú :"
    hang_ghi_chu = next(r for r in range(1, ws.max_row + 1)
                        if str(ws.cell(r, 1).value or "").startswith("Ghi chú"))

    # openpyxl cố ý xoá cạnh BÊN TRONG vùng gộp — Excel không vẽ vạch giữa một ô
    # đã gộp. Nên chỉ đòi những cạnh thật sự nằm ở biên ngoài của ô người ta nhìn thấy.
    gop = [m for m in ws.merged_cells.ranges if 3 <= m.min_row < hang_ghi_chu]

    def canh_can_co(r: int, c: int) -> set:
        for m in gop:
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                return {ten for ten, o_bien in (("left", c == m.min_col),
                                                ("right", c == m.max_col),
                                                ("top", r == m.min_row),
                                                ("bottom", r == m.max_row)) if o_bien}
        return {"left", "right", "top", "bottom"}

    thieu = []
    for r in range(3, hang_ghi_chu):
        for c in range(1, 6):
            b = ws.cell(r, c).border
            co = {k for k in ("left", "right", "top", "bottom") if getattr(b, k).style}
            can = canh_can_co(r, c)
            if not can <= co:
                thieu.append(f"{ws.cell(r, c).coordinate} thiếu {','.join(sorted(can - co))}")
    assert not thieu, "ô hở cạnh viền: " + " · ".join(thieu)


def test_bang_du_rong_va_can_giua_tren_a4_ngang():
    """
    Bảng phải lấp gần kín bề ngang A4 xoay ngang và nằm giữa trang.

    Trước đây rộng 22,2cm trong khi vùng in là 27,2cm — thừa 5cm dồn hết về một
    bên, in ra nhìn lệch hẳn.
    """
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)
    ws = _mo_file(db)

    tong = sum(ws.column_dimensions[c].width or 0 for c in "ABCDE")
    rong_cm = ((tong * 7 + 5) / 96) * 2.54          # đơn vị width → px → inch → cm
    vung_in_cm = 29.7 - 2 * 1.27                     # A4 ngang trừ lề 0,5 inch mỗi bên

    assert 25.0 <= rong_cm <= vung_in_cm, (
        f"bảng rộng {rong_cm:.1f}cm, cần nằm trong 25,0–{vung_in_cm:.1f}cm")
    assert ws.print_options.horizontalCentered, "phải bật căn giữa ngang khi in"

    # Tỉ lệ giữa các cột giữ nguyên như mẫu, chỉ phóng to đều
    w = {c: ws.column_dimensions[c].width for c in "ABCDE"}
    assert abs(w["C"] - w["E"]) < 0.5, "cột Nhân viên 1 và Lãnh đạo vốn bằng nhau"
    assert w["C"] > w["D"] > w["A"] > w["B"], "thứ tự độ rộng cột phải như mẫu"


def test_chieu_cao_hang_du_cho_co_chu_khong_de_len_vach():
    """
    Hàng phải cao hơn cỡ chữ × 1,33 (khoảng cách dòng của Excel).

    Hàng thấp hơn thì chữ bị ép sát và tràn đè lên vạch kẻ — nhìn ra thành "ô
    thiếu kẻ phía trên", đúng lỗi đã gặp ở ô tiêu đề "NHÂN VIÊN".
    """
    db = _make_db(_standard_staff())
    generate_schedule_for_week(db, MONDAY, seed=1)
    ws = _mo_file(db)

    # (hàng, cỡ chữ dùng ở hàng đó)
    for r, co_chu in ((1, 24), (3, 18)):
        cao = ws.row_dimensions[r].height
        can = co_chu * 1.33
        assert cao is not None and cao >= can, (
            f"hàng {r} cao {cao}pt nhưng chữ {co_chu}pt cần ít nhất {can:.1f}pt")


def test_ngay_khai_nghi_le_thi_file_ghi_nhan_nghi_le():
    """
    Ngày ĐÃ KHAI nghỉ lễ thì file xuất ra phải ghi nhãn "Nghỉ lễ".

    Ngày không có ca vì lý do khác (chưa xếp, thiếu người) thì để trống — không
    được ghi "Nghỉ lễ" cho nó, vì hôm đó không phải ngày được nghỉ.
    """
    db = _make_db(_standard_staff())
    # Thứ 3 hổng vì hết Lãnh đạo; thứ 4 nghỉ lễ (phải khai vào DB thì engine mới
    # bỏ qua ngày đó, chỉ truyền holiday_map cho hàm dựng file là chưa đủ)
    for sid in (1, 2):
        db.execute("INSERT INTO duty_absences (staff_id, absence_date) VALUES (?, '2026-08-11')",
                   (sid,))
    db.execute("INSERT INTO duty_special_days (date, day_type, label, is_confirmed) "
               "VALUES ('2026-08-12', 'holiday', 'Lễ thử', 1)")
    db.commit()
    generate_schedule_for_week(db, MONDAY, seed=1)
    ws = _mo_file(db, holiday_map={"2026-08-12": "Lễ thử"})

    hang_ghi_chu = next(r for r in range(1, ws.max_row + 1)
                        if str(ws.cell(r, 1).value or "").startswith("Ghi chú"))

    chu = {str(ws.cell(r, 2).value): str(ws.cell(r, 3).value or "")
           for r in range(4, hang_ghi_chu) if ws.cell(r, 2).value}

    # Ngày khai nghỉ lễ: có nhãn, kèm tên ngày lễ đã khai
    assert chu.get("12/08/2026", "") == "(Nghỉ lễ: Lễ thử)"
    # Ngày hổng lịch vì thiếu người: KHÔNG được gán nhãn nghỉ lễ
    assert "Nghỉ lễ" not in chu.get("11/08/2026", "")


def test_ngay_le_khong_khai_ghi_chu_van_co_chu_nghi_le():
    """
    Ngày lễ khai mà BỎ TRỐNG ô Ghi chú vẫn phải ghi "Nghỉ lễ".

    Lỗi thật đã gặp: cột `label` trong DB là NULL khi người dùng không nhập ghi
    chú. Tầng API dựng holiday_map bằng `h.get("label", "")` — nhưng khoá "label"
    LUÔN tồn tại nên .get() trả về None chứ không phải "", và hàm dựng file lại
    suy "có phải ngày lễ không" từ chính giá trị đó → ô ra trắng.
    """
    db = _make_db(_standard_staff())
    db.execute("INSERT INTO duty_special_days (date, day_type, label, is_confirmed) "
               "VALUES ('2026-08-12', 'holiday', NULL, 1)")
    db.commit()
    generate_schedule_for_week(db, MONDAY, seed=1)

    # Dựng holiday_map y hệt tầng API để bắt được lỗi ở đúng chỗ nó xảy ra
    from backend.services.duty_constraint_service import list_special_days
    hmap = {h["date"]: (h.get("label") or "") for h in
            list_special_days(db, day_type="holiday", year=YEAR)}
    assert hmap["2026-08-12"] == "", "nhãn None phải quy về chuỗi rỗng"

    ws = _mo_file(db, holiday_map=hmap)
    o = next(str(ws.cell(r, 3).value or "") for r in range(4, ws.max_row + 1)
             if str(ws.cell(r, 2).value or "") == "12/08/2026")
    assert o == "(Nghỉ lễ)", f"ngày lễ không ghi chú phải ra '(Nghỉ lễ)', đang là {o!r}"


# ══════════════════════════════════════════════════════════════
# 9. Hai luật mềm bổ sung: không trực thứ 6 2 lần/tháng,
#    Lãnh đạo không trực quá 2 ca/tuần
# ══════════════════════════════════════════════════════════════

# 4 thứ 6 của tháng 8/2026 (14/08 là thứ 6 — xem hằng số FRIDAY ở đầu file)
FRIDAYS_08_2026 = ["2026-08-07", "2026-08-14", "2026-08-21", "2026-08-28"]


def _gen_friday_channel(db, date_str: str, seed: int) -> dict:
    """Sinh + LƯU ca thứ 6 — luật tránh lặp trong tháng đọc lại từ duty_shifts,
    nên phải ghi xuống DB mới thấy hiệu lực ở thứ 6 kế tiếp (khác _gen_one, vốn
    chỉ dùng cho test 1 ca đơn lẻ không cần lịch sử)."""
    shifts, warns = _generate_ca(
        db, date_str, YEAR, "LD_friday", "NV_friday", "friday", random.Random(seed)
    )
    if shifts:
        _save_shift(db, shifts[0])
        db.commit()
    return {"shift": shifts[0] if shifts else None, "warnings": warns}


def test_khong_trung_thu6_trong_thang_khi_du_nguoi():
    """Pool đủ rộng — ai cũng phải được ưu tiên trước người đã trực thứ 6 rồi,
    nên trong 4 thứ 6 của tháng không ai bị lặp."""
    staff = [(i, f"LD {i}", "truong_phong", 0, 0, i) for i in range(1, 5)]
    staff += [(10 + i, f"NV {i}", "chuyen_vien", 1 if i % 2 == 0 else 0, 0, 10 + i)
              for i in range(1, 13)]
    db = _make_db(staff)

    da_dung: dict = {}
    for ds in FRIDAYS_08_2026:
        r = _gen_friday_channel(db, ds, seed=1)
        assert r["shift"] is not None
        for sid in _members(r["shift"]):
            da_dung[sid] = da_dung.get(sid, 0) + 1

    trung = {sid: n for sid, n in da_dung.items() if n > 1}
    assert not trung, f"đủ người mà vẫn có người trực thứ 6 2 lần/tháng: {trung}"


def test_buoc_trung_thu6_khi_thieu_nguoi_van_len_canh_bao():
    """Pool quá nhỏ so với số thứ 6 trong tháng — không tránh được thì vẫn phải
    lập ca đủ người, kèm đúng loại cảnh báo để người phân lịch biết mà xử lý."""
    staff = [
        (1, "LD Một", "truong_phong", 1, 0, 1),
        (2, "LD Hai", "pho_phong",    0, 0, 2),
        (3, "NV Ba",   "chuyen_vien", 1, 0, 3),
        (4, "NV Bốn",  "chuyen_vien", 0, 0, 4),
        (5, "NV Năm",  "chuyen_vien", 0, 0, 5),
    ]
    db = _make_db(staff)

    co_canh_bao_trung = False
    for ds in FRIDAYS_08_2026:
        r = _gen_friday_channel(db, ds, seed=1)
        assert r["shift"] is not None, f"pool nhỏ vẫn phải đủ người ngày {ds}"
        assert len(_members(r["shift"])) == 3
        if any(w["type"] == "trung_thu6_thang" for w in r["warnings"]):
            co_canh_bao_trung = True
    assert co_canh_bao_trung, (
        "pool chỉ 2 LD/3 NV cho 4 thứ 6 trong tháng phải buộc lặp lại ít nhất 1 lần")


def test_lanh_dao_khong_qua_2_ca_tuan_khi_du_nguoi():
    """3 Lãnh đạo cho 1 tuần 5 ngày — đủ người để không ai vượt tối đa 2 ca/tuần."""
    staff = [
        (1, "LD Một", "truong_phong", 0, 0, 1),
        (2, "LD Hai", "pho_phong",    0, 0, 2),
        (3, "LD Ba",  "pho_phong",    0, 0, 3),
    ] + [(10 + i, f"NV {i}", "chuyen_vien", 1 if i <= 2 else 0, 0, 10 + i) for i in range(1, 5)]
    db = _make_db(staff)
    generate_schedule_for_week(db, MONDAY, seed=3)

    dem: dict = {}
    for r in db.execute("SELECT leader_ids FROM duty_shifts"):
        for sid in _leaders(dict(r)):
            dem[sid] = dem.get(sid, 0) + 1
    assert all(n <= 2 for n in dem.values()), f"có Lãnh đạo vượt 2 ca/tuần: {dem}"


def test_lanh_dao_qua_2_ca_tuan_khi_thieu_nguoi_van_len_canh_bao():
    """Chỉ 2 Lãnh đạo cho tuần 5 ngày — về mặt toán học không thể tránh được (5
    ngày ÷ 2 người), phải lập đủ ca và lên đúng cảnh báo thay vì âm thầm vượt."""
    staff = [
        (1, "LD Một", "truong_phong", 0, 0, 1),
        (2, "LD Hai", "pho_phong",    0, 0, 2),
    ] + [(10 + i, f"NV {i}", "chuyen_vien", 1 if i <= 2 else 0, 0, 10 + i) for i in range(1, 5)]
    db = _make_db(staff)
    result = generate_schedule_for_week(db, MONDAY, seed=3)

    assert result["created"] == 5, "đủ 2 LD + 4 NV cho cả 5 ngày, không ngày nào bị bỏ"
    assert any(w["type"] == "ld_qua_tai_tuan" for w in result["warnings"]), (
        "chỉ 2 Lãnh đạo cho 5 ngày phải buộc ai đó vượt 2 ca/tuần, kèm cảnh báo")
