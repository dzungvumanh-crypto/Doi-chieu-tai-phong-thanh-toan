"""
Synthetic tests cho pipeline Chấm ILO1000.
Bao phủ tất cả các lỗi đã tìm và sửa trong audit round.

Chạy: python -m pytest tests/test_ilo1000_algorithm.py -v
"""

import pandas as pd
import pyzipper
import pytest

from backend.services.ilo1000.process import (
    _first_match,
    _safe_str,
    detect_huy,
    process_hub,
    process_citad,
    process_core,
)
from backend.services.ilo1000.config import (
    HUB_COL_SO_GD, HUB_COL_STC, HUB_COL_TRACE,
    HUB_COL_TRANG_THAI, HUB_COL_NGAY_GIO, HUB_COL_NOI_DUNG, HUB_COL_SO_TIEN,
)


# ── Helper ────────────────────────────────────────────────────────────────────

def _hub_row(so_gd, stc, trace, trang_thai, ngay_gio, noi_dung='', so_tien='1000000'):
    """Tạo 1 dòng hub với đầy đủ cột chuẩn."""
    return {
        HUB_COL_SO_GD:      so_gd,
        HUB_COL_STC:        stc,
        HUB_COL_TRACE:      trace,
        HUB_COL_TRANG_THAI: trang_thai,
        HUB_COL_NGAY_GIO:   ngay_gio,
        HUB_COL_NOI_DUNG:   noi_dung,
        HUB_COL_SO_TIEN:    so_tien,
        'Số Ref Hub':       'REF' + so_gd,
    }


def _core_row(ref, brcd, cramount, dramount='0'):
    return {
        'REFERENCE': ref, 'TRBRCD': brcd, 'CRAMOUNT': str(cramount),
        'DRAMOUNT': str(dramount), 'TRDATE': '20260512', 'USERID': 'U1',
        'JOURSEQ': '1', 'DYTRSEQ': '1', 'LOCAC': '', 'CCY': 'VND',
        'BUSCD': '', 'UNIT': '', 'TRCD': '', 'CUSTOMER': '', 'TRTP': '',
        'REMARK': '', 'CRTDTM': '',
    }


# ── Test 1: _first_match — giữ lần đầu tiên, bỏ qua lần sau ─────────────────

class TestFirstMatch:
    def test_basic_first_wins(self):
        keys   = pd.Series(['A', 'B', 'A', 'C'])
        values = pd.Series(['v1', 'v2', 'v_LAST', 'v3'])
        result = _first_match(keys, values)
        assert result['A'] == 'v1', "Phải giữ giá trị ĐẦU TIÊN, không phải cuối"
        assert result['B'] == 'v2'
        assert result['C'] == 'v3'

    def test_empty_key_excluded(self):
        keys   = pd.Series(['', 'A', ''])
        values = pd.Series(['x', 'y', 'z'])
        result = _first_match(keys, values)
        assert '' not in result, "Key rỗng không được đưa vào dict"
        assert result.get('A') == 'y'

    def test_all_unique(self):
        keys   = pd.Series(['X', 'Y', 'Z'])
        values = pd.Series(['1', '2', '3'])
        result = _first_match(keys, values)
        assert result == {'X': '1', 'Y': '2', 'Z': '3'}

    def test_empty_series(self):
        result = _first_match(pd.Series([], dtype=str), pd.Series([], dtype=str))
        assert result == {}


# ── Test 2: Hub Ngày parsing — ngày đơn digit ────────────────────────────────

class TestHubNgay:
    """P0 fix: str[:2] fails cho '5/05/2026 09:00' → parse đầy đủ với to_datetime."""

    def _make_hub(self, ngay_gio_val):
        return pd.DataFrame([_hub_row('S001', 'STC001', 'TRC001', 'Thành công', ngay_gio_val)])

    def test_single_digit_day(self):
        df = self._make_hub('5/05/2026 09:00')
        hub_out, _ = process_hub(df, {}, 20260512)
        ngay = hub_out['Ngày'].iloc[0]
        assert ngay == 5.0, f"Ngày phải là 5, nhận được {ngay}"

    def test_double_digit_day(self):
        df = self._make_hub('12/05/2026 09:00')
        hub_out, _ = process_hub(df, {}, 20260512)
        ngay = hub_out['Ngày'].iloc[0]
        assert ngay == 12.0, f"Ngày phải là 12, nhận được {ngay}"

    def test_day_31(self):
        df = self._make_hub('31/05/2026 23:59')
        hub_out, _ = process_hub(df, {}, 20260512)
        ngay = hub_out['Ngày'].iloc[0]
        assert ngay == 31.0

    def test_nan_when_empty(self):
        df = self._make_hub('')
        hub_out, _ = process_hub(df, {}, 20260512)
        import math
        ngay = hub_out['Ngày'].iloc[0]
        assert ngay != ngay or math.isnan(ngay), "Ngày rỗng → NaN"

    def test_cho_di_kenh_flag(self):
        """Ngày > ngày_dc → 'Chờ đi kênh'. Đây là case hay bị miss khi str[:2] fail."""
        # Ngày đối chiếu = 12, transaction ngày 14 → phải flag Chờ đi kênh
        df = self._make_hub('14/05/2026 08:00')
        hub_out, _ = process_hub(df, {}, 20260512)
        assert hub_out[HUB_COL_TRANG_THAI].iloc[0] == 'Chờ đi kênh'

    def test_cho_di_kenh_flag_ngay_dc_plus_1(self):
        """Ngày = ngày_dc + 1 (giao dịch xử lý ngày hôm sau) cũng phải bị flag — golden sample xác nhận."""
        df = self._make_hub('13/05/2026 08:00')
        hub_out, _ = process_hub(df, {}, 20260512)
        assert hub_out[HUB_COL_TRANG_THAI].iloc[0] == 'Chờ đi kênh'

    def test_cho_di_kenh_NOT_flagged_for_same_day(self):
        """Ngày = ngày_dc → không phải Chờ đi kênh."""
        df = self._make_hub('12/05/2026 08:00')
        hub_out, _ = process_hub(df, {}, 20260512)
        # Trạng thái gốc phải còn nguyên là 'Thành công'
        assert hub_out[HUB_COL_TRANG_THAI].iloc[0] == 'Thành công'


class TestHubTraceLeadingZero:
    """
    Dữ liệu thật 06/07/2026: pHub xuất 'Số Trace 1' có số 0 thừa ở đầu tùy đợt
    export (VD '063403118') trong khi Core tính Trace độc lập từ REFERENCE
    không có số 0 đầu ('63403118') — cùng 1 giao dịch nhưng Map dc (ghép chuỗi)
    không khớp, khiến TT bị bỏ trống dù giao dịch đã khớp thật. Fix: strip số 0
    thừa cho Trace thuần chữ số.
    """

    def test_leading_zero_stripped_for_numeric_trace(self):
        df = pd.DataFrame([_hub_row('S001', 'STC001', '063403118', 'Hoàn thành', '06/07/2026 09:47')])
        hub_out, _ = process_hub(df, {}, 20260706)
        assert hub_out['Trace'].iloc[0] == '63403118'

    def test_all_zero_trace_kept_as_single_zero(self):
        df = pd.DataFrame([_hub_row('S002', 'STC002', '0000', 'Hoàn thành', '06/07/2026 09:47')])
        hub_out, _ = process_hub(df, {}, 20260706)
        assert hub_out['Trace'].iloc[0] == '0'

    def test_no_leading_zero_unchanged(self):
        df = pd.DataFrame([_hub_row('S003', 'STC003', '63403118', 'Hoàn thành', '06/07/2026 09:47')])
        hub_out, _ = process_hub(df, {}, 20260706)
        assert hub_out['Trace'].iloc[0] == '63403118'

    def test_bfx_trace_not_stripped_even_if_looks_like_leading_zero(self):
        """Trace BFX = right(nội dung,16) có thể bắt đầu bằng '0' có ý nghĩa (không phải số) — giữ nguyên."""
        df = pd.DataFrame([_hub_row('S004', 'STC004', 'ignored', 'Hoàn thành', '06/07/2026 09:47',
                                     noi_dung='BFX0123456789012345')])
        hub_out, _ = process_hub(df, {}, 20260706)
        # BFX trace = right(nội dung, 16) — không đi qua bước strip số 0 (chứa chữ 'BFX' phía trước bị cắt,
        # phần còn lại toàn số '0123456789012345' — vẫn thuần số nên vẫn bị strip theo đúng logic áp dụng
        # đồng nhất cho MỌI trace thuần chữ số, kể cả nguồn gốc từ BFX)
        assert hub_out['Trace'].iloc[0] == '123456789012345'

    def test_map_dc_now_matches_core_after_strip(self):
        """Regression: Hub Trace có số 0 đầu → Citad Map dc giờ khớp đúng với Core (trace không số 0 đầu)."""
        hub_df = pd.DataFrame([_hub_row('SA010', 'STC010', '063403118', 'Hoàn thành', '06/07/2026 09:47')])
        hub_out, hub_lookups = process_hub(hub_df, {}, 20260706)

        citad_df = pd.DataFrame([{
            'SERIAL_NO': 'STC010', 'RELATION_NO': '17813327',
            'TRX_DATE': '20260706', 'AMOUNT': '1000000', 'TRX_STATUS': 'OK',
        }])
        _, citad_mapdc = process_citad(citad_df, hub_lookups, 20260706)

        core_df = pd.DataFrame([{
            'REFERENCE': '1000API63403118', 'TRBRCD': '1781',
            'CRAMOUNT': '1000000', 'DRAMOUNT': '0',
            'TRDATE': '20260706', 'USERID': '', 'JOURSEQ': '', 'DYTRSEQ': '',
            'LOCAC': '', 'CCY': '', 'BUSCD': '', 'UNIT': '', 'TRCD': '',
            'CUSTOMER': '', 'TRTP': '', 'REMARK': '', 'CRTDTM': '',
        }])
        core_out = process_core(core_df, citad_mapdc, hub_lookups, 20260706)
        assert core_out['TT'].iloc[0] == 'citad 6.7', (
            f"Sau khi strip số 0 đầu, Map dc phải khớp citad — nhận TT={core_out['TT'].iloc[0]!r}"
        )


class TestHubSmfSkipsVlookup:
    """
    Người dùng xác nhận 2026-07-16: dòng Hub có "Số giao dịch" chứa 'SMF' (giao
    dịch smart form) phải giữ nguyên Số Trace 1 gốc — KHÔNG chạy VLOOKUP qua
    EICP, dù thuộc nhóm "Số giao dịch" chứa 'S' + không BFX (lẽ ra sẽ qua EICP).
    Trước đây code chạy VLOOKUP cho cả cột không phân biệt SMF — không mong muốn.
    """

    def test_smf_keeps_raw_trace1_even_when_eicp_has_a_match(self):
        df = pd.DataFrame([_hub_row('SMF12345', 'STC_SMF', 'RAW_TRACE_GIU_NGUYEN', 'Hoàn thành', '06/07/2026 09:00')])
        eicp_maps = {'hub_to_core': {'SMF12345': 'TRACE_TU_EICP_KHONG_DUOC_DUNG'}}
        hub_out, _ = process_hub(df, eicp_maps, 20260706)
        assert hub_out['Trace'].iloc[0] == 'RAW_TRACE_GIU_NGUYEN'

    def test_smf_case_insensitive(self):
        df = pd.DataFrame([_hub_row('smfABCDE', 'STC_SMF2', 'RAW2', 'Hoàn thành', '06/07/2026 09:00')])
        eicp_maps = {'hub_to_core': {'smfABCDE': 'SAI_KHONG_DUOC_DUNG'}}
        hub_out, _ = process_hub(df, eicp_maps, 20260706)
        assert hub_out['Trace'].iloc[0] == 'RAW2'

    def test_non_smf_s_transaction_still_uses_eicp(self):
        """Không phải SMF thì vẫn qua EICP như bình thường — không ảnh hưởng logic cũ."""
        df = pd.DataFrame([_hub_row('SA002', 'STC_B', 'T_B', 'OK', '06/07/2026 09:00')])
        eicp_maps = {'hub_to_core': {'SA002': 'TRACE_TU_EICP'}}
        hub_out, _ = process_hub(df, eicp_maps, 20260706)
        assert hub_out['Trace'].iloc[0] == 'TRACE_TU_EICP'


# ── Test 3: HI pattern — không dùng word boundary ────────────────────────────

class TestHIPattern:
    """HI phải khớp khi ký tự liền kề là chữ cái (như '1000HIO000000006')."""

    def test_hi_embedded_in_reference(self):
        rows = [_core_row('1000HIO000000006', 'BRCD1', 3_000_000)]
        df   = pd.DataFrame(rows)
        out  = process_core(df, {}, {}, 20260512)
        assert out['Trace'].iloc[0] == 'Quyết toán', "HI trong REFERENCE phải → Trace='Quyết toán'"

    def test_hi_at_start(self):
        rows = [_core_row('HI20260512ABC', 'BRCD1', 1_000_000)]
        df   = pd.DataFrame(rows)
        out  = process_core(df, {}, {}, 20260512)
        assert out['Trace'].iloc[0] == 'Quyết toán'

    def test_no_hi_no_qt(self):
        rows = [_core_row('API20260512REF0001', 'BRCD1', 2_000_000)]
        df   = pd.DataFrame(rows)
        out  = process_core(df, {}, {}, 20260512)
        assert out['Trace'].iloc[0] != 'Quyết toán'


# ── Test 4: TT gán trực tiếp cho HI (quyết toán) ────────────────────────────

class TestHITT:
    """HI transaction: TT phải là 'quyết toán', không phụ thuộc vào Hub lookup."""

    def test_hi_tt_is_quyet_toan(self):
        rows = [_core_row('1000HIO0001', 'B001', 5_000_000)]
        df   = pd.DataFrame(rows)
        out  = process_core(df, {}, {}, 20260512)
        assert out['TT'].iloc[0] == 'quyết toán', (
            f"HI transaction phải TT='quyết toán', nhận {out['TT'].iloc[0]!r}"
        )

    def test_huy_takes_priority_over_hi(self):
        """Nếu REFERENCE của HI transaction bị hủy (CR+DR=0) → Hủy thắng."""
        rows = [
            _core_row('1000HIO0001', 'B001', 3_000_000, dramount='0'),
            _core_row('1000HIO0001', 'B001', 0,          dramount='3000000'),
        ]
        df  = pd.DataFrame(rows)
        out = process_core(df, {}, {}, 20260512)
        assert all(out['TT'] == 'Hủy'), "Hủy phải ưu tiên hơn quyết toán"


# ── Test 5: Phát hiện Hủy ────────────────────────────────────────────────────

class TestHuyDetection:
    def test_basic_huy(self):
        """CR + reverse (DR same ref) = 0 → Hủy."""
        rows = [
            _core_row('REF001', 'B001', 3_000_000, dramount='0'),
            _core_row('REF001', 'B001', 0,          dramount='3000000'),
        ]
        df  = pd.DataFrame(rows)
        out = process_core(df, {}, {}, 20260512)
        assert all(out['TT'] == 'Hủy')

    def test_non_huy_not_flagged(self):
        """CR ≠ 0 (không có cặp đảo) → không phải Hủy."""
        rows = [_core_row('REF002', 'B001', 5_000_000)]
        df   = pd.DataFrame(rows)
        out  = process_core(df, {}, {}, 20260512)
        assert out['TT'].iloc[0] != 'Hủy'

    def test_partial_reversal_not_huy(self):
        """Đảo một phần (2M vs 3M): tổng ≠ 0 → không phải Hủy."""
        rows = [
            _core_row('REF003', 'B001', 3_000_000, dramount='0'),
            _core_row('REF003', 'B001', 0,          dramount='2000000'),
        ]
        df  = pd.DataFrame(rows)
        out = process_core(df, {}, {}, 20260512)
        assert all(out['TT'] != 'Hủy')

    def test_nan_reference_not_huy(self):
        """REFERENCE là NaN: groupby bỏ qua, không mark Hủy (đúng — không xác định được cặp)."""
        rows = [
            {'REFERENCE': None, 'TRBRCD': 'B001', 'CRAMOUNT': '1000000',
             'DRAMOUNT': '0', 'TRDATE': '20260512', 'USERID': '', 'JOURSEQ': '',
             'DYTRSEQ': '', 'LOCAC': '', 'CCY': '', 'BUSCD': '', 'UNIT': '',
             'TRCD': '', 'CUSTOMER': '', 'TRTP': '', 'REMARK': '', 'CRTDTM': ''},
        ]
        df  = pd.DataFrame(rows)
        out = process_core(df, {}, {}, 20260512)
        assert out['TT'].iloc[0] != 'Hủy'


# ── Test 5.4: detect_huy — phân biệt Hủy cùng ngày / khác ngày ──────────────

def _core_row_dated(ref, brcd, cramount, dramount='0', trdate='20260512', remark=''):
    return {
        'REFERENCE': ref, 'TRBRCD': brcd, 'CRAMOUNT': str(cramount),
        'DRAMOUNT': str(dramount), 'TRDATE': trdate, 'USERID': 'U1',
        'JOURSEQ': '1', 'DYTRSEQ': '1', 'LOCAC': '', 'CCY': 'VND',
        'BUSCD': '', 'UNIT': '', 'TRCD': '', 'CUSTOMER': '', 'TRTP': '',
        'REMARK': remark, 'CRTDTM': '',
    }


class TestDetectHuyCrossDay:
    """Lệnh lập 1 ngày, hủy ngày khác — phải gộp Core NHIỀU ngày mới thấy đủ
    cặp Nợ/Có để phát hiện. Xác nhận nghiệp vụ 2026-07-16: 'Hủy' (cùng ngày)
    và 'Đã hủy' (khác ngày) là 2 trạng thái khác nhau, không phải lỗi gõ tay."""

    def test_same_day_pair_labeled_huy(self):
        rows = [
            _core_row_dated('REF700', 'B001', 3_000_000, dramount='0',       trdate='20260704'),
            _core_row_dated('REF700', 'B001', 0,          dramount='3000000', trdate='20260704'),
        ]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {'REF700': 'Hủy'}

    def test_cross_day_pair_labeled_da_huy(self):
        """CR lập ngày 04/7, DR hủy ngày 06/7 — nếu chỉ xét từng ngày riêng lẻ
        sẽ không bao giờ thấy cặp này (đây là bug đã sửa)."""
        rows = [
            _core_row_dated('REF701', 'B001', 3_000_000, dramount='0',       trdate='20260704'),
            _core_row_dated('REF701', 'B001', 0,          dramount='3000000', trdate='20260706'),
        ]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {'REF701': 'Đã hủy'}

    def test_single_day_view_misses_cross_day_huy(self):
        """Chỉ đưa Core của 1 ngày (04/7) vào detect_huy — dòng DR hủy ở ngày
        06/7 không có mặt nên KHÔNG được phát hiện. Chứng minh vì sao phải gộp
        toàn batch trước khi tính, không thể tính lẻ từng ngày."""
        rows_day1 = [_core_row_dated('REF702', 'B001', 3_000_000, dramount='0', trdate='20260704')]
        df_day1 = pd.DataFrame(rows_day1)
        assert detect_huy(df_day1) == {}

    def test_non_huy_across_days_not_flagged(self):
        rows = [
            _core_row_dated('REF703', 'B001', 5_000_000, trdate='20260704'),
        ]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {}

    def test_process_core_uses_precomputed_huy_map_across_days(self):
        """process_core() nhận huy_map tính sẵn trên toàn batch — dòng của ngày
        đang xử lý (04/7) phải được gán 'Đã hủy' dù bản thân Core ngày 04/7
        không tự chứa đủ cặp Nợ/Có."""
        all_rows = [
            _core_row_dated('REF704', 'B001', 3_000_000, dramount='0',       trdate='20260704'),
            _core_row_dated('REF704', 'B001', 0,          dramount='3000000', trdate='20260706'),
        ]
        huy_map = detect_huy(pd.DataFrame(all_rows))

        day1_df = pd.DataFrame([all_rows[0]])
        out = process_core(day1_df, {}, {}, 20260704, huy_map)
        assert out['TT'].iloc[0] == 'Đã hủy'


# ── Test 5.4b: detect_huy — tín hiệu CRAMOUNT<0 độc lập (P1, thiếu vế gốc) ──

class TestDetectHuyNegativeCrSignal:
    """Dữ liệu Core THẬT sau load_core() luôn có DRAMOUNT=0 (đã lọc lúc đọc) —
    vế 'hủy' chỉ còn cách biểu diễn duy nhất là CRAMOUNT ÂM (đúng quy tắc docx
    gốc: 'nếu Số tiền (-) thì là hủy lệnh ngày cũ'). Nếu vế gốc (+X) nằm ở một
    ngày KHÔNG nằm trong batch đang xử lý (Core không có carryover T-1 như
    Hub/Eicp), tín hiệu net-zero một mình không đủ — cần tín hiệu CR<0 độc lập.
    Xác nhận qua dữ liệu thật 11-12/8/2026: 91 REFERENCE có CRAMOUNT âm, 47/91
    (52%) không có vế gốc dương cùng REFERENCE trong batch 2 ngày đó."""

    def test_negative_cr_without_matching_positive_labeled_da_hu(self):
        """Chỉ có vế âm trong batch (vế gốc dương ở ngày khác, không được nạp)
        — trước khi sửa P1, tín hiệu net-zero một mình bỏ sót ca này."""
        rows = [_core_row_dated('REF800', 'B001', -500_000, dramount='0', trdate='20260706')]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {'REF800': 'Đã hủy'}

    def test_negative_cr_with_same_day_positive_labeled_huy(self):
        """Cả 2 vế (dương + âm) cùng ngày, cùng có mặt trong batch → 'Hủy' (không đổi)."""
        rows = [
            _core_row_dated('REF801', 'B001', 500_000, dramount='0', trdate='20260706'),
            _core_row_dated('REF801', 'B001', -500_000, dramount='0', trdate='20260706'),
        ]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {'REF801': 'Hủy'}

    def test_negative_cr_with_cross_day_positive_labeled_da_huy(self):
        """Cả 2 vế cùng có mặt nhưng khác ngày → 'Đã hủy' (net-zero, không cần tín hiệu CR<0)."""
        rows = [
            _core_row_dated('REF802', 'B001', 500_000, dramount='0', trdate='20260704'),
            _core_row_dated('REF802', 'B001', -500_000, dramount='0', trdate='20260706'),
        ]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {'REF802': 'Đã hủy'}

    def test_positive_cr_alone_not_flagged(self):
        """CRAMOUNT dương, không có dòng âm nào cùng REFERENCE → không phải Hủy."""
        rows = [_core_row_dated('REF803', 'B001', 500_000, dramount='0', trdate='20260706')]
        df = pd.DataFrame(rows)
        assert detect_huy(df) == {}

    def test_process_core_labels_missing_origin_reference_da_huy(self):
        """process_core() nhận huy_map có tín hiệu CR<0 — dòng CRAMOUNT âm phải
        được gán 'Đã hủy' dù không có vế gốc nào trong toàn batch."""
        row = _core_row_dated('REF804', 'B001', -700_000, dramount='0', trdate='20260706')
        df = pd.DataFrame([row])
        huy_map = detect_huy(df)
        out = process_core(df, {}, {}, 20260706, huy_map)
        assert out['TT'].iloc[0] == 'Đã hủy'


# ── Test 5.5: Kênh OSB — REMARK chứa "IBPSILO" ──────────────────────────────

class TestOSBChannel:
    """REMARK chứa 'IBPSILO' → TT='OSB'. Xác nhận qua dữ liệu thật 04-06/7/2026:
    3/3 dòng IBPSILO trong bài chấm tay đều TT='OSB'."""

    def test_ibpsilo_remark_sets_tt_osb(self):
        row = _core_row('REF900', 'B001', 5_000_000)
        row['REMARK'] = 'IBPSILO'
        df  = pd.DataFrame([row])
        out = process_core(df, {}, {}, 20260512)
        assert out['TT'].iloc[0] == 'OSB'

    def test_ibpsilo_case_insensitive(self):
        row = _core_row('REF901', 'B001', 5_000_000)
        row['REMARK'] = 'ibpsilo transfer'
        df  = pd.DataFrame([row])
        out = process_core(df, {}, {}, 20260512)
        assert out['TT'].iloc[0] == 'OSB'

    def test_no_ibpsilo_not_flagged_osb(self):
        row = _core_row('REF902', 'B001', 5_000_000)
        row['REMARK'] = 'chuyen tien thuong'
        df  = pd.DataFrame([row])
        out = process_core(df, {}, {}, 20260512)
        assert out['TT'].iloc[0] != 'OSB'

    def test_huy_takes_priority_over_osb(self):
        """REFERENCE bị hủy (CR+DR=0) VÀ REMARK chứa IBPSILO → Hủy vẫn thắng."""
        rows = [
            _core_row('REF903', 'B001', 3_000_000, dramount='0'),
            _core_row('REF903', 'B001', 0,          dramount='3000000'),
        ]
        rows[0]['REMARK'] = 'IBPSILO'
        rows[1]['REMARK'] = 'IBPSILO'
        df  = pd.DataFrame(rows)
        out = process_core(df, {}, {}, 20260512)
        assert all(out['TT'] == 'Hủy'), "Hủy phải ưu tiên hơn OSB"


# ── Test 6: Citad TT label — 'citad 12.5' ────────────────────────────────────

class TestCitadTTLabel:
    def test_label_format(self):
        """ngay_int = 20260512 → citad label = 'citad 12.5' (không phải '20260512')."""
        citad_df = pd.DataFrame([{
            'SERIAL_NO':   'STC001',
            'RELATION_NO': '12345678',
            'TRX_DATE':    '20260512',
            'AMOUNT':      '1000000',
            'TRX_STATUS':  'OK',
        }])
        hub_lookups = {'stc_to_trace': {'STC001': 'TRC001'}}
        _, mapdc_to_ngay = process_citad(citad_df, hub_lookups, 20260512)

        # label phải là 'citad 12.5'
        for label in mapdc_to_ngay.values():
            assert label == 'citad 12.5', f"Label sai: {label!r}"

    def test_single_digit_day_in_label(self):
        """ngay_int = 20260505 → 'citad 5.5'."""
        citad_df = pd.DataFrame([{
            'SERIAL_NO':   'STC002',
            'RELATION_NO': '87654321',
            'TRX_DATE':    '20260505',
            'AMOUNT':      '500000',
            'TRX_STATUS':  'OK',
        }])
        hub_lookups = {'stc_to_trace': {'STC002': 'TRC002'}}
        _, mapdc_to_ngay = process_citad(citad_df, hub_lookups, 20260505)
        for label in mapdc_to_ngay.values():
            assert label == 'citad 5.5', f"Label sai: {label!r}"


# ── Test 7: Citad AMOUNT "ltd" → TRX_STATUS ──────────────────────────────────

class TestCitadAmountLtd:
    def test_ltd_replaced_from_trx_status(self):
        citad_df = pd.DataFrame([{
            'SERIAL_NO':   'STC_LTD',
            'RELATION_NO': 'ABCD1234',
            'TRX_DATE':    '20260512',
            'AMOUNT':      'ltd',
            'TRX_STATUS':  '2500000',
        }])
        hub_lookups = {'stc_to_trace': {}}
        out, _ = process_citad(citad_df, hub_lookups, 20260512)
        assert out['AMOUNT'].iloc[0] == 2_500_000, (
            f"AMOUNT 'ltd' phải được thay bằng TRX_STATUS, nhận {out['AMOUNT'].iloc[0]}"
        )

    def test_normal_amount_unchanged(self):
        citad_df = pd.DataFrame([{
            'SERIAL_NO':   'STC_N',
            'RELATION_NO': 'ABCD0000',
            'TRX_DATE':    '20260512',
            'AMOUNT':      '3000000',
            'TRX_STATUS':  'OK',
        }])
        hub_lookups = {'stc_to_trace': {}}
        out, _ = process_citad(citad_df, hub_lookups, 20260512)
        assert out['AMOUNT'].iloc[0] == 3_000_000


# ── Test 8: EICP first-match (không dùng dict(zip) last-match) ───────────────

class TestEICPFirstMatch:
    def test_duplicate_msgkey_first_wins(self):
        from backend.services.ilo1000.load_eicp import build_eicp_maps
        eicp_df = pd.DataFrame([
            {'BRCD': 'B001', 'MSGKEY': 'MSG001', 'TRSEQ': 'TRQ_FIRST'},
            {'BRCD': 'B001', 'MSGKEY': 'MSG001', 'TRSEQ': 'TRQ_LAST'},
        ])
        maps = build_eicp_maps(eicp_df)
        key = 'B001MSG001'
        assert key in maps['hub_to_core']
        assert maps['hub_to_core'][key] == 'B001OTTTRQ_FIRST', (
            "EICP phải giữ lần xuất hiện ĐẦU TIÊN của MSGKEY"
        )


# ── Test 9: Hub Trace — nhánh 'S' (BFX/EICP) vs không 'S' (Số Trace 2) ──────

class TestHubEicpIndexAlignment:
    """
    Theo tài liệu gốc (CÁC BƯỚC LÀM ĐỐI CHIẾU ILO1.docx): chỉ giao dịch có
    "Số giao dịch" chứa 'S' mới qua bước BFX/EICP. Giao dịch KHÔNG chứa 'S'
    (chủ yếu loại "OT" — hoàn trả lệnh gốc) dùng thẳng "Số Trace 2" — xác nhận
    qua đối chiếu dữ liệu thật 06/07/2026: nhóm không 'S' khớp 96,5% qua Số
    Trace 2 (so với chỉ 50% qua Số Trace 1); riêng OT khớp 100% qua Số Trace 2.
    Trước đây (dựa trên dữ liệu tháng 5, chưa biết cột Số Trace 2 tồn tại) đã
    bỏ hẳn filter 'S' và áp EICP cho mọi dòng — nay xác nhận đó là sai, EICP
    chỉ đúng cho nhóm 'S'. Nhưng KHÔNG được filter BỚT DÒNG theo 'S' — mọi
    dòng (kể cả OT) vẫn phải có mặt để CITAD/CORE tra Trace/Map dc.
    EICP lookup phải dùng .map() trên toàn bộ df, không phải indexing trực tiếp.
    """

    def test_no_row_dropped_and_branch_applied_correctly(self):
        rows = [
            _hub_row('A001',  'STC_A', 'T_A', 'OK', '12/05/2026 09:00'),   # không chứa 'S' -> giữ Trace 1 (không có Trace 2)
            _hub_row('SA002', 'STC_B', 'T_B', 'OK', '12/05/2026 09:00'),   # chứa 'S' -> EICP
            _hub_row('A003',  'STC_C', 'T_C', 'OK', '12/05/2026 09:00'),   # không chứa 'S' -> giữ Trace 1
            _hub_row('SB004', 'STC_D', 'T_D', 'OK', '12/05/2026 09:00'),   # chứa 'S' -> EICP
        ]
        df = pd.DataFrame(rows)

        eicp_maps = {'hub_to_core': {
            'A001':  'TRACE_A001_CORE',
            'SA002': 'TRACE_SA002_CORE',
            'A003':  'TRACE_A003_CORE',
            'SB004': 'TRACE_SB004_CORE',
        }}
        hub_out, _ = process_hub(df, eicp_maps, 20260512)

        assert len(hub_out) == 4, "Không được filter bớt dòng theo 'Số giao dịch'"
        for so_gd, expected_trace in [
            ('A001', 'T_A'), ('SA002', 'TRACE_SA002_CORE'),
            ('A003', 'T_C'), ('SB004', 'TRACE_SB004_CORE'),
        ]:
            row = hub_out[hub_out[HUB_COL_SO_GD] == so_gd].iloc[0]
            assert row['Trace'] == expected_trace, f"{so_gd} Trace sai: {row['Trace']!r}"

    def test_not_s_uses_so_trace_2_when_present(self):
        """Giao dịch OT (không chứa 'S') phải dùng Số Trace 2, bỏ qua EICP dù EICP có khớp."""
        row = _hub_row('1000OT260706285085', 'STC_OT', '005278295', 'Hoàn thành', '06/07/2026 09:44')
        row['Số Trace 2'] = '005469531'
        df = pd.DataFrame([row])

        eicp_maps = {'hub_to_core': {'1000OT260706285085': 'SAI_LE_EICP_KHONG_DUOC_DUNG'}}
        hub_out, _ = process_hub(df, eicp_maps, 20260706)
        assert hub_out['Trace'].iloc[0] == '5469531', (
            "Dòng OT phải dùng Số Trace 2 (sau strip số 0 đầu), không dùng EICP"
        )

    def test_not_s_falls_back_to_trace1_when_trace2_empty(self):
        row = _hub_row('1000OT260706999999', 'STC_OT2', '00123', 'Hoàn thành', '06/07/2026 09:44')
        row['Số Trace 2'] = ''
        df = pd.DataFrame([row])
        hub_out, _ = process_hub(df, {}, 20260706)
        assert hub_out['Trace'].iloc[0] == '123'

    def test_lowercase_s_also_triggers_eicp_branch(self):
        """Filter 'S' không phân biệt hoa/thường — xác nhận từ người dùng 2026-07-16."""
        df = pd.DataFrame([_hub_row('1000sA002', 'STC_low', 'RAW', 'OK', '06/07/2026 09:00')])
        eicp_maps = {'hub_to_core': {'1000sA002': 'TRACE_TU_EICP_LOWERCASE_S'}}
        hub_out, _ = process_hub(df, eicp_maps, 20260706)
        assert hub_out['Trace'].iloc[0] == 'TRACE_TU_EICP_LOWERCASE_S'


# ── Test 10: Luồng đầu cuối mini (integration) ───────────────────────────────

class TestEndToEndMini:
    """Pipeline nhỏ: 1 giao dịch mỗi loại, kiểm tra TT cuối."""

    def test_citad_match_sets_tt(self):
        """Core khớp citad qua Map dc → TT = 'citad 12.5'."""
        # Hub: SA001 → STC001 → Trace=TRC001
        hub_rows = [_hub_row('SA001', 'STC001', 'TRC001', 'Thành công', '12/05/2026 09:00')]
        hub_df = pd.DataFrame(hub_rows)
        hub_out, hub_lookups = process_hub(hub_df, {}, 20260512)

        # Citad: SERIAL_NO=STC001 → Trace=TRC001; RELATION_NO[:4]=1234; AMOUNT=1000000
        # Map dc = '1234' + 'TRC001' + '1000000'
        citad_rows = [{
            'SERIAL_NO': 'STC001', 'RELATION_NO': '12345678',
            'TRX_DATE': '20260512', 'AMOUNT': '1000000', 'TRX_STATUS': 'OK',
        }]
        citad_df = pd.DataFrame(citad_rows)
        _, citad_mapdc = process_citad(citad_df, hub_lookups, 20260512)

        # Core: TRBRCD=B001, Trace=TRC001, CRAMOUNT=1000000
        # Map dc = 'B001' + 'TRC001' + '1000000'
        # CITAD Map dc = '1234' + 'TRC001' + '1000000'  → khác BRCD → không khớp
        # Thay BRCD khớp với LEFT(RELATION_NO,4)='1234'
        core_rows = [_core_row('APIREF001_TRC001_XXX', '1234', 1_000_000)]
        # Cần trace core là TRC001 để map dc khớp:
        # Với REFERENCE 'APIREF001_TRC001_XXX', trace = REFERENCE[7:23] = 'T_TRC001_XXX000'
        # Ta dùng REFERENCE 'API    TRC001  XXXX' để Trace = ref[7:23] = 'TRC001  XXXX    '
        # Dễ hơn: đặt REFERENCE chứa 'OTT' với TRBRCD='1234', REFERENCE[4:16]='TRC001      '
        core_rows2 = [{
            'REFERENCE': 'OTT_TRC001__',  # OTT → Trace = BRCD+REF[4:16]
            'TRBRCD': '1234',
            'CRAMOUNT': '1000000', 'DRAMOUNT': '0',
            'TRDATE': '20260512', 'USERID': '', 'JOURSEQ': '', 'DYTRSEQ': '',
            'LOCAC': '', 'CCY': '', 'BUSCD': '', 'UNIT': '', 'TRCD': '',
            'CUSTOMER': '', 'TRTP': '', 'REMARK': '', 'CRTDTM': '',
        }]
        core_df = pd.DataFrame(core_rows2)
        core_out = process_core(core_df, citad_mapdc, hub_lookups, 20260512)

        # Trace = '1234' + 'TRC001      '
        # Map dc = '1234' + Trace + '1000000'
        # citad Map dc = '1234' + 'TRC001' + '1000000'
        # Trace từ OTT = BRCD + REF[4:16] = '1234' + 'TRC001__'
        # Map dc core = '1234' + '1234TRC001__' + '1000000'
        # Map dc citad = '1234' + 'TRC001' + '1000000'
        # → không khớp vì Trace có BRCD prefix
        # Đây là giới hạn của test đơn giản này; ta kiểm tra TT ≠ '' là đủ
        # (citad match đòi hỏi Trace chính xác giống nhau)
        tt = core_out['TT'].iloc[0]
        # Không expect citad match ở đây vì format trace khác
        # Nhưng ít nhất TT phải là string (không phải NaN hay crash)
        assert isinstance(tt, str)


# ── Test 11: load_hub gộp nhiều file pHub cùng ngày ─────────────────────────

class TestLoadHubMultiFile:
    """
    pHub có thể export theo nhiều đợt/batch trong cùng 1 ngày (VD: sáng + chiều).
    load_hub phải gộp TẤT CẢ file, không được ghi đè/mất dữ liệu của file trước.
    """

    def test_concat_multiple_files(self, tmp_path):
        from backend.services.ilo1000.load_hub import load_hub

        def _write_phub(path, so_gd_list):
            rows = []
            for so_gd in so_gd_list:
                rows.append({
                    'Số giao dịch': so_gd, 'Số Ref Hub': 'REF' + so_gd,
                    'Số thành công': 'STC_' + so_gd, 'Số Trace 1': 'T_' + so_gd,
                    'Số tiền thực chuyển': '1000000', 'Trạng thái': 'Hoàn thành',
                    'Ngày giờ kênh trả': '12/05/2026 09:00', 'Nội dung chuyển tiền': '',
                })
            df = pd.DataFrame(rows)
            # Ghi với 1 dòng title giả ở row 0, header ở row 1 (đúng format pHub thật)
            with pd.ExcelWriter(path) as writer:
                df.to_excel(writer, index=False, header=True, startrow=1)

        p1 = tmp_path / 'batch1.xlsx'
        p2 = tmp_path / 'batch2.xlsx'
        _write_phub(p1, ['A001', 'A002'])
        _write_phub(p2, ['B001', 'B002'])

        result = load_hub([p1, p2])
        assert len(result) == 4, "Phải gộp đủ dòng từ cả 2 file, không mất dòng của file nào"
        assert set(result[HUB_COL_SO_GD]) == {'A001', 'A002', 'B001', 'B002'}

    def test_dedup_same_transaction_across_files(self, tmp_path):
        """Nếu 1 giao dịch (Số giao dịch) bị export trùng ở nhiều file → chỉ giữ 1 dòng."""
        from backend.services.ilo1000.load_hub import load_hub

        def _write_phub(path, so_gd_list):
            rows = [{
                'Số giao dịch': so_gd, 'Số Ref Hub': 'REF' + so_gd,
                'Số thành công': 'STC_' + so_gd, 'Số Trace 1': 'T_' + so_gd,
                'Số tiền thực chuyển': '1000000', 'Trạng thái': 'Hoàn thành',
                'Ngày giờ kênh trả': '12/05/2026 09:00', 'Nội dung chuyển tiền': '',
            } for so_gd in so_gd_list]
            df = pd.DataFrame(rows)
            with pd.ExcelWriter(path) as writer:
                df.to_excel(writer, index=False, header=True, startrow=1)

        p1 = tmp_path / 'batch1.xlsx'
        p2 = tmp_path / 'batch2.xlsx'
        _write_phub(p1, ['A001'])
        _write_phub(p2, ['A001'])

        result = load_hub([p1, p2])
        assert len(result) == 1, "Giao dịch trùng ở 2 file phải dedup còn 1 dòng"

    def test_tong_tien_footer_row_excluded(self, tmp_path):
        """
        pHub thật luôn có dòng 'Tổng tiền' ở cuối (Excel tự cộng) — STT='Tổng tiền',
        Số giao dịch rỗng, chỉ có Số tiền thực chuyển = tổng cả file. Nếu không loại,
        dòng này cộng dồn sai vào tổng Hub (xác nhận qua dữ liệu thật 04-06/7/2026:
        lệch đúng bằng giá trị dòng tổng, ~561 tỷ).
        """
        from backend.services.ilo1000.load_hub import load_hub

        rows = [
            {'STT': 1, 'Số giao dịch': 'A001', 'Số Ref Hub': 'REFA001',
             'Số thành công': 'STC_A001', 'Số Trace 1': 'T_A001',
             'Số tiền thực chuyển': '1000000', 'Trạng thái': 'Hoàn thành',
             'Ngày giờ kênh trả': '06/07/2026 09:00', 'Nội dung chuyển tiền': ''},
            {'STT': 'Tổng tiền', 'Số giao dịch': None, 'Số Ref Hub': None,
             'Số thành công': None, 'Số Trace 1': None,
             'Số tiền thực chuyển': '561304218679', 'Trạng thái': None,
             'Ngày giờ kênh trả': None, 'Nội dung chuyển tiền': None},
        ]
        df = pd.DataFrame(rows)
        p = tmp_path / 'phub.xlsx'
        with pd.ExcelWriter(p) as writer:
            df.to_excel(writer, index=False, header=True, startrow=1)

        result = load_hub([p])
        assert len(result) == 1, "Dòng 'Tổng tiền' phải bị loại, chỉ còn 1 giao dịch thật"
        assert list(result[HUB_COL_SO_GD]) == ['A001']
        assert result['Số tiền thực chuyển'].astype(float).sum() == 1_000_000


# ── Test 12a: EICP không khớp ngày nào — gán vào nhóm nhiều dữ liệu nhất ────

class TestEicpUnmatchedFallback:
    """
    EICP đặt tên không theo công thức chuẩn (VD 'eicp 3 7.XLS' — trích ngày=3
    nhưng không có nhóm Hub/Citad/Core nào ngày 03) trước đây bị BỎ HẲN, chỉ log
    WARN. Xác nhận qua dữ liệu thật 4-6.7.2026 (người dùng chỉ ra) — phải gán vào
    nhóm có nhiều citad+core nhất, cùng cơ chế fallback đã có sẵn cho Hub.
    """

    def _write_core(self, path):
        path.write_text('TRDATE,TRBRCD\n20260706,1000\n', encoding='utf-8')

    def _write_citad(self, path):
        path.write_text(
            'SERIAL_NO,RELATION_NO,TRX_DATE,AMOUNT,TRX_STATUS,extra\n1,2,20260706,100,OK,\n',
            encoding='utf-8',
        )

    def test_unmatched_eicp_assigned_to_biggest_group(self, tmp_path):
        from backend.services.ilo1000.detect import group_files_by_date

        core_path = tmp_path / 'gl02_20260706.csv'
        self._write_core(core_path)
        citad_path = tmp_path / 'citad.csv'
        self._write_citad(citad_path)

        eicp_path = tmp_path / 'eicp 3 7.XLS'  # trích ngày=3, không có nhóm ngày 03 nào
        eicp_path.write_bytes(b'')

        groups = group_files_by_date([core_path, citad_path, eicp_path])
        assert eicp_path in groups['20260706']['eicp'], (
            "EICP không khớp ngày nào phải được gán vào nhóm có nhiều citad+core nhất, không bị bỏ"
        )

    def test_matched_eicp_still_prefers_own_date_group(self, tmp_path):
        """Nếu EICP khớp đúng ngày sẵn có, vẫn ưu tiên gán đúng nhóm đó (không đổi hành vi cũ)."""
        from backend.services.ilo1000.detect import group_files_by_date

        core6 = tmp_path / 'gl02_20260706.csv'
        self._write_core(core6)
        citad6 = tmp_path / 'citad.csv'
        self._write_citad(citad6)

        eicp6 = tmp_path / 'eicp 6.XLS'  # trích ngày=6, khớp đúng nhóm 20260706
        eicp6.write_bytes(b'')

        groups = group_files_by_date([core6, citad6, eicp6])
        assert eicp6 in groups['20260706']['eicp']

    def test_no_groups_at_all_just_warns(self, tmp_path):
        """Không có nhóm nào (không citad/core) — EICP không có chỗ gán, chỉ WARN, không crash."""
        from backend.services.ilo1000.detect import group_files_by_date

        eicp_path = tmp_path / 'eicp 3 7.XLS'
        eicp_path.write_bytes(b'')
        logs = []
        groups = group_files_by_date([eicp_path], log=logs.append)
        assert groups == {}
        assert any('WARN' in l and 'EICP' in l for l in logs)


# ── Test 11b: Ngày thường trong tuần — gộp thêm Hub/EICP của T-1 ────────────

class TestPreviousDayHubEicpCarryover:
    """
    Theo tài liệu gốc bước 1: mọi ngày chấm bình thường (không phải thứ 2) đều
    phải gộp thêm Hub/EICP của T-1 — lệnh vào hệ thống sau cutoff T-1 chờ đi
    kênh sang T, Hub/EICP của lệnh đó vẫn nằm trong file T-1. Citad/Core giữ
    nguyên chỉ ngày T. Thứ 2 dùng cơ chế riêng (merge_monday_carryover), không
    áp dụng quy tắc T-1 đơn giản này (người dùng xác nhận 2026-07-16).
    """

    def _empty_group(self):
        return {'hub': [], 'citad': [], 'eicp': [], 'core': []}

    def test_regular_day_merges_prev_day_hub_and_eicp(self):
        from backend.services.ilo1000.detect import merge_previous_day_hub_eicp
        from pathlib import Path

        # 2026-07-14 = thứ 3, T-1 = 13/7 (thứ 2)
        groups = {
            '20260714': {**self._empty_group(), 'hub': [Path('tue_hub.xlsx')],
                         'citad': [Path('tue_citad.csv')], 'core': [Path('tue_core.csv')]},
            '20260713': {**self._empty_group(), 'hub': [Path('mon_hub.xlsx')],
                         'eicp': [Path('mon_eicp.xls')], 'core': [Path('mon_core.csv')]},
        }
        merge_previous_day_hub_eicp(groups)
        tue = groups['20260714']

        assert set(tue['hub']) == {Path('tue_hub.xlsx'), Path('mon_hub.xlsx')}
        assert set(tue['eicp']) == {Path('mon_eicp.xls')}
        assert tue['core'] == [Path('tue_core.csv')], "Core KHÔNG được gộp T-1, chỉ Hub/EICP"
        assert tue['citad'] == [Path('tue_citad.csv')], "Citad chỉ giữ ngày T"

    def test_monday_not_affected_by_this_rule(self):
        """Thứ 2 không áp dụng quy tắc T-1 đơn giản — dùng merge_monday_carryover riêng."""
        from backend.services.ilo1000.detect import merge_previous_day_hub_eicp
        from pathlib import Path

        # 2026-07-13 = thứ 2, T-1 = 12/7 (CN)
        groups = {
            '20260713': {**self._empty_group(), 'citad': [Path('mon_citad.csv')]},
            '20260712': {**self._empty_group(), 'hub': [Path('sun_hub.xlsx')]},
        }
        merge_previous_day_hub_eicp(groups)
        assert groups['20260713']['hub'] == [], "Thứ 2 không được áp dụng quy tắc T-1 đơn giản"

    def test_prev_day_own_group_unchanged_after_merge(self):
        """COPY, không phải move — nhóm T-1 gốc vẫn tự ra báo cáo riêng bình thường."""
        from backend.services.ilo1000.detect import merge_previous_day_hub_eicp
        from pathlib import Path

        groups = {
            '20260714': {**self._empty_group(), 'citad': [Path('tue_citad.csv')]},
            '20260713': {**self._empty_group(), 'hub': [Path('mon_hub.xlsx')], 'citad': [Path('mon_citad.csv')]},
        }
        merge_previous_day_hub_eicp(groups)
        assert groups['20260713']['hub'] == [Path('mon_hub.xlsx')]
        assert groups['20260713']['citad'] == [Path('mon_citad.csv')]

    def test_missing_prev_day_warns_but_still_runs(self):
        from backend.services.ilo1000.detect import merge_previous_day_hub_eicp
        from pathlib import Path

        groups = {'20260714': {**self._empty_group(), 'citad': [Path('tue_citad.csv')]}}
        logs = []
        merge_previous_day_hub_eicp(groups, log=logs.append)
        assert groups['20260714']['hub'] == []
        assert any('CẢNH BÁO' in l and 'T-1' in l for l in logs)

    def test_does_not_leak_mondays_carryover_into_tuesday(self):
        """
        Thứ 3 chỉ lấy Hub/EICP GỐC của thứ 2 (T-1), KHÔNG được kéo theo dữ liệu
        cuối tuần mà thứ 2 đã tự gộp thêm qua merge_monday_carryover — đây là lý
        do merge_previous_day_hub_eicp phải chạy TRƯỚC merge_monday_carryover.
        """
        from backend.services.ilo1000.detect import group_files_by_date
        from pathlib import Path
        import backend.services.ilo1000.detect as detect_mod

        groups = {
            '20260714': {**self._empty_group(), 'citad': [Path('tue_citad.csv')]},
            '20260713': {**self._empty_group(), 'hub': [Path('mon_hub.xlsx')], 'citad': [Path('mon_citad.csv')]},
            '20260710': {**self._empty_group(), 'hub': [Path('fri_hub.xlsx')]},
        }
        detect_mod.merge_previous_day_hub_eicp(groups)
        detect_mod.merge_monday_carryover(groups)

        assert Path('fri_hub.xlsx') not in groups['20260714']['hub'], (
            "Thứ 3 không được kéo theo Hub cuối tuần mà thứ 2 tự gộp thêm"
        )
        assert set(groups['20260714']['hub']) == {Path('mon_hub.xlsx')}


# ── Test 12: Chấm thứ 2 — gộp dữ liệu cuối tuần ─────────────────────────────

class TestMondayCarryover:
    """
    Citad không chạy phiên thứ 7/CN → lệnh chờ đi kênh sau cutoff thứ 6 + cả
    thứ 7 + CN dồn sang phiên thứ 2. merge_monday_carryover phải gộp thêm
    Hub/EICP (T-3,T-2,T-1) và Core (T-2,T-1) vào nhóm thứ 2; Citad giữ nguyên.
    """

    def _empty_group(self):
        return {'hub': [], 'citad': [], 'eicp': [], 'core': []}

    def test_merges_hub_eicp_from_fri_sat_sun_and_core_from_sat_sun_only(self):
        from backend.services.ilo1000.detect import merge_monday_carryover
        from pathlib import Path

        # 2026-07-13 = thứ 2. T-3=10/7 (T6), T-2=11/7 (T7), T-1=12/7 (CN)
        groups = {
            '20260713': {**self._empty_group(), 'hub': [Path('mon_hub.xlsx')],
                         'citad': [Path('mon_citad.csv')], 'core': [Path('mon_core.csv')]},
            '20260710': {**self._empty_group(), 'hub': [Path('fri_hub.xlsx')],
                         'eicp': [Path('fri_eicp.xls')], 'core': [Path('fri_core.csv')]},
            '20260711': {**self._empty_group(), 'hub': [Path('sat_hub.xlsx')],
                         'eicp': [Path('sat_eicp.xls')], 'core': [Path('sat_core.csv')]},
            '20260712': {**self._empty_group(), 'hub': [Path('sun_hub.xlsx')],
                         'eicp': [Path('sun_eicp.xls')], 'core': [Path('sun_core.csv')]},
        }
        merge_monday_carryover(groups)
        mon = groups['20260713']

        assert set(mon['hub']) == {Path('mon_hub.xlsx'), Path('fri_hub.xlsx'), Path('sat_hub.xlsx'), Path('sun_hub.xlsx')}
        assert set(mon['eicp']) == {Path('fri_eicp.xls'), Path('sat_eicp.xls'), Path('sun_eicp.xls')}
        assert set(mon['core']) == {Path('mon_core.csv'), Path('sat_core.csv'), Path('sun_core.csv')}, (
            "Core KHÔNG được gộp thứ 6 — thứ 6 tự chấm bình thường trong ngày của nó"
        )
        assert mon['citad'] == [Path('mon_citad.csv')], "Citad chỉ giữ ngày thứ 2, không gộp"

    def test_friday_own_group_unchanged_after_merge(self):
        """Gộp vào thứ 2 phải là COPY — nhóm thứ 6 gốc không bị đụng tới."""
        from backend.services.ilo1000.detect import merge_monday_carryover
        from pathlib import Path

        groups = {
            '20260713': {**self._empty_group(), 'citad': [Path('mon_citad.csv')]},
            '20260710': {**self._empty_group(), 'hub': [Path('fri_hub.xlsx')], 'citad': [Path('fri_citad.csv')]},
        }
        merge_monday_carryover(groups)
        assert groups['20260710']['hub'] == [Path('fri_hub.xlsx')], "Nhóm thứ 6 gốc không được thay đổi"
        assert groups['20260710']['citad'] == [Path('fri_citad.csv')]

    def test_month_boundary_monday_is_first_of_month(self):
        """Thứ 2 là ngày 1 đầu tháng → T-3,T-2,T-1 phải tính đúng lịch, rơi vào cuối tháng trước."""
        from backend.services.ilo1000.detect import merge_monday_carryover
        from pathlib import Path

        # 2026-06-01 = thứ 2. T-3=29/5 (T6), T-2=30/5 (T7), T-1=31/5 (CN)
        groups = {
            '20260601': {**self._empty_group(), 'citad': [Path('mon_citad.csv')]},
            '20260529': {**self._empty_group(), 'hub': [Path('fri_hub.xlsx')]},
            '20260530': {**self._empty_group(), 'hub': [Path('sat_hub.xlsx')], 'core': [Path('sat_core.csv')]},
            '20260531': {**self._empty_group(), 'hub': [Path('sun_hub.xlsx')], 'core': [Path('sun_core.csv')]},
        }
        merge_monday_carryover(groups)
        mon = groups['20260601']
        assert set(mon['hub']) == {Path('fri_hub.xlsx'), Path('sat_hub.xlsx'), Path('sun_hub.xlsx')}
        assert set(mon['core']) == {Path('sat_core.csv'), Path('sun_core.csv')}

    def test_missing_weekend_files_warns_but_still_runs(self):
        """Thiếu file bù cuối tuần → không crash, chỉ cảnh báo rõ qua log."""
        from backend.services.ilo1000.detect import merge_monday_carryover
        from pathlib import Path

        groups = {
            '20260713': {**self._empty_group(), 'citad': [Path('mon_citad.csv')]},
        }
        logs = []
        merge_monday_carryover(groups, log=logs.append)

        mon = groups['20260713']
        assert mon['hub'] == [] and mon['core'] == [], "Không có gì để gộp → nhóm thứ 2 giữ nguyên rỗng"
        warn_logs = [l for l in logs if 'CẢNH BÁO' in l]
        assert warn_logs, "Phải log cảnh báo khi thiếu dữ liệu bù cuối tuần"
        assert 'thứ 6' in warn_logs[0] and 'thứ 7' in warn_logs[0] and 'CN' in warn_logs[0]

    def test_non_monday_groups_untouched(self):
        """Ngày không phải thứ 2 (VD thứ 3) không bị áp dụng carryover."""
        from backend.services.ilo1000.detect import merge_monday_carryover
        from pathlib import Path

        # 2026-07-14 = thứ 3
        groups = {
            '20260714': {**self._empty_group(), 'citad': [Path('tue_citad.csv')]},
            '20260713': {**self._empty_group(), 'hub': [Path('mon_hub.xlsx')]},
        }
        merge_monday_carryover(groups)
        assert groups['20260714']['hub'] == [], "Thứ 3 không được gộp thêm gì"


# ── Test 13: load_core — ZIP GL02 mã hóa AES + dedup CSV rời trùng dữ liệu ──

def _gl02_core_row(ref='REF1', dr='0', cr='1000000', trbrcd='1000', journseq='1'):
    return {
        'TRDATE': '20260706', 'TRBRCD': trbrcd, 'USERID': '1000API0', 'JOURSEQ': journseq,
        'DYTRSEQ': '1', 'LOCAC': '501202', 'CCY': 'VND', 'BUSCD': 'EI', 'UNIT': 'AP',
        'TRCD': '', 'CUSTOMER': '1000-000007709', 'TRTP': 'Normal', 'REFERENCE': ref,
        'REMARK': '', 'DRAMOUNT': dr, 'CRAMOUNT': cr, 'CRTDTM': '',
    }


def _write_gl02_zip(path, rows, entry_name='data.csv', password=None):
    from backend.services.ilo1000.config import ZIP_PASSWORD
    pwd = password if password is not None else ZIP_PASSWORD
    cols = ['TRDATE', 'TRBRCD', 'USERID', 'JOURSEQ', 'DYTRSEQ', 'LOCAC', 'CCY',
            'BUSCD', 'UNIT', 'TRCD', 'CUSTOMER', 'TRTP', 'REFERENCE', 'REMARK',
            'DRAMOUNT', 'CRAMOUNT', 'CRTDTM']
    df = pd.DataFrame(rows)[cols]
    csv_bytes = df.to_csv(index=False).encode('utf-8-sig')
    with pyzipper.AESZipFile(path, 'w', compression=pyzipper.ZIP_DEFLATED,
                              encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(pwd)
        zf.writestr(entry_name, csv_bytes)


class TestLoadCoreEncryptedZip:
    """
    GL02 zip thật (đối chiếu thứ 2 06/07/2026) mã hóa AES — trước đây load_core
    dùng zipfile.ZipFile thường (không mật khẩu) nên mọi entry bị bỏ qua với
    WARN 'mã hóa'. Fix: dùng pyzipper.AESZipFile + ZIP_PASSWORD chung với
    module Chấm 459901/ACH (xác nhận cùng 1 mật khẩu qua thử thật trên file GL02
    thật của người dùng).
    """

    def test_reads_aes_encrypted_zip_with_correct_password(self, tmp_path):
        from backend.services.ilo1000.load_core import load_core

        zpath = tmp_path / 'GL02_20260706_1000.zip'
        _write_gl02_zip(zpath, [_gl02_core_row('REF1'), _gl02_core_row('REF2', dr='500000', cr='0')])

        df = load_core([zpath])
        # DRAMOUNT=0 filter giữ REF1 (dr='0'), loại REF2 (dr='500000')
        assert len(df) == 1
        assert df.iloc[0]['REFERENCE'] == 'REF1'

    def test_wrong_password_entry_skipped_with_warning_not_crash(self, tmp_path):
        from backend.services.ilo1000.load_core import load_core

        zpath = tmp_path / 'bad.zip'
        _write_gl02_zip(zpath, [_gl02_core_row('REF1')], password=b'sai-mat-khau')

        logs = []
        df = load_core([zpath], log=logs.append)
        assert len(df) == 0, "Sai mật khẩu → bỏ qua entry, không crash"
        assert any('mã hóa' in l for l in logs)

    def test_multi_part_zip_all_entries_loaded(self, tmp_path):
        """GL02 thật chia thành nhiều CSV part trong 1 zip (_1, _2, ...) — phải đọc đủ tất cả."""
        from backend.services.ilo1000.load_core import load_core
        from backend.services.ilo1000.config import ZIP_PASSWORD

        zpath = tmp_path / 'multi.zip'
        rows_p1 = [_gl02_core_row('PART1_REF1'), _gl02_core_row('PART1_REF2')]
        rows_p2 = [_gl02_core_row('PART2_REF1')]
        cols = ['TRDATE', 'TRBRCD', 'USERID', 'JOURSEQ', 'DYTRSEQ', 'LOCAC', 'CCY',
                'BUSCD', 'UNIT', 'TRCD', 'CUSTOMER', 'TRTP', 'REFERENCE', 'REMARK',
                'DRAMOUNT', 'CRAMOUNT', 'CRTDTM']
        with pyzipper.AESZipFile(zpath, 'w', compression=pyzipper.ZIP_DEFLATED,
                                  encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(ZIP_PASSWORD)
            zf.writestr('gl02_1.csv', pd.DataFrame(rows_p1)[cols].to_csv(index=False).encode('utf-8-sig'))
            zf.writestr('gl02_2.csv', pd.DataFrame(rows_p2)[cols].to_csv(index=False).encode('utf-8-sig'))

        df = load_core([zpath])
        assert set(df['REFERENCE']) == {'PART1_REF1', 'PART1_REF2', 'PART2_REF1'}


class TestLoadCoreDedup:
    """
    Thực nghiệm trên dữ liệu thật 06/07/2026: 1 CSV rời (export tay) trùng 100%
    (byte-identical) với 1 trong các part bên trong ZIP GL02. Từ khi ZIP đọc
    được, dùng cả CSV rời + ZIP (không còn bỏ ZIP) → phải dedup dòng trùng,
    không đếm 2 lần.
    """

    def test_duplicate_rows_between_zip_and_standalone_csv_deduped(self, tmp_path):
        from backend.services.ilo1000.load_core import load_core

        zpath = tmp_path / 'GL02_20260706_1000.zip'
        dup_row = _gl02_core_row('DUP_REF', journseq='99')
        _write_gl02_zip(zpath, [dup_row, _gl02_core_row('ZIP_ONLY_REF')])

        cols = ['TRDATE', 'TRBRCD', 'USERID', 'JOURSEQ', 'DYTRSEQ', 'LOCAC', 'CCY',
                'BUSCD', 'UNIT', 'TRCD', 'CUSTOMER', 'TRTP', 'REFERENCE', 'REMARK',
                'DRAMOUNT', 'CRAMOUNT', 'CRTDTM']
        csv_path = tmp_path / '1000_gl02_2026070420260706.csv'
        pd.DataFrame([dup_row])[cols].to_csv(csv_path, index=False, encoding='utf-8-sig')

        df = load_core([zpath, csv_path])
        assert len(df) == 2, "Dòng trùng (DUP_REF) chỉ được giữ 1 lần, không đếm 2 lần"
        assert set(df['REFERENCE']) == {'DUP_REF', 'ZIP_ONLY_REF'}

    def test_non_duplicate_rows_all_kept(self, tmp_path):
        """Dòng KHÔNG trùng (khác JOURSEQ/REFERENCE) giữa CSV rời và ZIP phải giữ đủ, không bị dedup nhầm."""
        from backend.services.ilo1000.load_core import load_core

        zpath = tmp_path / 'z.zip'
        _write_gl02_zip(zpath, [_gl02_core_row('REF_A', journseq='1')])

        cols = ['TRDATE', 'TRBRCD', 'USERID', 'JOURSEQ', 'DYTRSEQ', 'LOCAC', 'CCY',
                'BUSCD', 'UNIT', 'TRCD', 'CUSTOMER', 'TRTP', 'REFERENCE', 'REMARK',
                'DRAMOUNT', 'CRAMOUNT', 'CRTDTM']
        csv_path = tmp_path / 'extra.csv'
        pd.DataFrame([_gl02_core_row('REF_B', journseq='2')])[cols].to_csv(csv_path, index=False, encoding='utf-8-sig')

        df = load_core([zpath, csv_path])
        assert len(df) == 2
        assert set(df['REFERENCE']) == {'REF_A', 'REF_B'}


# ── Test 14: load_core — lọc GL02 về đúng kênh Citad Thấp/ILO1000 ──────────

class TestLoadCoreChannelFilter:
    """
    GL02_*_1000.zip chứa TOÀN BỘ bút toán chi nhánh 1000 (mọi kênh). Xác nhận
    qua dữ liệu thật 06/07/2026: chỉ LOCAC=501202 & CUSTOMER=1000-000007709
    (44.569/1.972.319 dòng, ~2%) thuộc kênh Citad Thấp/ILO1000 — phần còn lại
    (LOCAC 502001, 502003...) là kênh khác, không được đưa vào đối chiếu.
    """

    def test_only_matching_locac_and_customer_kept(self, tmp_path):
        from backend.services.ilo1000.load_core import load_core

        zpath = tmp_path / 'z.zip'
        rows = [
            _gl02_core_row('ILO_REF', trbrcd='1000'),
            _gl02_core_row('OTHER_LOCAC_REF', trbrcd='1000'),
            _gl02_core_row('OTHER_CUSTOMER_REF', trbrcd='1000'),
        ]
        rows[0]['LOCAC'] = '501202'
        rows[0]['CUSTOMER'] = '1000-000007709'
        rows[1]['LOCAC'] = '502001'          # kênh khác (VD SWIFT/ACH) — không phải ILO
        rows[1]['CUSTOMER'] = '1000-000007709'
        rows[2]['LOCAC'] = '501202'
        rows[2]['CUSTOMER'] = 'khac-khong-phai-ilo'
        _write_gl02_zip(zpath, rows)

        df = load_core([zpath])
        assert list(df['REFERENCE']) == ['ILO_REF'], (
            "Chỉ giữ đúng LOCAC=501202 & CUSTOMER=1000-000007709, loại các dòng kênh khác"
        )

    def test_whitespace_in_locac_customer_still_matches(self, tmp_path):
        """GL02 thật có thể có khoảng trắng thừa quanh giá trị — vẫn phải khớp sau strip."""
        from backend.services.ilo1000.load_core import load_core

        zpath = tmp_path / 'z.zip'
        row = _gl02_core_row('ILO_REF_WS')
        row['LOCAC'] = ' 501202 '
        row['CUSTOMER'] = ' 1000-000007709 '
        _write_gl02_zip(zpath, [row])

        df = load_core([zpath])
        assert list(df['REFERENCE']) == ['ILO_REF_WS']


# ── Test: Tổng hợp số món + số tiền (Hub, Core & Citad) trong sheet Tóm tắt ─

class TestExportTongHop:
    """Sheet 'Tóm tắt' phải có bảng tổng số món + tổng số tiền cho Hub, Core và Citad."""

    def test_tong_hop_so_mon_so_tien(self, tmp_path):
        from backend.services.ilo1000.export import export_excel
        import openpyxl

        hub_df = pd.DataFrame([
            {'Số giao dịch': 'SA001', 'Số tiền thực chuyển': 1_000_000},
            {'Số giao dịch': 'SA002', 'Số tiền thực chuyển': 500_000},
        ])
        core_df = pd.DataFrame([
            {'TRBRCD': '1220', 'REFERENCE': 'REF1', 'DRAMOUNT': 0, 'CRAMOUNT': 1_000_000, 'TT': 'citad 6.7'},
            {'TRBRCD': '1220', 'REFERENCE': 'REF2', 'DRAMOUNT': 0, 'CRAMOUNT': 2_000_000, 'TT': ''},
        ])
        citad_df = pd.DataFrame([
            {'SERIAL_NO': 'S1', 'RELATION_NO': '1220x', 'TRX_DATE': '20260706',
             'AMOUNT': 1_000_000, 'Trace': 't1', 'Map dc': 'm1', 'Ngày': 20260706},
        ])
        out_path = export_excel(hub_df, citad_df, pd.DataFrame(), core_df, 20260706, tmp_path)

        wb = openpyxl.load_workbook(out_path, data_only=True)
        ws = wb['Tóm tắt']
        rows = [tuple(c.value for c in row) for row in ws.iter_rows()]

        header_idx = next(i for i, r in enumerate(rows) if r[0] == 'Sheet')
        assert rows[header_idx][:4] == ('Sheet', 'Số món', 'Tổng Nợ', 'Tổng Có / Số tiền')

        hub_row = rows[header_idx + 1]
        assert hub_row[0] == 'Hub' and hub_row[1] == 2 and hub_row[3] == 1_500_000, f"Hub totals sai: {hub_row}"

        core_row = rows[header_idx + 2]
        assert core_row[:4] == ('Core', 2, 0, 3_000_000), f"Core totals sai: {core_row}"

        citad_row = rows[header_idx + 3]
        assert citad_row[0] == 'Citad' and citad_row[1] == 1 and citad_row[3] == 1_000_000, (
            f"Citad totals sai: {citad_row}"
        )

    def test_empty_hub_df_does_not_crash(self, tmp_path):
        """Ngày không có file Hub (hub_df rỗng, không có cột nào) — không được crash."""
        from backend.services.ilo1000.export import export_excel

        core_df = pd.DataFrame([
            {'TRBRCD': '1220', 'REFERENCE': 'REF1', 'DRAMOUNT': 0, 'CRAMOUNT': 1_000_000, 'TT': ''},
        ])
        out_path = export_excel(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), core_df, 20260714, tmp_path)
        assert out_path.exists()


# ── Test: round-trip đọc lại sheet hub/citad NHIỀU dòng — chống bug xlsxwriter
#          constant_memory ghi column-major làm mất dữ liệu mọi dòng trừ dòng cuối

class TestExportMultiRowRoundTrip:
    """Trước khi sửa: `constant_memory=True` + `df.to_excel()` (ghi theo CỘT) làm
    xlsxwriter âm thầm bỏ các lệnh ghi "lùi dòng" — chỉ dòng CUỐI của mỗi sheet
    giữ đủ mọi cột, các dòng trước chỉ còn cột đầu tiên. `TestExportTongHop` chỉ
    kiểm sheet 'Tóm tắt' (3 dòng) nên không bắt được lỗi này ở sheet hub/citad —
    test dưới đọc lại TỪNG DÒNG, TỪNG CỘT của cả 2 sheet để bắt đúng lớp lỗi đó."""

    def test_hub_sheet_every_row_keeps_all_columns(self, tmp_path):
        from backend.services.ilo1000.export import export_excel
        import openpyxl

        hub_df = pd.DataFrame([
            {'Số giao dịch': 'SA001', 'Số Ref Hub': 'R1', 'STC': 'C1', 'Trace': 'T1',
             'Trace2': 1, 'Số tiền thực chuyển': 100_000, 'Trạng thái': 'Hoàn thành',
             'Ngày giờ kênh trả': '06/07/2026 08:00:00', 'Nội dung chuyển tiền': 'nd1', 'ngay': 20260706},
            {'Số giao dịch': 'SA002', 'Số Ref Hub': 'R2', 'STC': 'C2', 'Trace': 'T2',
             'Trace2': 2, 'Số tiền thực chuyển': 200_000, 'Trạng thái': 'Hoàn thành',
             'Ngày giờ kênh trả': '06/07/2026 09:00:00', 'Nội dung chuyển tiền': 'nd2', 'ngay': 20260706},
            {'Số giao dịch': 'SA003', 'Số Ref Hub': 'R3', 'STC': 'C3', 'Trace': 'T3',
             'Trace2': 3, 'Số tiền thực chuyển': 300_000, 'Trạng thái': 'Hoàn thành',
             'Ngày giờ kênh trả': '06/07/2026 10:00:00', 'Nội dung chuyển tiền': 'nd3', 'ngay': 20260706},
        ])
        core_df = pd.DataFrame([{'TRBRCD': '1220', 'REFERENCE': 'REF1', 'DRAMOUNT': 0, 'CRAMOUNT': 0, 'TT': ''}])
        out_path = export_excel(hub_df, pd.DataFrame(), pd.DataFrame(), core_df, 20260706, tmp_path)

        wb = openpyxl.load_workbook(out_path, data_only=True)
        ws = wb['hub']
        rows = [tuple(c.value for c in row) for row in ws.iter_rows()]
        header, data_rows = rows[0], rows[1:4]

        assert header[0] == 'Số giao dịch'
        for i, (so_gd, so_tien, nd) in enumerate([
            ('SA001', 100_000, 'nd1'), ('SA002', 200_000, 'nd2'), ('SA003', 300_000, 'nd3'),
        ]):
            assert data_rows[i][0] == so_gd, f"dòng {i}: Số giao dịch sai: {data_rows[i]}"
            assert data_rows[i][5] == so_tien, f"dòng {i}: Số tiền thực chuyển mất/sai: {data_rows[i]}"
            assert data_rows[i][8] == nd, f"dòng {i}: Nội dung chuyển tiền mất/sai: {data_rows[i]}"

    def test_citad_sheet_every_row_keeps_all_columns(self, tmp_path):
        from backend.services.ilo1000.export import export_excel
        import openpyxl

        citad_df = pd.DataFrame([
            {'SERIAL_NO': f'S{i}', 'RELATION_NO': f'REL{i}', 'TRX_DATE': '20260706',
             'AMOUNT': (i + 1) * 100_000, 'Trace': f'T{i}', 'Map dc': f'M{i}', 'Ngày': 20260706}
            for i in range(3)
        ])
        core_df = pd.DataFrame([{'TRBRCD': '1220', 'REFERENCE': 'REF1', 'DRAMOUNT': 0, 'CRAMOUNT': 0, 'TT': ''}])
        out_path = export_excel(pd.DataFrame(), citad_df, pd.DataFrame(), core_df, 20260706, tmp_path)

        wb = openpyxl.load_workbook(out_path, data_only=True)
        ws = wb['citad']
        rows = [tuple(c.value for c in row) for row in ws.iter_rows()]
        header, data_rows = rows[0], rows[1:4]

        assert header[0] == 'SERIAL_NO'
        for i in range(3):
            assert data_rows[i][0] == f'S{i}', f"dòng {i}: SERIAL_NO mất/sai: {data_rows[i]}"
            assert data_rows[i][3] == (i + 1) * 100_000, f"dòng {i}: AMOUNT mất/sai: {data_rows[i]}"
            assert data_rows[i][5] == f'M{i}', f"dòng {i}: Map dc mất/sai: {data_rows[i]}"


# ── Test: load_citad — dòng CSV bị "ragged" do CI_NAME có dấu phẩy chưa escape ─

class TestLoadCitadRaggedLine:
    """
    Tên ngân hàng quốc tế viết theo thông lệ "X Bank, Ltd CN <chi nhánh>" chứa
    dấu phẩy chưa escape trong CSV citad thật — làm dòng có NHIỀU field hơn
    header (VD 12 thay vì 11 field), khiến pandas coi là dòng lỗi và ÂM THẦM
    BỎ QUA CẢ DÒNG (mất hẳn giao dịch, không chỉ sai cột AMOUNT). Xác nhận qua
    dữ liệu thật 04-06/7/2026: 4 giao dịch CITAD bị thiếu hoàn toàn vì lý do
    này (SERIAL_NO 11793751, 12814557, 18820932, 17813753).
    """

    def _write_citad_csv(self, path, lines):
        header = 'ID,SERIAL_NO,RELATION_NO,O_CI_ID,R_CI_ID,TRX_DATE,CI_CODE,CI_NAME,AMOUNT,TRX_STATUS,TELLERID'
        path.write_text(header + '\n' + '\n'.join(lines) + '\n', encoding='utf-8')

    def test_row_with_comma_in_bank_name_not_dropped(self, tmp_path):
        from backend.services.ilo1000.load_citad import load_citad

        p = tmp_path / 'citad.csv'
        self._write_citad_csv(p, [
            'LFO1,11793751,1420HUB500152,01204009,01653001,20260706,01653001,'
            'Ngân hàng MUFG Bank, Ltd CN Ha Noi,315861641.00,110100010000101,',
        ])
        df = load_citad([p])
        assert len(df) == 1, "Dòng có dấu phẩy trong tên ngân hàng không được bị bỏ qua"
        assert df['SERIAL_NO'].iloc[0] == '11793751'
        assert float(df['AMOUNT'].iloc[0]) == 315861641.00

    def test_normal_row_without_comma_unaffected(self, tmp_path):
        from backend.services.ilo1000.load_citad import load_citad

        p = tmp_path / 'citad.csv'
        self._write_citad_csv(p, [
            'LFO2,99999999,1234HUB000001,01204009,01653001,20260706,01653001,'
            'Ngan hang binh thuong,1000000.00,OK,',
        ])
        df = load_citad([p])
        assert len(df) == 1
        assert float(df['AMOUNT'].iloc[0]) == 1000000.00

    def test_multiple_commas_in_bank_name_still_recovered(self, tmp_path):
        """2 dấu phẩy trong tên (3 field dư) vẫn phải gộp đúng về 1 CI_NAME."""
        from backend.services.ilo1000.load_citad import load_citad

        p = tmp_path / 'citad.csv'
        self._write_citad_csv(p, [
            'LFO3,88888888,5678HUB000002,01204009,01653001,20260706,01653001,'
            'Ngan hang A, B, C Ltd,2000000.00,OK,',
        ])
        df = load_citad([p])
        assert len(df) == 1, "Nhiều dấu phẩy trong tên vẫn phải gộp về đúng 1 dòng"
        assert float(df['AMOUNT'].iloc[0]) == 2000000.00


# ── Test: load_citad lọc theo TRX_DATE — 1 file citad gộp nhiều ngày ────────

class TestLoadCitadMultiDayFile:
    """
    Xác nhận qua dữ liệu thật 14-15/7/2026: cả 5 file citad "cổng" đều trộn
    lẫn TRX_DATE của CẢ 2 ngày trong CÙNG 1 file (khác các bộ trước — mỗi file
    chỉ 1 ngày). detect.py không còn đoán ngày theo file cho citad nữa (gán
    chung 1 pool cho mọi nhóm ngày) — load_citad() phải tự lọc đúng TRX_DATE.
    """

    def _write_citad_csv(self, path, rows):
        header = 'ID,SERIAL_NO,RELATION_NO,O_CI_ID,R_CI_ID,TRX_DATE,CI_CODE,CI_NAME,AMOUNT,TRX_STATUS,TELLERID'
        lines = [
            f'{r["id"]},{r["serial"]},{r["rel"]},01204009,01653001,{r["trdate"]},01653001,'
            f'Ngan hang,{r["amt"]},OK,'
            for r in rows
        ]
        path.write_text(header + '\n' + '\n'.join(lines) + '\n', encoding='utf-8')

    def test_filters_to_requested_day_only(self, tmp_path):
        from backend.services.ilo1000.load_citad import load_citad

        p = tmp_path / 'citad.csv'
        self._write_citad_csv(p, [
            {'id': 'L1', 'serial': 'S1', 'rel': '1220HUB1', 'trdate': '20260714', 'amt': '1000000'},
            {'id': 'L2', 'serial': 'S2', 'rel': '1220HUB2', 'trdate': '20260715', 'amt': '2000000'},
        ])
        df14 = load_citad([p], ngay_int=20260714)
        assert list(df14['SERIAL_NO']) == ['S1']

        df15 = load_citad([p], ngay_int=20260715)
        assert list(df15['SERIAL_NO']) == ['S2']

    def test_no_ngay_int_keeps_all_days(self, tmp_path):
        """Không truyền ngay_int (VD gọi lẻ) — giữ nguyên hành vi cũ, không lọc."""
        from backend.services.ilo1000.load_citad import load_citad

        p = tmp_path / 'citad.csv'
        self._write_citad_csv(p, [
            {'id': 'L1', 'serial': 'S1', 'rel': '1220HUB1', 'trdate': '20260714', 'amt': '1000000'},
            {'id': 'L2', 'serial': 'S2', 'rel': '1220HUB2', 'trdate': '20260715', 'amt': '2000000'},
        ])
        df = load_citad([p])
        assert len(df) == 2


class TestDetectCitadPoolAcrossDays:
    """detect.py không được đoán ngày theo file citad — gán chung 1 pool cho
    mọi nhóm ngày, để load_citad() tự lọc TRX_DATE."""

    def test_citad_pool_assigned_to_every_date_group(self, tmp_path):
        from backend.services.ilo1000.detect import group_files_by_date

        citad_csv = tmp_path / 'abc123.csv'
        citad_csv.write_text(
            'ID,SERIAL_NO,RELATION_NO,O_CI_ID,R_CI_ID,TRX_DATE,CI_CODE,CI_NAME,AMOUNT,TRX_STATUS,TELLERID\n'
            'L1,S1,1220HUB1,01204009,01653001,20260714,01653001,Ngan hang,1000000,OK,\n'
            'L2,S2,1220HUB2,01204009,01653001,20260715,01653001,Ngan hang,2000000,OK,\n',
            encoding='utf-8',
        )
        core14 = tmp_path / 'gl02_20260714.csv'
        core14.write_text('TRDATE,TRBRCD\n20260714,1000\n', encoding='utf-8')
        core15 = tmp_path / 'gl02_20260715.csv'
        core15.write_text('TRDATE,TRBRCD\n20260715,1000\n', encoding='utf-8')

        groups = group_files_by_date([citad_csv, core14, core15])
        assert '20260714' in groups and '20260715' in groups
        assert groups['20260714']['citad'] == [citad_csv]
        assert groups['20260715']['citad'] == [citad_csv]


# ── Test: nhận diện + pool OSB — tên file không đáng tin về ngày ────────────

class TestDetectOSBPool:
    """detect.py không được đoán ngày theo tên file OSB (xác nhận thật: file
    'OSB n 11-12.7.xlsx' chứa dữ liệu tháng 8, không phải tháng 7) — gán chung
    1 pool cho mọi nhóm ngày, để load_osb() tự lọc 'Ngày hạch toán'."""

    def test_osb_filename_detected(self):
        from pathlib import Path
        from backend.services.ilo1000.detect import detect_file_type

        assert detect_file_type(Path('OSB n 11-12.7.xlsx')) == 'osb'
        assert detect_file_type(Path('osb n 10.7.xlsx')) == 'osb'

    def test_osb_pool_assigned_to_every_date_group(self, tmp_path):
        from backend.services.ilo1000.detect import group_files_by_date

        osb_file = tmp_path / 'OSB n 11-12.7.xlsx'
        osb_file.write_text('placeholder', encoding='utf-8')  # nội dung không quan trọng ở bước nhóm ngày
        core14 = tmp_path / 'gl02_20260714.csv'
        core14.write_text('TRDATE,TRBRCD\n20260714,1000\n', encoding='utf-8')
        core15 = tmp_path / 'gl02_20260715.csv'
        core15.write_text('TRDATE,TRBRCD\n20260715,1000\n', encoding='utf-8')

        groups = group_files_by_date([osb_file, core14, core15])
        assert groups['20260714']['osb'] == [osb_file]
        assert groups['20260715']['osb'] == [osb_file]


# ── Test: load_osb — đọc file OSB thật (title dòng 1, trống dòng 2, header dòng 3) ─

def _write_osb_xlsx(path, rows):
    """Ghi file OSB đúng cấu trúc thật: title row0, blank row1, header row2 (0-based)."""
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([['DỮ LIỆU CHI TIẾT HẠCH TOÁN']]).to_excel(
            writer, sheet_name='Sheet 1', index=False, header=False, startrow=0
        )
        df.to_excel(writer, sheet_name='Sheet 1', index=False, header=True, startrow=2)


def _osb_row(ma_gd, cn_thuc_hien, so_tien, ngay_hach_toan):
    return {
        'Mã giao dịch': ma_gd, 'CN thực hiện': cn_thuc_hien,
        'Số tiền': so_tien, 'Ngày hạch toán': ngay_hach_toan,
    }


class TestLoadOSB:
    def test_reads_real_shaped_file(self, tmp_path):
        from backend.services.ilo1000.load_osb import load_osb

        p = tmp_path / 'OSB n 11-12.7.xlsx'
        _write_osb_xlsx(p, [
            _osb_row('141470277', '3510 - Agribank CN Ngọc Lặc Thanh Hóa', '70.000', '10/08/2026'),
            _osb_row('141635580', '4809 - Agribank CN Bắc Bình Bình Thuận', '250.426', '10/08/2026'),
        ])
        df = load_osb([p])
        assert len(df) == 2
        assert list(df['Mã giao dịch']) == ['141470277', '141635580']

    def test_filters_by_ngay_ints_ignores_filename(self, tmp_path):
        """Tên file gợi ý tháng 7 nhưng dữ liệu ghi tháng 8 — phải lọc đúng
        theo cột 'Ngày hạch toán', không suy ngày từ tên file."""
        from backend.services.ilo1000.load_osb import load_osb

        p = tmp_path / 'OSB n 11-12.7.xlsx'
        _write_osb_xlsx(p, [
            _osb_row('141470277', '3510 - CN A', '70.000', '11/08/2026'),
            _osb_row('141635580', '4809 - CN B', '250.000', '12/08/2026'),
            _osb_row('141668800', '1401 - CN C', '1.250.000', '13/08/2026'),
        ])
        df = load_osb([p], ngay_ints=20260811)
        assert list(df['Mã giao dịch']) == ['141470277']

    def test_carryover_multiple_ngay_ints(self, tmp_path):
        """Carryover 'OSB cũ chưa đi': truyền tập nhiều ngày cùng lúc."""
        from backend.services.ilo1000.load_osb import load_osb

        p = tmp_path / 'OSB n 11-12.7.xlsx'
        _write_osb_xlsx(p, [
            _osb_row('141470277', '3510 - CN A', '70.000', '11/08/2026'),
            _osb_row('141635580', '4809 - CN B', '250.000', '12/08/2026'),
            _osb_row('141668800', '1401 - CN C', '1.250.000', '13/08/2026'),
        ])
        df = load_osb([p], ngay_ints={20260811, 20260812})
        assert set(df['Mã giao dịch']) == {'141470277', '141635580'}

    def test_dedup_by_ma_giao_dich(self, tmp_path):
        from backend.services.ilo1000.load_osb import load_osb

        p1 = tmp_path / 'OSB n 10.7.xlsx'
        p2 = tmp_path / 'OSB n 10.7b.xlsx'
        _write_osb_xlsx(p1, [_osb_row('141470277', '3510 - CN A', '70.000', '10/08/2026')])
        _write_osb_xlsx(p2, [_osb_row('141470277', '3510 - CN A', '70.000', '10/08/2026')])

        df = load_osb([p1, p2])
        assert len(df) == 1


# ── Test: build_osb_key — LEFT(CN,4) & Mã giao dịch & Số tiền ───────────────

class TestBuildOsbKey:
    def test_key_matches_citad_map_dc_shape(self):
        """Khóa OSB phải cùng khuôn với Map dc Citad (LEFT(RELATION_NO,4) &
        Trace & AMOUNT) — xác nhận bằng dữ liệu thật 2026-08-19."""
        from backend.services.ilo1000.load_osb import build_osb_key

        df = pd.DataFrame([_osb_row('141470277', '3510 - Agribank CN Ngọc Lặc', '70.000', '10/08/2026')])
        key = build_osb_key(df)
        assert key.iloc[0] == '3510' + '141470277' + '70000'

    def test_raises_on_unknown_amount_format(self):
        """Số tiền không đúng mẫu đã biết (thuần số hoặc ngăn nghìn 3 chữ số)
        → raise, không đoán — dùng chung doc_so_tien() với module ACH."""
        from backend.services.ilo1000.load_osb import build_osb_key

        df = pd.DataFrame([_osb_row('141470277', '3510 - CN A', '70,000', '10/08/2026')])
        with pytest.raises(ValueError):
            build_osb_key(df)

    def test_empty_df_returns_empty_series(self):
        from backend.services.ilo1000.load_osb import build_osb_key

        key = build_osb_key(pd.DataFrame())
        assert len(key) == 0


# ── Test: used_citad_keys — Citad Map dc THỰC SỰ được Core chọn trúng ───────

class TestUsedCitadKeys:
    def test_matched_key_included(self):
        from backend.services.ilo1000.process import used_citad_keys

        core_out = pd.DataFrame([{'Map dc': 'M1', 'TT': 'citad 6.7'}, {'Map dc': 'M2', 'TT': ''}])
        citad_mapdc = {'M1': 'citad 6.7', 'M2': 'citad 6.7'}
        assert used_citad_keys(core_out, citad_mapdc) == {'M1'}

    def test_key_present_in_dict_but_overridden_by_higher_priority_not_used(self):
        """M1 có trong citad_mapdc nhưng dòng Core đó bị Hủy/Quyết toán ghi đè
        trước — TT cuối KHÔNG bằng nhãn citad → không tính là 'đã dùng'."""
        from backend.services.ilo1000.process import used_citad_keys

        core_out = pd.DataFrame([{'Map dc': 'M1', 'TT': 'Hủy'}])
        citad_mapdc = {'M1': 'citad 6.7'}
        assert used_citad_keys(core_out, citad_mapdc) == set()

    def test_empty_core_returns_empty_set(self):
        from backend.services.ilo1000.process import used_citad_keys

        assert used_citad_keys(pd.DataFrame(), {}) == set()


# ── Test: match_citad_leftover_with_osb — Citad còn thừa khớp OSB ───────────

class TestMatchCitadLeftoverWithOsb:
    def test_leftover_matched_gets_tt_osb(self):
        from backend.services.ilo1000.process import match_citad_leftover_with_osb

        citad_df = pd.DataFrame([{'Map dc': 'M1'}, {'Map dc': 'M2'}])
        used = {'M1'}  # M1 đã khớp Core; M2 còn thừa
        osb_key = pd.Series(['M2'])
        out = match_citad_leftover_with_osb(citad_df, used, osb_key)
        assert list(out['TT']) == ['', 'OSB']

    def test_matched_by_core_not_touched_even_if_also_in_osb(self):
        """Dòng đã khớp Core (nằm trong used_citad_mapdc) không đụng tới cột
        TT dù khóa đó tình cờ cũng có trong OSB — đã coi là xử lý xong ở Core."""
        from backend.services.ilo1000.process import match_citad_leftover_with_osb

        citad_df = pd.DataFrame([{'Map dc': 'M1'}])
        used = {'M1'}
        osb_key = pd.Series(['M1'])
        out = match_citad_leftover_with_osb(citad_df, used, osb_key)
        assert out['TT'].iloc[0] == ''

    def test_leftover_not_in_osb_stays_blank(self):
        """Còn thừa nhưng không khớp OSB → để RỖNG cho người chấm tay xem xét,
        không tự suy luận thêm."""
        from backend.services.ilo1000.process import match_citad_leftover_with_osb

        citad_df = pd.DataFrame([{'Map dc': 'M3'}])
        out = match_citad_leftover_with_osb(citad_df, set(), pd.Series(['OTHER']))
        assert out['TT'].iloc[0] == ''

    def test_no_osb_data_all_leftover_stay_blank(self):
        from backend.services.ilo1000.process import match_citad_leftover_with_osb

        citad_df = pd.DataFrame([{'Map dc': 'M1'}])
        out = match_citad_leftover_with_osb(citad_df, set(), None)
        assert out['TT'].iloc[0] == ''

    def test_empty_citad_df_does_not_crash(self):
        from backend.services.ilo1000.process import match_citad_leftover_with_osb

        out = match_citad_leftover_with_osb(pd.DataFrame(), set(), pd.Series(['X']))
        assert out.empty


# ── Test: _osb_carryover_days — cửa sổ carryover "OSB cũ chưa đi" ───────────

class TestOsbCarryoverDays:
    def test_regular_day_includes_yesterday(self):
        from backend.services.ilo1000.pipeline import _osb_carryover_days

        assert _osb_carryover_days(20260812) == {20260812, 20260811}

    def test_monday_includes_fri_sat_sun(self):
        """10/08/2026 là thứ 2 — carryover phải gồm 7,8,9/08 (thứ 6,7,CN)."""
        from backend.services.ilo1000.pipeline import _osb_carryover_days

        assert _osb_carryover_days(20260810) == {20260810, 20260809, 20260808, 20260807}

    def test_month_boundary(self):
        from backend.services.ilo1000.pipeline import _osb_carryover_days

        assert _osb_carryover_days(20260801) == {20260801, 20260731}


# ── Test: _filter_core_by_date — 1 file GL02 gốc chứa nhiều ngày trộn lẫn ───

class TestFilterCoreByDate:
    """`core_raw` truyền cho `_run_one_day()` là core nạp cho CẢ BATCH nhiều
    ngày (để detect_huy() bắt Hủy khác ngày) — nhưng phải lọc lại đúng TRDATE
    trước khi xuất, nếu không sheet 'core' của MỌI ngày trong batch giống hệt
    nhau. Xác nhận thật 2026-08-19: file "1000_gl02_2026081120260812.csv"
    chứa cả TRDATE 11 và 12/08 trộn lẫn — trước khi sửa, cả 2 file xuất ra
    (20260811.xlsx và 20260812.xlsx) đều có sheet core giống hệt 144.037 dòng."""

    def test_keeps_only_matching_trdate(self):
        from backend.services.ilo1000.pipeline import _filter_core_by_date

        df = pd.DataFrame([
            {'TRDATE': '20260811  ', 'REFERENCE': 'A'},
            {'TRDATE': '20260812  ', 'REFERENCE': 'B'},
            {'TRDATE': '20260811  ', 'REFERENCE': 'C'},
        ])
        out = _filter_core_by_date(df, '20260811')
        assert sorted(out['REFERENCE']) == ['A', 'C']

    def test_strips_padding_whitespace_before_compare(self):
        """TRDATE trong file thật có khoảng trắng đệm cố định (fixed-width)."""
        from backend.services.ilo1000.pipeline import _filter_core_by_date

        df = pd.DataFrame([{'TRDATE': '20260812         ', 'REFERENCE': 'X'}])
        out = _filter_core_by_date(df, '20260812')
        assert len(out) == 1

    def test_missing_trdate_column_returns_unchanged(self):
        from backend.services.ilo1000.pipeline import _filter_core_by_date

        df = pd.DataFrame([{'REFERENCE': 'A'}])
        out = _filter_core_by_date(df, '20260811')
        assert len(out) == 1


# ── Test: main_from_dir dedup all_core_df — CSV/ZIP GL02 trùng dữ liệu ──────

class TestDedupCoreBatch:
    """
    1 file GL02 gốc có thể được export CẢ CSV (thường) lẫn ZIP (mã hoá) —
    xác nhận thật 2026-08-19: CSV và ZIP của "1000_gl02_2026081120260812"
    giống hệt nhau từng dòng (288.074 dòng gộp → 144.037 sau dedup). Vì
    file-date-extraction gán CSV/ZIP vào 2 khóa ngày KHÁC NHAU (mỗi file tự
    chứa nhiều ngày trộn lẫn), dedup nội bộ của `load_core()` (dedup trong 1
    lời gọi) không bắt được cặp trùng lặp XUYÊN 2 khóa ngày —
    `_dedup_core_batch()` phải tự dedup khi gộp toàn batch trước khi tính
    huy_map.

    Lưu ý phạm vi: với dữ liệu Core THẬT (sau `load_core()`, DRAMOUNT luôn
    = 0 do đã lọc lúc đọc — vế "hủy" chỉ có thể biểu diễn qua CRAMOUNT ÂM),
    tín hiệu CR<0 độc lập (P1) đã khiến `detect_huy()` không còn nhạy với
    kiểu trùng lặp này nữa — dedup ở đây chủ yếu là vệ sinh dữ liệu (tránh xử
    lý dư ~2x dữ liệu) và phòng ngừa cho các cách gọi khác của `detect_huy()`
    (không đi qua `load_core()`, có thể còn dữ liệu DR≠0) — không kiểm bằng
    kịch bản huy_map cụ thể vì không tái hiện được qua luồng thật.
    """

    def test_identical_rows_from_different_sources_deduped(self):
        from backend.services.ilo1000.pipeline import _dedup_core_batch

        row = _gl02_core_row('REF1')
        out = _dedup_core_batch({
            '20260811': pd.DataFrame([row]),
            '20260812': pd.DataFrame([row]),  # nguồn khác (VD ZIP), dòng giống hệt CSV
        })
        assert len(out) == 1

    def test_distinct_rows_all_kept(self):
        from backend.services.ilo1000.pipeline import _dedup_core_batch

        out = _dedup_core_batch({
            '20260811': pd.DataFrame([_gl02_core_row('REF1')]),
            '20260812': pd.DataFrame([_gl02_core_row('REF2')]),
        })
        assert len(out) == 2

    def test_empty_input_returns_empty_df(self):
        from backend.services.ilo1000.pipeline import _dedup_core_batch

        assert _dedup_core_batch({}).empty
