"""
Khoá THỨ TỰ DÒNG của file Excel "Đối chiếu CITAD"
(`doi_chieu_citad_service.build_xlsx`).

Vì sao cần: toàn bộ vị trí từ dòng 12 trở xuống tính bằng biến `row` tăng
dần, nên thêm/bớt một kênh sẽ dịch mọi dòng phía dưới — kể cả dòng "Chênh
lệch" và khối chữ ký LẬP BẢNG / KIỂM SOÁT. Sai ở đây KHÔNG ném lỗi, chỉ ra
một file Excel trông vẫn bình thường nhưng lệch bố cục so với bản Phòng
Thanh toán vẫn in.

Đồng thời canh hai bất biến nghiệp vụ đã nhiều lần bị hiểu nhầm:
  - Kênh Ebanking đã bỏ hẳn khỏi bản in (20/08/2026) — payload cũ còn gửi
    `em`/`et` thì cũng KHÔNG được rò ra file.
  - Ebanking CHƯA BAO GIỜ được cộng vào tổng CITAD, nên bỏ dòng in không
    được phép làm đổi con số "Chênh lệch".

Không mock: ghi .xlsx thật vào bộ nhớ rồi đọc lại bằng openpyxl.
"""
import io

import openpyxl

from backend.schemas.doi_chieu_citad import ExportIn
from backend.services.doi_chieu_citad_service import build_xlsx

FK = ['di_ih_m', 'di_ih_t', 'di_il_m', 'di_il_t', 'den_ih_m', 'den_ih_t', 'den_il_m', 'den_il_t']
CONGS = [1, 9, 18, 17, 12]
CURS = ['VNĐ', 'USD', 'EUR']


def _export(**over) -> ExportIn:
    """Bộ dữ liệu mẫu CÂN BẰNG khi bỏ qua Ebanking: Cổng 1 (10 món/1.000) +
    Napas (3/300) + PSS-MDP (2/200) = PaymentHub VNĐ (15/1.500)."""
    gD = {str(c): {u: {f: 0.0 for f in FK} for u in CURS} for c in CONGS}
    gD['1']['VNĐ']['den_ih_m'] = 10
    gD['1']['VNĐ']['den_ih_t'] = 1000
    phD = {u: {f: 0.0 for f in FK} for u in CURS}
    phD['VNĐ']['den_ih_m'] = 15
    phD['VNĐ']['den_ih_t'] = 1500
    kw = dict(
        day_str='20/08/2026', sheet_name='20_08_2026', lb='Nguyễn Văn A', ks='Trần Thị B',
        gD=gD, phD=phD, nm=3, nt=300, sm=2, st=200,
    )
    kw.update(over)
    return ExportIn(**kw)


def _sheet(data: ExportIn):
    return openpyxl.load_workbook(io.BytesIO(build_xlsx(data))).active


def _labels(ws) -> list:
    """Nhãn cột A theo đúng thứ tự xuất hiện, bỏ dòng trống."""
    return [ws.cell(r, 1).value for r in range(8, ws.max_row + 1) if ws.cell(r, 1).value]


# ── Bố cục ────────────────────────────────────────────────────────────────
def test_thu_tu_dong_dung_va_du():
    ws = _sheet(_export())
    assert _labels(ws) == [
        'Payment EUR', 'Payment USD', 'Payment VNĐ', 'CITAD',
        'Cổng 1', 'Cổng 9', 'Cổng 18', 'Cổng 17', 'Cổng 12',
        'Napas', 'PSS - MDP', 'Chênh lệch',
        '                  LẬP BẢNG',
    ]


def test_khoi_chu_ky_nam_duoi_dong_chenh_lech():
    """LẬP BẢNG / KIỂM SOÁT phải cách "Chênh lệch" đúng 2 dòng, và tên người
    ký nằm cách đó 5 dòng nữa (chừa chỗ ký tay) — dịch sai thì bản in đè chữ."""
    ws = _sheet(_export())
    rows = {ws.cell(r, 1).value: r for r in range(8, ws.max_row + 1) if ws.cell(r, 1).value}
    r_diff = rows['Chênh lệch']
    r_sign = rows['                  LẬP BẢNG']
    assert r_sign == r_diff + 2
    assert ws.cell(r_sign, 8).value == 'KIỂM SOÁT'
    assert ws.cell(r_sign + 5, 2).value == 'Nguyễn Văn A'
    assert ws.cell(r_sign + 5, 8).value == 'Trần Thị B'
    assert ws.print_area == f"'20_08_2026'!$A$1:$J${r_sign + 5}"


# ── Ebanking đã bỏ hẳn khỏi bản in ────────────────────────────────────────
def test_khong_con_dong_ebanking_du_payload_cu_van_gui_em_et():
    """Session đã chấm TRƯỚC 14/08 vẫn còn `em`/`et` trong payload. Bản in
    mới không được hiện dòng Ebanking, cũng không được rò con số đó ra bất
    kỳ ô nào."""
    ws = _sheet(_export(em=99, et=9999))
    assert not [
        (r, c)
        for r in range(1, ws.max_row + 1)
        for c in range(1, 11)
        if isinstance(ws.cell(r, c).value, str) and 'banking' in ws.cell(r, c).value.lower()
    ]
    assert not [
        (r, c)
        for r in range(1, ws.max_row + 1)
        for c in range(1, 11)
        if ws.cell(r, c).value in (99, 99.0, 9999, 9999.0)
    ]


# ── Bất biến nghiệp vụ: em/et không đụng vào Chênh lệch ───────────────────
def test_em_et_khong_lam_doi_chenh_lech():
    """Ebanking chưa bao giờ được cộng vào tổng CITAD — cùng bộ dữ liệu, có
    hay không có `em`/`et` thì dòng Chênh lệch phải y hệt (và bằng 0)."""
    ws0 = _sheet(_export())
    ws1 = _sheet(_export(em=99, et=9999))

    def _diff(ws):
        r = next(r for r in range(8, ws.max_row + 1) if ws.cell(r, 1).value == 'Chênh lệch')
        return [ws.cell(r, c).value for c in range(3, 11)]

    assert _diff(ws0) == _diff(ws1) == [None] * 8  # 0 được ghi thành ô trống


def test_chenh_lech_van_bat_duoc_sai_lech():
    """Canh chiều ngược lại: test trên không phải luôn-xanh vì mọi thứ đều
    trống — lệch thật thì phải hiện đúng con số lệch."""
    ws = _sheet(_export(nm=5))  # Napas dư 2 món so với PaymentHub
    r = next(r for r in range(8, ws.max_row + 1) if ws.cell(r, 1).value == 'Chênh lệch')
    assert ws.cell(r, 7).value == 2
