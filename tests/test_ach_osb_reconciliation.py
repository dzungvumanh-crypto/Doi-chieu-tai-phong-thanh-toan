"""Test thuật toán Điểm 2 — b9_doi_chieu_osb.py (đối chiếu lệnh OSB qua file
Quyết toán OSB "QT").

Chạy: .venv\\Scripts\\python.exe -m pytest tests/test_ach_osb_reconciliation.py -v
"""
import pandas as pd
import pytest
import xlsxwriter
import openpyxl

from backend.services.ach.b9_doi_chieu_osb import xu_ly_qt, doi_chieu_osb_di, doi_chieu_osb_den
from backend.services.ach.pipeline import xuat_excel_osb, xuat_excel


_HEADER_QT = [
    'STT', 'Kênh', 'Mã dịch vụ', 'Dịch vụ', 'Chiều giao dịch', 'Mã giao dịch',
    'IPCAS Trace', 'Mã giao dịch gốc', '', 'Seq', 'CN thực hiện', 'CN hạch toán',
    'TK ghi nợ', 'TK ghi có', 'Tài khoản chuyên thu', 'Tiền tệ', 'Số tiền',
    'Số bút toán OSB', 'Ref nợ', 'Ref có', 'Ngày giao dịch', 'Ngày tổng hợp',
    'Ngày hạch toán', 'Nội dung giao dịch', 'Kiểu giao dịch',
]


def _viet_file_qt(path, rows, chieu_gd='GD đi', voi_sheet_config=False):
    """rows: list of (ma_giao_dich, cn_thuc_hien, so_tien, kieu_giao_dich)."""
    wb = xlsxwriter.Workbook(str(path))
    if voi_sheet_config:
        # File QT thật luôn có sheet 'Config' đứng TRƯỚC sheet dữ liệu (bug thật
        # phát hiện khi verify dữ liệu G: 2026-07-31: sheet_name=0 bắt trúng sheet
        # này thay vì sheet dữ liệu — xem _doc_sheet_du_lieu_qt()).
        ws_cfg = wb.add_worksheet('Config')
        ws_cfg.write_row(0, 3, ['Loại CN', ''])
        ws_cfg.write_row(1, 3, ['1', 'TTĐH'])
    ws = wb.add_worksheet('Sheet 1')
    ws.write_row(0, 0, ['DỮ LIỆU CHI TIẾT HẠCH TOÁN'])
    ws.write_row(2, 0, _HEADER_QT)
    for i, (ma_gd, cn, so_tien, kieu) in enumerate(rows, start=3):
        row = [
            i - 2, 'NAPASHUB', 'NAPASACH', 'Giao dịch qua kênh HUB', chieu_gd, ma_gd,
            '', ma_gd, '', 4, cn, '1000 - Tru so chinh', '519101', '502003',
            'NPO1000', 'VND', so_tien, 'REF', 'RefN', 'RefC', '27/07/2026',
            '27/07/2026', '27/07/2026', 'noi dung', kieu,
        ]
        ws.write_row(i, 0, row)
    wb.close()


# ── xu_ly_qt ─────────────────────────────────────────────────────────────────

class TestXuLyQt:
    def test_doc_file_qt_di_dung_dinh_dang(self, tmp_path):
        path = tmp_path / 'QT đi 27.07.xlsx'
        _viet_file_qt(path, [('142088610', '5612 - CN A', 200000, 'Normal')], chieu_gd='GD đi')
        chieu, df = xu_ly_qt(str(path))
        assert chieu == 'đi'
        assert len(df) == 1
        assert df.loc[0, 'CN_TRACE_TIEN'] == '5612' + '142088610' + '200000'

    def test_doc_file_qt_den_dung_dinh_dang(self, tmp_path):
        path = tmp_path / 'QT đến 27.07.xlsx'
        _viet_file_qt(path, [('092036658', '6360 - CN B', 10000000, 'Normal')], chieu_gd='GD về')
        chieu, df = xu_ly_qt(str(path))
        assert chieu == 'đến'
        assert len(df) == 1

    def test_ma_giao_dich_bo_so_0_dau_khi_tinh_khoa(self, tmp_path):
        path = tmp_path / 'QT đến 27.07.xlsx'
        _viet_file_qt(path, [('092036658', '6360 - CN B', 10000000, 'Normal')], chieu_gd='GD về')
        _, df = xu_ly_qt(str(path))
        assert df.loc[0, 'CN_TRACE_TIEN'] == '6360' + '92036658' + '10000000'

    def test_thieu_cot_bat_buoc_bao_loi(self, tmp_path):
        path = tmp_path / 'QT_thieu_cot.xlsx'
        wb = xlsxwriter.Workbook(str(path))
        ws = wb.add_worksheet('Sheet 1')
        ws.write_row(2, 0, ['STT', 'Kênh', 'Số tiền'])
        ws.write_row(3, 0, [1, 'NAPASHUB', 100])
        wb.close()
        with pytest.raises(ValueError, match='thiếu cột'):
            xu_ly_qt(str(path))

    def test_lan_ca_2_chieu_trong_1_file_bao_loi(self, tmp_path):
        path = tmp_path / 'QT_lan_chieu.xlsx'
        wb = xlsxwriter.Workbook(str(path))
        ws = wb.add_worksheet('Sheet 1')
        ws.write_row(2, 0, _HEADER_QT)
        ws.write_row(3, 0, [1, 'NAPASHUB', 'NAPASACH', 'x', 'GD đi', '111', '', '111', '', 4,
                            '5612 - CN A', '1000', '519101', '502003', 'NPO1000', 'VND', 1000,
                            'R', 'Rn', 'Rc', '27/07/2026', '27/07/2026', '27/07/2026', 'x', 'Normal'])
        ws.write_row(4, 0, [2, 'NAPASHUB', 'NAPASACH', 'x', 'GD về', '222', '', '222', '', 4,
                            '5612 - CN A', '1000', '519101', '502003', 'NPO1000', 'VND', 2000,
                            'R', 'Rn', 'Rc', '27/07/2026', '27/07/2026', '27/07/2026', 'x', 'Normal'])
        wb.close()
        with pytest.raises(ValueError, match='Chiều giao dịch'):
            xu_ly_qt(str(path))

    def test_cn_thuc_hien_sai_dinh_dang_bao_loi(self, tmp_path):
        path = tmp_path / 'QT_sai_cn.xlsx'
        _viet_file_qt(path, [('142088610', 'CN không có mã số', 200000, 'Normal')])
        with pytest.raises(ValueError, match='CN thực hiện'):
            xu_ly_qt(str(path))

    def test_so_tien_khong_phai_so_bao_loi(self, tmp_path):
        path = tmp_path / 'QT_sai_tien.xlsx'
        _viet_file_qt(path, [('142088610', '5612 - CN A', 'abc', 'Normal')])
        with pytest.raises(ValueError, match='Số tiền'):
            xu_ly_qt(str(path))

    def test_doc_dung_sheet_du_lieu_khi_co_sheet_config_dung_truoc(self, tmp_path):
        """Regression: file QT thật có sheet 'Config' đứng trước sheet dữ liệu —
        đọc cứng sheet_name=0 (vị trí) sẽ bắt trúng 'Config' và báo lỗi thiếu
        header. Phải dò đúng sheet có 'STT'+'Số tiền' bất kể vị trí."""
        path = tmp_path / 'QT đi 27.07.xlsx'
        _viet_file_qt(path, [('142088610', '5612 - CN A', 200000, 'Normal')], voi_sheet_config=True)
        chieu, df = xu_ly_qt(str(path))
        assert chieu == 'đi'
        assert len(df) == 1


# ── doi_chieu_osb_di / doi_chieu_osb_den ────────────────────────────────────

def _mis_di_thua_row(chi_nhanh, so_trace, so_tien, osb='O'):
    key_hub = chi_nhanh.strip() + so_trace + str(so_tien)
    return {'CHI_NHANH': chi_nhanh, 'SO_TRACE': so_trace, 'SO_TIEN': so_tien,
            'LOAI_LENH_OSB': osb, 'KEY_HUB': key_hub}


def _mis_den_thua_row(chi_nhanh, trace, so_tien, osb='1'):
    key = chi_nhanh.strip() + trace + str(so_tien)
    return {'CHI_NHANH': chi_nhanh, 'TRACE': trace, 'SO_TIEN': so_tien,
            'LOAI_LENH_OSB': osb, 'KEY_DEN_HUB': key}


def _qt_row(cn_trace_tien, so_tien):
    return {'CN_TRACE_TIEN': cn_trace_tien, 'SO_TIEN': so_tien}


class TestDoiChieuOsbDi:
    def test_dong_khop_vao_da_quyet_toan(self):
        df_mis = pd.DataFrame([_mis_di_thua_row('5612', '142088610', 200000)])
        df_qt  = pd.DataFrame([_qt_row('5612142088610200000', 200000)])
        khop, chua_khop = doi_chieu_osb_di(df_mis, df_qt)
        assert len(khop) == 1
        assert len(chua_khop) == 0

    def test_dong_khong_osb_bi_loai_khoi_ca_2_ket_qua(self):
        """Dòng MIS không phải OSB (LOAI_LENH_OSB != 'O') không được đưa vào so
        khớp — không xuất hiện trong đã quyết toán lẫn chưa khớp."""
        df_mis = pd.DataFrame([_mis_di_thua_row('5612', '142088610', 200000, osb='')])
        df_qt  = pd.DataFrame([_qt_row('5612142088610200000', 200000)])
        khop, chua_khop = doi_chieu_osb_di(df_mis, df_qt)
        assert len(khop) == 0
        assert len(chua_khop) == 1  # chỉ còn dòng QT thừa
        assert chua_khop.loc[0, 'NGUON'] == 'QT'

    def test_khong_khop_vao_chua_khop_ca_2_nguon(self):
        df_mis = pd.DataFrame([_mis_di_thua_row('5612', '142088610', 200000)])
        df_qt  = pd.DataFrame([_qt_row('9999000000000001', 999)])
        khop, chua_khop = doi_chieu_osb_di(df_mis, df_qt)
        assert len(khop) == 0
        assert len(chua_khop) == 2
        assert set(chua_khop['NGUON']) == {'MIS', 'QT'}

    def test_khong_mat_du_lieu_bat_bien(self):
        df_mis = pd.DataFrame([
            _mis_di_thua_row('5612', '142088610', 200000),
            _mis_di_thua_row('5406', '142088766', 14000000),
        ])
        df_qt = pd.DataFrame([_qt_row('5612142088610200000', 200000)])
        khop, chua_khop = doi_chieu_osb_di(df_mis, df_qt)
        n_mis_chua_khop = len(chua_khop[chua_khop['NGUON'] == 'MIS'])
        assert len(khop) + n_mis_chua_khop == len(df_mis)


class TestDoiChieuOsbDen:
    def test_dong_khop_vao_da_quyet_toan(self):
        df_mis = pd.DataFrame([_mis_den_thua_row('6360', '92036658', 10000000)])
        df_qt  = pd.DataFrame([_qt_row('636092036658' + '10000000', 10000000)])
        khop, chua_khop = doi_chieu_osb_den(df_mis, df_qt)
        assert len(khop) == 1
        assert len(chua_khop) == 0

    def test_dong_khong_osb_bi_loai(self):
        df_mis = pd.DataFrame([_mis_den_thua_row('6360', '92036658', 10000000, osb='')])
        df_qt  = pd.DataFrame([_qt_row('636092036658' + '10000000', 10000000)])
        khop, chua_khop = doi_chieu_osb_den(df_mis, df_qt)
        assert len(khop) == 0
        assert len(chua_khop) == 1
        assert chua_khop.loc[0, 'NGUON'] == 'QT'


# ── xuat_excel_osb() — file OSB riêng ───────────────────────────────────────

class TestXuatExcelOsb:
    def test_tao_du_5_sheet(self, tmp_path):
        df_osb_di_khop  = pd.DataFrame([{'KEY_HUB': 'k1', 'SO_TIEN': 200000, 'CHI_NHANH': '5612'}])
        df_di_chua_khop = pd.DataFrame([
            {'NGUON': 'MIS', 'KEY_HUB': 'k2', 'SO_TIEN': 5000, 'CHI_NHANH': '1111'},
            {'NGUON': 'QT', 'CN_TRACE_TIEN': 'k3', 'SO_TIEN': 7000},
        ])
        xuat_excel_osb(str(tmp_path), '16422', __import__('datetime').datetime(2026, 7, 27),
                       df_osb_di_khop=df_osb_di_khop, df_di_chua_khop=df_di_chua_khop)

        wb = openpyxl.load_workbook(tmp_path / '20260727_ACH_OSB.xlsx')
        assert set(wb.sheetnames) == {
            'TONG_KET', 'OSB_DI_DA_QUYET_TOAN', 'OSB_DEN_DA_QUYET_TOAN',
            'DI_CHUA_KHOP', 'DEN_CHUA_KHOP',
        }

    def test_bo_cot_khoa_noi_bo(self, tmp_path):
        df_osb_di_khop  = pd.DataFrame([{'KEY_HUB': 'k1', 'SO_TIEN': 200000, 'CHI_NHANH': '5612'}])
        df_di_chua_khop = pd.DataFrame([
            {'NGUON': 'MIS', 'KEY_HUB': 'k2', 'SO_TIEN': 5000},
            {'NGUON': 'QT', 'CN_TRACE_TIEN': 'k3', 'SO_TIEN': 7000},
        ])
        xuat_excel_osb(str(tmp_path), '16422', __import__('datetime').datetime(2026, 7, 27),
                       df_osb_di_khop=df_osb_di_khop, df_di_chua_khop=df_di_chua_khop)

        wb = openpyxl.load_workbook(tmp_path / '20260727_ACH_OSB.xlsx')
        header_khop = [c.value for c in next(wb['OSB_DI_DA_QUYET_TOAN'].iter_rows(max_row=1))]
        header_ck   = [c.value for c in next(wb['DI_CHUA_KHOP'].iter_rows(max_row=1))]
        assert 'KEY_HUB' not in header_khop
        assert 'KEY_HUB' not in header_ck
        assert 'CN_TRACE_TIEN' not in header_ck
        assert 'NGUON' in header_ck

    def test_tong_ket_dung_so_lieu(self, tmp_path):
        df_osb_di_khop  = pd.DataFrame([
            {'KEY_HUB': 'k1', 'SO_TIEN': 200000}, {'KEY_HUB': 'k2', 'SO_TIEN': 100000},
        ])
        df_di_chua_khop = pd.DataFrame([
            {'NGUON': 'MIS', 'KEY_HUB': 'k3', 'SO_TIEN': 5000},
            {'NGUON': 'QT', 'CN_TRACE_TIEN': 'k4', 'SO_TIEN': 7000},
            {'NGUON': 'QT', 'CN_TRACE_TIEN': 'k5', 'SO_TIEN': 8000},
        ])
        xuat_excel_osb(str(tmp_path), '16422', __import__('datetime').datetime(2026, 7, 27),
                       df_osb_di_khop=df_osb_di_khop, df_di_chua_khop=df_di_chua_khop)

        wb = openpyxl.load_workbook(tmp_path / '20260727_ACH_OSB.xlsx')
        ws = wb['TONG_KET']
        rows = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}
        assert rows['OSB đã quyết toán (MIS khớp QT đi)'] == 2
        assert rows['OSB chưa khớp — phía MIS_đi thừa'] == 1
        assert rows['OSB chưa khớp — phía QT đi'] == 2

    def test_khong_co_qt_nao_van_khong_loi_hien_thi_rong(self, tmp_path):
        path = xuat_excel_osb(str(tmp_path), '16422', __import__('datetime').datetime(2026, 7, 27))
        wb = openpyxl.load_workbook(path)
        assert wb['OSB_DI_DA_QUYET_TOAN']['A1'].value == '(Không có dữ liệu)'


# ── C2 (sign-off mục 60) — ghi chú OSB trên TONG_KET báo cáo chính ─────────

class TestGhiChuOsbTrenBaoCaoChinh:
    def test_dem_dung_so_luong_osb_trong_mis_thua(self, tmp_path):
        """Số lệnh OSB nằm trong MIS_đi/đến thừa không còn tóm tắt trên TONG_KET
        (mẫu mới 2026-08-07 không có dòng này) — vẫn đếm được đầy đủ từ cột
        LOAI_LENH_OSB có sẵn trên chính sheet MIS_DI_THUA/MIS_DEN_THUA."""
        df_mis_di_thua = pd.DataFrame({
            'SO_TIEN': [100, 200, 300],
            'LOAI_LENH_OSB': ['O', '', 'O'],
        })
        df_mis_den_thua = pd.DataFrame({
            'SO_TIEN': [400],
            'LOAI_LENH_OSB': ['1'],
        })
        output_path = str(tmp_path / 'doi_chieu_20260727.xlsx')
        xuat_excel(
            output_path, '16422',
            pd.DataFrame({'CHI_NHANH': ['0001'], 'SO_TIEN': [1]}),
            pd.DataFrame({'CRAMOUNT': [1]}), df_mis_di_thua,
            pd.DataFrame({'SO_TIEN': [1]}),
            pd.DataFrame({'SO_TIEN': [1]}), pd.DataFrame({'DRAMOUNT': [1]}), df_mis_den_thua,
            pd.DataFrame({'BRCD': ['0001'], 'STTLMAMT': [1], 'MSGREF': ['R'], 'SessionId': ['1'],
                          'PrcFlg': ['x'], 'KEY_GW': ['k']}),
        )
        wb = openpyxl.load_workbook(output_path)

        ws_di = wb['MIS_DI_THUA']
        header_di = [c.value for c in next(ws_di.iter_rows(max_row=1))]
        col_di = header_di.index('LOAI_LENH_OSB')
        n_osb_di = sum(1 for row in ws_di.iter_rows(min_row=2) if row[col_di].value == 'O')
        assert n_osb_di == 2

        ws_den = wb['MIS_DEN_THUA']
        header_den = [c.value for c in next(ws_den.iter_rows(max_row=1))]
        col_den = header_den.index('LOAI_LENH_OSB')
        n_osb_den = sum(1 for row in ws_den.iter_rows(min_row=2) if row[col_den].value == '1')
        assert n_osb_den == 1
