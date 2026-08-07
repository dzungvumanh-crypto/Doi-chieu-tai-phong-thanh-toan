"""
Test lớp xuất Excel của pipeline ACH (`xuat_excel()`), tập trung vào phần mới
wiring Requirement C.1a (GW-thừa, `tim_nhom_gw_thua()`) vào 2 sheet Excel
GW_THUA_XAC_DINH / GW_CAN_DOI_CHIEU (xem Implementation-notes.html mục 41 và
pipeline.py::xuat_excel).

Không mock — ghi file .xlsx thật ra tmp_path rồi đọc lại bằng openpyxl, đúng
nguyên tắc "không mock I/O" của skill doi-chieu.
"""
import pandas as pd
import openpyxl

from backend.services.ach.pipeline import xuat_excel
from backend.services.ach.b11_doi_chieu_cheo_ngay import KETQUA_THUONG_KHOP


def _synthetic_dfs():
    df_mis_di_khop = pd.DataFrame({'CHI_NHANH': ['0001'], 'SO_TIEN': [1_000_000]})
    df_npo_di_thua = pd.DataFrame({'CRAMOUNT': [500_000]})
    df_mis_di_thua = pd.DataFrame({'SO_TIEN': [200_000], 'LOAI_LENH_OSB': ['']})
    df_timeout     = pd.DataFrame({'SO_TIEN': [300_000]})
    df_mis_den_khop  = pd.DataFrame({'SO_TIEN': [400_000]})
    df_npo_den_thua  = pd.DataFrame({'DRAMOUNT': [100_000]})
    df_mis_den_thua  = pd.DataFrame({'SO_TIEN': [150_000], 'LOAI_LENH_OSB': ['']})
    df_gw_raw = pd.DataFrame({
        'BRCD': ['0001', '0002'], 'STTLMAMT': [1_000_000, 2_000_000],
        'MSGREF': ['REF1', 'REF2'], 'SessionId': ['16282', '16282'],
        'PrcFlg': ['Lệnh Hoàn thành', 'Lệnh Hoàn thành'],
        'KEY_GW': ['00011000000', '00022000000'],
    })

    # C.1a — GW thừa xác định chắc chắn (COUNT_MIS == 0 cho nhóm đó): toàn GW.
    df_gw_thua_xac_dinh = pd.DataFrame({
        'BRCD': ['0003'], 'STTLMAMT': [777_000], 'MSGREF': ['REF3'],
        'SessionId': ['16282'], 'PrcFlg': ['Lệnh Hoàn thành'],
        'KEY_GW': ['00037770000'], 'SOURCE': ['GW'], 'NHOM_CN_TIEN': ['00037770000'],
    })

    # C.1a — cần đối chiếu thủ công: trộn dòng GW + dòng MIS cùng nhóm.
    df_gw_can_doi_chieu = pd.DataFrame({
        'BRCD': ['0004', None], 'STTLMAMT': [888_000, None],
        'MSGREF': ['REF4', 'REF5'], 'SessionId': ['16282', None],
        'PrcFlg': ['Lệnh Hoàn thành', None],
        'KEY_GW': ['00048880000', None],
        'CHI_NHANH': [None, '0004'], 'SO_TIEN': [None, 888_000],
        'CN tiền Hub': [None, '00048880000'],
        'SOURCE': ['GW', 'MIS'], 'NHOM_CN_TIEN': ['00048880000', '00048880000'],
    })

    # Điểm 3 — điện đi huỷ trong ngày (2 dòng đối ứng) / khác ngày (1 dòng Cancel).
    df_dien_huy_trong_ngay = pd.DataFrame({
        'TRBRCD': ['1240', '1240'], 'CRAMOUNT': [3_000_000, -3_000_000],
        'TRTP': ['Normal', 'Cancel'],
    })
    df_dien_huy_khac_ngay = pd.DataFrame({
        'TRBRCD': ['1300'], 'CRAMOUNT': [-6_000_000], 'TRTP': ['Cancel'],
    })

    return dict(
        df_mis_di_khop=df_mis_di_khop, df_npo_di_thua=df_npo_di_thua,
        df_mis_di_thua=df_mis_di_thua, df_timeout=df_timeout,
        df_mis_den_khop=df_mis_den_khop, df_npo_den_thua=df_npo_den_thua,
        df_mis_den_thua=df_mis_den_thua, df_gw_raw=df_gw_raw,
        df_gw_thua_xac_dinh=df_gw_thua_xac_dinh,
        df_gw_can_doi_chieu=df_gw_can_doi_chieu,
        df_dien_huy_trong_ngay=df_dien_huy_trong_ngay,
        df_dien_huy_khac_ngay=df_dien_huy_khac_ngay,
    )


def test_gw_thua_sheets_present(tmp_path):
    """2 sheet mới GW_THUA_XAC_DINH / GW_CAN_DOI_CHIEU phải xuất hiện trong Excel."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_gw_thua_xac_dinh=dfs['df_gw_thua_xac_dinh'],
               df_gw_can_doi_chieu=dfs['df_gw_can_doi_chieu'])

    wb = openpyxl.load_workbook(output_path)
    assert 'GW_THUA_XAC_DINH' in wb.sheetnames
    assert 'GW_CAN_DOI_CHIEU' in wb.sheetnames


def test_gw_thua_sheets_drop_internal_key_columns(tmp_path):
    """KEY_GW (và 'CN tiền Hub' ở sheet trộn) là cột khóa nội bộ, không xuất ra Excel."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_gw_thua_xac_dinh=dfs['df_gw_thua_xac_dinh'],
               df_gw_can_doi_chieu=dfs['df_gw_can_doi_chieu'])

    wb = openpyxl.load_workbook(output_path)
    header_xac_dinh = [c.value for c in next(wb['GW_THUA_XAC_DINH'].iter_rows(max_row=1))]
    header_can_doi  = [c.value for c in next(wb['GW_CAN_DOI_CHIEU'].iter_rows(max_row=1))]

    assert 'KEY_GW' not in header_xac_dinh
    assert 'KEY_GW' not in header_can_doi
    assert 'CN tiền Hub' not in header_can_doi
    # Cột nghiệp vụ vẫn còn đủ để đối chiếu tay.
    assert 'SOURCE' in header_xac_dinh and 'SOURCE' in header_can_doi
    assert 'NHOM_CN_TIEN' in header_can_doi


def test_tong_ket_reports_gw_thua_counts(tmp_path):
    """GW thừa xác định + cần đối chiếu thủ công không còn tóm tắt trên TONG_KET
    (mẫu mới 2026-08-07 không có 2 dòng này) — vẫn phải đọc được đầy đủ số dòng ở
    sheet riêng GW_THUA_XAC_DINH/GW_CAN_DOI_CHIEU (dữ liệu không mất)."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_gw_thua_xac_dinh=dfs['df_gw_thua_xac_dinh'],
               df_gw_can_doi_chieu=dfs['df_gw_can_doi_chieu'])

    wb = openpyxl.load_workbook(output_path)
    assert wb['GW_THUA_XAC_DINH'].max_row - 1 == len(dfs['df_gw_thua_xac_dinh'])
    assert wb['GW_CAN_DOI_CHIEU'].max_row - 1 == len(dfs['df_gw_can_doi_chieu'])


def test_xuat_excel_khong_co_gw_thua_van_chay_duoc(tmp_path):
    """Khi không truyền df_gw_thua_xac_dinh/df_gw_can_doi_chieu (None mặc định) —
    Excel vẫn xuất được, sheet hiển thị '(Không có dữ liệu)', không lỗi."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260723.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'])

    wb = openpyxl.load_workbook(output_path)
    assert 'GW_THUA_XAC_DINH' in wb.sheetnames
    assert 'GW_CAN_DOI_CHIEU' in wb.sheetnames
    assert wb['GW_THUA_XAC_DINH']['A1'].value == '(Không có dữ liệu)'
    assert wb['GW_CAN_DOI_CHIEU']['A1'].value == '(Không có dữ liệu)'


# ── Điểm 3 (2026-07-31) — sheet DIEN_DI_HUY_TRONG_NGAY / DIEN_DI_HUY_KHAC_NGAY ──

def test_dien_huy_sheets_present(tmp_path):
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260731.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_dien_huy_trong_ngay=dfs['df_dien_huy_trong_ngay'],
               df_dien_huy_khac_ngay=dfs['df_dien_huy_khac_ngay'])

    wb = openpyxl.load_workbook(output_path)
    assert 'DIEN_DI_HUY_TRONG_NGAY' in wb.sheetnames
    assert 'DIEN_DI_HUY_KHAC_NGAY' in wb.sheetnames
    assert wb['DIEN_DI_HUY_TRONG_NGAY'].max_row == 3  # header + 2 dòng, KHÔNG có dòng tổng
    assert wb['DIEN_DI_HUY_KHAC_NGAY'].max_row == 3   # header + 1 dòng + 1 dòng TỔNG


def test_dien_huy_khac_ngay_co_dong_tong(tmp_path):
    """Sheet DIEN_DI_HUY_KHAC_NGAY phải kèm dòng TỔNG (số món + số tiền) cuối
    sheet — đúng yêu cầu PR gốc, khác DIEN_DI_HUY_TRONG_NGAY."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260731.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_dien_huy_khac_ngay=dfs['df_dien_huy_khac_ngay'])

    wb = openpyxl.load_workbook(output_path)
    ws = wb['DIEN_DI_HUY_KHAC_NGAY']
    dong_tong = [c.value for c in ws[3]]
    assert dong_tong[0] == 'TỔNG: 1 món'
    assert dong_tong[1] == '-6,000,000 VND'


def test_tong_ket_tach_dien_huy_khoi_npo_di_thua(tmp_path):
    """TONG_KET (mẫu mới) phải hiển thị đúng số dòng huỷ trong ngày/khác ngày dưới
    nhãn có ngày thật, và dòng 'IPCAS' (thay cho 'Tổng NPO_DI (cần đối)' cũ) phải
    cộng đủ cả 2 khoản huỷ (bảo toàn số học)."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260731.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_dien_huy_trong_ngay=dfs['df_dien_huy_trong_ngay'],
               df_dien_huy_khac_ngay=dfs['df_dien_huy_khac_ngay'])

    wb = openpyxl.load_workbook(output_path)
    ws = wb['TONG_KET']
    rows = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}

    ngay_display = '31/07/2026'
    assert rows[f'huỷ trong ngày {ngay_display}'] == 2
    assert rows[f'huỷ khác ngày {ngay_display}'] == 1
    # IPCAS (đi) = n_di_khop + n_npo_di_thua + n_huy_trong_ngay + n_huy_khac_ngay —
    # đúng công thức "Tổng NPO_DI (cần đối)" cũ, chỉ đổi tên hiển thị.
    n_di_khop = len(dfs['df_mis_di_khop'])
    n_npo_di_thua = len(dfs['df_npo_di_thua'])
    n_huy_trong_ngay = len(dfs['df_dien_huy_trong_ngay'])
    n_huy_khac_ngay = len(dfs['df_dien_huy_khac_ngay'])
    assert rows['IPCAS'] == n_di_khop + n_npo_di_thua + n_huy_trong_ngay + n_huy_khac_ngay


# ── Điểm 4 (2026-07-31) — cột GHI_CHU_T2 trên NPO_DI_THUA/NPO_DEN_THUA ─────────

def test_ghi_chu_t2_xuat_hien_tren_npo_di_thua_va_npo_den_thua(tmp_path):
    dfs = _synthetic_dfs()
    dfs['df_npo_di_thua']  = dfs['df_npo_di_thua'].assign(GHI_CHU_T2=['Hạch toán lệnh ngày T-2'])
    dfs['df_npo_den_thua'] = dfs['df_npo_den_thua'].assign(GHI_CHU_T2=[''])
    output_path = str(tmp_path / 'doi_chieu_20260731.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'])

    wb = openpyxl.load_workbook(output_path)
    header_di  = [c.value for c in next(wb['NPO_DI_THUA'].iter_rows(max_row=1))]
    header_den = [c.value for c in next(wb['NPO_DEN_THUA'].iter_rows(max_row=1))]
    assert 'GHI_CHU_T2' in header_di
    assert 'GHI_CHU_T2' in header_den

    di_col = header_di.index('GHI_CHU_T2') + 1
    assert wb['NPO_DI_THUA'].cell(row=2, column=di_col).value == 'Hạch toán lệnh ngày T-2'


def test_ghi_chu_t2_khong_xuat_hien_tren_sheet_huy_diem_3(tmp_path):
    """GHI_CHU_T2 chỉ có ý nghĩa trên NPO_DI_THUA/NPO_DEN_THUA — 2 sheet huỷ của
    Điểm 3 không nên bị ảnh hưởng bởi wiring cột mới của Điểm 4."""
    dfs = _synthetic_dfs()
    dfs['df_npo_di_thua'] = dfs['df_npo_di_thua'].assign(GHI_CHU_T2=[''])
    output_path = str(tmp_path / 'doi_chieu_20260731.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_dien_huy_trong_ngay=dfs['df_dien_huy_trong_ngay'],
               df_dien_huy_khac_ngay=dfs['df_dien_huy_khac_ngay'])

    wb = openpyxl.load_workbook(output_path)
    header_trong_ngay = [c.value for c in next(wb['DIEN_DI_HUY_TRONG_NGAY'].iter_rows(max_row=1))]
    header_khac_ngay  = [c.value for c in next(wb['DIEN_DI_HUY_KHAC_NGAY'].iter_rows(max_row=1))]
    assert 'GHI_CHU_T2' not in header_trong_ngay
    assert 'GHI_CHU_T2' not in header_khac_ngay


# ── Audit 2026-08-04 — sheet SESSION_NULL_BI_LOAI (giao dịch SESSION=NULL bị loại
# khỏi MIS_đi, mọi lý do — trước đây hoàn toàn vô hình ở báo cáo cuối) ──────────

def test_session_null_bi_loai_sheet_xuat_hien_va_co_du_lieu(tmp_path):
    dfs = _synthetic_dfs()
    df_bi_loai = pd.DataFrame({
        'REFHUB': ['R1'], 'MSGREF': ['MSG1'], 'SO_TIEN': [999_000],
        'LY_DO_CAN_KIEM_TRA': ['GW_SESSION_KHAC'],
    })
    output_path = str(tmp_path / 'doi_chieu_20260804.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_session_null_bi_loai=df_bi_loai)

    wb = openpyxl.load_workbook(output_path)
    assert 'SESSION_NULL_BI_LOAI' in wb.sheetnames
    ws = wb['SESSION_NULL_BI_LOAI']
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert 'LY_DO_CAN_KIEM_TRA' in header
    ly_do_col = header.index('LY_DO_CAN_KIEM_TRA') + 1
    assert ws.cell(row=2, column=ly_do_col).value == 'GW_SESSION_KHAC'


def test_tong_ket_bao_cao_so_luong_session_null_bi_loai(tmp_path):
    """Số dòng SESSION=NULL bị loại không còn tóm tắt trên TONG_KET (mẫu mới
    2026-08-07 không có dòng này) — vẫn đọc được đầy đủ ở sheet riêng
    SESSION_NULL_BI_LOAI (dữ liệu không mất)."""
    dfs = _synthetic_dfs()
    df_bi_loai = pd.DataFrame({
        'REFHUB': ['R1', 'R2'], 'SO_TIEN': [999_000, 111_000],
        'LY_DO_CAN_KIEM_TRA': ['GW_SESSION_KHAC', 'KHONG_TIM_THAY_TREN_GW_TAI_T-1'],
    })
    output_path = str(tmp_path / 'doi_chieu_20260804.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_session_null_bi_loai=df_bi_loai)

    wb = openpyxl.load_workbook(output_path)
    assert wb['SESSION_NULL_BI_LOAI'].max_row - 1 == 2


def test_khong_co_session_null_bi_loai_van_chay_duoc(tmp_path):
    """Không truyền df_session_null_bi_loai (None mặc định, hoặc DataFrame rỗng) —
    Excel vẫn xuất được, sheet SESSION_NULL_BI_LOAI hiện '(Không có dữ liệu)',
    không lỗi."""
    dfs = _synthetic_dfs()
    output_path = str(tmp_path / 'doi_chieu_20260804.xlsx')

    xuat_excel(output_path, '16282', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'])

    wb = openpyxl.load_workbook(output_path)
    assert 'SESSION_NULL_BI_LOAI' in wb.sheetnames
    assert wb['SESSION_NULL_BI_LOAI']['A1'].value == '(Không có dữ liệu)'


# ── TONG_KET mẫu mới (2026-08-07, đã cập nhật lại theo đúng file mẫu) ─────────
# Cấu trúc đúng cho mục "Điện thường T-2 hạch toán T-1" (đối xứng hệt mục "Điện
# OSB T-2" — đã làm đúng từ đầu):
#   1. 1 dòng DUY NHẤT mang cả nhãn "Điện thanh toán thường đi... ngày T-2 hạch
#      toán ngày T-1" lẫn số liệu T-2 thật (không tách dòng tiêu đề riêng).
#   2. NGAY SAU đó là dòng LẶP LẠI y hệt dòng "khớp NPO cùng ngày" phía trên
#      (cùng nhãn "Lệnh đi ngày T-1 hạch toán NPO ngày T-1", cùng số liệu
#      n_di_khop/s_di_khop) — KHÔNG được xoá (đã xoá nhầm 1 lần trong phiên,
#      người dùng phản hồi lại: "mong muốn có hai dòng giống nhau").

def test_dong_dien_thuong_t2_gop_1_dong_va_lap_lai_dong_khop_npo(tmp_path):
    dfs = _synthetic_dfs()
    df_ketqua_di_t2 = pd.DataFrame({
        'SO_TIEN': [100_000], 'KET_QUA': [KETQUA_THUONG_KHOP],
    })
    output_path = str(tmp_path / 'doi_chieu_20260803.xlsx')

    xuat_excel(output_path, '16454', dfs['df_mis_di_khop'], dfs['df_npo_di_thua'],
               dfs['df_mis_di_thua'], dfs['df_timeout'], dfs['df_mis_den_khop'],
               dfs['df_npo_den_thua'], dfs['df_mis_den_thua'], dfs['df_gw_raw'],
               df_ketqua_di_t2=df_ketqua_di_t2)

    wb = openpyxl.load_workbook(output_path)
    all_rows = [(row[0].value, row[1].value, row[2].value)
                for row in wb['TONG_KET'].iter_rows(min_row=2) if row[0].value]

    nhan_dien_thuong_t2 = ('Điện thanh toán thường đi (không phải OSB) ngày '
                           '02/08/2026 hạch toán ngày 03/08/2026')
    nhan_khop_cung_ngay = 'Lệnh đi ngày 03/08/2026 hạch toán NPO ngày 03/08/2026'
    nhan_phu_cu_sai     = 'Lệnh đi ngày 02/08/2026 hạch toán NPO ngày 03/08/2026'

    # Dòng "Điện thường T-2" mang đúng số liệu ngay trên chính nó.
    assert (nhan_dien_thuong_t2, 1, 100_000) in all_rows
    # Dòng "khớp NPO cùng ngày" xuất hiện ĐÚNG 2 LẦN — hai dòng giống nhau, cùng
    # số liệu thật (n_di_khop=1, s_di_khop=1_000_000 theo _synthetic_dfs).
    lan_xuat_hien = [r for r in all_rows if r[0] == nhan_khop_cung_ngay]
    assert lan_xuat_hien == [(nhan_khop_cung_ngay, 1, 1_000_000)] * 2
    # Nhãn sai đã xoá ở lần sửa trước (ngày T-2 lẫn vào vế "hạch toán NPO") —
    # không được tái xuất hiện.
    assert nhan_phu_cu_sai not in [r[0] for r in all_rows]
